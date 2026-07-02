# -*- coding: utf-8 -*-
"""生成《上海 / 长三角产业研学考察计划》Excel 排期表 + 联合主办合作框架（对等）。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import content as C

# ---------------------------------------------------------------- 样式
NAVY = "142B45"; BLUE = "1F4E79"; STEEL = "2E6DA4"; AMBER = "E88A1A"
GREEN = "2E7D53"; RED = "B03A2E"
LIGHT = "EEF3F8"; BAND1 = "E3ECF4"; BAND2 = "EDF3EA"; BAND3 = "F6EEE1"; WHITE = "FFFFFF"

F = "微软雅黑"
thin = Side(style="thin", color="C9D6E3")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="center")
WRAP_L = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def tfont(sz=11, color=WHITE, bold=True):
    return Font(name=F, size=sz, bold=bold, color=color)


def bfont(sz=10, color="222A33", bold=False):
    return Font(name=F, size=sz, bold=bold, color=color)


def fill(hexc):
    return PatternFill("solid", fgColor=hexc)


def header_row(ws, row, ncols, bg=NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg); cell.font = tfont(); cell.alignment = CENTER; cell.border = BORDER


wb = Workbook()

# ================================================================ Sheet 1 说明
ws = wb.active
ws.title = "说明"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 96

ws["B2"] = C.PROJECT_TITLE
ws["B2"].font = Font(name=F, size=18, bold=True, color=NAVY)
ws["B3"] = C.PROJECT_SUBTITLE
ws["B3"].font = Font(name=F, size=13, bold=True, color=AMBER)

rows = [
    ("联合主办（乙方）", "复旦大学住房政策研究中心 · 上海市科技企业联合会 · 上海市杨浦区科技企业联合会"),
    ("战略合作（甲方）", "中国钢铁工业协会（总会层面）"),
    ("合作层级", C.STRATEGIC_LEVEL),
    ("形式", "四期 · 每期两天半（2.5 天） · 上海（长三角延伸）"),
    ("规模", C.SCALE),
    ("成本口径", C.LOGISTICS),
    ("考察内容", "不动产标杆项目（考察项目本身，非开发企业） · 标杆园区 · 智能制造 · 科技企业 · 供需闭门撮合"),
    ("战略目标", C.STRATEGIC_GOAL),
    ("本表结构", "四期总览 → 各期排期（时间节点） → 合作框架（对等） → 对价与分成落位"),
]
r = 5
for k, v in rows:
    kc = ws.cell(row=r, column=2, value=k)
    kc.font = bfont(11, NAVY, True); kc.alignment = WRAP; kc.fill = fill(LIGHT); kc.border = BORDER
    vc = ws.cell(row=r, column=3, value=v)
    vc.font = bfont(11); vc.alignment = WRAP_L; vc.border = BORDER
    ws.row_dimensions[r].height = 30
    r += 1

# ================================================================ Sheet 2 四期总览
ws = wb.create_sheet("四期总览")
ws.sheet_view.showGridLines = False
for i, w in enumerate([10, 14, 26, 40, 34, 30], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.merge_cells("A1:F1")
ws["A1"] = "四期考察总览"; ws["A1"].font = Font(name=F, size=15, bold=True, color=NAVY)
ws["A1"].alignment = WRAP_L; ws.row_dimensions[1].height = 26
for c, h in enumerate(["期次", "时间窗口", "主题", "聚焦方向", "用钢新场景", "本期主题课"], start=1):
    ws.cell(row=2, column=c, value=h)
header_row(ws, 2, 6); ws.row_dimensions[2].height = 24
bands = [BAND1, BAND2, BAND3, LIGHT]
for i, t in enumerate(C.TOURS):
    row = 3 + i
    for c, v in enumerate([t["code"], t["window"], t["theme"], t["focus"], t["steel"], t["lecture"]], start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill(bands[i % 4]); cell.font = bfont(10, NAVY, c in (1, 3))
        cell.alignment = WRAP_L; cell.border = BORDER
    ws.row_dimensions[row].height = 62

start = 8
ws.cell(row=start, column=1, value="各期考察点清单").font = Font(name=F, size=13, bold=True, color=NAVY)
start += 1
for c, h in enumerate(["期次", "类别", "考察点（项目 / 园区 / 工厂）", "看点 · 用钢关联"], start=1):
    ws.cell(row=start, column=c, value=h)
header_row(ws, start, 4); ws.row_dimensions[start].height = 22
ws.column_dimensions["D"].width = 40
r = start + 1
for i, t in enumerate(C.TOURS):
    for j, (cat, name, note) in enumerate(t["sites"]):
        code = t["code"] if j == 0 else ""
        for c, v in enumerate([code, cat, name, note], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill(bands[i % 4] if code else WHITE)
            cell.font = bfont(10, STEEL if c == 2 else "222A33", c in (1, 2, 3))
            cell.alignment = WRAP_L; cell.border = BORDER
        ws.row_dimensions[r].height = 26
        r += 1

# ================================================================ 各期排期表
def schedule_rows(tour):
    idx = 0
    out = []
    for d, tm, act, note in C.DAY_TEMPLATE:
        act2, note2 = act, note
        if act.startswith("参访点"):
            cat, name, sn = tour["sites"][idx]
            act2 = "参访 " + act[3] + "：" + name
            note2 = "[" + cat + "] " + sn
            idx += 1
        elif "主题报告" in act:
            note2 = tour["lecture"]
        out.append((d, tm, act2, note2))
    return out


def make_schedule_sheet(tour):
    ws = wb.create_sheet(tour["code"] + "排期")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [12, 16, 40, 52]):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:D1")
    ws["A1"] = tour["code"] + " · " + tour["window"] + " · " + tour["theme"] + "（两天半）"
    ws["A1"].font = Font(name=F, size=14, bold=True, color=NAVY); ws["A1"].alignment = WRAP_L
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:D2")
    ws["A2"] = "聚焦：" + tour["focus"] + "　|　用钢新场景：" + tour["steel"]
    ws["A2"].font = bfont(10, STEEL, True); ws["A2"].alignment = WRAP_L; ws.row_dimensions[2].height = 30
    for c, h in enumerate(["日期", "时间", "环节", "说明 / 考察点看点"], start=1):
        ws.cell(row=3, column=c, value=h)
    header_row(ws, 3, 4); ws.row_dimensions[3].height = 22
    day_bg = {"第 1 天": BAND1, "第 2 天": BAND2, "第 3 天": BAND3}
    r = 4
    for d, tm, act, note in schedule_rows(tour):
        for c, v in enumerate([d, tm, act, note], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill(day_bg.get(d, WHITE))
            cell.font = bfont(10, NAVY if c == 1 else "222A33", c in (1, 3))
            cell.alignment = CENTER if c in (1, 2) else WRAP_L; cell.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1


for t in C.TOURS:
    make_schedule_sheet(t)

# ================================================================ Sheet 合作框架（对等）
ws = wb.create_sheet("合作框架")
ws.sheet_view.showGridLines = False
for col, w in zip("ABC", [4, 30, 92]):
    ws.column_dimensions[col].width = w
R = {"r": 2}


def title_at(txt, sz=16):
    ws.cell(row=R["r"], column=2, value=txt).font = Font(name=F, size=sz, bold=True, color=NAVY)
    R["r"] += 1


def intro(txt):
    ws.merge_cells(start_row=R["r"], start_column=2, end_row=R["r"], end_column=3)
    c = ws.cell(row=R["r"], column=2, value=txt); c.font = bfont(11); c.alignment = WRAP_L
    ws.row_dimensions[R["r"]].height = 44
    R["r"] += 1


def section(txt, bg=BLUE):
    ws.merge_cells(start_row=R["r"], start_column=2, end_row=R["r"], end_column=3)
    c = ws.cell(row=R["r"], column=2, value=txt)
    c.font = tfont(12, WHITE, True); c.fill = fill(bg); c.alignment = WRAP_L
    ws.row_dimensions[R["r"]].height = 24
    R["r"] += 1


def kv(k, v, kbg=LIGHT, kcolor=NAVY):
    kc = ws.cell(row=R["r"], column=2, value=k)
    kc.font = bfont(11, kcolor, True); kc.fill = fill(kbg); kc.alignment = WRAP_L; kc.border = BORDER
    vc = ws.cell(row=R["r"], column=3, value=v)
    vc.font = bfont(11); vc.alignment = WRAP_L; vc.border = BORDER
    txt = str(v)
    est = sum(max(1, (len(seg) // 44) + 1) for seg in txt.split("\n"))
    ws.row_dimensions[R["r"]].height = max(28, 17 * est + 6)
    R["r"] += 1


title_at("联合主办合作框架（对等 · 总会对总会 · 商务对商务）")
intro(C.FRAMEWORK_INTRO)
R["r"] += 1

section("一、合作层级与双方主体（对等）")
for role, name, duty in C.PARTIES:
    kv(role, name + "\n职责：" + duty)
R["r"] += 1

section("二、针对“各收各钱、自负盈亏”方案的指导性立场（红线）", bg=RED)
for i, t in enumerate(C.STANCE, start=1):
    kv("红线 " + str(i), t, kbg=BAND3, kcolor=RED)
R["r"] += 1

section("三、我方核心商业资产界定")
for a, b, c in C.ASSETS:
    kv(a, b + "\n价值点：" + c)
R["r"] += 1

section("四、三种商业对价方案（三选一或组合）", bg=STEEL)
for m in C.MODELS:
    kv(m["name"], "适用：" + m["case"] + "\n" + "\n".join("• " + x for x in m["rights"]))
R["r"] += 1

section("五、对外沟通话术（可直接用）", bg=GREEN)
kv("① 商业升维版（试水）", C.SCRIPT_UPGRADE)
kv("② 战略高位切入版（正式磋商）", C.SCRIPT_HIGH)
R["r"] += 1

section("六、我方底线", bg=RED)
for i, t in enumerate(C.BOTTOM_LINES, start=1):
    kv("底线 " + str(i), t, kbg=BAND3, kcolor=RED)
R["r"] += 1

section("七、对方反应 · 分级应对")
tier_bg = {"接受": BAND2, "犹豫": BAND3, "拒绝": "F5E1DE"}
for k, v in C.RESPONSE_TIERS:
    kv(k, v, kbg=tier_bg.get(k, LIGHT))
R["r"] += 1

section("八、目标里程碑")
for k, v in C.MILESTONES:
    kv(k, v)
kv("战略目标", C.STRATEGIC_GOAL, kbg=BAND3, kcolor=AMBER)

# ================================================================ Sheet 对价与分成落位
ws = wb.create_sheet("对价与分成落位")
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [3, 22, 20, 34, 34]):
    ws.column_dimensions[col].width = w
ws.merge_cells("B1:E1")
ws["B1"] = "对价与分成落位（明确 · 可执行 · 可追溯）"
ws["B1"].font = Font(name=F, size=15, bold=True, color=NAVY); ws["B1"].alignment = WRAP_L
ws.row_dimensions[1].height = 26
heads = ["收益类型", "计费方式", "我方对价 / 分成", "结算与凭证"]
for c, h in enumerate(heads, start=2):
    ws.cell(row=2, column=c, value=h)
for c in range(2, 6):
    cell = ws.cell(row=2, column=c)
    cell.fill = fill(NAVY); cell.font = tfont(); cell.alignment = CENTER; cell.border = BORDER
ws.row_dimensions[2].height = 24
r = 3
for i, row in enumerate(C.REVENUE_TABLE):
    for c, v in enumerate(row, start=2):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = fill(WHITE if i % 2 == 0 else LIGHT)
        col_color = {2: NAVY, 3: "222A33", 4: AMBER, 5: "5A636E"}[c]
        cell.font = bfont(11, col_color, c in (2, 4)); cell.alignment = WRAP_L; cell.border = BORDER
    ws.row_dimensions[r].height = 40
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
ws.cell(row=r, column=2, value="核心保护条款（数据与资源防火墙）").font = Font(name=F, size=13, bold=True, color=NAVY)
r += 1
for c, h in enumerate(["条款", "内容"], start=2):
    ws.cell(row=r, column=c, value=h)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
for c in range(2, 6):
    cell = ws.cell(row=r, column=c); cell.fill = fill(NAVY); cell.font = tfont(); cell.alignment = CENTER; cell.border = BORDER
ws.row_dimensions[r].height = 22
r += 1
for i, (k, v) in enumerate(C.PROTECTION):
    kc = ws.cell(row=r, column=2, value=str(i + 1) + "）" + k)
    kc.font = bfont(11, NAVY, True); kc.fill = fill(LIGHT); kc.alignment = WRAP_L; kc.border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    vc = ws.cell(row=r, column=3, value=v)
    vc.font = bfont(11); vc.alignment = WRAP_L; vc.border = BORDER
    for c in (4, 5):
        ws.cell(row=r, column=c).border = BORDER
    ws.row_dimensions[r].height = max(30, 17 * (len(v) // 70 + 1) + 6)
    r += 1
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
note = ws.cell(row=r, column=2, value="说明：会务与接待等直接运营成本可各自独立核算；上述四类对价须约定于《联合主办合作框架与对价清单》，并以名单共管库作为分佣与追溯依据（活动后 1 年追溯期）。")
note.font = bfont(10, "5A636E"); note.alignment = WRAP_L
ws.row_dimensions[r].height = 34

# ---------------------------------------------------------------- 打印布局：横向 + 适应页宽
for sh in wb.worksheets:
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sh.page_margins.left = sh.page_margins.right = 0.3
    sh.page_margins.top = sh.page_margins.bottom = 0.4

OUT = "output/中钢协_上海长三角产业研学考察计划_排期表.xlsx"
wb.save(OUT)
print("saved:", OUT, "sheets:", wb.sheetnames)
