#!/usr/bin/env python3
"""生成「WAIC UP! 北外滩之夜」双夜活动整合方案的 Excel 版本。

工作表：总览 / 两晚差异化 / Day1 议程 / Day2 议程 / WAIC 衔接与联动 / 物料与传播 / 执行 Checklist
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "WAIC2026北外滩双夜活动整合方案.xlsx"

NAVY = "1F4E79"
LIGHT = "DCE6F1"
WHITE = "FFFFFF"

thin = Side(style="thin", color="B0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color=NAVY)
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
STRIPE_FILL = PatternFill("solid", fgColor=LIGHT)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def add_sheet(wb, name, title, header, rows, widths, first_col_bold=False):
    ws = wb.create_sheet(name)
    ncols = len(header)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    for j, h in enumerate(header, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[3].height = 22

    for i, row in enumerate(rows):
        r = 4 + i
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(name="微软雅黑", size=11, bold=(first_col_bold and j == 1))
            c.alignment = WRAP
            c.border = BORDER
            if i % 2 == 1:
                c.fill = STRIPE_FILL

    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


wb = Workbook()
wb.remove(wb.active)

# ── 1. 总览 ──────────────────────────────────────────────
add_sheet(
    wb, "①总览", "「WAIC UP! 北外滩之夜」双夜活动整合方案 · 总览",
    ["项目", "内容"],
    [
        ["系列名称", "「WAIC UP! 北外滩之夜」（WAIC 2026 Side Events · DAY & NIGHT 系列风格）"],
        ["主办方", "复旦大学住房政策研究中心、上海市杨浦区科技企业联合会"],
        ["场地", "上海市北外滩 · 一滴水（滨江主厅 + 露台 + 分区社交空间）"],
        ["时间", "2026 年 7 月 17 日晚（Day 1）、7 月 18 日晚（Day 2）"],
        ["定位", "白天在展馆理解模型，晚上到北外滩追问 AI 的下一步"],
        ["差异化赛道", "Day1：AI 产业落地与建造者社交；Day2：AGI 前沿思辨（模型下一步 / 智能体 / AI for Science / 安全与治理）。与官方夜场错位：创作、开发者社区、全球游牧者、产品测评、音综均已被占坑"],
        ["呈现气质", "反形式主义 · 无 PPT · 跨圈层 · 滨江微醺 · 学术松弛"],
        ["核心思路", "17 日\u201c社交暖身型\u201d先破圈、积累人脉与报名池；18 日\u201c深度思辨型\u201d承接，把认识的人沉淀成深度讨论，形成两晚闭环"],
        ["主办方优势", "杨浦联合会：大创智、在线新经济生态园、环高校科技企业群、投融资与产业对接网络（对应 Day 1）；复旦大学：顶尖高校 AI 学术底蕴、教授与研究员网络、庞大校友圈（对应 Day 2）"],
        ["场地优势", "滨江夜景与陆家嘴天际线（社交与深夜思辨最佳背景）；主厅+露台+分区可同时支撑\u201c大场社交\u201d与\u201c圆桌+学术酒吧\u201d两种形态"],
        ["传播话术", "「用白天理解机器和模型，在晚上追问 AI 的下一步」「把 AI 谈进北外滩的夜色里」"],
    ],
    [16, 90], first_col_bold=True,
)

# ── 2. 两晚差异化 ────────────────────────────────────────
add_sheet(
    wb, "②两晚差异化", "两晚差异化一览（7/17 vs 7/18）",
    ["维度", "7月17日 Day1 · AI Builders Night 建造者之夜", "7月18日 Day2 · 通往 AGI 之夜"],
    [
        ["一句话定位", "暖身 · 社交 · 建造者连接（横向、广）", "深度 · 思辨 · AI 前沿（纵向、专）"],
        ["牵头主办", "杨浦区科技企业联合会主导", "复旦大学（学术资源）主导"],
        ["核心一问", "「你今年在造什么？」(What are you building this year?)", "「AI 的下一步会发生什么？」"],
        ["核心议题", "AI 产品、创业、出海、开源、投融资", "大模型下一步、智能体、AI for Science、安全对齐与治理"],
        ["主打形式", "闪电演讲 + 1 分钟快闪开放麦 + Mixer 社交", "跨界圆桌 Panel + 学术酒吧围炉"],
        ["目标受众", "创业者 / 开发者 / 产品经理 / 投资人", "学者 / 研究员 / AI 企业领袖 / 深度从业者"],
        ["建议规模", "120–150 人（开放报名 + 定向邀请）", "80–100 人（以定向邀请为主，精品化）"],
        ["现场氛围", "灵感盲盒、松弛微醺、快节奏", "严肃但松弛、思想碰撞、慢深度"],
        ["参考对标", "AI Founders Mixer / WAIC Afterparty / AI 创造者之夜", "AI 创作者之夜（议题错位到\u201cAGI 前沿\u201d而非\u201c内容创作\u201d）"],
        ["报名门槛", "低门槛，拉新扩圈", "定向邀请为主，控质量"],
        ["时间", "18:30–21:30（散场后无缝承接）", "19:00–21:30（散场后无缝承接）"],
    ],
    [14, 44, 44], first_col_bold=True,
)

# ── 3. Day1 议程 ─────────────────────────────────────────
add_sheet(
    wb, "③Day1议程 7-17", "Day 1 · 7月17日晚「AI Builders Night · 建造者之夜」议程（18:30–21:30 · 120–150人）",
    ["时段", "环节", "内容"],
    [
        ["18:00–18:30", "签到 + 滨江破冰", "精酿 / 特调 / 咖啡；入场名牌写\u201c我正在造的 AI 项目\u201d（What are you building this year?）；浦江夜景为背景"],
        ["18:30–18:40", "开场致辞", "双主办致辞，介绍系列与玩法规则"],
        ["18:40–19:30", "闪电演讲 × 4（每人10分钟）", "① AI 产品从 Demo 到 PMF 的实战；② AI 出海最新观察；③ 爆款 AI 应用拆解；④ 开源 / Agent 生态增长经验"],
        ["19:30–20:00", "1 分钟快闪开放麦", "报名即开麦：Demo 展示 / 招募搭档 / 疯狂想法，支持\u201c带货演讲\u201d"],
        ["20:00–20:30", "AMA 圆桌", "3–4 位创业与投资嘉宾，现场 Ask Me Anything"],
        ["20:30–21:30", "自由社交（游走式）", "无固定座位；设\u201c投资人蹲项目区 / 企业对接区 / 招聘与找搭档区\u201d"],
        ["嘉宾方向", "4–6 位核心，跨圈层", "AI 应用 / Agent 创业者、AI 出海创业者、开源 Maintainer / 独立开发者、资深 AI 投资人、大模型 / Infra 技术负责人"],
        ["差异化亮点", "—", "入场名牌破冰；投资人\u201c蹲项目\u201d独立区；杨浦产业资源对接区；滨江露台微醺社交"],
    ],
    [14, 26, 70], first_col_bold=True,
)

# ── 4. Day2 议程 ─────────────────────────────────────────
add_sheet(
    wb, "④Day2议程 7-18", "Day 2 · 7月18日晚「通往 AGI 之夜」议程（19:00–21:30 · 80–100人）",
    ["时段", "环节", "内容"],
    [
        ["18:30–19:00", "签到 + 迎宾", "滨江晚宴式茶歇；发放\u201cAGI 之问\u201d话题卡，入场即思考"],
        ["19:00–19:10", "开场", "复旦学者致辞定调，抛出核心议题"],
        ["19:10–20:10", "主题圆桌 Panel（无 PPT 深谈）", "「通往 AGI 的路上，我们走到哪了？」——学者 × 大模型企业 × 智能体产品 × 投资人，站在链条不同位置展开"],
        ["20:10–20:40", "学术酒吧 · 分主题围炉", "3 桌轮换：① 模型桌（Scaling 之后：架构、推理、端侧）；② 智能体桌（Agent 元年之后：工作流、评测、商业化）；③ 科学与治理桌（AI for Science、安全对齐、开源与监管），用知识下酒"],
        ["20:40–21:00", "观点碰撞 / 快问快答", "现场提问与辩题投票（示例辩题：「通往 AGI，缺的是算力还是想法？」），无标准答案的开放讨论"],
        ["21:00–21:30", "自由社交", "滨江夜景下的深度连接，沉淀两晚人脉"],
        ["核心议题", "—", "Scaling Law 之后大模型的下一步？智能体从 Demo 到可信赖生产力还差几步？AI 会先在哪个学科点燃下一次科学革命？安全、对齐与治理如何跟上能力曲线？"],
        ["嘉宾方向", "5 位跨界", "复旦 AI 方向学者（主持/定调）、大模型公司研究/技术负责人、智能体（Agent）产品创始人、AI for Science 研究员、AI 安全/治理方向学者或投资人"],
        ["差异化亮点", "—", "复旦学术定调赋予稀缺性；无 PPT 学术酒吧；学者×模型企业×产品×投资跨界同桌；辩题投票观众即参与者"],
    ],
    [14, 30, 70], first_col_bold=True,
)

# ── 5. WAIC 衔接与两晚联动 ───────────────────────────────
add_sheet(
    wb, "⑤WAIC衔接与联动", "与白天 WAIC 的衔接机制 & 两晚联动闭环",
    ["类别", "维度", "具体做法"],
    [
        ["WAIC 衔接", "时间衔接", "均安排在 WAIC 白天散场后（Day1 18:30 起 / Day2 19:00 起），无缝承接"],
        ["WAIC 衔接", "空间衔接", "提供主展馆 → 北外滩一滴水的散场接驳（大巴 / 拼车 / 打车指引），降低到场门槛"],
        ["WAIC 衔接", "内容衔接", "Day1 把白天看到的 Demo 与发布拉到\u201c真实落地\u201d检验；Day2 把白天的技术发布沉淀为\u201cAI 下一步\u201d的深度追问"],
        ["WAIC 衔接", "嘉宾衔接", "邀请部分白天参展 / 演讲嘉宾，晚上到北外滩继续\u201c脱稿深聊\u201d"],
        ["WAIC 衔接", "话术衔接", "「用白天理解机器和模型，在晚上追问 AI 的下一步」"],
        ["WAIC 衔接", "报名衔接", "设\u201c两晚连票 / 连报\u201d通道，17 日到场者定向邀约 18 日深度场"],
        ["两晚联动", "人群沉淀", "17 日开放场积累的报名池与到场者，定向邀约进入 18 日精品深度场"],
        ["两晚联动", "主题递进", "17 日「大家在造什么」→ 18 日「这些正在造的东西，会把 AI 带向哪里」"],
        ["两晚联动", "主办协同", "两晚均双方共同主办，各自牵头一晚，资源互补（杨浦产业动员 × 复旦学术定调）"],
        ["两晚联动", "品牌统一", "统一使用「WAIC UP! 北外滩之夜」系列视觉与话术"],
        ["两晚联动", "数据沉淀", "两晚形成完整参会名录、项目库与议题记录，可延伸为后续报告 / 沙龙"],
    ],
    [12, 14, 80], first_col_bold=True,
)

# ── 6. 物料与传播 ────────────────────────────────────────
add_sheet(
    wb, "⑥物料与传播", "物料清单 & 传播与报名",
    ["类别", "内容"],
    [
        ["现场氛围", "灯光 / 音乐 / 无固定座位布局 / 滨江露台"],
        ["互动物料", "Day1：\u201cWhat are you building\u201d入场名牌；Day2：\u201cAGI 之问\u201d话题卡、辩题投票牌"],
        ["品牌物料", "系列主视觉、背景板、指引牌、双主办 Logo 墙"],
        ["定制周边", "定制笔记本 / AI 名片 / 徽章"],
        ["餐饮", "精酿 / 特调 / 咖啡 / 茶歇（Day2 偏晚宴式）"],
        ["报名通道", "活动行 / Luma / 微信小程序；Day1 开放报名 + 定向邀请，Day2 以定向邀请为主"],
        ["接驳信息", "随报名确认推送\u201c主展馆 → 北外滩一滴水\u201d接驳与交通指引"],
        ["传播话术", "「用白天理解机器和模型，在晚上追问 AI 的下一步」「把 AI 谈进北外滩的夜色里」「第一晚认识正在造 AI 的人；第二晚想清楚 AI 会走向哪里」"],
        ["合作生态", "社区联合、品牌置换、生态伙伴（延续 Mixer / Afterparty 生态合作模式）"],
        ["命名备选（系列）", "WAIC UP! 北外滩之夜 / 一滴水 · AI 之夜 / AI Nightfall 北外滩"],
        ["命名备选（7/17）", "AI Builders Night 建造者之夜 / 你今年在造什么 · AI 之夜 / 北外滩 AI Mixer"],
        ["命名备选（7/18）", "通往 AGI 之夜 / AI 的下一步 · 深夜圆桌 / AGI 之问 · 学术酒吧夜"],
    ],
    [20, 86], first_col_bold=True,
)

# ── 7. 执行 Checklist ────────────────────────────────────
add_sheet(
    wb, "⑦执行Checklist", "执行 Checklist 与分工建议",
    ["序号", "事项", "建议牵头方", "状态"],
    [
        [1, "敲定两晚场地分区方案（主厅 / 露台 / 学术酒吧区）与一滴水档期", "双方共同 + 执行团队", "待办"],
        [2, "确认双主办致辞人与流程分工", "双方共同", "待办"],
        [3, "Day1 邀请 4–6 位跨圈层嘉宾（应用创业 / 出海 / 开源 / 投资 / 模型技术）", "杨浦区科技企业联合会", "待办"],
        [4, "Day2 邀请 5 位跨界圆桌嘉宾（学者 / 大模型 / 智能体 / AI for Science / 安全治理）", "复旦大学（学术资源）", "待办"],
        [5, "设计并印制入场名牌、话题卡、投票牌、主视觉物料", "双方共同 + 传播团队", "待办"],
        [6, "开通报名通道（Day1 开放 + Day2 定向）", "杨浦区科技企业联合会", "待办"],
        [7, "落实主展馆 → 北外滩接驳与交通指引", "执行团队", "待办"],
        [8, "确认餐饮（精酿 / 特调 / 茶歇 / 晚宴式）", "执行团队", "待办"],
        [9, "现场流程手册、主持稿、开放麦 / 圆桌规则", "双方共同", "待办"],
        [10, "两晚数据沉淀（名录 / 项目库 / 议题记录）方案", "双方共同", "待办"],
    ],
    [8, 60, 26, 10], first_col_bold=False,
)

wb.save(OUT)
print(f"Saved {OUT}, sheets: {wb.sheetnames}")
