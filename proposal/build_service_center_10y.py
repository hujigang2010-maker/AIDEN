#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
科创科技服务中心收益 · 2027–2036 十年数字表（给甲方）
三部分累加：①载体可申请 ②活动经费可申请 ③申报奖项可获得
依据 PPT「政策工具箱 / 政策申报能力 / 政策→现金」页口径。
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent
ART = Path("/opt/cursor/artifacts")

# —— 主题色（对齐方案紫金）——
PURPLE = "4A2C7A"
GOLD = "C4A35A"
PURPLE_LIGHT = "F3E9FF"
GOLD_LIGHT = "FFF8E7"
ROW_ALT = "FAF7FF"
GREEN = "1F7A4D"
INK = "2C2C2C"
MUTED = "666666"
WHITE = "FFFFFF"
LINE = "D9D0E8"

# —— 载体 / 活动年上限 ——
CARRIER_PLATFORM = 100  # 成果转化服务平台运营
CARRIER_BASE = 10       # 区级创业孵化基地运营
CARRIER_YEAR = CARRIER_PLATFORM + CARRIER_BASE  # 110
ACTIVITY_YEAR = 200     # YOUNG立方活动封顶

# —— 奖项单价（企业侧可获得，万元）——
UNIT = {
    "高企首认": 20,
    "专精特新(市级)": 10,
    "国家小巨人": 30,
    "区科技小巨人": 50,
    "区双创小巨人": 100,
}

# 分年申报成功家数假设（合理测算，用于甲方沟通；非承诺）
# (高企, 专精特新市级, 国家小巨人, 区科技小巨人, 区双创小巨人)
AWARD_CASES = {
    2027: (3, 2, 0, 0, 0),
    2028: (5, 3, 0, 1, 0),
    2029: (6, 4, 1, 1, 0),
    2030: (7, 5, 1, 1, 1),
    2031: (8, 5, 1, 2, 1),
    2032: (8, 6, 1, 2, 1),
    2033: (9, 6, 2, 2, 1),
    2034: (9, 6, 2, 2, 1),
    2035: (10, 7, 2, 2, 1),
    2036: (10, 7, 2, 2, 1),
}

YEARS = list(range(2027, 2037))


def award_amount(y: int) -> tuple[int, dict]:
    g, z, gj, qx, qc = AWARD_CASES[y]
    detail = {
        "高企首认": g * UNIT["高企首认"],
        "专精特新(市级)": z * UNIT["专精特新(市级)"],
        "国家小巨人": gj * UNIT["国家小巨人"],
        "区科技小巨人": qx * UNIT["区科技小巨人"],
        "区双创小巨人": qc * UNIT["区双创小巨人"],
    }
    return sum(detail.values()), detail, (g, z, gj, qx, qc)


def font(name="微软雅黑", size=11, bold=False, color=INK):
    return Font(name=name, size=size, bold=bold, color=color)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def border():
    s = Side(style="thin", color=LINE)
    return Border(left=s, right=s, top=s, bottom=s)


def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_range(ws, r1, c1, r2, c2, **kwargs):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            if "font" in kwargs:
                cell.font = kwargs["font"]
            if "fill" in kwargs:
                cell.fill = kwargs["fill"]
            if "align" in kwargs:
                cell.alignment = kwargs["align"]
            if "border" in kwargs:
                cell.border = kwargs["border"]


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build() -> Path:
    wb = Workbook()

    # ========== Sheet 1: 十年总表（主交付）==========
    ws = wb.active
    ws.title = "十年收益总表"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.print_title_rows = "1:5"

    # 顶栏
    ws.merge_cells("A1:H1")
    ws["A1"] = "上海创智汇 · 科创科技服务中心收益测算表"
    ws["A1"].font = font(size=18, bold=True, color=PURPLE)
    ws["A1"].alignment = align("left")

    ws.merge_cells("A2:H2")
    ws["A2"] = "2027–2036 年　载体政策 × 活动经费 × 申报奖项　｜　给甲方沟通版"
    ws["A2"].font = font(size=12, bold=True, color=GOLD)

    ws.merge_cells("A3:H3")
    ws["A3"] = (
        "整理自方案 PPT「政策工具箱 / 政策申报能力 / 政策→现金」页。"
        "本表按三部分累加：①载体可申请金额　②活动经费可申请金额　③申报奖项可获得金额（高企 / 专精特新 / 小巨人等）。"
        "单位：人民币万元。表内为可申请/可获得测算口径，非政府批复承诺；奖项家数为运营节奏下的合理假设。"
    )
    ws["A3"].font = font(size=9, color=MUTED)
    ws["A3"].alignment = align("left")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 42

    # KPI 摘要条
    kpi_labels = [
        ("A4", "载体十年合计"),
        ("C4", "活动十年合计"),
        ("E4", "奖项十年合计"),
        ("G4", "三部分十年总计"),
    ]
    for cell_ref, label in kpi_labels:
        ws[cell_ref] = label
        ws[cell_ref].font = font(size=9, color=MUTED)
        ws[cell_ref].alignment = align("left")

    # 表头
    headers = [
        "年份",
        "① 载体可申请金额",
        "② 活动经费可申请金额",
        "③ 申报奖项可获得金额",
        "三部分当年合计",
        "奖项构成（家数×单价）",
        "阶段",
        "备注",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(6, i, h)
        cell.font = font(size=11, bold=True, color=WHITE)
        cell.fill = fill(PURPLE)
        cell.alignment = align()
        cell.border = border()
    ws.row_dimensions[6].height = 36

    stage = {
        2027: "启动导入",
        2028: "成型提质",
        2029: "提升复制",
        2030: "稳定运营",
        2031: "稳定运营",
        2032: "稳定运营",
        2033: "品牌深化",
        2034: "品牌深化",
        2035: "持续放量",
        2036: "持续放量",
    }

    sum_c = sum_a = sum_w = 0
    data_start = 7
    for idx, y in enumerate(YEARS):
        r = data_start + idx
        award_sum, detail, counts = award_amount(y)
        g, z, gj, qx, qc = counts
        parts = []
        if g:
            parts.append(f"高企{g}×20={detail['高企首认']}")
        if z:
            parts.append(f"专精特新{z}×10={detail['专精特新(市级)']}")
        if gj:
            parts.append(f"国家小巨人{gj}×30={detail['国家小巨人']}")
        if qx:
            parts.append(f"区科技小巨人{qx}×50={detail['区科技小巨人']}")
        if qc:
            parts.append(f"区双创小巨人{qc}×100={detail['区双创小巨人']}")
        compose = "；".join(parts) if parts else "—"

        row_total = CARRIER_YEAR + ACTIVITY_YEAR + award_sum
        vals = [
            y,
            CARRIER_YEAR,
            ACTIVITY_YEAR,
            award_sum,
            row_total,
            compose,
            stage[y],
            "须先完成载体/服务中心挂牌认定；奖项以企业获批为准",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = font(size=10, bold=(c == 5), color=INK if c != 5 else PURPLE)
            cell.alignment = align("center" if c < 6 or c == 7 else "left")
            cell.border = border()
            if idx % 2:
                cell.fill = fill(ROW_ALT)
            if c == 5:
                cell.fill = fill(GOLD_LIGHT)
        sum_c += CARRIER_YEAR
        sum_a += ACTIVITY_YEAR
        sum_w += award_sum
        ws.row_dimensions[r].height = 32

    # 合计行
    r = data_start + 10
    totals = ["十年合计", sum_c, sum_a, sum_w, sum_c + sum_a + sum_w, "见「奖项明细」工作表", "—", "上限/测算加总，非保证获批"]
    for c, v in enumerate(totals, 1):
        cell = ws.cell(r, c, v)
        cell.font = font(size=11, bold=True, color=WHITE)
        cell.fill = fill(GOLD if c != 1 else PURPLE)
        if c == 1:
            cell.fill = fill(PURPLE)
        elif c == 5:
            cell.fill = fill("8B6914")
        else:
            cell.fill = fill(GOLD)
        cell.font = font(size=11, bold=True, color=WHITE)
        cell.alignment = align()
        cell.border = border()
    ws.row_dimensions[r].height = 28

    # KPI 数值
    ws["A5"] = sum_c
    ws["C5"] = sum_a
    ws["E5"] = sum_w
    ws["G5"] = sum_c + sum_a + sum_w
    for ref in ("A5", "C5", "E5", "G5"):
        ws[ref].font = font(size=16, bold=True, color=PURPLE if ref != "G5" else "8B6914")
        ws[ref].number_format = '#,##0" 万"'
    ws["A5"].number_format = '#,##0" 万"'
    ws.merge_cells("A5:B5")
    ws.merge_cells("C5:D5")
    ws.merge_cells("E5:F5")
    ws.merge_cells("G5:H5")
    for ref in ("A5", "C5", "E5", "G5"):
        ws[ref].fill = fill(PURPLE_LIGHT)
        ws[ref].alignment = align("left")
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 26

    # 脚注
    ws.merge_cells("A18:H18")
    ws["A18"] = (
        "口径说明：①载体＝成果转化服务平台运营补贴≤100万/年＋创业孵化基地运营≤10万/年；"
        "②活动＝YOUNG立方活动补贴（投入×最高50%，同年封顶≤200万）；"
        "③奖项＝入驻企业获批后可获得的政府奖励/资助金额合计（补贴主体多为企业；服务中心另按服务费2–8万/家及到账5%–15%成功分成取酬，见「平台收益对照」表）。"
        "2030年后假设政策同类续期，以当年有效文件为准。"
    )
    ws["A18"].font = font(size=8.5, color=MUTED)
    ws["A18"].alignment = align("left")
    ws.row_dimensions[18].height = 48

    set_widths(ws, [10, 16, 18, 18, 14, 48, 12, 28])
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = "A6:H16"

    # 图表数据区（隐藏旁侧）— 用主表数据做柱状图
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "2027–2036 三部分累加（万元）"
    chart.y_axis.title = "金额（万元）"
    chart.x_axis.title = None
    chart.style = 10
    data = Reference(ws, min_col=2, min_row=6, max_col=4, max_row=16)
    cats = Reference(ws, min_col=1, min_row=7, max_row=16)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 18
    chart.height = 9
    ws.add_chart(chart, "A20")

    # ========== Sheet 2: 奖项明细 ==========
    ws2 = wb.create_sheet("奖项申报明细")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = "申报奖项可获得金额 · 分年明细（高企 / 专精特新 / 小巨人）"
    ws2["A1"].font = font(size=14, bold=True, color=PURPLE)
    ws2.merge_cells("A1:L1")
    ws2["A2"] = "单价依据 PPT：高企首认20万；专精特新市级10万；国家小巨人30万；区科技小巨人50万；区双创小巨人100万。家数为服务中心陪跑节奏下的合理测算。"
    ws2["A2"].font = font(size=9, color=MUTED)
    ws2.merge_cells("A2:L2")
    ws2.row_dimensions[2].height = 28

    h2 = [
        "年份",
        "高企(家)",
        "高企金额",
        "专精特新(家)",
        "专精特新金额",
        "国家小巨人(家)",
        "国家小巨人金额",
        "区科技小巨人(家)",
        "区科技小巨人金额",
        "区双创小巨人(家)",
        "区双创小巨人金额",
        "奖项合计",
    ]
    for i, h in enumerate(h2, 1):
        cell = ws2.cell(4, i, h)
        cell.font = font(size=9, bold=True, color=WHITE)
        cell.fill = fill(PURPLE)
        cell.alignment = align()
        cell.border = border()
    ws2.row_dimensions[4].height = 34

    for idx, y in enumerate(YEARS):
        r = 5 + idx
        award_sum, detail, counts = award_amount(y)
        g, z, gj, qx, qc = counts
        row = [
            y,
            g,
            detail["高企首认"],
            z,
            detail["专精特新(市级)"],
            gj,
            detail["国家小巨人"],
            qx,
            detail["区科技小巨人"],
            qc,
            detail["区双创小巨人"],
            award_sum,
        ]
        for c, v in enumerate(row, 1):
            cell = ws2.cell(r, c, v)
            cell.font = font(size=10, bold=(c == 12), color=PURPLE if c == 12 else INK)
            cell.alignment = align()
            cell.border = border()
            if idx % 2:
                cell.fill = fill(ROW_ALT)
            if c == 12:
                cell.fill = fill(GOLD_LIGHT)

    r = 15
    # 合计
    totals2 = ["十年合计"]
    col_sums = [0] * 11
    for y in YEARS:
        _, detail, counts = award_amount(y)
        vals = [
            counts[0],
            detail["高企首认"],
            counts[1],
            detail["专精特新(市级)"],
            counts[2],
            detail["国家小巨人"],
            counts[3],
            detail["区科技小巨人"],
            counts[4],
            detail["区双创小巨人"],
            sum(detail.values()),
        ]
        for i, v in enumerate(vals):
            col_sums[i] += v
    for c, v in enumerate(["十年合计"] + col_sums, 1):
        cell = ws2.cell(r, c, v)
        cell.font = font(size=10, bold=True, color=WHITE)
        cell.fill = fill(GOLD)
        cell.alignment = align()
        cell.border = border()

    set_widths(ws2, [9, 10, 10, 12, 12, 13, 13, 14, 14, 14, 14, 11])
    ws2.freeze_panes = "B5"

    # ========== Sheet 3: 平台收益对照 ==========
    ws3 = wb.create_sheet("平台收益对照")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "科创科技服务中心 · 平台可实现现金对照（PPT「政策→现金」口径）"
    ws3["A1"].font = font(size=14, bold=True, color=PURPLE)
    ws3.merge_cells("A1:F1")
    ws3["A2"] = (
        "说明：主表③为「企业侧奖项可获得金额」；本表同步列示服务中心作为申报服务机构可沉淀的平台现金"
        "（申报服务费 + 成功分成示意 + 载体运营直接到平台部分），便于甲方理解「收益」双口径。"
    )
    ws3["A2"].font = font(size=9, color=MUTED)
    ws3.merge_cells("A2:F2")
    ws3.row_dimensions[2].height = 36

    h3 = [
        "年份",
        "载体运营到平台（中值）",
        "奖项申报服务费（中值）",
        "奖项成功分成示意（10%）",
        "平台现金小计",
        "对照：主表三部分合计（企业侧口径）",
    ]
    for i, h in enumerate(h3, 1):
        cell = ws3.cell(4, i, h)
        cell.font = font(size=10, bold=True, color=WHITE)
        cell.fill = fill(PURPLE)
        cell.alignment = align()
        cell.border = border()
    ws3.row_dimensions[4].height = 40

    # 服务费中值：高企/专精特新按 5万/家，科小不单列；成功分成按奖项金额10%
    fee_unit = 5
    for idx, y in enumerate(YEARS):
        r = 5 + idx
        award_sum, _, counts = award_amount(y)
        # 计费家数：高企+专精特新+各类小巨人
        fee_cases = sum(counts)
        carrier_mid = 15 if y == 2027 else 20  # PPT 满年约10–20，成熟取20；上限仍见主表
        if y >= 2029:
            carrier_mid = 20
        fee = fee_cases * fee_unit
        share = round(award_sum * 0.10, 1)
        plat = carrier_mid + fee + share
        main_total = CARRIER_YEAR + ACTIVITY_YEAR + award_sum
        for c, v in enumerate([y, carrier_mid, fee, share, plat, main_total], 1):
            cell = ws3.cell(r, c, v)
            cell.font = font(size=10, bold=(c in (5, 6)), color=PURPLE if c == 5 else INK)
            cell.alignment = align()
            cell.border = border()
            if idx % 2:
                cell.fill = fill(ROW_ALT)
            if c == 5:
                cell.fill = fill(GOLD_LIGHT)

    ws3.merge_cells("A16:F16")
    ws3["A16"] = (
        "平台现金小计＝载体运营到平台中值（10–20万区间取值）＋申报服务费中值（按5万/家×当年奖项家数）"
        "＋成功分成示意（按奖项获批金额×10%，落在PPT的5%–15%区间）。活动经费补贴通常按活动主办/投入主体申报，"
        "主表②已列可申请上限；若由服务中心主体申报，可再计入平台，需以申报主体资格为准。"
    )
    ws3["A16"].font = font(size=8.5, color=MUTED)
    ws3["A16"].alignment = align("left")
    ws3.row_dimensions[16].height = 48
    set_widths(ws3, [10, 18, 18, 18, 14, 28])

    # ========== Sheet 4: 政策依据 ==========
    ws4 = wb.create_sheet("政策依据与口径")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "政策依据与口径说明（对齐 PPT）"
    ws4["A1"].font = font(size=14, bold=True, color=PURPLE)
    ws4.merge_cells("A1:D1")

    rows = [
        ["序号", "类别", "政策要点 / 金额", "对应本表栏目"],
        ["1", "载体", "科技成果转化服务平台运营费补贴，经认定最高100万元/年", "① 载体 · 100万"],
        ["2", "载体", "区级创业孵化基地运营补贴，评估达标10万元/年", "① 载体 · 10万"],
        ["3", "活动", "YOUNG立方活动补贴：按实际投入最高50%，同一主体同年累计最高200万元", "② 活动经费 · 200万"],
        ["4", "奖项", "高新技术企业首次认定奖励：20万元/次（企业）", "③ 高企"],
        ["5", "奖项", "专精特新中小企业（市级）：10万元；国家级小巨人：30万元", "③ 专精特新 / 国家小巨人"],
        ["6", "奖项", "区级科技小巨人50万元；区双创小巨人100万元", "③ 区级小巨人"],
        ["7", "平台收益", "申报服务费：科小0.3–2万/家；高企/专精特新2–8万/家", "见「平台收益对照」"],
        ["8", "平台收益", "成功分成：补贴到账额5%–15%；载体运营满年约10–20万直接到平台", "见「平台收益对照」"],
        ["9", "另计", "服务业引导资金：重点≤总投入20%且≤300万（项目制，不按年重复列入主表）", "另案申报"],
        ["10", "前提", "挂牌杨浦科技企业服务中心；取得经认定载体/成果转化平台；入驻创新券服务机构平台", "兑现前置条件"],
    ]
    for i, row in enumerate(rows, 3):
        for j, v in enumerate(row, 1):
            cell = ws4.cell(i, j, v)
            if i == 3:
                cell.font = font(size=10, bold=True, color=WHITE)
                cell.fill = fill(PURPLE)
            else:
                cell.font = font(size=10)
                if i % 2 == 0:
                    cell.fill = fill(ROW_ALT)
            cell.alignment = align("center" if j in (1, 2, 4) else "left")
            cell.border = border()
        ws4.row_dimensions[i].height = 28
    set_widths(ws4, [8, 12, 62, 22])

    ws4["A15"] = "使用建议（给甲方）"
    ws4["A15"].font = font(size=12, bold=True, color=PURPLE)
    ws4.merge_cells("A16:D16")
    ws4["A16"] = (
        "1）对外沟通优先使用「十年收益总表」：三列金额分列清晰，最右为累加合计；\n"
        "2）奖项家数可按实际入驻与培育进度在「奖项申报明细」中调整后重算；\n"
        "3）若需强调服务中心自身进账，请同步出示「平台收益对照」表；\n"
        "4）所有数字为政策上限或合理测算，最终以当年申报指南、认定结果及拨付文件为准。"
    )
    ws4["A16"].font = font(size=10, color=INK)
    ws4["A16"].alignment = align("left")
    ws4.row_dimensions[16].height = 72

    # 保存
    ART.mkdir(parents=True, exist_ok=True)
    name_cn = "创智汇科创服务中心收益十年表-给甲方.xlsx"
    name_en = "chuangzhihui-service-center-income-2027-2036.xlsx"
    for base in (OUT, ART):
        wb.save(base / name_cn)
        wb.save(base / name_en)

    print(f"十年总计: 载体{sum_c} + 活动{sum_a} + 奖项{sum_w} = {sum_c+sum_a+sum_w}")
    for y in YEARS:
        a, _, _ = award_amount(y)
        print(f"  {y}: 载体{CARRIER_YEAR} + 活动{ACTIVITY_YEAR} + 奖项{a} = {CARRIER_YEAR+ACTIVITY_YEAR+a}")
    return OUT / name_cn


if __name__ == "__main__":
    build()
