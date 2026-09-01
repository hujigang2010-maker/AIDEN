#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成《创智汇 · 载体/活动政策 2027–2036 十年可申请金额表》."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent

# 政策口径（现行文件；2030 年后按「续期同类」假设）
# 载体：成果转化服务平台运营 ≤100万/年 + 区级创业孵化基地运营 10万/年（评估达标）
# 活动：YOUNG立方活动补贴 ≤实际投入50%，同一主体同年累计 ≤200万
CARRIER_PLATFORM = 100  # 成果转化平台运营
CARRIER_INCUBATOR = 10   # 孵化基地运营
ACTIVITY_CAP = 200       # YOUNG立方活动封顶

# 分年：认定节奏下的「建议申报/预期可实现」口径（非法律承诺）
# 载体建议：认定推进→满年→稳定；活动建议：按当年活动投入×50% 估算（封顶200）
SUGGEST = {
    # year: (载体建议中值或区间用中值, 活动投入假设, 活动建议=min(投入*0.5,200))
    2027: (20, 60, "认定推进/首年满年KPI；活动起步（沙龙+发布）"),
    2028: (80, 120, "成果转化平台认定落地；活动放量（含节展）"),
    2029: (100, 200, "双资质稳定；活动投入抬升，逼近政策封顶"),
    2030: (110, 280, "平台+孵化基地兼得；旗舰活动/漫展/峰会"),
    2031: (110, 300, "稳定申报；活动维持高位"),
    2032: (110, 320, "稳定申报；活动维持高位"),
    2033: (110, 350, "稳定申报；可叠加特别有影响力活动专项（另议）"),
    2034: (110, 350, "稳定申报（假设政策同类续期）"),
    2035: (110, 380, "稳定申报（假设政策同类续期）"),
    2036: (110, 400, "稳定申报；活动投入达封顶对应门槛"),
}


def style_header(cell):
    cell.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4A2C7A")
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


def style_cell(cell, bold=False):
    cell.font = Font(name="微软雅黑", size=10, bold=bold)
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")


def thin_border(ws, rows, cols):
    side = Side(style="thin", color="CCCCCC")
    border = Border(left=side, right=side, top=side, bottom=side)
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(r, c).border = border


def build() -> Path:
    wb = Workbook()

    # —— Sheet1：可申请上限（十年）——
    ws = wb.active
    ws.title = "01 可申请上限"
    ws["A1"] = "创智汇 · 载体政策 / 活动政策可申请金额表（2027–2036）"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="4A2C7A")
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "口径：列为政策「可申请上限」。载体＝成果转化服务平台运营补贴≤100万/年"
        "＋区级创业孵化基地运营补贴≤10万/年（评估达标、兼得时合计≤110万/年）；"
        "活动＝YOUNG立方活动补贴（按实际投入最高50%，同一主体同一年度累计≤200万）。"
        "现行文件有效期届满后，表内2030–2036按「同类政策续期」假设列示，以当年有效文件为准。"
    )
    ws["A2"].font = Font(name="微软雅黑", size=9, color="555555")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 48

    headers = [
        "年份",
        "载体政策可申请上限（万元）",
        "其中：成果转化平台运营（万元）",
        "其中：创业孵化基地运营（万元）",
        "活动政策可申请上限（万元）",
        "载体+活动合计上限（万元）",
        "备注",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        style_header(cell)
    ws.row_dimensions[3].height = 36

    years = list(range(2027, 2037))
    total_c = total_a = 0
    for idx, y in enumerate(years):
        r = 4 + idx
        # 2027：认定推进年——载体按「可启动申报」仍列政策上限，备注说明前提
        c_plat = CARRIER_PLATFORM
        c_inc = CARRIER_INCUBATOR
        c_sum = c_plat + c_inc
        a_cap = ACTIVITY_CAP
        note = "须先取得相应认定；活动按投入×50%计，封顶200"
        if y == 2027:
            note = "认定推进年：载体以上限列示，实际以当年是否完成认定为前提；活动可随落地节展申报"
        elif y >= 2030:
            note = "假设政策同类续期；以当年有效文件及申报指南为准"
        ws.cell(r, 1, y)
        ws.cell(r, 2, c_sum)
        ws.cell(r, 3, c_plat)
        ws.cell(r, 4, c_inc)
        ws.cell(r, 5, a_cap)
        ws.cell(r, 6, c_sum + a_cap)
        ws.cell(r, 7, note)
        for c in range(1, 8):
            style_cell(ws.cell(r, c))
        total_c += c_sum
        total_a += a_cap

    r = 14
    ws.cell(r, 1, "十年合计")
    ws.cell(r, 2, total_c)
    ws.cell(r, 3, CARRIER_PLATFORM * 10)
    ws.cell(r, 4, CARRIER_INCUBATOR * 10)
    ws.cell(r, 5, total_a)
    ws.cell(r, 6, total_c + total_a)
    ws.cell(r, 7, "上限加总；非保证获批金额")
    for c in range(1, 8):
        style_cell(ws.cell(r, c), bold=True)
        ws.cell(r, c).fill = PatternFill("solid", fgColor="F3E9FF")

    ws.row_dimensions[1].height = 24
    widths = [10, 18, 18, 18, 18, 16, 42]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    thin_border(ws, 14, 7)

    # —— Sheet2：建议申报（结合投入）——
    ws2 = wb.create_sheet("02 建议申报目标")
    ws2["A1"] = "分年建议申报 / 预期可实现口径（结合认定节奏与活动投入）"
    ws2["A1"].font = Font(name="微软雅黑", bold=True, size=13, color="4A2C7A")
    ws2.merge_cells("A1:H1")
    ws2["A2"] = (
        "说明：建议申报≠获批保证。载体建议取「认定后可实现」区间中值靠上限；"
        "活动建议＝min(当年活动投入假设×50%, 200)。本项目基线活动执行约30万/年，"
        "若仅按基线投入，活动可申约15万；下表按「做大节展/峰会」抬升投入测算可申空间。"
    )
    ws2["A2"].font = Font(name="微软雅黑", size=9, color="555555")
    ws2["A2"].alignment = Alignment(wrap_text=True)
    ws2.merge_cells("A2:H2")
    ws2.row_dimensions[2].height = 42

    h2 = [
        "年份",
        "载体建议申报（万元）",
        "活动投入假设（万元）",
        "活动建议申报（万元）",
        "当年建议合计（万元）",
        "对照：载体上限（万元）",
        "对照：活动上限（万元）",
        "节奏说明",
    ]
    for i, h in enumerate(h2, 1):
        style_header(ws2.cell(3, i, h))
    ws2.row_dimensions[3].height = 36

    sum_c = sum_a = 0
    for idx, y in enumerate(years):
        r = 4 + idx
        c_sug, invest, note = SUGGEST[y]
        a_sug = min(int(invest * 0.5), ACTIVITY_CAP)
        ws2.cell(r, 1, y)
        ws2.cell(r, 2, c_sug)
        ws2.cell(r, 3, invest)
        ws2.cell(r, 4, a_sug)
        ws2.cell(r, 5, c_sug + a_sug)
        ws2.cell(r, 6, CARRIER_PLATFORM + CARRIER_INCUBATOR)
        ws2.cell(r, 7, ACTIVITY_CAP)
        ws2.cell(r, 8, note)
        for c in range(1, 9):
            style_cell(ws2.cell(r, c))
        sum_c += c_sug
        sum_a += a_sug

    r = 14
    ws2.cell(r, 1, "十年合计")
    ws2.cell(r, 2, sum_c)
    ws2.cell(r, 3, "—")
    ws2.cell(r, 4, sum_a)
    ws2.cell(r, 5, sum_c + sum_a)
    ws2.cell(r, 6, (CARRIER_PLATFORM + CARRIER_INCUBATOR) * 10)
    ws2.cell(r, 7, ACTIVITY_CAP * 10)
    ws2.cell(r, 8, "建议口径合计")
    for c in range(1, 9):
        style_cell(ws2.cell(r, c), bold=True)
        ws2.cell(r, c).fill = PatternFill("solid", fgColor="F3E9FF")

    for i, w in enumerate([10, 16, 16, 16, 14, 14, 14, 40], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    thin_border(ws2, 14, 8)

    # —— Sheet3：政策依据 ——
    ws3 = wb.create_sheet("03 政策依据")
    ws3["A1"] = "政策依据与另计项目"
    ws3["A1"].font = Font(name="微软雅黑", bold=True, size=13, color="4A2C7A")
    ws3.merge_cells("A1:D1")
    rows = [
        ["类别", "政策要点", "可申请上限", "文件/说明"],
        [
            "载体",
            "科技成果转化服务平台运营费补贴（经认定）",
            "最高100万元/年",
            "《杨浦区加快科技服务业高质量发展若干措施》等",
        ],
        [
            "载体",
            "区级创业孵化基地运营补贴（评估达标）",
            "10万元/年（一次性评估口径）",
            "《杨浦区关于做好促进就业创业工作的实施意见》等",
        ],
        [
            "活动",
            "YOUNG立方·活动补贴（推动集聚区建设的活动/节展）",
            "按实际投入最高50%，同一主体同年累计最高200万元",
            "《杨浦区打造互联网优质内容创作集聚区支持政策》",
        ],
        [
            "另计·项目制",
            "服务业引导资金（载体改造/服务业项目）",
            "重点≤总投入20%且≤300万；一般≤15%且≤200万",
            "不按年重复列入本十年表；达标投资额另案申报",
        ],
        [
            "另计·企业侧",
            "入驻企业高企/专精特新/房租/创新券等",
            "按企业各自条件",
            "补贴主体为企业；平台按服务费+成功分成取酬，不计入本表载体/活动栏",
        ],
        [
            "前提",
            "载体认定 / 挂牌 / 创新券服务机构等",
            "—",
            "未取得认定前，载体运营类上限不可兑现；活动须符合 YOUNG立方认定条件",
        ],
    ]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            cell = ws3.cell(i, j, v)
            if i == 2:
                style_header(cell)
            else:
                style_cell(cell)
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    for i, w in enumerate([12, 42, 28, 42], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for r in range(3, 9):
        ws3.row_dimensions[r].height = 36
    thin_border(ws3, 8, 4)

    out_cn = OUT / "创智汇载体活动政策十年申请表.xlsx"
    out_en = OUT / "chuangzhihui-carrier-activity-policy-2027-2036.xlsx"
    wb.save(out_cn)
    wb.save(out_en)

    # Markdown
    md_lines = [
        "# 创智汇 · 载体 / 活动政策可申请金额表（2027–2036）",
        "",
        "## 一、可申请上限（政策天花板）",
        "",
        "| 年份 | 载体政策可申请上限（万元） | 其中：成果转化平台运营 | 其中：创业孵化基地运营 | 活动政策可申请上限（万元） | 载体+活动合计上限（万元） |",
        "|------|---------------------------|------------------------|------------------------|---------------------------|---------------------------|",
    ]
    for y in years:
        md_lines.append(
            f"| {y} | {CARRIER_PLATFORM + CARRIER_INCUBATOR} | {CARRIER_PLATFORM} | {CARRIER_INCUBATOR} | {ACTIVITY_CAP} | {CARRIER_PLATFORM + CARRIER_INCUBATOR + ACTIVITY_CAP} |"
        )
    md_lines.append(
        f"| **十年合计** | **{total_c}** | **{CARRIER_PLATFORM * 10}** | **{CARRIER_INCUBATOR * 10}** | **{total_a}** | **{total_c + total_a}** |"
    )
    md_lines += [
        "",
        "**口径说明**",
        "",
        "- **载体**：成果转化服务平台运营补贴最高 **100 万/年** + 区级创业孵化基地运营补贴 **10 万/年**（评估达标；兼得时合计 **110 万/年**）。",
        "- **活动**：YOUNG立方活动补贴按活动实际投入最高 **50%**，同一主体同一年度累计最高 **200 万**。",
        "- 2030–2036 按「同类政策续期」假设列示，**以当年有效文件为准**；未完成载体认定前，载体栏不可兑现。",
        "- **另计（不列入年表）**：服务业引导资金为项目制（重点 ≤300 万），不按年重复累计。",
        "",
        "## 二、分年建议申报目标（结合认定节奏与活动投入）",
        "",
        "| 年份 | 载体建议申报（万元） | 活动投入假设（万元） | 活动建议申报（万元） | 当年建议合计（万元） | 节奏说明 |",
        "|------|----------------------|----------------------|----------------------|----------------------|----------|",
    ]
    for y in years:
        c_sug, invest, note = SUGGEST[y]
        a_sug = min(int(invest * 0.5), ACTIVITY_CAP)
        md_lines.append(
            f"| {y} | {c_sug} | {invest} | {a_sug} | {c_sug + a_sug} | {note} |"
        )
    md_lines += [
        f"| **十年合计** | **{sum_c}** | — | **{sum_a}** | **{sum_c + sum_a}** | 建议口径合计 |",
        "",
        "> 若仅维持本项目基线活动执行费约 **30 万/年**，则活动可申约为 **15 万/年**（30×50%），远低于 200 万封顶；上表活动栏按「做大节展」抬升投入测算可申空间。",
        "",
    ]
    md_path = OUT / "创智汇载体活动政策十年申请表.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")
    print(f"Wrote {out_cn}")
    print(f"Wrote {md_path}")
    return out_cn


if __name__ == "__main__":
    build()
