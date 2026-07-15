# -*- coding: utf-8 -*-
"""生成《源信网络算力补贴合作方案》配套表格 (xlsx)。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "0B2A5B"
BLUE = "125CC4"
CYAN = "00B4D8"
LIGHT = "EAF2FB"
LIGHT2 = "DBE9FA"
GOLD = "F2A02D"
WHITE = "FFFFFF"
INK = "1B2433"

thin = Side(style="thin", color="B9CBE6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F = "Microsoft YaHei"

wb = Workbook()


def style_title(ws, text, ncols, color=NAVY):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=F, size=15, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34


def header_row(ws, row, headers, color=BLUE):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=F, size=11.5, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 30


def data_row(ws, row, values, bold_first=True, zebra=True):
    fill = LIGHT if (zebra and row % 2 == 0) else WHITE
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=j, value=v)
        is_first = (j == 1 and bold_first)
        c.font = Font(name=F, size=11, bold=is_first, color=NAVY if is_first else INK)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center" if j > 1 else "left",
                                vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 26


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def note(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=F, size=10, italic=True, color="5A6472")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 40


# ============== Sheet 1: 准入与申请条件 ==============
ws = wb.active
ws.title = "准入与申请条件"
set_widths(ws, [16, 40, 30])
style_title(ws, "一、准入与申请条件", 3)
header_row(ws, 2, ["条件类别", "申请门槛 / 标准", "说明"])
rows1 = [
    ["入驻载体", "甲级写字楼 / 重点产业园 / 科技载体", "纳入园区·楼宇白名单"],
    ["企业属性", "注册及纳税在杨浦区；科技型、AI / 数字经济企业优先", "面向 B 端企业客户"],
    ["租赁规模", "租赁面积 ≥ 200㎡ 或 ≥ 10 个工位", "高端办公场所门槛"],
    ["合作期限", "签约期 ≥ 12 个月", "并承诺使用大厂云算力服务"],
    ["配合事项", "配合补贴核销、案例展示、政策研究数据采集", "便于成效评估"],
    ["无门槛福利", "免费领取 5 万元消费券（无申请门槛）", "园区/楼宇内企业均可领取"],
]
r = 3
for v in rows1:
    data_row(ws, r, v); r += 1
note(ws, r, 3, "说明：高端办公场所满足上述门槛即可申请分档补贴；5 万元消费券为无门槛福利，可直接领取用于抵扣大厂云产品与算力消费。")

# ============== Sheet 2: 补贴标准表 ==============
ws2 = wb.create_sheet("补贴标准表")
set_widths(ws2, [16, 18, 14, 12, 14, 12])
style_title(ws2, "二、补贴标准一览（按企业规模分档 · 建议）", 6)
header_row(ws2, 2, ["企业规模（人）", "算力补贴额度(元/年)", "免费消费券",
                    "租金减免方案", "云/运营消费补贴", "大额签约折扣"])
rows2 = [
    ["0–50（含）", "5 万 token券", "5 万元", "免1补1", "8.5 折", "9.5 折"],
    ["50–100（含）", "15 万", "5 万元", "免2补1", "8.0 折", "9.0 折"],
    ["100–300（含）", "30 万", "5 万元", "免2补2", "7.5 折", "8.5 折"],
    ["300–500（含）", "60 万", "5 万元", "免3补2", "7.0 折", "8.5 折"],
    ["> 500", "100 万 + 定制", "5 万元", "免3补3", "6.5 折", "8.0 折"],
]
r = 3
for v in rows2:
    data_row(ws2, r, v); r += 1
note(ws2, r, 6, "说明：以上为建议方案，最终额度以大厂（火山引擎/腾讯云）补贴政策及三方协议为准；“免X补X”指首年租金由载体减免与补贴池共担的月数。")

# ============== Sheet 2b: 火山引擎园区独立政策 ==============
wsv = wb.create_sheet("火山引擎园区政策")
set_widths(wsv, [22, 30, 28])
style_title(wsv, "二·补：火山引擎园区独立政策（已沟通确认）", 3)
# 一、半年免费
header_row(wsv, 2, ["政策一", "内容", "说明"])
data_row(wsv, 3, ["无门槛·半年费用免费", "按预估半年费用一次性发放代金券", "园区企业 0 门槛即可享用"])
# 二、大客户额外折扣
wsv.cell(row=5, column=1, value="政策二：大客户额外折扣（除代金券外，可叠加）").font = Font(name=F, size=12, bold=True, color=NAVY)
wsv.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)
wsv.row_dimensions[5].height = 26
header_row(wsv, 6, ["累计消费（万元）", "额外折扣（除代金券外）", "备注"])
disc_rows = [
    ["0 – 10", "5 折 ～ 7 折", ""],
    ["10 – 30", "4.5 折 ～ 5 折", ""],
    ["30 – 50", "4 折 ～ 4.5 折", "折扣随累计"],
    ["50 – 100", "3.5 折 ～ 4 折", "消费提升而"],
    ["100 – 300", "3 折 ～ 3.5 折", "走低，越用"],
    ["300 – 500 +", "2.5 折 ～ 3 折", "越优惠"],
]
r = 7
for v in disc_rows:
    data_row(wsv, r, v, zebra=True); r += 1
note(wsv, r, 3, "说明：半年免费代金券与大客户额外折扣可叠加享受；具体折扣与额度以火山引擎最终政策为准。")
wsv.freeze_panes = "A3"

# ============== Sheet 3: 补贴与优惠方式 ==============
ws3 = wb.create_sheet("补贴与优惠方式")
set_widths(ws3, [18, 26, 40])
style_title(ws3, "三、补贴与优惠方式（三大支柱）", 3)
header_row(ws3, 2, ["支柱", "方式", "具体内容"])
rows3 = [
    ["租金减免", "免几个月 + 补几个月", "按合作规模采取“免X月+补X月”折扣，由载体让利与补贴池共担，签约越长、规模越大减免越多。"],
    ["算力补贴", "大厂算力直接补贴", "发放 token 算力券 / 代金额度，对接火山引擎、腾讯云补贴池，按企业规模分档授信。"],
    ["运营与产品消费补贴", "云服务 + 产品 + 运营", "云产品消费折扣与返券、大厂产品采购专项补贴、运营/培训/上云服务补贴。"],
]
r = 3
for v in rows3:
    data_row(ws3, r, v, zebra=True); r += 1

# ============== Sheet 4: 各方职责分工 ==============
ws4 = wb.create_sheet("各方职责分工")
set_widths(ws4, [26, 18, 44])
style_title(ws4, "四、各方角色与职责分工", 3)
header_row(ws4, 2, ["合作方", "定位", "主要职责"])
rows4 = [
    ["复旦大学住房政策研究中心", "政策智库·研究背书", "政策设计与合规研究、补贴成效评估与课题、产学研成果转化与背书。"],
    ["杨浦区科技企业联合会", "企业资源·组织协调", "对接区内企业与载体、组织申报与政策宣贯、汇集企业算力需求。"],
    ["源信网络", "运营落地·资源对接", "对接大厂算力补贴资源、统筹楼宇/园区合作、补贴发放与运营服务。"],
    ["大厂（火山引擎/腾讯云）", "算力与补贴提供方", "提供 token 算力补贴与生态资源，确认补贴政策与额度池。"],
    ["物业 / 楼宇 / 园区", "落地渠道载体", "承接并落地政策、提供租金减免、协助企业申报与核销。"],
]
r = 3
for v in rows4:
    data_row(ws4, r, v, zebra=True); r += 1

# ============== Sheet 5: 落地路径 ==============
ws5 = wb.create_sheet("落地路径")
set_widths(ws5, [14, 22, 46])
style_title(ws5, "五、落地路径与下一步", 3)
header_row(ws5, 2, ["阶段", "目标", "关键动作"])
rows5 = [
    ["第一阶段·试点", "跑通最小闭环", "选取 1–2 个标杆楼宇/园区，确定补贴池与白名单，落地首批企业。"],
    ["第二阶段·推广", "规模化复制", "总结试点成效，由联合会组织区内载体规模化申报与宣贯。"],
    ["第三阶段·复制", "标准化输出", "形成标准化合作模板与政策研究报告，向全区及更大范围复制。"],
    ["下一步行动", "立即推进", "①三方明确分工与补贴池规模 ②对接大厂确认算力补贴政策 ③确定首批试点载体与企业名单。"],
]
r = 3
for v in rows5:
    data_row(ws5, r, v, zebra=True); r += 1

# 冻结表头
for ws_ in [ws, ws2, ws3, ws4, ws5]:
    ws_.freeze_panes = "A3"

# ============== Sheet 6: 开放式可复制模式 ==============
ws6 = wb.create_sheet("开放式可复制模式")
set_widths(ws6, [16, 22, 46])
style_title(ws6, "六、开放式 · 可复制合作模式", 3)
header_row(ws6, 2, ["要素", "关键词", "说明"])
rows6 = [
    ["标准化模板", "开箱即用", "合作协议、准入条件、补贴标准、核销流程全部模板化，可直接套用。"],
    ["模块化组合", "自由拼装", "算力补贴 / 租金减免 / 消费补贴按载体需求自由组合。"],
    ["开放式接入", "动态扩容", "对大厂、物业、园区、企业开放，白名单动态扩容、平台化运营。"],
    ["可复制推广", "边际递减", "一套打法复制到多楼宇、多园区乃至跨区域，边际成本递减。"],
]
r = 3
for v in rows6:
    data_row(ws6, r, v, zebra=True); r += 1
note(ws6, r, 3, "目标：形成“一次设计、处处可用”的开放式算力普惠样板，可持续扩张为区域算力普惠生态。")
ws6.freeze_panes = "A3"

out = "源信网络算力补贴三方合作方案_配套表格.xlsx"
wb.save(out)
print("saved", out)
