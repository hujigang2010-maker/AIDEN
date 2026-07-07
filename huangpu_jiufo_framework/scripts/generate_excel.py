# -*- coding: utf-8 -*-
"""生成 Excel:项目概要与分工 / 服务框架(工作分解) / 报价明细。16万报价。"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import content_framework as C

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "deliverables")
FONT = "微软雅黑"
NAVY = "0F2A4A"
BLUE = "1F4E79"
GOLD = "C79A3B"
LIGHT = "EEF2F8"
LIGHTGOLD = "F6EEDA"
WHITE = "FFFFFF"
DARK = "222B35"

med = Side(style="thin", color="AAB4C0")
BORDER = Border(left=med, right=med, top=med, bottom=med)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def setup(ws, tab=GOLD):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.sheet_properties.tabColor = tab


def widths(ws, ws_widths):
    for j, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def banner(ws, ncols, title, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=15, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 34
    r = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(row=2, column=1, value=subtitle)
        c.font = Font(name=FONT, size=10, italic=True, color="5A5A5A")
        c.fill = PatternFill("solid", fgColor=LIGHTGOLD)
        c.alignment = CENTER
        ws.row_dimensions[2].height = 20
        r = 3
    return r


def header(ws, row, headers, fill=GOLD):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=FONT, size=10.5, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def drow(ws, row, values, center_cols=(), h=22, zebra=True, fill_hex=None):
    fill = None
    if fill_hex:
        fill = PatternFill("solid", fgColor=fill_hex)
    elif zebra and row % 2 == 0:
        fill = PatternFill("solid", fgColor=LIGHT)
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(name=FONT, size=10, color=DARK)
        c.alignment = CENTER if j in center_cols else LEFT
        c.border = BORDER
        if fill:
            c.fill = fill
    ws.row_dimensions[row].height = h


def total_row(ws, row, values, center_cols=()):
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        c.fill = PatternFill("solid", fgColor=LIGHTGOLD)
        c.alignment = CENTER if j in center_cols else LEFT
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def kv(ws, row, label, value, ncols):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name=FONT, size=10.5, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = CENTER
    c.border = BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    v = ws.cell(row=row, column=2, value=value)
    v.font = Font(name=FONT, size=10, color=DARK)
    v.alignment = LEFT
    for col in range(2, ncols + 1):
        ws.cell(row=row, column=col).border = BORDER


def note(ws, row, ncols, text, h=32):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=9.5, italic=True, color="666666")
    c.alignment = LEFT
    ws.row_dimensions[row].height = h


# --------------------------------------------------- 1 项目概要与分工
def sheet_overview(wb):
    ws = wb.active
    ws.title = "项目概要与分工"
    setup(ws)
    NC = 4
    widths(ws, [18, 30, 30, 22])
    r = banner(ws, NC, C.PROJECT_NAME + " · " + C.SUBTITLE,
               "提供方:" + C.PROVIDER_LINE + "　|　日期:" + C.DOC_DATE)
    for label, value in [
        ("项目名称", C.PROJECT_NAME),
        ("提供方（我方）", C.PROVIDER_LINE),
        ("第四部分承接方", C.PARTNER + "（我方不负责）"),
        ("提交对象", C.CLIENT),
        ("框架与分工", C.PREFACE),
        ("我方报价合计", f"{C.OUR_TOTAL} 万元（含税参考价，仅含第一至第三部分）"),
    ]:
        kv(ws, r, label, value, NC)
        nlines = len(str(value)) // 46 + 1
        ws.row_dimensions[r].height = max(24, min(nlines * 16 + 8, 150))
        r += 1


# --------------------------------------------------- 2 服务框架(工作分解)
def sheet_framework(wb):
    ws = wb.create_sheet("服务框架（工作分解）")
    setup(ws)
    widths(ws, [22, 44, 16, 10])
    r = banner(ws, 4, "策划服务框架 · 工作分解与分工")
    header(ws, r, ["服务部分", "工作项（子模块）", "负责方", "报价(万元)"])
    r += 1
    for (no, name, owner, fee, desc, subs) in C.PARTS:
        fee_txt = fee if fee is not None else "—"
        is_partner = fee is None
        fill = "F3E9E9" if is_partner else None
        n = len(subs)
        start = r
        for k, s in enumerate(subs):
            part_cell = f"第{no}部分  {name}" if k == 0 else ""
            drow(ws, r, [part_cell, s, owner if k == 0 else "", fee_txt if k == 0 else ""],
                 center_cols=(3, 4), h=24, zebra=False, fill_hex=fill)
            r += 1
        # 合并该部分的“服务部分/负责方/报价”单元格
        if n > 1:
            ws.merge_cells(start_row=start, start_column=1, end_row=r - 1, end_column=1)
            ws.merge_cells(start_row=start, start_column=3, end_row=r - 1, end_column=3)
            ws.merge_cells(start_row=start, start_column=4, end_row=r - 1, end_column=4)
    total_row(ws, r, ["我方合计（第一至第三部分）", "", "我方", C.OUR_TOTAL],
              center_cols=(3, 4))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)


# --------------------------------------------------- 3 报价明细
def sheet_quote(wb):
    ws = wb.create_sheet("报价明细")
    setup(ws)
    widths(ws, [20, 50, 34, 16, 12])
    r = banner(ws, 5, "策划服务报价明细",
               "含税参考价(人民币);仅第一至第三部分由我方报价,合计16万元")
    header(ws, r, ["服务部分", "主要工作内容", "主要成果", "负责方", "报价(万元)"])
    r += 1
    outputs = {
        "一": "《市场与内部深度洞察报告》",
        "二": "《总体战略与故事线策划》",
        "三": "《产业定位与招商建议书》",
        "四": "场景营造成果（由内里集交付）",
    }
    for (no, name, owner, fee, desc, subs) in C.PARTS:
        fee_txt = fee if fee is not None else "—"
        fill = "F3E9E9" if fee is None else None
        drow(ws, r, [f"第{no}部分  {name}", desc, outputs[no], owner, fee_txt],
             center_cols=(4, 5), h=46, zebra=(fill is None), fill_hex=fill)
        r += 1
    total_row(ws, r, ["我方合计", "第一至第三部分(我方负责范围)", "", "我方", C.OUR_TOTAL],
              center_cols=(4, 5))
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    r += 2
    for i, nt in enumerate(C.QUOTATION_NOTES):
        note(ws, r, 5, f"{i+1}. {nt}", h=22)
        r += 1


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_framework(wb)
    sheet_quote(wb)
    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, "广州黄埔九佛TOD全球自贸区_服务框架与报价.xlsx")
    wb.save(out)
    print("saved:", out)


if __name__ == "__main__":
    main()
