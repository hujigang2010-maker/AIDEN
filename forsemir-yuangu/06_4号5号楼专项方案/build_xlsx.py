"""4#+5# 楼 2 万方专项 — 合作收益测算表 (Excel).

Sheets:
  01 封面与说明
  02 对甲方贡献 (24 月 + 永续年金)
  03 基础月费构成 (1-3 人配置阶梯)
  04 招商佣金阶梯 (1.5-2 个月年租金)
  05 挂牌奖励明细 (5 项)
  06 6 场沙龙测算
  07 24 月双向账本 + ROI
  08 招商客户管道 (示例脱敏)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("4#5#楼合作收益测算表.xlsx")

PRIMARY = "142C5E"
ACCENT = "F27E2D"
GOLD = "C8993D"
LIGHT = "EAEEF5"
ALT = "F8F9FB"

THIN = Side(style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1 = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
H2 = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
H3 = Font(name="微软雅黑", size=11, bold=True, color=PRIMARY)
BODY = Font(name="微软雅黑", size=11, color="1F2A44")
NOTE = Font(name="微软雅黑", size=10, italic=True, color="55607A")
NUM = Font(name="Consolas", size=11, color="1F2A44")

FILL_PRIMARY = PatternFill("solid", fgColor=PRIMARY)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_ALT = PatternFill("solid", fgColor=ALT)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row, headers, fill=None):
    fill = fill or FILL_PRIMARY
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = H2; c.fill = fill; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 28


def write_row(ws, row, vals, num_cols=None, bold_total=False, fill=None):
    num_cols = num_cols or []
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.border = BORDER
        if j in num_cols:
            c.number_format = "#,##0"
            c.alignment = RIGHT
            c.font = NUM
        else:
            c.alignment = LEFT
            c.font = BODY
        if bold_total:
            c.fill = fill or FILL_ACCENT
            c.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        else:
            c.fill = FILL_ALT if row % 2 == 0 else FILL_LIGHT


def title_bar(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = H1; c.fill = FILL_PRIMARY
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40


def main():
    wb = Workbook()

    # ============= Sheet 1 封面 =============
    ws = wb.active
    ws.title = "01 封面与说明"
    title_bar(ws, "元谷项目 4#+5# 楼约 2 万方 · 招商运营合作收益测算表 v1.0", 8)
    set_widths(ws, [4, 22, 22, 22, 22, 22, 22, 22])

    rows = [
        ("", ""),
        ("文件版本", "v1.0 (专项报告专用)"),
        ("业务范围", "元谷 4# 楼 5F+ 潮玩产业集群 + 5# 楼 5F+ 潮玩产业集群, 共约 2 万㎡"),
        ("服务期限", "24 个月 (首期), 满租后可续约"),
        ("货币单位", "人民币 元 (除非另行注明)"),
        ("Sheet 索引", "02 对甲方贡献 / 03 月费构成 / 04 佣金阶梯 / 05 挂牌 / 06 沙龙 / 07 双向账本+ROI / 08 招商客户管道"),
        ("", ""),
        ("【商业条款简表】", ""),
        ("基础月费 (推荐)",   "12 万元/月 × 24 月 = 288 万元 (2 人配置)"),
        ("招商佣金",        "实际成交年租金的 1.5 / 1.75 / 2.0 个月 (按面积阶梯)"),
        ("挂牌奖励",        "5 项 × 30 万 = 150 万元一次性"),
        ("沙龙执行费",       "6 场 × 5 万 = 30 万元基础包 + 净利 30/70 分润"),
        ("满租推算 (甲方)",  "2 万㎡ × 365 × 2.2 ≈ 1,606 万元/年 (永续)"),
        ("", ""),
        ("【5 项挂牌】", ""),
        ("①", "AI 潮玩产业基地 (中国动漫集团)"),
        ("②", "潮玩次元商业专委会 (中国百货商业协会)"),
        ("③", "复旦大学住房政策研究中心 · 元谷分中心"),
        ("④", "上海市科技企业联合会 · 元谷产业基地"),
        ("⑤", "福布斯产业影响力奖 · 元谷专场"),
        ("", ""),
        ("【6 场产业沙龙】(每场 ≥ 30 个目标产业客户)", ""),
        ("#1 T+1 月", "AI + 潮玩 (借势 5/22 峰会)"),
        ("#2 T+3 月", "潮玩出海 (北欧 / 日韩 / 东南亚)"),
        ("#3 T+5 月", "投融资路演 (硬科技 + 潮玩)"),
        ("#4 T+7 月", "设计与创意 (上海交大 + 上海市科企联)"),
        ("#5 T+9 月", "内容 IP 与 Z 世代 (中百协 + 中动漫)"),
        ("#6 T+11 月", "政策补贴 (闵行科协 + 复旦住房政策中心)"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        if not k and not v:
            continue
        a = ws.cell(row=i, column=2, value=k); a.font = H3 if k.startswith("【") else BODY; a.alignment = LEFT
        b = ws.cell(row=i, column=3, value=v); b.font = BODY; b.alignment = LEFT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)

    # ============= Sheet 2 对甲方贡献 =============
    ws = wb.create_sheet("02 对甲方贡献")
    title_bar(ws, "对甲方 (森马) 的贡献 — 24 个月 + 永续年金", 7)
    set_widths(ws, [4, 30, 18, 18, 18, 18, 30])

    ws.cell(row=2, column=2, value="表 A · 24 个月期间甲方直接收入").font = H3
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    write_header(ws, 3, ["", "项目", "T+1 年 (元)", "T+2 年 (元)", "24 月合计 (元)", "永续年化 (元)", "假设"])
    contrib = [
        ("租金收入 (满租后年化)",        8_030_000, 16_060_000, 24_090_000, 16_060_000, "T+1 平均 1 万方满租, T+2 总 2 万方满租, 永续按 1,606 万/年"),
        ("物业费 (10 元/㎡/月)",           1_200_000,   2_400_000,   3_600_000,   2_400_000, "渐进入驻, 永续按满租"),
        ("停车增量 (1500+ 车位)",           200_000,     500_000,     700_000,     500_000, "保守估计"),
        ("水电税收 / 商业配套增量",          300_000,     800_000,   1_100_000,     800_000, "保守估计"),
    ]
    sum1 = 0; sum2 = 0; sum_total = 0; sum_perp = 0
    for i, (n, a, b, c, p, note) in enumerate(contrib, start=4):
        sum1 += a; sum2 += b; sum_total += c; sum_perp += p
        write_row(ws, i, [i - 3, n, a, b, c, p, note], num_cols=[3, 4, 5, 6])
    last = 4 + len(contrib)
    write_row(ws, last, ["合计", "24 个月甲方直接收入 (元)", sum1, sum2, sum_total, sum_perp, "永续口径仅供参考"], num_cols=[3, 4, 5, 6], bold_total=True)

    # 表 B 不可量化
    ws.cell(row=last + 2, column=2, value="表 B · 不可量化收益").font = H3
    ws.merge_cells(start_row=last + 2, start_column=2, end_row=last + 2, end_column=7)
    write_header(ws, last + 3, ["", "维度", "影响", "估值口径 (供参考)", "", "", "说明"])
    intangible = [
        ("品牌势能",      "森马由零售品牌 → 潮玩产业生态品牌", "亿元量级",  "上市公司估值 P/E 改善 0.5-1 倍"),
        ("资产增值",      "TOD 板块 + 产业认证后地块溢价", "千万元量级", "周边对标 8-12% 年化资产增值"),
        ("政策红利",      "科技时尚特色小镇 / AI 潮玩产业基地", "百万级", "政府补贴 / 税收返还"),
        ("数据资产",      "200+ 入驻企业的产业数据", "千万级", "可作为森马二期 / 5.2 万方招商的基础数据"),
    ]
    for i, (k, v, est, note) in enumerate(intangible, start=last + 4):
        write_row(ws, i, [i - (last + 3), k, v, est, "", "", note], num_cols=[])

    # 备注
    ws.cell(row=last + 4 + len(intangible) + 1, column=2,
            value="口径说明: 表 A 永续年化 ≈ 1,976 万元/年 (租金 + 物业 + 停车 + 配套), 按 8% 折现的资产估值 ≈ 2.47 亿元.").font = NOTE
    ws.merge_cells(start_row=last + 4 + len(intangible) + 1, start_column=2,
                   end_row=last + 4 + len(intangible) + 1, end_column=7)
    ws.row_dimensions[last + 4 + len(intangible) + 1].height = 36

    # ============= Sheet 3 基础月费构成 =============
    ws = wb.create_sheet("03 基础月费构成")
    title_bar(ws, "基础月费 (Retainer) · 1-3 人配置阶梯", 7)
    set_widths(ws, [4, 24, 14, 14, 16, 16, 30])

    # 三档配置
    ws.cell(row=2, column=2, value="三档人员配置 (1-3 人) 月度刚性成本对比").font = H3
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    write_header(ws, 3, ["", "成本项", "1 人轻配 (元/月)", "★ 2 人推荐 (元/月)", "3 人重配 (元/月)", "备注", ""])
    cost_items = [
        ("产业招商经理 (含绩效)",   32_000, 32_000, 32_000, "22K 底薪 + 10K 绩效"),
        ("国际合作 & 活动策划 (含绩效)", 0,    28_000, 28_000, "20K 底薪 + 8K 绩效"),
        ("基金投后 & 政府关系 (含绩效)", 0,        0,  28_000, "20K 底薪 + 8K 绩效"),
        ("CSO 顾问费 (胡教授)",     12_000, 25_000, 25_000, "1 人轻配按半数, 2/3 人按全额"),
        ("行政 / 财务分摊",         3_000,  5_000,  8_000, "合资公司层面摊销"),
        ("仲量联行爬楼数据接口",     3_000,  3_000,  4_000, "26K 已购入, 后续运维"),
        ("差旅 / 接待 / 物料",       8_000, 12_000, 18_000, "外事接待 + 爬楼差旅"),
        ("管理与协调费",            3_000,  5_000,  8_000, "月度汇报 + 外部协调"),
    ]
    total1 = total2 = total3 = 0
    for i, (name, c1, c2, c3, note) in enumerate(cost_items, start=4):
        total1 += c1; total2 += c2; total3 += c3
        write_row(ws, i, [i - 3, name, c1, c2, c3, note, ""], num_cols=[3, 4, 5])
    last = 4 + len(cost_items)
    write_row(ws, last, ["合计", "月度刚性成本", total1, total2, total3, "", ""], num_cols=[3, 4, 5], bold_total=True)

    write_header(ws, last + 2, ["", "档位", "刚性成本 (元/月)", "建议月费 (元/月)", "毛利垫 (元)", "毛利率", "推荐度"])
    options = [
        ("1 人轻配", total1, 60_000),
        ("★ 2 人推荐配", total2, 120_000),
        ("3 人重配", total3, 180_000),
    ]
    for i, (name, cost, fee) in enumerate(options, start=last + 3):
        margin = fee - cost
        rate = margin / fee
        rec = "★★★" if name.startswith("★") else ("★" if "1 人" in name else "★★")
        write_row(ws, i, [i - (last + 2), name, cost, fee, margin, f"{rate*100:.1f}%", rec], num_cols=[3, 4, 5])

    note = ("v1.0 推荐方案: 2 人配置 + 12 万元/月, 24 个月合计基础月费 288 万元. 该配置覆盖刚性成本 9.5 万 + 安全垫 2.5 万 (毛利率 21%), 用于差旅/接待/数据/物料.")
    ws.cell(row=last + 3 + len(options) + 1, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=last + 3 + len(options) + 1, start_column=2, end_row=last + 3 + len(options) + 1, end_column=7)
    ws.row_dimensions[last + 3 + len(options) + 1].height = 50

    # ============= Sheet 4 招商佣金阶梯 =============
    ws = wb.create_sheet("04 招商佣金阶梯")
    title_bar(ws, "招商佣金阶梯 · 1.5–2 个月年租金", 7)
    set_widths(ws, [4, 22, 14, 14, 14, 16, 32])

    write_header(ws, 3, ["", "面积档位", "日租金 (元/㎡/天)", "年租金 (元/㎡)", "佣金月数", "佣金 (元/㎡)", "说明"])
    tiers = [
        ("≤ 2,000㎡ 小型",        2.0, 1.5,  "小型潮玩 / 服务机构"),
        ("2,001-5,000㎡ 中型",    2.2, 1.75, "中型潮玩运营企业"),
        ("> 5,000㎡ 头部 / 央企", 2.5, 2.0,  "头部央企 / 行业协会"),
    ]
    for i, (tier, d, m, note) in enumerate(tiers, start=4):
        annual = round(d * 365)
        commission = round(annual * m / 12)
        write_row(ws, i, [i - 3, tier, d, annual, m, commission, note], num_cols=[4, 6])

    # 24 月场景
    ws.cell(row=8, column=2, value="24 个月成交场景预估").font = H3
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=7)
    write_header(ws, 9, ["", "场景", "成交面积 (㎡)", "平均日租金", "平均佣金 (元/㎡)", "佣金合计 (元)", "假设"])
    scenarios = [
        ("保守", 14_000, 2.0, 96,  "首年仅 4# 楼 + T+2 楼部分租出"),
        ("基础", 20_000, 2.2, 117, "T+1 满 4# 楼 + T+2 满 5# 楼"),
        ("乐观", 22_000, 2.4, 140, "提前满租 + 头部加成 + 返投奖励"),
    ]
    for i, (s, area, daily, avg, note) in enumerate(scenarios, start=10):
        total = area * avg
        write_row(ws, i, [i - 9, s, area, daily, avg, total, note], num_cols=[3, 4, 5, 6])

    note = "① 仲量联行爬楼大数据可使转化率 +30%; ② 追觅基金返投落地客户额外加成 +0.25 个月; ③ 客户名单归属保护期 24 个月。"
    ws.cell(row=14, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=14, start_column=2, end_row=14, end_column=7)
    ws.row_dimensions[14].height = 36

    # ============= Sheet 5 挂牌奖励明细 =============
    ws = wb.create_sheet("05 挂牌奖励明细")
    title_bar(ws, "5 项挂牌奖励 · 一次性激励", 7)
    set_widths(ws, [4, 32, 24, 14, 14, 14, 32])

    write_header(ws, 3, ["", "挂牌名称", "出牌方", "落地节点", "数量", "单项奖励 (元)", "对元谷的核心价值"])
    bands = [
        ("AI 潮玩产业基地",                "中国动漫集团",            "T+3 月", 1, 300_000, "牌照即招商, AI+潮玩双赛道核心抓手"),
        ("潮玩次元商业专委会",              "中国百货商业协会",         "T+3 月", 1, 300_000, "聚集潮玩零售生态 + 政府对话渠道"),
        ("上海市科技企业联合会 · 元谷产业基地", "上海市科技企业联合会",     "T+6 月", 1, 300_000, "上海科技企业生态导流 + 补贴申报"),
        ("复旦大学住房政策研究中心 · 元谷分中心", "复旦大学住房政策研究中心",  "T+9 月", 1, 300_000, "学术背书 + 政策研究 + 高净值人脉"),
        ("福布斯产业影响力奖 · 元谷专场",     "福布斯",                  "T+12 月 (每年)", 1, 300_000, "国际品牌势能 + 年度评选 IP"),
    ]
    total_award = 0
    for i, (n, body, node, qty, price, value) in enumerate(bands, start=4):
        total_award += qty * price
        write_row(ws, i, [i - 3, n, body, node, qty, qty * price, value], num_cols=[5, 6])
    write_row(ws, 4 + len(bands), ["合计", "5 项挂牌奖励合计", "", "", 5, total_award, "首年 4 项前置 + T+2 年福布斯续挂"], num_cols=[5, 6], bold_total=True)

    note = "结算节奏: 挂牌正式公告之日起 30 日内由甲方一次性支付; 续挂 (第 2 年起) 按 50% 收取."
    ws.cell(row=4 + len(bands) + 2, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=4 + len(bands) + 2, start_column=2, end_row=4 + len(bands) + 2, end_column=7)

    # ============= Sheet 6 6 场沙龙 =============
    ws = wb.create_sheet("06 沙龙测算")
    title_bar(ws, "6 场产业沙龙 · 每场 ≥ 30 目标产业客户", 8)
    set_widths(ws, [4, 22, 26, 12, 14, 14, 14, 30])

    write_header(ws, 3, ["", "序号 / 时间", "主题 / 联办", "目标客户", "执行成本 (元)", "营收预估 (元)", "单场净利 (元)", "说明"])
    salons = [
        ("#1 / T+1 月",  "AI+潮玩 (借势 5/22 峰会) / 中动漫 + AI 腾讯",       30, 50_000, 120_000, "首场, 借势主题鲜明"),
        ("#2 / T+3 月",  "潮玩出海 / 北欧会客厅 + 福布斯",                    30, 50_000, 110_000, "国际化主题, 出海路演"),
        ("#3 / T+5 月",  "投融资路演 / 追觅 + 招行 + 长江 + 金浦",             30, 50_000, 130_000, "金融嘉宾密集, 招商利器"),
        ("#4 / T+7 月",  "设计与创意 / 上海交大 + 上海市科企联",                30, 50_000, 100_000, "学术联办 + 产业转化"),
        ("#5 / T+9 月",  "内容 IP 与 Z 世代 / 中百协潮玩次元专委 + 中动漫",      30, 50_000, 110_000, "IP 产业生态"),
        ("#6 / T+11 月", "政策与小镇 / 闵行科协 + 复旦住房政策中心",            30, 50_000, 100_000, "政府关系深化"),
    ]
    total_cost = 0; total_rev = 0; total_net = 0
    for i, (time, theme, n, cost, rev, note) in enumerate(salons, start=4):
        net = rev - cost
        total_cost += cost; total_rev += rev; total_net += net
        write_row(ws, i, [i - 3, time, theme, n, cost, rev, net, note], num_cols=[4, 5, 6, 7])
    last = 4 + len(salons)
    write_row(ws, last, ["合计", "6 场沙龙合计", "", 180, total_cost, total_rev, total_net, "≈ 36 万 / 67 万 / 31 万元"], num_cols=[4, 5, 6, 7], bold_total=True)

    # 分润
    ws.cell(row=last + 2, column=2, value="净利分润机制 (合资公司 / 胡教授团队)").font = H3
    ws.merge_cells(start_row=last + 2, start_column=2, end_row=last + 2, end_column=8)
    write_header(ws, last + 3, ["", "项目", "金额 (元)", "甲方 (30%)", "丙方 (70%)", "", "", "说明"])
    write_row(ws, last + 4, ["1", "6 场沙龙合计净利", total_net, round(total_net*0.3), round(total_net*0.7), "", "", "胡教授团队享 70%"], num_cols=[3, 4, 5])

    note = "口径: 营收 = 赞助 + 票务 + 政府补贴; 执行成本 = 场地 + 物料 + 嘉宾接待 + 传播. 单场目标客户 ≥ 30 是 KPI 红线, 未达标的当场净利甲方扣回 50%."
    ws.cell(row=last + 6, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=last + 6, start_column=2, end_row=last + 6, end_column=8)
    ws.row_dimensions[last + 6].height = 36

    # ============= Sheet 7 24 月双向账本 =============
    ws = wb.create_sheet("07 双向账本+ROI")
    title_bar(ws, "24 个月双向账本 · 投入产出比 ROI", 7)
    set_widths(ws, [4, 32, 18, 18, 18, 18, 30])

    ws.cell(row=2, column=2, value="表 A · 胡教授团队 24 月结算金额 (向森马收取)").font = H3
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    write_header(ws, 3, ["", "收入类别", "T+1 年 (元)", "T+2 年 (元)", "24 月合计 (元)", "占比", "结算节奏"])
    incomes = [
        ("基础月费 (12 万 × 12)",         1_440_000, 1_440_000, 2_880_000, "按月预付"),
        ("招商佣金 (1.5-2 个月)",         1_170_000, 1_170_000, 2_340_000, "起租后 30 日内"),
        ("挂牌奖励 (5 × 30 万, T+2 半价)", 1_200_000,   300_000, 1_500_000, "挂牌公告后 30 日内"),
        ("沙龙净分润 (70% × 31 万)",        210_000,   210_000,   420_000, "按场结算"),
    ]
    tot1 = tot2 = tot_all = 0
    for i, (n, a, b, c, node) in enumerate(incomes, start=4):
        tot1 += a; tot2 += b; tot_all += c
        write_row(ws, i, [i - 3, n, a, b, c, f"{c/(2_880_000+2_340_000+1_500_000+420_000)*100:.0f}%", node], num_cols=[3, 4, 5])
    last = 4 + len(incomes)
    write_row(ws, last, ["合计", "胡教授团队 24 月收入合计 (元)", tot1, tot2, tot_all, "100%", "≈ 714 万元"], num_cols=[3, 4, 5], bold_total=True)

    # 表 B 甲方贡献
    ws.cell(row=last + 2, column=2, value="表 B · 甲方 (森马) 24 月直接收入").font = H3
    ws.merge_cells(start_row=last + 2, start_column=2, end_row=last + 2, end_column=7)
    write_header(ws, last + 3, ["", "项目", "T+1 年 (元)", "T+2 年 (元)", "24 月合计 (元)", "", "永续年化 (元)"])
    side_b = [
        ("租金收入",          8_030_000, 16_060_000, 24_090_000, "", 16_060_000),
        ("物业费 (10/㎡/月)", 1_200_000,  2_400_000,   3_600_000, "",  2_400_000),
    ]
    tot_b1 = tot_b2 = tot_b = tot_bp = 0
    for i, (n, a, b, c, _, p) in enumerate(side_b, start=last + 4):
        tot_b1 += a; tot_b2 += b; tot_b += c; tot_bp += p
        write_row(ws, i, [i - (last + 3), n, a, b, c, "", p], num_cols=[3, 4, 5, 7])
    write_row(ws, last + 4 + len(side_b), ["合计", "甲方 24 月直接收入 (元)", tot_b1, tot_b2, tot_b, "", tot_bp], num_cols=[3, 4, 5, 7], bold_total=True)

    # ROI
    ws.cell(row=last + 4 + len(side_b) + 2, column=2, value="表 C · ROI (甲方视角)").font = H3
    ws.merge_cells(start_row=last + 4 + len(side_b) + 2, start_column=2, end_row=last + 4 + len(side_b) + 2, end_column=7)
    write_header(ws, last + 4 + len(side_b) + 3, ["", "项目", "金额 (元)", "占比 / 倍数", "", "", "说明"])
    roi_rows = [
        ("甲方付出 (24 月)",   tot_all,                 "—", "", "", "= 表 A 合计"),
        ("甲方收入 (24 月)",   tot_b,                   "—", "", "", "= 表 B 合计"),
        ("净收入",             tot_b - tot_all,         "—", "", "", "= 表 B - 表 A"),
        ("ROI 投入产出比",     "",                      f"1 : {tot_b/tot_all:.2f}", "", "", "甲方付出 1 元 → 收入 X 元"),
        ("永续资产估值 (8% 折现)", round(tot_bp / 0.08), "", "", "", "按永续年金折现"),
    ]
    for i, (k, v, ratio, _, _, note) in enumerate(roi_rows, start=last + 4 + len(side_b) + 4):
        c1 = ws.cell(row=i, column=2, value=k); c1.font = BODY; c1.alignment = LEFT; c1.border = BORDER
        c2 = ws.cell(row=i, column=3, value=v); c2.font = NUM if isinstance(v, int) else BODY
        c2.alignment = RIGHT if isinstance(v, int) else LEFT; c2.border = BORDER
        if isinstance(v, int):
            c2.number_format = "#,##0"
        c3 = ws.cell(row=i, column=4, value=ratio); c3.font = BODY; c3.alignment = CENTER; c3.border = BORDER
        c4 = ws.cell(row=i, column=7, value=note); c4.font = BODY; c4.alignment = LEFT; c4.border = BORDER
        ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)
        for j in [2, 3, 7]:
            ws.cell(row=i, column=j).fill = FILL_ALT if i % 2 == 0 else FILL_LIGHT

    # ============= Sheet 8 客户管道 =============
    ws = wb.create_sheet("08 招商客户管道")
    title_bar(ws, "招商客户管道 (示例脱敏 + 5/22 峰会嘉宾衍生)", 7)
    set_widths(ws, [4, 22, 28, 14, 14, 16, 30])

    write_header(ws, 3, ["", "层级 / 类型", "代表客户 (脱敏)", "意向面积 (㎡)", "对接渠道", "预期成交节点", "状态"])
    pipeline = [
        ("L1 头部央企",    "中字头数字央企 A (AI 内容方向)",         2_000, "中国动漫集团 + 牌照",        "T+3 月",  "意向沟通中"),
        ("L1 头部央企",    "中字头文创央企 B",                       2_000, "中动漫 + 闵行科协",          "T+6 月",  "初步接触"),
        ("L1 行业协会",    "中国百货协会 潮玩次元专委",                1_500, "中百协联建",                 "T+3 月",  "已确认"),
        ("L2 中型 / AI",   "AI 潮玩品牌 X",                          1_500, "5/22 峰会 + 追觅基金",      "T+4 月",  "高意向"),
        ("L2 中型 / 国漫", "国漫公司 Y",                              1_200, "中动漫推荐",                 "T+5 月",  "高意向"),
        ("L2 中型 / 出海", "潮玩出海公司 Z",                          1_000, "北欧会客厅",                 "T+6 月",  "意向"),
        ("L2 中型 / 设计", "AI 设计公司 W",                            800, "AI 腾讯 + 上海交大",         "T+7 月",  "意向"),
        ("L3 小型",        "盲盒新锐品牌 (5 家)",                    1_500, "仲量联行爬楼数据",            "T+6-9 月", "数据筛选中"),
        ("L3 小型",        "设计师工作室 (10 家)",                  2_500, "5/22 峰会 + 沙龙 #4",        "T+7-12 月", "待沙龙转化"),
        ("L3 小型",        "潮玩零售商 (15 家)",                    3_750, "爬楼 + 牌照拉新",            "T+9-15 月", "数据筛选中"),
        ("L4 服务机构",    "财税/法律/IP/出海咨询 (15 家)",          3_000, "6 场沙龙 + 服务中心",        "T+6-12 月", "管道储备"),
    ]
    pipeline_total = 0
    for i, (l, n, a, ch, time, status) in enumerate(pipeline, start=4):
        pipeline_total += a
        write_row(ws, i, [i - 3, l, n, a, ch, time, status], num_cols=[4])
    write_row(ws, 4 + len(pipeline), ["合计", "招商管道意向面积 (㎡)", "", pipeline_total, "", "覆盖 4#+5# 楼 2 万方目标", ""], num_cols=[4], bold_total=True)

    note = "实际管道客户已超过 60 家 (本表为节选脱敏); 与 5/22 峰会嘉宾 (200+ VIP) 一一对接转化."
    ws.cell(row=4 + len(pipeline) + 2, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=4 + len(pipeline) + 2, start_column=2, end_row=4 + len(pipeline) + 2, end_column=7)
    ws.row_dimensions[4 + len(pipeline) + 2].height = 30

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
