"""Build the Yuangu cooperation revenue model `.xlsx`.

v1.1 changes:
  * 加入 Sheet "02b 市场租金对标"，引用 2024-2025 大零号湾 / 紫竹高新区
    实际成交租金 (1.5-2.5 元/㎡/天)，作为佣金测算基准。
  * 招商佣金阶梯日租金从 4.5/5.0/5.5 下调至 1.8/2.2/2.5。
  * 月费推荐区间从 30-50 下调至 20-35 万 (推荐 28 万)。
  * 新增 Sheet "05b 科技企业服务中心"——对入驻 / 区域内企业收费的
    增值服务，不向甲方收取，作为合资公司增量净利源 → 30% 归丙方分红。
  * Sheet "06 股权分红+三年累计" 重算，吸收服务中心净利贡献。
  * Sheet "07 敏感性分析" 月费档位调整为 15-35 万。
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("合作收益测算模型.xlsx")

PRIMARY = "FF142C5E"
ACCENT = "FFF27E2D"
LIGHT = "FFEAEEF5"
ALT = "FFF8F9FB"

THIN = Side(style="thin", color="FFB0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1 = Font(name="微软雅黑", size=18, bold=True, color="FFFFFFFF")
H2 = Font(name="微软雅黑", size=12, bold=True, color="FFFFFFFF")
H3 = Font(name="微软雅黑", size=11, bold=True, color=PRIMARY[2:])
BODY = Font(name="微软雅黑", size=11, color="FF1F2A44")
NOTE = Font(name="微软雅黑", size=10, italic=True, color="FF55607A")
NUM = Font(name="Consolas", size=11, color="FF1F2A44")
NUM_BOLD = Font(name="Consolas", size=11, bold=True, color=PRIMARY[2:])

FILL_PRIMARY = PatternFill("solid", fgColor=PRIMARY[2:])
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT[2:])
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT[2:])
FILL_ALT = PatternFill("solid", fgColor=ALT[2:])

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = H2
        c.fill = FILL_PRIMARY
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def write_row(ws, row, vals, num_cols=None, bold_total=False):
    num_cols = num_cols or []
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.border = BORDER
        if j in num_cols:
            c.number_format = "#,##0"
            c.alignment = RIGHT
            c.font = NUM_BOLD if bold_total else NUM
        else:
            c.alignment = LEFT
            c.font = BODY
        if bold_total:
            c.fill = FILL_ACCENT
            c.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
        elif row % 2 == 0:
            c.fill = FILL_ALT
        else:
            c.fill = FILL_LIGHT


def title_bar(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = H1
    c.fill = FILL_PRIMARY
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40


def main():
    wb = Workbook()

    # ============= Sheet 01 封面 =============
    ws = wb.active
    ws.title = "01 封面与说明"
    title_bar(ws, "元谷项目 · 胡教授团队 × 森马 联合运营合作收益测算模型 v1.1", 8)
    set_widths(ws, [4, 22, 22, 22, 22, 22, 22, 22])

    notes = [
        ("", ""),
        ("文件版本", "v1.1 (基于真实市场租金调研后的修订版)"),
        ("适用项目", "森马(上海)国际运营中心 元谷项目"),
        ("适用主体", "森马集团 × 危总团队 × 胡教授团队 联合运营合资公司"),
        ("货币单位", "人民币 元 (除非另行注明)"),
        ("测算口径", "首年现金口径; 不含税; 不含合资公司行政固定成本"),
        ("Sheet 索引", "02 总览 / 02b 市场租金对标 / 03 月费 / 04 招商佣金 / 05 活动+奖项 / 05b 科技企业服务中心 / 06 分红+三年累计 / 07 敏感性"),
        ("", ""),
        ("【v1.1 关键调整】", ""),
        ("市场租金对标", "大零号湾 / 紫竹高新区主流办公&研发租金 1.8-2.5 元/㎡/天 (Sheet 02b)"),
        ("月费区间", "由 30-50 万下调至 20-35 万 (推荐 28 万)"),
        ("招商佣金", "日租金基准从 4.5-5.5 下调到 1.8-2.5"),
        ("新增模块", "科技企业服务中心 (面向入驻企业 9 大类服务，对外收费，不向甲方收取)"),
        ("", ""),
        ("【收入结构】", ""),
        ("固定月费 Retainer", "甲方支付 → 覆盖团队 + 数据接口 (Sheet 03)"),
        ("专项奖项激励", "甲方支付 → 牌照 / 福布斯 / 小镇奖项 (Sheet 05)"),
        ("活动运营收入", "外部 + 甲方混合 → 赞助 + 票务 + 政府补贴 (Sheet 05)"),
        ("招商佣金", "甲方支付 → 实际成交年租金的 1-2.5 个月 (Sheet 04)"),
        ("科技企业服务中心", "外部支付 → 入驻企业 9 大类增值服务费 (Sheet 05b)"),
        ("股权分红", "合资公司净利润按股比分配 → 含服务中心利润 (Sheet 06)"),
    ]
    for i, (k, v) in enumerate(notes, start=2):
        if not k and not v:
            continue
        a = ws.cell(row=i, column=2, value=k)
        a.font = H3 if k.startswith("【") else BODY
        a.alignment = LEFT
        b = ws.cell(row=i, column=3, value=v)
        b.font = BODY
        b.alignment = LEFT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)

    # ============= Sheet 02 收入结构总览 =============
    ws = wb.create_sheet("02 收入结构总览")
    title_bar(ws, "首年收入结构总览(保守 / 基础 / 乐观 三档) — v1.1 含科技企业服务中心", 6)
    set_widths(ws, [4, 30, 18, 18, 18, 36])

    ws.cell(row=2, column=2,
            value="表 A · 胡教授团队首年现金口径 (5 项, 已去重; 服务中心利润已通过分红体现)").font = H3
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
    write_header(ws, 3, ["", "收入类别", "保守场景 (元)", "基础场景 (元)", "乐观场景 (元)", "测算口径备注"])
    rows = [
        ("固定月费 Retainer (年化, 甲方付)", 2_400_000, 3_360_000, 4_200_000, "20 / 28 / 35 万元 × 12 月 (Sheet 03)"),
        ("招商佣金 Commission (甲方付)",       720_000, 1_320_000, 2_640_000, "按实际成交年租金的 1-2.5 个月 (Sheet 04)"),
        ("活动运营收入 (外部+甲方)",         1_800_000, 2_400_000, 3_600_000, "10-15 场科技开放麦 + 1 届潮玩大赛"),
        ("专项奖项 / 挂牌激励 (甲方付)",       600_000, 1_500_000, 2_000_000, "牌照 / 福布斯 / 小镇奖项 (Sheet 05)"),
        ("股权分红 (合资公司净利, 含服务中心)", 720_000, 1_440_000, 2_880_000, "合资公司净利 × 30% (Sheet 06)"),
    ]
    sub_total_conservative = sum(r[1] for r in rows)
    sub_total_base = sum(r[2] for r in rows)
    sub_total_optimistic = sum(r[3] for r in rows)
    for i, (cat, c1, c2, c3, note) in enumerate(rows, start=4):
        write_row(ws, i, [i - 3, cat, c1, c2, c3, note], num_cols=[3, 4, 5])
    last = 4 + len(rows)
    write_row(ws, last, ["合计", "胡教授团队首年合计 (元)", sub_total_conservative, sub_total_base, sub_total_optimistic, "保守 / 基础 / 乐观"], num_cols=[3, 4, 5], bold_total=True)
    write_row(ws, last + 1, ["折合 (万元)", "胡教授团队首年合计 (万元)",
                              round(sub_total_conservative / 10000),
                              round(sub_total_base / 10000),
                              round(sub_total_optimistic / 10000),
                              "≈ 624 / 1,002 / 1,532 万元"], num_cols=[3, 4, 5], bold_total=True)

    # ---- 表 B: 合资公司层面增量收入 (信息项) ----
    info_row = last + 3
    ws.cell(row=info_row, column=2,
            value="表 B · 合资公司额外营收信息项 (来源于对外, 不直接计入胡教授团队总额, 通过分红 30% 间接受益)").font = H3
    ws.merge_cells(start_row=info_row, start_column=2, end_row=info_row, end_column=6)
    write_header(ws, info_row + 1,
                 ["", "项目", "保守 (元)", "基础 (元)", "乐观 (元)", "丙方 30% 分红影响 (基础)"])
    info_rows = [
        ("科技企业服务中心营收 (Sheet 05b)", 1_500_000, 3_000_000, 6_400_000, "经 38% 净利 × 30% ≈ 34 万元/年"),
        ("自营业态毛利 (选品/直播/设计中心)",   600_000, 1_200_000, 2_500_000, "已纳入分红线"),
    ]
    for i, (cat, c1, c2, c3, note) in enumerate(info_rows, start=info_row + 2):
        write_row(ws, i, [i - (info_row + 1), cat, c1, c2, c3, note], num_cols=[3, 4, 5])

    # ---- 占比小表 ----
    pct_row = info_row + 3 + len(info_rows)
    ws.cell(row=pct_row, column=2, value="基础场景下胡教授团队 5 项收入占比").font = H3
    ws.merge_cells(start_row=pct_row, start_column=2, end_row=pct_row, end_column=6)
    write_header(ws, pct_row + 1, ["", "收入类别", "金额 (元)", "占比", "支付方", "备注"])
    sources = ["对甲方 (森马)", "对甲方", "外部+甲方", "对甲方", "合资公司"]
    for i, ((cat, _c1, c2, _c3, _note), src) in enumerate(zip(rows, sources), start=pct_row + 2):
        pct = c2 / sub_total_base
        write_row(ws, i, [i - (pct_row + 1), cat, c2, f"{pct*100:.1f}%", src, ""], num_cols=[3])

    note_text = "口径说明: 表 A 为流向胡教授团队的实际现金 (避免与服务中心营收重复计算)。表 B 为合资公司层面的额外营收, 该部分对胡教授团队的影响通过股权分红 30% 一并体现在表 A 第 5 行 (因此分红额随服务中心规模扩大而提升)。该结构旨在向森马清晰展示: 合资公司有多元化对外收入, 而非仅依赖森马月费。"
    ws.cell(row=pct_row + 2 + len(rows), column=2, value=note_text).font = NOTE
    ws.merge_cells(start_row=pct_row + 2 + len(rows), start_column=2, end_row=pct_row + 2 + len(rows), end_column=6)
    ws.row_dimensions[pct_row + 2 + len(rows)].height = 60

    # ============= Sheet 02b 市场租金对标 =============
    ws = wb.create_sheet("02b 市场租金对标")
    title_bar(ws, "大零号湾 / 紫竹高新区主流园区办公&研发租金对标 (2024-2025)", 5)
    set_widths(ws, [4, 36, 18, 18, 28])

    write_header(ws, 3, ["", "园区 / 项目", "日租金 (元/㎡/天)", "年租金 (元/㎡)", "备注"])
    benchmarks = [
        ("零号湾全球创新创业集聚区", 2.25, "200-1000㎡ 灵活工位"),
        ("大零号湾科创成果转化中心", 2.25, "500-2500㎡ 研发办公 + 中试"),
        ("华谊万创新所",            2.20, "剑川路沿线"),
        ("上海人工智能产业园",        2.20, "AI 主题园区"),
        ("紫竹信息数码港 (5A 甲级)",  2.30, "东川路 555 号"),
        ("紫竹数字创意港",          2.50, "数字创意主题"),
        ("龙湖蓝海引擎·淡水河畔",     2.00, "1.5-2.5 元区间"),
        ("金领谷科技产业园",         2.15, "1.5-2.8 元区间"),
        ("东软软件园",              1.50, "成熟存量园区"),
        ("云境 443 未来产业社区 (高端)", 3.15, "2.3-4.0 元区间"),
        ("夏日汇国际中心 (高端)",     4.00, "3.5-4.5 元区间"),
    ]
    for i, (name, daily, note) in enumerate(benchmarks, start=4):
        annual = round(daily * 365)
        write_row(ws, i, [i - 3, name, daily, annual, note], num_cols=[4])

    avg_low = 1.8
    avg_mid = 2.2
    avg_high = 2.5
    sum_row = 4 + len(benchmarks)
    write_row(ws, sum_row, ["", "市场主流区间 (办公 / 研发)", f"{avg_low} - {avg_high}", f"{round(avg_low*365)} - {round(avg_high*365)}", "本测算所采用基准"], num_cols=[], bold_total=True)
    note = "数据来源：001zf / 汇租选址 / 聚多楼选址 / 租易网 / urlou 等平台 2024-2025 年公开放租信息汇总。元谷项目作为新建 TOD 综合体 + 政府重点扶持的科技时尚特色小镇，建议招商定价采用市场主流区间中位 (1.8-2.5 元/㎡/天)；商业 1F 临街铺位可单独按 3.0-5.0 元定价，但作为加权平均不影响整体测算。"
    ws.cell(row=sum_row + 2, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=sum_row + 2, start_column=2, end_row=sum_row + 2, end_column=5)
    ws.row_dimensions[sum_row + 2].height = 80

    # ============= Sheet 03 月费测算 =============
    ws = wb.create_sheet("03 月费测算")
    title_bar(ws, "固定月费 Retainer · 团队成本反算(下调后建议月费区间)", 6)
    set_widths(ws, [4, 26, 14, 14, 18, 32])

    write_header(ws, 3, ["", "成本项", "人数", "月度单价 (元)", "月度合计 (元)", "说明"])
    cost_items = [
        ("胡教授(CSO 顾问费)",     1, 50_000, "每周 2 个工作日 + 战略评审 (由 60K 调降)"),
        ("产业招商经理",          1, 32_000, "底薪 22K + 招商绩效 10K"),
        ("国际合作 & 活动策划",    1, 28_000, "底薪 20K + 活动绩效 8K"),
        ("基金投后 & 政府关系",    1, 28_000, "底薪 20K + 牌照绩效 8K"),
        ("行政 / 财务 (50% 分摊)", 1, 10_000, "合资公司共享岗位摊销"),
        ("仲量联行爬楼数据接口",   1, 4_000,  "在 26,000 元已购基础上的运维费"),
        ("差旅 / 接待 / 物料",     1, 14_000, "外事接待 + 爬楼差旅 + 物料"),
        ("管理与协调费",           1, 8_000,  "月度汇报 + 外部协调"),
    ]
    total_monthly = 0
    for i, (name, n, price, note) in enumerate(cost_items, start=4):
        sub = n * price
        total_monthly += sub
        write_row(ws, i, [i - 3, name, n, price, sub, note], num_cols=[3, 4, 5])
    last = 4 + len(cost_items)
    write_row(ws, last, ["合计", "团队月度刚性成本", "", "", total_monthly, f"约 {total_monthly//10000} 万元/月"], num_cols=[5], bold_total=True)

    margin_rows = [
        ("月费保底 (推荐, 谈判锚点)", 280_000, "≈ 团队成本的 1.6 倍, 留出适度安全垫"),
        ("月费区间下限 (红线)",      200_000, "覆盖核心 5 人 + 数据接口, 0 利润"),
        ("月费区间上限 (理想)",      350_000, "扩配 + 国际接待 + 政府关系强化"),
    ]
    ws.cell(row=last + 2, column=2, value="月费谈判区间 (v1.1 已下调)").font = H3
    write_header(ws, last + 3, ["", "档位", "金额 (元/月)", "年化 (元)", "毛利率 (相对成本)", "说明"])
    for i, (name, m, note) in enumerate(margin_rows, start=last + 4):
        annual = m * 12
        margin = (m - total_monthly) / m
        write_row(ws, i, [i - (last + 3), name, m, annual, f"{margin*100:.1f}%", note], num_cols=[3, 4])

    note = "原 v1.0 月费区间为 30-50 万 (推荐 40 万)。考虑到大零号湾实际办公租金 1.8-2.5 元/㎡/天 (年租 657-913 元/㎡)，对比同类园区运营商管理费水平 (年租金的 5-8%)，本版本将月费下调为 20-35 万 (推荐 28 万)，更符合森马成本可承受度，并以新增的科技企业服务中心填补差额。"
    ws.cell(row=last + 4 + len(margin_rows) + 1, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=last + 4 + len(margin_rows) + 1, start_column=2, end_row=last + 4 + len(margin_rows) + 1, end_column=6)
    ws.row_dimensions[last + 4 + len(margin_rows) + 1].height = 80

    # ============= Sheet 04 招商佣金阶梯 =============
    ws = wb.create_sheet("04 招商佣金阶梯")
    title_bar(ws, "招商佣金阶梯测算 (v1.1 按市场实际租金 1.8-2.5 元/㎡/天)", 7)
    set_widths(ws, [4, 24, 14, 14, 14, 16, 28])

    write_header(ws, 3, ["", "面积档位", "日租金 (元/㎡/天)", "年租金 (元/㎡)", "佣金月数", "佣金 (元/㎡)", "说明"])
    tiers = [
        ("≤ 2,000㎡ 小型 (上层办公)", 1.8, 1.0, "小型潮玩 / 服务机构, 多为 5F+"),
        ("2,001-5,000㎡ 中型",       2.2, 1.5, "中型潮玩运营企业"),
        ("> 5,000㎡ 头部 (综合)",    2.5, 2.5, "头部央企 / 行业协会"),
    ]
    for i, (tier, d, m, note) in enumerate(tiers, start=4):
        annual = round(d * 365)
        commission = round(annual * m / 12)
        write_row(ws, i, [i - 3, tier, d, annual, m, commission, note], num_cols=[4, 6])

    note1 = "说明：① 商业 1F 临街铺位 (如 IP 选品中心 / 主题街区) 可按 3.0-5.0 元/㎡/天单独定价，对应佣金按其实际成交价计算；② 仲量联行爬楼大数据可使佣金转化率提升 30%；③ 追觅基金返投落地客户额外计 0.5 个月加成；④ 招商主导权由合资公司独家持有 5 年。"
    ws.cell(row=8, column=2, value=note1).font = NOTE
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=7)
    ws.row_dimensions[8].height = 56

    write_header(ws, 10, ["", "场景", "成交面积 (㎡)", "平均佣金 (元/㎡)", "成交家数", "佣金合计 (元)", "说明"])
    scenarios = [
        ("保守", 6_000,  120, 12, "首年成交不足；以中小型为主"),
        ("基础", 12_000, 110, 22, "稳态推进；中小+中型 + 1 家头部"),
        ("乐观", 24_000, 110, 35, "返投基金 + 牌照拉动；头部签约 3 家"),
    ]
    for i, (s, area, avg, n, note) in enumerate(scenarios, start=11):
        total = area * avg
        write_row(ws, i, [i - 10, s, area, avg, n, total, note], num_cols=[3, 4, 5, 6])

    note2 = "v1.1 提醒: 佣金额下降是因为日租金对标真实市场水平后下调约 60%。要补足合资公司收入，必须强化“科技企业服务中心”对外收费 (Sheet 05b)。"
    ws.cell(row=15, column=2, value=note2).font = NOTE
    ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=7)
    ws.row_dimensions[15].height = 38

    # ============= Sheet 05 活动 + 奖项 =============
    ws = wb.create_sheet("05 活动+奖项")
    title_bar(ws, "活动运营收入 + 专项奖项激励 测算", 7)
    set_widths(ws, [4, 28, 12, 14, 14, 16, 28])

    write_header(ws, 3, ["", "活动 / 奖项", "频次/年", "单场收入 (元)", "成本 (元)", "净收入 (元)", "说明"])
    activities = [
        ("科技开放麦 (基础场)",       12, 80_000,  30_000, "赞助 + 票务 + 园区分摊"),
        ("科技开放麦 (大场)",          2, 250_000, 80_000, "森马联合发布 / 国际嘉宾"),
        ("北欧创新国际会客厅外事接待", 6, 120_000, 50_000, "外事接待 + 路演"),
        ("全国潮玩设计大赛 (年度)",   1, 800_000, 300_000, "省级补贴 + 头部赞助"),
        ("福布斯榜单 元谷专场发布",   1, 600_000, 200_000, "联合森马 + 福布斯方"),
        ("AI 共享设计中心 工作坊",    12, 30_000, 10_000, "AI 腾讯背书 + 中型企业付费"),
    ]
    total_net = 0
    for i, (name, f, rev, cost, note) in enumerate(activities, start=4):
        net = (rev - cost) * f
        total_net += net
        write_row(ws, i, [i - 3, name, f, rev, cost, net, note], num_cols=[4, 5, 6])
    write_row(ws, 4 + len(activities), ["合计", "活动年度净收入", "", "", "", total_net, "≈ 240-360 万元"], num_cols=[6], bold_total=True)

    base = 4 + len(activities) + 2
    ws.cell(row=base, column=2, value="专项奖项 / 挂牌激励 (一次性)").font = H3
    write_header(ws, base + 1, ["", "奖项 / 挂牌", "数量", "单项激励 (元)", "合计 (元)", "", "说明"])
    awards = [
        ("AI 潮玩产业基地 牌照挂牌",      1, 300_000, "中国动漫集团"),
        ("潮玩次元商业专委会 牌照挂牌",   1, 300_000, "中国百货商业协会"),
        ("上海科技时尚特色小镇 政府奖项",  1, 400_000, "市/区级"),
        ("福布斯榜单上榜 / 评选挂牌",     2, 200_000, "U30 / 创新榜 / 园区榜"),
        ("国家级科技 / 文创奖项",         1, 300_000, "如国家文创基金 / 工信部专项"),
    ]
    award_total = 0
    for i, (n, qty, price, note) in enumerate(awards, start=base + 2):
        total = qty * price
        award_total += total
        write_row(ws, i, [i - (base + 1), n, qty, price, total, "", note], num_cols=[4, 5])
    write_row(ws, base + 2 + len(awards), ["合计", "首年挂牌奖励合计", "", "", award_total, "", "≈ 150-200 万元"], num_cols=[5], bold_total=True)

    # ============= Sheet 05b 科技企业服务中心 =============
    ws = wb.create_sheet("05b 科技企业服务中心")
    title_bar(ws, "科技企业服务中心 — 对入驻企业 9 大类增值服务费 (对外收费, 不向甲方收取)", 8)
    set_widths(ws, [4, 28, 28, 14, 14, 14, 14, 30])

    pos_text = "定位：合资公司在元谷 4# 楼 5F+ 或 5# 楼 5F+ 设立『元谷科技企业服务中心』，对入驻企业及大零号湾区域内潜在企业提供 9 大类标准化 + 项目制服务。该中心收入与森马月费相互独立，由合资公司单独核算。"
    ws.cell(row=2, column=2, value=pos_text).font = NOTE
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 36

    write_header(ws, 4, ["", "服务大类", "代表服务", "收费方式", "起价 (元)", "上限 (元)", "毛利率", "对应需求场景"])
    services = [
        ("一. 注册落户",         "公司注册 / 园区落户 / 政策红利申报代办",      "一次性",        5_000,    30_000, 0.55, "新企业落地 / 跨区迁入"),
        ("二. 财税法",           "代理记账 / 法律顾问 / 税筹与税收优惠申请",    "月费 + 项目制", 1_500,    10_000, 0.40, "中小型潮玩企业刚需"),
        ("三. 知识产权",         "商标 / 专利 / 潮玩 IP 维权",                "单件 + 项目制", 1_500,    150_000, 0.50, "潮玩 IP 高度依赖"),
        ("四. 政府补贴申报",      "高新技术 / 专精特新 / 创新券 / 文创基金",    "项目制 + 提成", 30_000,   150_000, 0.60, "上海科技小巨人 / 闵行专项"),
        ("五. 人才与签证",       "居住证积分 / 留学生落户 / 外籍工作签证",      "单人",          5_000,    20_000, 0.35, "国际化 + 海归人才回流"),
        ("六. 投融资",           "路演对接 / FA 顾问 / 并购顾问",             "月费 + 提成",   5_000,    300_000, 0.65, "追觅基金 + 外部 LP 联动"),
        ("七. 品牌与公关",        "媒体投放 / KOL / 海外推广",                "项目制",       50_000,    1_000_000, 0.30, "潮玩出海 + 北欧外事"),
        ("八. 数字化工具",        "SaaS 包 / AI 设计工作站",                 "月订阅",          800,    8_000,  0.55, "AI 腾讯生态导流 + 共享设计中心"),
        ("九. 培训与认证",        "潮玩产业认证 / 出海实操营",                "单课",          2_000,    20_000, 0.55, "中国动漫集团 / 中百协联办"),
    ]
    for i, (cat, svc, charge, low, high, gm, scene) in enumerate(services, start=5):
        write_row(ws, i, [i - 4, cat, svc, charge, low, high, f"{gm*100:.0f}%", scene], num_cols=[5, 6])

    base = 5 + len(services) + 2
    ws.cell(row=base, column=2, value="三年营收预测 (基础场景, 服务渗透率渐进)").font = H3
    write_header(ws, base + 1, ["", "年度", "入驻企业数", "活跃客户数 (渗透)", "ARPU (元/户/年)", "营收 (元)", "净利率", "净利 (元)"])
    forecasts = [
        ("T+1 年", 60,  30, 50_000,  0.35),
        ("T+2 年", 90,  50, 60_000,  0.38),
        ("T+3 年", 120, 80, 80_000,  0.40),
    ]
    cum_revenue = 0
    cum_net = 0
    for i, (year, total_co, active, arpu, gm) in enumerate(forecasts, start=base + 2):
        rev = active * arpu
        net = round(rev * gm)
        cum_revenue += rev
        cum_net += net
        penetration = active / total_co
        write_row(ws, i, [i - (base + 1), year, total_co, f"{active} ({penetration*100:.0f}%)", arpu, rev, f"{gm*100:.0f}%", net], num_cols=[3, 5, 6, 8])
    write_row(ws, base + 2 + len(forecasts),
              ["合计", "三年累计", "", "", "", cum_revenue, "", cum_net],
              num_cols=[6, 8], bold_total=True)

    note = ("分润机制：服务中心营收完全归合资公司，胡教授团队按 30% 持股享有净利分红 (T+3 年达 67 万元/年)，并通过 CSO 顾问费定额分摊管理工作。该收入流不与甲方月费形成竞争，是合资公司增量价值的核心来源。三年累计营收约 1,090 万元、累计净利约 415 万元、丙方累计获分红约 124 万元。")
    ws.cell(row=base + 4 + len(forecasts), column=2, value=note).font = NOTE
    ws.merge_cells(start_row=base + 4 + len(forecasts), start_column=2, end_row=base + 4 + len(forecasts), end_column=8)
    ws.row_dimensions[base + 4 + len(forecasts)].height = 80

    # ============= Sheet 06 股权分红 + 三年累计 =============
    ws = wb.create_sheet("06 股权分红+三年累计")
    title_bar(ws, "股权分红预测 + 三年累计现金流 (基础场景, 已计入科技企业服务中心)", 7)
    set_widths(ws, [4, 28, 16, 16, 16, 16, 28])

    ws.cell(row=2, column=2,
            value="表 A · 胡教授团队三年现金收入 (5 项, 已去重)").font = H3
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)

    write_header(ws, 3, ["", "项目", "T+1 年 (元)", "T+2 年 (元)", "T+3 年 (元)", "三年合计 (元)", "假设"])
    rows_3y = [
        ("固定月费 Retainer (年化)",       3_360_000, 3_600_000, 3_960_000, "月费 28 → 30 → 33 万"),
        ("招商佣金",                      1_320_000, 2_200_000, 3_300_000, "成交面积 1.2 / 2.0 / 3.0 万㎡"),
        ("活动运营净收入",                2_400_000, 3_000_000, 3_600_000, "活动+1 场/年 + 大赛升级"),
        ("奖项 / 挂牌激励",               1_500_000,   800_000, 1_500_000, "牌照前置, 第二年保持"),
        ("合资公司分红 (30%, 含服务中心)", 1_440_000, 2_700_000, 4_500_000, "净利 30% (服务中心 38% 毛利贡献)"),
    ]
    sums = [0, 0, 0]
    for i, (n, a, b, c, note) in enumerate(rows_3y, start=4):
        s = a + b + c
        sums[0] += a
        sums[1] += b
        sums[2] += c
        write_row(ws, i, [i - 3, n, a, b, c, s, note], num_cols=[3, 4, 5, 6])
    last = 4 + len(rows_3y)
    write_row(ws, last,
              ["合计", "胡教授团队年度收入 (元)", sums[0], sums[1], sums[2], sum(sums), "三年合计现金口径"],
              num_cols=[3, 4, 5, 6], bold_total=True)
    write_row(ws, last + 1,
              ["折合", "胡教授团队年度收入 (万元)",
               round(sums[0] / 10000), round(sums[1] / 10000), round(sums[2] / 10000),
               round(sum(sums) / 10000), f"三年累计 ≈ {round(sum(sums)/10000):,} 万元"],
              num_cols=[3, 4, 5, 6], bold_total=True)

    # ---- 合资公司层面 P&L ----
    base = last + 3
    ws.cell(row=base, column=2, value="表 B · 合资公司 P&L 反推 (含科技企业服务中心)").font = H3
    ws.merge_cells(start_row=base, start_column=2, end_row=base, end_column=7)
    write_header(ws, base + 1, ["", "项目", "T+1 (元)", "T+2 (元)", "T+3 (元)", "三年合计 (元)", "说明"])
    jv_rows = [
        ("合资公司总收入",       13_500_000, 21_000_000, 30_400_000, "招商分润+活动+直营毛利+服务中心营收"),
        ("  其中:服务中心营收",   1_500_000,  3_000_000,  6_400_000, "Sheet 05b"),
        ("合资公司经营成本",      8_700_000, 12_000_000, 15_400_000, "团队+运营+服务中心成本"),
        ("合资公司净利润",        4_800_000,  9_000_000, 15_000_000, "基础场景"),
        ("胡教授团队 30% 分红",   1_440_000,  2_700_000,  4_500_000, "= 上述净利润 × 30%"),
    ]
    for i, (n, a, b, c, note) in enumerate(jv_rows, start=base + 2):
        s = a + b + c
        write_row(ws, i, [i - (base + 1), n, a, b, c, s, note], num_cols=[3, 4, 5, 6])

    # ============= Sheet 07 敏感性分析 =============
    ws = wb.create_sheet("07 敏感性分析")
    title_bar(ws, "敏感性分析: 月费 × 招商成交面积 → 首年总收入 (v1.1)", 9)
    set_widths(ws, [4, 26, 14, 14, 14, 14, 14, 14, 14])

    monthly_options = [150_000, 200_000, 280_000, 350_000, 400_000]
    area_options = [4_000, 8_000, 12_000, 16_000, 20_000, 24_000]

    write_header(ws, 3, ["", "月费 \\ 成交面积 (㎡)"] + [f"{a:,}" for a in area_options])
    for i, m in enumerate(monthly_options, start=4):
        row = [i - 3, f"月费 {m:,} 元"]
        for a in area_options:
            avg_commission = 110
            commission = a * avg_commission
            activity_award = 3_900_000
            service_center = 3_000_000
            dividend = 1_440_000
            total = m * 12 + commission + activity_award + service_center + dividend
            row.append(total)
        write_row(ws, i, row, num_cols=list(range(3, 3 + len(area_options))))

    note = "口径: 单元格 = 月费×12 + 成交面积×110 元/㎡ (佣金平均) + 活动+奖项 (390 万) + 科技企业服务中心 (300 万) + 分红 (144 万)。"
    ws.cell(row=4 + len(monthly_options) + 1, column=2, value=note).font = NOTE
    ws.merge_cells(start_row=4 + len(monthly_options) + 1, start_column=2, end_row=4 + len(monthly_options) + 1, end_column=2 + len(area_options))
    ws.row_dimensions[4 + len(monthly_options) + 1].height = 36

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
