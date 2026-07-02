"""元谷招商排期与服务费用 — 脱敏简版 (发危建平总).

5 个 Sheet:
  01 封面与说明
  02 招商排期 (5 阶段 + 两节点)
  03 月度签约率推进 (动态租金)
  04 动态租金平衡模型
  05 服务费用构成 (无月费)
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("元谷招商排期与服务费用(脱敏版).xlsx")

PRIMARY = "0F244E"; ACCENT = "F27E2D"; GREEN = "1E8E5A"; RED = "C0392B"
LIGHT = "EAEEF5"; ALT = "F4F6FA"

THIN = Side(style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H1 = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
H2 = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
H3 = Font(name="微软雅黑", size=11, bold=True, color=PRIMARY)
BODY = Font(name="微软雅黑", size=10, color="212B42")
NOTE = Font(name="微软雅黑", size=9, italic=True, color="667086")
NUM = Font(name="Consolas", size=10, color="212B42")
FILL_PRIMARY = PatternFill("solid", fgColor=PRIMARY)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_ALT = PatternFill("solid", fgColor=ALT)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header(ws, row, headers, fill=None):
    fill = fill or FILL_PRIMARY
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = H2; c.fill = fill; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 26


def row(ws, r, vals, num_cols=None, bold=False, fill=None):
    num_cols = num_cols or []
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=j, value=v); c.border = BORDER
        if j in num_cols:
            c.number_format = "#,##0"; c.alignment = RIGHT; c.font = NUM
        else:
            c.alignment = LEFT; c.font = BODY
        if bold:
            c.fill = fill or FILL_ACCENT; c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        else:
            c.fill = FILL_ALT if r % 2 == 0 else FILL_LIGHT


def title_bar(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text); c.font = H1; c.fill = FILL_PRIMARY
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 38


def main():
    wb = Workbook()

    # ===== Sheet 1 封面 =====
    ws = wb.active; ws.title = "01 封面与说明"
    title_bar(ws, "元谷 2 万方招商排期与服务费用 · 脱敏版 (呈:危建平总)", 7)
    set_widths(ws, [4, 22, 24, 24, 24, 24, 24])
    notes = [
        ("", ""),
        ("文件版本", "脱敏外发版 v1.0"),
        ("出品方", "胡教授团队 (代表复旦大学住房政策研究中心、上海市科技企业联合会)"),
        ("服务方式", "第三方专业招商运营服务 (不涉及新设公司)"),
        ("业务范围", "元谷 4# 楼 5F+ + 5# 楼 5F+ 共约 2 万㎡产业研发办公"),
        ("协议期限", "24 个月 (首期)"),
        ("Sheet 索引", "02 招商排期 / 03 月度签约率 / 04 动态租金模型 / 05 服务费用构成"),
        ("", ""),
        ("【两个硬节点】", ""),
        ("节点一", "2026/9/30 → 累计签约 2,000㎡"),
        ("节点二", "2027/5/1 → 项目开业, 累计签约率 50%+"),
        ("全周期目标", "出租率 90%+"),
        ("", ""),
        ("【动态租金平衡】", ""),
        ("招商期 (满租前)", "平均 1.5 - 1.8 元/㎡/天 (先低价招满)"),
        ("长期稳定 (满租后)", "2.0 - 2.5 元/㎡/天 (后抬升)"),
        ("保底锚点", "2.0 元/㎡/天 含物业 (危建平总已签)"),
        ("", ""),
        ("【服务费用 · 无月费】", ""),
        ("招商佣金", "实际成交年租金 1.5 / 1.75 / 2.0 个月 (核心收入)"),
        ("沙龙执行费", "6 场打包 30 万 (单独收取不分润, 一次性付清)"),
        ("挂牌费", "10 万/项 可选 (最多 5 项, 挂牌前一次性)"),
        ("超额奖励", "出租率 ≥ 90% → 20 万 (正常分期)"),
    ]
    for i, (k, v) in enumerate(notes, start=2):
        if not k and not v: continue
        a = ws.cell(row=i, column=2, value=k); a.font = H3 if k.startswith("【") else BODY; a.alignment = LEFT
        b = ws.cell(row=i, column=3, value=v); b.font = BODY; b.alignment = LEFT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)

    # ===== Sheet 2 招商排期 =====
    ws = wb.create_sheet("02 招商排期")
    title_bar(ws, "招商排期 — 5 个阶段 + 两个硬节点", 6)
    set_widths(ws, [4, 18, 22, 18, 24, 30])
    header(ws, 3, ["", "阶段", "时间窗口", "累计签约目标", "核心动作", "里程碑"])
    stages = [
        ("阶段 0 准备", "启动 - 第 3 周", "0", "团队就位 / 资源接入 / 物料", "启动 + 政府首访"),
        ("阶段 1 抢节点", "第 3 周 - 9/30", "2,300㎡(含擦边球)", "牌照锚定 + 直播基地 + 设计中心", "★ 9/30 节点达成"),
        ("阶段 2 加速", "10/1 - 12/31", "5,500㎡ (27%)", "爬楼大数据全开 + 沙龙 + 挂牌", "资本招商落地"),
        ("阶段 3 开业", "1/1 - 5/1", "10,400㎡ (52%)", "大客户决战 + 沙龙 + 挂牌", "★ 5/1 开业典礼"),
        ("阶段 4 满租", "5/2 - 次年", "18,000㎡+ (90%)", "沙龙 IP 化 + 启动抬租", "满租 + 抬租"),
    ]
    for i, s in enumerate(stages, start=4):
        row(ws, i, [i - 3] + list(s))

    # ===== Sheet 3 月度签约率 =====
    ws = wb.create_sheet("03 月度签约率")
    title_bar(ws, "月度签约率推进 (招商期按 1.7 元/㎡/天 估算)", 7)
    set_widths(ws, [4, 16, 16, 16, 14, 18, 26])
    header(ws, 3, ["", "月份", "本月新增(㎡)", "累计签约(㎡)", "签约率", "累计年租金(元)", "里程碑"])
    monthly = [
        ("第 1 月", 300, 300, "1.5%", "战队就位 + 物料"),
        ("第 2 月", 800, 1100, "5.5%", "牌照锚定 + 大客户接触"),
        ("第 3 月 (9/30)", 1200, 2300, "11.5%", "★ 硬节点一达成"),
        ("第 4 月", 1000, 3300, "16.5%", "爬楼大数据启动"),
        ("第 5 月", 1000, 4300, "21.5%", "沙龙 + 挂牌"),
        ("第 6 月", 1200, 5500, "27.5%", "Q4 收官"),
        ("第 7 月", 1200, 6700, "33.5%", "新年首单"),
        ("第 8 月", 1000, 7700, "38.5%", "—"),
        ("第 9 月", 1200, 8900, "44.5%", "沙龙"),
        ("第 10 月 (5/1)", 1500, 10400, "52.0%", "★ 硬节点二达成"),
        ("第 11 月", 1500, 11900, "59.5%", "开业 + 启动抬租"),
        ("第 12 月", 1300, 13200, "66.0%", "持续推进"),
    ]
    unit = 1.7 * 365  # 招商期年租金/㎡
    for i, (m, add, accum, rate, note) in enumerate(monthly, start=4):
        row(ws, i, [i - 3, m, add, accum, rate, round(accum * unit), note], num_cols=[3, 4, 6])
    ws.cell(row=4 + len(monthly) + 1, column=2,
            value="说明:招商期按 1.7 元/㎡/天 估算;满租后启动抬租至 2.0-2.5 元, 累计年租金将进一步提升。").font = NOTE
    ws.merge_cells(start_row=4 + len(monthly) + 1, start_column=2, end_row=4 + len(monthly) + 1, end_column=7)

    # ===== Sheet 4 动态租金模型 =====
    ws = wb.create_sheet("04 动态租金模型")
    title_bar(ws, "动态租金平衡模型 (先低价招满 · 后抬升)", 6)
    set_widths(ws, [4, 20, 18, 20, 20, 24])
    header(ws, 3, ["", "阶段", "日租金(元/㎡/天)", "年租金(元/㎡)", "2 万㎡ 年化(元)", "策略说明"])
    rents = [
        ("招商初期 (0-9 月)", 1.5, "先低价快速招满"),
        ("加速期 (9-18 月)", 1.7, "维持低价冲量"),
        ("满租稳定 (18-24 月)", 2.1, "满租后启动抬升"),
        ("长期稳定 (24 月+)", 2.4, "推向稳定上限"),
        ("保底锚点 (危建平总已签)", 2.0, "含物业费, 长期底线"),
    ]
    for i, (stage, daily, note) in enumerate(rents, start=4):
        annual = round(daily * 365)
        total = round(annual * 20000)
        bold = stage.startswith("保底")
        row(ws, i, [i - 3, stage, daily, annual, total, note], num_cols=[3, 4, 5], bold=bold, fill=FILL_ACCENT if bold else None)
    ws.cell(row=4 + len(rents) + 1, column=2,
            value="逻辑:招商期 1.5-1.8 元先招满形成产业氛围;满租后借势逐步抬升至长期稳定 2.0-2.5 元, 长期锚定保底 2.0 元(含物业), 实现项目方租金动态增收。").font = NOTE
    ws.merge_cells(start_row=4 + len(rents) + 1, start_column=2, end_row=4 + len(rents) + 1, end_column=6)
    ws.row_dimensions[4 + len(rents) + 1].height = 40

    # 满租收益对照
    base = 4 + len(rents) + 3
    ws.cell(row=base, column=2, value="满租期年租金收益对照 (2 万㎡)").font = H3
    ws.merge_cells(start_row=base, start_column=2, end_row=base, end_column=6)
    header(ws, base + 1, ["", "租金水平", "年租金(元/㎡)", "年化租金(元)", "说明", ""])
    full = [
        ("招商期均值 1.5 元", 1.5, "快速招满"),
        ("招商期均值 1.8 元", 1.8, "招商后期"),
        ("保底 2.0 元(含物业)", 2.0, "长期底线"),
        ("长期稳定 2.5 元", 2.5, "稳定上限"),
    ]
    for i, (label, daily, note) in enumerate(full, start=base + 2):
        annual = round(daily * 365)
        row(ws, i, [i - (base + 1), label, annual, round(annual * 20000), note, ""], num_cols=[3, 4])

    # ===== Sheet 5 服务费用构成 =====
    ws = wb.create_sheet("05 服务费用构成")
    title_bar(ws, "服务费用构成 (无月费 · 按成果付费)", 6)
    set_widths(ws, [4, 22, 30, 22, 20, 8])
    header(ws, 3, ["", "费用类别", "标准", "支付方式", "估算金额", ""])
    fees = [
        ("月费", "★ 无 (不设固定月费, 不签对赌)", "—", "0"),
        ("招商佣金(核心)", "实际成交年租金的 1.5 / 1.75 / 2.0 个月(按面积)", "起租后 30 日内", "约 180-240 万"),
        ("沙龙执行费", "6 场打包 30 万(单独收取, 不分润)", "一次性付清(或分两次)", "30 万"),
        ("挂牌费(可选)", "10 万/项 × 选定项数(最多 5 项)", "挂牌前一次性付清", "0-50 万"),
        ("超额奖励(适度)", "出租率 ≥ 90% → 20 万", "正常分期支付", "20 万"),
    ]
    for i, f in enumerate(fees, start=4):
        row(ws, i, [i - 3] + list(f) + [""])
    ws.cell(row=4 + len(fees) + 1, column=2,
            value="合计区间(满租 + 全部挂牌):约 280-340 万元 (24 个月)。其中招商佣金随动态租金浮动:招商期低价时佣金较低, 抬租后续签佣金提升。").font = NOTE
    ws.merge_cells(start_row=4 + len(fees) + 1, start_column=2, end_row=4 + len(fees) + 1, end_column=6)
    ws.row_dimensions[4 + len(fees) + 1].height = 40

    base = 4 + len(fees) + 3
    ws.cell(row=base, column=2, value="对项目方价值 (我方按成果收费, 利益绑定)").font = H3
    ws.merge_cells(start_row=base, start_column=2, end_row=base, end_column=6)
    header(ws, base + 1, ["", "项目", "金额", "说明", "", ""])
    value = [
        ("满租期年租金(2.0 元保底)", "约 1,460 万/年", "2 万㎡ × 2.0 元含物业 × 365"),
        ("满租期年租金(2.5 元上限)", "约 1,825 万/年", "2 万㎡ × 2.5 元 × 365"),
        ("我方 24 月服务费", "约 280-340 万", "完全按招商成果收取"),
        ("资产增值", "数千万级", "产业认证 + TOD 板块溢价(不计入)"),
    ]
    for i, (k, v, note) in enumerate(value, start=base + 2):
        row(ws, i, [i - (base + 1), k, v, note, "", ""])

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
