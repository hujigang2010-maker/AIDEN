"""元谷招商排期总表 — Excel.

8 个 Sheet:
  01 封面与说明
  02 总排期 (5 个阶段)
  03 月度签约率推进 (12 个月明细)
  04 9/30 节点深拆 (2,000㎡ 来源)
  05 双方分工矩阵 (RACI)
  06 12 月费用明细
  07 KPI 与对赌
  08 风险与对冲
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("元谷招商排期总表.xlsx")

PRIMARY = "142C5E"; ACCENT = "F27E2D"; RED = "C0392B"; GOLD = "C8993D"
LIGHT = "EAEEF5"; ALT = "F8F9FB"; GREEN = "27AE60"

THIN = Side(style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H1 = Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
H2 = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
H3 = Font(name="微软雅黑", size=11, bold=True, color=PRIMARY)
BODY = Font(name="微软雅黑", size=10, color="1F2A44")
NOTE = Font(name="微软雅黑", size=9, italic=True, color="55607A")
NUM = Font(name="Consolas", size=10, color="1F2A44")

FILL_PRIMARY = PatternFill("solid", fgColor=PRIMARY)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT)
FILL_RED = PatternFill("solid", fgColor=RED)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_ALT = PatternFill("solid", fgColor=ALT)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row, headers, fill=None):
    fill = fill or FILL_PRIMARY
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = H2; c.fill = fill; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 26


def write_row(ws, row, vals, num_cols=None, bold_total=False, fill=None):
    num_cols = num_cols or []
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.border = BORDER
        if j in num_cols:
            c.number_format = "#,##0"
            c.alignment = RIGHT
            c.font = NUM
        else:
            c.alignment = LEFT
            c.font = BODY
        if bold_total:
            c.fill = fill or FILL_ACCENT
            c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        else:
            c.fill = FILL_ALT if row % 2 == 0 else FILL_LIGHT


def title_bar(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = H1; c.fill = FILL_PRIMARY
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 38


def main():
    wb = Workbook()

    # ===== Sheet 1 封面 =====
    ws = wb.active; ws.title = "01 封面与说明"
    title_bar(ws, "元谷 2 万方 IP+AI 双轨招商 — 排期总表 v1.0", 8)
    set_widths(ws, [4, 22, 24, 24, 24, 24, 24, 24])

    notes = [
        ("", ""),
        ("文件版本", "v1.0 (基于 2026/6 会议纪要 + IP+AI 双轨战略调整)"),
        ("业务范围", "元谷 4# 楼 5F+ + 5# 楼 5F+ 共约 2 万㎡产业研发办公"),
        ("服务期限", "24 个月 (首期, 2026/6 - 2028/6)"),
        ("Sheet 索引", "02 总排期 / 03 月度签约率 / 04 9-30 节点深拆 / 05 RACI / 06 费用 / 07 KPI / 08 风险"),
        ("", ""),
        ("【两个硬节点】", ""),
        ("节点一", "2026/9/30 → 2,000㎡ 签约 (T+100 天)"),
        ("节点二", "2027/5/1 → 项目开业, 50%+ 签约 (T+314 天)"),
        ("全周期目标", "2028/5/1 → 90%+ 满租 (T+679 天)"),
        ("", ""),
        ("【战略主轴调整】", ""),
        ("原方向", "纯二次元 (政府/资本不投, 招商面窄, 难以达成节点)"),
        ("新方向", "★ IP + AI 双轨 (AI 主轴拿政策/基金/税收, IP 副轴承接潮玩 + 文化)"),
        ("先例", "杨浦区已有 1 万方 'AI + IP' 项目通过先例"),
        ("", ""),
        ("【行业转化率新基准】", ""),
        ("当前", "110 : 1 (招 1 单要见 110 家潜客)"),
        ("演变", "45:1 → 68:1 → 110:1 (AI 替代办公 + 经济下行)"),
        ("应对", "★ 招 1 家 2,000-20,000㎡ 大户 ≫ 招 10 家 200㎡ 小户"),
    ]
    for i, (k, v) in enumerate(notes, start=2):
        if not k and not v: continue
        a = ws.cell(row=i, column=2, value=k); a.font = H3 if k.startswith("【") else BODY; a.alignment = LEFT
        b = ws.cell(row=i, column=3, value=v); b.font = BODY; b.alignment = LEFT
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)

    # ===== Sheet 2 总排期 =====
    ws = wb.create_sheet("02 总排期 5 阶段")
    title_bar(ws, "总排期 — 5 个阶段 (按 2026/6/22 起算)", 7)
    set_widths(ws, [4, 18, 22, 16, 18, 30, 30])
    write_header(ws, 3, ["", "阶段", "时间窗口", "天数 (累计)", "阶段目标 (累计签约)", "核心动作", "里程碑"])
    stages = [
        ("阶段 0 准备", "2026/6/22 - 2026/7/15", "T+23 天", "0",                  "团队就位 / 资源接入 / 物料 1.0", "MOU 签订 + 闵行区科委首次拜访"),
        ("阶段 1 抢节点", "2026/7/15 - 2026/9/30", "T+100 天", "★ 2,300㎡ (含擦边球)", "L1 牌照锚定 + L2 资本招商 + 直播基地 + 设计中心", "★ 硬节点一 2,000㎡ 达成"),
        ("阶段 2 加速",  "2026/10/1 - 2026/12/31", "T+192 天", "5,500㎡ (≈ 27%)", "L3 爬楼大数据全开 + 沙龙 #2/#3 + 挂牌 ③/④",  "L2 资本招商累计 ≥ 3 家"),
        ("阶段 3 开业",  "2027/1/1 - 2027/5/1",  "T+314 天", "★ 10,400㎡ (≈ 52%)", "大客户决战 + 沙龙 #4/#5/#6 + 挂牌 ⑤",        "★ 硬节点二 5/1 开业典礼"),
        ("阶段 4 满租",  "2027/5/2 - 2028/5/1",  "T+679 天", "★ 18,000㎡+ (≥ 90%)", "沙龙 IP 化 + 续约 + 5.2 万方二期承接",       "全周期满租, 续约启动"),
    ]
    for i, (n, t, d, target, actions, milestone) in enumerate(stages, start=4):
        write_row(ws, i, [i - 3, n, t, d, target, actions, milestone])

    # ===== Sheet 3 月度签约率 =====
    ws = wb.create_sheet("03 月度签约率推进")
    title_bar(ws, "月度签约率推进表 (12 个月明细, 按年租金 803 元/㎡ 估算)", 8)
    set_widths(ws, [4, 16, 16, 16, 14, 18, 18, 24])
    write_header(ws, 3, ["", "月份", "本月新增 (㎡)", "累计签约 (㎡)", "累计签约率", "本月年租金 (元)", "累计年租金 (元)", "里程碑 / 备注"])
    monthly = [
        ("2026/7",  300,   300,    "1.5%",  "战队就位 + 物料 1.0"),
        ("2026/8",  800,   1100,   "5.5%",  "L1 牌照锚定 + 大客户接触"),
        ("2026/9",  1200,  2300,   "11.5%", "★ 9/30 硬节点达成"),
        ("2026/10", 1000,  3300,   "16.5%", "L3 爬楼启动 + 沙龙 #2 出海"),
        ("2026/11", 1000,  4300,   "21.5%", "沙龙 #3 投融资 + 挂牌 ③"),
        ("2026/12", 1200,  5500,   "27.5%", "Q4 攻坚 + 挂牌 ④"),
        ("2027/1",  1200,  6700,   "33.5%", "新年首单 + 沙龙 #4"),
        ("2027/2",  1000,  7700,   "38.5%", "—"),
        ("2027/3",  1200,  8900,   "44.5%", "沙龙 #5"),
        ("2027/4",  1500,  10400,  "52.0%", "★ 5/1 节点提前达成"),
        ("2027/5",  1500,  11900,  "59.5%", "★ 开业典礼 + 挂牌 ⑤"),
        ("2027/6",  1300,  13200,  "66.0%", "Q3 推进"),
    ]
    annual_unit = 803
    cum = 0
    for i, (m, add, accum, rate, note) in enumerate(monthly, start=4):
        annual_month = add * annual_unit
        annual_cum = accum * annual_unit
        write_row(ws, i, [i - 3, m, add, accum, rate, annual_month, annual_cum, note],
                  num_cols=[3, 4, 6, 7])

    # ===== Sheet 4 9/30 节点深拆 =====
    ws = wb.create_sheet("04 9-30 节点深拆")
    title_bar(ws, "9/30 节点深拆 — 2,000㎡ 怎么来 (T+100 天必达)", 7)
    set_widths(ws, [4, 28, 16, 14, 18, 18, 28])
    write_header(ws, 3, ["", "签约来源", "面积估算 (㎡)", "落地概率", "对节点贡献期望值", "责任方", "兜底方案 / 备注"])
    sources = [
        ("1-2 家中大型 AI/IP 客户",       1500, 0.6,  "甲乙联合", "若未签, 启动直播 + 设计中心兜底"),
        ("共享直播基地 (4# 楼 4F)",       900,  0.9, "胡教授主导 + 森马场地", "已具备意愿, 7 月启动硬件"),
        ("AI 共享设计中心 (4# 楼 5F)",    650,  0.8, "胡教授 + 腾讯 + 上海交大", "联动腾讯/AI 腾讯生态联合挂牌"),
        ("产业服务中心入驻 (5# 楼 5F)",   400,  1.0, "★ 乙方自营", "100% 计入"),
        ("首批小型 AI/IP 客户 3-5 家",   1050, 0.5,  "胡教授招商团队", "L4 沙龙 + 福布斯/腾讯导流"),
    ]
    total_expected = 0
    for i, (n, area, prob, owner, note) in enumerate(sources, start=4):
        expected = round(area * prob)
        total_expected += expected
        write_row(ws, i, [i - 3, n, area, f"{prob*100:.0f}%", expected, owner, note], num_cols=[3, 5])
    write_row(ws, 4 + len(sources), ["合计", "中性场景期望面积", "", "", total_expected, "", "实际计入 ≥ 2,000㎡ 概率 > 95%"], num_cols=[5], bold_total=True)

    # ===== Sheet 5 RACI 双方分工 =====
    ws = wb.create_sheet("05 双方分工 RACI")
    title_bar(ws, "双方分工矩阵 (RACI: R=主责 / A=审批 / C=咨询 / I=知会)", 6)
    set_widths(ws, [4, 32, 20, 20, 20, 28])
    write_header(ws, 3, ["", "工作项", "森马侧 (威总 / 发哥 / 周志超)", "胡教授团队 (胡教授 + 2 人战队)", "联合", "备注"])
    raci = [
        ("战略制定 (IP+AI 双轨)",     "A 威总最终拍板",   "★ R 胡教授起草战略",     "C 发哥提建议",  "已基本对齐"),
        ("合资公司 / 合作协议",       "★ R 发哥 + 法务",  "A 胡教授确认条款",      "—",            "—"),
        ("产业招商 (2 万方)",         "I 知会",          "★ R 胡教授团队",        "C 大客户拜访",  "胡教授全权"),
        ("商业招商 (5.2 万方)",       "★ R 发哥/周志超",  "—",                     "—",            "本方案不涉及"),
        ("政府对接 (闵行区科委/商务委)", "C 提供森马背书",  "★ R 胡教授主导",        "联合拜访",      "胡教授有市/区科委背景"),
        ("腾讯算力补贴 / CVC 基金",   "I",               "★ R 胡教授对接",        "—",            "胡教授资源"),
        ("欧洲驻沪机构对接",          "I",               "★ R 胡教授对接",        "—",            "德/以/北欧"),
        ("福布斯采访 / 榜单",          "C",               "★ R 胡教授对接",        "联合发布",      "胡教授资源"),
        ("AI 潮玩产业基地挂牌",        "A 森马挂牌仪式",  "★ R 牵线 + 仪式",       "联合发布",      "中国动漫集团"),
        ("潮玩次元商业专委会挂牌",     "A",               "★ R",                   "联合发布",      "中国百货商业协会"),
        ("复旦住房政策中心挂牌",       "I",               "★ R 胡教授本人渠道",    "联合发布",      "胡教授任职"),
        ("上海市科技企业联合会挂牌",   "I",               "★ R 胡教授本人渠道",    "联合发布",      "胡教授任职"),
        ("福布斯产业影响力奖挂牌",     "A 高层出席",      "★ R 胡教授对接",        "联合发布",      "—"),
        ("6 场产业沙龙",              "C 品牌联合宣传",  "★ R 全部执行",          "森马高层出席",  "每月 1 场"),
        ("PR 传播 / 主流媒体",         "C 森马商业部",    "★ R 政府官网/新华网",   "联合发声",      "—"),
        ("5/1 开业典礼",              "★ R 森马集团 + 政府高层", "C 胡教授+复旦+科企联出席", "联合主办", "—"),
        ("24 月报告 / 续约谈判",       "A 续约审批",      "★ R 胡教授起草报告",    "—",            "T+24 月节点"),
    ]
    for i, (item, semir, hu, joint, note) in enumerate(raci, start=4):
        write_row(ws, i, [i - 3, item, semir, hu, joint, note])

    # ===== Sheet 6 12 月费用 =====
    ws = wb.create_sheet("06 12 月费用明细")
    title_bar(ws, "12 个月费用明细 (向森马提案口径)", 6)
    set_widths(ws, [4, 28, 18, 22, 22, 32])
    write_header(ws, 3, ["", "费用类别", "金额 (元)", "支付节奏", "森马承担比例", "备注 / 计算逻辑"])
    fees = [
        ("团队月费 12 万 × 12",          1_440_000, "按月预付",          "100% 森马",  "2 人配置 + CSO 顾问"),
        ("招商佣金 (累计 11,900㎡)",     1_400_000, "起租后 30 日内",     "100% 森马",  "11,900 × 803 × 1.75/12 ≈ 140 万"),
        ("5 项挂牌 (一次性)",            1_500_000, "挂牌 30 日内",      "100% 森马",  "30 万/项 × 5 项"),
        ("6 场沙龙执行",                  300_000,  "按场结算",          "100% 森马",  "5 万/场 × 6 场"),
        ("5/1 开业典礼",                  800_000,  "开业前 30 日内",     "100% 森马",  "森马高端开业仪式"),
        ("装修补贴 + 物料 + 设计",         500_000,  "实报实销",          "100% 森马",  "首批 5 项业态 + 装修补贴包"),
        ("接待 + 交通 + 政府关系",          200_000,  "月度报销",          "100% 森马",  "差旅 / 礼物 / 媒体投放"),
        ("直播 + AI 设计中心硬件",         700_000,  "一次性",            "100% 森马",  "9/30 擦边球必备"),
        ("12 月合计",                    6_840_000, "", "", "向森马提案口径"),
    ]
    for i, row in enumerate(fees, start=4):
        bold = (i == 4 + len(fees) - 1)
        write_row(ws, i, [i - 3 if not bold else "合计"] + list(row), num_cols=[3], bold_total=bold)

    # 配套 ROI
    ws.cell(row=4 + len(fees) + 2, column=2, value="同期甲方现金回收 (12 个月口径)").font = H3
    ws.merge_cells(start_row=4 + len(fees) + 2, start_column=2, end_row=4 + len(fees) + 2, end_column=6)
    write_header(ws, 4 + len(fees) + 3, ["", "类别", "金额 (元)", "说明", "", ""])
    cashflows = [
        ("租金 (12 个月累计)",     9_560_000, "签约面积按月渐进进入起租, 取累计签约 × 中位起租周期"),
        ("物业费 (10 元/㎡/月)",   900_000,   "签约面积渐进 × 10/月 × 12 月"),
        ("12 月甲方现金回收合计",  10_460_000, "★ 单 12 月已现金回正 (1,046 万 vs 684 万)"),
    ]
    for i, (n, v, note) in enumerate(cashflows, start=4 + len(fees) + 4):
        bold = (i == 4 + len(fees) + 4 + len(cashflows) - 1)
        write_row(ws, i, [i - (4 + len(fees) + 3) if not bold else "合计", n, v, note, "", ""], num_cols=[3], bold_total=bold)

    # ===== Sheet 7 KPI 与对赌 =====
    ws = wb.create_sheet("07 KPI 与对赌")
    title_bar(ws, "KPI 与业绩对赌 (写进协议)", 6)
    set_widths(ws, [4, 32, 18, 20, 20, 32])
    write_header(ws, 3, ["", "KPI 指标", "T+6 月目标", "T+12 月目标", "T+24 月目标", "未达标处理"])
    kpis = [
        ("累计签约面积 (㎡)",        "4,000",       "10,000",     "18,000+",     "差额 ≥ 20% → 月费保留 50%"),
        ("挂牌数量 (项)",            "★ 2 (AI 潮玩基地 + 潮玩专委)", "★ 4",       "★ 5 (含福布斯)", "首挂未完成 → 推延奖励"),
        ("产业沙龙场次",             "3 场",         "6 场",       "12 场+",     "单场少于 30 客户 → 执行费扣 50%"),
        ("CVC 基金落地企业 (家)",    "1",            "3",          "5+",         "未达标 → 切换 IP 主轴比例"),
        ("腾讯算力补贴企业 (家)",     "2",            "5",          "10+",       "—"),
        ("政府对接 (闵行科委 + 商务委 + 街道)", "2 次", "6 次",        "12 次+",     "—"),
        ("媒体声量 (亿次曝光)",       "0.5",          "2",          "5+",         "—"),
        ("9/30 硬节点 (2,000㎡)",     "★ 必达",       "—",          "—",          "★ 未达 → 协议解除选项"),
        ("5/1 开业 (50% 签约)",       "—",           "★ 必达",      "—",          "★ 未达 → 阶段 4 延长"),
    ]
    for i, row in enumerate(kpis, start=4):
        write_row(ws, i, [i - 3] + list(row))

    # 超额奖励
    ws.cell(row=4 + len(kpis) + 2, column=2, value="超额奖励").font = H3
    ws.merge_cells(start_row=4 + len(kpis) + 2, start_column=2, end_row=4 + len(kpis) + 2, end_column=6)
    write_header(ws, 4 + len(kpis) + 3, ["", "条件", "奖励金额 (元)", "支付节奏", "", ""])
    bonuses = [
        ("24 月满租率 ≥ 95%",       1_000_000, "T+24 月一次性"),
        ("9/30 节点提前达成 (T+90 天前)", 200_000, "达成确认后 15 日内"),
        ("5/1 开业签约率 ≥ 60%",    500_000, "5/1 当月内"),
    ]
    for i, (n, v, node) in enumerate(bonuses, start=4 + len(kpis) + 4):
        write_row(ws, i, [i - (4 + len(kpis) + 3), n, v, node, "", ""], num_cols=[3])

    # ===== Sheet 8 风险与对冲 =====
    ws = wb.create_sheet("08 风险与对冲")
    title_bar(ws, "风险识别与对冲机制", 6)
    set_widths(ws, [4, 32, 20, 18, 18, 32])
    write_header(ws, 3, ["", "风险事项", "影响", "概率", "影响度", "对冲机制"])
    risks = [
        ("9/30 节点 2,000㎡ 未达",         "森马失信 + 项目延期",   "中",   "高",   "★ 擦边球 (直播 + 设计中心) 兜底 + KPI 对赌"),
        ("5/1 开业 50% 未达",              "开业氛围不足",         "中低", "中高", "★ 阶段 4 满租期延长 + 沙龙 IP 化加速"),
        ("110:1 → 130:1 (转化率恶化)",     "招商节奏放缓 30%",     "中",   "中",   "★ 团队扩 1 人 + 月费上调至 18 万"),
        ("闵行区政策变化",                  "补贴/牌照不落地",       "低",   "中高", "★ 备份: 杨浦区 / 漕河泾 / 张江"),
        ("AI 行业资本退潮",                 "AI 客户付款下降",       "中",   "中",   "★ 切换 IP 主轴比例 + 出海企业引入"),
        ("森马商业部协同延误",              "1#/6# 商业氛围不足",   "中",   "低",   "★ 5.2 万方与 2 万方独立排期"),
        ("CVC 基金落地延后",                 "L2 资本招商无支撑",    "低",   "中",   "★ 备份: 追觅 / 招行 / 长江"),
        ("腾讯算力补贴政策变化",             "AI 招商话术失效",      "低",   "中",   "★ 切换其他云厂商: 阿里 / 华为"),
        ("欧洲驻沪机构对接受阻",             "国际化故事弱化",        "低",   "低",   "★ 备份: 日韩 / 东南亚"),
        ("胡教授时间投入不足",               "战略执行力打折",        "中",   "高",   "★ CSO 任期不少于 5 年 + 期权绑定"),
    ]
    for i, row in enumerate(risks, start=4):
        write_row(ws, i, [i - 3] + list(row))

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
