"""
生成"附件3：应聘人员登记表"（奇瑞控股公司）Excel 文件。

数据来源：汤先生（胡继刚）的简历。凡简历中不明确的信息一律留空。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 样式定义
# ---------------------------------------------------------------------------
MAX_COL = 11  # A-K

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

LABEL_FILL = PatternFill("solid", fgColor="E8EEF7")
HEADER_FILL = PatternFill("solid", fgColor="C9D6EA")
TITLE_FONT = Font(name="宋体", size=16, bold=True)
HEADER_FONT = Font(name="宋体", size=11, bold=True)
LABEL_FONT = Font(name="宋体", size=10, bold=True)
VALUE_FONT = Font(name="宋体", size=10)
SMALL_FONT = Font(name="宋体", size=9)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "应聘人员登记表"

# 列宽
widths = [13, 15, 15, 13, 15, 12, 11, 8, 11, 8, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w


def cell(row, col, value="", *, font=VALUE_FONT, align=LEFT, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    if fill is not None:
        c.fill = fill
    return c


def merge(row1, col1, row2, col2, value="", *, font=VALUE_FONT, align=LEFT, fill=None):
    ws.merge_cells(start_row=row1, start_column=col1, end_row=row2, end_column=col2)
    return cell(row1, col1, value, font=font, align=align, fill=fill)


def label(row, col, value):
    return cell(row, col, value, font=LABEL_FONT, align=CENTER, fill=LABEL_FILL)


def mlabel(row1, col1, row2, col2, value):
    return merge(row1, col1, row2, col2, value, font=LABEL_FONT, align=CENTER, fill=LABEL_FILL)


def header(row1, col1, row2, col2, value):
    return merge(row1, col1, row2, col2, value, font=HEADER_FONT, align=CENTER, fill=HEADER_FILL)


r = 1

# ---------------------------------------------------------------------------
# 标题与说明
# ---------------------------------------------------------------------------
merge(r, 1, r, MAX_COL, "附件3：应聘人员登记表", font=TITLE_FONT, align=CENTER)
ws.row_dimensions[r].height = 30
r += 1
merge(r, 1, r, MAX_COL, "（本表由参加应聘人员详细填写）", font=SMALL_FONT, align=CENTER)
r += 1
merge(r, 1, r, MAX_COL,
      "本人自愿应聘奇瑞控股公司，保证以下内容属实，接受奇瑞控股公司对本人的背景调查，并承担提供虚假信息的后果。",
      font=SMALL_FONT, align=LEFT)
ws.row_dimensions[r].height = 24
r += 1
merge(r, 1, r, MAX_COL, "我是第 ____ 次参加奇瑞控股应聘活动。", font=SMALL_FONT, align=LEFT)
r += 1
merge(r, 1, r, MAX_COL, "承诺人：___________________          日期：_____________________",
      font=SMALL_FONT, align=LEFT)
r += 1
merge(r, 1, r, MAX_COL,
      "填写说明：1、本表对招聘录用有重要参考价值，务必认真填写，不要遗漏；2、本表提交后，概不退回。",
      font=SMALL_FONT, align=LEFT)
ws.row_dimensions[r].height = 24
r += 1

# ---------------------------------------------------------------------------
# 基本信息 + 照片
# ---------------------------------------------------------------------------
photo_top = r
# 应聘公司/职位/待遇
label(r, 1, "应聘公司/板块")
merge(r, 2, r, 3, "")
label(r, 4, "应聘职位")
merge(r, 5, r, 6, "政府关系 / 会务会展执行")
label(r, 7, "希望待遇")
merge(r, 8, r, 10, "30-60K×12薪（面议）")
r += 1

# 姓名/身份证/出生/性别
label(r, 1, "姓　名")
cell(r, 2, "胡继刚", align=CENTER)
label(r, 3, "身份证号码/护照号")
merge(r, 4, r, 5, "220283198704250614", align=CENTER)
label(r, 6, "出生年月")
cell(r, 7, "1987.4", align=CENTER)
label(r, 8, "性　别")
merge(r, 9, r, 10, "男", align=CENTER)
r += 1

# 婚姻/政治面貌/病史/身高/体重
label(r, 1, "婚姻状况")
cell(r, 2, "已婚", align=CENTER)
label(r, 3, "政治面貌")
cell(r, 4, "九三学社（民主党派）", align=CENTER)
label(r, 5, "传染病史/其它病史")
cell(r, 6, "无", align=CENTER)
label(r, 7, "身高(cm)")
cell(r, 8, "180", align=CENTER)
label(r, 9, "体重(kg)")
cell(r, 10, "75", align=CENTER)
r += 1

# 外语/国籍/民族/籍贯
label(r, 1, "外语水平")
cell(r, 2, "雅思6.0", align=CENTER)
label(r, 3, "国　籍")
cell(r, 4, "中国", align=CENTER)
label(r, 5, "民　族")
cell(r, 6, "汉族", align=CENTER)
label(r, 7, "籍　贯")
merge(r, 8, r, 10, "青岛", align=CENTER)
r += 1

# 应聘状况
label(r, 1, "应聘状况")
merge(r, 2, r, 10, "☑在职    □退伍    □失业    □退休    □其它", align=LEFT)
r += 1

# 教育方式
label(r, 1, "教育方式")
merge(r, 2, r, 10, "☑全日制普通高校    □自学考试    □成人高校    □网络学院    □电视大学", align=LEFT)
photo_bottom = r
# 照片区（右侧竖跨）
merge(photo_top, 11, photo_bottom, 11, "照\n片", align=CENTER, font=LABEL_FONT, fill=LABEL_FILL)
r += 1

# 最高学历
label(r, 1, "最高学历/学位")
cell(r, 2, "硕士研究生", align=CENTER)
label(r, 3, "所学专业/毕业时间")
merge(r, 4, r, 5, "工商管理（财务金融方向）/ 2021.6", align=CENTER)
label(r, 6, "毕业院校")
merge(r, 7, r, 11, "复旦大学", align=CENTER)
r += 1

# 参加工作时间/现工作单位/资格证书/职称
label(r, 1, "参加工作时间（第一次）")
cell(r, 2, "2011", align=CENTER)
label(r, 3, "现工作单位及岗位")
merge(r, 4, r, 5, "复旦大学住房政策研究中心 / 秘书长", align=CENTER)
label(r, 6, "资格证书")
cell(r, 7, "高级工程师", align=CENTER)
label(r, 8, "职称")
merge(r, 9, r, 11, "副教授级高级工程师", align=CENTER)
r += 1

# 联系电话/手机/E-mail
label(r, 1, "联系电话")
cell(r, 2, "13262607888", align=CENTER)
label(r, 3, "手机")
cell(r, 4, "13262607888", align=CENTER)
label(r, 5, "E-mail")
merge(r, 6, r, 11, "262782809@qq.com", align=CENTER)
r += 1

# 户口所在地
label(r, 1, "户口所在地")
merge(r, 2, r, 11, "上海", align=LEFT)
r += 1

# 特长/爱好
label(r, 1, "特长")
merge(r, 2, r, 5, "资源整合", align=LEFT)
label(r, 6, "爱好")
merge(r, 7, r, 11, "马拉松、游泳、滑雪", align=LEFT)
r += 1

# 家庭住址/邮编/紧急联系人
label(r, 1, "家庭住址")
merge(r, 2, r, 5, "上海市杨浦区爱国路389号", align=LEFT)
label(r, 6, "邮政编码")
cell(r, 7, "", align=CENTER)
label(r, 8, "紧急联系人及电话")
merge(r, 9, r, 11, "马喜艳  18653268620", align=LEFT)
r += 1

# 现居住地
mlabel(r, 1, r, 2, "现本人居(租)住地详细地址")
merge(r, 3, r, 11, "上海市杨浦区爱国路389号", align=LEFT)
r += 1

# ---------------------------------------------------------------------------
# 主要工作经历
# ---------------------------------------------------------------------------
header(r, 1, r, MAX_COL, "主 要 工 作 经 历（含参军入伍经历）")
r += 1

# 表头
label(r, 1, "起止时间")
mlabel(r, 2, r, 4, "任职单位（或部队）名称/性质")
mlabel(r, 5, r, 6, "职位（或兵种军衔）")
mlabel(r, 7, r, 8, "证明人/联系方式")
label(r, 9, "薪资")
mlabel(r, 10, r, 11, "离职原因")
r += 1

work_rows = [
    ("2011.6-2015.10", "中南建设集团", "高级经理"),
    ("2015.10-2017.2", "新城控股集团", "资深专业经理"),
    ("2017.3-2021.5", "万科企业集团股份有限公司", "副总经理"),
    ("2021.6-至今", "复旦大学住房政策研究中心", "秘书长"),
]
for start, org, pos in work_rows:
    cell(r, 1, start, align=CENTER)
    merge(r, 2, r, 4, org, align=LEFT)
    merge(r, 5, r, 6, pos, align=CENTER)
    merge(r, 7, r, 8, "")
    cell(r, 9, "")
    merge(r, 10, r, 11, "")
    r += 1

# 业绩自评
header(r, 1, r, MAX_COL, "业绩自评（含受过何种奖惩）")
r += 1
achievement = (
    "房地产龙头企业15年以上工作经验，土木建筑科班出身，14年投资拓展行业经验，专注华东（江浙沪、皖鲁闽）区域市场。\n"
    "主导过靖江印象城（商业综合体）、嘉兴海宁、镇江大港、常州金坛、临沂鲁商新都会、上海龙湖滟澜山（法拍收并购）、"
    "嘉兴平湖及上海南桥冷链物流园等项目的土地获取与收并购落地，覆盖重资产勾地、轻资产品牌输出、委托代建、股权收并购、"
    "法院拍卖及不良资产处置等多种拓展形式。\n"
    "副教授级高级工程师，上海市人才引进认证；兼任上海市工商联房地产商会秘书长、上海市杨浦区科技企业联合会执行会长、"
    "复旦管院不动产资产管理协会创始理事长等社会职务；担任多家机构评审专家。\n"
    "发表研究论文4篇（含知网、《住宅与房地产》《中国新技术新产品》等），拥有实用新型专利2项。"
)
merge(r, 1, r, MAX_COL, achievement, align=LEFT_TOP)
ws.row_dimensions[r].height = 100
r += 1

# ---------------------------------------------------------------------------
# 表2-1 学校教育经历及培训经历
# ---------------------------------------------------------------------------
header(r, 1, r, MAX_COL, "学校教育经历（从高中开始填写）及培训经历")
r += 1

# 教育经历表头
mlabel(r, 1, r, 2, "起止时间")
mlabel(r, 3, r, 5, "教育单位")
mlabel(r, 6, r, 8, "教育专业")
mlabel(r, 9, r, 11, "获证情况")
r += 1

edu_rows = [
    ("2018/09-2021/06", "复旦大学", "工商管理（财务金融方向）", "硕士研究生"),
    ("2007/09-2011/07", "中国海洋大学", "土木工程", "本科（工学学士）"),
    ("", "", "", ""),
]
for start, org, major, cert in edu_rows:
    merge(r, 1, r, 2, start, align=CENTER)
    merge(r, 3, r, 5, org, align=LEFT)
    merge(r, 6, r, 8, major, align=LEFT)
    merge(r, 9, r, 11, cert, align=LEFT)
    r += 1

# 培训经历表头
mlabel(r, 1, r, 2, "起止时间")
mlabel(r, 3, r, 5, "培训单位")
mlabel(r, 6, r, 8, "培训项目")
mlabel(r, 9, r, 11, "培训内容")
r += 1

train_rows = [
    ("", "上海交通大学", "城市治理数字化转型—技术与实践高级研修班", ""),
    ("", "复旦大学·怡安", "人力资源和风险管理实践课程研修班", ""),
    ("", "复旦大学·花旗", "银行实践课程研修班", ""),
    ("", "", "", ""),
]
for start, org, proj, content in train_rows:
    merge(r, 1, r, 2, start, align=CENTER)
    merge(r, 3, r, 5, org, align=LEFT)
    merge(r, 6, r, 8, proj, align=LEFT)
    merge(r, 9, r, 11, content, align=LEFT)
    r += 1

# ---------------------------------------------------------------------------
# 家庭成员及主要社会关系
# ---------------------------------------------------------------------------
header(r, 1, r, MAX_COL, "家庭成员及主要社会关系（子女数：___人  其中男：___人，女：___人）")
r += 1

label(r, 1, "称谓")
mlabel(r, 2, r, 3, "姓名")
label(r, 4, "年龄")
mlabel(r, 5, r, 6, "文化程度")
mlabel(r, 7, r, 8, "工作单位")
label(r, 9, "职位")
mlabel(r, 10, r, 11, "联系电话")
r += 1

family_rows = [
    ("配偶", "马喜艳", "", "", "", "", "18653268620"),
    ("", "", "", "", "", "", ""),
    ("", "", "", "", "", "", ""),
    ("", "", "", "", "", "", ""),
]
for rel, name, age, edu, org, pos, phone in family_rows:
    cell(r, 1, rel, align=CENTER)
    merge(r, 2, r, 3, name, align=CENTER)
    cell(r, 4, age, align=CENTER)
    merge(r, 5, r, 6, edu, align=CENTER)
    merge(r, 7, r, 8, org, align=LEFT)
    cell(r, 9, pos, align=CENTER)
    merge(r, 10, r, 11, phone, align=CENTER)
    r += 1

# ---------------------------------------------------------------------------
# 就职调查
# ---------------------------------------------------------------------------
header(r, 1, r, MAX_COL, "就职调查")
r += 1

survey = [
    ("1、对奇瑞集团及应聘公司的认识？", ""),
    ("2、对应聘岗位的理解？", ""),
    ("3、什么岗位能最大限度地发挥您的才智？",
     "投资拓展、政府关系、资源整合及产业招商类岗位。"),
    ("4、您最不喜欢的工作？", ""),
    ("5、您的性格特点？",
     "行事果断、稳健、理智，进取心强，思路清晰、逻辑性强，抗压与执行力强，亲和力强，善于团队合作。"),
]
for q, a in survey:
    merge(r, 1, r, MAX_COL, q, align=LEFT)
    r += 1
    merge(r, 1, r, MAX_COL, a, align=LEFT_TOP)
    ws.row_dimensions[r].height = 34
    r += 1

# Q6 档案及社保
merge(r, 1, r, MAX_COL, "6、档案及社保关系？", align=LEFT)
r += 1
merge(r, 1, r, MAX_COL, "档案：□原单位    □市社保局    □本人持有    □其它", align=LEFT)
r += 1
merge(r, 1, r, MAX_COL,
      "社保：□原单位    □自己缴纳    种类：□养老  □医疗  □失业  □工伤  □生育", align=LEFT)
r += 1

# Q7
merge(r, 1, r, MAX_COL, "7、公司内有无亲属及亲属关系、姓名、部门？", align=LEFT)
r += 1
merge(r, 1, r, MAX_COL, "", align=LEFT_TOP)
ws.row_dimensions[r].height = 30
r += 1

# Q8
merge(r, 1, r, MAX_COL, "8、招聘信息来源？", align=LEFT)
r += 1
merge(r, 1, r, MAX_COL,
      "□招聘网站    □朋友或同学介绍    □公司员工介绍    □内部招聘    □其它", align=LEFT)
r += 1

# 证件审查/体检
merge(r, 1, r, MAX_COL,
      "证件审查情况记录：□身份证    □护照    □毕业证    □学位证    □其他证件", align=LEFT)
r += 1
merge(r, 1, r, MAX_COL, "体检结果：□合格    □不合格", align=LEFT)
r += 1
merge(r, 1, r, MAX_COL, "审查人：___________________          日期：_____________________",
      align=LEFT)
r += 1

last_row = r - 1

# ---------------------------------------------------------------------------
# 统一加边框、默认行高
# ---------------------------------------------------------------------------
for row in range(7, last_row + 1):
    for col in range(1, MAX_COL + 1):
        c = ws.cell(row=row, column=col)
        c.border = BORDER
        if ws.row_dimensions[row].height is None:
            ws.row_dimensions[row].height = 22

# 打印设置
ws.print_area = f"A1:{get_column_letter(MAX_COL)}{last_row}"
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

out = "附件3-应聘人员登记表-胡继刚.xlsx"
wb.save(out)
print(f"saved: {out}  (rows: {last_row})")
