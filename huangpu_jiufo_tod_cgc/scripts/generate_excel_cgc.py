# -*- coding: utf-8 -*-
"""生成 Excel 附件数据表:出口TOP20 / 消费类出口TOP20 / 楼层功能布局。"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import content_cgc as C

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "deliverables")
FONT = "微软雅黑"
NAVY = "0F2A4A"
LIGHT = "EEF2F8"
GOLD = "C79A3B"

thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def setup_print(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.sheet_view.showGridLines = False


def widths(ws, ws_widths):
    for j, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def title_row(ws, text, ncols, row=1, h=30):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CENTER
    ws.row_dimensions[row].height = h


def header_row(ws, headers, row):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=FONT, size=10.5, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=GOLD)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 24


def data_row(ws, row, values, center_cols=(), zebra=True):
    fill = PatternFill("solid", fgColor=LIGHT) if (zebra and row % 2 == 0) else None
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(name=FONT, size=10)
        c.alignment = CENTER if j in center_cols else LEFT
        c.border = BORDER
        if fill:
            c.fill = fill


def note_row(ws, row, ncols, note):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value="说明：" + note)
    c.font = Font(name=FONT, size=9.5, italic=True, color="666666")
    c.alignment = LEFT
    ws.row_dimensions[row].height = 42


def sheet_export(wb):
    ws = wb.active
    ws.title = "出口TOP20"
    setup_print(ws)
    widths(ws, [6, 22, 22, 16, 46])
    title_row(ws, "2025年广州出口TOP20品类", 5)
    header_row(ws, ["排名", "品类", "代表品牌", "广州口岸交易额", "核心说明"], 2)
    r = 3
    for row in C.EXPORT_TOP20:
        data_row(ws, r, row, center_cols=(1, 4))
        ws.row_dimensions[r].height = 22
        r += 1
    note_row(ws, r, 5, C.EXPORT_TOP20_NOTE)


def sheet_consumer(wb):
    ws = wb.create_sheet("消费类出口TOP20")
    setup_print(ws)
    widths(ws, [6, 20, 14, 10, 26, 20])
    title_row(ws, "2025年前10月广州消费类出口20强（单位：亿元）", 6)
    header_row(ws, ["排名", "品类", "出口额", "同比增速", "核心出口品牌", "核心市场"], 2)
    r = 3
    for row in C.CONSUMER_TOP20:
        data_row(ws, r, row, center_cols=(1, 3, 4))
        ws.row_dimensions[r].height = 22
        r += 1
    note_row(ws, r, 6, C.CONSUMER_TOP20_NOTE)


def sheet_floor(wb):
    ws = wb.create_sheet("楼层功能布局")
    setup_print(ws)
    widths(ws, [16, 26, 42, 40])
    title_row(ws, "全球自贸365街区 · 楼层功能定位与业态布局", 4)
    header_row(ws, ["区位楼层", "功能定位", "核心业态 / 服务", "数据支撑与参考案例"], 2)
    r = 3
    for row in C.FLOOR_LAYOUT:
        data_row(ws, r, row, center_cols=(1,))
        ws.row_dimensions[r].height = 40
        r += 1


def main():
    wb = Workbook()
    sheet_export(wb)
    sheet_consumer(wb)
    sheet_floor(wb)
    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, "广州黄埔九佛TOD全球自贸365街区_附件数据表.xlsx")
    wb.save(out)
    print("saved:", out)


if __name__ == "__main__":
    main()
