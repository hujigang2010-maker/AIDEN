"""生成 Excel：AI+机器人轻资产创业活动 · 倒排计划与运营管理表

Sheets：
  00 总览 Dashboard
  01 会议纪要梗概
  02 七至十二月倒排计划
  03 寒假主活动执行
  04 合作方对接清单
  05 预算明细
  06 品牌调研跟踪
  07 风险预案
  08 KPI 考核
  09 待讨论事项
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# 紫色主题
PRIMARY = "4A148C"
ACCENT = "7B1FA2"
LIGHT = "F3E5F5"
LAVENDER = "CE93D8"
GOLD = "E1BEE7"
GREY = "6A5A7A"
WHITE = "FFFFFF"
DARK = "2D1B3D"
GREEN = "1E8449"
ORANGE = "E67E22"
YELLOW = "F1C40F"
RED = "C0392B"


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_row(ws, row, headers, fill=PRIMARY, color=WHITE, height=26):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="微软雅黑", size=11, bold=True, color=color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[row].height = height


def write_row(ws, row, data, fill=None, bold=False, size=10, wrap=True, align="left"):
    for i, v in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name="微软雅黑", size=size, bold=bold, color=DARK)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border = thin()
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def section_banner(ws, row, text, span=10, fill=ACCENT, color=WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", size=12, bold=True, color=color)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24


def big_title(ws, span_col="L", text="", sub=""):
    ws.merge_cells(f"A1:{span_col}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(name="微软雅黑", size=18, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A2:{span_col}2")
    c2 = ws["A2"]
    c2.value = sub
    c2.font = Font(name="微软雅黑", size=10, italic=True, color=GREY)
    c2.fill = PatternFill("solid", fgColor=LIGHT)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20


def add_status_validation(ws, col_letter, start_row, end_row):
    dv = DataValidation(
        type="list",
        formula1='"未开始,进行中,已完成,延期,取消"',
        allow_blank=True,
    )
    dv.error = "请从下拉列表选择状态"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def build_dashboard(wb):
    ws = wb.active
    ws.title = "00 总览Dashboard"
    big_title(
        ws,
        "M",
        "AI+机器人轻资产创业活动 · 总览 Dashboard",
        "寒假研学营为主节点  |  七至十二月倒排筹备  |  讨论稿 V1.0",
    )

    section_banner(ws, 4, "一、项目基本信息", span=13, fill=PRIMARY)
    info = [
        ["项目名称", "「智创未来」寒假 AI+机器人科创研学营（青岛黄岛试点）"],
        ["核心策略", "轻资产 · 研学切入 · 资源池×产品×客群对齐 · 游击战营销"],
        ["主活动窗口", "2026年1月中下旬—2月（寒假）"],
        ["筹备周期", "2025年7月—12月（倒排六个月）"],
        ["目标客群", "研学学生/教育机构 + 活动公司/商场 B 端"],
        ["暂缓事项", "展厅建设、国企混改、50万级代理买断"],
        ["建议合作", "高校/中航科幻科普中心（内容背书）+ 黄岛区科技企业（设备演示）"],
        ["轻资产预算", "37—60 万元（首年验证期，不含代理买断）"],
    ]
    for i, (k, v) in enumerate(info, 5):
        ws.cell(row=i, column=1, value=k).font = Font(name="微软雅黑", bold=True, color=DARK)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=1).border = thin()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=13)
        c = ws.cell(row=i, column=2, value=v)
        c.font = Font(name="微软雅黑", color=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin()

    section_banner(ws, 14, "二、月度里程碑一览", span=13, fill=ACCENT)
    header_row(ws, 15, ["月份", "阶段", "关键交付", "状态", "负责人", "备注"])
    milestones = [
        ("7月", "战略对齐", "策划案定稿、品牌调研启动、研学资源盘点", "未开始", "", "本表讨论起点"),
        ("8月", "资源对接", "协会入会、高校/企业意向书、团队分工", "未开始", "", ""),
        ("9月", "试点验证", "2-3场体验活动、代理政策比选、试点复盘", "未开始", "", ""),
        ("10月", "产品封装", "课程包、PPT/视频物料、报价体系", "未开始", "", ""),
        ("11月", "招生预热", "机构签约、报名数据、体验课开放日", "未开始", "", "设最低开班线"),
        ("12月", "冲刺彩排", "场地确认、全流程彩排、寒假最终排期", "未开始", "", ""),
        ("1-2月", "寒假执行", "3-5期研学营、收入结算、复盘", "未开始", "", "主活动窗口"),
    ]
    for i, row in enumerate(milestones, 16):
        write_row(ws, i, list(row))
    add_status_validation(ws, "D", 16, 22)

    section_banner(ws, 24, "三、核心 KPI 目标", span=13, fill=PRIMARY)
    header_row(ws, 25, ["维度", "指标", "目标值", "实际值", "完成率", "状态"])
    kpis = [
        ("招生", "寒假营总人次", 150),
        ("收入", "活动总收入（元）", 150000),
        ("渠道", "签约研学机构（家）", 3),
        ("合作", "联合高校/企业（家）", 2),
        ("转化", "算力/课程续费用户", 30),
        ("成本", "单场活动毛利率（%）", 25),
    ]
    for i, (dim, kpi, target) in enumerate(kpis, 26):
        ws.cell(row=i, column=1, value=dim)
        ws.cell(row=i, column=2, value=kpi)
        ws.cell(row=i, column=3, value=target)
        ws.cell(row=i, column=4, value=0)
        ws.cell(row=i, column=5, value=f"=IFERROR(D{i}/C{i},0)")
        ws.cell(row=i, column=5).number_format = "0.0%"
        ws.cell(row=i, column=6, value=f'=IF(D{i}=0,"未开始",IF(E{i}>=1,"达成",IF(E{i}>=0.7,"进行中","风险")))')
        for col in range(1, 7):
            c = ws.cell(row=i, column=col)
            c.font = Font(name="微软雅黑", size=10, color=DARK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin()
            if col == 1:
                c.fill = PatternFill("solid", fgColor=LIGHT)

    rng = "F26:F31"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="equal",
            formula=['"达成"'],
            fill=PatternFill("solid", fgColor="D4EDDA"),
            font=Font(color="155724", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="equal",
            formula=['"风险"'],
            fill=PatternFill("solid", fgColor="F8D7DA"),
            font=Font(color="721C24", bold=True),
        ),
    )

    set_widths(ws, [14, 22, 14, 12, 10, 12, 20])
    ws.freeze_panes = "A4"


def build_summary(wb):
    ws = wb.create_sheet("01 会议纪要梗概")
    big_title(ws, "F", "会议纪要梗概", "基于《AI与机器人创业方向探讨》两份材料整理")

    sections = [
        ("行业转型", [
            "地产行业将长期横盘/下滑，不宜再作主业，仅作辅助变现",
            "中国经济进入资本科技驱动阶段，须转向轻资产科技赛道",
            "团队存在地产思维惯性：依赖政府关系、重资产展厅、走捷径",
        ]),
        ("资源与客群", [
            "三要素对齐：资源池、产品、客群（如同当年卖房逻辑）",
            "最大可控资源：研学渠道（学校/教育机构），非政府/国企",
            "政府资源不可控，代理商易被追本溯源跨级对接厂家",
            "聚焦 B 端：婚庆、活动公司、短视频制作方、研学机构",
        ]),
        ("轻资产运营", [
            "反对建展厅：C 端无法收费，B 端更愿看总部或 PPT/视频",
            "游击战 > 阵地战：走出去办活动，五四广场/商场/企业均可",
            "可加入青岛人工智能/机器人协会（3-5万副会长）横向比选",
        ]),
        ("产品研判", [
            "机器人：表演/引流为主，宇树九万九、智谱等头部优选",
            "库存风险极高（机器狗一年四代），建议租赁+提成勿压货",
            "AI算力：小微团队是主力客群，即梦等可拿4-5折代理",
            "成功案例：青岛萝卜快跑，小工作室+免费培训+流量包续费",
        ]),
        ("战略建议", [
            "暂缓展厅、混改、50万深蓝/20-30万智巨人等非头部买断",
            "机器人与AI客群不同，双线须统一客群或聚焦单一方向",
            "先找需求再找产品，快速验证MVP，不必急于求成",
            "建议参观人工智能大会等行业活动后再最终决策",
        ]),
    ]

    row = 4
    for title, items in sections:
        section_banner(ws, row, title, span=6, fill=ACCENT)
        row += 1
        for it in items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = ws.cell(row=row, column=1, value=f"• {it}")
            c.font = Font(name="微软雅黑", size=10, color=DARK)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = thin()
            ws.row_dimensions[row].height = 28
            row += 1
        row += 1

    set_widths(ws, [18, 18, 18, 18, 18, 18])
    ws.freeze_panes = "A4"


def build_timeline(wb):
    ws = wb.create_sheet("02 七至十二月倒排计划")
    big_title(ws, "J", "七至十二月倒排计划", "以2026寒假为主节点逆向拆解  |  每周任务可跟踪")

    headers = [
        "序号", "月份", "周次", "任务类别", "具体任务", "交付物",
        "负责人", "协作方", "计划完成日", "状态", "备注",
    ]
    header_row(ws, 4, headers)

    tasks = [
        # 7月
        (1, "7月", "W1", "战略", "召开策划案讨论会，确认MVP方向", "会议纪要", "", "全体", "7/7", "未开始", "本表起点"),
        (2, "7月", "W1", "调研", "盘点研学资源：机构数、年接待量、单价", "资源盘点表", "", "研学渠道", "7/10", "未开始", ""),
        (3, "7月", "W2", "调研", "调研宇树、智谱、即梦、DeepSeek代理政策", "品牌对比表", "", "项目负责人", "7/17", "未开始", ""),
        (4, "7月", "W2", "决策", "确认暂缓展厅/混改，轻资产路径", "决策记录", "", "全体", "7/18", "未开始", "关键决策"),
        (5, "7月", "W3", "筹备", "确定寒假活动名称、定位、定价区间", "产品Brief", "", "项目负责人", "7/24", "未开始", ""),
        (6, "7月", "W4", "团队", "明确5-8人分工，建立周报机制", "分工表", "", "项目负责人", "7/31", "未开始", ""),
        # 8月
        (7, "8月", "W1", "合作", "联系青岛人工智能/机器人协会", "协会资料", "", "项目负责人", "8/7", "未开始", "3-5万副会长"),
        (8, "8月", "W1", "合作", "拜访高校/中航科幻类科普中心", "合作意向书", "", "内容与课程", "8/10", "未开始", "内容背书"),
        (9, "8月", "W2", "合作", "拜访黄岛区科技企业2-3家", "企业拜访记录", "", "项目负责人", "8/17", "未开始", "设备演示"),
        (10, "8月", "W2", "合作", "评估扬州等地设备租赁补充方案", "租赁方案", "", "活动执行", "8/17", "未开始", "备选"),
        (11, "8月", "W3", "品牌", "横向比选机器人演示供应商", "供应商短名单", "", "项目负责人", "8/24", "未开始", ""),
        (12, "8月", "W4", "品牌", "洽谈一级代理/提成模式（非买断）", "代理条款草案", "", "项目负责人", "8/31", "未开始", ""),
        # 9月
        (13, "9月", "W1", "试点", "策划首场周末体验课（30人以内）", "活动方案", "", "活动执行", "9/7", "未开始", "MVP验证"),
        (14, "9月", "W2", "试点", "执行首场体验活动+数据采集", "试点复盘报告", "", "全体", "9/14", "未开始", ""),
        (15, "9月", "W3", "试点", "执行第2-3场试点（不同场地）", "试点数据汇总", "", "活动执行", "9/21", "未开始", "游击战验证"),
        (16, "9月", "W4", "决策", "根据试点确定主代理品牌与合作方", "合作确认书", "", "项目负责人", "9/30", "未开始", "关键决策"),
        # 10月
        (17, "10月", "W1", "产品", "寒假课程大纲定稿（机器人+AI模块）", "课程大纲", "", "内容与课程", "10/10", "未开始", ""),
        (18, "10月", "W2", "物料", "制作宣传PPT、演示视频、招生海报", "物料包", "", "市场运营", "10/17", "未开始", ""),
        (19, "10月", "W3", "商务", "确定报价体系、合同模板、退费规则", "商务文件包", "", "项目负责人", "10/24", "未开始", ""),
        (20, "10月", "W4", "合作", "与高校/企业签署联合框架（如有）", "合作协议", "", "项目负责人", "10/31", "未开始", ""),
        # 11月
        (21, "11月", "W1", "招生", "研学机构批量推介+签约", "机构合同", "", "研学渠道", "11/7", "未开始", ""),
        (22, "11月", "W2", "招生", "家长社群/短视频招生预热", "报名链接", "", "市场运营", "11/14", "未开始", ""),
        (23, "11月", "W3", "活动", "体验课开放日（收费筛选）", "转化数据", "", "活动执行", "11/21", "未开始", ""),
        (24, "11月", "W4", "决策", "评估报名达最低开班线（建议≥30人/期）", "招生报告", "", "全体", "11/30", "未开始", "未达标则延期"),
        # 12月
        (25, "12月", "W1", "运营", "确认寒假场地档期（多点位）", "场地合同", "", "活动执行", "12/7", "未开始", ""),
        (26, "12月", "W2", "运营", "师资/志愿者培训、安全预案", "执行手册", "", "内容与课程", "12/14", "未开始", ""),
        (27, "12月", "W3", "运营", "全流程彩排（含设备/签到/课程）", "彩排记录", "", "全体", "12/21", "未开始", ""),
        (28, "12月", "W4", "冲刺", "寒假最终排期发布、物料就位", "排期表", "", "项目负责人", "12/28", "未开始", "进入主活动"),
    ]

    for i, t in enumerate(tasks, 5):
        write_row(ws, i, t, fill=LIGHT if i % 2 == 0 else None)
    add_status_validation(ws, "J", 5, 5 + len(tasks) - 1)

    # 按月份条件格式
    month_colors = {"7月": "EDE7F6", "8月": "E1BEE7", "9月": "D1C4E9", "10月": "CE93D8", "11月": "BA68C8", "12月": "AB47BC"}
    for i, t in enumerate(tasks, 5):
        month = t[1]
        if month in month_colors:
            ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=month_colors[month])

    set_widths(ws, [6, 8, 8, 10, 36, 18, 10, 12, 12, 10, 16])
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{4 + len(tasks)}"


def build_winter_camp(wb):
    ws = wb.create_sheet("03 寒假主活动执行")
    big_title(ws, "I", "寒假主活动执行计划", "2026年1月中下旬—2月  |  3-5期研学营")

    section_banner(ws, 4, "活动排期（模板，具体日期待定）", span=9)
    header_row(ws, 5, ["期次", "计划日期", "场地", "主题", "目标人数", "已报名", "状态", "收入（元）", "备注"])

    camps = [
        ("第1期", "2026/1/18（周六）", "黄岛商场中庭（待定）", "机器人奇遇记", 40, 0, "未开始", 0, "首场"),
        ("第2期", "2026/1/25（周六）", "社区活动中心（待定）", "AI小导演", 40, 0, "未开始", 0, ""),
        ("第3期", "2026/2/1（周六）", "高校科普场地（待定）", "智创未来一日营", 50, 0, "未开始", 0, "联合高校"),
        ("第4期", "2026/2/8（周六）", "黄岛商场/户外（待定）", "机器人挑战赛", 40, 0, "未开始", 0, "视报名增设"),
        ("第5期", "2026/2/15（周六）", "待定", "寒假收官营", 40, 0, "未开始", 0, "视报名增设"),
    ]
    for i, c in enumerate(camps, 6):
        write_row(ws, i, c)
    add_status_validation(ws, "G", 6, 10)

    section_banner(ws, 12, "单日执行流程", span=9, fill=PRIMARY)
    flow = [
        ("09:00-09:30", "签到破冰", "活动执行", "签到表、名牌、保险确认"),
        ("09:30-10:30", "机器人表演与互动", "合作企业技术", "样机、安全围栏"),
        ("10:45-11:45", "AI短视频小课堂", "内容与课程", "即梦/类似工具账号"),
        ("12:00-13:30", "午餐休息", "活动执行", "盒饭或家长自理"),
        ("13:30-15:00", "分组科创挑战", "内容与课程", "任务卡、奖品"),
        ("15:00-15:30", "成果展示+结业", "全体", "证书、合影、问卷"),
    ]
    header_row(ws, 13, ["时段", "环节", "负责角色", "物料/备注"], fill=ACCENT)
    for i, f in enumerate(flow, 14):
        write_row(ws, i, f, fill=LIGHT if i % 2 == 0 else None)

    set_widths(ws, [10, 14, 22, 16, 10, 10, 10, 12, 18])
    ws.freeze_panes = "A5"


def build_partners(wb):
    ws = wb.create_sheet("04 合作方对接清单")
    big_title(ws, "H", "合作方对接清单", "高校中心 + 黄岛区科技企业  |  建议引入")

    section_banner(ws, 4, "A类：高校/中航科幻科普中心（建议引入）", span=8, fill=ACCENT)
    header_row(ws, 5, ["机构名称", "对接人", "角色定位", "合作内容", "状态", "计划完成", "负责人", "备注"])
    partners_a = [
        ("高校科创/科普中心（待定）", "", "课程背书+师资", "联合研学实践基地、科普讲座", "未接触", "9月", "", "提升公信力"),
        ("中航科幻类科普中心（待定）", "", "场景内容", "航空航天+机器人跨界体验模块", "未接触", "9月", "", "差异化内容"),
    ]
    for i, p in enumerate(partners_a, 6):
        write_row(ws, i, p)

    section_banner(ws, 9, "B类：黄岛区科技企业（建议引入）", span=8, fill=PRIMARY)
    header_row(ws, 10, ["企业名称", "对接人", "可提供", "合作模式", "状态", "计划完成", "负责人", "备注"])
    partners_b = [
        ("黄岛区人工智能/机器人企业1", "", "机器人样机+技术讲解", "单场演示分成", "未接触", "8月", "", "优先本地"),
        ("黄岛区人工智能/机器人企业2", "", "设备租赁", "按场结算", "未接触", "8月", "", ""),
        ("青岛人工智能协会", "", "行业资源+品牌", "副会长/会员", "未接触", "8月", "", "3-5万/年"),
        ("扬州区科技企业（备选）", "", "设备租赁补充", "异地租赁", "未接触", "9月", "", "仅作备选"),
    ]
    for i, p in enumerate(partners_b, 11):
        write_row(ws, i, p)

    section_banner(ws, 16, "C类：不建议作为核心依赖", span=8, fill=RED)
    header_row(ws, 17, ["机构", "原设想", "风险", "建议处理方式", "状态", "", "", ""])
    partners_c = [
        ("国企混改", "展厅+场地+资金", "易被架空、决策慢", "仅作场地赞助，不参与股权", "暂缓", "", "", ""),
        ("政府直接合作", "政策/补贴/获客", "资源不可控", "不依赖，可了解孵化补贴信息", "暂缓", "", "", ""),
    ]
    for i, p in enumerate(partners_c, 18):
        write_row(ws, i, p)

    add_status_validation(ws, "E", 6, 14)
    set_widths(ws, [24, 10, 16, 24, 10, 10, 10, 20])
    ws.freeze_panes = "A5"


def build_budget(wb):
    ws = wb.create_sheet("05 预算明细")
    big_title(ws, "G", "预算明细（轻资产版）", "首年验证期  |  不含50万级代理买断")

    header_row(ws, 4, ["类别", "项目", "预算（万元）", "已支出", "余额", "支付时间", "备注"])
    items = [
        ("一次性", "协会/行业资源（副会长等）", 4, 0, "=C5-D5", "8月", ""),
        ("变动", "机器人设备租赁（全年）", 12, 0, "=C6-D6", "按场", "非购置"),
        ("变动", "场地与活动执行", 8, 0, "=C7-D7", "按月", "多点位"),
        ("一次性", "课程与物料开发", 4, 0, "=C8-D8", "10月", "PPT/视频/教材"),
        ("变动", "市场招生与宣发", 4, 0, "=C9-D9", "11月", ""),
        ("固定", "人员与日常运营", 12, 0, "=C10-D10", "按月", "5-8人"),
        ("储备", "预备金", 5, 0, "=C11-D11", "随时", ""),
        ("", "合计", "=SUM(C5:C11)", "=SUM(D5:D11)", "=C12-D12", "", "37-60万区间"),
    ]
    for i, it in enumerate(items, 5):
        write_row(ws, i, it, bold=(it[0] == ""))
        if isinstance(it[2], str) and it[2].startswith("="):
            ws.cell(row=i, column=3).value = it[2]
        if isinstance(it[3], str) and it[3].startswith("="):
            ws.cell(row=i, column=4).value = it[3]
        if isinstance(it[4], str) and it[4].startswith("="):
            ws.cell(row=i, column=5).value = it[4]

    # 对比
    section_banner(ws, 14, "与重资产方案对比", span=7, fill=ACCENT)
    header_row(ws, 15, ["方案", "初期投入", "风险", "建议"], fill=PRIMARY)
    compare = [
        ("轻资产研学营（本方案）", "37-60万", "低", "✅ 推荐"),
        ("深蓝机器人区域代理", "50万+", "高（库存贬值）", "❌ 暂缓"),
        ("智巨人AI代理", "20-30万", "中（非头部）", "❌ 暂缓"),
        ("国企混改+展厅", "100万+", "极高", "❌ 暂缓"),
    ]
    for i, c in enumerate(compare, 16):
        write_row(ws, i, c)

    set_widths(ws, [10, 28, 12, 12, 12, 12, 24])
    ws.freeze_panes = "A5"


def build_brand_research(wb):
    ws = wb.create_sheet("06 品牌调研跟踪")
    big_title(ws, "H", "品牌调研跟踪表", "头部优先  |  一级代理+提成  |  勿压库存")

    header_row(ws, 4, ["品类", "品牌", "代理费/模式", "优势", "劣势", "调研状态", "优先级", "备注"])
    brands = [
        ("人形机器人", "宇树", "租赁+代理提成", "头部、价格适中（约9.9万起）", "表演场景为主", "待调研", "P0", "会议纪要推荐"),
        ("人形机器人", "智谱", "待了解", "头部品牌", "待了解", "待调研", "P0", ""),
        ("人形机器人", "深蓝科技", "50万黄岛区加盟", "有实体工厂", "价格高、买断风险", "已接触", "P2", "暂缓买断"),
        ("AI视频/算力", "字节即梦", "一级代理4-5折", "头部、短视频场景匹配", "需谈代理", "待调研", "P0", "会议纪要推荐"),
        ("AI算力", "DeepSeek", "待了解", "国内头部", "待了解", "待调研", "P1", ""),
        ("AI视频", "智巨人", "20-30万加盟", "已接触", "非头部", "已接触", "P3", "不建议"),
        ("行业资源", "青岛AI/机器人协会", "3-5万副会长", "横向比选资源", "需时间积累", "待接触", "P1", ""),
    ]
    for i, b in enumerate(brands, 5):
        write_row(ws, i, b, fill=LIGHT if i % 2 == 0 else None)

    add_status_validation(ws, "F", 5, 11)
    set_widths(ws, [12, 14, 18, 22, 18, 10, 8, 18])
    ws.freeze_panes = "A5"


def build_risk(wb):
    ws = wb.create_sheet("07 风险预案")
    big_title(ws, "F", "风险预案", "会议纪要警示项 + 应对策略")

    header_row(ws, 4, ["风险项", "等级", "触发信号", "应对策略", "责任人", "状态"])
    risks = [
        ("双线作战精力分散", "高", "机器人与AI团队互相抢资源", "统一研学客群，分渠道运营", "", "监控中"),
        ("机器人库存贬值", "高", "设备购入后滞销", "只租不买，成交后结算", "", "监控中"),
        ("政府资源被架空", "高", "合作方绕过对接厂家", "不依赖政府获客", "", "已规避"),
        ("招生不达预期", "中", "11月报名<30人/期", "延期或合并期次", "", "监控中"),
        ("品牌代理政策变化", "中", "厂家提高门槛", "多品牌比选", "", "监控中"),
        ("免费课转化低", "中", "体验课到课率<50%", "改收费筛选", "", "监控中"),
        ("安全事故", "高", "活动现场意外", "保险+预案+围栏", "", "监控中"),
    ]
    for i, r in enumerate(risks, 5):
        write_row(ws, i, r)
        level = r[1]
        color = {"高": "F8D7DA", "中": "FFF3CD", "低": "D4EDDA"}.get(level, LIGHT)
        ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=color)

    set_widths(ws, [22, 8, 24, 28, 10, 10])
    ws.freeze_panes = "A5"


def build_kpi(wb):
    ws = wb.create_sheet("08 KPI考核")
    big_title(ws, "G", "KPI 考核表", "寒假验证期  |  每周更新")

    header_row(ws, 4, ["考核周期", "指标", "目标", "实际", "完成率", "考核人", "备注"])
    kpis = [
        ("7月", "策划案定稿", "1份", 0, "=IFERROR(D5/C5,0)", "", ""),
        ("7月", "品牌调研完成", "≥4家", 0, "=IFERROR(D6/C6,0)", "", ""),
        ("8月", "合作意向书", "≥3份", 0, "=IFERROR(D7/C7,0)", "", ""),
        ("9月", "试点活动场次", "≥3场", 0, "=IFERROR(D8/C8,0)", "", ""),
        ("10月", "课程物料完成度", "100%", 0, "=IFERROR(D9/C9,0)", "", ""),
        ("11月", "寒假预报名人数", "≥120人", 0, "=IFERROR(D10/C10,0)", "", ""),
        ("12月", "彩排完成", "1次全流程", 0, "=IFERROR(D11/C11,0)", "", ""),
        ("1-2月", "寒假营总收入", "≥15万元", 0, "=IFERROR(D12/C12,0)", "", ""),
        ("1-2月", "研学机构签约", "≥3家", 0, "=IFERROR(D13/C13,0)", "", ""),
    ]
    for i, k in enumerate(kpis, 5):
        write_row(ws, i, k)
        ws.cell(row=i, column=5).number_format = "0.0%"

    set_widths(ws, [10, 22, 12, 12, 10, 10, 20])
    ws.freeze_panes = "A5"


def build_discussion(wb):
    ws = wb.create_sheet("09 待讨论事项")
    big_title(ws, "E", "待讨论事项", "供团队评审  |  讨论后请在结论列填写")

    header_row(ws, 4, ["序号", "讨论议题", "建议方案", "结论（待填）", "决策人"])
    topics = [
        (1, "是否确认以寒假研学营为第一个MVP？", "是，暂缓展厅与混改", "", ""),
        (2, "机器人与AI算力如何组合？", "统一研学客群，分团队运营", "", ""),
        (3, "是否引入高校/中航科幻中心？", "建议引入，9月前签框架", "", ""),
        (4, "是否引入黄岛区科技企业？", "建议引入2-3家演示伙伴", "", ""),
        (5, "扬州区科技企业是否参与？", "仅作设备租赁备选", "", ""),
        (6, "代理品牌优先调研哪家？", "宇树+即梦+智谱", "", ""),
        (7, "轻资产预算37-60万是否可接受？", "可分阶段投入", "", ""),
        (8, "寒假主档期选1月还是2月？", "建议两月各排2-3期", "", ""),
        (9, "是否赴行业展会补看品牌？", "建议关注下一届人工智能大会", "", ""),
    ]
    for i, t in enumerate(topics, 5):
        write_row(ws, i, t, fill=LIGHT if i % 2 == 0 else None)

    set_widths(ws, [6, 36, 28, 28, 12])
    ws.freeze_panes = "A5"


def main():
    wb = Workbook()
    build_dashboard(wb)
    build_summary(wb)
    build_timeline(wb)
    build_winter_camp(wb)
    build_partners(wb)
    build_budget(wb)
    build_brand_research(wb)
    build_risk(wb)
    build_kpi(wb)
    build_discussion(wb)

    import os
    out = "/workspace/output/AI机器人创业活动倒排计划.xlsx"
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"已生成: {out}")


if __name__ == "__main__":
    main()
