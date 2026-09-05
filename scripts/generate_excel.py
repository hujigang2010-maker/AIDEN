# -*- coding: utf-8 -*-
"""生成前期费用、交付与分工表。"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import content as C

OUT = Path(__file__).resolve().parent.parent / "output" / "港大经管上海中心_前期费用与交付清单.xlsx"

GREEN = "003D2E"
GOLD = "C4A35A"
CREAM = "F7F4EC"
WHITE = "FFFFFF"
LIGHT = "E8EFEA"
DARK = "1A2420"

thin = Border(
    left=Side(style="thin", color="C5D0CA"),
    right=Side(style="thin", color="C5D0CA"),
    top=Side(style="thin", color="C5D0CA"),
    bottom=Side(style="thin", color="C5D0CA"),
)
head_font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
title_font = Font(name="微软雅黑", size=16, bold=True, color=GREEN)
sub_font = Font(name="微软雅黑", size=11, color="5B6B64")
cell_font = Font(name="微软雅黑", size=10.5, color=DARK)
money_font = Font(name="微软雅黑", size=14, bold=True, color=GREEN)
fill_head = PatternFill("solid", fgColor=GREEN)
fill_gold = PatternFill("solid", fgColor=GOLD)
fill_cream = PatternFill("solid", fgColor=CREAM)
fill_light = PatternFill("solid", fgColor=LIGHT)
fill_white = PatternFill("solid", fgColor=WHITE)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = head_font
        cell.fill = fill_head
        cell.alignment = center
        cell.border = thin


def style_row(ws, row, cols, fill=fill_white):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = cell_font
        cell.fill = fill
        cell.alignment = left if col > 1 else center
        cell.border = thin


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_title(ws, title, subtitle, widths, last_col="D"):
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = left
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = sub_font
    ws["A2"].alignment = left
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:4"
    ws.oddHeader.left.text = C.PROJECT_NAME
    ws.oddFooter.right.text = "第 &P 页 / 共 &N 页"


def build(path: Path | None = None) -> Path:
    path = path or OUT
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # 1 封面
    ws = wb.active
    ws.title = "00-封面"
    freeze_title(
        ws,
        "外滩·出海课堂  前期费用与交付清单",
        f"{C.THEIR_UNIT}  ×  {C.OUR_PARTIES}    {C.VERSION}  {C.DATE_CN}    依据{C.MEETING_DATE}交流",
        [18, 42, 42, 28],
        "D",
    )
    ws["A4"] = "项目"
    ws["B4"] = "内容"
    style_header(ws, 4, 2)
    rows = [
        ("产品名称", C.PROJECT_NAME),
        ("合作方", f"{C.THEIR_LEGAL} / {C.THEIR_UNIT}"),
        ("接口人", f"{C.THEIR_CONTACT}  {C.THEIR_TITLE}"),
        ("电话 / 邮箱", f"{C.THEIR_TEL}  {C.THEIR_EMAIL}"),
        ("场地", C.THEIR_ADDR),
        ("联合策划方", C.OUR_PARTIES),
        ("结算主体", C.OUR_SETTLEMENT),
        ("费用名称", C.FEE_NAME),
        ("费用金额", f"{C.FEE_AMOUNT_CN}（¥{C.FEE_AMOUNT:,}）"),
        ("支付时点", f"协议生效后 {C.FEE_DAYS} 个工作日内一次性支付"),
        ("服务周期", f"{C.PLAN_DAYS} 天 + 首场闭门课"),
        ("商业原则", "只收前期费用；主赛道出海；不碰学费分成；不承诺录取人数"),
    ]
    for i, (a, b) in enumerate(rows, 5):
        ws.cell(i, 1, a)
        ws.cell(i, 2, b)
        style_row(ws, i, 2, fill_cream if i == 13 else (fill_light if i % 2 == 0 else fill_white))
        ws.row_dimensions[i].height = 22
    for r in range(5, 17):
        ws.merge_cells(f"B{r}:D{r}")
    ws["A18"] = C.ONE_LINER
    ws.merge_cells("A18:D20")
    ws["A18"].font = Font(name="微软雅黑", size=12, italic=True, color=GREEN)
    ws["A18"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 2 费用构成
    ws = wb.create_sheet("01-前期费用构成")
    freeze_title(ws, "前期联合策划服务费构成（包干）", "对外只报一个数：88,000 元。含外滩出海首场 + 一次主任体验组织。", [12, 28, 18, 55])
    for col, h in enumerate(["序号", "模块", "对外包干（元）", "覆盖说明"], 1):
        ws.cell(4, col, h)
    style_header(ws, 4, 4)
    items = C.FEE_BREAKDOWN
    last_item = 4 + len(items)
    for i, row in enumerate(items, 5):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
        style_row(ws, i, 4, fill_light if i % 2 == 0 else fill_white)
        ws.cell(i, 3).number_format = '#,##0'
        ws.cell(i, 3).alignment = center
        ws.row_dimensions[i].height = 28
    total_row = last_item + 1
    ws.cell(total_row, 1, "")
    ws.cell(total_row, 2, "合计（前期费用）")
    ws.cell(total_row, 3, f"=SUM(C5:C{last_item})")
    ws.cell(total_row, 4, "一次性支付，包干，不含协议外增项")
    for col in range(1, 5):
        ws.cell(total_row, col).font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
        ws.cell(total_row, col).fill = fill_head
        ws.cell(total_row, col).alignment = center if col != 4 else left
        ws.cell(total_row, col).border = thin
    ws.cell(total_row, 3).number_format = '#,##0'
    skip = total_row + 2
    ws.cell(skip, 1, "不包含")
    ws.cell(skip, 1).font = head_font
    ws.cell(skip, 1).fill = fill_gold
    ws.merge_cells(f"A{skip}:D{skip}")
    start_nc = skip + 1
    for i, t in enumerate(C.FEE_NOT_COVERED):
        r = start_nc + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2, t)
        ws.merge_cells(f"B{r}:D{r}")
        style_row(ws, r, 4, fill_cream)
    sub_h = start_nc + len(C.FEE_NOT_COVERED) + 1
    ws.cell(sub_h, 1, "后续场次（不自动生效）")
    ws.merge_cells(f"A{sub_h}:D{sub_h}")
    ws.cell(sub_h, 1).font = head_font
    ws.cell(sub_h, 1).fill = fill_head
    sub_r = sub_h + 1
    ws.cell(sub_r, 1, "—")
    ws.cell(sub_r, 2, "第二场及以后")
    ws.cell(sub_r, 3, 68000)
    ws.cell(sub_r, 4, "另签确认单，含执行不含新一轮 90 天策划")
    style_row(ws, sub_r, 4, fill_light)
    ws.cell(sub_r, 3).number_format = '#,##0'

    # 3 交付
    ws = wb.create_sheet("02-90天交付清单")
    freeze_title(ws, "90 天交付清单（验收用）", "费用到账后启动。名单为工作名单，不保证每一人均出席。", [14, 18, 48, 22])
    for col, h in enumerate(["阶段", "节点", "交付物", "验收方式"], 1):
        ws.cell(4, col, h)
    style_header(ws, 4, 4)
    deliver = [
        ("D0–D7", "签约与资料", "签署齐套、费用到账、出海课资料、主任窗口与首场档期", "银行回单 + 邮件"),
        ("D8–D21", "画像与名单", "出海客群画像、邀约话术、定向名单 ≥80 人初稿", "乙方 5 个工作日书面意见，逾期视为通过"),
        ("D22–D45", "外滩首场", "出海闭门课现场、签到表、议程", "签到表 + 现场完成"),
        ("D22–D45", "主任体验", "高规格出海场书面邀约（出席以其公务为准）", "书面邀约 + 活动举办"),
        ("D22–D45", "会后 7 日", "纪要、A/B/C 分级名单、面试窗口协助记录", "邮件提交"),
        ("D46–D90", "备忘", "复盘；出海课联合推广备忘；原创谷/海外模块清单（不验收）", "邮件提交"),
    ]
    for i, row in enumerate(deliver, 5):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
        style_row(ws, i, 4, fill_light if i % 2 == 0 else fill_white)
        ws.row_dimensions[i].height = 32
    kpi_h = 5 + len(deliver) + 1
    ws.cell(kpi_h, 1, "工作目标（非违约条款）")
    ws.merge_cells(f"A{kpi_h}:D{kpi_h}")
    ws.cell(kpi_h, 1).font = head_font
    ws.cell(kpi_h, 1).fill = fill_gold
    for i, (k, v) in enumerate(C.KPI):
        r = kpi_h + 1 + i
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        ws.merge_cells(f"B{r}:D{r}")
        style_row(ws, r, 4, fill_cream)

    # 4 首场
    ws = wb.create_sheet("03-首场执行")
    freeze_title(ws, "首场闭门课执行表", C.FIRST_EVENT["theme"], [12, 28, 55, 18])
    meta = [
        ("主题", C.FIRST_EVENT["theme"]),
        ("时间", C.FIRST_EVENT["when"]),
        ("地点", C.FIRST_EVENT["where"]),
        ("人群", C.FIRST_EVENT["who"]),
        ("规模", C.FIRST_EVENT_SIZE),
    ]
    for col, h in enumerate(["项", "内容", "备注", "状态"], 1):
        ws.cell(4, col, h)
    style_header(ws, 4, 4)
    r = 5
    for k, v in meta:
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        ws.cell(r, 3, "签约后书面确认")
        ws.cell(r, 4, "待确认")
        style_row(ws, r, 4, fill_light if r % 2 == 0 else fill_white)
        ws.merge_cells(f"B{r}:B{r}")
        r += 1
    ws.cell(r, 1, "议程")
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(r, 1).font = head_font
    ws.cell(r, 1).fill = fill_head
    r += 1
    for line in C.FIRST_EVENT["agenda"]:
        t, *rest = line.split("  ", 1)
        ws.cell(r, 1, t)
        ws.cell(r, 2, rest[0] if rest else line)
        ws.cell(r, 3, "")
        ws.cell(r, 4, "草案")
        style_row(ws, r, 4, fill_cream)
        ws.row_dimensions[r].height = 22
        r += 1

    # 5 分工
    ws = wb.create_sheet("04-双方分工")
    freeze_title(ws, "双方分工", "招生官收口，我方组织到场。", [22, 55, 22, 18])
    for col, h in enumerate(["责任方", "事项", "产出", "时点"], 1):
        ws.cell(4, col, h)
    style_header(ws, 4, 4)
    duties = []
    for item in C.ROLES["港大经管上海中心"]:
        duties.append(("港大经管上海中心", item, "场地/口径/面试", "全程"))
    for item in C.ROLES["联合策划方"]:
        duties.append(("联合策划方", item, "策划/名单/现场/纪要", "90天内"))
    for i, row in enumerate(duties, 5):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
        style_row(ws, i, 4, fill_light if i % 2 == 0 else fill_white)
        ws.row_dimensions[i].height = 28

    # 6 付款
    ws = wb.create_sheet("05-付款与开票")
    freeze_title(ws, "付款与开票（签署时填写）", "账户信息与发票抬头不一致的，以书面确认为准。", [22, 50, 22, 22])
    for col, h in enumerate(["字段", "填写", "责任", "备注"], 1):
        ws.cell(4, col, h)
    style_header(ws, 4, 4)
    pay = [
        ("费用名称", C.FEE_NAME, "双方", "正文锁定"),
        ("金额（小写）", C.FEE_AMOUNT, "双方", "包干"),
        ("金额（大写）", C.FEE_AMOUNT_CN, "双方", "以协议为准"),
        ("支付时点", f"生效后 {C.FEE_DAYS} 个工作日", "乙方", "一次性"),
        ("开户名称", "", "甲方结算主体", "签署时填"),
        ("开户银行", "", "甲方结算主体", "签署时填"),
        ("账号", "", "甲方结算主体", "签署时填"),
        ("纳税人识别号", "", "甲方结算主体", "签署时填"),
        ("发票类型", "", "甲方", "增值税发票"),
        ("发票抬头", C.THEIR_LEGAL, "乙方", "可改"),
        ("乙方有权签署机构", "", "乙方", "学院/中心/指定法人"),
        ("接口人", f"{C.THEIR_CONTACT} / 甲方项目经理", "双方", "生效后 3 个工作日书面确认"),
    ]
    for i, row in enumerate(pay, 5):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
        style_row(ws, i, 4, fill_cream if row[1] == "" else (fill_light if i % 2 == 0 else fill_white))
        ws.row_dimensions[i].height = 24
        if i == 6:
            ws.cell(i, 2).number_format = '#,##0'
            ws.cell(i, 2).font = money_font

    ws = wb.create_sheet("06-对方需求摘录")
    freeze_title(
        ws,
        f"对方需求摘录（{C.MEETING_DATE} 外滩中心交流）",
        "整理自洽谈记录。对外方案采用出海、经管学院层级、主任体验；不采用闲聊中的非正式跨境口径。",
        [18, 22, 55, 22],
    )
    for col, h in enumerate(["类型", "要点", "说明", "本期如何用"], 1):
        ws.cell(4, col, h)
    style_header(ws, 4, 4)
    need_rows = [
        ("听懂", C.HEARD[0][0], C.HEARD[0][1], "协议第一条分层"),
        ("听懂", C.HEARD[1][0], C.HEARD[1][1], "首场与方案主赛道"),
        ("听懂", C.HEARD[2][0], C.HEARD[2][1], "主任体验含在前期费"),
        ("听懂", C.HEARD[3][0], C.HEARD[3][1], "接口人权限边界"),
        ("需求", C.THEIR_NEEDS[0][0], C.THEIR_NEEDS[0][1], "层 A 外滩出海课"),
        ("需求", C.THEIR_NEEDS[1][0], C.THEIR_NEEDS[1][1], "场地与体验叙事"),
        ("需求", C.THEIR_NEEDS[2][0], C.THEIR_NEEDS[2][1], "层 B 主任体验"),
        ("需求", C.THEIR_NEEDS[3][0], C.THEIR_NEEDS[3][1], "层 C 备忘不验收"),
    ]
    for i, row in enumerate(need_rows, 5):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
        style_row(ws, i, 4, fill_light if i % 2 == 0 else fill_white)
        ws.row_dimensions[i].height = 36

    wb.save(path)
    return path


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else OUT
    p = build(out)
    print(f"已生成 {p}")
