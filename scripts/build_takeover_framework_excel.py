#!/usr/bin/env python3
"""生成《接手失败项目的通用框架》执行清单 Excel。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INK = "0F172A"
TEAL = "0F766E"
AMBER = "D97706"
SOFT = "F1F5F9"
MINT = "ECFDF5"
WHITE = "FFFFFF"
SLATE = "475569"

thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def style_header(ws, row, cols, fill_color=TEAL):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(name="微软雅黑", bold=True, color=WHITE, size=11)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def style_body(ws, start_row, end_row, cols):
    font = Font(name="微软雅黑", size=10, color=INK)
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=SOFT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title(ws, title, subtitle, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="微软雅黑", bold=True, size=16, color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = Font(name="微软雅黑", size=10, color=SLATE)
    ws.row_dimensions[2].height = 20


def build_excel(output_path: Path) -> None:
    wb = Workbook()

    # Sheet 1: 总览
    ws = wb.active
    ws.title = "框架总览"
    add_title(ws, "接手失败项目的通用框架", "全量摸底 → 按自己的逻辑重构 → 完成后及时转向", 5)
    headers = ["步骤", "核心问题", "建议周期", "关键输出", "完成标志"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 5)
    rows = [
        ["01 全量摸底", "现在到底是什么状态？", "3–5 天时间盒", "真相图 / 风险清单 / 停做清单", "干系人对现状陈述基本无异议"],
        ["02 逻辑重构", "按谁的逻辑前进？", "1–2 周", "新成功定义 / 杀留重写矩阵 / 主路径", "最短闭环可演示"],
        ["03 及时转向", "什么时候算接手完成？", "完成后 3–5 天交接", "完成标准 / Owner / playbook", "关闭接手专项，进入创造或交班"],
    ]
    for r_i, row in enumerate(rows):
        for c_i, val in enumerate(row, 1):
            ws.cell(row=5 + r_i, column=c_i, value=val)
    style_body(ws, 5, 7, 5)
    set_widths(ws, [16, 24, 16, 36, 32])
    ws.row_dimensions[4].height = 24
    for r in range(5, 8):
        ws.row_dimensions[r].height = 40

    # Sheet 2: 摸底清单
    ws2 = wb.create_sheet("01_全量摸底清单")
    add_title(ws2, "步骤一｜全量摸底清单", "只读优先；用绿/黄/红标注；事实与观点分离", 6)
    h2 = ["类别", "检查项", "状态(绿/黄/红)", "证据/链接", "负责人", "备注"]
    for i, h in enumerate(h2, 1):
        ws2.cell(row=4, column=i, value=h)
    style_header(ws2, 4, 6, TEAL)
    items2 = [
        ["目标与承诺", "原始目标与当前目标是否一致", "", "", "", ""],
        ["目标与承诺", "对外承诺清单（客户/老板/合作方）", "", "", "", ""],
        ["资产", "代码仓库与部署环境盘点", "", "", "", ""],
        ["资产", "数据资产与权限账号盘点", "", "", "", ""],
        ["资产", "文档、合同、供应商清单", "", "", "", ""],
        ["债务", "技术债 TOP10", "", "", "", ""],
        ["债务", "合规/安全/信誉风险点", "", "", "", ""],
        ["干系人", "决策人 / 影响人 / 执行人地图", "", "", "", ""],
        ["约束", "时间、预算、红线约束", "", "", "", ""],
        ["输出", "一页真相图（可复用/改造/冻结/待查）", "", "", "", ""],
        ["输出", "必须立刻停做的动作清单", "", "", "", ""],
        ["输出", "失败主因假设（含证据强度）", "", "", "", ""],
    ]
    for r_i, row in enumerate(items2):
        for c_i, val in enumerate(row, 1):
            ws2.cell(row=5 + r_i, column=c_i, value=val)
    style_body(ws2, 5, 16, 6)
    set_widths(ws2, [14, 40, 16, 28, 12, 24])
    for r in range(5, 17):
        ws2.row_dimensions[r].height = 28

    # Sheet 3: 重构决策
    ws3 = wb.create_sheet("02_杀留重写矩阵")
    add_title(ws3, "步骤二｜杀 / 留 / 重写决策矩阵", "默认不信任旧路线；先压缩到一条主路径", 7)
    h3 = ["模块/事项", "当前价值", "维护成本", "风险", "决策(杀/留/重写)", "下一步动作", "Owner"]
    for i, h in enumerate(h3, 1):
        ws3.cell(row=4, column=i, value=h)
    style_header(ws3, 4, 7, INK)
    for r in range(5, 20):
        for c in range(1, 8):
            ws3.cell(row=r, column=c, value="")
        # preset hints in first empty template rows via comments? keep blank for fill-in
    style_body(ws3, 5, 19, 7)
    # seed examples
    examples = [
        ["示例：无人维护旁路功能", "低", "高", "中", "杀", "下线并切断入口", ""],
        ["示例：核心结算链路", "高", "中", "高", "留", "写清契约，最小改动", ""],
        ["示例：主流程编排层", "高", "高", "高", "重写", "按新逻辑做最小切片", ""],
    ]
    for r_i, row in enumerate(examples):
        for c_i, val in enumerate(row, 1):
            ws3.cell(row=5 + r_i, column=c_i, value=val)
    set_widths(ws3, [28, 12, 12, 12, 16, 28, 12])
    for r in range(5, 20):
        ws3.row_dimensions[r].height = 26

    # Sheet 4: 完成与转向
    ws4 = wb.create_sheet("03_完成与转向")
    add_title(ws4, "步骤三｜接手完成标准与转向清单", "没有完成定义，就会永远停在清理态", 5)
    h4 = ["完成标准", "是否达成(是/否)", "证据", "Owner", "备注"]
    for i, h in enumerate(h4, 1):
        ws4.cell(row=4, column=i, value=h)
    style_header(ws4, 4, 5, AMBER)
    criteria = [
        ["主路径可演示且可复现", "", "", "", ""],
        ["关键风险已封堵或有明确 Owner", "", "", "", ""],
        ["干系人书面接受新目标与边界", "", "", "", ""],
        ["文档与职责完成交接", "", "", "", ""],
        ["接手专项看板已关闭或冻结", "", "", "", ""],
        ["playbook 已沉淀并可复用", "", "", "", ""],
        ["清理任务 WIP 已清零或移交", "", "", "", ""],
        ["下一阶段目标（创造/交班/关停）已确认", "", "", "", ""],
    ]
    for r_i, row in enumerate(criteria):
        for c_i, val in enumerate(row, 1):
            ws4.cell(row=5 + r_i, column=c_i, value=val)
    style_body(ws4, 5, 12, 5)
    set_widths(ws4, [40, 16, 28, 12, 24])
    for r in range(5, 13):
        ws4.row_dimensions[r].height = 28

    # Sheet 5: 反模式
    ws5 = wb.create_sheet("反模式对照")
    add_title(ws5, "反模式对照表", "框架一半价值在「坚决不做什么」", 3)
    h5 = ["反模式", "为什么危险", "替代动作"]
    for i, h in enumerate(h5, 1):
        ws5.cell(row=4, column=i, value=h)
    style_header(ws5, 4, 3, TEAL)
    antis = [
        ["边摸边大改", "无法归因新旧问题", "宣布只读期，改动一律登记"],
        ["讨好式继承", "双目标并行撕碎资源", "书面重定目标与非目标"],
        ["完美主义清理", "错过窗口期", "设完成定义与 WIP 上限"],
        ["工具崇拜", "无逻辑的效率幻觉", "先定操作系统再选工具"],
        ["口头交接", "问题回流到接手人", "书面完成标准 + Owner"],
        ["英雄主义", "经验不可迁移", "沉淀 playbook 并交班"],
    ]
    for r_i, row in enumerate(antis):
        for c_i, val in enumerate(row, 1):
            ws5.cell(row=5 + r_i, column=c_i, value=val)
    style_body(ws5, 5, 10, 3)
    set_widths(ws5, [18, 28, 32])
    for r in range(5, 11):
        ws5.row_dimensions[r].height = 30

    # Sheet 6: 场景速查
    ws6 = wb.create_sheet("场景速查")
    add_title(ws6, "场景速查：AI 项目管理 / 产品接手", "同一框架，不同落地动作", 4)
    h6 = ["场景", "步骤", "关键动作", "注意点"]
    for i, h in enumerate(h6, 1):
        ws6.cell(row=4, column=i, value=h)
    style_header(ws6, 4, 4, INK)
    scenes = [
        ["AI 项目管理", "摸底", "盘点指标/数据/Agent边界/评测/成本", "分清演示效果与可上线能力"],
        ["AI 项目管理", "重构", "重定人机分工；杀无效链路；建验收回退", "工具服从逻辑，先定编排"],
        ["AI 项目管理", "转向", "主链路达标后冻结探索；交稳态团队", "探索预算单列"],
        ["产品接手", "摸底", "用户路径 + 收入/留存/成本 + 访谈", "第 1 周完成真相图"],
        ["产品接手", "重构", "重写北极星与非目标；砍到主路径", "第 2–3 周统一叙事"],
        ["产品接手", "转向", "演示切片；明确 Owner/SLA；关专项看板", "第 4 周进入常态迭代"],
    ]
    for r_i, row in enumerate(scenes):
        for c_i, val in enumerate(row, 1):
            ws6.cell(row=5 + r_i, column=c_i, value=val)
    style_body(ws6, 5, 10, 4)
    set_widths(ws6, [16, 10, 42, 28])
    for r in range(5, 11):
        ws6.row_dimensions[r].height = 32

    wb.save(str(output_path))
    print(f"已生成: {output_path}")


if __name__ == "__main__":
    out = Path(__file__).resolve().parents[1] / "deliverables" / "接手失败项目的通用框架_执行清单.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    build_excel(out)
