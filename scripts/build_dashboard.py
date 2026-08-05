# -*- coding: utf-8 -*-
"""从《终表-上海板块学区融合总表-0805.xlsx》生成交互式 HTML 分析看板。

输出：output/上海板块学区融合分析看板-0805.html（单文件，无外部依赖，可离线打开）
用法：python3 scripts/build_dashboard.py
"""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "终表-上海板块学区融合总表-0805.xlsx"
OUT = ROOT / "output" / "上海板块学区融合分析看板-0805.html"

LEVELS = ["⭐⭐⭐ 强烈推荐", "⭐⭐ 可考虑", "⭐ 谨慎", "❌ 排除"]
LEVEL_KEY = {"⭐⭐⭐ 强烈推荐": "strong", "⭐⭐ 可考虑": "consider", "⭐ 谨慎": "caution", "❌ 排除": "exclude"}


def num(v):
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def load_rows():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["融合总表"]
    rows = []
    for r in ws.iter_rows(min_row=6, values_only=True):
        if r[7] is None:
            continue
        rows.append({
            "level": str(r[0] or ""), "conclusion": str(r[1] or ""), "memo": str(r[2] or ""),
            "picked": str(r[3] or ""), "score": num(r[4]), "fit": num(r[5]),
            "district": str(r[6] or ""), "name": str(r[7] or ""), "street": str(r[8] or ""),
            "primary": str(r[9] or ""), "middle": str(r[10] or ""), "enroll": str(r[11] or ""),
            "five_year": str(r[12] or ""), "hollow": num(r[13]), "newcomer": num(r[14]),
            "compete": num(r[15]), "pri_q": num(r[16]), "mid_q": num(r[17]),
            "price": str(r[18] or ""), "profile": str(r[19] or ""),
            "source": str(r[20] or ""), "note": str(r[21] or ""),
        })
    shortlist = []
    for r in wb["入选看房清单"].iter_rows(min_row=5, values_only=True):
        if r[0] is None:
            continue
        shortlist.append({
            "seq": str(r[0]), "level": str(r[1] or ""), "conclusion": str(r[2] or ""),
            "district": str(r[3] or ""), "name": str(r[4] or ""),
            "primary": str(r[5] or ""), "middle": str(r[6] or ""),
            "price": str(r[7] or ""), "score": num(r[8]), "reason": str(r[9] or ""),
        })
    districts = []
    for r in wb["十六区策略总览"].iter_rows(min_row=4, values_only=True):
        if r[1] is None:
            continue
        districts.append({
            "memo": str(r[0] or ""), "name": str(r[1] or ""), "enroll_type": str(r[2] or ""),
            "profile": str(r[3] or ""), "compete": str(r[4] or ""), "protect": str(r[5] or ""),
            "pressure": str(r[6] or ""), "strategy": str(r[7] or ""),
        })
    return rows, shortlist, districts


def build_stats(rows):
    level_count = {lv: 0 for lv in LEVELS}
    district_matrix = {}
    picked = 0
    for r in rows:
        if r["level"] in level_count:
            level_count[r["level"]] += 1
        d = district_matrix.setdefault(r["district"], {lv: 0 for lv in LEVELS})
        if r["level"] in d:
            d[r["level"]] += 1
        if r["picked"] == "是":
            picked += 1
    return {
        "total": len(rows), "level_count": level_count, "picked": picked,
        "district_matrix": district_matrix,
    }


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>上海板块 × 学区融合分析看板 · 0805</title>
<style>
:root{
  --navy:#0f2a43; --navy2:#16395b; --gold:#c9a227; --bg:#f2f4f7; --card:#ffffff;
  --ink:#1d2733; --sub:#5b6a7a; --line:#e3e8ee;
  --c-strong:#1a7f37; --c-consider:#b07d10; --c-caution:#c25e00; --c-exclude:#b3261e;
}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"PingFang SC","Microsoft YaHei","Source Han Sans SC","Noto Sans CJK SC",sans-serif;
  background:var(--bg);color:var(--ink);line-height:1.6}
.hero{background:linear-gradient(135deg,#0c2237 0%,#16395b 55%,#1d4a75 100%);color:#fff;padding:44px 28px 36px}
.hero-inner{max-width:1240px;margin:0 auto}
.hero .org{font-size:13px;letter-spacing:2px;color:#d9c37a;margin-bottom:10px}
.hero h1{font-size:30px;font-weight:700;letter-spacing:1px}
.hero .sub{margin-top:10px;font-size:14px;color:#b9c8d8;max-width:880px}
.hero .meta{margin-top:16px;font-size:12.5px;color:#8fa5bb;display:flex;gap:18px;flex-wrap:wrap}
.wrap{max-width:1240px;margin:0 auto;padding:26px 22px 60px}
.kpis{display:grid;grid-template-columns:repeat(6,1fr);gap:14px;margin-top:-30px}
.kpi{background:var(--card);border-radius:12px;padding:16px 14px;box-shadow:0 4px 16px rgba(15,42,67,.08);
  border-top:4px solid var(--navy2);text-align:center}
.kpi .v{font-size:30px;font-weight:800;color:var(--navy)}
.kpi .l{font-size:12.5px;color:var(--sub);margin-top:2px}
.kpi.s3{border-top-color:var(--c-strong)} .kpi.s3 .v{color:var(--c-strong)}
.kpi.s2{border-top-color:var(--c-consider)} .kpi.s2 .v{color:var(--c-consider)}
.kpi.s1{border-top-color:var(--c-caution)} .kpi.s1 .v{color:var(--c-caution)}
.kpi.s0{border-top-color:var(--c-exclude)} .kpi.s0 .v{color:var(--c-exclude)}
.kpi.sp{border-top-color:var(--gold)} .kpi.sp .v{color:var(--gold)}
section{margin-top:30px}
h2{font-size:19px;color:var(--navy);border-left:5px solid var(--gold);padding-left:10px;margin-bottom:6px}
.sec-note{font-size:12.5px;color:var(--sub);margin-bottom:14px}
.cards3{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}
.idea{background:var(--card);border-radius:12px;padding:18px;box-shadow:0 2px 10px rgba(15,42,67,.06)}
.idea h3{font-size:15px;color:var(--navy2);margin-bottom:8px}
.idea p{font-size:13px;color:var(--sub)}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.panel{background:var(--card);border-radius:12px;padding:18px;box-shadow:0 2px 10px rgba(15,42,67,.06)}
.panel h3{font-size:14.5px;color:var(--navy2);margin-bottom:4px}
.panel .hint{font-size:12px;color:var(--sub);margin-bottom:10px}
.legend{display:flex;gap:14px;flex-wrap:wrap;font-size:12px;color:var(--sub);margin:6px 0 4px}
.legend i{display:inline-block;width:11px;height:11px;border-radius:3px;margin-right:5px;vertical-align:-1px}
.filters{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.filters select,.filters input[type=text]{padding:8px 10px;border:1px solid var(--line);border-radius:8px;
  font-size:13px;background:#fff;color:var(--ink);outline:none}
.filters input[type=text]{width:200px}
.filters label{font-size:13px;color:var(--sub);display:flex;align-items:center;gap:5px}
.filters .count{margin-left:auto;font-size:13px;color:var(--navy2);font-weight:600}
.tablebox{overflow-x:auto;border:1px solid var(--line);border-radius:12px;background:#fff;max-height:640px;overflow-y:auto}
table{border-collapse:collapse;width:100%;font-size:12.5px;white-space:nowrap}
thead th{position:sticky;top:0;background:var(--navy);color:#fff;padding:9px 10px;text-align:left;
  font-weight:600;z-index:2}
tbody td{padding:8px 10px;border-bottom:1px solid var(--line);vertical-align:top;max-width:280px;
  overflow:hidden;text-overflow:ellipsis}
tbody tr:hover{background:#f4f8fc}
tr.picked{box-shadow:inset 4px 0 0 var(--gold)}
.badge{display:inline-block;padding:2px 8px;border-radius:20px;font-size:11.5px;font-weight:600;color:#fff}
.b-strong{background:var(--c-strong)} .b-consider{background:var(--c-consider)}
.b-caution{background:var(--c-caution)} .b-exclude{background:var(--c-exclude)}
.pill{display:inline-block;padding:1px 7px;border-radius:4px;font-size:11px;background:#eef3f8;color:var(--navy2);border:1px solid var(--line)}
.pill.gold{background:#fbf3d9;border-color:#e6d286;color:#7a6210;font-weight:700}
.tag-memo{font-size:11.5px;color:var(--sub)}
.route{display:flex;gap:10px;flex-wrap:wrap;margin-top:4px}
.route .stop{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:10px 14px;font-size:13px;
  display:flex;align-items:center;gap:8px;box-shadow:0 2px 8px rgba(15,42,67,.05)}
.route .stop b{color:var(--navy)}
.route .no{background:var(--gold);color:#fff;font-weight:800;border-radius:50%;width:22px;height:22px;
  display:inline-flex;align-items:center;justify-content:center;font-size:12px}
.tour{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}
.tcard{background:var(--card);border-radius:12px;padding:16px;box-shadow:0 2px 10px rgba(15,42,67,.06);
  border-top:4px solid var(--c-strong);position:relative}
.tcard.consider{border-top-color:var(--c-consider)}
.tcard .rank{position:absolute;top:10px;right:14px;font-size:26px;font-weight:800;color:#e8edf3}
.tcard h4{font-size:15px;color:var(--navy)}
.tcard .school{font-size:12.5px;color:var(--sub);margin:6px 0}
.tcard .price{font-size:13px;font-weight:700;color:var(--c-consider)}
.tcard .reason{font-size:12px;color:var(--sub);margin-top:8px;border-top:1px dashed var(--line);padding-top:8px}
.tcard .sc{font-size:12px;color:var(--navy2);font-weight:700}
svg text{font-family:inherit}
.foot{margin-top:40px;background:var(--navy);color:#9fb2c6;font-size:12px;padding:22px 28px;line-height:1.8}
.foot .in{max-width:1240px;margin:0 auto}
.foot b{color:#d9c37a}
@media (max-width:1000px){.kpis{grid-template-columns:repeat(3,1fr)}.cards3,.tour,.grid2{grid-template-columns:1fr}}
</style>
</head>
<body>
<header class="hero">
  <div class="hero-inner">
    <div class="org">复旦大学住房政策研究中心团队 · 决策辅助工具</div>
    <h1>上海板块 × 学区融合分析看板</h1>
    <div class="sub">以「人口空心化优先、避开新人口导入高竞争区」的差异化选房策略为主线，对全市 106 个板块样本做双源融合评估：量化评分底座（融合分析表）× 策略校准层（差异化选房纪要），选出本轮看房名单并逐板块标记。</div>
    <div class="meta">
      <span>数据日期：2026-08-05</span>
      <span>数据源①《上海板块学区融合分析》评分体系</span>
      <span>数据源②《差异化选房地图》纪要策略标记</span>
      <span>版本：终表-0805</span>
    </div>
  </div>
</header>
<div class="wrap">

  <div class="kpis" id="kpis"></div>

  <section>
    <h2>策略核心：跳出内卷的差异化选房逻辑</h2>
    <div class="sec-note">购房目标是孩子的优质学位资格，而非房产增值——两套逻辑指向完全不同的板块。</div>
    <div class="cards3">
      <div class="idea"><h3>① 选「人口空心化」老板块</h3><p>居住人口少于学校配套容量（学校在、人口出），对口校不易触发超额排序与统筹，人户一致更稳。全市 168 个板块中，黄浦、杨浦滨江最符合该特征。</p></div>
      <div class="idea"><h3>② 避开新上海人导入新区</h3><p>导入型新区适龄儿童集中出生，升学竞争极大。嘉定、松江、闵行大部分、浦东大部分板块排除；九亭属早期高竞争段位，明确不推荐。</p></div>
      <div class="idea"><h3>③ 先小学、后初中，分步解决</h3><p>不必一步到位买「小初双优」。先锁定小学学区（2027 年入学），小学阶段结束后再置换初中学区；高中靠统考，无需用学区房提前锁定。</p></div>
    </div>
  </section>

  <section>
    <h2>全市板块融合结果总览</h2>
    <div class="sec-note">106 个样本板块的推荐等级分布、行政区结构与「空心化 × 新上海人占比」策略定位。</div>
    <div class="grid2">
      <div class="panel">
        <h3>推荐等级构成</h3>
        <div class="hint">强烈推荐 15 · 可考虑 23 · 谨慎 21 · 排除 47（排除近半，策略高度聚焦）</div>
        <div id="donut"></div>
        <div class="legend">
          <span><i style="background:var(--c-strong)"></i>⭐⭐⭐ 强烈推荐</span>
          <span><i style="background:var(--c-consider)"></i>⭐⭐ 可考虑</span>
          <span><i style="background:var(--c-caution)"></i>⭐ 谨慎</span>
          <span><i style="background:var(--c-exclude)"></i>❌ 排除</span>
        </div>
      </div>
      <div class="panel">
        <h3>行政区 × 推荐等级（板块数）</h3>
        <div class="hint">黄浦、杨浦集中了绝大多数强烈推荐；近郊导入区几乎全线排除。</div>
        <div id="stacked"></div>
      </div>
    </div>
    <div class="grid2" style="margin-top:16px">
      <div class="panel">
        <h3>策略定位图：人口空心化 × 新上海人占比</h3>
        <div class="hint">右上金色区域为策略甜区（空心化 ≥4 且新上海人占比 ≤2）；气泡越大综合评分越高，悬停查看板块。</div>
        <div id="scatter"></div>
        <div class="legend">
          <span><i style="background:var(--c-strong)"></i>强烈推荐</span>
          <span><i style="background:var(--c-consider)"></i>可考虑</span>
          <span><i style="background:var(--c-caution)"></i>谨慎</span>
          <span><i style="background:var(--c-exclude)"></i>排除</span>
        </div>
      </div>
      <div class="panel">
        <h3>本轮入选 18 板块 · 综合评分</h3>
        <div class="hint">金色为「双源一致 / 纪要上调」的强烈推荐，琥珀色为纪要次优入选。</div>
        <div id="top18"></div>
      </div>
    </div>
  </section>

  <section>
    <h2>本轮入选看房清单（18 板块）</h2>
    <div class="sec-note">建议看房顺序：潍坊新村（世纪大道南）→ 杨浦鞍山 / 滨江 → 黄浦老城厢 → 虹口 / 曹杨等次优观察。</div>
    <div class="route">
      <div class="stop"><span class="no">1</span><span><b>浦东·潍坊新村</b>（世纪大道南、崂山一至十村，200–300万）</span></div>
      <div class="stop"><span class="no">2</span><span><b>杨浦·鞍山 / 东外滩滨江</b>（学校在、人口出）</span></div>
      <div class="stop"><span class="no">3</span><span><b>黄浦·老城厢组团</b>（豫园/老西门/小东门）</span></div>
      <div class="stop"><span class="no">4</span><span><b>虹口·北外滩 / 四川北路 · 普陀·曹杨</b>（次优观察）</span></div>
    </div>
    <div class="tour" id="tour" style="margin-top:14px"></div>
  </section>

  <section>
    <h2>全市 106 板块融合总表（可筛选）</h2>
    <div class="sec-note">金色左缘行为本轮入选板块；「评分」为五维量化模型结果，「契合分」为与纪要策略的贴合度（百分制）。</div>
    <div class="panel">
      <div class="filters">
        <select id="fDistrict"><option value="">全部行政区</option></select>
        <select id="fLevel"><option value="">全部推荐等级</option></select>
        <select id="fMemo"><option value="">全部策略标记</option></select>
        <input type="text" id="fSearch" placeholder="搜索板块 / 学校 / 街道…">
        <label><input type="checkbox" id="fPicked"> 仅看本轮入选</label>
        <span class="count" id="fCount"></span>
      </div>
      <div class="tablebox">
        <table id="mainTable">
          <thead><tr>
            <th>推荐等级</th><th>融合结论</th><th>纪要策略标记</th><th>入选</th><th>综合评分</th><th>策略契合分</th>
            <th>行政区</th><th>板块</th><th>对口小学（代表）</th><th>对口初中（代表）</th><th>入学方式</th><th>五年一户</th>
            <th>空心化</th><th>新沪占比</th><th>升学竞争</th><th>小学质量</th><th>初中质量</th><th>总价区间</th><th>人口结构画像</th><th>融合备注</th>
          </tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </section>

  <section>
    <h2>十六区策略分层</h2>
    <div class="sec-note">区级视角服务板块筛选：先看区的入学对口类型与竞争生态，再落到具体板块。</div>
    <div class="tablebox" style="max-height:none">
      <table id="distTable">
        <thead><tr><th>策略标记</th><th>行政区</th><th>入学对口类型</th><th>人口画像</th><th>区内竞争</th><th>本区保护</th><th>中考相对压力</th><th>策略结论</th></tr></thead>
        <tbody></tbody>
      </table>
    </div>
  </section>
</div>

<footer class="foot"><div class="in">
  <b>重要提示：</b>本看板为决策辅助工具，非官方划片文件。对口学校、入户年限、五年一户等以市教委及各区教育局当年招生文件为准；
  购房前务必核对当年招生简章与对口地段表。<br>
  出品：复旦大学住房政策研究中心团队 ｜ 数据源：《终表-上海板块学区融合总表-0805.xlsx》 ｜ 生成日期：2026-08-05
</div></footer>

<script>
const DATA = __DATA_JSON__;
const STATS = __STATS_JSON__;
const SHORTLIST = __SHORTLIST_JSON__;
const DISTRICTS16 = __DISTRICTS16_JSON__;
const LC = {"⭐⭐⭐ 强烈推荐":"#1a7f37","⭐⭐ 可考虑":"#b07d10","⭐ 谨慎":"#c25e00","❌ 排除":"#b3261e"};
const LK = {"⭐⭐⭐ 强烈推荐":"strong","⭐⭐ 可考虑":"consider","⭐ 谨慎":"caution","❌ 排除":"exclude"};
const LEVELS = Object.keys(LC);
const esc = s => String(s??"").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
const fmt = v => v==null ? "—" : (Number.isInteger(v)? v : v.toFixed(2).replace(/\\.?0+$/,""));

/* ---------- KPI ---------- */
document.getElementById("kpis").innerHTML = `
  <div class="kpi"><div class="v">${STATS.total}</div><div class="l">板块样本</div></div>
  <div class="kpi s3"><div class="v">${STATS.level_count["⭐⭐⭐ 强烈推荐"]}</div><div class="l">⭐⭐⭐ 强烈推荐</div></div>
  <div class="kpi s2"><div class="v">${STATS.level_count["⭐⭐ 可考虑"]}</div><div class="l">⭐⭐ 可考虑</div></div>
  <div class="kpi s1"><div class="v">${STATS.level_count["⭐ 谨慎"]}</div><div class="l">⭐ 谨慎</div></div>
  <div class="kpi s0"><div class="v">${STATS.level_count["❌ 排除"]}</div><div class="l">❌ 排除</div></div>
  <div class="kpi sp"><div class="v">${STATS.picked}</div><div class="l">本轮入选看房</div></div>`;

/* ---------- 环形图 ---------- */
(function(){
  const vals = LEVELS.map(l=>STATS.level_count[l]);
  const total = vals.reduce((a,b)=>a+b,0);
  const cx=110, cy=110, r=78, ir=50;
  let ang=-Math.PI/2, paths="";
  vals.forEach((v,i)=>{
    const a2 = ang + v/total*Math.PI*2;
    const x1=cx+r*Math.cos(ang), y1=cy+r*Math.sin(ang), x2=cx+r*Math.cos(a2), y2=cy+r*Math.sin(a2);
    const x3=cx+ir*Math.cos(a2), y3=cy+ir*Math.sin(a2), x4=cx+ir*Math.cos(ang), y4=cy+ir*Math.sin(ang);
    const large = (a2-ang)>Math.PI?1:0;
    paths += `<path d="M${x1},${y1} A${r},${r} 0 ${large} 1 ${x2},${y2} L${x3},${y3} A${ir},${ir} 0 ${large} 0 ${x4},${y4} Z" fill="${LC[LEVELS[i]]}"><title>${LEVELS[i]}：${v} 个（${(v/total*100).toFixed(1)}%）</title></path>`;
    ang=a2;
  });
  document.getElementById("donut").innerHTML =
    `<svg viewBox="0 0 220 220" style="width:100%;max-width:260px;display:block;margin:0 auto">${paths}
     <text x="${cx}" y="${cy-4}" text-anchor="middle" font-size="26" font-weight="800" fill="#0f2a43">${total}</text>
     <text x="${cx}" y="${cy+16}" text-anchor="middle" font-size="11" fill="#5b6a7a">板块样本</text></svg>`;
})();

/* ---------- 行政区堆叠条形图 ---------- */
(function(){
  const m = STATS.district_matrix;
  const rows = Object.keys(m).map(d=>({d, ...m[d], t: LEVELS.reduce((s,l)=>s+m[d][l],0)})).sort((a,b)=>b.t-a.t);
  const maxT = Math.max(...rows.map(r=>r.t));
  const W=560, barH=20, gap=7, left=46, top=6;
  const H = top + rows.length*(barH+gap);
  let svg = `<svg viewBox="0 0 ${W} ${H}" style="width:100%">`;
  rows.forEach((r,i)=>{
    const y = top+i*(barH+gap);
    svg += `<text x="${left-6}" y="${y+barH/2+4}" text-anchor="end" font-size="11.5" fill="#1d2733">${r.d}</text>`;
    let x = left;
    LEVELS.forEach(l=>{
      const w = r[l]/maxT*(W-left-34);
      if(w>0){
        svg += `<rect x="${x}" y="${y}" width="${w}" height="${barH}" fill="${LC[l]}" rx="2"><title>${r.d} · ${l}：${r[l]} 个</title></rect>`;
        if(w>14) svg += `<text x="${x+w/2}" y="${y+barH/2+4}" text-anchor="middle" font-size="10.5" fill="#fff">${r[l]}</text>`;
        x += w;
      }
    });
    svg += `<text x="${x+5}" y="${y+barH/2+4}" font-size="11" fill="#5b6a7a">${r.t}</text>`;
  });
  svg += `</svg>`;
  document.getElementById("stacked").innerHTML = svg;
})();

/* ---------- 散点图：空心化 × 新沪占比 ---------- */
(function(){
  const pts = DATA.filter(r=>r.hollow!=null && r.newcomer!=null);
  const W=560, H=360, L=44, R=16, T=14, B=40;
  const X = v => L + (v-0.5)/5*(W-L-R);
  const Y = v => T + (1-(v-0.5)/5)*(H-T-B);
  let svg = `<svg viewBox="0 0 ${W} ${H}" style="width:100%">`;
  // 甜区
  svg += `<rect x="${X(3.5)}" y="${T}" width="${X(5.5)-X(3.5)}" height="${Y(2.5)-T}" fill="#c9a227" opacity="0.14" rx="6"></rect>`;
  svg += `<text x="${X(4.98)}" y="${T+16}" text-anchor="end" font-size="11" fill="#9a7b12" font-weight="700">策略甜区：空心化高 × 新沪占比低</text>`;
  for(let i=1;i<=5;i++){
    svg += `<line x1="${X(i)}" y1="${T}" x2="${X(i)}" y2="${H-B}" stroke="#e8edf3"/>`;
    svg += `<text x="${X(i)}" y="${H-B+16}" text-anchor="middle" font-size="10.5" fill="#8fa0b0">${i}</text>`;
    svg += `<line x1="${L}" y1="${Y(i)}" x2="${W-R}" y2="${Y(i)}" stroke="#e8edf3"/>`;
    svg += `<text x="${L-6}" y="${Y(i)+4}" text-anchor="end" font-size="10.5" fill="#8fa0b0">${i}</text>`;
  }
  svg += `<text x="${(L+W-R)/2}" y="${H-6}" text-anchor="middle" font-size="11.5" fill="#5b6a7a">人口空心化程度（1→5 越高越空心化）→</text>`;
  svg += `<text x="14" y="${(T+H-B)/2}" text-anchor="middle" font-size="11.5" fill="#5b6a7a" transform="rotate(-90 14 ${(T+H-B)/2})">新上海人占比（1→5 越高越导入）→</text>`;
  pts.forEach((r,i)=>{
    // 确定性抖动避免完全重叠
    const jx = ((i*37)%17-8)*1.6, jy = ((i*53)%17-8)*1.6;
    const rad = 4 + (r.score? (r.score-1.3)*3.4 : 1.5);
    const cx = X(r.hollow)+jx, cy = Y(r.newcomer)+jy;
    svg += `<circle cx="${cx}" cy="${cy}" r="${Math.max(3.5,rad)}" fill="${LC[r.level]}" opacity="0.78" stroke="#fff" stroke-width="1"><title>${r.district}·${r.name}｜${r.level}｜评分 ${fmt(r.score)}｜空心化 ${r.hollow}｜新沪 ${r.newcomer}｜${r.price}</title></circle>`;
    if(r.picked==="是" && r.level===LEVELS[0])
      svg += `<text x="${cx}" y="${cy-Math.max(3.5,rad)-3}" text-anchor="middle" font-size="9.5" fill="#0f2a43" font-weight="700">${r.name}</text>`;
  });
  svg += `</svg>`;
  document.getElementById("scatter").innerHTML = svg;
})();

/* ---------- 入选 18 板块评分条 ---------- */
(function(){
  const arr = [...SHORTLIST].sort((a,b)=>(b.score??-1)-(a.score??-1));
  const W=560, barH=17, gap=6, left=118, top=4, maxS=4.5;
  const H = top+arr.length*(barH+gap);
  let svg = `<svg viewBox="0 0 ${W} ${H}" style="width:100%">`;
  arr.forEach((r,i)=>{
    const y=top+i*(barH+gap);
    const w = r.score? r.score/maxS*(W-left-46) : 6;
    const col = r.level.includes("强烈") ? "#c9a227" : "#b07d10";
    svg += `<text x="${left-6}" y="${y+barH/2+4}" text-anchor="end" font-size="11" fill="#1d2733">${r.district}·${r.name.length>7?r.name.slice(0,7)+"…":r.name}</text>`;
    svg += `<rect x="${left}" y="${y}" width="${Math.max(w,6)}" height="${barH}" rx="3" fill="${col}" opacity="${r.level.includes("强烈")?0.95:0.6}"><title>${r.district}·${r.name}｜${fmt(r.score)}</title></rect>`;
    svg += `<text x="${left+Math.max(w,6)+5}" y="${y+barH/2+4}" font-size="10.5" fill="#5b6a7a">${fmt(r.score)}</text>`;
  });
  svg += `</svg>`;
  document.getElementById("top18").innerHTML = svg;
})();

/* ---------- 看房清单卡片 ---------- */
document.getElementById("tour").innerHTML = SHORTLIST.map(r=>`
  <div class="tcard ${r.level.includes("强烈")?"":"consider"}">
    <div class="rank">${esc(r.seq)}</div>
    <span class="badge b-${LK[r.level]||"consider"}">${esc(r.level)}</span>
    <span class="pill gold">${esc(r.conclusion)}</span>
    <h4 style="margin-top:8px">${esc(r.district)} · ${esc(r.name)}</h4>
    <div class="school">小学：${esc(r.primary)}<br>初中：${esc(r.middle)}</div>
    <div class="price">${esc(r.price)} <span class="sc">｜综合评分 ${fmt(r.score)}</span></div>
    <div class="reason">${esc(r.reason)}</div>
  </div>`).join("");

/* ---------- 主表 ---------- */
const memoBadge = m => {
  if(!m || m==="—") return '<span class="tag-memo">—</span>';
  const color = m.includes("优选") ? "#1a7f37" : m.includes("次优") ? "#b07d10" : m.includes("排除") ? "#b3261e" : m.includes("不符") ? "#c25e00" : "#5b6a7a";
  return `<span class="pill" style="color:${color};border-color:${color}33;background:${color}11">${esc(m)}</span>`;
};
const scoreCell = v => v==null ? "—" : `<b style="color:${v>=4?"#1a7f37":v>=3?"#b07d10":v>=2?"#c25e00":"#b3261e"}">${fmt(v)}</b>`;
const dimCell = v => v==null ? "—" : v;

function renderTable(){
  const d = document.getElementById("fDistrict").value;
  const l = document.getElementById("fLevel").value;
  const m = document.getElementById("fMemo").value;
  const p = document.getElementById("fPicked").checked;
  const q = document.getElementById("fSearch").value.trim().toLowerCase();
  const rows = DATA.filter(r =>
    (!d || r.district===d) && (!l || r.level===l) && (!m || r.memo===m) &&
    (!p || r.picked==="是") &&
    (!q || (r.name+r.district+r.primary+r.middle+r.street+r.note).toLowerCase().includes(q))
  );
  document.querySelector("#mainTable tbody").innerHTML = rows.map(r=>`
    <tr class="${r.picked==="是"?"picked":""}">
      <td><span class="badge b-${LK[r.level]||"consider"}">${esc(r.level)}</span></td>
      <td>${esc(r.conclusion)}</td><td>${memoBadge(r.memo)}</td>
      <td>${r.picked==="是"?'<span class="pill gold">入选</span>':""}</td>
      <td>${scoreCell(r.score)}</td><td>${r.fit==null?"—":fmt(r.fit)}</td>
      <td>${esc(r.district)}</td><td><b>${esc(r.name)}</b></td>
      <td title="${esc(r.primary)}">${esc(r.primary)}</td><td title="${esc(r.middle)}">${esc(r.middle)}</td>
      <td>${esc(r.enroll)}</td><td>${esc(r.five_year)}</td>
      <td>${dimCell(r.hollow)}</td><td>${dimCell(r.newcomer)}</td><td>${dimCell(r.compete)}</td>
      <td>${dimCell(r.pri_q)}</td><td>${dimCell(r.mid_q)}</td>
      <td>${esc(r.price)}</td><td title="${esc(r.profile)}">${esc(r.profile)}</td><td title="${esc(r.note)}">${esc(r.note)}</td>
    </tr>`).join("");
  document.getElementById("fCount").textContent = `共 ${rows.length} / ${DATA.length} 个板块`;
}
(function(){
  const ds=[...new Set(DATA.map(r=>r.district))];
  const ms=[...new Set(DATA.map(r=>r.memo).filter(x=>x&&x!=="—"))];
  document.getElementById("fDistrict").innerHTML += ds.map(x=>`<option>${esc(x)}</option>`).join("");
  document.getElementById("fLevel").innerHTML += LEVELS.map(x=>`<option>${esc(x)}</option>`).join("");
  document.getElementById("fMemo").innerHTML += ms.map(x=>`<option>${esc(x)}</option>`).join("");
  ["fDistrict","fLevel","fMemo","fPicked","fSearch"].forEach(id=>
    document.getElementById(id).addEventListener("input", renderTable));
  renderTable();
})();

/* ---------- 十六区表 ---------- */
document.querySelector("#distTable tbody").innerHTML = DISTRICTS16.map(r=>{
  const color = r.memo.includes("优选") ? "#1a7f37" : r.memo.includes("次优") ? "#b07d10" : r.memo.includes("排除") ? "#b3261e" : r.memo.includes("不符") ? "#c25e00" : "#5b6a7a";
  return `<tr>
    <td><span class="pill" style="color:${color};border-color:${color}33;background:${color}11">${esc(r.memo)}</span></td>
    <td><b>${esc(r.name)}</b></td><td>${esc(r.enroll_type)}</td>
    <td style="white-space:normal;min-width:200px">${esc(r.profile)}</td>
    <td>${esc(r.compete)}</td><td>${esc(r.protect)}</td><td>${esc(r.pressure)}</td>
    <td style="white-space:normal;min-width:220px">${esc(r.strategy)}</td></tr>`;
}).join("");
</script>
</body>
</html>
"""


def main():
    rows, shortlist, districts = load_rows()
    stats = build_stats(rows)
    html = (HTML_TEMPLATE
            .replace("__DATA_JSON__", json.dumps(rows, ensure_ascii=False))
            .replace("__STATS_JSON__", json.dumps(stats, ensure_ascii=False))
            .replace("__SHORTLIST_JSON__", json.dumps(shortlist, ensure_ascii=False))
            .replace("__DISTRICTS16_JSON__", json.dumps(districts, ensure_ascii=False)))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding="utf-8")
    print(f"已生成 {OUT}（{len(html)/1024:.0f} KB，{stats['total']} 板块，入选 {stats['picked']}）")


if __name__ == "__main__":
    main()
