"""生成 Excel：学生实习赋能计划 · 可行性论证数据表

Sheets：
  00 总览Dashboard
  01 人群需求分析
  02 商业实习-大厂清单
  03 商业实习-科技机器人
  04 战略取舍-取消开证明
  05 收费标准与定价
  06 盈利测算
  07 落地实施计划
  08 风险与合规
  09 KPI跟踪
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

PRIMARY = "0B3D5C"
ACCENT = "1A7A6D"
LIGHT = "E8F3F1"
SOFT = "D0E8E4"
GOLD = "C4A35A"
GREY = "5A6A72"
WHITE = "FFFFFF"
DARK = "1A2A33"
ORANGE = "C06A2F"
RED = "B54A4A"
YELLOW = "FFF3CD"
GREEN_BG = "D4EDDA"
RED_BG = "F8D7DA"


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_row(ws, row, headers, fill=PRIMARY, color=WHITE, height=26):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="微软雅黑", size=11, bold=True, color=color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[row].height = height


def write_row(ws, row, data, fill=None, bold=False, size=10, wrap=True, align="left"):
    for i, v in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name="微软雅黑", size=size, bold=bold, color=DARK)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border = thin()
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        elif row % 2 == 0:
            c.fill = PatternFill("solid", fgColor=LIGHT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def section_banner(ws, row, text, span=10, fill=ACCENT, color=WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", size=12, bold=True, color=color)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row=row, column=col).border = thin()
    ws.row_dimensions[row].height = 24


def big_title(ws, span_col="L", text="", sub=""):
    ws.merge_cells(f"A1:{span_col}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(name="微软雅黑", size=18, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A2:{span_col}2")
    c2 = ws["A2"]
    c2.value = sub
    c2.font = Font(name="微软雅黑", size=10, italic=True, color=GREY)
    c2.fill = PatternFill("solid", fgColor=LIGHT)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20


def add_status_validation(ws, col_letter, start_row, end_row):
    dv = DataValidation(
        type="list",
        formula1='"未开始,进行中,已完成,延期,取消"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def style_kv(ws, row, key, value, span_end=8):
    c1 = ws.cell(row=row, column=1, value=key)
    c1.font = Font(name="微软雅黑", bold=True, color=DARK, size=10)
    c1.fill = PatternFill("solid", fgColor=LIGHT)
    c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c1.border = thin()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span_end)
    c2 = ws.cell(row=row, column=2, value=value)
    c2.font = Font(name="微软雅黑", color=DARK, size=10)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = thin()
    for col in range(3, span_end + 1):
        ws.cell(row=row, column=col).border = thin()


def build_dashboard(wb):
    ws = wb.active
    ws.title = "00 总览Dashboard"
    big_title(
        ws, "L",
        "学生实习赋能计划 · 总览 Dashboard",
        "高中生 / 大学生 / 毕业生  |  高价值商业实习 × 就业转化  |  已取消开证明低价档  |  V1.1",
    )

    section_banner(ws, 4, "一、项目基本信息", span=12, fill=PRIMARY)
    info = [
        ["项目名称", "学生实习赋能计划（商业企业课题 + 就业转化）"],
        ["目标人群", "高中生（申请向）· 大学生（履历向）· 毕业生（就业转化向）"],
        ["核心目标", "辅助就业转化；提升考学与申报竞争力；建立可持续盈利模式"],
        ["产品结构", "只做高价值：商业标准 / 商业进阶 / 就业转化（已停售开证明）"],
        ["已停售", "199 / 699 / 1,680 开证明三档（费用低、成本与风险高）"],
        ["试点周期", "6 个月（建议）"],
        ["基准情景", "年付费 200 人 · 均客单 8,240 元 · 营收约 164.8 万 · 毛利率约 38%"],
        ["盈亏平衡", "固定成本约 28–35 万/年；取消低价档后基准情景更易盈余"],
        ["配套交付", "论证 PPT + 流程专项 PPT + 本 Excel"],
    ]
    for i, (k, v) in enumerate(info, 5):
        style_kv(ws, i, k, v, span_end=12)

    section_banner(ws, 15, "二、产品线收入结构（基准情景 · 已无开证明档）", span=12, fill=ACCENT)
    header_row(ws, 16, ["产品线", "人数占比", "人数", "客单价(元)", "收入(元)", "收入占比", "毛利率", "毛利(元)"])
    # 人数 200；仅高价值三档
    lines = [
        ("商业标准(科技/具身)", 0.50, 5480, 0.34),
        ("商业进阶(上海AI/头部)", 0.30, 11800, 0.42),
        ("毕业生就业转化包", 0.20, 9800, 0.40),
    ]
    total_people = 200
    for i, (name, pct, price, margin) in enumerate(lines, 17):
        people = round(total_people * pct)
        revenue = people * price
        profit = round(revenue * margin)
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=pct).number_format = "0%"
        ws.cell(row=i, column=3, value=people)
        ws.cell(row=i, column=4, value=price)
        ws.cell(row=i, column=5, value=revenue)
        ws.cell(row=i, column=6, value=f"=E{i}/SUM($E$17:$E$19)")
        ws.cell(row=i, column=6).number_format = "0.0%"
        ws.cell(row=i, column=7, value=margin).number_format = "0%"
        ws.cell(row=i, column=8, value=profit)
        for col in range(1, 9):
            c = ws.cell(row=i, column=col)
            c.font = Font(name="微软雅黑", size=10, color=DARK)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin()
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=LIGHT)

    # totals
    ws.cell(row=20, column=1, value="合计").font = Font(name="微软雅黑", bold=True, color=WHITE)
    for col in range(1, 9):
        c = ws.cell(row=20, column=col)
        c.fill = PatternFill("solid", fgColor=PRIMARY)
        c.font = Font(name="微软雅黑", bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
    ws.cell(row=20, column=3, value="=SUM(C17:C19)")
    ws.cell(row=20, column=5, value="=SUM(E17:E19)")
    ws.cell(row=20, column=6, value=1).number_format = "0%"
    ws.cell(row=20, column=8, value="=SUM(H17:H19)")
    for col in (3, 5, 8):
        ws.cell(row=20, column=col).font = Font(name="微软雅黑", bold=True, color=WHITE)
        ws.cell(row=20, column=col).fill = PatternFill("solid", fgColor=PRIMARY)

    chart = BarChart()
    chart.type = "col"
    chart.title = "各产品线收入（元）"
    chart.y_axis.title = "收入"
    data = Reference(ws, min_col=5, min_row=16, max_row=19)
    cats = Reference(ws, min_col=1, min_row=17, max_row=19)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "A22")

    section_banner(ws, 38, "三、三情景摘要（详见「06 盈利测算」）", span=12, fill=PRIMARY)
    header_row(ws, 39, ["情景", "付费人数", "均客单(元)", "营收(万元)", "毛利率", "毛利(万元)", "是否覆盖固定成本"])
    scenarios = [
        ("保守", 100, 7500, 75.0, "35%", 26.3, "临界偏稳"),
        ("基准", 200, 8240, 164.8, "38%", 62.6, "是"),
        ("乐观", 300, 9000, 270.0, "42%", 113.4, "是，可扩张"),
    ]
    for i, row in enumerate(scenarios, 40):
        write_row(ws, i, list(row), align="center")

    set_widths(ws, [24, 12, 12, 14, 14, 12, 12, 14, 12, 12, 12, 12])
    ws.freeze_panes = "A4"


def build_audience(wb):
    ws = wb.create_sheet("01 人群需求分析")
    big_title(ws, "J", "三类目标人群需求分析", "高中生 · 大学生 · 毕业生 — 痛点 / 诉求 / 交付 / 付费")

    section_banner(ws, 4, "一、人群对比总表", span=10, fill=PRIMARY)
    header_row(ws, 5, [
        "人群", "核心目标", "关键痛点", "关键交付物", "决策者",
        "付费意愿", "价格敏感带(元)", "转化出口", "推荐产品线", "备注",
    ])
    rows = [
        ["高中生", "海外/名校申请加分", "缺可核验实践；家长怕水经历",
         "真实企业课题+结业证明+可选推荐信", "家长为主", "中高", "5,480–11,800",
         "大学录取 / 活动列表", "商业标准/进阶", "不卖开证明；走项目档"],
        ["大学生", "丰富履历、校招竞争力", "大厂难进；履历同质化",
         "品牌实习经历、项目成果、导师评价", "个人+家庭", "中", "5,480–15,800",
         "暑期转正 / 校招 Offer", "商业标准/进阶", "与专业匹配更重要"],
        ["毕业生", "转入正式就业岗位", "空窗期；社招门槛高",
         "过渡实习+作品集+内推辅导", "个人", "中高（急切）", "6,800–12,800",
         "全职 Offer", "就业转化包", "强调转化率与售后"],
    ]
    for i, r in enumerate(rows, 6):
        write_row(ws, i, r)
        ws.row_dimensions[i].height = 40

    section_banner(ws, 10, "二、需求细分（可勾选落地优先级）", span=10, fill=ACCENT)
    header_row(ws, 11, ["人群", "需求条目", "重要度(1-5)", "当前供给缺口", "建议动作", "优先级", "负责人", "状态"])
    details = [
        ["高中生", "海外申请真实课题经历", 5, "缺正规高价值通道", "导入商业标准/进阶项目档", "P0", "", "未开始"],
        ["高中生", "结业证明与推荐信", 4, "不可再靠低价开证明", "仅随项目交付；推荐信加购", "P0", "", "未开始"],
        ["大学生", "大厂/上海AI品牌经历", 5, "名额稀缺、信息不对称", "建立分层企业合作池", "P0", "", "未开始"],
        ["大学生", "科技/具身智能硬科技履历", 5, "学生不知如何进入", "上海具身智能清单优先签约", "P0", "", "未开始"],
        ["大学生", "简历与面试辅导", 3, "分散在其他机构", "作为项目内增值加购", "P2", "", "未开始"],
        ["毕业生", "过渡实习岗位", 5, "毕业后无人管", "就业转化包 4–12 周", "P0", "", "未开始"],
        ["毕业生", "内推与 Offer 转化", 5, "缺少闭环", "合作企业内推+复盘", "P0", "", "未开始"],
        ["战略", "停售开证明低价档", 5, "费用低风险高", "199/699/1680 一律停售", "P0", "", "进行中"],
        ["共性", "真实课题可核验", 5, "行业信任危机", "过程材料留存+企业评语", "P0", "", "未开始"],
    ]
    for i, r in enumerate(details, 12):
        write_row(ws, i, r, align="center")
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    add_status_validation(ws, "H", 12, 20)

    set_widths(ws, [10, 22, 12, 18, 28, 12, 14, 16, 16, 18])
    ws.freeze_panes = "A4"


def build_fortune(wb):
    ws = wb.create_sheet("02 商业实习-大厂清单")
    big_title(ws, "K", "商业实习项目 · 世界500强与知名大厂合作池",
              "论证用清单：落地需商务签约；状态可下拉更新")

    section_banner(ws, 4, "企业清单", span=11, fill=PRIMARY)
    header_row(ws, 5, [
        "序号", "层级", "企业名称", "类型", "方向示例",
        "建议产品形态", "适合人群", "合作难度", "对接状态", "预估通道成本(元/人)", "备注",
    ])
    companies = [
        ["S", "Google", "全球科技", "产品/数据/工程", "远程课题+导师营", "大学生/优生", "高", "未开始", 4500, "需合规表述为合作课题/合作方"],
        ["S", "Microsoft", "全球科技", "云/开发/数据分析", "远程课题+导师营", "大学生", "高", "未开始", 4200, ""],
        ["S", "Amazon", "全球科技", "运营/数据/云", "项目制", "大学生", "高", "未开始", 4000, ""],
        ["S", "Meta", "全球科技", "产品/增长/内容", "远程课题", "大学生", "高", "未开始", 4500, ""],
        ["S", "腾讯", "互联网大厂", "产品运营/研发/市场", "暑期营/项目制", "大学生", "中高", "未开始", 3500, "优先拓展"],
        ["S", "阿里巴巴", "互联网大厂", "电商运营/技术/数据", "项目制", "大学生", "中高", "未开始", 3500, ""],
        ["S", "字节跳动", "互联网大厂", "内容/产品/算法应用", "项目制", "大学生", "中高", "未开始", 3800, "含火山引擎协同"],
        ["S", "美团", "互联网大厂", "本地生活运营/数据", "项目制", "大学生", "中", "未开始", 2800, ""],
        ["A", "火山引擎", "云与AI", "大模型应用/数据", "课题实训营", "大学生", "中", "未开始", 3000, "重点开拓"],
        ["A", "阿里云", "云与AI", "云原生/解决方案", "课题实训营", "理工大学生", "中", "未开始", 3000, ""],
        ["A", "腾讯云", "云与AI", "云与安全", "课题实训营", "理工大学生", "中", "未开始", 3000, ""],
        ["A", "华为云", "云与AI", "云/鸿蒙生态应用", "课题实训营", "理工大学生", "中高", "未开始", 3200, ""],
        ["A", "小米", "消费电子", "市场/用户研究", "短期项目实习", "大学生", "中", "未开始", 2500, ""],
        ["A", "OPPO", "消费电子", "产品市场", "短期项目实习", "大学生", "中", "未开始", 2400, ""],
        ["A", "vivo", "消费电子", "品牌/用户洞察", "短期项目实习", "大学生", "中", "未开始", 2400, ""],
        ["B", "IBM", "综合500强", "咨询/数字化", "校企联合课题", "大学生/毕业生", "中", "未开始", 2200, ""],
        ["B", "SAP", "综合500强", "企业软件实施", "项目助理", "信息管理/商科", "中", "未开始", 2200, ""],
        ["B", "西门子", "综合500强", "工业数字化", "课题实习", "理工", "中高", "未开始", 2500, ""],
        ["B", "联合利华", "综合500强", "品牌/市场", "项目制", "商科", "中", "未开始", 2000, ""],
        ["B", "宝洁", "综合500强", "品牌管理", "项目制", "商科", "中高", "未开始", 2200, ""],
        ["B", "蚂蚁集团", "金融科技", "风控/运营/数据", "远程项目实习", "大学生/毕业生", "中高", "未开始", 3000, ""],
        ["B", "京东科技", "金融科技", "数科产品/运营", "项目制", "大学生", "中", "未开始", 2600, ""],
        ["B", "平安科技", "金融科技", "科技赋能/数据", "项目制", "大学生", "中", "未开始", 2400, ""],
    ]
    for i, r in enumerate(companies, 6):
        write_row(ws, i, [i - 5] + r, align="center")
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=i, column=11).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    add_status_validation(ws, "I", 6, 5 + len(companies))

    dv2 = DataValidation(type="list", formula1='"低,中,中高,高"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"H6:H{5 + len(companies)}")

    section_banner(ws, 31, "分层定价挂钩建议", span=11, fill=ACCENT)
    header_row(ws, 32, ["层级", "建议学员报价(元)", "通道成本占比目标", "产品名称建议", "最短周期", "是否含导师答辩"])
    tiers = [
        ["S 级", "8,800–15,800", "≤40%", "大厂/全球科技通道营", "4–6 周", "是"],
        ["A 级", "5,980–9,800", "≤35%", "云与 AI / 消费电子项目营", "4 周", "是"],
        ["B 级", "3,980–6,980", "≤30%", "综合 500 强 / 金科项目实习", "4 周", "可选"],
    ]
    for i, r in enumerate(tiers, 33):
        write_row(ws, i, r, align="center")

    set_widths(ws, [6, 8, 14, 12, 20, 18, 14, 10, 10, 18, 28])
    ws.freeze_panes = "A6"


def build_tech(wb):
    ws = wb.create_sheet("03 商业实习-科技机器人")
    big_title(ws, "K", "商业实习项目 · 机器人及其他科技公司",
              "硬科技差异化履历；适合理工与交叉背景学生")

    section_banner(ws, 4, "企业与课题清单", span=11, fill=PRIMARY)
    header_row(ws, 5, [
        "序号", "类别", "企业名称", "适合人群", "岗位/课题方向",
        "产品形态", "建议定价(元)", "合作难度", "对接状态", "城市/远程", "备注",
    ])
    rows = [
        ["人形/工业机器人", "优必选", "理工大学生", "算法/机械/测试", "实岗或课题", 5980, "中", "未开始", "深圳/混合", ""],
        ["人形/工业机器人", "宇树科技", "理工大学生", "嵌入式/运动控制", "课题实训", 6480, "中高", "未开始", "杭州/混合", "热门，名额紧"],
        ["人形/工业机器人", "节卡机器人", "机械/自动化", "应用工程师助理", "项目制", 4980, "中", "未开始", "上海", ""],
        ["人形/工业机器人", "埃斯顿", "自动化相关", "工业机器人应用", "项目制", 4980, "中", "未开始", "南京", ""],
        ["自动驾驶/出行", "小马智行", "计算机/车辆", "感知标注/仿真支持", "远程+驻场", 6800, "中高", "未开始", "广州/远程", ""],
        ["自动驾驶/出行", "文远知行", "计算机相关", "运营支持/数据", "项目制", 6500, "中高", "未开始", "广州", ""],
        ["自动驾驶/出行", "Momenta", "算法相关", "数据闭环支持", "课题", 6800, "高", "未开始", "苏州/远程", ""],
        ["具身智能/AI硬件", "智元机器人", "硕本交叉", "数据采集/应用demo", "课题营", 7200, "中高", "未开始", "上海", ""],
        ["具身智能/AI硬件", "银河通用", "AI/机器人", "具身数据与评测", "课题", 7000, "中高", "未开始", "北京", ""],
        ["具身智能/AI硬件", "云深处科技", "机械/控制", "四足机器人应用", "项目制", 5800, "中", "未开始", "杭州", ""],
        ["企业服务SaaS", "用友", "商科/信管", "实施助理/客户成功", "项目制", 3980, "低", "未开始", "远程/北京", "易起量"],
        ["企业服务SaaS", "金蝶", "商科/信管", "实施/支持", "项目制", 3980, "低", "未开始", "远程/深圳", ""],
        ["企业服务SaaS", "销售易", "市场营销/信管", "售前支持/运营", "远程实习", 4200, "中", "未开始", "远程", ""],
        ["企业服务SaaS", "北森", "人力/心理/信管", "HR SaaS 运营支持", "远程实习", 4200, "中", "未开始", "远程", ""],
        ["半导体/芯片生态", "寒武纪", "电子信息", "工具链文档/资料研究", "课题研究", 5500, "高", "未开始", "北京/远程", "表述需谨慎"],
        ["半导体/芯片生态", "地平线", "电子/计算机", "嵌入式工具支持", "课题", 5800, "高", "未开始", "上海/远程", ""],
        ["生物科技/医疗AI", "联影智能", "生医/AI交叉", "标注/产品助理", "项目制", 5200, "中", "未开始", "上海", ""],
        ["生物科技/医疗AI", "推想科技", "生医/AI", "文献/数据标注", "远程项目", 4800, "中", "未开始", "远程", ""],
        ["生物科技/医疗AI", "数坤科技", "生医交叉", "产品助理支持", "项目制", 5000, "中", "未开始", "北京", ""],
    ]
    for i, r in enumerate(rows, 6):
        write_row(ws, i, [i - 5] + r, align="center")
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    add_status_validation(ws, "I", 6, 5 + len(rows))

    section_banner(ws, 27, "赛道组合建议（试点优先）", span=11, fill=ACCENT)
    header_row(ws, 28, ["优先级", "赛道", "首批目标签约数", "理由", "主要客群"])
    focus = [
        ["P0", "企业服务 SaaS", 4, "合作门槛低、远程友好、易标准化交付", "商科/综合类大学生"],
        ["P0", "人形/工业机器人", 3, "履历辨识度高、家长与学生兴趣强", "理工大学生"],
        ["P1", "具身智能/AI 硬件", 2, "热点赛道，支撑进阶定价", "优生/交叉背景"],
        ["P1", "医疗 AI", 2, "差异化，适合生医交叉", "相关专业"],
        ["P2", "自动驾驶/芯片", 2, "品牌强但对接周期长", "对口专业"],
    ]
    for i, r in enumerate(focus, 29):
        write_row(ws, i, r, align="center")
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    set_widths(ws, [6, 16, 14, 14, 22, 12, 12, 10, 10, 12, 16])
    ws.freeze_panes = "A6"


def build_strategy(wb):
    ws = wb.create_sheet("04 战略取舍-取消开证明")
    big_title(ws, "I", "战略取舍：取消「开证明」业务",
              "费用低 · 成本与风险高 → 一律停售 199 / 699 / 1,680，聚焦高价值商业与就业档")

    section_banner(ws, 4, "一、已停售产品（不再报价、不再交付）", span=9, fill=RED)
    header_row(ws, 5, [
        "原产品代码", "原产品名称", "原定价(元)", "停售原因", "替代方案", "状态",
    ], fill=RED)
    stopped = [
        ["GW-01", "线上实习·证明版", 199, "客单过低；核验/舆情成本高", "导入商业标准项目 5,480", "已停售"],
        ["GW-02", "证明+申报支持", 699, "费用低、运营带宽占用大", "随项目交付结业证明", "已停售"],
        ["GW-03", "证明+推荐信协助", 1680, "单独售卖风险高、易被质疑水经历", "仅项目学员加购 +1,500", "已停售"],
    ]
    for i, r in enumerate(stopped, 6):
        write_row(ws, i, r, align="center", fill=RED_BG)
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 32

    section_banner(ws, 10, "二、为什么取消（决策依据）", span=9, fill=ORANGE)
    reasons = [
        ["客单价", "199–1,680 元档难以覆盖真实企业对接与合规成本"],
        ["风险", "「开证明」易引发不实经历质疑，品牌与舆情风险不对等"],
        ["机会成本", "同样教务/企业带宽可售 5,480–11,800 高价值档"],
        ["战略聚焦", "三类人群需求改由真实企业课题满足，证明只作交付物"],
    ]
    header_row(ws, 11, ["维度", "说明"])
    for i, r in enumerate(reasons, 12):
        write_row(ws, i, r)

    section_banner(ws, 17, "三、在售高价值产品（替代组合）", span=9, fill=ACCENT)
    header_row(ws, 18, [
        "产品代码", "产品名称", "标准价(元)", "交付物", "周期", "主要人群", "备注",
    ])
    live = [
        ["BZ-01", "科技/具身智能项目实习", 5480, "真实课题+周报+结业证明", "4–6 周", "高中优生/大学生", "主推"],
        ["BZ-02", "上海 AI / 头部通道营", 11800, "品牌项目+导师+答辩+证明", "4–6 周", "大学生优生", "高毛利"],
        ["JY-01", "毕业生就业转化包", 9800, "过渡实习+内推辅导+证明", "6–12 周", "毕业生", "转化出口"],
        ["ZZ-01", "推荐信加购（不可单卖）", 1500, "中/英推荐信", "随项目", "在营学员", "可选"],
    ]
    for i, r in enumerate(live, 19):
        write_row(ws, i, r, align="center")

    section_banner(ws, 24, "四、对外话术红线", span=9, fill=PRIMARY)
    lines = [
        "不得再宣传或报价任何「开证明 / 盖章证明 / 低价实习证明」单独产品。",
        "结业证明仅作为高价值项目履约交付物，须基于真实课题完成情况出具。",
        "推荐信不得单独售卖；仅可对已参加项目的学员加购，且内容须基于真实表现。",
    ]
    for i, t in enumerate(lines, 25):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
        c = ws.cell(row=i, column=1, value=f"• {t}")
        c.font = Font(name="微软雅黑", size=10, color=DARK)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.fill = PatternFill("solid", fgColor=LIGHT)
        ws.row_dimensions[i].height = 28

    set_widths(ws, [12, 26, 12, 36, 22, 16, 12, 12, 12])
    ws.freeze_panes = "A4"


def build_pricing(wb):
    ws = wb.create_sheet("05 收费标准与定价")
    big_title(ws, "L", "收费标准与定价机制",
              "仅高价值档 · 成本加成 · 支付节奏 · 折扣政策 · 已取消开证明")

    section_banner(ws, 4, "一、面向 C 端标准价目表（在售）", span=12, fill=PRIMARY)
    header_row(ws, 5, [
        "产品线", "套餐名称", "建议定价下限", "建议定价上限", "标准价(用于测算)",
        "包含内容", "不含内容", "目标人群", "最短周期", "退费政策要点", "推荐渠道", "上架优先级",
    ])
    price_rows = [
        ["商业", "科技/具身智能项目实习", 3980, 6980, 5480,
         "4–6周真实课题+周报+结业证明", "包录取/包转正", "高中优生/大学生", "4 周",
         "定金不退，企业未确认前可协商", "高校社团", "P0"],
        ["商业", "上海AI/头部通道营", 8800, 15800, 11800,
         "品牌项目+导师+答辩+证明", "包转正", "大学生/优生", "4–6 周",
         "分期节点退费", "KOL/机构", "P0"],
        ["就业", "毕业生过渡实习包", 6800, 12800, 9800,
         "实习+内推辅导+复盘+证明", "包 Offer", "毕业生", "6–12 周",
         "入营评估后按阶段退", "社招渠道", "P0"],
        ["增值", "推荐信加购", 1500, 1500, 1500,
         "中/英推荐信（仅项目学员）", "不可单独购买", "在营/结业学员", "随项目",
         "启动后不退", "项目内转化", "P1"],
        ["B端", "学校/教培团购", "标准价×0.6", "标准价×0.8", "按协议",
         "批量名额+统一课题交付", "开证明/造假包装", "学校/机构", "按学期",
         "合同约定", "商务BD", "P1"],
    ]
    for i, r in enumerate(price_rows, 6):
        write_row(ws, i, r, align="center")
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=i, column=7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 36

    section_banner(ws, 12, "二、已停售（勿再对外报价）", span=12, fill=RED)
    header_row(ws, 13, ["原套餐", "原标准价", "状态", "说明"], fill=RED)
    for i, r in enumerate([
        ["线上实习·证明版", 199, "已停售", "开证明低价档取消"],
        ["证明+申报支持", 699, "已停售", "开证明低价档取消"],
        ["证明+推荐信协助", 1680, "已停售", "开证明低价档取消"],
    ], 14):
        write_row(ws, i, r, align="center", fill=RED_BG)

    section_banner(ws, 18, "三、定价公式与机制", span=12, fill=ACCENT)
    header_row(ws, 19, ["机制", "说明", "参数/规则", "示例"])
    mechanisms = [
        ["成本加成", "学员价 ≥ 通道成本 + 运营分摊 + 目标毛利", "商业线目标毛利 34–42%", "通道 2400 → 学员价约 5,480"],
        ["三维锚点", "企业品牌 × 课题真实度 × 辅导强度", "每维 1–5 分加权", "上海AI头部+强导师 → 进阶价"],
        ["交付绑定", "结业证明随项目交付，不单卖", "禁止开证明产品线", "证明=履约结果"],
        ["支付节奏", "定金 30% + 企业确认 40% + 结业 30%", "与交付物绑定", "降低纠纷与坏账"],
        ["折扣纪律", "早鸟/团购折扣 ≤ 20%", "需审批", "保护品牌与单价"],
        ["动态调价", "旺季（寒暑假）上浮 10–15%", "淡季赠增值不降标价", "暑假通道营可上浮"],
    ]
    for i, r in enumerate(mechanisms, 20):
        write_row(ws, i, r)
        ws.row_dimensions[i].height = 30

    section_banner(ws, 27, "四、单人成本拆解（商业标准套餐示意，标准价 5,480）", span=12, fill=PRIMARY)
    header_row(ws, 28, ["成本项", "金额(元)", "占比", "备注"])
    costs = [
        ("企业通道/课题合作费", 1800),
        ("导师辅导与批改", 600),
        ("运营与客服分摊", 450),
        ("获客 CAC", 900),
        ("平台与合规", 250),
        ("其他/预备费", 200),
    ]
    for i, (name, amt) in enumerate(costs, 29):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=amt)
        ws.cell(row=i, column=3, value=f"=B{i}/5480")
        ws.cell(row=i, column=3).number_format = "0.0%"
        ws.cell(row=i, column=4, value="")
        for col in range(1, 5):
            c = ws.cell(row=i, column=col)
            c.font = Font(name="微软雅黑", size=10, color=DARK)
            c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
            c.border = thin()
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=LIGHT)
    ws.cell(row=35, column=1, value="成本合计").font = Font(name="微软雅黑", bold=True, color=WHITE)
    ws.cell(row=35, column=2, value="=SUM(B29:B34)")
    ws.cell(row=35, column=3, value="=B35/5480")
    ws.cell(row=35, column=3).number_format = "0.0%"
    ws.cell(row=35, column=4, value="毛利 ≈ 标准价 − 成本合计")
    for col in range(1, 5):
        c = ws.cell(row=35, column=col)
        c.fill = PatternFill("solid", fgColor=PRIMARY)
        c.font = Font(name="微软雅黑", bold=True, color=WHITE)
        c.border = thin()
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=37, column=1, value="单人毛利(元)")
    ws.cell(row=37, column=2, value="=5480-B35")
    ws.cell(row=38, column=1, value="单人毛利率")
    ws.cell(row=38, column=2, value="=B37/5480")
    ws.cell(row=38, column=2).number_format = "0.0%"
    for r in (37, 38):
        for col in (1, 2):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="微软雅黑", bold=True, color=DARK)
            c.border = thin()
            c.fill = PatternFill("solid", fgColor=SOFT)

    set_widths(ws, [12, 24, 14, 14, 16, 32, 18, 14, 10, 22, 12, 10])
    ws.freeze_panes = "A4"


def build_finance(wb):
    ws = wb.create_sheet("06 盈利测算")
    big_title(ws, "J", "盈利模型与财务测算",
              "可调参数区（改黄色单元格）→ 自动汇总三情景｜已取消开证明，仅高价值三档")

    section_banner(ws, 4, "一、可调假设参数", span=10, fill=PRIMARY)
    header_row(ws, 5, ["参数", "保守", "基准", "乐观", "单位", "说明"])
    params = [
        ["付费总人数", 100, 200, 300, "人", "年度或滚动 12 个月"],
        ["商业标准人数占比", 0.55, 0.50, 0.45, "%", "科技/具身项目"],
        ["商业进阶人数占比", 0.25, 0.30, 0.35, "%", "上海AI/头部通道"],
        ["就业包人数占比", 0.20, 0.20, 0.20, "%", "毕业生"],
        ["商业标准客单", 5200, 5480, 6200, "元", ""],
        ["商业进阶客单", 10800, 11800, 13000, "元", ""],
        ["就业包客单", 9000, 9800, 11000, "元", ""],
        ["综合毛利率", 0.35, 0.38, 0.42, "%", "收入结构加权后"],
        ["固定成本", 280000, 320000, 380000, "元/年", "人员+平台+合规+办公"],
        ["获客占收入比", 0.20, 0.18, 0.15, "%", "渠道与内容投放"],
    ]
    for i, r in enumerate(params, 6):
        for col, v in enumerate(r, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = Font(name="微软雅黑", size=10, color=DARK)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin()
            if col in (2, 3, 4) and i <= 15:
                c.fill = PatternFill("solid", fgColor=YELLOW)
            if col in (2, 3, 4) and i in (7, 8, 9, 13, 15):
                c.number_format = "0%"
            if col == 1:
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.fill = PatternFill("solid", fgColor=LIGHT)

    section_banner(ws, 17, "二、三情景自动测算", span=10, fill=ACCENT)
    header_row(ws, 18, ["指标", "保守", "基准", "乐观", "公式说明"])
    labels = [
        ("商业标准人数", "=ROUND(B6*B7,0)", "=ROUND(C6*C7,0)", "=ROUND(D6*D7,0)", "总人数×标准占比"),
        ("商业进阶人数", "=ROUND(B6*B8,0)", "=ROUND(C6*C8,0)", "=ROUND(D6*D8,0)", "总人数×进阶占比"),
        ("就业包人数", "=ROUND(B6*B9,0)", "=ROUND(C6*C9,0)", "=ROUND(D6*D9,0)", "总人数×就业占比"),
        ("标准收入(元)", "=B19*B10", "=C19*C10", "=D19*D10", "人数×客单"),
        ("进阶收入(元)", "=B20*B11", "=C20*C11", "=D20*D11", ""),
        ("就业收入(元)", "=B21*B12", "=C21*C12", "=D21*D12", ""),
        ("总营收(元)", "=B22+B23+B24", "=C22+C23+C24", "=D22+D23+D24", ""),
        ("总营收(万元)", "=B25/10000", "=C25/10000", "=D25/10000", ""),
        ("毛利(元)", "=B25*B13", "=C25*C13", "=D25*D13", "营收×综合毛利率"),
        ("毛利(万元)", "=B27/10000", "=C27/10000", "=D27/10000", ""),
        ("固定成本(元)", "=B14", "=C14", "=D14", ""),
        ("经营盈余(元)", "=B27-B29", "=C27-C29", "=D27-D29", "毛利−固定成本"),
        ("经营盈余(万元)", "=B30/10000", "=C30/10000", "=D30/10000", ""),
        ("是否盈亏平衡", '=IF(B30>=0,"是","否")', '=IF(C30>=0,"是","否")', '=IF(D30>=0,"是","否")', ""),
    ]
    for i, (label, b, c, d, note) in enumerate(labels, 19):
        ws.cell(row=i, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True, color=DARK)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=3, value=c)
        ws.cell(row=i, column=4, value=d)
        ws.cell(row=i, column=5, value=note)
        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            cell.border = thin()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="微软雅黑", size=10, color=DARK, bold=(col == 1))
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
        if i in (22, 23, 24, 25, 27, 29, 30):
            for col in (2, 3, 4):
                ws.cell(row=i, column=col).number_format = "#,##0"
        if i in (26, 28, 31):
            for col in (2, 3, 4):
                ws.cell(row=i, column=col).number_format = "0.0"

    # highlight surplus
    for col in range(1, 5):
        ws.cell(row=30, column=col).fill = PatternFill("solid", fgColor=SOFT)
        ws.cell(row=31, column=col).fill = PatternFill("solid", fgColor=SOFT)

    section_banner(ws, 34, "三、敏感因素与管理动作", span=10, fill=PRIMARY)
    header_row(ws, 35, ["敏感因子", "影响", "监控指标", "应对动作"])
    sens = [
        ["大厂通道成本上升", "进阶套餐毛利下降", "通道成本/学员价", "备选科技公司池；调整进阶定价"],
        ["获客 CAC 过高", "盈余被侵蚀", "CAC、付费转化率", "加强 B 端团购与转介绍激励"],
        ["标准→进阶升级不足", "收入结构偏弱", "升级转化率", "结业前 7 日内升级优惠（≤15%）"],
        ["退费率升高", "口碑与利润双杀", "退费率、NPS", "分期交付+服务协议+补学机制"],
        ["企业履约不稳", "交付失败", "履约率、替补启用次数", "每赛道 ≥2 家备份"],
        ["重回低价开证明", "风险回潮", "是否出现违规报价", "价目表审计；违者下架"],
    ]
    for i, r in enumerate(sens, 36):
        write_row(ws, i, r)
        ws.row_dimensions[i].height = 28

    set_widths(ws, [18, 14, 14, 14, 28, 12, 12, 12, 12, 12])
    ws.freeze_panes = "A4"


def build_plan(wb):
    ws = wb.create_sheet("07 落地实施计划")
    big_title(ws, "J", "6 个月试点落地实施计划", "里程碑 · 任务拆解 · 状态跟踪")

    section_banner(ws, 4, "里程碑与任务", span=10, fill=PRIMARY)
    header_row(ws, 5, [
        "月份", "阶段", "任务", "交付物", "负责人",
        "开始", "结束", "状态", "依赖", "备注",
    ])
    tasks = [
        ["M1", "基建", "高价值三档产品封装与话术", "产品说明书（无开证明）", "", "", "", "未开始", "", ""],
        ["M1", "基建", "下架开证明相关页面/报价", "停售确认清单", "", "", "", "进行中", "", "199/699/1680"],
        ["M1", "基建", "报名表单/收款/协议", "电子协议+表单", "", "", "", "未开始", "", ""],
        ["M2", "供给", "签约上海AI/具身 5–8 家", "合作意向/协议", "", "", "", "未开始", "M1", ""],
        ["M2", "供给", "组建导师池 ≥ 10 人", "导师名单", "", "", "", "未开始", "", ""],
        ["M2", "供给", "确认 AI 头部通道 1–2 条", "通道说明", "", "", "", "未开始", "", ""],
        ["M3", "获客", "高校社团/机构渠道试点", "渠道清单", "", "", "", "未开始", "M1", ""],
        ["M3", "获客", "商业标准档首批开营", "30 人开营", "", "", "", "未开始", "M2", ""],
        ["M4", "放量", "进阶通道营开营", "案例与口碑", "", "", "", "未开始", "M3", ""],
        ["M4", "放量", "标准→进阶升级追踪", "升级率报表", "", "", "", "未开始", "M3", ""],
        ["M5", "转化", "毕业生就业包上线", "产品页+辅导流程", "", "", "", "未开始", "M2", ""],
        ["M5", "转化", "Offer/升学反馈收集", "案例 ≥10", "", "", "", "未开始", "M4", ""],
        ["M6", "复盘", "财务与 NPS 复盘", "复盘报告", "", "", "", "未开始", "全部", ""],
        ["M6", "复盘", "是否扩张决策会", "Go/No-Go 纪要", "", "", "", "未开始", "复盘报告", ""],
    ]
    for i, r in enumerate(tasks, 6):
        write_row(ws, i, r, align="center")
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    add_status_validation(ws, "H", 6, 5 + len(tasks))

    set_widths(ws, [8, 10, 28, 18, 10, 12, 12, 10, 12, 18])
    ws.freeze_panes = "A6"


def build_risk(wb):
    ws = wb.create_sheet("08 风险与合规")
    big_title(ws, "I", "风险登记与合规要点", "先守住可信度与合规，再谈规模化")

    section_banner(ws, 4, "风险登记册", span=9, fill=PRIMARY)
    header_row(ws, 5, [
        "编号", "风险", "类别", "可能性", "影响",
        "风险等级", "应对措施", "责任人", "状态",
    ])
    risks = [
        ["R01", "出具不实实习证明引发投诉/舆情", "合规", "中", "高", "高",
         "真实任务+编号核验+抽查+拒绝虚假需求", "", "未开始"],
        ["R02", "企业合作中断导致无法交付", "供给", "中", "高", "高",
         "每赛道≥2备份；合同约定替补方案", "", "未开始"],
        ["R03", "价格战与「水项目」冲击", "市场", "高", "中", "高",
         "强化可核验与成果物；案例与 NPS 公开", "", "未开始"],
        ["R04", "退费与消费纠纷", "运营", "中", "中", "中",
         "分期交付、书面协议、补学/延期机制", "", "未开始"],
        ["R05", "未成年人信息与监护人同意缺失", "合规", "中", "高", "高",
         "隐私告知书+监护人同意；最小必要采集", "", "未开始"],
        ["R06", "推荐信/文书虚假陈述", "合规", "中", "高", "高",
         "仅基于真实表现；审核清单；拒绝包装造假", "", "未开始"],
        ["R07", "品牌方名称使用不当", "法务", "中", "高", "高",
         "对外话术经法务审核；避免虚假「官方实习」表述", "", "未开始"],
        ["R08", "获客成本失控", "财务", "中", "中", "中",
         "周度 CAC 看板；达标渠道加码，未达标停投", "", "未开始"],
    ]
    for i, r in enumerate(risks, 6):
        write_row(ws, i, r, align="center")
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=i, column=7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 36
        if r[5] == "高":
            ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor=RED_BG)
    add_status_validation(ws, "I", 6, 13)

    section_banner(ws, 15, "对外话术红线（摘要）", span=9, fill=ORANGE)
    lines = [
        "不得承诺「包录取 / 包 Offer / 包进某公司正式编制」。",
        "涉及知名企业时，须准确表述合作形态（联合课题/合作方项目/导师来自生态等），禁止虚假「官方实习生」宣传。",
        "证明内容必须与学员真实完成的任务一致，保留过程材料备查。",
        "推荐信仅描述可验证事实与导师真实评价。",
    ]
    for i, t in enumerate(lines, 16):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
        c = ws.cell(row=i, column=1, value=f"• {t}")
        c.font = Font(name="微软雅黑", size=10, color=DARK)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.fill = PatternFill("solid", fgColor=LIGHT)
        ws.row_dimensions[i].height = 28

    set_widths(ws, [8, 32, 10, 10, 10, 10, 36, 10, 10])
    ws.freeze_panes = "A6"


def build_kpi(wb):
    ws = wb.create_sheet("09 KPI跟踪")
    big_title(ws, "H", "试点期 KPI 跟踪表", "填写实际值后自动计算完成率与状态")

    section_banner(ws, 4, "核心指标", span=8, fill=PRIMARY)
    header_row(ws, 5, ["维度", "指标", "目标值", "实际值", "完成率", "状态", "更新频率", "备注"])
    kpis = [
        ("供给", "签约合作企业（家）", 15),
        ("供给", "可用导师数（人）", 10),
        ("销售", "付费学员（人）", 200),
        ("销售", "商业标准学员（人）", 100),
        ("销售", "标准→进阶升级率", 0.25),
        ("交付", "企业课题履约率", 0.95),
        ("交付", "课程完成率", 0.90),
        ("财务", "综合毛利率", 0.35),
        ("财务", "退费率（低于为佳）", 0.08),
        ("合规", "开证明违规报价次数", 0),
        ("口碑", "NPS", 40),
        ("结果", "就业/升学正向反馈率", 0.60),
        ("结果", "可复用成功案例数", 15),
    ]
    for i, (dim, kpi, target) in enumerate(kpis, 6):
        ws.cell(row=i, column=1, value=dim)
        ws.cell(row=i, column=2, value=kpi)
        ws.cell(row=i, column=3, value=target)
        ws.cell(row=i, column=4, value=0)
        if target == 0:
            # 违规次数目标为 0：实际为 0 即达成
            ws.cell(row=i, column=5, value='=IF(D{0}=0,1,0)'.format(i))
            ws.cell(row=i, column=6,
                    value=f'=IF(D{i}=0,"达成","风险")')
        else:
            ws.cell(row=i, column=5, value=f"=IFERROR(D{i}/C{i},0)")
            ws.cell(row=i, column=6,
                    value=f'=IF(D{i}=0,"未开始",IF(E{i}>=1,"达成",IF(E{i}>=0.7,"进行中","风险")))')
        ws.cell(row=i, column=5).number_format = "0.0%"
        ws.cell(row=i, column=7, value="双周" if dim in ("销售", "财务", "合规") else "月")
        ws.cell(row=i, column=8,
                value="退费率目标为上限，实际应低于目标" if "退费" in kpi
                else ("目标为 0，出现即违规" if "开证明" in kpi else ""))
        for col in range(1, 9):
            c = ws.cell(row=i, column=col)
            c.font = Font(name="微软雅黑", size=10, color=DARK)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin()
            if col == 4:
                c.fill = PatternFill("solid", fgColor=YELLOW)
        if isinstance(target, float) and target <= 1 and target > 0:
            ws.cell(row=i, column=3).number_format = "0%"
            ws.cell(row=i, column=4).number_format = "0%"

    rng = "F6:F18"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"达成"'],
                   fill=PatternFill("solid", fgColor=GREEN_BG),
                   font=Font(color="155724", bold=True)),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"风险"'],
                   fill=PatternFill("solid", fgColor=RED_BG),
                   font=Font(color="721C24", bold=True)),
    )

    section_banner(ws, 20, "使用说明", span=8, fill=ACCENT)
    notes = [
        "黄色单元格为实际值录入区；完成率与状态自动计算。",
        "「退费率」目标 8% 为上限：实际填写时应尽量低于 8%；状态公式仍按比值计算，需人工结合解读。",
        "「开证明违规报价次数」目标为 0：出现即记为风险，用于防止低价档回流。",
        "建议每两周同步 Dashboard 与本表，作为可行性复盘输入。",
    ]
    for i, t in enumerate(notes, 21):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        c = ws.cell(row=i, column=1, value=f"• {t}")
        c.font = Font(name="微软雅黑", size=10, color=DARK)
        c.fill = PatternFill("solid", fgColor=LIGHT)

    set_widths(ws, [10, 28, 12, 12, 10, 10, 10, 36])
    ws.freeze_panes = "A6"


def main():
    wb = Workbook()
    build_dashboard(wb)
    build_audience(wb)
    build_fortune(wb)
    build_tech(wb)
    build_strategy(wb)
    build_pricing(wb)
    build_finance(wb)
    build_plan(wb)
    build_risk(wb)
    build_kpi(wb)

    out = Path(__file__).resolve().parents[1] / "exports"
    out.mkdir(parents=True, exist_ok=True)
    path = out / "学生实习赋能计划_可行性论证数据表.xlsx"
    wb.save(str(path))
    print(f"已生成: {path}")
    return path


if __name__ == "__main__":
    main()
