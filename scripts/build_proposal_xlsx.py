#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成给同浦汇的业务承接执行台账 Excel。"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
import proposal_data as D

NAVY = "0E2240"
GOLD = "C9A227"
CREAM = "FAF6EE"
GOLD_PALE = "F4EBD3"
WHITE = "FFFFFF"
INK = "243044"
TEAL = "1F6B5C"
RED = "8B2E2E"
GRAY = "6B7280"
LINE = "E7DDC6"

OUT = Path(__file__).resolve().parents[1] / "output" / "同浦汇_30场活动与科技企业服务中心筹备_执行台账.xlsx"

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
thick_gold = Border(
    left=Side(style="thin", color=GOLD),
    right=Side(style="thin", color=GOLD),
    top=Side(style="thin", color=GOLD),
    bottom=Side(style="thin", color=GOLD),
)


def font(size=11, bold=False, color=INK, name="微软雅黑"):
    return Font(name=name, size=size, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_header(ws, row, cols, fill_color=NAVY):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = font(10, True, WHITE)
        cell.fill = fill(fill_color)
        cell.alignment = align("center")
        cell.border = thin


def style_body(ws, r0, r1, cols):
    for r in range(r0, r1 + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.font = font(10)
            cell.alignment = align("left" if c > 1 else "center")
            cell.border = thin
            if r % 2 == 0:
                if not (cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb not in (None, "00000000", "0E2240")):
                    cell.fill = fill(GOLD_PALE if r % 2 == 0 else WHITE)


def widths(ws, mapping):
    for k, v in mapping.items():
        ws.column_dimensions[k].width = v


def freeze_title(ws, title, subtitle, cols, tab_color=GOLD):
    ws.sheet_properties.tabColor = tab_color
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c1 = ws.cell(1, 1, title)
    c1.font = font(16, True, WHITE)
    c1.fill = fill(NAVY)
    c1.alignment = align("left")
    c2 = ws.cell(2, 1, subtitle)
    c2.font = font(10, False, NAVY)
    c2.fill = fill(GOLD_PALE)
    c2.alignment = align("left")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:3"
    ws.sheet_view.showGridLines = False
    ws.oddFooter.left.text = D.CONFIDENTIAL
    ws.oddFooter.right.text = "第 &P 页"


def build():
    wb = Workbook()

    # —— 0 封面 ——
    ws = wb.active
    ws.title = "00-封面说明"
    ws.sheet_properties.tabColor = GOLD
    ws.sheet_view.showGridLines = False
    widths(ws, {c: 22 for c in "ABCDEFGHI"})
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 55
    ws.merge_cells("A1:C1")
    ws["A1"] = D.DOC_TITLE + "　·　" + D.DOC_SUBTITLE
    ws["A1"].font = font(18, True, WHITE)
    ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = align()
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:C2")
    ws["A2"] = f"提交 {D.DOC_FOR}　｜　{D.DOC_FROM}　｜　{D.DOC_DATE}　｜　周期 {D.DOC_PERIOD}"
    ws["A2"].font = font(11, False, NAVY)
    ws["A2"].fill = fill(GOLD_PALE)
    ws.row_dimensions[2].height = 22

    meta = [
        ("提交对象", D.DOC_FOR),
        ("提出方", D.DOC_FROM),
        ("联合", D.DOC_COFROM),
        ("学术支持", D.DOC_SUPPORT),
        ("同浦汇联系人", D.PARTIES["联系人"]),
        ("业主", D.PARTIES["业主"]),
        ("园区", D.PARTIES["园区"]),
        ("新赛道", D.NEW_POSITIONING),
        ("政策包装", D.POLICY_PACKAGING),
        ("工作包 A", "30 场活动全案执行（30 万 / 年，每场 ≤30 人）"),
        ("工作包 B", "科技企业服务中心 90 天筹备（筹备不另向同浦汇收费）"),
        ("活动内部分成", "服中心 70% / 同浦汇 30%（建议，待确认）"),
        ("政策分成", "同浦汇 38% / 服中心 62%（锁版）"),
        ("付款", D.COMMERCIAL["付款"]),
        ("机密", D.CONFIDENTIAL),
    ]
    ws["A4"] = "字段"
    ws["B4"] = "内容"
    ws.merge_cells("B4:C4")
    style_header(ws, 4, 3)
    for i, (k, v) in enumerate(meta):
        r = 5 + i
        ws.cell(r, 1, k).font = font(10, True, WHITE)
        ws.cell(r, 1).fill = fill(TEAL if i % 2 == 0 else NAVY)
        ws.cell(r, 1).alignment = align("center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(r, 2, v).font = font(10)
        ws.cell(r, 2).alignment = align()
        for c in range(1, 4):
            ws.cell(r, c).border = thin
        ws.row_dimensions[r].height = 22

    ws["A21"] = "工作表索引"
    ws["A21"].font = font(13, True, NAVY)
    index = [
        ("00-封面说明", "口径、主体、分成总览"),
        ("01-昨日共识", "8 月 31 日交流要点，作为共同事实"),
        ("02-分工矩阵", "同浦汇 / 服中心 / 园区 / 复旦 / 业主"),
        ("03-30场总表", "30 场排期、线条、状态（可勾选）"),
        ("04-单场标准", "人数、负责人、节奏、转化、加购"),
        ("05-服中心90天", "挂牌筹备甘特与前置条件"),
        ("06-商务付款", "30 万拆分到季度与内部结算"),
        ("07-政策分成", "上限测算与 38/62"),
        ("08-边界风险", "不承诺清单 + 审核三句口径"),
        ("09-下一步", "请同浦汇确认的五件事"),
    ]
    ws["A22"], ws["B22"], ws["C22"] = "工作表", "用途", "对应文件"
    style_header(ws, 22, 3)
    for i, (a, b) in enumerate(index):
        r = 23 + i
        ws.cell(r, 1, a).font = font(10, True, NAVY)
        ws.cell(r, 2, b)
        ws.cell(r, 3, "PPT 汇报 + Word 正文")
        for c in range(1, 4):
            ws.cell(r, c).font = font(10) if c > 1 else font(10, True, NAVY)
            ws.cell(r, c).alignment = align()
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill(CREAM if i % 2 else WHITE)

    # —— 1 昨日共识 ——
    ws = wb.create_sheet("01-昨日共识")
    freeze_title(ws, "昨日共识（2026-08-31）", "创智汇赛道调整交流 · 约 36 分钟 · 作为双方共同事实，不另做预测承诺", 3)
    widths(ws, {"A": 8, "B": 22, "C": 88})
    ws["A3"], ws["B3"], ws["C3"] = "序号", "类型", "要点"
    style_header(ws, 3, 3)
    tags = ["赛道", "赛道", "目标", "风险", "口径", "资源", "出海", "政策", "空间", "领馆", "节奏"]
    for i, t in enumerate(D.YESTERDAY):
        r = 4 + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2, tags[i] if i < len(tags) else "其他")
        ws.cell(r, 3, t)
        for c in range(1, 4):
            ws.cell(r, c).font = font(10, c == 2, NAVY if c < 3 else INK)
            ws.cell(r, c).alignment = align("center" if c < 3 else "left")
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill(GOLD_PALE if i % 2 == 0 else WHITE)
        ws.row_dimensions[r].height = 36
    r = 16
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(r, 1, f"定位切换：{D.OLD_POSITIONING}  →  {D.NEW_POSITIONING}　｜　政策包装：{D.POLICY_PACKAGING}")
    ws.cell(r, 1).font = font(12, True, WHITE)
    ws.cell(r, 1).fill = fill(TEAL)
    ws.cell(r, 1).alignment = align()
    ws.row_dimensions[r].height = 28

    # —— 2 分工 ——
    ws = wb.create_sheet("02-分工矩阵")
    freeze_title(ws, "分工矩阵", "同浦汇管关系，服中心管交付；园区管销售促成", 4, TEAL)
    widths(ws, {"A": 22, "B": 70, "C": 22, "D": 22})
    ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "主体", "责任", "主责/协同", "是否承接范围内"
    style_header(ws, 3, 4, TEAL)
    flags = [
        ("同浦汇", "主责客户关系", "接口保留，不替换"),
        ("科技企业服务中心", "主责执行", "是，工作包 A+B"),
        ("科技企业联合会", "协同", "联合协办"),
        ("复旦住房政策研究中心", "按场次确认", "学术支持，不替代接口"),
        ("创智汇 / 园区", "主责销售", "不纳入承接承诺"),
        ("杨浦科创集团", "业主确认", "口径与场地条件"),
    ]
    for i, ((who, what), (w2, flag, note)) in enumerate(zip(D.ROLES, flags)):
        r = 4 + i
        ws.cell(r, 1, who)
        ws.cell(r, 2, what)
        ws.cell(r, 3, flag)
        ws.cell(r, 4, note)
        for c in range(1, 5):
            ws.cell(r, c).font = font(10, c == 1, NAVY if c == 1 else INK)
            ws.cell(r, c).alignment = align()
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill(WHITE if i % 2 else GOLD_PALE)
        ws.row_dimensions[r].height = 32

    # —— 3 三十场 ——
    ws = wb.create_sheet("03-30场总表")
    freeze_title(
        ws,
        "30 场活动执行总表（计入年包）",
        "每场 ≤30 人　｜　负责人 ≤30%　｜　状态可下拉　｜　另计价项目见「04-单场标准」",
        14,
    )
    headers = [
        "序号", "编号", "月份", "线条", "活动名称", "本场作用",
        "建议周次", "人数上限", "负责人占比上限", "主责", "同浦汇协同", "状态", "线索条数", "备注",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, len(headers))
    widths(ws, {
        "A": 6, "B": 8, "C": 12, "D": 12, "E": 36, "F": 38,
        "G": 12, "H": 10, "I": 14, "J": 16, "K": 16, "L": 12, "M": 10, "N": 18,
    })
    theme_fill = {
        "启动": "1B3A6B",
        "智能建造": "1F6B5C",
        "政策": "6B4F1D",
        "出海准备": "0E2240",
        "展示": "5C3D8F",
        "收官": "8B2E2E",
    }
    dv = DataValidation(type="list", formula1='"未启动,策划中,名单确认,已执行,月报已交,延期,取消"', allow_blank=True)
    ws.add_data_validation(dv)
    for i, (code, month, theme, name, why) in enumerate(D.EVENTS):
        r = 4 + i
        vals = [
            i + 1, code, month, theme, name, why,
            "", 30, "≤30%", "服中心执行", "带客/回访", "未启动", "", "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = font(10)
            cell.alignment = align("center" if c in (1, 2, 3, 7, 8, 9, 12, 13) else "left")
            cell.border = thin
            cell.fill = fill(WHITE if i % 2 else CREAM)
        # theme color on 线条
        tf = theme_fill.get(theme, NAVY)
        ws.cell(r, 4).fill = fill(tf)
        ws.cell(r, 4).font = font(10, True, WHITE)
        dv.add(ws.cell(r, 12))
        ws.row_dimensions[r].height = 28
    # summary
    r = 35
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    ws.cell(r, 1, "场次合计：30　｜　启动 5　智能建造 7　政策 4　出海准备 6　展示 5　收官 3　｜　另计价不计入本表")
    ws.cell(r, 1).font = font(11, True, WHITE)
    ws.cell(r, 1).fill = fill(NAVY)
    # count by theme
    ws["A37"] = "线条"
    ws["B37"] = "场次"
    style_header(ws, 37, 2)
    for i, (name, n, when, what) in enumerate(D.THEMES):
        ws.cell(38 + i, 1, name)
        ws.cell(38 + i, 2, n)
        ws.cell(38 + i, 1).font = font(10, True, NAVY)
        ws.cell(38 + i, 2).font = font(10, True, GOLD)
        ws.cell(38 + i, 2).alignment = align("center")
        for c in range(1, 3):
            ws.cell(38 + i, c).border = thin
            ws.cell(38 + i, c).fill = fill(GOLD_PALE)
    chart = BarChart()
    chart.type = "col"
    chart.title = "30 场结构"
    chart.y_axis.title = "场次"
    data = Reference(ws, min_col=2, min_row=37, max_row=43)
    cats = Reference(ws, min_col=1, min_row=38, max_row=43)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.legend = None
    chart.style = 10
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "D37")

    # —— 4 单场标准 ——
    ws = wb.create_sheet("04-单场标准")
    freeze_title(ws, "单场标准 / 转化 / 加购", D.EVENT_STANDARD["节奏"], 3)
    widths(ws, {"A": 22, "B": 55, "C": 40})
    ws["A3"], ws["B3"], ws["C3"] = "项目", "口径", "说明"
    style_header(ws, 3, 3)
    std_rows = [
        ("人数", D.EVENT_STANDARD["人数"], "签到表核验"),
        ("负责人占比", D.EVENT_STANDARD["负责人占比"], "最后锁版，覆盖早期 50% 口径"),
        ("全年触达", D.EVENT_STANDARD["触达"], "30×约 20 人量级，不写千人场"),
        ("执行节奏", D.EVENT_STANDARD["节奏"], "档期灵活，不是一个月一场"),
        ("转化闭环", D.EVENT_STANDARD["转化"], "活动不对租金去化对赌"),
        ("交付物", D.EVENT_STANDARD["交付"], "回访只交摘要"),
    ]
    for i, row in enumerate(std_rows):
        r = 4 + i
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v).font = font(10, c == 1, NAVY if c == 1 else INK)
            ws.cell(r, c).alignment = align()
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill(WHITE if i % 2 else GOLD_PALE)
        ws.row_dimensions[r].height = 32
    ws["A11"] = "另计价（不进 30 万包）"
    ws["A11"].font = font(13, True, WHITE)
    ws.merge_cells("A11:C11")
    ws["A11"].fill = fill(RED)
    ws["A12"], ws["B12"], ws["C12"] = "项目", "内容", "计价原则"
    style_header(ws, 12, 3, RED)
    for i, row in enumerate(D.EXTRA_PRICE):
        r = 13 + i
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v).font = font(10)
            ws.cell(r, c).alignment = align()
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill("F8E8E8")
        ws.row_dimensions[r].height = 28

    # —— 5 90天 ——
    ws = wb.create_sheet("05-服中心90天")
    freeze_title(ws, "科技企业服务中心筹备 · 90 天", "挂牌仪式另计价　｜　筹备不另向同浦汇收费", 8, TEAL)
    headers = ["阶段", "窗口", "主题", "交付", "W1-2", "W3-6", "W7-10", "W11-12"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 8, TEAL)
    widths(ws, {"A": 8, "B": 12, "C": 16, "D": 48, "E": 10, "F": 10, "G": 10, "H": 10})
    gantt = [
        ("1", *D.CENTER_90[0], "●", "", "", ""),
        ("2", *D.CENTER_90[1], "", "●", "", ""),
        ("3", *D.CENTER_90[2], "", "", "●", ""),
        ("4", *D.CENTER_90[3], "", "", "", "●"),
    ]
    for i, row in enumerate(gantt):
        r = 4 + i
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = font(10, c >= 5, WHITE if c >= 5 and v else INK)
            cell.alignment = align("center" if c != 4 else "left")
            cell.border = thin
            if c >= 5 and v == "●":
                cell.fill = fill(GOLD)
            else:
                cell.fill = fill(WHITE if i % 2 else CREAM)
        ws.row_dimensions[r].height = 40
    ws["A9"] = "前置条件（政策）"
    ws.merge_cells("A9:H9")
    ws["A9"].font = font(12, True, WHITE)
    ws["A9"].fill = fill(NAVY)
    ws.merge_cells("A10:H10")
    ws["A10"] = D.POLICY_CAP["前置条件"]
    ws["A10"].alignment = align()
    ws["A10"].font = font(11)
    ws["A10"].fill = fill(GOLD_PALE)
    ws.row_dimensions[10].height = 28
    ws.merge_cells("A12:H12")
    ws["A12"] = (
        f"空间：{D.SPACE['项目']} {D.SPACE['面积']}　｜　3F {D.SPACE['3F']}　｜　5F {D.SPACE['5F']}　｜　"
        f"专题 {D.SPACE['展厅专题']}　｜　办公 {D.SPACE['办公租金']}　物业 {D.SPACE['物业']}"
    )
    ws["A12"].font = font(10, True, NAVY)
    ws["A12"].alignment = align()
    ws.row_dimensions[12].height = 36

    # —— 6 商务 ——
    ws = wb.create_sheet("06-商务付款")
    freeze_title(ws, "商务付款与内部结算", "对园区一口价　｜　内部两套分成互不混用", 6)
    widths(ws, {"A": 22, "B": 18, "C": 16, "D": 16, "E": 16, "F": 28})
    ws["A3"] = "对园区 · 活动年包 30 万元拆分"
    ws.merge_cells("A3:F3")
    ws["A3"].font = font(12, True, WHITE)
    ws["A3"].fill = fill(NAVY)
    heads = ["节点", "比例", "金额（万元）", "累计（万元）", "挂钩条件", "备注"]
    for i, h in enumerate(heads, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, 6)
    pays = [
        ("签约后 7 日内 · 启动款", 0.50, "启动策划与前 8 场", "到账后 T+14 交执行手册"),
        ("2026 Q3", 0.10, "Q3 场次完成度 + 月报", "8–9 月为主"),
        ("2026 Q4", 0.10, "Q4 场次完成度 + 月报", "10–12 月"),
        ("2027 Q1", 0.10, "Q1 场次完成度 + 月报", "1–3 月"),
        ("2027 Q2", 0.10, "Q2 场次完成度 + 月报", "4–6 月；7 月收官计入年终"),
        ("年终收官尾款", 0.10, "30 场核验 + 年终月报", "完成后 15 日内"),
    ]
    acc = 0
    for i, (name, pct, cond, note) in enumerate(pays):
        r = 5 + i
        amt = round(30 * pct, 2)
        acc = round(acc + amt, 2)
        ws.cell(r, 1, name)
        ws.cell(r, 2, pct)
        ws.cell(r, 2).number_format = "0%"
        ws.cell(r, 3, amt)
        ws.cell(r, 4, acc)
        ws.cell(r, 5, cond)
        ws.cell(r, 6, note)
        for c in range(1, 7):
            ws.cell(r, c).font = font(10)
            ws.cell(r, c).alignment = align("center" if c in (2, 3, 4) else "left")
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill(WHITE if i % 2 else GOLD_PALE)
        ws.cell(r, 3).number_format = "0.00"
        ws.cell(r, 4).number_format = "0.00"
    ws.cell(11, 1, "合计")
    ws.cell(11, 2, 1)
    ws.cell(11, 2).number_format = "0%"
    ws.cell(11, 3, 30)
    ws.cell(11, 4, 30)
    for c in range(1, 7):
        ws.cell(11, c).font = font(11, True, WHITE)
        ws.cell(11, c).fill = fill(TEAL)
        ws.cell(11, c).alignment = align("center" if c > 1 else "left")
        ws.cell(11, c).border = thin

    ws["A13"] = "内部结算（建议，待同浦汇确认）"
    ws.merge_cells("A13:F13")
    ws["A13"].font = font(12, True, WHITE)
    ws["A13"].fill = fill(TEAL)
    for i, h in enumerate(["收入类型", "基数（万元）", "服中心", "同浦汇", "服中心金额", "同浦汇金额"], 1):
        ws.cell(14, i, h)
    style_header(ws, 14, 6, TEAL)
    # 活动 70/30
    ws["A15"] = "活动年包"
    ws["B15"] = 30
    ws["C15"] = 0.70
    ws["D15"] = 0.30
    ws["E15"] = "=B15*C15"
    ws["F15"] = "=B15*D15"
    ws["A16"] = "政策收益（到账后，示例不作承诺）"
    ws["B16"] = ""
    ws["C16"] = 0.62
    ws["D16"] = 0.38
    ws["E16"] = '=IF(B16="","待到账",B16*C16)'
    ws["F16"] = '=IF(B16="","待到账",B16*D16)'
    ws["A17"] = "服中心筹备"
    ws["B17"] = 0
    ws["C17"] = "不另收费"
    ws["D17"] = "—"
    ws["E17"] = 0
    ws["F17"] = 0
    for r in range(15, 18):
        for c in range(1, 7):
            ws.cell(r, c).font = font(10)
            ws.cell(r, c).alignment = align("center" if c > 1 else "left")
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill(CREAM)
        if r == 15:
            ws.cell(r, 2).number_format = "0.00"
            ws.cell(r, 3).number_format = "0%"
            ws.cell(r, 4).number_format = "0%"
            ws.cell(r, 5).number_format = "0.00"
            ws.cell(r, 6).number_format = "0.00"
        if r == 16:
            ws.cell(r, 3).number_format = "0%"
            ws.cell(r, 4).number_format = "0%"
    ws.merge_cells("A19:F19")
    ws["A19"] = (
        "招商佣金：2 个月净租金（首年不重复），由园区销售闭环触发，不计入上表。"
        "门票/赞助可冲抵活动成本，不替代 30 万打包价。"
    )
    ws["A19"].alignment = align()
    ws["A19"].font = font(10, False, NAVY)
    ws.row_dimensions[19].height = 32

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "活动年包付款节奏（万元）"
    data = Reference(ws, min_col=3, min_row=4, max_row=10)
    cats = Reference(ws, min_col=1, min_row=5, max_row=10)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.style = 10
    chart2.legend = None
    chart2.width = 18
    chart2.height = 8
    ws.add_chart(chart2, "A21")

    # —— 7 政策 ——
    ws = wb.create_sheet("07-政策分成")
    freeze_title(
        ws,
        "政策工具箱与分成（上限测算，非保证获批）",
        D.POLICY_CAP["前置条件"],
        6,
    )
    widths(ws, {"A": 28, "B": 16, "C": 16, "D": 16, "E": 18, "F": 22})
    ws["A3"], ws["B3"], ws["C3"], ws["D3"], ws["E3"], ws["F3"] = (
        "项目", "年上限（万元）", "十年（万元）", "同浦汇 38%", "服中心 62%", "备注"
    )
    style_header(ws, 3, 6)
    rows = [
        ("载体（平台 100+基地 10）", 110, 1100, "认定后才可申报"),
        ("活动 YOUNG立方封顶", 200, 2000, "投入×50%，年包 30 万大约对应约 15 万量级"),
        ("申报奖项（培育节奏测算）", "80→530", 3820, "高企/专精特新/小巨人等，按到账"),
        ("三部分合计上限", "390→840", 6920, "上限不是保底"),
    ]
    for i, (a, b, c, f) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, a)
        ws.cell(r, 2, b)
        ws.cell(r, 3, c)
        if isinstance(c, (int, float)):
            ws.cell(r, 4, round(c * 0.38, 1))
            ws.cell(r, 5, round(c * 0.62, 1))
        else:
            ws.cell(r, 4, "—")
            ws.cell(r, 5, "—")
        ws.cell(r, 6, f)
        for col in range(1, 7):
            ws.cell(r, col).font = font(10, i == 3, WHITE if i == 3 else INK)
            ws.cell(r, col).alignment = align("center" if col > 1 else "left")
            ws.cell(r, col).border = thin
            ws.cell(r, col).fill = fill(NAVY if i == 3 else (WHITE if i % 2 else GOLD_PALE))
        ws.row_dimensions[r].height = 28
    ws.merge_cells("A9:F9")
    ws["A9"] = "服务业引导资金 ≤300 万为项目制，未计入十年年表。成功费以资金到账为准，不向企业承诺必然获批。"
    ws["A9"].font = font(10, False, RED)
    ws["A9"].alignment = align()

    # —— 8 风险 ——
    ws = wb.create_sheet("08-边界风险")
    freeze_title(ws, "不承诺清单与审核口径", "转交审核方时建议原样保留第六章三句话", 2, RED)
    widths(ws, {"A": 12, "B": 100})
    ws["A3"], ws["B3"] = "类型", "条文"
    style_header(ws, 3, 2, RED)
    items = [("不承接 / 不承诺", t) for t in D.NOT_TAKEOVER] + [
        ("审核口径 1", "我们不是在招「付不起租金的施工队来租办公室」。"),
        ("审核口径 2", "我们是在用科技企业服务中心，把智能建造产品、模块化建筑和绿色低碳能力组织成可出海的集群。"),
        ("审核口径 3", "业主是杨浦科创集团；政府对接使用科技与产业服务口径，不以「中建」名义包装。"),
        ("KPI 场次", D.KPI[0][1]),
        ("KPI 人数", D.KPI[1][1]),
        ("KPI 月报", D.KPI[2][1]),
        ("KPI 服中心", D.KPI[3][1]),
        ("KPI 转化", D.KPI[4][1]),
    ]
    for i, (a, b) in enumerate(items):
        r = 4 + i
        ws.cell(r, 1, a)
        ws.cell(r, 2, b)
        ws.cell(r, 1).font = font(10, True, WHITE)
        ws.cell(r, 1).fill = fill(RED if "口径" in a else (TEAL if a.startswith("KPI") else NAVY))
        ws.cell(r, 1).alignment = align("center")
        ws.cell(r, 2).font = font(10)
        ws.cell(r, 2).alignment = align()
        ws.cell(r, 2).border = thin
        ws.cell(r, 1).border = thin
        ws.cell(r, 2).fill = fill(WHITE if i % 2 else CREAM)
        ws.row_dimensions[r].height = 28

    # —— 9 下一步 ——
    ws = wb.create_sheet("09-下一步")
    freeze_title(ws, "请同浦汇确认", "五件事确认后，启动款到账 14 日内交付细化执行手册", 4)
    widths(ws, {"A": 8, "B": 70, "C": 16, "D": 28})
    ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "序号", "确认事项", "状态", "同浦汇意见"
    style_header(ws, 3, 4)
    dv2 = DataValidation(type="list", formula1='"待确认,同意,调整后同意,不同意"', allow_blank=True)
    ws.add_data_validation(dv2)
    for i, t in enumerate(D.NEXT_STEPS):
        r = 4 + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2, t)
        ws.cell(r, 3, "待确认")
        ws.cell(r, 4, "")
        for c in range(1, 5):
            ws.cell(r, c).font = font(11)
            ws.cell(r, c).alignment = align("center" if c != 2 else "left")
            ws.cell(r, c).border = thin
            ws.cell(r, c).fill = fill(GOLD_PALE)
        dv2.add(ws.cell(r, 3))
        ws.row_dimensions[r].height = 32
    ws.merge_cells("A10:D10")
    ws["A10"] = f"联系人：{D.PARTIES['联系人']}　｜　我方：{D.DOC_FROM}　｜　{D.DOC_DATE}"
    ws["A10"].font = font(11, True, NAVY)
    ws["A10"].alignment = align()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Excel 已写入 {OUT}　工作表 {wb.sheetnames}")
    return OUT


if __name__ == "__main__":
    build()
