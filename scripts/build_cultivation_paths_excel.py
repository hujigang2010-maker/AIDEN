#!/usr/bin/env python3
"""生成 Excel：《东方修行路径对照表》

Sheets：
  00 总览
  01 三教总对照
  02 佛家路径
  03 道家路径
  04 儒家路径
  05 维度深比
  06 入门路线图
  07 误区速查
  08 术语简释
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 青墨主题（墨色 + 青绿，避开常见紫/奶油陶土配色）
PRIMARY = "1B3A4B"  # 墨青
ACCENT = "2F6F5E"  # 青绿
LIGHT = "E8F0ED"  # 雾青底
MIST = "D5E5DE"
GOLD = "A67C3D"  # 沉金点缀
CINNABAR = "8B3A2F"  # 朱砂警示
GREY = "5A6A72"
WHITE = "FFFFFF"
DARK = "1A2A32"
BUDDHA = "3A5A7A"  # 佛家栏色
DAO = "2F6F5E"  # 道家栏色
RU = "7A5A3A"  # 儒家栏色


def thin():
    s = Side(style="thin", color="B8C8C0")
    return Border(left=s, right=s, top=s, bottom=s)


def header_row(ws, row, headers, fill=PRIMARY, color=WHITE, height=28):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="微软雅黑", size=11, bold=True, color=color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[row].height = height


def write_row(ws, row, data, fill=None, bold=False, size=10, wrap=True, align="left", color=DARK):
    for i, v in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name="微软雅黑", size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border = thin()
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        elif row % 2 == 0:
            c.fill = PatternFill("solid", fgColor=LIGHT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def section_banner(ws, row, text, span=10, fill=ACCENT, color=WHITE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="微软雅黑", size=12, bold=True, color=color)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).border = thin()
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[row].height = 26


def big_title(ws, span_col="L", text="", sub=""):
    ws.merge_cells(f"A1:{span_col}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(name="微软雅黑", size=18, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A2:{span_col}2")
    c2 = ws["A2"]
    c2.value = sub
    c2.font = Font(name="微软雅黑", size=10, italic=True, color=GREY)
    c2.fill = PatternFill("solid", fgColor=LIGHT)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22


def build_overview(wb):
    ws = wb.active
    ws.title = "00 总览"
    big_title(
        ws,
        "H",
        "东方修行路径对照表 · 总览",
        "佛家（禅宗 / 净土 / 唯识）· 道家（内丹 / 符箓 / 导引）· 儒家（修身 / 慎独）｜对照维度：目标 · 方法 · 入门门槛 · 典型误区",
    )

    section_banner(ws, 4, "一、文档用途与使用方式", span=8, fill=PRIMARY)
    tips = [
        ["用途", "为初学者与跨传统对照者提供结构化地图，非宗教劝信，亦非替代师承指导。"],
        ["读法", "先看「01 三教总对照」建立大局；再按兴趣进入 02–04 分路径；用 06–07 做入门与避坑。"],
        ["原则", "尊重各传统原典语境；对照是为理解差异与互补，而非强行统一或优劣排名。"],
        ["边界", "涉及身心修炼者请量力而行；有身心疾患者应咨询专业医疗，勿以修炼替代治疗。"],
    ]
    for i, (k, v) in enumerate(tips, 5):
        ws.cell(row=i, column=1, value=k).font = Font(name="微软雅黑", bold=True, color=DARK)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=MIST)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=1).border = thin()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        c = ws.cell(row=i, column=2, value=v)
        c.font = Font(name="微软雅黑", color=DARK, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin()
        ws.row_dimensions[i].height = 32

    section_banner(ws, 10, "二、路径速览（点击对应工作表深入）", span=8, fill=ACCENT)
    header_row(ws, 11, ["传统", "路径", "一句话目标", "核心方法关键词", "入门难度", "典型误区关键词", "工作表", "色标"])
    rows = [
        ["佛家", "禅宗", "见性成佛 / 直指本心", "坐禅、公案、话头、日常即道", "中高（需师资与耐力）", "口头禅、求玄、厌弃日常", "02 佛家路径", "墨蓝"],
        ["佛家", "净土", "信愿持名、往生净土", "念佛、发愿、回向、福慧双修", "低（门槛最低之一）", "只求感应、废止伦理", "02 佛家路径", "墨蓝"],
        ["佛家", "唯识", "转识成智、了达心识", "经论研习、观心、止观配教理", "高（教理门槛高）", "空谈名相、知解障", "02 佛家路径", "墨蓝"],
        ["道家", "内丹", "性命双修、炼精化气", "筑基、周天、性命功法", "高（需明师与次第）", "盲练周天、求神通", "03 道家路径", "青绿"],
        ["道家", "符箓", "通神济世、正一科仪", "斋醮、符咒、科仪、积德", "中（需门派传承）", "迷信符力、功利祈请", "03 道家路径", "青绿"],
        ["道家", "导引", "形气神和、延年保健", "导引术、吐纳、站桩、动功", "低中（可渐进自学入门）", "逞强过量、迷信特效", "03 道家路径", "青绿"],
        ["儒家", "修身", "成德成人、齐家治国", "格物、诚意、克己复礼", "低中（生活即道场）", "外求功名、道德表演", "04 儒家路径", "沉棕"],
        ["儒家", "慎独", "独处亦正、意念自清", "省察、戒惧、日省其身", "中（难在持续）", "苛责自我、形式化反省", "04 儒家路径", "沉棕"],
    ]
    for i, r in enumerate(rows, 12):
        fill = LIGHT if i % 2 == 0 else WHITE
        write_row(ws, i, r, fill=fill, size=10)
        ws.row_dimensions[i].height = 36

    section_banner(ws, 21, "三、对照维度说明", span=8, fill=PRIMARY)
    dims = [
        ["目标", "该路径所指向的最终或阶段性成就（解脱、合道、成德等）。"],
        ["方法", "日常可操作的核心修习手段与次第结构。"],
        ["入门门槛", "经典阅读、师承、时间投入、身心条件等起步要求。"],
        ["典型误区", "初学与进阶阶段最常见的理解偏差与实践陷阱。"],
    ]
    header_row(ws, 22, ["维度", "含义"])
    for i, r in enumerate(dims, 23):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 28
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

    section_banner(ws, 28, "四、版本信息", span=8, fill=GOLD)
    meta = [
        ["版本", "V1.0"],
        ["定位", "知识对照 / 学习地图 / 讨论底稿"],
        ["说明", "内容基于公开经典与通识梳理，各派内部亦有多支传承，表中取代表性概括。"],
    ]
    for i, (k, v) in enumerate(meta, 29):
        ws.cell(row=i, column=1, value=k).font = Font(name="微软雅黑", bold=True, color=DARK)
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=MIST)
        ws.cell(row=i, column=1).border = thin()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        c = ws.cell(row=i, column=2, value=v)
        c.font = Font(name="微软雅黑", color=DARK)
        c.border = thin()
        c.alignment = Alignment(wrap_text=True, vertical="center")

    set_widths(ws, [12, 12, 28, 32, 22, 26, 14, 10])


def build_three_traditions(wb):
    ws = wb.create_sheet("01 三教总对照")
    big_title(
        ws,
        "F",
        "三教总对照：佛 · 道 · 儒",
        "从问题意识、终极指向到日常落点的宏观对照",
    )
    header_row(
        ws,
        4,
        ["对照项", "佛家", "道家", "儒家", "互补提示", "备注"],
    )
    rows = [
        [
            "核心问题",
            "苦从何来？如何离苦得乐、了生死？",
            "人如何与道合真、全生保真？",
            "人如何成德、安身立命于人伦？",
            "问题不同，答案不可硬比高下",
            "先辨问题再选路径",
        ],
        [
            "终极目标",
            "涅槃解脱 / 成佛（诸宗表述不同）",
            "得道合真 / 长生久视（义理与功夫并重）",
            "成圣成贤 / 内圣外王",
            "解脱、合道、成德可并行理解",
            "终极语汇不可简单等同",
        ],
        [
            "人性/心性观",
            "缘起无我；心性本净或如来藏诸说",
            "道法自然；性命双修、形神并重",
            "性善可学；气质之性可变化",
            "都重视「心」的转化",
            "术语体系差异大",
        ],
        [
            "主要方法气质",
            "戒定慧、止观、信愿行",
            "虚静、吐纳、丹道、科仪",
            "格致诚正、克己、礼乐",
            "静定与人伦实践可互补",
            "方法可借鉴，体系勿混炖",
        ],
        [
            "与社会关系",
            "出离心与菩萨道（入世度生）并存",
            "隐逸全真与济世度人并存",
            "家国天下、责任伦理为核心",
            "儒重入世，佛道可出可入",
            "现代人常需入世场景落地",
        ],
        [
            "经典入口",
            "《心经》《金刚经》《阿弥陀经》《成唯识论》等",
            "《道德经》《庄子》《黄庭》《悟真篇》等",
            "《论语》《大学》《中庸》《孟子》",
            "各选一部精读胜过泛览",
            "配合可靠注释",
        ],
        [
            "入门门槛总评",
            "净土偏低；禅中高；唯识偏高",
            "导引偏低；符箓需传承；内丹偏高",
            "修身门槛低；慎独难在坚持",
            "可按精力与机缘组合",
            "见 06 入门路线图",
        ],
        [
            "典型共同误区",
            "求神通、废伦理、口头开悟",
            "盲练、迷信符力、逆生理硬练",
            "道德表演、苛责、外求功名",
            "三教皆忌「自欺」",
            "见 07 误区速查",
        ],
    ]
    for i, r in enumerate(rows, 5):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, size=10)
        ws.row_dimensions[i].height = 48
    set_widths(ws, [14, 32, 32, 28, 26, 18])


def build_buddhism(wb):
    ws = wb.create_sheet("02 佛家路径")
    big_title(
        ws,
        "H",
        "佛家路径对照：禅宗 · 净土 · 唯识",
        "目标 · 方法 · 入门门槛 · 典型误区（代表性概括，不覆盖一切支派）",
    )
    header_row(
        ws,
        4,
        ["路径", "目标", "核心方法", "日常功课示例", "入门门槛", "适合倾向", "典型误区", "纠偏提示"],
        fill=BUDDHA,
    )
    rows = [
        [
            "禅宗",
            "直指人心、见性成佛；在日用中彻见本来",
            "坐禅、参话头/公案、默照或看话；保任于行住坐卧",
            "每日静坐 20–40 分钟；提起一则话头；做事不忘觉照",
            "中高：需明眼师承或可靠团体；对「放下知见」有耐心",
            "喜简捷、耐寂寞、能在生活中用功者",
            "①口头禅、以机锋代实修 ②厌弃伦理与责任 ③求玄求异象当开悟",
            "以戒为基；开悟须保任；日常责任即道场",
        ],
        [
            "净土",
            "信愿持名，求生极乐；现世安心，临终有靠",
            "执持名号、发菩提心、读诵经典、诸善回向",
            "定课念佛（可计数）；早晚课；行住坐卧默持；回向众生",
            "低：方法简明，老幼皆可；关键在信愿真切与持续",
            "事务繁忙、求稳妥、重视他力加持者",
            "①只求感应神迹 ②废止做人做事 ③轻视因果与戒律",
            "信愿行为三资粮；念佛与伦理福德并进",
        ],
        [
            "唯识",
            "了达心识结构，转识成智；以教理支撑止观",
            "研习唯识名相与种子熏习；配合止观观察心所起灭",
            "系统读论（如《三十颂》入门）；日记心行；止观练习",
            "高：名相繁密，需长期经论基础与辅导",
            "思辨强、喜体系化、愿下苦功读经论者",
            "①知解障：懂名相却不观心 ②炫博轻修 ③以理论否定他宗",
            "教观双运；懂一点就用一点；对治傲慢",
        ],
    ]
    for i, r in enumerate(rows, 5):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, size=10)
        ws.row_dimensions[i].height = 90

    section_banner(ws, 9, "佛家共通基础（三路径皆建议具备）", span=8, fill=BUDDHA)
    header_row(ws, 10, ["基础项", "要点", "为什么重要", "最低可行做法", "", "", "", ""], fill=ACCENT)
    commons = [
        ["戒律/伦理", "不伤害、诚实、节制、责任", "定慧大厦的地基", "先守五戒精神于生活"],
        ["因果正见", "业由心造、可转可净", "避免投机与自欺", "做事负责，日省己过"],
        ["出离心与慈悲", "不耽生死轮回；利他愿力", "方向端正，不被神通诱拐", "每周做一件利他实事"],
        ["善知识", "可依止的人/团体/清净传承", "减少盲修与极端", "多方了解再决定亲近"],
    ]
    for i, r in enumerate(commons, 11):
        write_row(ws, i, r + ["", "", "", ""], fill=LIGHT if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 36
        ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=8)

    set_widths(ws, [10, 28, 32, 30, 26, 22, 32, 24])


def build_daoism(wb):
    ws = wb.create_sheet("03 道家路径")
    big_title(
        ws,
        "H",
        "道家路径对照：内丹 · 符箓 · 导引",
        "目标 · 方法 · 入门门槛 · 典型误区（代表性概括）",
    )
    header_row(
        ws,
        4,
        ["路径", "目标", "核心方法", "日常功课示例", "入门门槛", "适合倾向", "典型误区", "纠偏提示"],
        fill=DAO,
    )
    rows = [
        [
            "内丹",
            "性命双修：炼精化气、炼气化神、炼神还虚，合道成真",
            "筑基清虚、调息守窍、周天运转、性功（神）与命功（气）次第",
            "在明师指导下的静功/动功；清心寡欲；作息与食饮有节",
            "高：须明师口传心授；身心条件与德行要求高；忌急求速成",
            "有长期志向、能守戒、愿循次第者",
            "①无师盲练周天致伤 ②以气感当成就 ③求长生神通而废德行",
            "先德行筑基；有病先医；不秘传者不强求",
        ],
        [
            "符箓",
            "通神达真、济世度人；正一科仪中安立身心与社群秩序",
            "授箓传承、斋醮科仪、符咒法术、积功累德、经典持诵",
            "依门派功课持诵；参与科仪学习；日常积德改过",
            "中：多需门派传承与合法合规场域；礼仪与经典基础",
            "重视仪式、愿服务社群、能守规戒者",
            "①迷信符力可代道德 ②功利祈请、恐吓他人 ③非法敛财",
            "以德为本；符为助缘；守法守戒",
        ],
        [
            "导引",
            "形正气畅、神清体健；为修道或养生奠基",
            "导引舒筋、吐纳调息、站桩、八段锦/五禽戏等动功",
            "每日 15–30 分钟导引或站桩；自然呼吸；松静为本",
            "低中：可循可靠教材/教师渐进；对年龄友好",
            "希望从身体入手、改善作息与气力者",
            "①逞强过量、屏气硬顶 ②迷信七日开天眼等夸张宣传 ③以出汗为效",
            "松柔渐进；不适即停；效果看长期稳定",
        ],
    ]
    for i, r in enumerate(rows, 5):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, size=10)
        ws.row_dimensions[i].height = 90

    section_banner(ws, 9, "道家共通提醒", span=8, fill=DAO)
    notes = [
        ["自然与节度", "道法自然不是放纵，而是合于节律；过犹不及。"],
        ["性命与伦理", "丹道、符法皆以积德清心为基；无德而求术，易入歧途。"],
        ["身体安全", "胸闷、头晕、心悸等异常应停止并就医，勿硬扛为「考验」。"],
        ["传承辨别", "警惕高价速成、神秘恐吓、要求断绝家庭与医疗者。"],
    ]
    header_row(ws, 10, ["主题", "说明", "", "", "", "", "", ""], fill=ACCENT)
    for i, (k, v) in enumerate(notes, 11):
        write_row(ws, i, [k, v, "", "", "", "", "", ""], fill=LIGHT if i % 2 == 0 else WHITE)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        ws.row_dimensions[i].height = 30

    set_widths(ws, [10, 28, 32, 30, 26, 22, 32, 24])


def build_confucianism(wb):
    ws = wb.create_sheet("04 儒家路径")
    big_title(
        ws,
        "H",
        "儒家路径对照：修身 · 慎独",
        "目标 · 方法 · 入门门槛 · 典型误区（与佛道并置时的「日用人伦」路径）",
    )
    header_row(
        ws,
        4,
        ["路径", "目标", "核心方法", "日常功课示例", "入门门槛", "适合倾向", "典型误区", "纠偏提示"],
        fill=RU,
    )
    rows = [
        [
            "修身",
            "变化气质、成德成人；推己及人，齐家处世",
            "格物致知、诚意正心、克己复礼、居敬穷理；在人伦事务中历练",
            "读《论语》《大学》精要；一事当前先问义利；日行一善且改一过",
            "低中：无需出家或秘传；难在把道理落在角色责任里",
            "在家庭/职场中求安身立命、重责任伦理者",
            "①道德表演与指责他人 ②外求功名当修身 ③空谈心性废事功",
            "先责己；修身在事上磨；知行合一",
        ],
        [
            "慎独",
            "独处暗室亦能戒慎恐惧；意念未发之际保持清明正直",
            "省察克治、戒惧慎微、日记心过、勿自欺；表里如一",
            "每晚三问：今日欺心否？对人亏欠否？明日一念如何立？",
            "中：方法简单，难在无人监督时的真实与持续",
            "敏感于内心动机、愿深耕诚意者",
            "①苛责成焦虑抑郁 ②形式化写反省而无改过 ③以慎独逃避群体责任",
            "温柔而坚定；省察为了改过，不是自我折磨",
        ],
    ]
    for i, r in enumerate(rows, 5):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, size=10)
        ws.row_dimensions[i].height = 95

    section_banner(ws, 8, "儒家与佛道的对照位置", span=8, fill=RU)
    header_row(ws, 9, ["议题", "儒家侧重", "可与佛道互补处", "", "", "", "", ""], fill=ACCENT)
    comps = [
        ["场域", "家国、职场、朋友人伦是主道场", "佛道的静定可滋养诚意；勿用出世逃避责任"],
        ["动力", "耻感、责任、志于道", "慈悲与虚静可软化苛严与紧绷"],
        ["成就观", "成德与事功相即", "解脱/合道视野可扩大生命格局，避免功利窄化"],
        ["风险", "沦为教条或精致利己", "以慎独防自欺；以实事检验学问"],
    ]
    for i, r in enumerate(comps, 10):
        write_row(ws, i, r + ["", "", "", ""], fill=LIGHT if i % 2 == 0 else WHITE)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)
        ws.row_dimensions[i].height = 36

    set_widths(ws, [10, 28, 32, 30, 26, 22, 32, 24])


def build_dimension_deep(wb):
    ws = wb.create_sheet("05 维度深比")
    big_title(
        ws,
        "I",
        "四维深比：目标 · 方法 · 门槛 · 误区",
        "将八条路径放在同一维度下横比，便于快速筛选",
    )

    section_banner(ws, 4, "A. 目标对照", span=9, fill=PRIMARY)
    header_row(ws, 5, ["路径", "近程目标", "远程目标", "可验证的日常信号", "不可急求的信号"])
    goals = [
        ["禅宗", "心念较稳、反应减少被动", "见性并保任", "做事清楚、少抱怨", "神通、异象"],
        ["净土", "心有所归、焦虑降低", "往生净土、圆满菩提", "定课稳定、性情柔和", "立即感应、梦兆攀比"],
        ["唯识", "能辨心所、少被情绪拖走", "转识成智", "观心记录变细", "口头名相娴熟"],
        ["内丹", "精气神较充、欲望有度", "还虚合道", "睡眠饮食改善、心静", "急速周天、发热发麻当成就"],
        ["符箓", "礼仪安顿、服务他人", "通真济世", "功课恭谨、言行收敛", "符到病除式承诺"],
        ["导引", "柔顺有力、呼吸顺畅", "形神俱妙之基", "体态松、少疼痛", "七日开慧等噱头"],
        ["修身", "一事上能克己", "成德推及家国", "关系改善、责任扛起", "名誉与点赞"],
        ["慎独", "独处少自欺", "意念精纯、表里如一", "反省能改过", "严厉自责本身"],
    ]
    for i, r in enumerate(goals, 6):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 32

    section_banner(ws, 15, "B. 方法对照（可操作性）", span=9, fill=ACCENT)
    header_row(ws, 16, ["路径", "主修", "辅修", "是否强依赖师承", "每日最低剂量（示意）"])
    methods = [
        ["禅宗", "坐禅 / 话头", "经典、礼仪、劳动", "强烈建议", "静坐 20 分 + 日用觉照"],
        ["净土", "持名念佛", "读经、放生/义工、回向", "可依经典自学起步，有善友更好", "定课数百至数千声（量力）"],
        ["唯识", "经论 + 观心", "止观、辩论/请益", "建议有人辅导", "读 30 分 + 观心日记"],
        ["内丹", "次第功法", "德行、导引、经典", "必须", "仅按师嘱；无师不自创"],
        ["符箓", "科仪与符法", "持诵、积德", "必须（门派）", "依门派功课"],
        ["导引", "动功 / 站桩 / 吐纳", "作息、食饮", "教师更佳，可稳健自学入门", "15–30 分松柔练习"],
        ["修身", "事上磨练", "读经、礼、友朋切磋", "不必秘传，需益友", "一事克己 + 读经一章"],
        ["慎独", "省察戒惧", "日记、静坐片刻", "不必秘传", "晚间三问 5–10 分"],
    ]
    for i, r in enumerate(methods, 17):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 30

    section_banner(ws, 26, "C. 入门门槛评分（示意，1低–5高）", span=9, fill=PRIMARY)
    header_row(ws, 27, ["路径", "经典难度", "师承依赖", "时间稳定要求", "身心风险", "综合入门门槛", "说明"])
    barriers = [
        ["禅宗", 3, 4, 4, 2, 4, "不难懂口号，难在真参实修"],
        ["净土", 2, 2, 3, 1, 2, "方法易，信愿持续是关键"],
        ["唯识", 5, 3, 4, 1, 5, "教理门槛最高"],
        ["内丹", 4, 5, 5, 4, 5, "无师高风险"],
        ["符箓", 3, 5, 3, 2, 4, "传承与合规是关键"],
        ["导引", 2, 2, 3, 2, 2, "过量是主要风险"],
        ["修身", 2, 1, 3, 1, 2, "难在落地而非理解"],
        ["慎独", 2, 1, 4, 2, 3, "易苛责，需心理弹性"],
    ]
    for i, r in enumerate(barriers, 28):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, align="center")
        ws.cell(row=i, column=7).alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws.row_dimensions[i].height = 28

    section_banner(ws, 37, "D. 误区热度（常见程度示意）", span=9, fill=CINNABAR)
    header_row(ws, 38, ["路径", "误区1", "误区2", "误区3", "高危信号（应停/应远离）"])
    pitfalls = [
        ["禅宗", "口头禅", "以狂放当解脱", "轻视因果", "鼓励违法、毁家庭、否定医疗"],
        ["净土", "感应攀比", "废事废人", "只要佛不要戒", "恐吓「不念即堕」敛财"],
        ["唯识", "名相游戏", "贬低实修宗派", "知而不行", "以理论攻击取代修行"],
        ["内丹", "网络自练周天", "气感崇拜", "秘传高价速成", "胸闷心悸仍强练；禁医"],
        ["符箓", "符力迷信", "恐吓驱邪营销", "无德求术", "违法科仪、高价消灾"],
        ["导引", "逞强屏气", "特效广告", "疼痛当进步", "剧烈眩晕、关节损伤"],
        ["修身", "指责他人", "道德表演", "读书不做事", "以儒学名义压迫他人"],
        ["慎独", "自我苛责", "反省形式化", "离群索居", "持续抑郁焦虑不求助"],
    ]
    for i, r in enumerate(pitfalls, 39):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 32

    set_widths(ws, [10, 22, 22, 22, 28, 18, 28, 12, 12])


def build_onboarding(wb):
    ws = wb.create_sheet("06 入门路线图")
    big_title(
        ws,
        "G",
        "入门路线图：按情境选择起步",
        "先选「问题与生活方式」，再选路径；可组合，但勿同时深钻多条高门槛路线",
    )
    section_banner(ws, 4, "一、情境 → 建议起步", span=7, fill=PRIMARY)
    header_row(ws, 5, ["你的情境", "更贴合的问题", "建议主路径", "可搭配", "30天起步包", "慎入", "原因"])
    scenarios = [
        [
            "工作家庭忙、心不安",
            "如何安心、有依靠",
            "净土（持名）",
            "儒家修身（责任不废）",
            "每日定课念佛 + 一件负责之事",
            "同时猛攻禅与丹道",
            "精力分散，易半途",
        ],
        [
            "喜欢坐下来安静",
            "如何看清念头",
            "禅宗（稳健坐禅）",
            "导引放松身体",
            "坐禅20分/日 + 身体松柔",
            "无师参险怪公案",
            "易落玄谈或压念",
        ],
        [
            "爱读书、喜体系",
            "心如何运作",
            "唯识入门",
            "短时止观",
            "《三十颂》导读 + 观心日记",
            "只攒术语不观心",
            "知解障",
        ],
        [
            "身体紧、睡眠差",
            "如何形气调和",
            "导引 / 站桩",
            "儒家作息与节欲",
            "导引20分 + 固定睡眠",
            "高强度闭气功法",
            "伤身",
        ],
        [
            "有明师机缘、长线志向",
            "性命如何双修",
            "内丹（依师）",
            "导引导入、德行筑基",
            "仅做师嘱筑基内容",
            "网课速成周天",
            "高风险",
        ],
        [
            "重视仪式与服务",
            "如何在敬畏中安顿",
            "符箓门派正规学习",
            "积德、持诵",
            "了解传承背景与规戒",
            "商业驱邪套路",
            "伦理与法律风险",
        ],
        [
            "想在关系里成长",
            "如何成人成德",
            "修身",
            "慎独",
            "《大学》纲领 + 事上克己",
            "只批判社会不改己",
            "变成道德表演",
        ],
        [
            "易自欺、表里不一",
            "如何在独处中诚实",
            "慎独",
            "短暂静坐",
            "晚间三问日记",
            "无限上纲自责",
            "伤及心理健康",
        ],
    ]
    for i, r in enumerate(scenarios, 6):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, size=10)
        ws.row_dimensions[i].height = 42

    section_banner(ws, 15, "二、组合原则（现代人常用）", span=7, fill=ACCENT)
    rules = [
        ["一主一辅", "同时只深耕一条主路径，另一条作辅助滋养。"],
        ["身→心→理", "身体躁动时先导引/作息；心乱再定课；理路后置深化。"],
        ["入世不废", "有家庭职场责任者，优先能嵌入日程的净土/修身/导引。"],
        ["高门槛单列", "内丹、符箓、深度禅堂用功，单独评估师承与身心条件。"],
        ["季度复盘", "每 90 天问：更平静？更负责？更少自欺？无效则调整而非加码玄术。"],
    ]
    header_row(ws, 16, ["原则", "说明", "", "", "", "", ""], fill=GOLD)
    for i, (k, v) in enumerate(rules, 17):
        write_row(ws, i, [k, v, "", "", "", "", ""], fill=LIGHT if i % 2 == 0 else WHITE)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        ws.row_dimensions[i].height = 28

    set_widths(ws, [18, 18, 18, 20, 32, 18, 16])


def build_pitfalls(wb):
    ws = wb.create_sheet("07 误区速查")
    big_title(
        ws,
        "F",
        "误区速查与纠偏",
        "按「症状」反查，便于自检与提醒他人",
    )
    header_row(ws, 4, ["症状/念头", "可能落入的误区", "相关路径", "纠偏动作", "是否建议外求帮助", "备注"])
    rows = [
        ["我觉得已经开悟，别人都愚痴", "慢心 / 口头禅", "禅宗、唯识", "回到戒行与服务；请益善知识验证", "是（师友）", "真见地含谦卑"],
        ["不念佛/不练功就恐慌被罚", "恐吓信仰", "净土、符箓、内丹", "区分经典正说与营销话术；恢复日常伦理与责任", "是", "远离恐吓型群体"],
        ["胸闷心悸还坚持练", "逆生理硬练", "内丹、导引", "立即停止；就医", "是（医疗）", "健康优先"],
        ["花大价钱买速成密法", "秘传消费主义", "内丹、符箓、禅", "冷静对比公开正统资源；不因焦虑付费", "视情况", "正法不靠高价恐吓"],
        ["念佛后什么都不做了", "废事废人", "净土", "恢复家庭/工作责任；回向与做事并行", "可与善友谈", "信愿不废人伦"],
        ["只会讲名相，心念依旧", "知解障", "唯识、禅", "减少输入，增加观心与改过", "可", "教观双运"],
        ["用道德标准攻击家人", "修身变武器", "儒家", "先责己三天再开口；练习倾听", "可（家庭沟通）", "成德非胜人"],
        ["反省日记越写越自我厌恶", "苛责式慎独", "慎独", "改「我有罪」为「下一步改一寸」；严重则心理咨询", "是（必要时医疗）", "温柔的诚实"],
        ["追求光、震动、出体", "神通导向", "禅、丹、导引", "放下追求；回归方法本身与德行", "师承场景下请益", "境界不作炫耀"],
        ["因为修行与亲友彻底决裂（非必要）", "极端隔离", "各路径", "评估是否被控制；重建健康边界", "是", "警惕封闭操控"],
    ]
    for i, r in enumerate(rows, 5):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, size=10)
        ws.row_dimensions[i].height = 40
    set_widths(ws, [28, 16, 14, 36, 18, 16])


def build_glossary(wb):
    ws = wb.create_sheet("08 术语简释")
    big_title(
        ws,
        "D",
        "术语简释",
        "便于跨传统阅读时减少望文生义（极简通识义，非学术定论）",
    )
    header_row(ws, 4, ["术语", "所属", "极简义", "易混点"])
    terms = [
        ["见性", "禅宗", "彻见自心本性（各家诠说有别）", "不等于性格脾气或情绪宣泄"],
        ["话头", "禅宗", "提持一则疑问用心参究", "不是口头反复念句子玩概念"],
        ["信愿行", "净土", "深信、切愿、实行持名等", "缺一则易偏感应或空谈"],
        ["回向", "佛家", "把功德定向给众生/菩提", "不是交易式许愿清单"],
        ["转识成智", "唯识", "改造杂染心识为智慧", "不是智商提升术"],
        ["种子熏习", "唯识", "心识潜能被行为反复强化", "可用于理解习惯养成"],
        ["性命双修", "内丹", "性（神）与命（气身）同修", "偏一不可称双修"],
        ["周天", "内丹", "精气沿任督等路线运转的功法概念", "不可按网络图自行硬通"],
        ["筑基", "内丹/道", "清虚、节欲、培补的基础阶段", "基础未稳勿求高功"],
        ["符箓", "道教", "符图与授箓传承相关法务", "不等于任意画符"],
        ["斋醮", "道教", "斋戒与祭祀科仪", "重恭敬与规范"],
        ["导引", "道家养生", "以姿势与呼吸引导气血", "近于今日气功养生，体系多样"],
        ["吐纳", "道家", "呼吸吐故纳新", "忌勉强憋气"],
        ["克己复礼", "儒家", "约束私欲以回复礼之中节", "不是压抑一切情感"],
        ["慎独", "儒家", "独处时亦谨慎不自欺", "不是自我监控到病态"],
        ["格物致知", "儒家", "在事物上穷理以明善", "朱子/阳明诠释不同"],
        ["内圣外王", "儒家", "内成德、外善治的理想", "不是鼓励权力欲"],
        ["出离心", "佛家", "厌离轮回苦、求解脱的动机", "不是仇视人间责任"],
        ["保任", "禅宗", "悟后在日用中护持", "开悟不是终点"],
        ["积德", "道/儒共通", "改过迁善、利人利物", "万术之基"],
    ]
    for i, r in enumerate(terms, 5):
        write_row(ws, i, r, fill=LIGHT if i % 2 == 0 else WHITE, size=10)
        ws.row_dimensions[i].height = 28
    set_widths(ws, [14, 12, 40, 28])


def main():
    wb = Workbook()
    build_overview(wb)
    build_three_traditions(wb)
    build_buddhism(wb)
    build_daoism(wb)
    build_confucianism(wb)
    build_dimension_deep(wb)
    build_onboarding(wb)
    build_pitfalls(wb)
    build_glossary(wb)

    out_dir = Path(__file__).resolve().parents[1] / "deliverables"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "东方修行路径对照表.xlsx"
    wb.save(out)
    print(f"已生成: {out}")


if __name__ == "__main__":
    main()
