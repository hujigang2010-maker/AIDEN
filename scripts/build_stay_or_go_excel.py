#!/usr/bin/env python3
"""生成《发展与去留》决策矩阵 Excel。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference


INK = "0F172A"
TEAL = "0F766E"
AMBER = "D97706"
ROSE = "BE1231"
SOFT = "F1F5F9"
MINT = "ECFDF5"
CREAM = "FFF7ED"
WHITE = "FFFFFF"
SLATE = "475569"

thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def font(size=11, bold=False, color=INK, name="微软雅黑"):
    return Font(name=name, size=size, bold=bold, color=color)


def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_header(ws, row, cols, bg=INK):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill(bg)
        cell.font = font(11, True, WHITE)
        cell.alignment = align("center")
        cell.border = thin


def style_range(ws, r1, r2, c1, c2, bg=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.alignment = align()
            cell.font = font()
            if bg:
                cell.fill = fill(bg)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(output_path: Path) -> None:
    wb = Workbook()

    # ---- 封面 ----
    cover = wb.active
    cover.title = "使用说明"
    cover["A1"] = "发展与去留 · 2026 中美判断备忘录"
    cover["A1"].font = font(20, True, INK)
    cover.merge_cells("A1:F1")
    cover["A2"] = "个人决策表 · 默认深耕，出海做成期权 · 不构成移民/投资/法律建议"
    cover["A2"].font = font(12, False, TEAL)
    cover.merge_cells("A2:F2")
    cover.row_dimensions[1].height = 28
    cover.row_dimensions[2].height = 20

    notes = [
        ("工作表", "用途"),
        ("1. 使用说明", "读完再填。分数只帮助显形，不代替判断。"),
        ("2. 能力溢价对照", "先看你的能力放在哪边更值钱。"),
        ("3. 适配打分", "1–5 分自评。预填示例为「本地有网络、海外无书面机会」，自动汇总象限。"),
        ("4. 走的触发清单", "硬条件 / 事业 / 心理。未同时点亮，不到果断的时候。"),
        ("5. 90天行动", "可勾选。把判断做成行为。"),
        ("6. 对那段话", "原句与改写对照。收藏修辞，不让修辞代替战略。"),
    ]
    for r, row in enumerate(notes, start=4):
        cover.cell(r, 1, row[0])
        cover.cell(r, 2, row[1])
        cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    style_header(cover, 4, 2)
    style_range(cover, 5, 10, 1, 2, SOFT)
    for r in range(5, 11):
        cover.cell(r, 1).font = font(11, True, TEAL)
        cover.row_dimensions[r].height = 22

    cover["A12"] = "填表纪律"
    cover["A12"].font = font(14, True, INK)
    rules = [
        "先填「能力溢价对照」和「适配打分」，再看触发清单。不要先被那段话带着走。",
        "打分用「过去 12 个月已经发生的事实」，不用「我希望明年发生的故事」。",
        "「去美国」三件套缺一不可：书面机会、签证路径、三年财务缓冲。缺一件，就还在路径 A 或 B。",
        "宏观不适（觉得大环境差）不等于个人错配（你在这里已经做不成想做的事）。",
        "本表可每月复盘一次。若连续三个月「留」的主航道仍无订单，再上调「走」的权重。",
    ]
    for i, t in enumerate(rules, start=13):
        cover.cell(i, 1, f"{i - 12}. {t}")
        cover.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        cover.cell(i, 1).font = font(11)
        cover.cell(i, 1).alignment = align()
        cover.row_dimensions[i].height = 24
    set_widths(cover, [22, 28, 18, 18, 18, 18])
    cover.row_dimensions[13].height = 32
    cover.row_dimensions[14].height = 32
    cover.row_dimensions[15].height = 32
    cover.row_dimensions[16].height = 32
    cover.row_dimensions[17].height = 32
    cover.freeze_panes = "A5"
    cover.print_title_rows = "1:2"

    # ---- 能力溢价 ----
    cap = wb.create_sheet("能力溢价对照")
    cap["A1"] = "能力结构对照：溢价在哪边，迁移成本有多高"
    cap["A1"].font = font(16, True, INK)
    cap.merge_cells("A1:E1")
    headers = ["能力类型", "在中国的溢价", "在美国的溢价", "迁移成本", "若这是你的主能力，默认建议"]
    for i, h in enumerate(headers, start=1):
        cap.cell(3, i, h)
    style_header(cap, 3, 5)
    rows = [
        ["Frontier 研究 / 训练", "中（少数实验室）", "极高", "中（有论文与推荐时）", "有锁定再走；无锁定则不必为高地迁徙"],
        ["大厂工程 / 算法岗", "高，但内卷", "高，但签证卡脖子", "高", "有书面 offer + 签证再评估，否则留"],
        ["产业场景转化 / 落地", "很高（政府+制造+园区）", "中低（要重做信用）", "极高", "默认深耕中国，出海只做项目驻留"],
        ["高校 / 园区 / 商会网络", "核心资产", "接近归零", "极高", "默认留下；走 = 资产归零后再创业"],
        ["不动产 + 产业空间", "存量时代仍有专业位", "需牌照与本地信用", "极高", "转向存量 / 场景载体，而不是换国家"],
        ["英语内容 / 跨境合作", "稀缺加分", "入场券，不是护城河", "中", "作为双根结构的工具，不作为迁徙理由"],
    ]
    highlight = {3, 4, 5}
    for i, row in enumerate(rows, start=4):
        for c, val in enumerate(row, start=1):
            cap.cell(i, c, val)
        bg = MINT if i - 1 in highlight else SOFT
        style_range(cap, i, i, 1, 5, bg)
        cap.cell(i, 1).font = font(11, True)
        cap.row_dimensions[i].height = 36
    cap["A11"] = "使用：圈出你最强的 1–2 项。若落在绿色三行，2026 年默认路径是 A 深耕，而不是 C 迁徙。"
    cap["A11"].font = font(11, False, SLATE)
    cap.merge_cells("A11:E11")
    cap.row_dimensions[11].height = 28
    set_widths(cap, [28, 28, 28, 26, 42])
    cap.freeze_panes = "A4"
    cap.auto_filter.ref = "A3:E9"

    # ---- 适配打分 ----
    score = wb.create_sheet("适配打分")
    score["A1"] = "适配打分（1–5）。请只依据过去 12 个月的事实。"
    score["A1"].font = font(16, True, INK)
    score.merge_cells("A1:G1")
    score["A2"] = "1=完全不符合  3=中性  5=非常符合。预填为「本地有网络、海外无书面机会」的示例分，请改成你的事实分。"
    score["A2"].font = font(10, False, SLATE)
    score.merge_cells("A2:G2")

    sh = ["维度", "具体问题", "留/走", "权重", "你的打分(1-5)", "加权分", "填写提示"]
    for i, h in enumerate(sh, start=1):
        score.cell(4, i, h)
    style_header(score, 4, 7)

    items = [
        ("本地主航道", "过去 12 个月，你是否仍能把一件事推进到可报价 / 可试点？", "留", 1.2, "有真实线索打 4–5；只有会议打 2–3；完全停摆打 1"),
        ("网络可变现", "高校 / 园区 / 企业 / 资本关系，是否仍能约到关键人并形成下一步？", "留", 1.2, "能约到且有下一步=5；只能吃饭叙旧=2"),
        ("新动能重叠", "你的工作是否已经接到 AI / 先进制造 / 数智服务，而不是纯增量地产？", "留", 1.1, "已有项目=5；正在转=3；仍在旧剧本=1"),
        ("订单与现金流", "未来 12 个月，本地是否看得到养活自己的路径？", "留", 1.3, "已有合同=5；有高概率线索=3；完全不确定=1"),
        ("意义与心安", "在本地做事时，你是否仍能感到「这事值得我认真做」？", "留", 1.0, "这是心安指标。长期 1–2 分，留下也会空转"),
        ("书面机会", "是否已有美国（或海外）雇主 / 机构的书面机会？", "走", 1.4, "有正式书面=5；口头意向=3；完全没有=1"),
        ("签证路径", "签证路径是否清楚，且不依赖纯粹抽签？", "走", 1.3, "路径清楚=5；要赌 H-1B=2；没研究过=1"),
        ("能力可迁移", "你最强的能力，到美国是否仍能直接变现，而不是归零重来？", "走", 1.2, "论文/工程履历=4–5；本地网络型=1–2"),
        ("财务缓冲", "是否有 3 年财务缓冲，不把全部流动性押上？", "走", 1.1, "有 3 年=5；有 1 年=3；没有=1"),
        ("不可替代性", "去了是否能做这里做不了的事，而不是换个地方重复焦虑？", "走", 1.2, "能明确说出「只有那边能做」=5；说不清=1"),
    ]
    for i, (dim, q, side, w, hint) in enumerate(items, start=5):
        score.cell(i, 1, dim)
        score.cell(i, 2, q)
        score.cell(i, 3, side)
        score.cell(i, 4, w)
        score.cell(i, 5, 4 if side == "留" else 2)
        score.cell(i, 6, f"=D{i}*E{i}")
        score.cell(i, 7, hint)
        bg = MINT if side == "留" else CREAM
        style_range(score, i, i, 1, 7, bg)
        score.cell(i, 1).font = font(11, True)
        score.cell(i, 4).number_format = "0.0"
        score.cell(i, 5).number_format = "0"
        score.cell(i, 6).number_format = "0.00"
        score.row_dimensions[i].height = 36

    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=False)
    dv.error = "请输入 1 到 5 的整数"
    dv.errorTitle = "打分超范围"
    dv.prompt = "1–5"
    dv.promptTitle = "打分"
    score.add_data_validation(dv)
    dv.add("E5:E14")

    score.cell(16, 1, "留的加权总分")
    score.cell(16, 2, "=SUMIF(C5:C14,\"留\",F5:F14)")
    score.cell(16, 3, "满分 29.0")
    score.cell(17, 1, "走的加权总分")
    score.cell(17, 2, "=SUMIF(C5:C14,\"走\",F5:F14)")
    score.cell(17, 3, "满分 31.0")
    score.cell(18, 1, "差值（留 − 走）")
    score.cell(18, 2, "=B16-B17")
    score.cell(19, 1, "当前象限（自动）")
    score.cell(19, 2, '=IF(AND(B16>=18,B17<18),"深耕中国",IF(AND(B16>=18,B17>=18),"有限期出海",IF(AND(B16<18,B17>=18),"才考虑迁徙","先造航道")))')
    score.cell(20, 1, "一句话建议")
    score.cell(
        20,
        2,
        '=IF(B19="深耕中国","默认留下，90 天把一条主航道做成可报价项目。",IF(B19="有限期出海","可以去 12–36 个月，写返回条款，不斩根。",IF(B19="才考虑迁徙","走之前仍要补齐书面机会与签证路径。","先不要走，先把本地一件事做实。")))',
    )
    for r in range(16, 21):
        score.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        style_range(score, r, r, 1, 7, SOFT)
        score.cell(r, 1).font = font(11, True, TEAL)
        score.cell(r, 2).font = font(12, True, INK)
        score.row_dimensions[r].height = 26
    score.cell(16, 2).number_format = "0.00"
    score.cell(17, 2).number_format = "0.00"
    score.cell(18, 2).number_format = "0.00"
    score.row_dimensions[20].height = 36

    # 条件格式：象限
    score.conditional_formatting.add(
        "B19",
        FormulaRule(formula=['B19="深耕中国"'], fill=fill(MINT), font=font(12, True, TEAL)),
    )
    score.conditional_formatting.add(
        "B19",
        FormulaRule(formula=['B19="有限期出海"'], fill=fill(CREAM), font=font(12, True, AMBER)),
    )
    score.conditional_formatting.add(
        "B19",
        FormulaRule(formula=['B19="才考虑迁徙"'], fill=fill("FFE4E6"), font=font(12, True, ROSE)),
    )

    chart = BarChart()
    chart.type = "col"
    chart.title = "留 vs 走 加权总分"
    chart.y_axis.title = "加权分"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 35
    data = Reference(score, min_col=2, min_row=16, max_row=17)
    cats = Reference(score, min_col=1, min_row=16, max_row=17)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.shape = 4
    chart.legend = None
    chart.style = 10
    chart.width = 12
    chart.height = 6
    score.add_chart(chart, "A22")

    score["A38"] = "读法：预填示例会落在「深耕中国」。请务必改成你的事实分。差值 > 4 且象限为「深耕中国」，不要被诀别辞带走。"
    score["A38"].font = font(10, False, SLATE)
    score.merge_cells("A38:G38")
    set_widths(score, [16, 52, 10, 10, 16, 12, 44])
    score.freeze_panes = "A5"
    score.auto_filter.ref = "A4:G14"

    # ---- 触发清单 ----
    trig = wb.create_sheet("走的触发清单")
    trig["A1"] = "什么情况下才该走：未同时点亮，就还不到「果断」"
    trig["A1"].font = font(16, True, INK)
    trig.merge_cells("A1:D1")
    th = ["类别", "条件", "是否点亮", "证据（填写具体事实，禁止空话）"]
    for i, h in enumerate(th, start=1):
        trig.cell(3, i, h)
    style_header(trig, 3, 4, ROSE)
    conditions = [
        ("硬条件", "已有雇主或机构的书面机会"),
        ("硬条件", "签证路径清楚，不靠赌抽签"),
        ("硬条件", "3 年财务缓冲，不押上全部流动性"),
        ("事业条件", "去了能做这里做不了的事"),
        ("事业条件", "不是逃避失败，而是升级能力"),
        ("事业条件", "走之前本地资产有交接，而非蒸发"),
        ("心理条件", "能忍受身份不确定与可能的社会降级"),
        ("心理条件", "不需要靠「封存故乡」才能行动"),
        ("心理条件", "家人知情同意，不是一个人的诗"),
    ]
    for i, (cat, cond) in enumerate(conditions, start=4):
        trig.cell(i, 1, cat)
        trig.cell(i, 2, cond)
        trig.cell(i, 3, "未点亮")
        trig.cell(i, 4, "")
        bg = MINT if cat == "硬条件" else CREAM if cat == "事业条件" else SOFT
        style_range(trig, i, i, 1, 4, bg)
        trig.cell(i, 1).font = font(11, True)
        trig.row_dimensions[i].height = 28
    yn = DataValidation(type="list", formula1='"已点亮,未点亮"', allow_blank=False)
    trig.add_data_validation(yn)
    yn.add("C4:C12")

    trig.cell(14, 1, "已点亮条数")
    trig.cell(14, 2, '=COUNTIF(C4:C12,"已点亮")')
    trig.cell(15, 1, "硬条件是否齐")
    trig.cell(15, 2, '=IF(COUNTIF(C4:C6,"已点亮")=3,"齐","不齐")')
    trig.cell(16, 1, "是否进入「果断」")
    trig.cell(16, 2, '=IF(AND(COUNTIF(C4:C6,"已点亮")=3,COUNTIF(C7:C9,"已点亮")=3,COUNTIF(C10:C12,"已点亮")>=2),"可以认真谈走","还不到果断的时候")')
    trig.cell(17, 1, "建议动作")
    trig.cell(
        17,
        2,
        '=IF(B16="可以认真谈走","进入路径 C：谈 offer、签证、返回条款，仍不接受「不要回来」作为人生协议。","留在路径 A 或 B。缺的硬条件用 90 天去补证据，而不是补情绪。")',
    )
    for r in range(14, 18):
        trig.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        style_range(trig, r, r, 1, 4, SOFT)
        trig.cell(r, 1).font = font(11, True, ROSE)
        trig.cell(r, 2).font = font(12, True, INK)
        trig.row_dimensions[r].height = 28
    trig.row_dimensions[17].height = 40
    trig.conditional_formatting.add(
        "C4:C12",
        FormulaRule(formula=['C4="已点亮"'], fill=fill(MINT), font=font(11, True, TEAL)),
    )
    set_widths(trig, [14, 42, 14, 48])
    trig.freeze_panes = "A4"

    # ---- 90天 ----
    act = wb.create_sheet("90天行动")
    act["A1"] = "90 天行动清单：前 30 天收束，中 30 天验证，后 30 天只开一条海外期权"
    act["A1"].font = font(16, True, INK)
    act.merge_cells("A1:F1")
    ah = ["阶段", "序号", "动作", "完成标准", "状态", "记录（日期 / 结果）"]
    for i, h in enumerate(ah, start=1):
        act.cell(3, i, h)
    style_header(act, 3, 6, TEAL)
    actions = [
        ("D1–30 收束", "A1", "只留一条主航道，写清卖什么、卖给谁、凭什么是你", "一页纸，能在 3 分钟内讲完", "未开始"),
        ("D1–30 收束", "A2", "停掉「既要出国又要本地爆」的并行幻想", "日历上删除所有无主航道会议", "未开始"),
        ("D1–30 收束", "A3", "把地产旧能力映射到产业空间 / 存量资产管理", "列出 3 件可立刻做的事", "未开始"),
        ("D31–60 验证", "B1", "推进 1 个可报价、可试点的 AI 场景项目", "出现报价单或试点纪要", "未开始"),
        ("D31–60 验证", "B2", "周复盘：线索 / 会议 / 订单哪一项在动", "连续 4 周有书面复盘", "未开始"),
        ("D31–60 验证", "B3", "宏观新闻降权，客户反馈升权", "每周客户对话 ≥ 宏观阅读时间", "未开始"),
        ("D61–90 期权", "C1", "准备英文一页纸 + 两个项目案例", "PDF 可外发", "未开始"),
        ("D61–90 期权", "C2", "锁定 1–2 个跨境合作接口（机构 / 校友 / 展会）", "至少一次实质对话记录", "未开始"),
        ("D61–90 期权", "C3", "没有书面机会，就不谈移民时间表", "家庭会议达成这条纪律", "未开始"),
    ]
    for i, row in enumerate(actions, start=4):
        for c, val in enumerate(row, start=1):
            act.cell(i, c, val)
        bg = MINT if row[0].startswith("D1") else CREAM if row[0].startswith("D31") else SOFT
        style_range(act, i, i, 1, 6, bg)
        act.cell(i, 1).font = font(11, True)
        act.row_dimensions[i].height = 32
    st = DataValidation(type="list", formula1='"未开始,进行中,已完成,已放弃"', allow_blank=False)
    act.add_data_validation(st)
    st.add("E4:E12")
    act.cell(14, 1, "已完成")
    act.cell(14, 2, '=COUNTIF(E4:E12,"已完成")&" / 9"')
    act.cell(15, 1, "进行中")
    act.cell(15, 2, '=COUNTIF(E4:E12,"进行中")')
    act.cell(16, 1, "卡点提示")
    act.cell(16, 2, '=IF(COUNTIF(E4:E6,"已完成")<3,"先完成收束，再谈验证。",IF(COUNTIF(E7:E9,"已完成")<1,"还没有报价或试点，出国讨论暂停。","可以开始做海外期权，但仍禁止无书面机会就定移民时间表。"))')
    for r in range(14, 17):
        act.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        style_range(act, r, r, 1, 6, SOFT)
        act.cell(r, 1).font = font(11, True, TEAL)
        act.cell(r, 2).font = font(12, True, INK)
        act.row_dimensions[r].height = 28
    act.row_dimensions[16].height = 40
    act.conditional_formatting.add(
        "E4:E12",
        FormulaRule(formula=['E4="已完成"'], fill=fill(MINT), font=font(11, True, TEAL)),
    )
    act.conditional_formatting.add(
        "E4:E12",
        FormulaRule(formula=['E4="已放弃"'], fill=fill("FFE4E6"), font=font(11, True, ROSE)),
    )
    set_widths(act, [18, 10, 48, 32, 12, 28])
    act.freeze_panes = "A4"
    act.auto_filter.ref = "A3:F12"

    # ---- 那段话 ----
    talk = wb.create_sheet("对那段话")
    talk["A1"] = "那段话：收藏修辞，不让修辞代替战略"
    talk["A1"].font = font(16, True, INK)
    talk.merge_cells("A1:C1")
    talk.cell(3, 1, "原句（诀别）")
    talk.cell(3, 2, "改写（战略）")
    talk.cell(3, 3, "为什么要改")
    style_header(talk, 3, 3, INK)
    pairs = [
        ("不要回来", "可以回来，也可以不回来；回来与否由价值与心安决定，不由誓言决定", "「不要回来」是走了之后的自我保护，不能充当走之前的可行性研究"),
        ("不要想念我们", "可以想家，但不让想家做决定；每周只留一个固定时段处理思念", "禁止想念会让思念改道成暗伤；管理想念才是成人"),
        ("不要回头", "可以回头看，但每周只看一次，看完继续做主航道", "完全不回头会失去校准；频繁回头会失去行动"),
        ("不要写信", "可以写信，把思念变成连接，而不是抽签式倾诉", "连接是资产；封存是销毁资产"),
        ("不要向乡愁屈服", "乡愁不是软肋，失控的乡愁才是", "原句把情感污名化，容易让人用决绝掩盖尚未想清的事"),
        ("不要落叶归根", "不必归根，也不必斩根", "归根不是失败；斩根也不是成功"),
        ("而要落地生根", "先在能创造价值的地方生根；生根不等于切断旧根", "双根结构比斩根结构更适合有护照、有职业选择的人"),
        ("心安之处即是家", "保留。心安要自己养，不能靠切断记忆来换", "这是原话里唯一应当完整保留的内核"),
        ("把关于这里的一切都封存起来", "把「这里」编码进能力，而不是封存进遗忘", "对网络型从业者，这里的一切往往就是资产负债表"),
        ("别让思念成为你的软肋", "别让失控的思念成为软肋；健康的思念是导航，不是枷锁", "软肋来自失控，不来自还有所爱"),
    ]
    for i, (a, b, c) in enumerate(pairs, start=4):
        talk.cell(i, 1, a)
        talk.cell(i, 2, b)
        talk.cell(i, 3, c)
        style_range(talk, i, i, 1, 3, MINT if i == 11 else SOFT)
        talk.cell(i, 1).font = font(11, True)
        talk.row_dimensions[i].height = 42
    talk["A15"] = "可带走的三句"
    talk["A15"].font = font(14, True, TEAL)
    talk.merge_cells("A15:C15")
    talk["A16"] = "别回头，向前走。不归根，也可以不斩根。心安处，即吾乡。"
    talk["A16"].font = font(14, True, INK)
    talk.merge_cells("A16:C16")
    talk.row_dimensions[16].height = 28
    talk["A18"] = "家是养出来的，不是逃出来的。如果必须把过去全部封存才能迈步，说明迈步的理由还不够硬，硬的是情绪。"
    talk["A18"].font = font(11, False, SLATE)
    talk.merge_cells("A18:C18")
    talk.row_dimensions[18].height = 32
    set_widths(talk, [28, 52, 52])
    talk.freeze_panes = "A4"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"已生成：{output_path}")


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    build_excel(root / "deliverables" / "发展与去留_2026中美判断备忘录_决策表.xlsx")
