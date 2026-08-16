#!/usr/bin/env python3
"""生成 WorkBuddy 银行引荐行动表：银行对照、十天节奏、约见准备、口径、决策树。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


NAVY = "0B2F5B"
GOLD = "C4A35A"
WHITE = "FFFFFF"
DARK = "1F2A37"
SOFT = "E8EEF5"
GREEN = "2F6B4F"
RED = "A63D2F"
AMBER = "B57A2A"
LIGHT = "F3F6FA"

thin = Border(
    left=Side(style="thin", color="C5CDD6"),
    right=Side(style="thin", color="C5CDD6"),
    top=Side(style="thin", color="C5CDD6"),
    bottom=Side(style="thin", color="C5CDD6"),
)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(size=11, bold=False, color=DARK, name="Microsoft YaHei"):
    return Font(name=name, size=size, bold=bold, color=color)


def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def paint_title(ws, cell, text, *, size=16):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=WHITE)
    ws[cell].fill = fill(NAVY)
    ws[cell].alignment = align("left", "center")


def header_row(ws, row, values, *, fill_color=NAVY, font_color=WHITE):
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font(size=11, bold=True, color=font_color)
        cell.fill = fill(fill_color)
        cell.alignment = align("center", "center")
        cell.border = thin


def write_row(ws, row, values, *, bg=None, bold=False, color=DARK, height=32):
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = font(size=11, bold=bold, color=color)
        cell.fill = fill(bg or WHITE)
        cell.alignment = align("left", "center")
        cell.border = thin
    ws.row_dimensions[row].height = height


def style_status(cell, text: str) -> None:
    mapping = {
        "主线": (NAVY, WHITE),
        "第二主线": (GREEN, WHITE),
        "探索": (AMBER, WHITE),
        "进行中": (AMBER, WHITE),
        "待确认": (AMBER, WHITE),
        "预热": (SOFT, NAVY),
        "会后决定": (SOFT, NAVY),
        "完成": (GREEN, WHITE),
        "本周不做": (RED, WHITE),
        "红线": (RED, WHITE),
    }
    bg, fg = mapping.get(text, (WHITE, DARK))
    cell.fill = fill(bg)
    cell.font = font(size=11, bold=True, color=fg)
    cell.alignment = align("center", "center")


def build_workbook(output_path: Path) -> None:
    wb = Workbook()

    # —— 1. 结论卡 ——
    ws = wb.active
    ws.title = "结论卡"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:F1")
    paint_title(ws, "A1", "  WorkBuddy 银行引荐判断 · 结论卡    2026-08-16")
    set_col_widths(ws, {"A": 18, "B": 22, "C": 28, "D": 28, "E": 28, "F": 22})
    ws.row_dimensions[2].height = 8

    ws.merge_cells("A3:F3")
    ws["A3"] = "一句话：可以引荐，但不要本周并行约见。先把 8 月 21 日杨行长这场走实；上海银行做第二主线预热，中国银行只做探索。不要用新银行催腾讯。"
    ws["A3"].font = font(size=13, bold=True, color=NAVY)
    ws["A3"].fill = fill(SOFT)
    ws["A3"].alignment = align("left", "center")
    ws.row_dimensions[3].height = 48

    header_row(ws, 5, ["事项", "对象", "本周怎么做", "会后怎么做", "不要做", "优先级"])
    rows = [
        ["主线约见", "泰隆上海分行杨行长", "确认时间、出席、可说边界；准备一页纸", "按会上信号推进试点和牵头人", "不要请当场定采购额；不要提另外两家", "P0"],
        ["已知验证", "东口支行行长", "会上用作铺垫：一线已验证话题", "如需试点，可回支行落网点", "不要把支行见面说成分行已立项", "P0"],
        ["第二主线", "上海银行", "只探询人选和条线，不约正式会", "杨行长场空则正式引荐；有方向则继续预热", "不要会前拉腾讯去见", "P1"],
        ["探索接触", "中国银行", "最多问是否方便会后聊 30 分钟", "仅在腾讯能讲清边界后再安排", "不要当催单对手盘；不要对“整个中行”引荐", "P2"],
        ["腾讯卡点", "腾讯 WorkBuddy 团队", "问清是缺银行样本，还是内部方案没齐", "病因 A 才加线；病因 B 先补说法", "不要指责“你们太慢”", "P0"],
    ]
    for i, r in enumerate(rows):
        write_row(ws, 6 + i, r, bg=LIGHT if i % 2 else WHITE, height=46)
        prio = ws.cell(row=6 + i, column=6)
        if r[5] == "P0":
            style_status(prio, "主线")
            prio.value = "P0"
        elif r[5] == "P1":
            style_status(prio, "第二主线")
            prio.value = "P1"
        else:
            style_status(prio, "探索")
            prio.value = "P2"

    ws.merge_cells("A12:F12")
    ws["A12"] = "会前必须问腾讯的五句话"
    ws["A12"].font = font(size=12, bold=True, color=WHITE)
    ws["A12"].fill = fill(NAVY)
    questions = [
        "1. 杨行长若问“你们希望我们做什么”，标准回答是哪一句？",
        "2. 首期试点，最低希望银行做什么、最高不能承诺什么？",
        "3. 价格、发票、结算，会上能不能讲，还是一律会后书面给？",
        "4. 数据存在哪、日志谁能看、是否出域，谁来答、答到哪一步？",
        "5. 最近想得比较长，卡点是缺银行样本，还是内部方案没齐？",
    ]
    for i, q in enumerate(questions):
        row = 13 + i
        write_row(ws, row, [q, "", "", "", "", ""], bg=WHITE if i % 2 else LIGHT, height=28)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    ws.merge_cells("A19:F19")
    ws["A19"] = "说明：本表供引荐人自用。1,500 万等数字仅来自既有讨论框架，不是采购承诺。杨行长、东口支行按现有约见口径填写。"
    ws["A19"].font = font(size=9, color="5C6B7A")
    ws["A19"].alignment = align("left", "center")
    ws.row_dimensions[19].height = 24
    ws.freeze_panes = "A6"
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # —— 2. 三家银行对照 ——
    ws = wb.create_sheet("三家银行对照")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    paint_title(ws, "A1", "  三家银行对照：能级不同，不能同一周拉上台")
    ws.row_dimensions[1].height = 28
    set_col_widths(ws, {"A": 20, "B": 32, "C": 32, "D": 32, "E": 18})
    header_row(ws, 3, ["维度", "泰隆银行", "上海银行", "中国银行", "本周策略"])
    compare = [
        ["角色定位", "本周唯一主线", "最合适的第二主线", "品牌型探索，不当对手盘", "主次分开"],
        ["机构类型", "城商行，主做中小微", "总部在上海的城商行", "国有大行，流程长、合规严", "先城商后大行"],
        ["与 WorkBuddy 贴合度", "客群最贴，已有讨论框架", "叙事可复用泰隆那套", "需另做总行/分行/条线故事", "不复制 1500 万话术"],
        ["当前进度", "已对接；东口支行行长已见；周五见杨行长", "尚未正式引荐", "尚未正式引荐", "只推进泰隆"],
        ["决策半径", "分行行长可定方向、试点、牵头人", "相对短，适合先分行部门或支行", "长，必须落到具体支行或条线", "中行不对“整个中行”引荐"],
        ["本周动作", "正式约见，要方向", "轻量探询人选和条线", "最多问是否愿会后聊 30 分钟", "不并行约见"],
        ["会后动作", "按信号推进试点", "场空则正式引荐；有方向则预热", "仅腾讯能讲清边界后再安排", "用会后决策树"],
        ["适合层级", "上海分行杨行长", "分行公司金融/零售权益/科技，或支行一把手", "上海某支行、普惠、公司金融或权益部门", "宁小勿大"],
        ["主要风险", "会上仍“再研究”；金额被当成承诺", "会前见面会冲泰隆主线", "第一印象变成没准备好，以后更难约", "保护第一印象"],
        ["对腾讯的价值", "最近、最可能立项的一家", "验证模式可复制", "国有大行认知，不解决本周节奏", "不把中行当催单"],
    ]
    for i, r in enumerate(compare):
        write_row(ws, 4 + i, r, bg=LIGHT if i % 2 else WHITE, height=40)
        style_status(ws.cell(row=4 + i, column=5), r[4] if r[4] in ("主线", "第二主线", "探索") else "")
        ws.cell(row=4 + i, column=5).value = r[4]
        ws.cell(row=4 + i, column=5).alignment = align("center", "center")
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # —— 3. 十天节奏 ——
    ws = wb.create_sheet("十天节奏")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    paint_title(ws, "A1", "  十天行动节奏  8月16日（日）— 8月25日（二）")
    ws.row_dimensions[1].height = 28
    set_col_widths(ws, {"A": 16, "B": 14, "C": 18, "D": 36, "E": 22, "F": 14, "G": 22})
    header_row(ws, 3, ["日期", "星期", "对象", "动作", "产出", "状态", "备注"])
    plan = [
        ["2026-08-16", "日", "自己 / 腾讯", "把可说/不可说和五问发给腾讯；问清卡点", "判断病因 A 或 B", "进行中", "今日"],
        ["2026-08-16", "日", "杨行长会务", "确认周五时间、地点、分行出席名单", "会务清单", "待确认", ""],
        ["2026-08-17", "一", "腾讯", "书面确认出席人和会上标准回答", "可说边界点头", "待确认", ""],
        ["2026-08-17", "一", "上海银行", "只整理人选和条线，不发正式邀约", "预热名单", "预热", "本周不约见"],
        ["2026-08-17", "一", "中国银行", "只整理具体支行/条线人选", "预热名单", "预热", "不对整个中行"],
        ["2026-08-18", "二", "材料", "准备一页纸：背景、试点建议、请行长定三件事", "一页纸", "待确认", ""],
        ["2026-08-19", "三", "腾讯", "模拟提问：采购额、数据安全、和谁签合同", "答问清单", "待确认", ""],
        ["2026-08-20", "四", "上海银行 / 中行", "最多各发一条预热，把正式会放到 8/25 后", "探询消息", "预热", "可跳过"],
        ["2026-08-21", "五", "杨行长", "开场用东口支行铺垫；盯牵头人、试点、下次谁来", "会后半页纪要", "待确认", "P0 主场"],
        ["2026-08-21", "五", "腾讯", "当天同步纪要和判断", "内部复盘", "待确认", ""],
        ["2026-08-24", "一", "自己", "按决策树决定辅线升级还是按住", "会后判断", "会后决定", ""],
        ["2026-08-25", "二", "上海银行", "若周五场空：正式引荐；若有方向：继续预热", "引荐或按住", "会后决定", ""],
        ["2026-08-25", "二", "中国银行", "默认继续探索；仅腾讯明确要大行样本时再约", "暂缓或一人探询", "会后决定", ""],
    ]
    for i, r in enumerate(plan):
        write_row(ws, 4 + i, r, bg=LIGHT if i % 2 else WHITE, height=36)
        style_status(ws.cell(row=4 + i, column=6), r[5])
    dv = DataValidation(type="list", formula1='"进行中,待确认,预热,完成,会后决定,本周不做"', allow_blank=True)
    dv.add("F4:F16")
    ws.add_data_validation(dv)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:G16"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # —— 4. 杨行长约见准备 ——
    ws = wb.create_sheet("杨行长约见准备")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    paint_title(ws, "A1", "  8月21日 泰隆上海分行杨行长约见准备")
    ws.row_dimensions[1].height = 28
    set_col_widths(ws, {"A": 16, "B": 22, "C": 40, "D": 18, "E": 22})
    header_row(ws, 3, ["模块", "项目", "内容", "负责人", "状态"])
    brief = [
        ["会务", "日期", "2026-08-21（周五）", "引荐人", "待确认"],
        ["会务", "时间 / 地点", "待填", "引荐人", "待确认"],
        ["会务", "分行出席", "杨行长；建议请东口支行行长或相关条线同事陪同", "引荐人", "待确认"],
        ["会务", "腾讯出席", "待填（能讲产品和能定边界的人要到）", "腾讯", "待确认"],
        ["目标", "P0 原则同意", "作为中小微增值服务方向继续论证", "全场", "待确认"],
        ["目标", "P0 指定牵头", "明确部门或人，避免会后再研究", "杨行长", "待确认"],
        ["目标", "P1 划定试点", "部分网点或客群，先小后大", "杨行长", "待确认"],
        ["目标", "P1 约下次", "商务/科技/合规谁来、何时", "双方", "待确认"],
        ["议程", "0-5 分钟", "引荐人说明来意：支行已见，请分行定方向", "引荐人", "待确认"],
        ["议程", "5-20 分钟", "腾讯讲 WorkBuddy 是什么、银行能用在哪", "腾讯", "待确认"],
        ["议程", "20-35 分钟", "只讲试点怎么做，不讲全年总包", "腾讯 + 引荐人", "待确认"],
        ["议程", "35-50 分钟", "提问", "杨行长一侧", "待确认"],
        ["议程", "50-60 分钟", "请行长定三件事：是否继续、谁牵头、试点范围", "杨行长", "待确认"],
        ["可说", "铺垫", "东口支行行长已交流，一线认为有差异化权益价值", "引荐人", "进行中"],
        ["可说", "主方向", "积分兑换、开户/拜访赠 AI；员工应用和联名卡可研究", "腾讯", "待确认"],
        ["可说", "路径", "先试点后扩量；兑换码自助激活；不接核心系统", "腾讯", "待确认"],
        ["不可说", "金额", "不把 1,500 万说成已定预算或报价承诺", "全场", "红线"],
        ["不可说", "口径", "不说行业首创、腾讯联合产品等未授权表述", "全场", "红线"],
        ["不可说", "辅线", "不提中国银行、上海银行，不用别家催泰隆", "全场", "红线"],
        ["不可说", "采购", "不请杨行长当场定采购金额", "全场", "红线"],
        ["会后", "当天", "半页纪要：方向、牵头人、试点、下次、未决问题", "引荐人", "待确认"],
    ]
    for i, r in enumerate(brief):
        write_row(ws, 4 + i, r, bg=LIGHT if i % 2 else WHITE, height=34)
        style_status(ws.cell(row=4 + i, column=5), r[4])
    dv2 = DataValidation(type="list", formula1='"进行中,待确认,完成,红线"', allow_blank=True)
    dv2.add("E4:E24")
    ws.add_data_validation(dv2)
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # —— 5. 口径卡 ——
    ws = wb.create_sheet("口径卡")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    paint_title(ws, "A1", "  口径卡：同一周，四套话不能串")
    ws.row_dimensions[1].height = 28
    set_col_widths(ws, {"A": 18, "B": 70, "C": 28})
    header_row(ws, 3, ["对象", "建议原话", "切忌"])
    lines = [
        [
            "对腾讯",
            "泰隆下周五见杨行长是本周唯一正式场。支行已经见过，分行这场是定方向的。中国银行和上海银行我可以帮你们预热人选，不占你们会前时间。等泰隆信号出来再决定要不要正式约。你们如果卡在方案和合规，我先不把大行带上场。",
            "不要说“你们太慢，我再找两家催催”",
        ],
        [
            "对上海银行（预热）",
            "腾讯 WorkBuddy 在看银行渠道，怎么把企业 AI 做成客户权益和员工提效。上海已经有一家城商行在交流。想先了解你们公司金融或零售权益条线，是否方便做一次业务沟通。不是采购谈判，先看场景合不合。",
            "不要说腾讯已经要采购，不要报 1500 万",
        ],
        [
            "对中国银行（探索）",
            "想介绍一个企业 AI 工具，看是否适合作为客户增值或网点服务补充。先找上海地区具体支行或普惠/公司金融条线做一次轻量交流，不谈总行级合作，也不谈采购额。",
            "不要对“中国银行”整体引荐，不要承诺复制泰隆规模",
        ],
        [
            "对泰隆杨行长",
            "东口支行已经沟通过。今天请杨行长看两件事：这个方向值不值得在上海试点；如果值得，谁来牵头。金额、系统对接、联名卡都可以后置，先定方向。",
            "不要提中行和上海银行，不要逼当场定金额",
        ],
    ]
    for i, r in enumerate(lines):
        write_row(ws, 4 + i, r, bg=LIGHT if i % 2 else WHITE, height=78)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # —— 6. 会后决策树 ——
    ws = wb.create_sheet("会后决策树")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    paint_title(ws, "A1", "  用杨行长会的结果决定要不要引荐（会后填写）")
    ws.row_dimensions[1].height = 28
    set_col_widths(ws, {"A": 28, "B": 18, "C": 32, "D": 32, "E": 22})
    header_row(ws, 3, ["如果会上出现", "信号", "泰隆下一步", "辅线怎么做", "是否正式引荐"])
    tree = [
        ["原则同意，且指定牵头人", "强正向", "继续深挖试点方案和部门对接", "上海银行保持预热；中行暂缓", "暂不"],
        ["态度正向，但要内部再议", "弱正向", "给 1—2 周，约定回复节点", "同时约上海银行做探索性沟通", "上海银行：是；中行：否"],
        ["礼貌、无牵头、无下次时间", "空转", "主线降为观察，不再加码催", "正式引荐上海银行；中行一人探询", "是"],
        ["腾讯答不利落（价格/合规/主体）", "准备不足", "先停，避免银行留下差印象", "停止一切新引荐，先补一页标准说法", "否，先补课"],
        ["明确否定或合规不可做", "终止", "礼貌收口，保留联系", "上海银行可另开，叙事不绑泰隆", "上海银行：视腾讯状态"],
    ]
    for i, r in enumerate(tree):
        write_row(ws, 4 + i, r, bg=LIGHT if i % 2 else WHITE, height=48)
        sig = ws.cell(row=4 + i, column=2)
        mapping = {
            "强正向": "完成",
            "弱正向": "进行中",
            "空转": "探索",
            "准备不足": "红线",
            "终止": "红线",
        }
        style_status(sig, mapping.get(r[1], r[1]))
        sig.value = r[1]

    ws.merge_cells("A10:E10")
    ws["A10"] = "会后记录（现场填写）"
    ws["A10"].font = font(size=12, bold=True, color=WHITE)
    ws["A10"].fill = fill(NAVY)
    labels = [
        ["实际信号", ""],
        ["牵头人 / 部门", ""],
        ["试点范围", ""],
        ["下次时间", ""],
        ["对腾讯的判断（病因 A / B）", ""],
        ["是否启动上海银行正式引荐", ""],
        ["是否对中国银行做一人探询", ""],
        ["备注", ""],
    ]
    for i, (k, v) in enumerate(labels):
        row = 11 + i
        write_row(ws, row, [k, v, "", "", ""], bg=WHITE if i % 2 else LIGHT, height=28)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"已生成：{output_path}")


if __name__ == "__main__":
    out = (
        Path(__file__).resolve().parents[1]
        / "deliverables"
        / "WorkBuddy银行引荐_三家银行对照与十天行动表_20260816.xlsx"
    )
    build_workbook(out)
