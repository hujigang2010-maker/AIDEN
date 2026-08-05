#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将《上海板块学区融合分析》与《上海学区板块差异化选房地图》大融合，
输出一张商务视觉的综合总表 Excel。
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output"
OUT_XLSX = OUT_DIR / "上海板块学区融合总表.xlsx"

UPLOAD_CANDIDATES = list(Path("/home/ubuntu/.cursor/projects/workspace/uploads").glob("*.xlsx"))
MAP_XLSX = OUT_DIR / "上海学区板块差异化选房地图.xlsx"

# —— 商务配色（深蓝灰主色，避免紫/奶油 AI 默认风）——
C = {
    "navy": "1B2A4A",
    "navy2": "243B5C",
    "teal": "0E7490",
    "teal_soft": "E0F2FE",
    "ink": "0F172A",
    "muted": "64748B",
    "line": "CBD5E1",
    "zebra": "F8FAFC",
    "white": "FFFFFF",
    "gold": "B45309",
    "gold_bg": "FEF3C7",
    "green": "0F766E",
    "green_bg": "CCFBF1",
    "blue": "1D4ED8",
    "blue_bg": "DBEAFE",
    "amber": "C2410C",
    "amber_bg": "FFEDD5",
    "red": "B91C1C",
    "red_bg": "FEE2E2",
    "slate_bg": "F1F5F9",
    "kpi_bg": "EEF2FF",
}

GRADE_STYLE = {
    "⭐⭐⭐ 强烈推荐": ("green", "green_bg"),
    "⭐⭐ 可考虑": ("blue", "blue_bg"),
    "⭐ 谨慎": ("amber", "amber_bg"),
    "❌ 排除": ("red", "red_bg"),
}

MARK_STYLE = {
    "★优选锁定": ("green", "green_bg"),
    "☆次优关注": ("blue", "blue_bg"),
    "○可观察": ("gold", "gold_bg"),
    "△条件不符": ("amber", "amber_bg"),
    "✕策略排除": ("red", "red_bg"),
}

# 板块名称归一化别名：上传文件名 -> 地图文件可能名
ALIASES = {
    "潍坊新村": ["潍坊新村（陆家嘴南）", "潍坊新村", "潍坊"],
    "东外滩/杨浦滨江": ["杨浦滨江", "东外滩/杨浦滨江", "东外滩"],
    "豫园": ["豫园/小东门", "豫园"],
    "小东门": ["豫园/小东门", "小东门"],
    "老西门": ["老西门/蓬莱", "老西门"],
    "南京东路": ["人民广场/南京东路", "南京东路"],
    "外滩": ["人民广场/南京东路", "外滩"],
    "鞍山": ["鞍山/四平", "鞍山"],
    "黄兴公园": ["黄兴公园/延吉", "黄兴公园"],
    "四川北路": ["四川北路/临平路", "四川北路"],
    "陆家嘴": ["陆家嘴金融城核心", "陆家嘴"],
    "花木": ["花木/世纪公园", "花木"],
    "三林": ["三林/御桥", "三林"],
    "周浦": ["周浦/康桥", "周浦"],
    "川沙": ["川沙/唐镇", "川沙"],
    "唐镇": ["川沙/唐镇", "唐镇"],
    "莘庄": ["莘庄"],
    "七宝": ["七宝"],
    "古美": ["古美/梅陇", "古美"],
    "梅陇": ["古美/梅陇", "梅陇"],
    "颛桥": ["颛桥/浦江", "颛桥"],
    "浦江": ["颛桥/浦江", "浦江"],
    "静安寺": ["南京西路/静安寺", "静安寺"],
    "南京西路": ["南京西路/静安寺", "南京西路"],
    "武宁": ["长风/武宁", "武宁"],
    "长风": ["长风/武宁", "长风"],
    "共康": ["共康/高境", "共康"],
    "淞南": ["淞宝/吴淞", "淞南"],
    "石化": ["石化/山阳", "石化"],
    "金山新城": ["石化/山阳", "金山新城"],
    "安亭": ["安亭/马陆", "安亭"],
    "徐家汇": ["衡复/徐家汇", "徐家汇"],
    "华泾": ["华泾/漕河泾", "华泾"],
    "漕河泾": ["华泾/漕河泾", "漕河泾"],
}


def thin() -> Border:
    s = Side(style="thin", color=C["line"])
    return Border(left=s, right=s, top=s, bottom=s)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, size=11, color=None, name="微软雅黑"):
    return Font(name=name, bold=bold, size=size, color=color or C["ink"])


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[（(].*?[）)]", "", s)
    s = s.replace("/", "").replace("·", "").replace(" ", "").replace("－", "-")
    return s


def load_upload(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["板块学区融合"]
    rows = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 17)]
        if not vals[3]:
            continue
        rows.append(
            {
                "grade": vals[0],
                "score": vals[1],
                "district": vals[2],
                "plate": vals[3],
                "street": vals[4],
                "primary": vals[5],
                "middle": vals[6],
                "entry": vals[7],
                "five_year": vals[8],
                "hollow": vals[9],
                "newcomer": vals[10],
                "competition": vals[11],
                "primary_q": vals[12],
                "middle_q": vals[13],
                "price": vals[14],
                "note": vals[15] or "",
            }
        )
    # 评分说明
    legend = []
    ws2 = wb["评分说明"]
    for r in range(1, ws2.max_row + 1):
        legend.append([ws2.cell(r, c).value for c in range(1, 4)])
    return rows, legend


def load_map(path: Path) -> tuple[list[dict], list[dict]]:
    wb = load_workbook(path, data_only=True)
    plates = []
    ws = wb["全市板块标记总表"]
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 14)]
        if not vals[3]:
            continue
        plates.append(
            {
                "mark": vals[0],
                "selected": vals[1],
                "district": vals[2],
                "plate": vals[3],
                "primary": vals[4],
                "middle": vals[5],
                "pop": vals[6],
                "comp": vals[7],
                "hollow": vals[8],
                "fit": vals[9],
                "price": vals[10],
                "policy": vals[11],
                "reason": vals[12],
            }
        )
    districts = []
    ws3 = wb["十六区策略总览"]
    for r in range(3, ws3.max_row + 1):
        vals = [ws3.cell(r, c).value for c in range(1, 9)]
        if vals[1]:
            districts.append(
                {
                    "mark": vals[0],
                    "district": vals[1],
                    "entry": vals[2],
                    "pop": vals[3],
                    "comp": vals[4],
                    "protect": vals[5],
                    "zhongkao": vals[6],
                    "conclusion": vals[7],
                }
            )
    return plates, districts


def build_map_index(map_plates: list[dict]) -> dict[str, list[dict]]:
    idx = defaultdict(list)
    for p in map_plates:
        key = f"{p['district']}|{norm(p['plate'])}"
        idx[key].append(p)
        # also index by district + each alias fragment
        for frag in re.split(r"[/、]", p["plate"]):
            idx[f"{p['district']}|{norm(frag)}"].append(p)
    return idx


def find_map_match(district: str, plate: str, idx: dict) -> dict | None:
    candidates = []
    names = ALIASES.get(plate, [plate]) + [plate]
    for name in names:
        key = f"{district}|{norm(name)}"
        candidates.extend(idx.get(key, []))
        # partial contains
        for k, items in idx.items():
            if not k.startswith(district + "|"):
                continue
            kn = k.split("|", 1)[1]
            nn = norm(name)
            if nn and (nn in kn or kn in nn):
                candidates.extend(items)
    # dedupe by plate name
    seen = set()
    uniq = []
    for c in candidates:
        if c["plate"] not in seen:
            seen.add(c["plate"])
            uniq.append(c)
    if not uniq:
        return None
    # prefer exact-ish
    for c in uniq:
        if norm(plate) in norm(c["plate"]) or norm(c["plate"]) in norm(plate):
            return c
    return uniq[0]


def unify_grade(upload_grade: str, map_mark: str | None, selected: str | None) -> tuple[str, str]:
    """返回 (融合推荐等级, 融合结论标签)"""
    if map_mark == "★优选锁定" or (selected == "是" and map_mark in ("★优选锁定", "☆次优关注")):
        if upload_grade == "⭐⭐⭐ 强烈推荐":
            return "⭐⭐⭐ 强烈推荐", "双源一致·优选"
        if map_mark == "★优选锁定":
            return "⭐⭐⭐ 强烈推荐", "纪要优选·上调"
        return "⭐⭐ 可考虑", "纪要入选·次优"
    if map_mark == "△条件不符":
        return "⭐ 谨慎", "年限条件不符"
    if map_mark == "✕策略排除" or upload_grade == "❌ 排除":
        return "❌ 排除", "策略排除"
    if upload_grade:
        tag = {
            "⭐⭐⭐ 强烈推荐": "融合分析优选",
            "⭐⭐ 可考虑": "可观察备选",
            "⭐ 谨慎": "谨慎评估",
            "❌ 排除": "策略排除",
        }.get(upload_grade, "待核验")
        return upload_grade, tag
    return "⭐ 谨慎", "仅地图侧"


def merge_rows(upload_rows: list[dict], map_plates: list[dict]) -> list[dict]:
    idx = build_map_index(map_plates)
    matched_map_plates = set()
    fused = []

    for u in upload_rows:
        m = find_map_match(u["district"], u["plate"], idx)
        if m:
            matched_map_plates.add(m["plate"])
        mark = m["mark"] if m else None
        selected = m["selected"] if m else "否"
        # 纪要硬规则覆盖：九亭/嘉定/松江/闵行大部分等已在上传标记；鞍山地图为优选则上调
        grade, verdict = unify_grade(u["grade"], mark, selected if m else None)

        # 鞍山特殊：地图明确优选，上传为谨慎 → 融合后强烈推荐并说明
        if u["district"] == "杨浦" and "鞍山" in u["plate"] and mark == "★优选锁定":
            grade, verdict = "⭐⭐⭐ 强烈推荐", "纪要挂户逻辑·上调"

        # 崇明城桥等远郊：空心但资源弱，保持可考虑但标注通勤
        note_parts = [u["note"]]
        if m and m.get("reason") and m["reason"] not in (u["note"] or ""):
            note_parts.append(f"【地图】{m['reason']}")
        if m and m.get("policy"):
            note_parts.append(f"【政策】{m['policy']}")

        fused.append(
            {
                "grade": grade,
                "verdict": verdict,
                "mark": mark or "—",
                "selected": "否",  # 稍后按等级与标记统一回填
                "score": u["score"],
                "fit": m["fit"] if m else None,
                "district": u["district"],
                "plate": u["plate"],
                "street": u["street"],
                "primary": u["primary"] or (m["primary"] if m else ""),
                "middle": u["middle"] or (m["middle"] if m else ""),
                "entry": u["entry"],
                "five_year": u["five_year"],
                "hollow": u["hollow"],
                "newcomer": u["newcomer"],
                "competition": u["competition"],
                "primary_q": u["primary_q"],
                "middle_q": u["middle_q"],
                "price": u["price"] or (m["price"] if m else ""),
                "pop": (m["pop"] if m else ""),
                "source": "双源融合" if m else "融合分析表",
                "note": "；".join([p for p in note_parts if p]),
            }
        )

    # 补入地图有、上传没有的板块
    upload_keys = {(r["district"], norm(r["plate"])) for r in upload_rows}
    for m in map_plates:
        if m["plate"] in matched_map_plates:
            continue
        # check fuzzy already covered
        covered = False
        for d, p in upload_keys:
            if d != m["district"]:
                continue
            if norm(p) in norm(m["plate"]) or norm(m["plate"]) in norm(p):
                covered = True
                break
            for alias_src, aliases in ALIASES.items():
                if m["plate"] in aliases or any(norm(a) == norm(m["plate"]) for a in aliases):
                    if norm(alias_src) == p or alias_src == p:
                        covered = True
        if covered:
            continue

        grade, verdict = unify_grade(None, m["mark"], m["selected"])
        # map-only grade from mark
        if m["mark"] == "★优选锁定":
            grade = "⭐⭐⭐ 强烈推荐"
        elif m["mark"] == "☆次优关注":
            grade = "⭐⭐ 可考虑"
        elif m["mark"] == "○可观察":
            grade = "⭐⭐ 可考虑"
        elif m["mark"] == "△条件不符":
            grade = "⭐ 谨慎"
        else:
            grade = "❌ 排除"

        fused.append(
            {
                "grade": grade,
                "verdict": verdict if verdict != "仅地图侧" else "地图补录",
                "mark": m["mark"],
                "selected": m["selected"],
                "score": None,
                "fit": m["fit"],
                "district": m["district"],
                "plate": m["plate"],
                "street": "—",
                "primary": m["primary"],
                "middle": m["middle"],
                "entry": "—",
                "five_year": "—",
                "hollow": m["hollow"],
                "newcomer": None,
                "competition": m["comp"],
                "primary_q": None,
                "middle_q": None,
                "price": f"{m['price']}万" if m["price"] and "万" not in str(m["price"]) else m["price"],
                "pop": m["pop"],
                "source": "差异化地图补录",
                "note": f"{m['reason']}；【政策】{m['policy']}",
            }
        )

    # 入选标记修正：强烈推荐默认入选看房清单
    for row in fused:
        if row["grade"] == "⭐⭐⭐ 强烈推荐":
            row["selected"] = "是"
        if row["mark"] == "☆次优关注":
            row["selected"] = "是"

    # 排序：等级权重 + 综合分 + 契合分
    grade_rank = {"⭐⭐⭐ 强烈推荐": 0, "⭐⭐ 可考虑": 1, "⭐ 谨慎": 2, "❌ 排除": 3}
    fused.sort(
        key=lambda r: (
            grade_rank.get(r["grade"], 9),
            -(r["score"] or 0),
            -(r["fit"] or 0),
            r["district"],
            r["plate"],
        )
    )
    return fused


def style_header_row(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill(C["navy"])
        cell.font = font(bold=True, size=10, color=C["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()


def apply_grade_cell(cell, grade: str):
    style = GRADE_STYLE.get(grade)
    if style:
        fg, bg = style
        cell.fill = fill(C[bg])
        cell.font = font(bold=True, size=10, color=C[fg])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin()


def apply_mark_cell(cell, mark: str):
    style = MARK_STYLE.get(mark)
    if style:
        fg, bg = style
        cell.fill = fill(C[bg])
        cell.font = font(bold=True, size=10, color=C[fg])
    else:
        cell.font = font(size=10, color=C["muted"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin()


def build_workbook(fused: list[dict], districts: list[dict], legend: list) -> Workbook:
    wb = Workbook()

    # ========== Sheet1 融合总表 ==========
    ws = wb.active
    ws.title = "融合总表"
    ws.sheet_view.showGridLines = False

    # 标题带
    ws.merge_cells("A1:V1")
    ws["A1"] = "上海板块 × 学区融合总表（双源整合）"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=18, color=C["white"])
    ws["A1"].fill = fill(C["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:V2")
    ws["A2"] = (
        "数据源①《上海板块学区融合分析》评分体系　｜　数据源②《差异化选房地图》纪要策略标记　｜　"
        "融合逻辑：空心化优先、避开新人口导入高竞争区；对口与年限以教育局当年文件为准"
    )
    ws["A2"].font = font(size=9, color=C["muted"])
    ws["A2"].fill = fill(C["slate_bg"])
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # KPI 行
    n = len(fused)
    c_grade = Counter(r["grade"] for r in fused)
    n_sel = sum(1 for r in fused if r["selected"] == "是")
    kpis = [
        ("板块样本", str(n)),
        ("强烈推荐", str(c_grade.get("⭐⭐⭐ 强烈推荐", 0))),
        ("可考虑", str(c_grade.get("⭐⭐ 可考虑", 0))),
        ("谨慎", str(c_grade.get("⭐ 谨慎", 0))),
        ("排除", str(c_grade.get("❌ 排除", 0))),
        ("本轮入选看房", str(n_sel)),
    ]
    for i, (label, val) in enumerate(kpis):
        col = 1 + i * 2
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        cell = ws.cell(3, col, f"{label}  {val}")
        cell.fill = fill(C["teal_soft"])
        cell.font = font(bold=True, size=10, color=C["teal"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin()
        ws.cell(3, col + 1).border = thin()
        ws.cell(3, col + 1).fill = fill(C["teal_soft"])
    ws.row_dimensions[3].height = 26

    headers = [
        "融合推荐等级",
        "融合结论",
        "纪要策略标记",
        "本轮入选",
        "综合评分",
        "策略契合分",
        "行政区",
        "板块",
        "对应街道",
        "对口小学（代表）",
        "对口初中（代表）",
        "入学方式",
        "五年一户",
        "人口空心化",
        "新上海人占比",
        "升学竞争",
        "小学质量",
        "初中质量",
        "总价区间",
        "人口结构画像",
        "数据来源",
        "融合备注",
    ]
    header_row = 5
    for c, h in enumerate(headers, 1):
        ws.cell(header_row, c, h)
    style_header_row(ws, header_row, len(headers))
    ws.row_dimensions[header_row].height = 32

    for i, r in enumerate(fused):
        row = header_row + 1 + i
        vals = [
            r["grade"],
            r["verdict"],
            r["mark"],
            r["selected"],
            r["score"],
            r["fit"],
            r["district"],
            r["plate"],
            r["street"],
            r["primary"],
            r["middle"],
            r["entry"],
            r["five_year"],
            r["hollow"],
            r["newcomer"],
            r["competition"],
            r["primary_q"],
            r["middle_q"],
            r["price"],
            r["pop"],
            r["source"],
            r["note"],
        ]
        zebra = i % 2 == 1
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v if v is not None and v != "" else "—")
            cell.border = thin()
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center" if c <= 6 else "left")
            cell.font = font(size=9)
            if zebra and c not in (1, 3, 4):
                cell.fill = fill(C["zebra"])

        apply_grade_cell(ws.cell(row, 1), r["grade"])
        apply_mark_cell(ws.cell(row, 3), r["mark"])
        # 入选
        sel = ws.cell(row, 4)
        if r["selected"] == "是":
            sel.fill = fill(C["green_bg"])
            sel.font = font(bold=True, size=10, color=C["green"])
        else:
            sel.font = font(size=9, color=C["muted"])
        sel.alignment = Alignment(horizontal="center", vertical="center")
        # 分数强调
        if isinstance(r["score"], (int, float)):
            ws.cell(row, 5).font = font(bold=True, size=10, color=C["navy"])
        if isinstance(r["fit"], (int, float)):
            ws.cell(row, 6).font = font(bold=True, size=10, color=C["teal"])
        ws.row_dimensions[row].height = 38

    widths = [14, 14, 12, 10, 10, 10, 8, 16, 14, 22, 20, 10, 10, 10, 10, 10, 10, 10, 12, 22, 12, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "I6"
    ws.auto_filter.ref = f"A{header_row}:V{header_row + len(fused)}"
    ws.print_title_rows = "1:5"

    # ========== Sheet2 入选看房清单 ==========
    ws2 = wb.create_sheet("入选看房清单")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:J1")
    ws2["A1"] = "本轮入选看房清单（融合后）"
    ws2["A1"].font = Font(name="微软雅黑", bold=True, size=16, color=C["white"])
    ws2["A1"].fill = fill(C["teal"])
    ws2["A1"].alignment = Alignment(vertical="center")
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells("A2:J2")
    ws2["A2"] = "建议顺序：潍坊新村（世纪大道南）→ 杨浦鞍山/滨江 → 黄浦老城厢 → 虹口/曹杨等次优观察"
    ws2["A2"].font = font(size=9, color=C["muted"])
    ws2["A2"].fill = fill(C["slate_bg"])

    h2 = ["序号", "融合推荐等级", "融合结论", "行政区", "板块", "对口小学", "对口初中", "总价区间", "综合评分", "关键理由"]
    for c, h in enumerate(h2, 1):
        ws2.cell(4, c, h)
    style_header_row(ws2, 4, len(h2))
    # override header to teal for this sheet
    for c in range(1, len(h2) + 1):
        ws2.cell(4, c).fill = fill(C["navy2"])

    selected = [r for r in fused if r["selected"] == "是"]
    for i, r in enumerate(selected, 1):
        row = 4 + i
        vals = [i, r["grade"], r["verdict"], r["district"], r["plate"], r["primary"], r["middle"], r["price"], r["score"], r["note"]]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row, c, v if v is not None else "—")
            cell.border = thin()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = font(size=9)
            if i % 2 == 0 and c != 2:
                cell.fill = fill(C["zebra"])
        apply_grade_cell(ws2.cell(row, 2), r["grade"])
        ws2.row_dimensions[row].height = 42
    for i, w in enumerate([6, 14, 14, 8, 16, 26, 24, 12, 10, 55], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A4:J{4 + len(selected)}"

    # ========== Sheet3 十六区策略 ==========
    ws3 = wb.create_sheet("十六区策略总览")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:H1")
    ws3["A1"] = "十六区策略总览（服务板块筛选）"
    ws3["A1"].font = Font(name="微软雅黑", bold=True, size=16, color=C["white"])
    ws3["A1"].fill = fill(C["navy"])
    ws3.row_dimensions[1].height = 30
    h3 = ["纪要策略标记", "行政区", "入学对口类型", "人口画像", "区内竞争", "本区保护", "中考相对压力", "策略结论"]
    for c, h in enumerate(h3, 1):
        ws3.cell(3, c, h)
    style_header_row(ws3, 3, len(h3))
    for i, d in enumerate(districts):
        row = 4 + i
        vals = [d["mark"], d["district"], d["entry"], d["pop"], d["comp"], d["protect"], d["zhongkao"], d["conclusion"]]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row, c, v)
            cell.border = thin()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = font(size=9)
            if i % 2 == 1 and c != 1:
                cell.fill = fill(C["zebra"])
        apply_mark_cell(ws3.cell(row, 1), d["mark"])
        ws3.row_dimensions[row].height = 36
    for i, w in enumerate([12, 8, 18, 40, 12, 10, 22, 42], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A4"

    # ========== Sheet4 评分与图例 ==========
    ws4 = wb.create_sheet("评分图例与融合规则")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:C1")
    ws4["A1"] = "评分权重 · 等级定义 · 融合规则"
    ws4["A1"].font = Font(name="微软雅黑", bold=True, size=16, color=C["white"])
    ws4["A1"].fill = fill(C["navy"])
    ws4.row_dimensions[1].height = 30

    ws4["A3"] = "一、融合分析表评分维度"
    ws4["A3"].font = font(bold=True, size=12, color=C["teal"])
    for c, h in enumerate(["维度", "权重", "说明"], 1):
        cell = ws4.cell(4, c, h)
        cell.fill = fill(C["navy2"])
        cell.font = font(bold=True, size=10, color=C["white"])
        cell.border = thin()
    # from legend first block
    rr = 5
    for row in legend:
        if row[0] in (None, "推荐等级", "重要提示") or (row[0] and str(row[0]).startswith("⭐")):
            if row[0] == "推荐等级":
                break
            continue
        if row[0] and row[1]:
            for c, v in enumerate(row[:3], 1):
                cell = ws4.cell(rr, c, v)
                cell.border = thin()
                cell.font = font(size=9)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws4.row_dimensions[rr].height = 28
            rr += 1

    rr += 1
    ws4.cell(rr, 1, "二、推荐等级定义").font = font(bold=True, size=12, color=C["teal"])
    rr += 1
    for c, h in enumerate(["等级", "含义"], 1):
        cell = ws4.cell(rr, c, h)
        cell.fill = fill(C["navy2"])
        cell.font = font(bold=True, size=10, color=C["white"])
        cell.border = thin()
    rr += 1
    grade_defs = [
        ("⭐⭐⭐ 强烈推荐", "符合人口空心化 + 低竞争；或纪要明确优选 / 双源一致上调"),
        ("⭐⭐ 可考虑", "部分指标良好，或纪要次优入选；需核验对口与入户年限"),
        ("⭐ 谨慎", "学区/竞争/年限存在短板，或条件暂不符（如九年一贯需提前3年）"),
        ("❌ 排除", "纪要排除的高竞争导入区（嘉定/松江/闵行大部分/浦东大部分等）"),
    ]
    for g, desc in grade_defs:
        ws4.cell(rr, 1, g)
        apply_grade_cell(ws4.cell(rr, 1), g)
        ws4.cell(rr, 2, desc).border = thin()
        ws4.cell(rr, 2).font = font(size=9)
        ws4.cell(rr, 2).alignment = Alignment(wrap_text=True, vertical="center")
        ws4.row_dimensions[rr].height = 30
        rr += 1

    rr += 1
    ws4.cell(rr, 1, "三、双源融合规则").font = font(bold=True, size=12, color=C["teal"])
    rr += 1
    rules = [
        "以融合分析表的街道/对口校/五维评分为主数据底座（覆盖更全）。",
        "以差异化选房地图的纪要策略标记、是否入选、政策提示为决策校准层。",
        "两侧均指向优选 →「双源一致·优选」；地图优选而上传偏低 → 上调并标注「纪要优选·上调」。",
        "地图补录独有板块（如塘桥、复旦附属九年一贯）写入总表，来源标注「差异化地图补录」。",
        "综合评分与策略契合分并存：前者看量化模型，后者看纪要策略贴合度。",
        "重要提示：本表为决策辅助，非官方划片；买房前务必核对当年招生简章与对口地段表。",
    ]
    for text in rules:
        ws4.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
        cell = ws4.cell(rr, 1, "• " + text)
        cell.font = font(size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.fill = fill(C["slate_bg"])
        ws4.row_dimensions[rr].height = 26
        rr += 1

    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 72

    # ========== Sheet5 冲突与上调说明 ==========
    ws5 = wb.create_sheet("双源校准说明")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:E1")
    ws5["A1"] = "双源校准要点（合并时发生上调/补录的板块）"
    ws5["A1"].font = Font(name="微软雅黑", bold=True, size=14, color=C["white"])
    ws5["A1"].fill = fill(C["navy"])
    ws5.row_dimensions[1].height = 28
    for c, h in enumerate(["行政区", "板块", "融合结论", "融合等级", "说明"], 1):
        ws5.cell(3, c, h)
    style_header_row(ws5, 3, 5)
    calib = [r for r in fused if r["verdict"] in ("纪要优选·上调", "纪要挂户逻辑·上调", "地图补录", "年限条件不符", "双源一致·优选")]
    # also show 双源一致 only top ones - actually list meaningful calib
    calib = [r for r in fused if any(k in r["verdict"] for k in ("上调", "补录", "年限", "双源一致"))]
    for i, r in enumerate(calib):
        row = 4 + i
        for c, v in enumerate([r["district"], r["plate"], r["verdict"], r["grade"], r["note"][:120]], 1):
            cell = ws5.cell(row, c, v)
            cell.border = thin()
            cell.font = font(size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        apply_grade_cell(ws5.cell(row, 4), r["grade"])
        ws5.row_dimensions[row].height = 36
    for i, w in enumerate([8, 18, 16, 14, 70], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    return wb


def main() -> None:
    if not UPLOAD_CANDIDATES:
        raise SystemExit("未找到上传的融合分析 xlsx")
    upload_path = UPLOAD_CANDIDATES[0]
    if not MAP_XLSX.exists():
        raise SystemExit(f"未找到地图文件: {MAP_XLSX}")

    upload_rows, legend = load_upload(upload_path)
    map_plates, districts = load_map(MAP_XLSX)
    fused = merge_rows(upload_rows, map_plates)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(fused, districts, legend)
    wb.save(OUT_XLSX)

    c = Counter(r["grade"] for r in fused)
    n_sel = sum(1 for r in fused if r["selected"] == "是")
    print(f"上传板块: {len(upload_rows)} | 地图板块: {len(map_plates)} | 融合后: {len(fused)}")
    print(f"等级分布: {dict(c)} | 入选: {n_sel}")
    print(f"已生成: {OUT_XLSX}")


if __name__ == "__main__":
    main()
