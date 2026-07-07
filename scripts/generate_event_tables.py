# -*- coding: utf-8 -*-
"""生成《首场·出海东南亚 总领事商务论坛》收费/赞助/预算/执行表（Excel）。

用法：
    python3 scripts/generate_event_tables.py [输出路径.xlsx]
"""
import sys
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import event_content as C

NAVY = "FF0B2A4A"
GOLD = "FFC9A24B"
LIGHT = "FFF1F4F8"
WHITE = "FFFFFFFF"

THIN = Side(style="thin", color="FFBFC8D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
TOTAL_FILL = PatternFill("solid", fgColor=GOLD)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
HEAD_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color=NAVY)
CELL_FONT = Font(name="Microsoft YaHei", size=10.5, color="FF1E2933")
NOTE_FONT = Font(name="Microsoft YaHei", size=10, italic=True, color="FF5B6B7B")
BOLD_CELL = Font(name="Microsoft YaHei", size=10.5, bold=True, color="FF1E2933")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_sheet(ws, tdata):
    headers = tdata["headers"]
    rows = tdata["rows"]
    ncol = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tc = ws.cell(row=1, column=1, value=tdata["title"])
    tc.font = TITLE_FONT
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    hrow = 3
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=hrow, column=ci, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.alignment = WRAP; c.border = BORDER
    ws.row_dimensions[hrow].height = 26
    for ri, row in enumerate(rows):
        r = hrow + 1 + ri
        is_total = str(row[0]) in ("合计",)
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = BOLD_CELL if (is_total or ci == 1) else CELL_FONT
            c.alignment = LEFT if ci == 1 else WRAP
            c.border = BORDER
            if is_total:
                c.fill = TOTAL_FILL
            elif ri % 2 == 1:
                c.fill = ALT_FILL
        ws.row_dimensions[r].height = 32
    note = tdata.get("note")
    if note:
        nrow = hrow + 1 + len(rows) + 1
        ws.merge_cells(start_row=nrow, start_column=1, end_row=nrow, end_column=ncol)
        nc = ws.cell(row=nrow, column=1, value="说明：" + note)
        nc.font = NOTE_FONT; nc.alignment = LEFT
        ws.row_dimensions[nrow].height = 36
    for ci in range(1, ncol + 1):
        col = get_column_letter(ci)
        maxlen = len(str(headers[ci - 1]))
        for row in rows:
            maxlen = max(maxlen, len(str(row[ci - 1])))
        ws.column_dimensions[col].width = min(52, max(12, maxlen * 1.9 + 2))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def write_overview(ws):
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value=f"{C.PROJECT_NAME}（{C.VERSION}）")
    t.font = Font(name="Microsoft YaHei", size=16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:C2")
    st = ws.cell(row=2, column=1, value=C.PROJECT_SUBTITLE + " · " + C.PROJECT_TAG)
    st.font = Font(name="Microsoft YaHei", size=11, color="FF5B6B7B")
    rows = [("工作表", "内容")] + [(t["title"].split("·")[0].strip(),
            t["title"].split("·", 1)[1].strip()) for t in C.ALL_TABLES]
    start = 4
    for ri, (a, b) in enumerate(rows):
        r = start + ri
        ca = ws.cell(row=r, column=1, value=a)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        cb = ws.cell(row=r, column=2, value=b)
        if ri == 0:
            ca.fill = HEAD_FILL; ca.font = HEAD_FONT; ca.alignment = WRAP
            cb.fill = HEAD_FILL; cb.font = HEAD_FONT; cb.alignment = WRAP
        else:
            ca.font = BOLD_CELL; ca.alignment = LEFT
            cb.font = CELL_FONT; cb.alignment = LEFT
            fill = ALT_FILL if ri % 2 else PatternFill("solid", fgColor=WHITE)
            ca.fill = fill; cb.fill = fill
        ca.border = BORDER; cb.border = BORDER
        ws.row_dimensions[r].height = 24
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.sheet_view.showGridLines = False
    nr = start + len(rows) + 1
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=3)
    nc = ws.cell(row=nr, column=1,
                 value="说明：金额均为人民币，用于测算示意，最终以协议与合同为准。")
    nc.font = NOTE_FONT; nc.alignment = LEFT


def build(path):
    wb = Workbook()
    write_overview(wb.active)
    wb.active.title = "总览"
    for i, tdata in enumerate(C.ALL_TABLES, start=1):
        safe_key = tdata['key']
        for ch in '/\\?*[]:':
            safe_key = safe_key.replace(ch, '·')
        name = f"{i:02d}·{safe_key}"[:31]
        ws = wb.create_sheet(title=name)
        write_sheet(ws, tdata)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"已生成 Excel：{path}（{len(wb.sheetnames)} 个工作表）")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else \
        "output/首场出海论坛-收费赞助与预算表.xlsx"
    build(out)
