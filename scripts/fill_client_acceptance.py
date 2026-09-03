"""生成给甲方的《绿城·潮鸣外滩》营销验收单。

以空白「基础营销验收单」为封面版式，从 2026-05-22 峰会拍立享相册
（约 425 张）中筛选绿城 / 潮鸣字样可辨的精华照片，嵌进表格。

筛选原则：
- 只收录绿城中国 / 绿城·潮鸣 / GREENTOWN 可核验露出；
- 不把美年大健康、泰隆银行、腾讯云、蔚来、长江商学院等他方展位写成绿城；
- 不把晚宴席卡姓名卡写成「潮鳴」logo 桌卡；
- 不与 2026 年 3 月开盘花束验收混用；
- 参观邀约、证书、朋友圈九宫格无成片的，黄底标待补，不拔高。
"""
from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image as PILImage

ROOT = Path("/workspace")
TEMPLATE = ROOT / "templates" / "基础营销验收单.xlsx"
PHOTO_DIR = ROOT / "deliverables" / "精华照片"
OUT = ROOT / "deliverables" / "绿城·潮鸣外滩-营销验收单（甲方交付）.xlsx"
PDF_OUT = ROOT / "deliverables" / "绿城·潮鸣外滩-营销验收单（甲方交付）.pdf"
ZIP_OUT = ROOT / "deliverables" / "下载包" / "绿城潮鸣外滩-营销验收单（甲方交付）.zip"
THUMB = Path("/tmp/acceptance-thumbs-client")

ALBUM = "https://live.pailixiang.com/album/a12523836160"
VIDEO = "https://pan.baidu.com/s/1S8cNdqnCR_anuVPD6vlhjg"
VIDEO_CODE = "8888"

FONT = "微软雅黑"
TITLE_FONT = "黑体"
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")
LEFT_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
GREEN = PatternFill("solid", fgColor="1F6B4A")
GREEN_LIGHT = PatternFill("solid", fgColor="E8F5E9")
GREEN_MID = PatternFill("solid", fgColor="C8E6C9")
AMBER = PatternFill("solid", fgColor="FFF8E1")
WHITE = PatternFill("solid", fgColor="FFFFFF")
HEADER_BAR = PatternFill("solid", fgColor="F5F5F5")
WHITE_FONT = Font(name=TITLE_FONT, size=16, bold=True, color="FFFFFF")

SHOOT_TIME = {
    "941A7785.JPG": "2026-05-22 13:20:53",
    "941A7663.JPG": "2026-05-22 12:54:38",
    "941A7769.JPG": "2026-05-22 13:17:12",
    "941A7768.JPG": "2026-05-22 13:16:51",
    "941A7776.JPG": "2026-05-22 13:18:16",
    "941A7832.JPG": "2026-05-22 13:37:45",
    "941A7825.JPG": "2026-05-22 13:36:38",
    "941A8810.JPG": "2026-05-22 19:50:04",
    "941A7777.JPG": "2026-05-22 13:19:01",
    "941A7779.JPG": "2026-05-22 13:19:08",
    "941A7782.JPG": "2026-05-22 13:19:19",
    "941A7851.JPG": "2026-05-22 13:40:52",
    "941A8857.JPG": "2026-05-22 20:25:29",
    "941A8881.JPG": "2026-05-22 20:32:54",
    "941A8859.JPG": "2026-05-22 20:26:10",
    "941A7849.JPG": "2026-05-22 13:40:40",
    "941A7836.JPG": "2026-05-22 13:39:31",
    "00_banner.jpg": "主视觉 KV（相册 Banner）",
}

# 首页 9 张：均为绿城 / 潮鸣可核验露出
COVER = [
    ("941A7785.JPG", "①议程看板【绿城·潮鸣】晚宴冠名", "议程看板：【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴"),
    ("941A7663.JPG", "②主背景板：晚宴冠名·绿城中国", "主背景板全幅：晚宴冠名战略合作伙伴·绿城中国"),
    ("941A7769.JPG", "③展台立牌：绿城中国 GREENTOWN", "主会场展台立牌：绿城中国 / GREENTOWN 可辨"),
    ("941A7768.JPG", "④展台接待：绿城礼袋与物料", "展台接待：绿城立牌、礼袋与纸质物料，嘉宾问询"),
    ("941A7776.JPG", "⑤主背景板合影核验冠名位", "主背景板合影：晚宴冠名位「绿城中国」可核验"),
    ("941A7832.JPG", "⑥主会场双屏（同款主视觉）", "主会场双屏主视觉与背景板同款，含晚宴冠名·绿城中国"),
    ("941A7825.JPG", "⑦主持现场：主会场露出", "主持环节主会场全幅露出，便于核验现场执行"),
    ("941A8810.JPG", "⑧晚宴大屏：绿城中国品牌片", "晚宴弧形大屏播放绿城中国「理想生活综合服务商」"),
    ("00_banner.jpg", "⑨主视觉 KV：战略合作含绿城", "主视觉 KV：战略合作伙伴名单含绿城中国"),
]

EXTRA = [
    ("941A7777.JPG", "主背景板合影（晚宴冠名位特写）"),
    ("941A7779.JPG", "主背景板双人合影核验冠名位"),
    ("941A7782.JPG", "主背景板合影：晚宴冠名战略合作伙伴·绿城中国"),
    ("941A7851.JPG", "主办致辞全景：主会场双屏与讲台"),
    ("941A7836.JPG", "主办致辞：姚志勇，主会场主视觉"),
    ("941A7849.JPG", "主办致辞讲台特写"),
    ("941A8857.JPG", "冠名晚宴合影（江景厅）"),
    ("941A8881.JPG", "冠名晚宴桌次（江景厅圆桌）"),
    ("941A8859.JPG", "冠名晚宴现场举杯"),
]

MODULE_PHOTOS = [
    (
        "1. 晚宴冠名和专属权益　¥55,000　☑ 已执行（冠名 / 品牌片）",
        [
            ("941A7785.JPG", "议程看板【绿城·潮鸣】晚宴"),
            ("941A7663.JPG", "主背景板：晚宴冠名·绿城中国"),
            ("941A8810.JPG", "晚宴大屏绿城中国品牌片"),
            ("941A8857.JPG", "冠名晚宴合影（江景厅）"),
            ("941A8859.JPG", "冠名晚宴现场举杯"),
            ("941A8881.JPG", "冠名晚宴桌次（江景厅）"),
        ],
    ),
    (
        "2. 主会场权益　¥30,000　☑ 已执行（展台 / 双屏）",
        [
            ("941A7769.JPG", "展台立牌：绿城中国 GREENTOWN"),
            ("941A7768.JPG", "展台接待：礼袋与物料"),
            ("941A7832.JPG", "主会场双屏同款主视觉"),
            ("00_banner.jpg", "主视觉 KV 含绿城中国"),
            ("941A7825.JPG", "主持现场"),
            ("941A7851.JPG", "主办致辞全景"),
            ("941A7836.JPG", "主办致辞：姚志勇"),
            ("941A7849.JPG", "主办致辞讲台特写"),
        ],
    ),
    (
        "3. 专场项目参观邀约　¥5,000　☐ 待录像核验（无专项成片）",
        [
            ("941A7825.JPG", "主持口播位已具备"),
            ("941A7776.JPG", "主背景板冠名位（非参观成片）"),
        ],
    ),
    (
        "4. 宣发配合　¥10,000　☑ 现场摄影已交　☐ 朋友圈 / 回顾视频待补",
        [
            ("941A7777.JPG", "主背景板合影（执行回执用）"),
            ("941A7779.JPG", "主背景板双人合影核验冠名"),
            ("941A7782.JPG", "主背景板合影核验冠名"),
            ("941A8881.JPG", "冠名晚宴桌次"),
        ],
    ),
]

# 全部精华照片按重要顺序分类（一级最重要）
RANKED = [
    {
        "level": "一级",
        "name": "核心品牌露出（必看）",
        "desc": "绿城 / 潮鸣 / GREENTOWN 字样可直接核验，对应报价主体，建议甲方优先复核。",
        "photos": [
            ("941A7785.JPG", "议程看板【绿城·潮鸣】星耀北外滩晚宴", "晚宴冠名"),
            ("941A7663.JPG", "主背景板：晚宴冠名战略合作伙伴·绿城中国", "晚宴冠名"),
            ("941A7769.JPG", "展台立牌：绿城中国 GREENTOWN", "主会场"),
            ("941A8810.JPG", "晚宴大屏播放绿城中国品牌片", "晚宴冠名"),
        ],
    },
    {
        "level": "二级",
        "name": "主会场执行（重点）",
        "desc": "展台接待、双屏、主持、主视觉，证明主会场权益已落地。",
        "photos": [
            ("941A7768.JPG", "展台接待：绿城立牌、礼袋与物料", "主会场"),
            ("941A7832.JPG", "主会场双屏同款主视觉（含晚宴冠名·绿城中国）", "主会场"),
            ("941A7825.JPG", "主持现场：主会场露出", "主会场"),
            ("00_banner.jpg", "主视觉 KV：战略合作名单含绿城中国", "主会场/宣发"),
        ],
    },
    {
        "level": "三级",
        "name": "冠名位合影核验",
        "desc": "主背景板合影，可复核晚宴冠名位「绿城中国」。",
        "photos": [
            ("941A7776.JPG", "主背景板合影核验冠名位", "晚宴冠名"),
            ("941A7777.JPG", "主背景板合影（冠名位特写）", "晚宴冠名"),
            ("941A7779.JPG", "主背景板双人合影核验冠名位", "晚宴冠名"),
            ("941A7782.JPG", "主背景板合影：晚宴冠名战略合作伙伴·绿城中国", "晚宴冠名"),
        ],
    },
    {
        "level": "四级",
        "name": "主办致辞现场",
        "desc": "主会场致辞与讲台，补充现场执行；不作参观邀约专项成片。",
        "photos": [
            ("941A7851.JPG", "主办致辞全景：主会场双屏与讲台", "主会场"),
            ("941A7836.JPG", "主办致辞：姚志勇，主会场主视觉", "主会场"),
            ("941A7849.JPG", "主办致辞讲台特写", "主会场"),
        ],
    },
    {
        "level": "五级",
        "name": "冠名晚宴氛围",
        "desc": "江景厅冠名晚宴现场。席卡为嘉宾姓名卡，不作「潮鳴」logo 桌卡证据。",
        "photos": [
            ("941A8857.JPG", "冠名晚宴合影（江景厅）", "晚宴冠名"),
            ("941A8881.JPG", "冠名晚宴桌次（江景厅圆桌）", "晚宴冠名"),
            ("941A8859.JPG", "冠名晚宴现场举杯", "晚宴冠名"),
        ],
    },
]


def all_ranked_photos():
    items = []
    for cat in RANKED:
        for photo in cat["photos"]:
            items.append((cat, photo))
    return items


def place_photo_grid(
    ws, start_row, photos, tag: str, img_h=118, img_w=168, paint_to_col=4
) -> int:
    """把照片按每行 4 张嵌进表格，返回下一空行。"""
    letters = ("A", "B", "C", "D")
    row = start_row
    for start in range(0, len(photos), 4):
        chunk = photos[start : start + 4]
        img_row = row
        cap_row = row + 1
        ws.row_dimensions[img_row].height = img_h
        ws.row_dimensions[cap_row].height = 36
        paint_range(ws, img_row, cap_row, 1, paint_to_col, WHITE)
        for i, item in enumerate(chunk):
            name, cap = item[0], item[1]
            t = thumb(resolve_src(name), THUMB / f"{tag}_{name}", max_w=420, max_h=300)
            place_image(ws, t, f"{letters[i]}{img_row}", img_w, img_h)
            fill_cell(ws, f"{letters[i]}{cap_row}", cap, 8, align=LEFT)
        row = cap_row + 1
    return row


def resolve_src(name: str) -> Path:
    p = PHOTO_DIR / name
    if p.exists() and p.stat().st_size > 2000:
        return p
    raise FileNotFoundError(name)


def thumb(src: Path, dest: Path, max_w=520, max_h=390, quality=86) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    im = PILImage.open(src)
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    im.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
    dest = dest.with_suffix(".jpg")
    im.save(dest, "JPEG", quality=quality, optimize=True)
    return dest


def place_image(ws, path: Path, cell: str, width: int, height: int) -> None:
    img = XLImage(str(path))
    img.width = width
    img.height = height
    img.anchor = cell
    ws.add_image(img)


def fill_cell(ws, coord, text, size=10, bold=False, align=WRAP, fill=None, font_name=FONT):
    cell = ws[coord]
    cell.value = text
    cell.font = Font(name=font_name, size=size, bold=bold)
    cell.alignment = align
    cell.border = THIN
    if fill is not None:
        cell.fill = fill


def paint_range(ws, start_row, end_row, start_col, end_col, fill=WHITE):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            if fill is not None:
                cell.fill = fill


def unmerge_if(ws, coord: str) -> None:
    hits = [str(rng) for rng in list(ws.merged_cells.ranges) if coord in rng]
    for rng in hits:
        ws.unmerge_cells(rng)


def setup_print(ws, area: str, landscape: bool = False, fit_height: int = 1) -> None:
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = area
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.55, bottom=0.55, header=0.2, footer=0.2
    )


def fill_cover(ws) -> None:
    ws.title = "营销验收单"
    fill_cell(ws, "B3", "晚宴冠名战略合作伙伴\n（绿城中国 | 绿城·潮鸣外滩）", 11, True)
    fill_cell(ws, "D3", "人民币 100,000 元整\n（含税 · 四模块整体打包）", 11, True)
    fill_cell(
        ws,
        "B4",
        "主办：上海市杨浦区科技企业联合会\n赞助：上海绿城泓盛建设发展有限公司\n项目品牌：绿城中国 · 绿城·潮鸣外滩",
        10,
    )
    fill_cell(ws, "D4", "2026年5月22日\n13:00 – 20:30\n上海·北外滩·一滴水", 11, True)
    fill_cell(
        ws,
        "B5",
        "【活动】第四届 / 2026人工智能商业化落地与硬核投资破局峰会。\n"
        "【身份】晚宴冠名战略合作伙伴。\n"
        "【晚宴】【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴。\n"
        "【四模块】①晚宴冠名和专属权益 ¥55,000；②主会场权益 ¥30,000；"
        "③专场项目参观邀约 ¥5,000；④宣发配合 ¥10,000。含税合计 ¥100,000，整体打包验收。\n"
        "【现场已核验】主背景板「晚宴冠名战略合作伙伴·绿城中国」、议程看板晚宴冠名【绿城·潮鸣】、"
        "展台立牌「绿城中国 GREENTOWN」、主会场双屏同款主视觉、晚宴大屏绿城中国品牌片。\n"
        "【证书】事项载明颁发“2026年度智慧人居新质资产领军企业 暨卓越战略合作伙伴”；"
        "本相册未见证书特写，不以照片推定已颁。",
        8,
        align=LEFT_TOP,
    )
    ws.row_dimensions[5].height = 118

    fill_cell(
        ws,
        "B13",
        "【露出】本页 9 张为封面摘要；全部 18 张已按一级至五级嵌入「按重要顺序分类」。\n"
        "【营销评估】非常好（☐）　良好（☑）　一般（☐）　较差（☐）\n"
        "【打包验收】四模块整体打包。参观邀约无单独成片、朋友圈九宫格 / 回顾视频未附本表，"
        "见「权益核验清单」黄底项。已核验露出效果良好，符合约定要求。",
        9,
        align=LEFT_TOP,
    )
    ws.row_dimensions[13].height = 52
    ws.row_dimensions[14].height = 18
    ws.row_dimensions[15].height = 18
    ws.row_dimensions[14].hidden = True
    ws.row_dimensions[15].hidden = True

    fill_cell(
        ws,
        "B16",
        "①本页 9 张为封面摘要，含图注；"
        "②「按重要顺序分类」将 18 张精华图全部按五级嵌入表格；"
        "③「权益核验清单」按报价单 4 模块嵌全部对应照片；"
        "④「绿城·潮鸣外滩露出」为大图（同序分级）；"
        "⑤「效果评估表」按绿城中国策划活动类效果评估表结构；"
        "⑥「赞助权益执行回执」供绿城复核盖章；"
        "⑦「照片索引」含重要等级、拍摄时间与模块对照；"
        f"⑧拍立享直播 {ALBUM}（相册约 425 张，本表为精华筛选）；"
        f"⑨全程录像 {VIDEO} 提取码 {VIDEO_CODE}。"
        "本表不与 2026年3月开盘花束验收混用。"
        "未把美年大健康 / 泰隆银行 / 腾讯云 / 蔚来展位当作绿城露出。",
        8,
        align=LEFT,
    )
    ws.row_dimensions[16].height = 108
    fill_cell(ws, "C17", "胡继刚\n13262607888\n（主办执行）\n签字：____________", 10, align=LEFT)
    fill_cell(
        ws,
        "C18",
        "绿城对接：笪浩 18678408669\n营销负责人签字：____________\n日期：____________",
        10,
        align=LEFT,
    )

    unmerge_if(ws, "B6")
    paint_range(ws, 6, 12, 2, 4, WHITE)
    ws.merge_cells("B12:D12")
    fill_cell(
        ws,
        "B12",
        "图①–⑨对应「权益核验清单」。他方品牌展位、晚宴席卡姓名卡不作为潮鸣 logo 证据。",
        8,
        align=LEFT,
    )

    photo_rows = (6, 8, 10)
    cap_rows = (7, 9, 11)
    cols = ("B", "C", "D")
    for i, (name, short, _full) in enumerate(COVER):
        r, c = divmod(i, 3)
        fill_cell(ws, f"{cols[c]}{cap_rows[r]}", short, 7, align=LEFT)
        t = thumb(resolve_src(name), THUMB / f"cover_{name}", max_w=440, max_h=320)
        place_image(ws, t, f"{cols[c]}{photo_rows[r]}", 172, 112)

    for r, h in [(6, 90), (7, 34), (8, 90), (9, 34), (10, 90), (11, 34), (12, 22)]:
        ws.row_dimensions[r].height = h

    ws["A20"] = "照片直播（拍立享）"
    ws["A20"].font = Font(name=FONT, size=9, bold=True)
    ws["B20"] = ALBUM
    ws["B20"].hyperlink = ALBUM
    ws["B20"].font = Font(name=FONT, size=9, color="0563C1", underline="single")
    ws.merge_cells("B20:D20")
    ws["A21"] = "全程录像（百度网盘）"
    ws["A21"].font = Font(name=FONT, size=9, bold=True)
    ws["B21"] = f"{VIDEO}  提取码：{VIDEO_CODE}"
    ws["B21"].hyperlink = VIDEO
    ws["B21"].font = Font(name=FONT, size=9, color="0563C1", underline="single")
    ws.merge_cells("B21:D21")

    setup_print(ws, "A2:D18", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_properties.tabColor = "1F6B4A"


def add_ranked_sheet(wb) -> None:
    ws = wb.create_sheet("按重要顺序分类", 1)
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 28

    total = len(all_ranked_photos())
    ws.merge_cells("A1:D1")
    fill_cell(
        ws,
        "A1",
        f"绿城·潮鸣外滩 · 精华照片按重要顺序分类（{total} 张全部嵌入本表）",
        16,
        True,
        WRAP,
        GREEN,
        TITLE_FONT,
    )
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        "排序：一级核心品牌露出 → 二级主会场执行 → 三级冠名位合影 → 四级主办致辞 → 五级晚宴氛围。"
        "只收录绿城 / 潮鸣 / GREENTOWN 可核验露出。本页表格已嵌入全部精华原图，可直接给甲方按等级复核。",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws.row_dimensions[2].height = 36

    headers = ["重要等级", "分类", "张数", "核验要点"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[3].height = 22

    banner_fill = {
        "一级": GREEN,
        "二级": GREEN_MID,
        "三级": GREEN_LIGHT,
        "四级": HEADER_BAR,
        "五级": WHITE,
    }
    for i, cat in enumerate(RANKED, 4):
        vals = (cat["level"], cat["name"], str(len(cat["photos"])), cat["desc"])
        fill = banner_fill.get(cat["level"], WHITE)
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(
                name=FONT,
                size=10,
                bold=(col <= 3),
                color="FFFFFF" if cat["level"] == "一级" else "000000",
            )
            cell.alignment = LEFT if col == 4 else WRAP
            cell.border = THIN
            cell.fill = fill
        ws.row_dimensions[i].height = 36

    sum_row = 4 + len(RANKED)
    ws.merge_cells(f"A{sum_row}:D{sum_row}")
    fill_cell(
        ws,
        f"A{sum_row}",
        f"合计 {total} 张，全部按上表顺序嵌在下方各分类单元格中。参观邀约 / 证书 / 朋友圈无成片，不在本页拔高。",
        9,
        True,
        LEFT,
        GREEN_LIGHT,
    )
    paint_range(ws, sum_row, sum_row, 1, 4, GREEN_LIGHT)
    ws.row_dimensions[sum_row].height = 24

    row = sum_row + 2
    for idx, cat in enumerate(RANKED, 1):
        n = len(cat["photos"])
        fill = banner_fill.get(cat["level"], GREEN_MID)
        ws.merge_cells(f"A{row}:D{row}")
        fill_cell(
            ws,
            f"A{row}",
            f"{cat['level']}｜{cat['name']}（{n} 张）",
            13,
            True,
            LEFT,
            fill,
            TITLE_FONT,
        )
        paint_range(ws, row, row, 1, 4, fill)
        if cat["level"] == "一级":
            ws[f"A{row}"].font = Font(name=TITLE_FONT, size=13, bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = GREEN
        ws.row_dimensions[row].height = 26
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        fill_cell(ws, f"A{row}", cat["desc"], 9, False, LEFT, GREEN_LIGHT)
        paint_range(ws, row, row, 1, 4, GREEN_LIGHT)
        ws.row_dimensions[row].height = 22
        row += 1
        photos = [(p[0], p[1]) for p in cat["photos"]]
        row = place_photo_grid(
            ws,
            row,
            photos,
            tag=f"rank{idx}",
            img_h=132,
            img_w=186,
            paint_to_col=4,
        )
        row += 1

    last = row - 1
    setup_print(ws, f"A1:D{last}", landscape=True, fit_height=0)
    ws.print_title_rows = "1:2"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F6B4A"


def add_rights_sheet(wb) -> None:
    ws = wb.create_sheet("权益核验清单", 2)
    widths = [6, 18, 36, 12, 42, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    fill_cell(ws, "A1", "绿城·潮鸣外滩 · 晚宴冠名四模块权益核验清单", 16, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 6, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    fill_cell(
        ws,
        "A2",
        "活动：2026-05-22 上海·北外滩·一滴水　身份：晚宴冠名战略合作伙伴　"
        "金额：含税 ¥100,000（报价单 4 模块整体打包，不拆子项验收）　"
        "依据：绿城中国-潮鸣外滩-10万元晚宴冠名服务报价单　"
        "本页按四模块嵌全部对应照片；若按重要程度查阅，见工作表「按重要顺序分类」。",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 6, GREEN_MID)
    ws.row_dimensions[2].height = 36

    headers = ["序号", "服务模块", "约定内容", "金额（元）", "现场照片核验", "结论"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[3].height = 22

    rows = [
        (
            "1",
            "晚宴冠名和专属权益",
            "晚宴主题冠名、专场PPT/视频、主持人口播、席卡等物料植入",
            "55,000",
            "图①议程看板晚宴冠名【绿城·潮鸣】；图②⑤主背景板「晚宴冠名战略合作伙伴·绿城中国」；"
            "图⑧晚宴大屏播放绿城中国品牌片（理想生活综合服务商 / LIGHT HEAT POWER）。"
            "现场席卡为嘉宾姓名卡，相册未见单独「潮鳴」logo 桌卡特写。",
            "通过（冠名/品牌片）",
        ),
        (
            "2",
            "主会场权益",
            "1号位展台、手拎袋项目物料、视频轮播/奖项颁发画面",
            "30,000",
            "图③展台立牌「绿城中国 GREENTOWN」；图④展台接待（礼袋与物料）；"
            "图⑥主会场双屏同款主视觉；图⑨主视觉KV含绿城中国。奖项颁发证书未见特写。",
            "通过（展台/双屏）",
        ),
        (
            "3",
            "专场项目参观邀约",
            "论坛/颁奖结束后主持人口播 + 动线引导，邀约前往案场",
            "5,000",
            "图⑦证明主会场主持人口播位已具备。相册未见「参观邀约 / 案场接驳」专项成片，"
            "请以全程录像及现场执行回执为准。",
            "待录像核验",
        ),
        (
            "4",
            "宣发配合",
            "现场摄影、朋友圈九宫格、回顾视频项目鸣谢、执行回执",
            "10,000",
            "现场摄影：本表精华照片 + 拍立享约 425 张 + 网盘录像。本工作簿「赞助权益执行回执」即执行报告。"
            "朋友圈九宫格、回顾视频片头鸣谢未附，不以照片推定已发。",
            "摄影通过；其余待补",
        ),
        (
            "—",
            "合计（含税）",
            "4 模块整体打包，一次性验收",
            "100,000",
            "已核验项以绿城 / 潮鸣字样可辨露出为准；黄底项需录像或另行回执，不在本表拔高为已完成。",
            "主体已交，部分待补",
        ),
        (
            "—",
            "证书",
            "2026年度智慧人居新质资产领军企业 暨卓越战略合作伙伴",
            "—",
            "事项描述已载明颁发。本相册未见证书 / 颁奖特写，不作为已交付证据。",
            "待补照片",
        ),
    ]
    fills = {
        "通过（冠名/品牌片）": GREEN_LIGHT,
        "通过（展台/双屏）": GREEN_LIGHT,
        "待录像核验": AMBER,
        "摄影通过；其余待补": AMBER,
        "主体已交，部分待补": AMBER,
        "待补照片": AMBER,
    }
    for i, vals in enumerate(rows, 4):
        result = vals[-1]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(name=FONT, size=9, bold=(col in (2, 4, 6)))
            cell.alignment = LEFT if col in (3, 5) else WRAP
            cell.border = THIN
            cell.fill = fills.get(result, WHITE) if col == 6 else WHITE
        ws.row_dimensions[i].height = 76

    note = 10
    ws.merge_cells(f"A{note}:F{note}")
    fill_cell(
        ws,
        f"A{note}",
        "验收口径：报价单约定「不逐项贴照片、四模块整体打包」。本清单把精华照片映射到模块，便于甲方复核，"
        "不等于把未成片的子项改成已交付。营销评估预勾「良好」，请绿城营销负责人复核签字。"
        "本清单仅对应 2026-05-22 峰会，不与 3 月开盘花束验收混用。",
        9,
        False,
        LEFT,
        GREEN_LIGHT,
    )
    paint_range(ws, note, note, 1, 6, GREEN_LIGHT)
    ws.row_dimensions[note].height = 48

    ev_title = note + 2
    ws.merge_cells(f"A{ev_title}:F{ev_title}")
    fill_cell(
        ws,
        f"A{ev_title}",
        "四模块现场证据照片（各模块全部对应图已嵌入，可直接给甲方核对）",
        12,
        True,
        LEFT,
        GREEN,
        TITLE_FONT,
    )
    paint_range(ws, ev_title, ev_title, 1, 6, GREEN)
    ws[f"A{ev_title}"].font = Font(name=TITLE_FONT, size=12, bold=True, color="FFFFFF")
    ws[f"A{ev_title}"].fill = GREEN
    ws.row_dimensions[ev_title].height = 24

    row = ev_title + 1
    for title, photos in MODULE_PHOTOS:
        ws.merge_cells(f"A{row}:F{row}")
        fill_cell(ws, f"A{row}", f"{title}　本模块 {len(photos)} 张（全部嵌入）", 11, True, LEFT, GREEN_MID)
        paint_range(ws, row, row, 1, 6, GREEN_MID)
        ws.row_dimensions[row].height = 22
        row = place_photo_grid(
            ws,
            row + 1,
            photos,
            tag=f"mod{title[:1]}",
            img_h=118,
            img_w=168,
            paint_to_col=6,
        )
        if "待录像" in title:
            ws.merge_cells(f"A{row}:F{row}")
            fill_cell(
                ws,
                f"A{row}",
                "相册未见参观 / 接驳专项成片。上图仅证明主持口播位与冠名位，不作为参观邀约已完成证据。",
                8,
                align=LEFT,
                fill=AMBER,
            )
            paint_range(ws, row, row, 1, 6, AMBER)
            ws.row_dimensions[row].height = 28
            row += 1

    last = row - 1
    setup_print(ws, f"A1:F{last}", landscape=True, fit_height=0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FFF8E1"


def add_photo_sheet(wb) -> None:
    ws = wb.create_sheet("绿城·潮鸣外滩露出")
    ws.merge_cells("A1:D1")
    fill_cell(
        ws,
        "A1",
        "绿城中国｜绿城·潮鸣外滩 现场露出精华照片（按重要顺序 · 甲方交付）",
        14,
        True,
        WRAP,
        GREEN,
        TITLE_FONT,
    )
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 28

    total = len(all_ranked_photos())
    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        f"活动时间：2026-05-22　地点：上海·北外滩·一滴水　"
        f"本页 {total} 张按一级→五级排列（与「按重要顺序分类」同序）　"
        f"完整相册约 425 张：{ALBUM}　录像：{VIDEO}（提取码 {VIDEO_CODE}）",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws["A2"].hyperlink = ALBUM
    ws.row_dimensions[2].height = 36
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 42

    banner_fill = {
        "一级": GREEN,
        "二级": GREEN_MID,
        "三级": GREEN_LIGHT,
        "四级": HEADER_BAR,
        "五级": WHITE,
    }
    row = 4
    seq = 1
    col_pair = [("A", "B"), ("C", "D")]
    for cat in RANKED:
        fill = banner_fill.get(cat["level"], GREEN_MID)
        ws.merge_cells(f"A{row}:D{row}")
        fill_cell(
            ws,
            f"A{row}",
            f"{cat['level']}｜{cat['name']}（{len(cat['photos'])} 张）　{cat['desc']}",
            11,
            True,
            LEFT,
            fill,
        )
        paint_range(ws, row, row, 1, 4, fill)
        if cat["level"] == "一级":
            ws[f"A{row}"].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = GREEN
        ws.row_dimensions[row].height = 28
        row += 1

        photos = [(p[0], p[1]) for p in cat["photos"]]
        i = 0
        while i < len(photos):
            ws.row_dimensions[row].height = 168
            ws.row_dimensions[row + 1].height = 36
            paint_range(ws, row, row + 1, 1, 4, WHITE)
            for j, (img_col, cap_col) in enumerate(col_pair):
                if i + j >= len(photos):
                    break
                name, cap = photos[i + j]
                t = thumb(resolve_src(name), THUMB / f"sheet_{name}", max_w=640, max_h=480)
                place_image(ws, t, f"{img_col}{row}", 290, 205)
                cell = ws[f"{img_col}{row + 1}"]
                ws.merge_cells(f"{img_col}{row + 1}:{cap_col}{row + 1}")
                cell.value = f"{seq + j}. {cap}"
                cell.font = Font(name=FONT, size=9)
                cell.alignment = LEFT
                cell.border = THIN
            seq += min(2, len(photos) - i)
            i += 2
            row += 2

    note_row = row + 1
    ws.merge_cells(f"A{note_row}:D{note_row}")
    fill_cell(
        ws,
        f"A{note_row}",
        "说明：本页只收录绿城中国、绿城·潮鸣、GREENTOWN 可辨露出及冠名晚宴现场。"
        "未收录美年大健康、泰隆银行、腾讯云、蔚来、长江商学院等他方展位。"
        "完整 425 张以拍立享为准；MP4 以网盘「5.22」为准。不与开盘花束混用。",
        9,
        False,
        LEFT,
        WHITE,
    )
    paint_range(ws, note_row, note_row, 1, 4, WHITE)
    ws.row_dimensions[note_row].height = 42
    setup_print(ws, f"A1:D{note_row}", landscape=True, fit_height=0)
    ws.print_title_rows = "1:2"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F6B4A"


def add_eval_sheet(wb) -> None:
    ws = wb.create_sheet("效果评估表", 3)
    for col, width in enumerate([16, 24, 24, 36], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.merge_cells("A1:D1")
    fill_cell(ws, "A1", "绿城中国策划活动类效果评估表", 16, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        "（冠名、活动赞助、活动执行等）　项目：绿城·潮鸣外滩　活动：2026-05-22 晚宴冠名",
        10,
        False,
        WRAP,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws.row_dimensions[2].height = 22

    fill_cell(ws, "A3", "活动主题", 11, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B3:D3")
    fill_cell(ws, "B3", "【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴", 13, True)
    paint_range(ws, 3, 3, 1, 4, WHITE)
    ws["A3"].fill = GREEN_LIGHT
    ws.row_dimensions[3].height = 28

    fill_cell(ws, "A4", "活动举办时间", 11, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B4:D4")
    fill_cell(ws, "B4", "2026年5月22日 13:00–20:30　　上海·北外滩·一滴水", 12, True)
    paint_range(ws, 4, 4, 1, 4, WHITE)
    ws["A4"].fill = GREEN_LIGHT
    ws.row_dimensions[4].height = 28

    ws.merge_cells("A5:A12")
    fill_cell(ws, "A5", "活动\n现场氛围", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 5, 12, 1, 4, WHITE)
    ws["A5"].fill = GREEN_LIGHT
    eval_photos = [COVER[0], COVER[1], COVER[2], COVER[7]]
    caps = [p[1] for p in eval_photos]
    ws.merge_cells("B5:C5")
    fill_cell(ws, "B5", caps[0], 8, align=LEFT)
    fill_cell(ws, "D5", caps[1], 8, align=LEFT)
    ws.merge_cells("B6:C8")
    ws.merge_cells("D6:D8")
    ws.merge_cells("B9:C9")
    fill_cell(ws, "B9", caps[2], 8, align=LEFT)
    fill_cell(ws, "D9", caps[3], 8, align=LEFT)
    ws.merge_cells("B10:C12")
    ws.merge_cells("D10:D12")
    for r, h in [(5, 28), (6, 78), (7, 78), (8, 78), (9, 28), (10, 78), (11, 78), (12, 78)]:
        ws.row_dimensions[r].height = h
    anchors = [("B6", 340, 280), ("D6", 250, 280), ("B10", 340, 280), ("D10", 250, 280)]
    for (name, _s, _f), (cell, w, h) in zip(eval_photos, anchors):
        t = thumb(resolve_src(name), THUMB / f"eval_{name}", max_w=800, max_h=600)
        place_image(ws, t, cell, w, h)

    ws.merge_cells("A13:A16")
    fill_cell(ws, "A13", "策划负责人填写", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 13, 16, 1, 4, WHITE)
    ws["A13"].fill = GREEN_LIGHT
    ws.merge_cells("B13:D13")
    fill_cell(ws, "B13", "总体活动评价：", 11, True, LEFT, GREEN_MID)
    ws.merge_cells("B14:D15")
    fill_cell(
        ws,
        "B14",
        "本次活动为「2026人工智能商业化落地与硬核投资破局峰会」晚宴冠名。"
        "绿城中国 | 绿城·潮鸣外滩作为晚宴冠名战略合作伙伴，主背景板、议程看板、"
        "展台立牌、主会场双屏及晚宴大屏品牌片均完成品牌露出。"
        "按照约定完成整体策划与执行，现场氛围良好，活动取得圆满成功。",
        11,
        False,
        LEFT_TOP,
        WHITE,
    )
    ws.merge_cells("B16:D16")
    fill_cell(
        ws,
        "B16",
        "策划负责人签名：胡继刚 ____________________　　日期：2026-05-22　　（请补签）",
        10,
        False,
        LEFT,
        HEADER_BAR,
    )
    ws.row_dimensions[13].height = 24
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 36
    ws.row_dimensions[16].height = 32

    ws.merge_cells("A17:A21")
    fill_cell(ws, "A17", "营销负责人填写", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 17, 21, 1, 4, WHITE)
    ws["A17"].fill = GREEN_LIGHT
    ws.merge_cells("B17:D17")
    fill_cell(ws, "B17", "对活动【总体效果】评估（文字说明）：", 11, True, LEFT, GREEN_MID)
    ws.merge_cells("B18:D18")
    fill_cell(
        ws,
        "B18",
        "绿城中国 / 潮鸣字样可辨露出已核验，效果良好，符合约定要求。"
        "参观邀约口播、朋友圈九宫格、回顾视频、证书特写见执行回执待补项，不在此拔高。",
        11,
        False,
        LEFT,
        WHITE,
    )
    ws.merge_cells("B19:D19")
    fill_cell(
        ws,
        "B19",
        "非常好（☐）　　良好（☑）　　一般（☐）　　较差（☐）",
        13,
        True,
        WRAP,
        AMBER,
    )
    ws.merge_cells("B20:D20")
    fill_cell(
        ws,
        "B20",
        "勾选说明：按《基础营销验收单》模板默认口径「效果良好，符合要求」预勾「良好」，"
        "不拔高为「非常好」。请绿城营销负责人复核。",
        9,
        False,
        LEFT,
        AMBER,
    )
    ws.merge_cells("B21:D21")
    fill_cell(
        ws,
        "B21",
        "营销负责人签名：____________________　　日期：____________________"
        "　　对接：笪浩 18678408669（评估表请补签）",
        10,
        False,
        LEFT,
        HEADER_BAR,
    )
    ws.row_dimensions[17].height = 24
    ws.row_dimensions[18].height = 42
    ws.row_dimensions[19].height = 28
    ws.row_dimensions[20].height = 42
    ws.row_dimensions[21].height = 32

    ws.merge_cells("A22:D22")
    fill_cell(
        ws,
        "A22",
        "详见附件：本工作簿「按重要顺序分类」（18 张分级）「营销验收单」「权益核验清单」「赞助权益执行回执」及拍立享相册、网盘录像。"
        "合作金额含税 ¥100,000，四模块整体打包。本表仅对应 5 月 22 日峰会。",
        9,
        False,
        LEFT,
        WHITE,
    )
    paint_range(ws, 22, 22, 1, 4, WHITE)
    ws.row_dimensions[22].height = 36
    setup_print(ws, "A1:D22", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C8E6C9"


def add_receipt_sheet(wb) -> None:
    ws = wb.create_sheet("赞助权益执行回执", 4)
    for col, w in enumerate([22, 28, 22, 28], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("A1:D1")
    fill_cell(ws, "A1", "赞助权益执行回执", 18, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        "依据《绿城中国-潮鸣外滩-10万元晚宴冠名服务报价单》：大会结束后 7 个自然日内，"
        "由主办向绿城一次性出具本回执，作为四模块整体验收依据。",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws.row_dimensions[2].height = 32

    pairs = [
        ("出具方（主办）", "上海市杨浦区科技企业联合会", "接收方（赞助）", "上海绿城泓盛建设发展有限公司"),
        ("对接人", "胡继刚　13262607888", "对接人", "笪浩　18678408669"),
        ("活动名称", "2026人工智能商业化落地与硬核投资破局峰会", "活动时间/地点", "2026-05-22　上海·北外滩·一滴水"),
        ("合作身份", "晚宴冠名战略合作伙伴", "合作金额", "人民币 100,000 元整（含税）"),
        ("晚宴场次", "【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴", "项目品牌", "绿城中国 · 绿城·潮鸣外滩"),
    ]
    r = 3
    for a, b, c, d in pairs:
        fill_cell(ws, f"A{r}", a, 10, True, WRAP, GREEN_LIGHT)
        fill_cell(ws, f"B{r}", b, 10, False, LEFT, WHITE)
        fill_cell(ws, f"C{r}", c, 10, True, WRAP, GREEN_LIGHT)
        fill_cell(ws, f"D{r}", d, 10, False, LEFT, WHITE)
        ws.row_dimensions[r].height = 28
        r += 1

    ws.merge_cells("A8:D8")
    fill_cell(ws, "A8", "四模块执行确认（整体打包，不拆子项计价）", 11, True, LEFT, GREEN)
    paint_range(ws, 8, 8, 1, 4, GREEN)
    ws["A8"].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws["A8"].fill = GREEN
    ws.row_dimensions[8].height = 24

    headers = ["模块", "金额（元）", "执行结论", "代表性证据"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(9, i, h)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[9].height = 22

    mods = [
        (
            "1. 晚宴冠名和专属权益",
            "55,000",
            "☑ 已执行（冠名/品牌片）",
            "议程看板【绿城·潮鸣】晚宴；主背景板「晚宴冠名战略合作伙伴·绿城中国」；晚宴大屏绿城中国品牌片",
        ),
        (
            "2. 主会场权益",
            "30,000",
            "☑ 已执行（展台/双屏）",
            "展台立牌「绿城中国 GREENTOWN」；展台接待礼袋物料；主会场双屏同款主视觉；主视觉KV含绿城中国",
        ),
        (
            "3. 专场项目参观邀约",
            "5,000",
            "☐ 待录像核验",
            "主持人口播位已具备。相册未见参观/接驳专项成片，请以全程录像确认是否口播执行。",
        ),
        (
            "4. 宣发配合",
            "10,000",
            "☑ 摄影已交　☐ 其余待补",
            "现场摄影+本回执。朋友圈九宫格、回顾视频片头鸣谢未附本回执。",
        ),
        (
            "合计（含税）",
            "100,000",
            "主体已交，部分待补",
            "四模块整体打包验收。已核验项以绿城/潮鸣字样可辨露出为准。",
        ),
    ]
    fills = {
        "☑ 已执行（冠名/品牌片）": GREEN_LIGHT,
        "☑ 已执行（展台/双屏）": GREEN_LIGHT,
        "☐ 待录像核验": AMBER,
        "☑ 摄影已交　☐ 其余待补": AMBER,
        "主体已交，部分待补": AMBER,
    }
    for i, row in enumerate(mods, 10):
        for col, val in enumerate(row, 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(name=FONT, size=9, bold=(col in (1, 2, 3)))
            cell.alignment = LEFT if col == 4 else WRAP
            cell.border = THIN
            cell.fill = fills.get(row[2], WHITE) if col == 3 else WHITE
        ws.row_dimensions[i].height = 48

    ws.merge_cells("A15:D15")
    fill_cell(
        ws,
        "A15",
        "证据目录：①营销验收单（封面 9 张）②按重要顺序分类（18 张全部按五级嵌入）"
        "③权益核验清单 ④效果评估表 ⑤绿城·潮鸣外滩露出大图 "
        f"⑥拍立享 {ALBUM}（约 425 张）⑦网盘录像 {VIDEO} 提取码 {VIDEO_CODE}。"
        "证书特写未见，不以本回执推定证书已颁。",
        9,
        False,
        LEFT,
        WHITE,
    )
    paint_range(ws, 15, 15, 1, 4, WHITE)
    ws.row_dimensions[15].height = 52

    ws.merge_cells("A16:D16")
    fill_cell(
        ws,
        "A16",
        "总体结论：晚宴冠名与主会场露出已按约定落地，效果良好，符合要求（预勾「良好」）。"
        "参观邀约、朋友圈九宫格、回顾视频、证书特写为待补项，不影响四模块整体打包之主体验收，"
        "由绿城营销负责人复核后签字盖章。本回执不与 2026 年 3 月开盘花束验收混用。",
        10,
        False,
        LEFT_TOP,
        GREEN_LIGHT,
    )
    paint_range(ws, 16, 16, 1, 4, GREEN_LIGHT)
    ws.row_dimensions[16].height = 56

    fill_cell(ws, "A17", "主办方（出具）", 10, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B17:B18")
    fill_cell(
        ws,
        "B17",
        "单位：上海市杨浦区科技企业联合会\n授权代表：胡继刚\n签字：____________\n日期：2026-05-22\n（公章）",
        10,
        False,
        LEFT_TOP,
        WHITE,
    )
    fill_cell(ws, "C17", "赞助方（接收）", 10, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("D17:D18")
    fill_cell(
        ws,
        "D17",
        "单位：上海绿城泓盛建设发展有限公司\n授权代表：____________________\n对接：笪浩 18678408669\n签字：____________　日期：____________\n（公章）",
        10,
        False,
        LEFT_TOP,
        WHITE,
    )
    paint_range(ws, 17, 18, 1, 4, WHITE)
    ws["A17"].fill = GREEN_LIGHT
    ws["C17"].fill = GREEN_LIGHT
    ws.merge_cells("A17:A18")
    ws.merge_cells("C17:C18")
    ws.row_dimensions[17].height = 36
    ws.row_dimensions[18].height = 72

    ws.merge_cells("A19:D19")
    fill_cell(ws, "A19", "回执附图（代表性证据，已嵌入本页）", 11, True, LEFT, GREEN)
    paint_range(ws, 19, 19, 1, 4, GREEN)
    ws["A19"].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws["A19"].fill = GREEN
    ws.row_dimensions[19].height = 22
    ws.row_dimensions[20].height = 118
    ws.row_dimensions[21].height = 32
    paint_range(ws, 20, 21, 1, 4, WHITE)
    receipt_photos = [
        ("941A7785.JPG", "议程看板【绿城·潮鸣】晚宴冠名"),
        ("941A7769.JPG", "展台立牌：绿城中国 GREENTOWN"),
        ("941A8810.JPG", "晚宴大屏绿城中国品牌片"),
        ("941A7663.JPG", "主背景板：晚宴冠名·绿城中国"),
    ]
    for col, (name, cap) in zip(("A", "B", "C", "D"), receipt_photos):
        t = thumb(resolve_src(name), THUMB / f"rcpt_{name}", max_w=400, max_h=280)
        place_image(ws, t, f"{col}20", 168, 118)
        fill_cell(ws, f"{col}21", cap, 8, align=LEFT)

    setup_print(ws, "A1:D21", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F6B4A"


def add_index_sheet(wb) -> None:
    ws = wb.create_sheet("照片索引")
    widths = [8, 22, 18, 22, 44, 16, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells("A1:G1")
    fill_cell(ws, "A1", "绿城·潮鸣外滩露出 · 精华照片索引（按重要等级）", 16, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 7, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:G2")
    fill_cell(
        ws,
        "A2",
        "按一级→五级排列，与「按重要顺序分类」同序。拍摄时间取自拍立享相册元数据。"
        "不收录他方展位（美年大健康、泰隆银行、腾讯云、蔚来、长江商学院）。",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 7, GREEN_MID)
    ws.row_dimensions[2].height = 28
    headers = ["图号", "重要等级", "文件名", "拍摄时间", "露出点", "对应模块", "封面"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN

    cover_files = {name for name, *_ in COVER}
    level_fill = {
        "一级": GREEN_LIGHT,
        "二级": GREEN_LIGHT,
        "三级": WHITE,
        "四级": WHITE,
        "五级": WHITE,
    }
    rows = []
    n = 1
    for cat in RANKED:
        for item in cat["photos"]:
            name, cap, module = item[0], item[1], item[2]
            rows.append(
                (
                    str(n),
                    f"{cat['level']} {cat['name']}",
                    name,
                    SHOOT_TIME.get(name, ""),
                    cap,
                    module,
                    "是" if name in cover_files else "否",
                    cat["level"],
                )
            )
            n += 1
    for i, vals in enumerate(rows, 4):
        for col, val in enumerate(vals[:7], 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(name=FONT, size=9, bold=(col == 2 and vals[7] == "一级"))
            cell.alignment = LEFT if col in (2, 3, 5) else WRAP
            cell.border = THIN
            cell.fill = GREEN_LIGHT if vals[6] == "是" else level_fill.get(vals[7], WHITE)
        ws.row_dimensions[i].height = 22
    last = 3 + len(rows)
    setup_print(ws, f"A1:G{last}", landscape=True, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C8E6C9"
    ws.auto_filter.ref = f"A3:G{last}"


def make_zip() -> None:
    ZIP_OUT.parent.mkdir(parents=True, exist_ok=True)
    used = [n for n, *_ in COVER] + [n for n, _ in EXTRA]
    with zipfile.ZipFile(ZIP_OUT, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(OUT, OUT.name)
        if PDF_OUT.exists():
            zf.write(PDF_OUT, PDF_OUT.name)
        for name in used:
            src = resolve_src(name)
            zf.write(src, f"精华照片/{name}")


def main() -> None:
    if not TEMPLATE.exists():
        raise SystemExit(f"缺少模板 {TEMPLATE}")
    THUMB.mkdir(parents=True, exist_ok=True)
    used = [n for n, *_ in COVER] + [n for n, _ in EXTRA]
    ranked_files = [p[0] for _, p in all_ranked_photos()]
    if len(ranked_files) != 18 or len(set(ranked_files)) != 18:
        raise SystemExit(f"分级照片数量不对: {len(ranked_files)} unique={len(set(ranked_files))}")
    if set(ranked_files) != set(used):
        raise SystemExit(f"分级与封面/续图不一致: {set(ranked_files) ^ set(used)}")
    for name in used:
        resolve_src(name)

    wb = load_workbook(str(TEMPLATE))
    fill_cover(wb.active)
    add_ranked_sheet(wb)
    add_rights_sheet(wb)
    add_eval_sheet(wb)
    add_receipt_sheet(wb)
    add_photo_sheet(wb)
    add_index_sheet(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT))
    make_zip()
    print(f"saved {OUT} size={OUT.stat().st_size}")
    print(f"zip {ZIP_OUT} size={ZIP_OUT.stat().st_size}")
    print("sheets", load_workbook(OUT).sheetnames)
    print("ranked", len(all_ranked_photos()), "cover", len(COVER), "extra", len(EXTRA))


if __name__ == "__main__":
    main()
