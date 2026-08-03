#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 chinajoy.net 官网 API / 新闻页公开联系方式并入招商 Excel。
数据源：
- https://www.chinajoy.net （2020api.chinajoy.net）
- https://btb.chinajoy.net 新闻页公开联系人
"""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

XLSX = Path("/workspace/output/2026ChinaJoy_招商引资参展商名录.xlsx")
DATA = Path("/tmp/cj_web/data")
RAW_OUT = Path("/workspace/output/chinajoy_official_raw")

HEADER_FILL = PatternFill("solid", fgColor="0B3D5C")
WEB_FILL = PatternFill("solid", fgColor="D5F5E3")  # 官网新增高亮
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY = Font(name="微软雅黑", size=10)
TITLE = Font(name="微软雅黑", bold=True, size=14, color="0B3D5C")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sheet(wb, title, headers, rows, widths, highlight=True):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = BODY
            cell.alignment = WRAP
            cell.border = THIN
            if highlight:
                cell.fill = WEB_FILL
        ws.row_dimensions[i].height = 28
    autofit(ws, widths)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1+len(rows)}"
    return ws


def clean(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def strip_html(s):
    return clean(re.sub(r"<[^>]+>", " ", s or ""))


def map_category(code: str, area: str) -> str:
    code = (code or "").upper()
    hall = code.split("-")[0] if code else (area or "").upper()
    if hall.startswith("N5"):
        return "骁龙主题馆"
    if hall.startswith("N"):
        return "游戏风云"
    if hall.startswith("W"):
        return "BTOB商务"
    if hall.startswith("E3"):
        return "潮流生活/甜次元/大赛"
    if hall.startswith("E4"):
        return "次元世界/创作者/摄影"
    if hall.startswith("E5"):
        return "魔玩天地"
    if hall.startswith("E6"):
        return "发烧硬件/前沿科技/极致装备/主舞台"
    if hall.startswith("E7"):
        return "发烧硬件"
    return "其他"


# 官网新闻页/招商稿补充的联系人（API contact 未全覆盖展区招商）
EXTRA_CONTACTS = [
    # 总邮箱
    ("综合总机/公共邮箱", "主办方公共邮箱", "", "cj.connect@howellexpo.net", "", "官网页脚/全站", "https://www.chinajoy.net / btb.chinajoy.net", "招商总入口"),
    # BTOB
    ("BTOB商务洽谈馆参展/赞助", "戚先生", "15801155131", "weber.qi@howellexpo.net", "", "BTOB", "btb.chinajoy.net/news/63658", "BTOB招商"),
    ("BTOB商务洽谈馆参展/赞助", "刘女士", "15810836246", "emily_liu@howellexpo.net", "", "BTOB", "btb.chinajoy.net/news/63658", "BTOB招商"),
    ("BTOB商务洽谈馆参展/赞助", "杨女士", "15321612022", "yangyanping@howellexpo.net", "", "BTOB", "btb.chinajoy.net/news/63658", "BTOB招商"),
    ("BTOB商务洽谈馆参展/赞助", "朱女士", "13811516015", "toma.zhu@howellexpo.net", "", "BTOB", "btb.chinajoy.net/news/63658", "BTOB招商"),
    ("BTOB商务洽谈馆参展/赞助", "刘先生", "18610552880", "liubaichen@howellexpo.net", "", "BTOB", "btb.chinajoy.net/news/63658", "BTOB招商"),
    ("BTOB商务洽谈馆参展/赞助", "张女士", "18810420832", "zhangsui@howellexpo.net", "", "BTOB", "btb.chinajoy.net/news/63658", "BTOB招商"),
    ("BTOB商务洽谈馆参展/赞助", "武先生", "13910527667", "william@howellexpo.net", "", "BTOB", "btb.chinajoy.net/news/63658", "BTOB招商"),
    # Indie / Game Connection
    ("ChinaJoy-Game Connection INDIE GAME 展区", "杨女士", "", "clara_yang@howellexpo.net", "clarayangqian", "W4 Indie", "btb.chinajoy.net/news/63719", "独立游戏B2B"),
    ("ChinaJoy-Game Connection INDIE GAME 展区", "赵先生", "", "lzhao@connection-events.com", "d113913144", "W4 Indie", "btb.chinajoy.net/news/63719", "独立游戏B2B"),
    ("ChinaJoy-Game Connection INDIE GAME 展区", "郭先生", "", "salesasia@connection-events.com", "travisguo1900", "W4 Indie", "btb.chinajoy.net/news/63667", "独立游戏B2B"),
    ("ChinaJoy-Game Connection INDIE GAME 展区（赞助/展商）", "葛女士", "", "salesasia@connection-events.com", "hakunagj", "W4 Indie", "btb.chinajoy.net/news/63719", "独立游戏B2B"),
    # 会议补充（寇女士等新闻页）
    ("全球游戏产业大会赞助/合作", "寇女士", "15210122472", "yali_kou@howellexpo.net", "", "同期会议", "btb.chinajoy.net/news/63709", "会议招商"),
    # 甜次元（招商稿公开）
    ("甜次元参展/赞助/合作", "王女士", "13910697624", "yvonne_wang@howellexpo.net", "", "E3 甜次元", "甜次元招商公开稿", "女性向展区"),
    ("甜次元参展/赞助/合作", "都女士", "13521800631", "", "", "E3 甜次元", "甜次元招商公开稿", "女性向展区"),
    ("甜次元参展/赞助/合作", "佘先生", "13520516873", "", "", "E3 甜次元", "甜次元招商公开稿", "女性向展区"),
    # Express（历史公开页仍常用）
    ("ChinaJoy Express 试玩区参展/赞助", "李女士", "18500508519", "grace_li@howellexpo.net", "", "N1 Express", "公开招商联系（与CGDC联系人同号，业务线可能兼任）", "独立试玩区"),
    # 指定搭建公司（服务商，招商配套）
    ("指定搭建公司", "孙宴辰 / 北京亿阳万泰", "13910758661", "13910758661@163.com", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "刘丽丽 / 上海丽华展览", "13671545818", "linaliu@sh-lihua.cn", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "王全喜 / 禧山国际", "13910097559", "564445499@qq.com", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "姚陈龙 / 全橙展览", "13801645543", "ycllogan@medesign.com.cn", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "石磊 / 上海褒茂", "13601690521", "stone@bowmore-sh.com", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "王刚 / 上海凯丰", "13218066218", "wanggang@kaifengculture.com", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "陈胤鹏 / 星图飞扬", "18518786788", "jjohnnychen@inspirestar.top", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "叶昕 / 绘事后素", "13910518975", "huishi2000@126.com", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    ("指定搭建公司", "李昊 / 墨德纬宣", "18820011747", "heaven.li@youngs-group.com", "", "搭建服务", "btb.chinajoy.net/news/63656", "展台搭建"),
    # 指定经纪
    ("指定经纪公司", "魏亚爽 / 北京德万文化", "13718535802", "annie.wei@dewan-china.com", "", "演艺经纪", "btb.chinajoy.net/news/63663", "模特/Coser/主持"),
    ("指定经纪公司", "周登科 / 北京德万文化", "15210172089", "snake.zhou@dewan-china.com", "", "演艺经纪", "btb.chinajoy.net/news/63663", "模特/Coser/主持"),
    ("指定经纪公司", "陈秋妍 / 上海星锐派", "17302174447", "3926134083@qq.com", "", "演艺经纪", "btb.chinajoy.net/news/63663", "模特/Coser/主持"),
    ("指定经纪公司", "陶俊敏 / 上海星锐派", "18521531757", "360925767@qq.com", "", "演艺经纪", "btb.chinajoy.net/news/63663", "模特/Coser/主持"),
]


def load_api_contacts():
    rows = []
    path = DATA / "contacts_all.json"
    if not path.exists():
        return rows
    for c in json.loads(path.read_text(encoding="utf-8")):
        rows.append(
            (
                clean(c.get("title_zh")),
                clean(c.get("name_zh")),
                clean(c.get("phone")),
                clean(c.get("email")),
                "",
                "同期会议（官网API）",
                "https://www.chinajoy.net → /v2/website/contact/us",
                "会议赞助合作",
            )
        )
    return rows


def build_exhibitor_rows():
    items = json.loads((DATA / "exhibitors_unique.json").read_text(encoding="utf-8"))
    rows = []
    for i, e in enumerate(sorted(items, key=lambda x: (x.get("code") or "", x.get("name_zh") or "")), 1):
        code = clean(e.get("code"))
        area = clean(e.get("_area_query"))
        rows.append(
            (
                i,
                map_category(code, area),
                area,
                code,
                clean(e.get("name_zh")),
                clean(e.get("name_en")),
                clean(e.get("website")),
                strip_html(e.get("introduction_zh")),
                e.get("id"),
                "chinajoy.net 官网 API /v2/exhibitor/list",
            )
        )
    return rows


def build_speaker_rows():
    items = json.loads((DATA / "speakers_all.json").read_text(encoding="utf-8"))
    # dedupe by name+topic+start
    seen = set()
    rows = []
    for s in items:
        key = (s.get("name_zh"), s.get("topic_zh"), s.get("start_at"))
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            (
                clean(s.get("date")),
                clean(s.get("start_at")),
                clean(s.get("end_at")),
                clean(s.get("theme_zh")),
                clean(s.get("topic_zh")),
                clean(s.get("name_zh")),
                clean(s.get("name_en")),
                clean(s.get("title_zh")),
                strip_html(s.get("introduction_zh"))[:500],
                "官网 /v4/schedule",
            )
        )
    return rows


def build_partner_rows():
    items = json.loads((DATA / "partners_flat.json").read_text(encoding="utf-8"))
    rows = []
    for p in items:
        rows.append(
            (
                clean(p.get("category_zh")),
                clean(p.get("name_zh")),
                clean(p.get("name_en")),
                clean(p.get("jump_url")),
                "官网 /v2/website/partnership",
            )
        )
    return rows


def build_conf_rows():
    rows = []
    for f in sorted(DATA.glob("conf_*.json")):
        if f.name.startswith("conf_brand"):
            continue
        try:
            j = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        d = j.get("data") or {}
        if not d.get("name_zh"):
            continue
        rows.append(
            (
                d.get("id"),
                clean(d.get("name_zh")),
                clean(d.get("theme_zh")),
                clean(d.get("venue_zh")),
                strip_html(d.get("introduction_zh"))[:800],
                clean(d.get("schedule_url")),
                "官网 /v3/conf/{id}",
            )
        )
    return rows


def copy_raw():
    RAW_OUT.mkdir(parents=True, exist_ok=True)
    for name in [
        "exhibitors_unique.json",
        "contacts_all.json",
        "speakers_all.json",
        "partners_flat.json",
        "schedules_all.json",
        "home.json",
        "partnership.json",
        "exhibitor_area.json",
        "ticket_cat.json",
    ]:
        src = DATA / name
        if src.exists():
            (RAW_OUT / name).write_text(src.read_text(encoding="utf-8"), encoding="utf-8")


def main():
    copy_raw()
    wb = load_workbook(XLSX)

    # 1) 官网展商
    ex_rows = build_exhibitor_rows()
    write_sheet(
        wb,
        "官网展商名录",
        [
            "序号",
            "展区分类(推断)",
            "查询馆区",
            "展位号",
            "企业中文名",
            "企业英文名",
            "官网网址",
            "企业简介",
            "官网ID",
            "数据来源",
        ],
        ex_rows,
        [6, 22, 10, 14, 28, 28, 28, 55, 10, 28],
    )

    # 2) 联系方式
    contact_rows = []
    seen = set()
    for row in load_api_contacts() + EXTRA_CONTACTS:
        key = (row[0], row[1], row[2], row[3], row[4])
        if key in seen:
            continue
        seen.add(key)
        contact_rows.append(row)
    write_sheet(
        wb,
        "官网联系方式",
        ["业务板块", "联系人", "电话", "邮箱", "微信", "适用展区/场景", "来源", "招商用途"],
        contact_rows,
        [32, 22, 16, 32, 18, 18, 36, 18],
    )

    # 3) 嘉宾
    sp_rows = build_speaker_rows()
    write_sheet(
        wb,
        "官网嘉宾与演讲",
        ["日期", "开始", "结束", "板块/主题", "议题", "嘉宾姓名", "英文名", "职务", "简介摘要", "来源"],
        sp_rows,
        [12, 20, 20, 18, 40, 14, 16, 36, 45, 16],
    )

    # 4) 合作伙伴
    write_sheet(
        wb,
        "官网合作伙伴",
        ["合作类别", "名称中文", "名称英文", "跳转链接", "来源"],
        build_partner_rows(),
        [36, 36, 40, 30, 28],
    )

    # 5) 同期会议
    write_sheet(
        wb,
        "官网同期会议",
        ["会议ID", "会议名称", "主题", "地点", "简介", "日程链接", "来源"],
        build_conf_rows(),
        [10, 28, 36, 28, 55, 28, 18],
    )

    # 6) 更新总览
    ws0 = wb["展会总览"]
    r = ws0.max_row + 2
    ws0.cell(row=r, column=1, value="官网数据更新").font = Font(name="微软雅黑", bold=True, size=10)
    ws0.cell(
        row=r,
        column=2,
        value=(
            f"{datetime.now().strftime('%Y-%m-%d %H:%M')} 从 https://www.chinajoy.net 及相关 API/新闻页拉取："
            f"官网展商 {len(ex_rows)} 家；联系人 {len(contact_rows)} 条；嘉宾演讲去重后 {len(sp_rows)} 条；"
            f"合作伙伴类别条目 {len(build_partner_rows())}；原始 JSON 见 output/chinajoy_official_raw/。"
            "官网展商接口一般不返回企业电话，联系方式以主办招商负责人及指定服务商公开信息为主。"
        ),
    ).alignment = WRAP
    ws0.row_dimensions[r].height = 70

    # 7) 数据来源追加
    ws_src = wb["数据来源与声明"]
    last = ws_src.max_row + 1
    for c, v in enumerate(
        [
            "ChinaJoy 官网 API",
            "2020api.chinajoy.net（home/contact/exhibitor/schedule/partnership/conf）",
            "https://www.chinajoy.net/#/navStation/home",
            "展商、联系人、嘉宾、伙伴",
        ],
        1,
    ):
        cell = ws_src.cell(row=last, column=c, value=v)
        cell.font = BODY
        cell.border = THIN
        cell.fill = WEB_FILL
    last += 1
    for c, v in enumerate(
        [
            "ChinaJoy BTOB 新闻",
            "btb.chinajoy.net 公开招商/指定服务商通讯录",
            "https://btb.chinajoy.net/",
            "BTOB/Indie/搭建/经纪联系人",
        ],
        1,
    ):
        cell = ws_src.cell(row=last, column=c, value=v)
        cell.font = BODY
        cell.border = THIN
        cell.fill = WEB_FILL

    # 8) 刷新统计补充
    if "名录统计" in wb.sheetnames:
        ws = wb["名录统计"]
        rr = ws.max_row + 2
        ws.cell(row=rr, column=1, value="官网展商名录条目").font = Font(name="微软雅黑", bold=True)
        ws.cell(row=rr, column=2, value=len(ex_rows))
        ws.cell(row=rr + 1, column=1, value="官网联系方式条目").font = Font(name="微软雅黑", bold=True)
        ws.cell(row=rr + 1, column=2, value=len(contact_rows))
        ws.cell(row=rr + 2, column=1, value="官网嘉宾演讲条目").font = Font(name="微软雅黑", bold=True)
        ws.cell(row=rr + 2, column=2, value=len(sp_rows))
        # hall stats
        cats = Counter(r[1] for r in ex_rows)
        rr += 4
        ws.cell(row=rr, column=1, value="官网展商分类分布").font = TITLE
        rr += 1
        for cat, n in cats.most_common():
            ws.cell(row=rr, column=1, value=cat).font = BODY
            ws.cell(row=rr, column=2, value=n).font = BODY
            rr += 1

    wb.save(XLSX)
    print("updated", XLSX)
    print("exhibitors", len(ex_rows), "contacts", len(contact_rows), "speakers", len(sp_rows))


if __name__ == "__main__":
    main()
