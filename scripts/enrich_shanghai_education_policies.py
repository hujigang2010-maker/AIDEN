#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
补充上海幼升小 / 小升初 / 中考政策，写入《上海板块学区融合总表.xlsx》。
政策口径以市教委近年实施意见与政策问答为主，区级细则每年以4月公布为准。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_XLSX = ROOT / "output" / "上海板块学区融合总表.xlsx"

C = {
    "navy": "1B2A4A",
    "navy2": "243B5C",
    "teal": "0E7490",
    "teal_soft": "E0F2FE",
    "ink": "0F172A",
    "muted": "64748B",
    "line": "CBD5E1",
    "zebra": "F8FAFC",
    "white": "FFFFFF",
    "green": "0F766E",
    "green_bg": "CCFBF1",
    "blue": "1D4ED8",
    "blue_bg": "DBEAFE",
    "amber": "C2410C",
    "amber_bg": "FFEDD5",
    "red": "B91C1C",
    "red_bg": "FEE2E2",
    "slate": "F1F5F9",
    "gold": "B45309",
    "gold_bg": "FEF3C7",
}


def thin() -> Border:
    s = Side(style="thin", color=C["line"])
    return Border(left=s, right=s, top=s, bottom=s)


def fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


def font(bold=False, size=10, color=None):
    return Font(name="微软雅黑", bold=bold, size=size, color=color or C["ink"])


def title_bar(ws, text: str, cols: int, color: str = "navy"):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(1, 1, text)
    cell.font = Font(name="微软雅黑", bold=True, size=16, color=C["white"])
    cell.fill = fill(C[color])
    cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False


def subtitle(ws, text: str, cols: int, row: int = 2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row, 1, text)
    cell.font = font(size=9, color=C["muted"])
    cell.fill = fill(C["slate"])
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def header_row(ws, row: int, headers: list[str], color: str = "navy2"):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = fill(C[color])
        cell.font = font(bold=True, size=10, color=C["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()
    ws.row_dimensions[row].height = 30


def write_table(ws, start_row: int, headers: list[str], rows: list[list], widths: list[int], header_color="navy2"):
    header_row(ws, start_row, headers, header_color)
    for i, row_vals in enumerate(rows):
        r = start_row + 1 + i
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin()
            cell.font = font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left" if c > 1 else "center")
            if i % 2 == 1:
                cell.fill = fill(C["zebra"])
        ws.row_dimensions[r].height = max(36, 18 + 12 * max(str(x).count("\n") for x in row_vals))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return start_row + 1 + len(rows)


def section_title(ws, row: int, text: str, cols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row, 1, text)
    cell.font = font(bold=True, size=12, color=C["teal"])
    cell.fill = fill(C["teal_soft"])
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24
    return row + 1


def tag_cell(cell, text: str, kind: str):
    mapping = {
        "ok": ("green", "green_bg"),
        "warn": ("amber", "amber_bg"),
        "bad": ("red", "red_bg"),
        "info": ("blue", "blue_bg"),
        "note": ("gold", "gold_bg"),
    }
    fg, bg = mapping.get(kind, ("ink", "slate"))
    cell.value = text
    cell.fill = fill(C[bg])
    cell.font = font(bold=True, size=9, color=C[fg])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin()


# —— 政策数据 ——
YOUXIAO_ROWS = [
    ["核心原则", "免试就近入学；严禁以考试/竞赛/培训成绩或证书作为招生依据；公民办同步招生，超额摇号"],
    ["报名系统", "「一网通办」义务教育入学专栏 或 上海市义务教育入学报名系统（shrxbm.edu.sh.gov.cn）；亦可经「随申办」办理"],
    ["入学对象（以当年文件为准）", "年满6周岁适龄儿童（如2026年对象为2019.9.1—2020.8.31出生）；特殊原因可申请推迟一年"],
    ["信息登记", "通常4月中下旬完成；在园儿童由幼儿园协助登记，未入园按区指定方式登记；登记后明确一名关联监护人"],
    ["公民办二选一", "公办报名与民办报名只能选择其一；提交后不可改报、不可放弃志愿后再报另一类"],
    ["公办路径", "按户籍/对口地段就近入学；人户一致优先；超额时按入户年限、房产登记时间、与户主关系等排序（各区细则不同）"],
    ["民办路径", "网上报名 → 报名人数≤计划则全部录取，超过计划则电脑随机录取；可填1个志愿+1个调剂志愿"],
    ["人户一致", "儿童户籍地址与本人或直系亲属自有住宅类房产地址一致，通常为公办最优入学类别"],
    ["人户分离", "确有困难不能在户籍地入学，可凭有效《本市户籍人户分离人员居住登记凭证》申请居住地入学；人户一致优先后再统筹"],
    ["五年一户", "部分区/校实行「同一地址五年内仅享有一次同校对口入学机会」；热门小学常见预警，买房前必须核验"],
    ["随迁子女", "儿童持有效居住证/登记凭证；父母一方持有效居住证，且满足社保满6个月或灵活就业登记连续3年等条件（以当年文件为准）"],
    ["验证与告知", "公办通常5月中下旬分批验证；通过后发送入学告知；民办未录取者按「人户一致优先、同类排序靠后」统筹进公办"],
    ["校园开放日", "约4月中旬；不与招生挂钩，不得测试/收简历/变相报名"],
]

XIAOSHENGCHU_ROWS = [
    ["核心原则", "免试入学；公办初中采取小学划片对口、居住地段对口或电脑派位等方式；公民办同步招生"],
    ["信息核对", "系统对接小学五年级学籍；约4月中下旬核对户籍/居住地址等，截止日后不可随意更改"],
    ["三种公办方式", "①小学划片对口（学籍对口）②居住地段对口（户籍/房产对口）③电脑派位；各区采用组合不同"],
    ["学籍对口区（常见）", "黄浦、杨浦、虹口、普陀等：小学毕业后按学籍对口进入初中，小学选择影响更大"],
    ["户籍对口区（常见）", "浦东、闵行、宝山、嘉定、松江、青浦、奉贤、金山、崇明等：更强调户籍/居住地段对口"],
    ["学籍+派位区（常见）", "徐汇、长宁、静安等：学籍对口为主，部分热门学校结合电脑派位"],
    ["回户籍/居住地", "跨区就读小学生可申请回户籍（居住）地读初中；网上填申请表，由目标区统筹安排；有明确截止日"],
    ["民办初中", "约5月中旬网上报名；1个志愿+1个调剂；超额电脑摇号；未录取按对口生源优先、同类靠后统筹公办"],
    ["民办一贯制直升", "一贯制学校通常先征求本校直升意愿并完成直升录取，再进行校外招生"],
    ["购买学位民办校", "部分民办校纳入区购买学位政策，仍属民办、同步招生、超额摇号；学费与生均标准差额规则见简章"],
    ["外省市回沪读六年级", "本市户籍或符合条件非本市户籍，在外地读五年级需回沪的，向户籍/居住地区登记后统筹安排"],
    ["与买房关系", "不必一步买「小初双优」；可先锁定小学，再视小升初规则置换初中学区（纪要策略）"],
]

ZHONGKAO_ROWS = [
    ["考试构成", "初中学业水平考试：语文、数学、外语、道法、历史、体育与健身、综合测试（物理/化学/跨学科/实验等），总分750分"],
    ["综合素质评价", "纪实报告评价；名额分配批次「合格」赋50分（满分800=750+50）；自主招生等也需结合综评"],
    ["录取三大批次", "依次：①自主招生 ②名额分配综合评价 ③统一招生；前一批次录取后不再进入后批次同路径"],
    ["自主招生", "市实验性示范性高中约10%计划；特色高中不超过15%；含体艺骨干、国际课程班、中职校自主招生等；可兼报类别但普高签约预录通常限1所"],
    ["名额分配到区", "市实验性示范性高中该批次约占本校计划65%中的「到区」部分；按报名所在区投档，区与区之间分数线差异大"],
    ["名额分配到校", "分配到不选择生源初中；通常要求在籍在读满3年；校内竞争，弱校也可能冲市重点"],
    ["统一招生", "普高+中职；平行志愿（常见1—15志愿）+征求志愿；按学业考试成绩投档"],
    ["中本贯通", "仅限符合条件的上海户籍；平行志愿；成绩须达普高最低投档控制线"],
    ["志愿填报时点", "学业考试后、成绩公布前填报各批次志愿（以当年考试院安排为准）"],
    ["报名资格关键", "中考报名资格与户籍/学籍/居住证积分等挂钩；随迁子女报考高中阶段学校有专门规定，小升初阶段就应提前了解"],
    ["与选区关系", "中考按报名所在区竞争「名额分配到区」；区内外头部高中供给、报名人数决定难度。黄浦等区常被视为「中考相对友好」区之一"],
    ["纪要提醒", "高中统考，不必用学区房提前锁定高中；升学主矛盾在初升高，应优先布局小学+初中"],
]

CALENDAR_YOU = [
    ["约3月底", "市教委发布当年义务教育招生实施意见与政策问答"],
    ["4月7日前后", "各区公布实施方案、对口划片、人户分离细则"],
    ["4月中旬", "校园开放日（不挂钩招生）"],
    ["4月中下旬", "幼升小入学信息登记；人户分离居住地入学申请截止"],
    ["5月上旬", "公办/民办小学报名（公民办二选一；民办通常更短窗口）"],
    ["5月中下旬", "公办验证；民办超额摇号；陆续发放入学告知"],
    ["5月底", "公办第二批验证（含民办未录取统筹）"],
    ["8月中旬前", "发放入学通知书；开学前完成学籍对接"],
]

CALENDAR_CHU = [
    ["约3月底", "市级招生意见发布；各区随后公布初中划片方案"],
    ["4月中下旬", "五年级信息核对；回户籍/居住地申请截止；一贯制征求直升意愿"],
    ["5月上旬", "民办一贯制直升录取（约）"],
    ["5月中旬", "民办初中网上报名"],
    ["5月中下旬", "民办摇号/调剂；公办按对口或派位分配；发放入学信息"],
    ["8月", "入学通知与报到"],
]

CALENDAR_ZK = [
    ["约2—3月", "市教委发布高中阶段招生若干意见；学校公布自主招生方案"],
    ["3—5月", "中招报名、资格确认；体艺等专项资格"],
    ["4—5月", "体育测试、外语听说、理化实验等（具体日程以考试院为准）"],
    ["6月中旬", "学业考试笔试"],
    ["考试后—出分前", "填报自主招生/名额分配/统一招生等志愿"],
    ["出分后", "按批次投档录取；公布各控制线与名额分配分数线"],
]

DISTRICT_ENTRY = [
    ["黄浦", "学籍对口", "强", "人户一致优先；全区五年一户较常见", "中考相对友好、人口少", "★★★ 策略优选区"],
    ["杨浦", "学籍对口", "强", "人户一致；热门校有入户/年限要求", "滨江与鞍山等老破小逻辑", "★★★ 策略优选区"],
    ["虹口", "学籍对口", "较强", "核验对口与年限", "北外滩等可作次优", "★★ 次优观察"],
    ["普陀", "学籍对口", "中等", "部分学校五年一户/预警", "曹杨等老工人新村可看", "★★ 可考虑"],
    ["徐汇", "学籍+派位", "中等", "热门双学区高门槛；九年一贯常需提前年限", "资源顶尖但竞争与总价高", "★ 谨慎/条件核验"],
    ["长宁", "学籍+派位", "较强", "派位规则需看当年细则", "公办较均衡", "★★ 可观察"],
    ["静安", "学籍+派位特征", "强", "热门一贯制竞争激烈", "总价门槛高", "★★ 可观察"],
    ["浦东", "户籍对口", "较强", "热门小学常见入户年限预警/五年一户", "仅潍坊等少数老板块契合空心化策略", "分化极大"],
    ["闵行", "户籍对口", "弱（民办开放）", "年限与竞争双高", "新上海人集中，纪要多数排除", "✕ 策略排除"],
    ["宝山", "户籍对口", "中等", "核验对口", "人口压力大", "✕ 策略排除"],
    ["嘉定", "户籍对口", "尚可", "新城热门校年限趋严", "纪要排除", "✕ 策略排除"],
    ["松江", "户籍对口", "弱", "九亭等卷度极高", "纪要点名排除九亭", "✕ 策略排除"],
    ["青浦", "户籍对口", "中等", "核验对口", "导入型为主", "✕ 主线排除"],
    ["奉贤", "户籍对口", "中等", "核验对口", "远郊，通勤与资源需权衡", "○ 观察"],
    ["金山", "户籍对口", "中等", "核验对口", "竞争低但资源与通勤弱", "○ 观察"],
    ["崇明", "户籍对口", "中等", "核验对口", "极端空心化，多数家庭不适配", "○ 观察"],
]

BUY_TIPS = [
    ["幼升小买房", "优先保证「人户一致」且满足目标校入户/房产年限；明年入学则倒推能否赶上登记截止日与验证要求"],
    ["五年一户", "同一套房子可能五年内只能送一孩进同一所对口小学；二孩家庭或准备出售/挂靠者务必查清"],
    ["小升初", "学籍对口区：小学选择几乎决定初中；户籍对口区：更可考虑小学后再搬家换初中"],
    ["九年一贯", "复旦附属等常见要求户籍房产一致且提前约3年；时间不够不要硬冲"],
    ["中考", "不靠学区房锁定高中；关注报名所在区的名额分配结构与竞争生态"],
    ["随迁/积分", "非沪籍路径与沪籍完全不同；积分、居住证、社保断缴风险要单独评估"],
    ["材料清单（常见）", "户口簿、房产证/租赁备案、入学信息登记表、人户分离凭证（如有）、居住证及社保材料（随迁）"],
    ["权威查询", "市教委官网、各区教育局当年实施方案、shrxbm.edu.sh.gov.cn、上海教育考试院（中考）"],
]


def build_youxiao(wb):
    name = "幼升小政策"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    title_bar(ws, "上海 · 幼升小（幼儿园→小学）政策要点", 3, "teal")
    subtitle(ws, "依据市教委义务教育招生实施意见及政策问答整理；2026年口径与2025年总体保持一致。具体日期、对口、年限以当年各区4月细则为准。", 3)
    r = section_title(ws, 4, "一、政策要点总览", 3)
    r = write_table(ws, r, ["政策项", "内容要点"], [[a, b] for a, b in YOUXIAO_ROWS], [22, 88], "teal")
    r += 1
    r = section_title(ws, r, "二、典型时间线（日期每年微调）", 3)
    r = write_table(ws, r, ["阶段", "事项"], CALENDAR_YOU, [18, 70], "navy2")
    r += 1
    r = section_title(ws, r, "三、入学优先级（理解「抢学位」的本质）", 3)
    hierarchy = [
        ["1", "人户一致（户籍=自有住房地址）", "最优，对口确定性最高"],
        ["2", "人户分离但符合居住地入学条件", "次优，可能被统筹"],
        ["3", "集体户口 / 廉租房等特殊路径", "按区细则就近或统筹"],
        ["4", "随迁子女（居住证+社保等）", "另行排序，热门校机会通常更弱"],
        ["5", "材料不齐 / 错过节点", "风险最大，可能无法按意愿入学"],
    ]
    r = write_table(ws, r, ["优先级", "类别", "说明"], hierarchy, [10, 36, 40], "navy2")
    # color first col
    for i in range(len(hierarchy)):
        tag_cell(ws.cell(r - len(hierarchy) + i, 1), hierarchy[i][0], "info" if i < 2 else ("warn" if i < 4 else "bad"))
    r += 1
    r = section_title(ws, r, "四、与本选房策略的衔接", 3)
    tips = [
        ["空心化板块", "居住适龄人口少 → 对口校更不易触发「超额排序/统筹」，人户一致更「稳」"],
        ["潍坊等老破小", "总价友好，但须核对明珠/浦师附小等当年入户年限与五年一户"],
        ["不要只看校名", "先看能否在入学年满足人户一致+年限，再看小学口碑"],
    ]
    write_table(ws, r, ["关注点", "操作建议"], tips, [18, 72], "navy2")


def build_xiaoshengchu(wb):
    name = "小升初政策"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 2)
    title_bar(ws, "上海 · 小升初政策要点", 3)
    subtitle(ws, "公办初中入学方式因区而异（对口 / 派位）。买房前先确认目标区属于哪一类，再决定「一步到位」还是「小学后再置换」。", 3)
    r = section_title(ws, 4, "一、政策要点总览", 3)
    r = write_table(ws, r, ["政策项", "内容要点"], XIAOSHENGCHU_ROWS, [22, 88])
    r += 1
    r = section_title(ws, r, "二、典型时间线", 3)
    r = write_table(ws, r, ["阶段", "事项"], CALENDAR_CHU, [18, 70], "teal")
    r += 1
    r = section_title(ws, r, "三、十六区公办小升初方式对照（购房决策用）", 3)
    r = write_table(
        ws,
        r,
        ["行政区", "常见入学方式", "本区保护体感", "年限/对口提示", "与纪要策略关系", "策略标签"],
        DISTRICT_ENTRY,
        [10, 14, 14, 28, 28, 16],
    )
    # tag last col
    start = r - len(DISTRICT_ENTRY)
    for i, row in enumerate(DISTRICT_ENTRY):
        label = row[-1]
        kind = "ok" if "★" in label and "✕" not in label else ("bad" if "✕" in label else ("warn" if "谨慎" in label else "info"))
        if "排除" in label:
            kind = "bad"
        elif "优选" in label:
            kind = "ok"
        elif "次优" in label or "可考虑" in label or "观察" in label:
            kind = "info"
        tag_cell(ws.cell(start + i, 6), label, kind)
    r += 1
    r = section_title(ws, r, "四、决策分流（结合纪要）", 3)
    flow = [
        ["目标在学籍对口区", "小学对口几乎绑定初中 → 更需一次选对小学；或接受对口初中普通再谋民办摇号"],
        ["目标在户籍对口区", "更适合「先小学、后置换初中」；搬家时间点卡在小升初信息核对前"],
        ["准备冲民办初中", "公民办同步；未摇中回公办时可能排序靠后，需有保底对口"],
        ["准备读一贯制", "核对直升规则与入户年限；时间不够则放弃"],
    ]
    write_table(ws, r, ["情形", "建议"], flow, [22, 78], "navy2")


def build_zhongkao(wb):
    name = "中考政策"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 3)
    title_bar(ws, "上海 · 中考（初升高）政策要点", 3, "navy2")
    subtitle(ws, "中考是全市学业考试+分批次录取。学区房无法「锁定」高中；真正影响的是报名区竞争生态与名额分配结构。分数线每年变化，下表讲规则不报预测分。", 3)
    r = section_title(ws, 4, "一、政策要点总览", 3)
    r = write_table(ws, r, ["政策项", "内容要点"], ZHONGKAO_ROWS, [22, 88], "navy2")
    r += 1
    r = section_title(ws, r, "二、录取批次结构", 3)
    batches = [
        ["第1批", "自主招生", "市实验性示范性/特色高中自招、体艺、国际课程班、中职校自招等"],
        ["第2批", "名额分配综合评价", "到区（区间竞争）+ 到校（校内竞争，满3年学籍）"],
        ["第3批", "统一招生", "普高与中职平行志愿 + 征求志愿"],
    ]
    r = write_table(ws, r, ["顺序", "批次", "主要内容"], batches, [10, 22, 70], "teal")
    for i in range(3):
        tag_cell(ws.cell(r - 3 + i, 1), batches[i][0], "info")
    r += 1
    r = section_title(ws, r, "三、典型时间线", 3)
    r = write_table(ws, r, ["阶段", "事项"], CALENDAR_ZK, [22, 70])
    r += 1
    r = section_title(ws, r, "四、对选房/选区的含义", 3)
    meaning = [
        ["不要为高中买学区房", "高中靠中考；学区房解决的是义务教育入场券"],
        ["区的选择仍有意义", "影响初中资源、名额分配到区的竞争池、中考「性价比」"],
        ["到校名额", "在薄弱初中满3年也可能冲市重点 → 与「差异化、降竞争」思路相容"],
        ["户籍与中本贯通", "中本贯通通常仅上海户籍；落户时间表要与升学规划对齐"],
    ]
    write_table(ws, r, ["结论", "说明"], meaning, [22, 70], "navy2")


def build_overview(wb):
    name = "升学政策总览"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    title_bar(ws, "上海升学政策总览（幼升小 · 小升初 · 中考）", 4)
    subtitle(ws, "本册政策页服务于学区选房决策。权威原文以市教委、各区教育局、上海教育考试院当年文件为准。", 4)

    r = section_title(ws, 4, "一、三阶段对照", 4)
    cmp_rows = [
        ["幼升小", "免试就近；人户一致优先", "公民办二选一；民办摇号", "人户一致+入户年限+五年一户", "现在必须解决（明年入学）"],
        ["小升初", "对口 / 派位 / 回户籍", "公民办同步；民办摇号", "学籍对口区几乎绑定小学", "可分步：小学后再置换"],
        ["中考", "全市统考+分批次录取", "自招→名额分配→统一招生", "报名区竞争与名额结构", "不必用学区房锁定高中"],
    ]
    r = write_table(ws, r, ["阶段", "入学/录取逻辑", "关键机制", "买房最相关变量", "纪要策略位置"], cmp_rows, [12, 24, 24, 28, 22])
    r += 1
    r = section_title(ws, r, "二、购房决策检查清单", 4)
    r = write_table(ws, r, ["检查项", "操作要点"], BUY_TIPS, [18, 80], "teal")
    r += 1
    r = section_title(ws, r, "三、本册其他政策页", 4)
    nav = [
        ["幼升小政策", "信息登记、人户一致、公民办报名、验证与统筹"],
        ["小升初政策", "三种入学方式、十六区对照、回户籍、民办初中"],
        ["中考政策", "750分构成、三大批次、名额分配到区/到校"],
    ]
    write_table(ws, r, ["工作表", "内容"], nav, [18, 60], "navy2")

    # 来源
    r = ws.max_row + 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(r, 1, "主要参考：上海市教委义务教育招生实施意见及政策问答；上海市教委高中阶段学校招生工作若干意见；上海教育考试院中招规则公开信息。").font = font(size=8, color=C["muted"])


def remove_old_policy_sheets(wb):
    for name in list(wb.sheetnames):
        if name in ("幼升小政策", "小升初政策", "中考政策", "升学政策总览"):
            del wb[name]


def main():
    if not OUT_XLSX.exists():
        raise SystemExit(f"请先生成融合总表: {OUT_XLSX}")
    wb = load_workbook(OUT_XLSX)
    remove_old_policy_sheets(wb)
    build_overview(wb)
    build_youxiao(wb)
    build_xiaoshengchu(wb)
    build_zhongkao(wb)
    # keep 融合总表 first
    order = [
        "融合总表",
        "升学政策总览",
        "幼升小政策",
        "小升初政策",
        "中考政策",
        "入选看房清单",
        "十六区策略总览",
        "评分图例与融合规则",
        "双源校准说明",
    ]
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
    wb.save(OUT_XLSX)
    print("已更新政策页:", [s for s in wb.sheetnames if "政策" in s or "升学" in s])
    print("全部工作表:", wb.sheetnames)
    print("输出:", OUT_XLSX)


if __name__ == "__main__":
    main()
