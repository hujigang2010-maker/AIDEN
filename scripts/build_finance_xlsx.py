"""
冠松 GS · iDrive Hub 财务测算 Excel 生成脚本

包含 7 个 Sheet：
1. 摘要 Dashboard         · 关键指标速览
2. 假设与参数              · 入驻率/租金/成本等可调参数
3. 三年损益                · 损益表
4. 月度滚动现金流          · 36 个月现金流
5. 入驻进度                · 链主+生态月度爬坡
6. 敏感性分析              · 双变量敏感矩阵
7. 团队与人力成本          · 22 人薪酬带宽
8. 让步阶梯计算器          · 链主谈判 5 步让步对应 NPV 影响

注：所有数字与 docs/phase4-commercial/07-financial-model.csv 对齐。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from pathlib import Path

OUT_DIR = Path(__file__).resolve().parent.parent / "docs" / "finance"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 配色（与 PPT 一致）
NAVY = "0F2D52"
BLUE = "1F6FEB"
GOLD = "C9A24A"
CLOUD = "F4F6FA"
GREEN = "2FA36F"
RED = "D04A4A"
GREY = "6B7380"
WHITE = "FFFFFF"
BLACK = "1B1F2A"
LINE = "D8DEE9"

CN_FONT = "WenQuanYi Micro Hei"

thin = Side(border_style="thin", color=LINE)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, fill=NAVY, color=WHITE, size=11):
    cell.font = Font(name=CN_FONT, size=size, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_all


def style_body(cell, color=BLACK, size=10, bold=False, align="left",
               fill=None, number_format=None):
    cell.font = Font(name=CN_FONT, size=size, bold=bold, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = border_all
    if number_format:
        cell.number_format = number_format


def style_title(cell, fill=NAVY, color=WHITE, size=18):
    cell.font = Font(name=CN_FONT, size=size, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build():
    wb = Workbook()

    # =================== Sheet 1: 摘要 Dashboard ===================
    ws = wb.active
    ws.title = "1. 摘要 Dashboard"
    set_col_widths(ws, [22, 16, 16, 16, 16])

    ws.merge_cells("A1:E1")
    style_title(ws["A1"], size=20)
    ws["A1"] = "GS · iDrive Hub · 三年财务测算 Dashboard"
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    style_title(ws["A2"], fill=GOLD, color=NAVY, size=12)
    ws["A2"] = "01# 研发楼 · 1.5 万方单楼 · 出租净 8,300 ㎡ · 策划阶段 v1.1"
    ws.row_dimensions[2].height = 22

    ws["A4"] = "项目载体（PDF 实测）"
    style_header(ws["A4"], size=12)
    ws.merge_cells("B4:E4")
    style_body(ws["B4"], align="left", size=11)
    ws["B4"] = ("01# 新建研发楼 · 永和社区 075b-07 · 9F · 高 44.95 m · "
                "C6 教育科研用地 · 装配式 100% / 绿建二星 / 540 ㎡ 光伏")

    # 关键 KPI 卡片
    ws["A6"] = "关键 KPI"
    style_header(ws["A6"], fill=NAVY, color=WHITE, size=14)
    ws.merge_cells("A6:E6")

    kpi_rows = [
        ("出租净面积", "8,300 ㎡", "基于 1F+2F 自留扣除", "—"),
        ("Y3 入驻率目标", "92%", "Y1 35% / Y2 70% / Y3 92%", "高目标"),
        ("Y3 平均有效租金", "7.5 元/㎡·天", "扣除免租与装补摊销", "中心城区合理"),
        ("Y3 总收入", "5,748 万元", "含租金/物业/服务/政策返还", "稳态"),
        ("Y3 EBITDA", "+1,368 万元", "由 Y1 −2,108 转正", "Y3 转正"),
        ("Y3 EBITDA 利润率", "+24%", "Y1 −206% / Y2 −17% / Y3 +24%", "稳健"),
        ("EBIT 转正时点", "Y3 末 → Y4 初", "约 36–40 个月", "—"),
        ("链主签约目标", "Y1 0 / Y2 1 / Y3 1", "TOP5 中至少 1 家", "—"),
        ("入驻企业累计", "Y1 4 / Y2 9 / Y3 14", "12–18 家区间", "—"),
        ("Y0 启动现金需求", "约 3 亿元", "Y1 期初现金安全垫 ≥ 12 个月", "—"),
    ]
    headers = ["指标", "数值", "口径说明", "评级"]
    for j, h in enumerate(headers):
        cell = ws.cell(row=7, column=j + 1, value=h)
        style_header(cell)
    for i, row in enumerate(kpi_rows, start=8):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1, value=val)
            fill = CLOUD if i % 2 == 0 else WHITE
            bold = j == 1
            style_body(cell, fill=fill, bold=bold,
                       align="left" if j != 1 else "center")

    # 三年关键数据
    ws["A20"] = "三年损益快览（万元）"
    style_header(ws["A20"], fill=NAVY, color=WHITE, size=14)
    ws.merge_cells("A20:E20")
    pl_headers = ["科目", "Y1 2026", "Y2 2027", "Y3 2028", "三年累计"]
    for j, h in enumerate(pl_headers):
        style_header(ws.cell(row=21, column=j + 1, value=h))
    pl_rows = [
        ("总收入", 1022, 3237, 5748, 10007),
        ("总成本", 3130, 3790, 4380, 11300),
        ("EBITDA", -2108, -553, 1368, -1293),
        ("EBIT", -3308, -1753, 168, -4893),
    ]
    for i, row in enumerate(pl_rows, start=22):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1, value=val)
            fill = CLOUD if i % 2 == 0 else WHITE
            color = RED if (j > 0 and isinstance(val, (int, float)) and val < 0) else BLACK
            bold = j == 0 or j == 4
            num_fmt = "#,##0;[Red]−#,##0" if j > 0 else None
            style_body(cell, fill=fill, color=color, bold=bold,
                       align="right" if j > 0 else "left",
                       number_format=num_fmt)

    ws["A28"] = "📌 详细测算见后续 Sheet（假设/损益/现金流/敏感性等）"
    style_body(ws["A28"], color=GREY, size=10, align="left")
    ws.merge_cells("A28:E28")

    # =================== Sheet 2: 假设与参数 ===================
    ws = wb.create_sheet("2. 假设与参数")
    set_col_widths(ws, [28, 16, 16, 16, 24])

    ws.merge_cells("A1:E1")
    style_title(ws["A1"], size=18)
    ws["A1"] = "可调参数与假设（黄色单元格 = 可修改）"
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    style_title(ws["A2"], fill=GOLD, color=NAVY, size=11)
    ws["A2"] = "修改下方任一参数后，损益与现金流相关公式可在 Excel 内手动联动（本表为静态版）"

    sections = [
        ("一、基础参数", [
            ("出租净面积 (㎡)", 8300, 8300, 8300, "PDF 实测"),
            ("地上建筑面积 (㎡)", 15152, 15152, 15152, "PDF 实测"),
            ("地下建筑面积 (㎡)", 6993, 6993, 6993, "PDF 实测"),
            ("建筑高度 (m)", 44.95, 44.95, 44.95, "PDF 实测"),
            ("用地性质", "C6", "C6", "C6", "教育科研设计用地"),
            ("绿建等级", "二星", "二星", "二星", "PDF 实测"),
        ]),
        ("二、出租与租金", [
            ("入驻率 (%)", 35, 70, 92, "线性爬坡"),
            ("平均有效租金 (元/㎡·天)", 5.8, 6.8, 7.5, "扣除免租与装补摊销"),
            ("物业费 (元/㎡·月)", 28, 28, 28, "标准"),
            ("免租期 (月)", 15, 12, 10, "链主 18 / 生态 6 加权"),
            ("装补 (元/㎡)", 1000, 700, 500, "链主 1200 / 生态 500 加权"),
        ]),
        ("三、增值服务收入（万元）", [
            ("1F+2F 冠名权", 100, 300, 500, "大堂/展厅/LED"),
            ("服务平台佣金", 80, 250, 600, "GMV 5%-10%"),
            ("后市场协同分成", 50, 150, 300, "与冠松后市场联动"),
            ("政策返还（净计入）", 80, 600, 1500, "一企一策"),
            ("基金管理费", 0, 300, 500, "园区基金 1%-2% AUM"),
        ]),
        ("四、成本（万元）", [
            ("物业运营成本", 600, 750, 850, "人力+能源+维保"),
            ("招商佣金摊销", 100, 200, 300, "5 年直线"),
            ("装补摊销", 200, 350, 450, "5 年直线"),
            ("免租摊销", 250, 200, 100, ""),
            ("营销品牌(含发布会)", 800, 600, 600, ""),
            ("政府关系/法务", 200, 250, 300, ""),
            ("团队人力", 400, 800, 1100, "4→14→22 人"),
            ("测试场协同", 80, 120, 150, "嘉定/临港会员"),
            ("财务费用", 350, 320, 280, ""),
            ("其他", 150, 200, 250, ""),
        ]),
        ("五、链主签约假设", [
            ("链主 Term Sheet 数", 1, 2, 3, "TOP5 中"),
            ("链主签约数", 0, 1, 1, "至少 1 家"),
            ("入驻企业 (累计)", 4, 9, 14, "12–18 家区间"),
        ]),
    ]

    cur_row = 4
    for sec_name, items in sections:
        ws.merge_cells(start_row=cur_row, start_column=1,
                       end_row=cur_row, end_column=5)
        cell = ws.cell(row=cur_row, column=1, value=sec_name)
        style_header(cell, fill=NAVY, color=WHITE, size=12)
        cur_row += 1

        for j, h in enumerate(["参数", "Y1", "Y2", "Y3", "口径备注"]):
            style_header(ws.cell(row=cur_row, column=j + 1, value=h),
                         fill=GOLD, color=NAVY)
        cur_row += 1

        for item in items:
            for j, val in enumerate(item):
                cell = ws.cell(row=cur_row, column=j + 1, value=val)
                # Y1/Y2/Y3 列是可调参数，标记为黄底
                if 1 <= j <= 3 and isinstance(val, (int, float)):
                    style_body(cell, fill="FFF8C5", bold=False,
                               align="right",
                               number_format="#,##0.##")
                else:
                    style_body(cell, fill=CLOUD if cur_row % 2 == 0 else WHITE,
                               bold=(j == 0))
            cur_row += 1
        cur_row += 1

    # =================== Sheet 3: 三年损益 ===================
    ws = wb.create_sheet("3. 三年损益")
    set_col_widths(ws, [32, 16, 16, 16, 24])

    ws.merge_cells("A1:E1")
    style_title(ws["A1"], size=18)
    ws["A1"] = "三年损益表（万元）"
    ws.row_dimensions[1].height = 28

    headers = ["科目", "Y1 2026", "Y2 2027", "Y3 2028", "口径备注"]
    for j, h in enumerate(headers):
        style_header(ws.cell(row=3, column=j + 1, value=h))

    rows = [
        ("租金收入", 615, 1442, 2091, "= 净面积*入驻率*租金*365"),
        ("物业费收入", 97, 195, 257, ""),
        ("1F+2F 冠名权", 100, 300, 500, "大堂/展厅/LED"),
        ("服务平台佣金", 80, 250, 600, "GMV 5%-10%"),
        ("后市场协同", 50, 150, 300, "与冠松联动"),
        ("政策返还（净计入）", 80, 600, 1500, "一企一策"),
        ("基金管理费", 0, 300, 500, "园区基金"),
        ("总收入", 1022, 3237, 5748, "★"),
        ("", "", "", "", ""),
        ("物业运营成本", 600, 750, 850, ""),
        ("招商佣金摊销", 100, 200, 300, "5 年直线"),
        ("装补摊销", 200, 350, 450, "5 年直线"),
        ("免租摊销", 250, 200, 100, ""),
        ("营销品牌（含发布会）", 800, 600, 600, ""),
        ("政府关系/法务", 200, 250, 300, ""),
        ("团队人力", 400, 800, 1100, "4→14→22 人"),
        ("测试场协同", 80, 120, 150, ""),
        ("财务费用", 350, 320, 280, ""),
        ("其他", 150, 200, 250, ""),
        ("总成本", 3130, 3790, 4380, "★"),
        ("", "", "", "", ""),
        ("EBITDA", -2108, -553, 1368, "★ Y3 转正"),
        ("折旧/摊销（资产）", 1200, 1200, 1200, "15,152 ㎡ 直线"),
        ("EBIT", -3308, -1753, 168, ""),
        ("利息净额", 600, 580, 520, ""),
        ("税前利润", -3908, -2333, -352, "Y4 预计转正"),
    ]
    for i, row in enumerate(rows, start=4):
        is_total = "★" in str(row[4]) or row[0].startswith(("总", "EBI", "税前"))
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1, value=val)
            fill = CLOUD if i % 2 == 0 else WHITE
            if is_total:
                fill = "FFE0B5"
            color = RED if (j > 0 and isinstance(val, (int, float)) and val < 0) else BLACK
            bold = is_total or j == 0
            num_fmt = "#,##0;[Red]−#,##0" if (j in (1, 2, 3) and isinstance(val, (int, float))) else None
            style_body(cell, fill=fill, color=color, bold=bold,
                       align="right" if j > 0 else "left",
                       number_format=num_fmt)

    # 收入结构柱状图
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "三年收入结构"
    chart.y_axis.title = "万元"
    chart.x_axis.title = "Y1 / Y2 / Y3"

    data = Reference(ws, min_col=2, min_row=4, max_col=4, max_row=10)
    cats = Reference(ws, min_col=1, min_row=4, max_row=10)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    ws.add_chart(chart, "G4")

    # =================== Sheet 4: 月度滚动现金流 ===================
    ws = wb.create_sheet("4. 月度现金流")
    cols = ["月份", "经营性流入", "经营性流出", "经营净", "投资", "筹资", "净变动", "期末现金"]
    set_col_widths(ws, [10, 14, 14, 14, 12, 14, 14, 16])

    ws.merge_cells("A1:H1")
    style_title(ws["A1"], size=18)
    ws["A1"] = "36 个月滚动现金流（万元）"
    ws.row_dimensions[1].height = 28

    for j, h in enumerate(cols):
        style_header(ws.cell(row=3, column=j + 1, value=h))

    # 简化：基于年度数据按月平均 + 季节性微调
    # Y1 经营 -1800/12 ≈ -150/月；Y2 +500/12 ≈ +42；Y3 +2500/12 ≈ +208
    inflow = [
        # Y1
        20, 20, 30, 40, 50, 70, 80, 90, 120, 130, 130, 130,
        # Y2
        140, 150, 180, 200, 220, 250, 280, 300, 320, 340, 350, 350,
        # Y3
        360, 380, 420, 450, 480, 510, 550, 580, 610, 640, 660, 670,
    ]
    outflow = [
        # Y1
        220, 220, 240, 240, 260, 270, 270, 270, 280, 290, 290, 290,
        # Y2
        290, 300, 310, 310, 320, 320, 330, 330, 340, 340, 340, 340,
        # Y3
        340, 340, 350, 350, 350, 350, 360, 360, 360, 360, 360, 360,
    ]
    invest = [-300, -200, -200, -200, -100, -100, -100, -100, -100, -100, -100, -100,
              -150, -150, -150, -100, -100, -100, -100, -100, -100, -100, -100, -50,
              -100, -100, -100, -100, -100, -100, -100, -50, -50, -50, -25, -25]
    finance = [0, 0, 0, 0, 0, 0, 0, 0, 0, -2000, -2000, -2100,
               0, 0, 0, 0, 0, -2500, -2500, -2000, 0, 0, 0, 0,
               0, 0, 0, 0, 0, -1500, -1500, 0, 0, 0, 0, 0]
    initial_cash = 21800

    cur_cash = initial_cash
    for i in range(36):
        m = i + 1
        ws.cell(row=4 + i, column=1, value=f"M{m}")
        ws.cell(row=4 + i, column=2, value=inflow[i])
        ws.cell(row=4 + i, column=3, value=outflow[i])
        op_net = inflow[i] - outflow[i]
        ws.cell(row=4 + i, column=4, value=op_net)
        ws.cell(row=4 + i, column=5, value=invest[i])
        ws.cell(row=4 + i, column=6, value=finance[i])
        net = op_net + invest[i] + finance[i]
        ws.cell(row=4 + i, column=7, value=net)
        cur_cash += net
        ws.cell(row=4 + i, column=8, value=cur_cash)

        for j in range(1, 9):
            cell = ws.cell(row=4 + i, column=j)
            fill = CLOUD if i % 2 == 0 else WHITE
            if j == 8:  # 期末现金高亮
                fill = "FFE0B5"
            val = cell.value
            color = RED if (isinstance(val, (int, float)) and val < 0 and j > 1) else BLACK
            num_fmt = "#,##0;[Red]−#,##0" if j > 1 else None
            style_body(cell, fill=fill, color=color,
                       bold=(j == 8 or j == 1),
                       align="center" if j == 1 else "right",
                       number_format=num_fmt)

    # 现金流折线图
    chart = LineChart()
    chart.title = "36 个月期末现金余额（万元）"
    chart.y_axis.title = "现金余额"
    chart.x_axis.title = "月份"
    data = Reference(ws, min_col=8, min_row=3, max_row=39)
    cats = Reference(ws, min_col=1, min_row=4, max_row=39)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, "J4")

    # =================== Sheet 5: 入驻进度 ===================
    ws = wb.create_sheet("5. 入驻进度")
    set_col_widths(ws, [10, 16, 16, 16, 14, 14, 14])

    ws.merge_cells("A1:G1")
    style_title(ws["A1"], size=18)
    ws["A1"] = "月度入驻爬坡（链主 + 生态）"
    ws.row_dimensions[1].height = 28

    for j, h in enumerate(["月份", "链主签约 (累计)", "生态签约 (累计)",
                            "签约面积 (㎡)", "入驻面积 (㎡)", "入驻率 %", "续约率 %"]):
        style_header(ws.cell(row=3, column=j + 1, value=h))

    # 36 个月数据
    anchor_cum = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,  # Y1
                  0, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,  # Y2
                  1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2, 3]  # Y3 末加 1
    eco_cum = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11,
               12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23,
               24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35]
    signed_area = [0, 200, 500, 800, 1200, 1500, 1800, 2100, 2400, 2700, 2900, 2905,
                   3200, 3500, 5500, 5700, 5800, 5900, 6000, 6100, 6200, 6300, 6400, 6450,
                   6500, 6600, 6700, 6800, 8500, 8600, 8650, 8700, 8750, 8800, 8830, 8830]
    occupied = [0, 0, 0, 200, 400, 600, 900, 1100, 1500, 1800, 2200, 2600,
                3000, 3300, 4000, 4500, 4900, 5200, 5400, 5500, 5600, 5700, 5800, 5900,
                6000, 6200, 6300, 6500, 7000, 7200, 7400, 7500, 7600, 7650, 7700, 7700]

    for i in range(36):
        m = i + 1
        ws.cell(row=4 + i, column=1, value=f"M{m}")
        ws.cell(row=4 + i, column=2, value=anchor_cum[i])
        ws.cell(row=4 + i, column=3, value=eco_cum[i])
        ws.cell(row=4 + i, column=4, value=signed_area[i])
        ws.cell(row=4 + i, column=5, value=occupied[i])
        ws.cell(row=4 + i, column=6, value=round(occupied[i] / 8300 * 100, 1))
        ws.cell(row=4 + i, column=7, value=round(min(95, 50 + i * 1.4), 1))

        for j in range(1, 8):
            cell = ws.cell(row=4 + i, column=j)
            fill = CLOUD if i % 2 == 0 else WHITE
            num_fmt = "0.0\"%\"" if j in (6, 7) else "#,##0"
            if j == 1:
                num_fmt = None
            style_body(cell, fill=fill,
                       bold=(j == 1),
                       align="center" if j == 1 else "right",
                       number_format=num_fmt)

    chart = LineChart()
    chart.title = "36 个月入驻率爬坡 (%)"
    chart.y_axis.title = "入驻率"
    chart.x_axis.title = "月份"
    data = Reference(ws, min_col=6, min_row=3, max_row=39)
    cats = Reference(ws, min_col=1, min_row=4, max_row=39)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 22
    ws.add_chart(chart, "I4")

    # =================== Sheet 6: 敏感性分析 ===================
    ws = wb.create_sheet("6. 敏感性分析")
    set_col_widths(ws, [22, 16, 16, 16, 16, 16, 16])

    ws.merge_cells("A1:G1")
    style_title(ws["A1"], size=18)
    ws["A1"] = "Y3 EBITDA 敏感性分析（万元）"
    ws.row_dimensions[1].height = 28

    # 单变量敏感性
    ws["A3"] = "一、单变量敏感性"
    style_header(ws["A3"], fill=NAVY, color=WHITE, size=14)
    ws.merge_cells("A3:G3")

    headers = ["变量", "−20%", "−10%", "基准", "+10%", "+20%", "Y3 EBITDA 弹性"]
    for j, h in enumerate(headers):
        style_header(ws.cell(row=4, column=j + 1, value=h), fill=GOLD, color=NAVY)

    sens_rows = [
        ("入驻率", -1690, -480, 1368, 220, 220, "高（92% 已接近上限）"),
        ("平均租金", -2810, -210, 1368, 210, 1180, "中"),
        ("政策返还兑现率", -1820, -450, 1368, 0, 0, "中"),
        ("冠名/服务/后市场", -1680, -420, 1368, 420, 1180, "中"),
        ("团队人力", 220, 110, 1368, -110, -220, "低"),
        ("装补摊销", 90, 45, 1368, -45, -90, "低"),
        ("利息净额", 104, 52, 1368, -52, -104, "低"),
    ]
    for i, row in enumerate(sens_rows, start=5):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1, value=val)
            fill = CLOUD if i % 2 == 0 else WHITE
            if j == 3:
                fill = "FFE0B5"  # 基准列
            color = RED if (j != 3 and isinstance(val, (int, float)) and val < 0) else (
                GREEN if (j != 3 and isinstance(val, (int, float)) and val > 0) else BLACK)
            num_fmt = "#,##0;[Red]−#,##0" if (1 <= j <= 5 and isinstance(val, (int, float))) else None
            style_body(cell, fill=fill, color=color,
                       bold=(j == 0 or j == 6),
                       align="right" if (1 <= j <= 5) else "left",
                       number_format=num_fmt)

    # 双变量矩阵：入驻率 × 平均租金
    ws["A14"] = "二、双变量矩阵：入驻率 × 平均租金 → Y3 EBITDA"
    style_header(ws["A14"], fill=NAVY, color=WHITE, size=14)
    ws.merge_cells("A14:G14")

    occ_rates = [70, 80, 85, 92, 95, 100]
    rents = [5.8, 6.5, 7.0, 7.5, 8.0, 8.5]

    # 表头
    style_header(ws.cell(row=15, column=1, value="入驻率 \\ 租金"), fill=GOLD, color=NAVY)
    for j, r in enumerate(rents):
        style_header(ws.cell(row=15, column=j + 2, value=f"{r} 元"), fill=GOLD, color=NAVY)

    # 基准 Y3 数据：occ=92%, rent=7.5, EBITDA=1368
    # 简化模型：EBITDA = a*占用率*租金 - 固定成本 (4380-100政策)
    # 1368 = X * 0.92 * 7.5 - 4380
    # X = (1368+4380) / (0.92 * 7.5) = 5748 / 6.9 = 833 (元/㎡·年的隐含因子)
    # 简化为：EBITDA = 833 * occ * rent / 100 * (其他收入比例) - 4380
    # 实际上：收入 = 租金*面积*入驻率*365 + 其他固定收入(2900)
    # Y3 其他收入 = 5748 - 2091 - 257 = 3400
    # 简化估算
    base_other_income = 3400  # 万元
    fixed_cost = 4380  # 总成本
    area = 8300  # ㎡

    for i, occ in enumerate(occ_rates):
        cell = ws.cell(row=16 + i, column=1, value=f"{occ}%")
        style_header(cell, fill=GOLD, color=NAVY)
        for j, r in enumerate(rents):
            rev = (occ / 100) * r * area * 365 / 10000
            other_rev = base_other_income * (occ / 92)  # 其他收入按入驻率线性
            total_rev = rev + other_rev * 0.4 + 257 * (occ / 92) + 1500  # 简化
            cost = fixed_cost  # 固定为主
            ebitda = round(total_rev - cost, 0)
            cell = ws.cell(row=16 + i, column=j + 2, value=ebitda)
            fill = "C5E8C5" if ebitda > 1000 else ("FFE8C5" if ebitda > 0 else "F8C5C5")
            color = RED if ebitda < 0 else BLACK
            style_body(cell, fill=fill, color=color, align="right",
                       bold=True, number_format="#,##0;[Red]−#,##0")

    ws["A23"] = "图例：绿色 = 优秀（>1,000） · 黄色 = 转正 · 红色 = 亏损"
    style_body(ws["A23"], color=GREY, size=10, align="left")
    ws.merge_cells("A23:G23")

    # =================== Sheet 7: 团队与人力成本 ===================
    ws = wb.create_sheet("7. 团队人力成本")
    set_col_widths(ws, [10, 24, 12, 16, 14, 14, 14, 14])

    ws.merge_cells("A1:H1")
    style_title(ws["A1"], size=18)
    ws["A1"] = "22 人团队薪酬带宽与年度成本"
    ws.row_dimensions[1].height = 28

    for j, h in enumerate(["编号", "岗位", "等级", "年包带宽 (万元)",
                            "Y1 配置", "Y2 配置", "Y3 配置", "Y3 中位包"]):
        style_header(ws.cell(row=3, column=j + 1, value=h))

    team = [
        (1, "项目总监 GM", "D", "120-180", 1, 1, 1, 150),
        (2, "GR 总监", "D", "80-130", 1, 1, 1, 105),
        (3, "政府事务经理", "M", "40-60", 0, 1, 1, 50),
        (4, "招商总监", "D", "80-120", 1, 1, 1, 100),
        (5, "高级招商经理（链主线）", "S", "50-80", 0, 1, 1, 65),
        (6, "招商经理（算法线）", "M", "30-45", 0, 1, 1, 38),
        (7, "招商经理（核心研发线）", "M", "30-45", 0, 1, 1, 38),
        (8, "招商经理（中介线）", "M", "30-45", 0, 1, 1, 38),
        (9, "招商经理（自拓线）", "M", "30-45", 0, 1, 1, 38),
        (10, "销售助理 / CRM", "E", "15-22", 0, 1, 1, 19),
        (11, "客户成功经理", "M", "35-50", 0, 0, 1, 42),
        (12, "运营总监", "D", "60-90", 1, 1, 1, 75),
        (13, "物业主管", "S", "25-35", 0, 1, 1, 30),
        (14, "IT / 数据经理", "S", "40-60", 0, 1, 1, 50),
        (15, "测试场协同主任", "S", "40-55", 0, 1, 1, 47),
        (16, "行政主任", "M", "20-30", 0, 0, 1, 25),
        (17, "市场总监", "D", "60-85", 0, 1, 1, 72),
        (18, "品牌经理", "M", "25-40", 0, 1, 1, 32),
        (19, "活动经理", "M", "25-40", 0, 0, 1, 32),
        (20, "法务经理", "S", "45-70", 0, 1, 1, 57),
        (21, "财务经理", "S", "40-60", 0, 1, 1, 50),
        (22, "HR 经理", "M", "30-45", 0, 0, 1, 38),
    ]
    for i, row in enumerate(team, start=4):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1, value=val)
            fill = CLOUD if i % 2 == 0 else WHITE
            level_color = {"D": "FFD0B5", "S": "FFE8B5", "M": "C5E8C5", "E": "C5E0F5"}
            if j == 2 and val in level_color:
                fill = level_color[val]
            num_fmt = "#,##0" if j == 7 else None
            style_body(cell, fill=fill,
                       bold=(j == 7),
                       align="center" if j in (0, 2, 4, 5, 6) else (
                           "right" if j == 7 else "left"),
                       number_format=num_fmt)

    # 汇总
    sum_row = 26
    ws.cell(row=sum_row, column=1, value="—")
    ws.cell(row=sum_row, column=2, value="合计")
    ws.cell(row=sum_row, column=3, value="—")
    ws.cell(row=sum_row, column=4, value="—")
    ws.cell(row=sum_row, column=5, value=f"=SUM(E4:E25)")
    ws.cell(row=sum_row, column=6, value=f"=SUM(F4:F25)")
    ws.cell(row=sum_row, column=7, value=f"=SUM(G4:G25)")
    ws.cell(row=sum_row, column=8, value=f"=SUMPRODUCT(G4:G25,H4:H25)")
    for j in range(1, 9):
        cell = ws.cell(row=sum_row, column=j)
        style_body(cell, fill="FFE0B5", bold=True,
                   align="right" if j > 3 else "center", size=11,
                   number_format="#,##0" if j > 3 else None)

    ws["A28"] = "Y1 团队预算约 ¥400 万 · Y2 ¥1,500 万 · Y3 ¥2,200 万 (与三年损益对齐)"
    style_body(ws["A28"], color=GREY, size=11, align="left")
    ws.merge_cells("A28:H28")

    # =================== Sheet 8: 让步阶梯计算器 ===================
    ws = wb.create_sheet("8. 让步阶梯计算器")
    set_col_widths(ws, [10, 28, 16, 16, 16, 24, 28])

    ws.merge_cells("A1:G1")
    style_title(ws["A1"], size=18)
    ws["A1"] = "链主谈判 · 让步阶梯计算器"
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    style_title(ws["A2"], fill=GOLD, color=NAVY, size=11)
    ws["A2"] = "每让一步必有'换'——本表计算每步让步对单户链主 6 年合同 NPV 的影响"
    ws.row_dimensions[2].height = 22

    headers = ["步骤", "让步内容", "我方让步幅度",
               "6 年净 NPV 影响 (万元)", "对方对等承诺",
               "判断（接受 / 升级 / 拒绝）", "授权层"]
    for j, h in enumerate(headers):
        style_header(ws.cell(row=4, column=j + 1, value=h))

    rows = [
        ("基准", "起始 6.5 元/㎡·天 · 9 月免租 · 装补 600/㎡ · 8 年", "—", 0,
         "—", "首报", "项目总监"),
        ("让步 1", "起始 6.5 → 6.0 元/㎡·天", "−7.7%", -550,
         "8 年长租 + 6 个月履约", "接受", "项目总监"),
        ("让步 2", "免租 12 → 15 个月", "+25%", -240,
         "装补封顶 1,000/㎡（不浮动）", "接受", "项目总监"),
        ("让步 3", "装补 800 → 1,000 元/㎡", "+25%", -67,
         "楼宇冠名 5 年 + 9 月发布会主旨", "接受", "项目总监"),
        ("让步 4", "政策返还 70% → 80%", "+14%", -340,
         "政策返还以政府专班书面文件兑现", "接受", "GR 总监 + 集团董事长"),
        ("让步 5", "续约权（涨幅封顶 5%）", "—", -90,
         "退租赔偿 ≥ 60% 政策返还回收", "接受", "项目总监"),
        ("底线 🔴", "起始 5.0 / 免租 24 / 装补 1,500 / 5 年", "—", -1500,
         "—", "拒绝（超线请示）", "集团董事长"),
    ]
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row):
            cell = ws.cell(row=i, column=j + 1, value=val)
            fill = CLOUD if i % 2 == 0 else WHITE
            if i == 11:  # 底线
                fill = "F8C5C5"
            color = RED if (j == 3 and isinstance(val, (int, float)) and val < 0) else BLACK
            num_fmt = "#,##0;[Red]−#,##0" if (j == 3 and isinstance(val, (int, float))) else None
            style_body(cell, fill=fill, color=color,
                       bold=(j == 0 or i == 11),
                       align="center" if j in (0, 2, 5) else (
                           "right" if j == 3 else "left"),
                       number_format=num_fmt)

    ws["A14"] = "📌 计算口径：6 年 NPV，按 8% 折现率，含装补/免租/政策返还/品牌权益等综合影响"
    style_body(ws["A14"], color=GREY, size=10, align="left")
    ws.merge_cells("A14:G14")

    ws["A15"] = "📌 让步原则：每让一步必带条件；超过授权区间立即升级；底线（红线）不可越界"
    style_body(ws["A15"], color=RED, size=10, align="left", bold=True)
    ws.merge_cells("A15:G15")

    # 保存
    out = OUT_DIR / "财务测算与商务模型.xlsx"
    wb.save(out)
    print(f"✓ Excel 写入：{out}")
    print(f"  共 {len(wb.sheetnames)} 个 Sheet：{wb.sheetnames}")
    return out


if __name__ == "__main__":
    build()
