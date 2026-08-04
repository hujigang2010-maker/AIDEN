#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将多个 ChinaJoy 招商引资 Excel 版本融合为全量资料补充材料。

优先级（同字段冲突时）：全量资料补充版 > 增强完善版 > 联系方式补充 > 终版 > 招商引资数据库
对标结构：全量资料补充版；其余版本仅做空值补齐与独有内容追加。
"""

from __future__ import annotations

import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "output" / "chinajoy_merge_sources"
REPO_XLSX = ROOT / "output" / "2026ChinaJoy_招商引资参展商名录.xlsx"
OUT = ROOT / "output" / "ChinaJoy2026_招商引资总表_六版融合全量版.xlsx"

# 视觉：深蓝商务风（沿用全量版主色，避免紫色模板感）
C_HEADER = "1F4E79"
C_TITLE = "2E5496"
C_SUB = "D6E3F0"
C_ALT = "F5F8FC"
C_WHITE = "FFFFFF"
C_YELLOW = "FFF2CC"
C_GREEN = "E2EFDA"
C_ORANGE = "FCE4D6"
C_RED_SOFT = "F8CBAD"
C_A = "C6EFCE"
C_B = "FFEB9C"
C_C = "D9D9D9"
C_TOUCH = "DDEBF7"

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def norm_name(s: object) -> str:
    t = str(s or "").strip().lower()
    t = t.replace("（", "(").replace("）", ")")
    t = re.sub(r"\s+", "", t)
    t = t.replace("有限公司", "").replace("股份有限公司", "")
    return t


def has_phone(s: object) -> bool:
    t = str(s or "").strip()
    if not t or t in {"-", "无", "N/A"}:
        return False
    if t.startswith("电玩") or t.startswith("环球"):
        return False
    return bool(re.search(r"\d{5,}", t))


def has_email(s: object) -> bool:
    return "@" in str(s or "")


def nonempty(s: object) -> bool:
    return s is not None and str(s).strip() not in {"", "-", "None", "nan"}


def richer(a: object, b: object) -> object:
    """二选一：优先非空且更长的文本。"""
    if not nonempty(a):
        return b if nonempty(b) else a
    if not nonempty(b):
        return a
    sa, sb = str(a).strip(), str(b).strip()
    if has_phone(sb) and not has_phone(sa):
        return b
    if has_email(sb) and not has_email(sa):
        return b
    return b if len(sb) > len(sa) else a


def read_table(path: Path, sheet: str, header_keywords: tuple[str, ...] = ("序号", "品牌/企业", "姓名")):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        vals = [str(c).strip() if c is not None else "" for c in r]
        if any(k in vals for k in header_keywords):
            header_idx = i
            break
    headers_raw = list(rows[header_idx])
    while headers_raw and not nonempty(headers_raw[-1]):
        headers_raw.pop()
    headers = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(headers_raw)]
    data = []
    for r in rows[header_idx + 1 :]:
        if all(not nonempty(c) for c in r[: len(headers)]):
            continue
        data.append({headers[j]: (r[j] if j < len(r) else None) for j in range(len(headers))})
    return headers, data


def style_header_row(ws: Worksheet, row: int, start_col: int, end_col: int):
    fill = PatternFill("solid", fgColor=C_HEADER)
    font = Font(name="微软雅黑", size=10, bold=True, color=C_WHITE)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def style_data_cell(cell, alt: bool = False, fill_color: str | None = None):
    cell.font = Font(name="微软雅黑", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = THIN
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    elif alt:
        cell.fill = PatternFill("solid", fgColor=C_ALT)


def autosize(ws: Worksheet, min_w: int = 8, max_w: int = 42, sample_rows: int = 40):
    for col in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for row in range(1, min((ws.max_row or 1), sample_rows) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            maxlen = max(maxlen, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, maxlen + 2))


def apply_table_polish(ws: Worksheet, header_row: int, priority_col: int | None = None, phone_cols: list[int] | None = None):
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    style_header_row(ws, header_row, 1, max_col)
    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
    for r in range(header_row + 1, max_row + 1):
        alt = (r - header_row) % 2 == 0
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            fill = None
            if priority_col and c == priority_col:
                pv = str(cell.value or "")
                if pv.startswith("A"):
                    fill = C_A
                elif pv.startswith("B"):
                    fill = C_B
                elif pv.startswith("C"):
                    fill = C_C
            if phone_cols and c in phone_cols and has_phone(cell.value):
                fill = C_GREEN
            style_data_cell(cell, alt=alt, fill_color=fill)


def load_priority_versions() -> list[tuple[str, Path]]:
    return [
        ("全量资料补充版", SRC / "全量资料补充版.xlsx"),
        ("增强完善版", SRC / "增强完善版.xlsx"),
        ("联系方式补充", SRC / "联系方式补充.xlsx"),
        ("终版", SRC / "终版.xlsx"),
        ("招商引资数据库", SRC / "招商引资数据库.xlsx"),
    ]


def merge_sheet03_fields(base_wb) -> dict:
    """对 03 表按品牌做字段级补齐，返回统计。"""
    ws = base_wb["03 参展商总名录·精编422"]
    # 找到表头行
    header_row = 3
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    while headers and headers[-1] is None:
        headers.pop()
    col_idx = {str(h): i + 1 for i, h in enumerate(headers) if h}

    # 读取各版本 03 / 数据库
    sources: list[tuple[str, list[dict]]] = []
    for label, path in load_priority_versions()[:4]:
        if not path.exists():
            continue
        _, data = read_table(path, "03 参展商总名录·精编422")
        sources.append((label, data))

    db_path = SRC / "招商引资数据库.xlsx"
    db_map: dict[str, dict] = {}
    if db_path.exists():
        _, db_rows = read_table(db_path, "参展商数据库", ("ID", "品牌/展商"))
        for r in db_rows:
            key = norm_name(r.get("品牌/展商"))
            if key:
                db_map[key] = r

    fill_count = 0
    db_fill = 0
    brand_col = col_idx.get("品牌/企业")
    merge_fields = [
        "联系电话",
        "公开渠道线索",
        "信息来源",
        "核实状态",
        "备注",
        "新库招商总分",
        "新库优先级",
        "招商匹配评分",
        "活动激活潜力",
        "建议招商切入点",
        "新库联系人",
        "新库电话/微信",
        "新库邮箱",
        "官网/公开渠道",
        "置信度及来源",
        "规模层级",
        "企业类型",
        "主营产品/代表作品",
        "活动与嘉宾",
        "招商价值标签",
        "潜在合作方向",
    ]

    # 建立其他版本索引（低优先级在前，高优先级后写覆盖空值时用 richer）
    # 实际：从低到高依次 richer 到当前单元格
    other_maps: list[tuple[str, dict[str, dict]]] = []
    for label, data in reversed(sources[1:]):  # 终版→联系补充→增强（全量已是底）
        m = {norm_name(r.get("品牌/企业")): r for r in data if r.get("品牌/企业")}
        other_maps.append((label, m))

    for row in range(header_row + 1, ws.max_row + 1):
        brand = ws.cell(row=row, column=brand_col).value if brand_col else None
        if not brand:
            continue
        key = norm_name(brand)
        for _label, mmap in other_maps:
            src = mmap.get(key)
            if not src:
                continue
            for field in merge_fields:
                if field not in col_idx or field not in src:
                    continue
                cell = ws.cell(row=row, column=col_idx[field])
                new_v = richer(cell.value, src.get(field))
                if new_v != cell.value and nonempty(new_v):
                    cell.value = new_v
                    fill_count += 1

        # 数据库字段映射补齐
        db = db_map.get(key)
        if not db:
            # 弱匹配：去空格包含
            for dk, dv in db_map.items():
                if key and dk and (key in dk or dk in key):
                    db = dv
                    break
        if db:
            mapping = {
                "新库招商总分": "招商总分",
                "新库优先级": "优先级",
                "招商匹配评分": "匹配评分(1-5)",
                "活动激活潜力": "激活潜力(1-5)",
                "建议招商切入点": "建议招商切入点",
                "官网/公开渠道": "官网",
                "置信度及来源": "置信度",
                "公开渠道线索": "微博/公众号/小红书/视频号",
            }
            for dst, src_k in mapping.items():
                if dst not in col_idx:
                    continue
                cell = ws.cell(row=row, column=col_idx[dst])
                new_v = richer(cell.value, db.get(src_k))
                if new_v != cell.value and nonempty(new_v):
                    cell.value = new_v
                    db_fill += 1

    # 从 sheet15 追加独立触达列（若尚无）
    extra_headers = [
        "触达邮箱",
        "工商主体",
        "统一社会信用代码",
        "法定代表人",
        "注册地址",
        "主电话(工商资料)",
        "其他电话(工商资料)",
        "触达完整度",
    ]
    start_extra = len(headers) + 1
    for i, h in enumerate(extra_headers):
        cell = ws.cell(row=header_row, column=start_extra + i, value=h)
        cell.fill = PatternFill("solid", fgColor=C_TITLE)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=C_WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

    s15 = base_wb["15 企业联系方式与资料"]
    s15_rows = list(s15.iter_rows(values_only=True))
    s15_h = [str(c) if c else f"c{i}" for i, c in enumerate(s15_rows[0])]
    s15_map = {}
    for r in s15_rows[1:]:
        if not any(nonempty(c) for c in r):
            continue
        d = {s15_h[i]: r[i] if i < len(r) else None for i in range(len(s15_h))}
        s15_map[norm_name(d.get("原品牌/企业"))] = d

    for row in range(header_row + 1, ws.max_row + 1):
        brand = ws.cell(row=row, column=brand_col).value if brand_col else None
        info = s15_map.get(norm_name(brand), {})
        email = info.get("邮箱")
        remark = str(ws.cell(row=row, column=col_idx.get("备注", 16)).value or "")
        m = re.search(r"[\w.+-]+@[\w.-]+\.\w+", remark)
        if not has_email(email) and m:
            email = m.group(0)
        phone_main = info.get("主电话") or ws.cell(row=row, column=col_idx["联系电话"]).value
        values = [
            email,
            info.get("实际匹配主体") or info.get("Claude规范主体"),
            info.get("统一社会信用代码"),
            info.get("法定代表人"),
            info.get("注册地址"),
            info.get("主电话"),
            info.get("其他电话"),
            None,
        ]
        score = 0
        if has_phone(phone_main) or has_phone(values[5]):
            score += 2
        if has_email(email):
            score += 2
        if nonempty(values[1]):
            score += 1
        if nonempty(ws.cell(row=row, column=col_idx.get("官网/公开渠道", 1)).value) or nonempty(
            ws.cell(row=row, column=col_idx.get("公开渠道线索", 1)).value
        ):
            score += 1
        values[7] = f"{score}/6"
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=start_extra + i, value=v)
            style_data_cell(cell, alt=(row % 2 == 0), fill_color=C_GREEN if i in (0, 5) and nonempty(v) else None)

    return {"field_fills": fill_count, "db_fills": db_fill, "extra_cols": len(extra_headers)}


def enrich_contacts_sheet(base_wb) -> int:
    """向 01 追加仓库/官网独有联系人。"""
    if not REPO_XLSX.exists():
        return 0
    wb = load_workbook(REPO_XLSX, data_only=True)
    rows = list(wb["官网联系方式"].iter_rows(values_only=True))
    wb.close()
    repo_contacts = rows[1:]

    ws = base_wb["01 招商联系方式"]
    existing = "\n".join(
        "|".join(str(c) if c is not None else "" for c in r) for r in ws.iter_rows(values_only=True)
    )

    # 找「官方招商与分板块联系人」表头后的数据末尾（网络渠道之前）
    insert_at = None
    network_row = None
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        # 避免误匹配封面导语里的「15个网络渠道」
        if v.startswith("网络渠道") or "网络渠道（" in v or "网络渠道(" in v:
            network_row = r
            break
    if network_row is None:
        insert_at = ws.max_row + 2
    else:
        insert_at = network_row

    added = 0
    # 在网络渠道前插入区块
    block_title_row = insert_at
    # 先腾地方：插入足够行
    to_add = []
    for r in repo_contacts:
        cat, name, phone, email, wechat, area, source, usage = (list(r) + [None] * 8)[:8]
        tokens = [t for t in [phone, email, wechat, name] if nonempty(t)]
        if not tokens:
            continue
        if any(str(t) in existing for t in tokens):
            # 补微信到备注场景：若姓名+邮箱在但微信息不在，仍追加一行标注微信
            if nonempty(wechat) and str(wechat) not in existing:
                pass
            else:
                continue
        to_add.append((cat, name, phone, email, wechat, area, source, usage))

    if not to_add:
        return 0

    need = len(to_add) + 3
    ws.insert_rows(block_title_row, amount=need)
    r = block_title_row
    ws.cell(row=r, column=1, value="【融合补录】官网/指定服务商联系人（来自仓库官网联系方式表）")
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=11, bold=True, color=C_TITLE)
    r += 1
    headers = ["类别/板块", "联系人", "电话", "邮箱/微信", "适用场景/展区", "备注来源"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = PatternFill("solid", fgColor=C_HEADER)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=C_WHITE)
        cell.border = THIN
    r += 1
    for cat, name, phone, email, wechat, area, source, usage in to_add:
        contact = " / ".join([x for x in [str(email or "").strip(), f"微信:{wechat}" if nonempty(wechat) else ""] if x])
        vals = [cat, name, phone, contact or email, area or usage, source]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            style_data_cell(cell, fill_color=C_ORANGE)
        r += 1
        added += 1
        existing += f"|{phone}|{email}|{wechat}|{name}"
    return added


def add_reach_priority_sheet(base_wb) -> int:
    """新建「可触达优先清单」——园区招商最关心的触达排序表。"""
    # 删除旧表若重跑
    if "可触达优先清单" in base_wb.sheetnames:
        del base_wb["可触达优先清单"]
    ws = base_wb.create_sheet("可触达优先清单", 1)

    ws.merge_cells("A1:L1")
    ws["A1"] = "可触达优先清单 · 园区招商引资触达排序"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color=C_TITLE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        "筛选规则：有电话/邮箱优先；同档按新库优先级与招商总分排序。"
        "用途：园区BD可直接按此表外呼/发函，再回写「16 招商跟进工作台」。"
    )
    ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")

    headers = [
        "排序",
        "触达等级",
        "品牌/企业",
        "参展分类",
        "展位号",
        "规模层级",
        "新库优先级",
        "招商总分",
        "联系电话",
        "触达邮箱",
        "官网/渠道",
        "建议招商切入点",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, len(headers))

    src = base_wb["03 参展商总名录·精编422"]
    hrow = 3
    headers3 = [src.cell(row=hrow, column=c).value for c in range(1, src.max_column + 1)]
    idx = {str(h): i + 1 for i, h in enumerate(headers3) if h}

    rows = []
    for r in range(hrow + 1, src.max_row + 1):
        brand = src.cell(row=r, column=idx["品牌/企业"]).value
        if not brand:
            continue
        phone = src.cell(row=r, column=idx.get("联系电话", 1)).value
        email = src.cell(row=r, column=idx.get("触达邮箱", 1)).value if "触达邮箱" in idx else None
        phone2 = src.cell(row=r, column=idx.get("主电话(工商资料)", 1)).value if "主电话(工商资料)" in idx else None
        use_phone = phone if has_phone(phone) else phone2
        website = src.cell(row=r, column=idx.get("官网/公开渠道", 1)).value
        channel = src.cell(row=r, column=idx.get("公开渠道线索", 1)).value
        prio = str(src.cell(row=r, column=idx.get("新库优先级", 1)).value or "")
        score = src.cell(row=r, column=idx.get("新库招商总分", 1)).value
        try:
            score_n = float(score) if score is not None and str(score).strip() != "" else -1
        except Exception:
            score_n = -1
        touch = 0
        if has_phone(use_phone):
            touch += 2
        if has_email(email):
            touch += 2
        if nonempty(website) or nonempty(channel):
            touch += 1
        if touch == 0:
            continue
        if touch >= 4:
            grade = "S-可直接触达"
        elif touch >= 3:
            grade = "A-电话或邮箱"
        else:
            grade = "B-渠道跟进"
        prio_rank = 0 if prio.startswith("A") else (1 if prio.startswith("B") else 2)
        rows.append(
            (
                -touch,
                prio_rank,
                -score_n,
                grade,
                brand,
                src.cell(row=r, column=idx.get("参展分类", 1)).value,
                src.cell(row=r, column=idx.get("展位号", 1)).value,
                src.cell(row=r, column=idx.get("规模层级", 1)).value,
                prio,
                score,
                use_phone,
                email,
                website or channel,
                src.cell(row=r, column=idx.get("建议招商切入点", 1)).value,
            )
        )

    rows.sort()
    for i, item in enumerate(rows, 1):
        (_, _, _, grade, brand, cat, booth, scale, prio, score, phone, email, web, tip) = item
        vals = [i, grade, brand, cat, booth, scale, prio, score, phone, email, web, tip]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=3 + i, column=c, value=v)
            fill = None
            if c == 2:
                fill = C_A if grade.startswith("S") else (C_TOUCH if grade.startswith("A") else C_B)
            if c == 7 and str(prio).startswith("A"):
                fill = C_A
            if c in (9, 10) and nonempty(v):
                fill = C_GREEN
            style_data_cell(cell, alt=i % 2 == 0, fill_color=fill)

    apply_table_polish(ws, 3, priority_col=7, phone_cols=[9, 10])
    autosize(ws)
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["L"].width = 36
    return len(rows)


def add_followup_sheet(base_wb) -> None:
    if "16 招商跟进工作台" in base_wb.sheetnames:
        del base_wb["16 招商跟进工作台"]
    ws = base_wb.create_sheet("16 招商跟进工作台")
    ws.merge_cells("A1:N1")
    ws["A1"] = "招商跟进工作台（可直接填报）· 融合自《招商引资数据库》"
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color=C_TITLE)

    headers = [
        "品牌/展商",
        "分类",
        "联系人",
        "职务",
        "电话/微信",
        "邮箱",
        "首次接触日",
        "当前阶段",
        "需求/合作方向",
        "预计投资/合作额(万元)",
        "下次跟进日",
        "负责人",
        "备注",
        "数据来源",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 1, len(headers))

    # 预填：可触达优先清单前 80 家
    reach = base_wb["可触达优先清单"]
    r_out = 3
    for r in range(4, min(reach.max_row, 83) + 1):
        brand = reach.cell(row=r, column=3).value
        cat = reach.cell(row=r, column=4).value
        phone = reach.cell(row=r, column=9).value
        email = reach.cell(row=r, column=10).value
        tip = reach.cell(row=r, column=12).value
        vals = [
            brand,
            cat,
            "",
            "",
            phone,
            email,
            "",
            "未联系",
            tip,
            "",
            "",
            "",
            "",
            "六版融合·可触达优先清单",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_out, column=c, value=v)
            fill = C_YELLOW if c in (3, 4, 7, 8, 10, 11, 12, 13) else None
            style_data_cell(cell, alt=r_out % 2 == 0, fill_color=fill)
        r_out += 1

    # 再附加数据库工作台空模板说明行
    ws.cell(row=r_out + 1, column=1, value="说明：黄色列为建议人工补录；请外呼/加微信后回写进度。")
    ws.cell(row=r_out + 1, column=1).font = Font(name="微软雅黑", size=9, italic=True, color="666666")
    autosize(ws)
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:N{r_out - 1}"


def add_db_unmatched_sheet(base_wb) -> int:
    """把数据库中未能并入精编的线索写入补录表。"""
    db_path = SRC / "招商引资数据库.xlsx"
    if not db_path.exists():
        return 0
    _, db_rows = read_table(db_path, "参展商数据库", ("ID", "品牌/展商"))
    ws3 = base_wb["03 参展商总名录·精编422"]
    hrow = 3
    brand_col = None
    for c in range(1, ws3.max_column + 1):
        if ws3.cell(row=hrow, column=c).value == "品牌/企业":
            brand_col = c
            break
    existing = {norm_name(ws3.cell(row=r, column=brand_col).value) for r in range(hrow + 1, ws3.max_row + 1)}

    unmatched = []
    for r in db_rows:
        b = str(r.get("品牌/展商") or "").strip()
        if not b:
            continue
        key = norm_name(b)
        if key in existing:
            continue
        if any(key in e or e in key for e in existing if e):
            continue
        unmatched.append(r)

    if "17 数据库未匹配线索" in base_wb.sheetnames:
        del base_wb["17 数据库未匹配线索"]
    ws = base_wb.create_sheet("17 数据库未匹配线索")
    ws.merge_cells("A1:J1")
    ws["A1"] = f"数据库独有线索（未出现在精编421）· 共 {len(unmatched)} 条 · 建议作为补录池"
    ws["A1"].font = Font(name="微软雅黑", size=13, bold=True, color=C_TITLE)

    headers = [
        "品牌/展商",
        "展馆",
        "参展分类",
        "展位号",
        "业务类型",
        "招商总分",
        "优先级",
        "建议招商切入点",
        "官网",
        "置信度",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, 1, len(headers))

    for i, r in enumerate(unmatched, 1):
        vals = [
            r.get("品牌/展商"),
            r.get("展馆"),
            r.get("参展分类"),
            r.get("展位号"),
            r.get("业务类型"),
            r.get("招商总分"),
            r.get("优先级"),
            r.get("建议招商切入点"),
            r.get("官网"),
            r.get("置信度"),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=2 + i, column=c, value=v)
            fill = C_A if str(r.get("优先级") or "").startswith("A") else None
            style_data_cell(cell, alt=i % 2 == 0, fill_color=fill)
    apply_table_polish(ws, 2, priority_col=7)
    autosize(ws)
    return len(unmatched)


def add_merge_notes_sheet(base_wb, stats: dict) -> None:
    if "00 融合说明" in base_wb.sheetnames:
        del base_wb["00 融合说明"]
    ws = base_wb.create_sheet("00 融合说明", 0)
    ws["A1"] = "ChinaJoy 2026 招商引资总表 · 六版融合全量版"
    ws["A1"].font = Font(name="微软雅黑", size=18, bold=True, color=C_TITLE)
    ws.merge_cells("A1:B1")

    lines = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("对标底表", "ChinaJoy2026_招商引资总表_全量资料补充版.xlsx"),
        ("融合优先级", "全量资料补充版 > 增强完善版 > 联系方式补充 > 终版 > 招商引资数据库（副本与正本相同，只取一份）"),
        ("融合原则", "底表结构不变；其他版本仅补空值/更长文本；独有联系人与未匹配线索另表追加"),
        ("园区用途", "先看「可触达优先清单」外呼，再用「01 招商联系方式」找主办对接，进度写入「16 招商跟进工作台」"),
        ("03字段补齐次数", stats.get("field_fills")),
        ("数据库字段回填次数", stats.get("db_fills")),
        ("01补录联系人", stats.get("contacts_added")),
        ("可触达企业数", stats.get("reach_count")),
        ("数据库未匹配线索", stats.get("unmatched")),
        ("新增列", "触达邮箱/工商主体/信用代码/法人/注册地址/工商电话/触达完整度"),
        ("样式", "深蓝表头、绿底=可拨打/可邮、黄底=待人工补录、A/B/C优先级色标、冻结首行首列+筛选"),
        ("注意", "公开信息整理，外联前请二次核验；不代表官方授权名录"),
    ]
    ws["A3"] = "项目"
    ws["B3"] = "说明"
    style_header_row(ws, 3, 1, 2)
    for i, (k, v) in enumerate(lines, 4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        style_data_cell(ws.cell(row=i, column=1), alt=i % 2 == 0, fill_color=C_SUB)
        style_data_cell(ws.cell(row=i, column=2), alt=i % 2 == 0)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 88


def update_cover(base_wb, stats: dict) -> None:
    ws = base_wb["封面·使用说明"]
    ws["B2"] = "ChinaJoy 2026 招商引资总表 · 六版融合全量版"
    ws["B2"].font = Font(name="微软雅黑", size=20, bold=True, color=C_TITLE)
    # B4 可能在合并区内，只改左上角单元格
    ws["B4"] = (
        f"对标全量资料补充版，融合增强/联系补充/终版/数据库；"
        f"可触达企业 {stats.get('reach_count')} 家｜补录联系人 {stats.get('contacts_added')} ｜"
        f"数据库未匹配线索 {stats.get('unmatched')}"
    )
    # 导航追加：插在原导航末行之后、页脚之前
    nav_start = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value or "") == "工作表导航":
            nav_start = r
            break
    if nav_start:
        last = nav_start
        for r in range(nav_start + 1, ws.max_row + 1):
            val = ws.cell(row=r, column=2).value
            if nonempty(val) and not str(val).startswith("生成日期"):
                last = r
            else:
                break
        extras = [
            ("00 融合说明", "版本优先级与使用路径"),
            ("可触达优先清单", "★园区招商首选：有电话/邮箱的企业排序"),
            ("16 招商跟进工作台", "外呼进度填报（黄格人工补录）"),
            ("17 数据库未匹配线索", "精编外补录池"),
        ]
        # 解除可能挡住的合并，并插入行
        to_unmerge = [str(m) for m in ws.merged_cells.ranges if m.min_row > last]
        for m in to_unmerge:
            ws.unmerge_cells(m)
        ws.insert_rows(last + 1, amount=len(extras))
        for i, (name, desc) in enumerate(extras, 1):
            row = last + i
            c2 = ws.cell(row=row, column=2, value=name)
            c3 = ws.cell(row=row, column=3, value=desc)
            c2.font = Font(name="微软雅黑", size=10, bold=True)
            c3.font = Font(name="微软雅黑", size=9)
            c2.fill = PatternFill("solid", fgColor=C_ORANGE)
            c3.fill = PatternFill("solid", fgColor=C_ORANGE)
        # 更新页脚
        for r in range(last + len(extras) + 1, ws.max_row + 2):
            v = str(ws.cell(row=r, column=2).value or "")
            if v.startswith("生成日期") or "数据来源" in v:
                ws.cell(
                    row=r,
                    column=2,
                    value="生成日期：2026-08-04  ·  六版融合全量版  ·  详见「00 融合说明」「13 数据来源与声明」",
                )
                break


def polish_key_sheets(base_wb) -> None:
    # 对关键表重新套用表头样式与筛选
    targets = {
        "03 参展商总名录·精编422": 3,
        "04 参展商全量·官网966": 3,
        "15 企业联系方式与资料": 1,
        "05 活动与特别嘉宾": 3,
        "06 明星·社团·嘉宾": 3,
        "08 网络渠道": 3,
        "02 参展分类总览": 3,
        "09 招商价值分析": 3,
        "10 高价值目标·产业集群": 3,
        "11 招商策略与跟进": 3,
    }
    for name, hrow in targets.items():
        if name not in base_wb.sheetnames:
            continue
        ws = base_wb[name]
        max_col = ws.max_column or 1
        # 确认 hrow 是表头
        vals = [ws.cell(row=hrow, column=c).value for c in range(1, min(8, max_col) + 1)]
        if not any(v in ("序号", "参展分类", "品牌/企业", "企业中文名", "原品牌/企业") for v in vals):
            # 尝试找
            found = None
            for r in range(1, 6):
                vals = [ws.cell(row=r, column=c).value for c in range(1, min(8, max_col) + 1)]
                if any(v in ("序号", "参展分类", "品牌/企业", "企业中文名", "原品牌/企业") for v in vals):
                    found = r
                    break
            if found:
                hrow = found
            else:
                continue
        prio_col = None
        phone_cols = []
        for c in range(1, max_col + 1):
            h = str(ws.cell(row=hrow, column=c).value or "")
            if "优先级" in h:
                prio_col = c
            if "电话" in h or "邮箱" in h:
                phone_cols.append(c)
        apply_table_polish(ws, hrow, priority_col=prio_col, phone_cols=phone_cols)
        # 标题行
        if hrow >= 2:
            title = ws.cell(row=1, column=1)
            if title.value:
                title.font = Font(name="微软雅黑", size=14, bold=True, color=C_TITLE)


def merge_db_channels_and_events(base_wb) -> dict:
    """把数据库里可能更完整的渠道/活动追加到现有表（按名称去重）。"""
    db = SRC / "招商引资数据库.xlsx"
    stats = {"channels": 0, "events": 0}
    if not db.exists():
        return stats

    # 渠道
    if "08 网络渠道" in base_wb.sheetnames:
        _, db_ch = read_table(db, "渠道与联系", ("品牌/机构", "渠道类型"))
        ws = base_wb["08 网络渠道"]
        # 找表头
        hrow = 3
        existing = set()
        for r in range(1, ws.max_row + 1):
            existing.add(norm_name(ws.cell(row=r, column=1).value) + "|" + norm_name(ws.cell(row=r, column=3).value))
        # 找末行
        end = ws.max_row
        for r in db_ch:
            key = norm_name(r.get("品牌/机构")) + "|" + norm_name(r.get("账号/关键词"))
            if key in existing:
                continue
            end += 1
            vals = [
                r.get("品牌/机构"),
                r.get("渠道类型"),
                r.get("账号/关键词"),
                r.get("公开链接/入口"),
                r.get("可切入事项"),
                r.get("核验状态"),
                "数据库融合补录",
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=end, column=c, value=v)
                style_data_cell(cell, fill_color=C_ORANGE)
            stats["channels"] += 1
            existing.add(key)

    # 活动
    if "05 活动与特别嘉宾" in base_wb.sheetnames:
        _, db_ev = read_table(db, "活动与嘉宾", ("日期", "活动"))
        ws = base_wb["05 活动与特别嘉宾"]
        hrow = 3
        existing = {norm_name(ws.cell(row=r, column=3).value) for r in range(hrow + 1, ws.max_row + 1)}
        # 探测活动列
        headers = [ws.cell(row=hrow, column=c).value for c in range(1, ws.max_column + 1)]
        # 简单：若活动名不在已有文本中则追加
        all_text = "\n".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, min(6, ws.max_column) + 1)
        )
        end = ws.max_row
        for r in db_ev:
            name = str(r.get("活动") or "").strip()
            if not name or name in all_text:
                continue
            end += 1
            # 按常见列写入
            vals = [
                end - hrow,
                r.get("类型") or "活动",
                name,
                r.get("特别嘉宾/参与方"),
                f"{r.get('日期') or ''} {r.get('地点/展馆') or ''}".strip(),
                r.get("招商价值/合作方式"),
                r.get("来源URL") or "数据库融合补录",
            ]
            for c, v in enumerate(vals, 1):
                if c <= ws.max_column:
                    cell = ws.cell(row=end, column=c, value=v)
                    style_data_cell(cell, fill_color=C_ORANGE)
            stats["events"] += 1
            all_text += "\n" + name
    return stats


def main():
    base_src = SRC / "全量资料补充版.xlsx"
    if not base_src.exists():
        raise SystemExit(f"缺少底表: {base_src}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base_src, OUT)
    wb = load_workbook(OUT)

    print("1/7 字段级融合 03…")
    stats = merge_sheet03_fields(wb)
    print("  ", stats)

    print("2/7 补录 01 联系人…")
    stats["contacts_added"] = enrich_contacts_sheet(wb)
    print("  contacts_added", stats["contacts_added"])

    print("3/7 渠道/活动补录…")
    extra = merge_db_channels_and_events(wb)
    stats.update(extra)

    print("4/7 可触达优先清单…")
    stats["reach_count"] = add_reach_priority_sheet(wb)
    print("  reach", stats["reach_count"])

    print("5/7 跟进工作台 + 未匹配线索…")
    add_followup_sheet(wb)
    stats["unmatched"] = add_db_unmatched_sheet(wb)
    print("  unmatched", stats["unmatched"])

    print("6/7 融合说明与封面…")
    add_merge_notes_sheet(wb, stats)
    update_cover(wb, stats)

    print("7/7 样式抛光…")
    polish_key_sheets(wb)

    # 工作表顺序微调：00、封面、可触达、01…
    desired = [
        "00 融合说明",
        "封面·使用说明",
        "可触达优先清单",
        "01 招商联系方式",
        "02 参展分类总览",
        "03 参展商总名录·精编422",
        "04 参展商全量·官网966",
        "05 活动与特别嘉宾",
        "06 明星·社团·嘉宾",
        "07 同期会议与合作伙伴",
        "08 网络渠道",
        "09 招商价值分析",
        "10 高价值目标·产业集群",
        "11 招商策略与跟进",
        "12 招商信息汇总",
        "15 企业联系方式与资料",
        "16 招商跟进工作台",
        "17 数据库未匹配线索",
        "14 新增材料合并",
        "13 数据来源与声明",
    ]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(OUT)
    print("已生成:", OUT)
    print("统计:", stats)
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
