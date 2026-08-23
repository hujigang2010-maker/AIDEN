# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from content import AGENT_JOIN, EVENT, FORWARD_LONG, FORWARD_REPLAY, FORWARD_SHORT, GUESTS

OUT = Path(__file__).resolve().parents[1] / "exports" / "飞书DemoDay4_观摩记录表.xlsx"

NAVY = "143A7A"
BLUE = "3370FF"
GOLD = "C9A227"
WHITE = "FFFFFF"
LIGHT = "EEF3FB"
YELLOW = "FFF6D8"
GREY = "F4F6FA"
THIN = Border(
    left=Side(style="thin", color="C5D0E0"),
    right=Side(style="thin", color="C5D0E0"),
    top=Side(style="thin", color="C5D0E0"),
    bottom=Side(style="thin", color="C5D0E0"),
)


def paint_header(ws, row, cols, fill=NAVY):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="微软雅黑", bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def style_range(ws, r1, r2, c1, c2, fill=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name="微软雅黑", size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def sheet_info(wb):
    ws = wb.active
    ws.title = "活动信息"
    ws.row_dimensions[1].height = 28
    headers = ["字段", "内容", "来源", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    paint_header(ws, 1, 4)
    rows = [
        ("系列", EVENT["series"], "公开检索", "飞书官方 B 站持续更新回放"),
        ("场次", EVENT["title_share"], "分享日历 INIT_DATA", "标题带「2」，更像加场"),
        ("开始", "2026-07-02 11:00 GMT+8", "分享日历", "周四"),
        ("结束", "2026-07-02 12:30 GMT+8", "分享日历", "90 分钟"),
        ("组织者", EVENT["organizer"], "分享日历", ""),
        ("状态", EVENT["status_zh"], "分享日历", "先点原链接确认回放/加场"),
        ("分享链接", EVENT["share_url"], "用户原文", "上一场易约满"),
        ("预告嘉宾 1", GUESTS[0]["name"], "用户预告", "非官方嘉宾表"),
        ("预告嘉宾 2", GUESTS[1]["name"], "用户预告", "非官方嘉宾表"),
        ("预告新能力", AGENT_JOIN["name"], "用户预告", "按系列惯例大概率现场展示"),
        ("第二期回放", EVENT["replay_ep2"]["url"], "B 站飞书账号", EVENT["replay_ep2"]["bvid"]),
    ]
    for i, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
        fill = LIGHT if i % 2 == 0 else WHITE
        style_range(ws, i, i, 1, 4, fill)
        ws.cell(i, 1).font = Font(name="微软雅黑", bold=True, size=11, color=NAVY)
        ws.row_dimensions[i].height = 32
    widths(ws, {1: 16, 2: 62, 3: 22, 4: 28})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:D12"


def sheet_watch(wb):
    ws = wb.create_sheet("观摩记录")
    headers = ["序号", "Demo / 环节", "日常麻烦", "人做哪一步", "Agent 做哪一步", "权限/入口", "可抄步骤", "我能不能复现", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    paint_header(ws, 1, len(headers), BLUE)
    presets = [
        "开场 / 产品同学介绍",
        "张咋啦 Demo",
        "向阳乔木 Demo",
        "Agent 入会演示",
        "其他 Builder Demo 1",
        "其他 Builder Demo 2",
        "Q&A / 彩蛋",
    ]
    for i, name in enumerate(presets, 2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, name)
        ws.cell(i, 8, "待看")
        style_range(ws, i, i, 1, 9, YELLOW if "Agent" in name else (LIGHT if i % 2 == 0 else WHITE))
        ws.row_dimensions[i].height = 36
    for r in range(9, 16):
        ws.cell(r, 1, r - 1)
        ws.cell(r, 8, "待看")
        style_range(ws, r, r, 1, 9, WHITE)
        ws.row_dimensions[r].height = 28
    widths(ws, {1: 8, 2: 22, 3: 22, 4: 22, 5: 22, 6: 16, 7: 28, 8: 14, 9: 18})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I15"
    ws.row_dimensions[1].height = 24


def sheet_agent(wb):
    ws = wb.create_sheet("Agent入会清单")
    headers = ["检查项", "现场观察", "通过?", "证据（截图/时间点）", "会后复现结果"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    paint_header(ws, 1, 5, GOLD)
    items = AGENT_JOIN["watch_questions"] + [
        "入会是否出现在参会列表",
        "是否需要等候室 / 主持人放行",
        "会中消息是否可见",
        "纪要出口在会中还是会后",
        "测试会能否复现，正式会是否禁止",
    ]
    for i, q in enumerate(items, 2):
        ws.cell(i, 1, q)
        ws.cell(i, 3, "待观察")
        style_range(ws, i, i, 1, 5, LIGHT if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 28
    widths(ws, {1: 42, 2: 28, 3: 12, 4: 28, 5: 24})
    ws.freeze_panes = "A2"


def sheet_copy(wb):
    ws = wb.create_sheet("转发口径")
    headers = ["版本", "适用场景", "正文"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h)
    paint_header(ws, 1, 3)
    data = [
        ("短口径", "微信群 / 飞书群一句话", FORWARD_SHORT),
        ("完整口径", "朋友圈 / 长公告", FORWARD_LONG),
        ("回放口径", "日历显示已结束时", FORWARD_REPLAY),
    ]
    for i, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
        style_range(ws, i, i, 1, 3, LIGHT if i % 2 == 0 else WHITE)
        ws.row_dimensions[i].height = 96
        ws.cell(i, 1).font = Font(name="微软雅黑", bold=True, size=11, color=NAVY)
    widths(ws, {1: 14, 2: 22, 3: 88})


def build():
    wb = Workbook()
    sheet_info(wb)
    sheet_watch(wb)
    sheet_agent(wb)
    sheet_copy(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    return OUT


if __name__ == "__main__":
    build()
