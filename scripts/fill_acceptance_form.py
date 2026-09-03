"""填写《绿城·潮鸣外滩-营销验收单(含现场照片)-2》。

以空白「基础营销验收单」为版式，嵌入 2026-05-22 峰会现场绿城/潮鸣露出照片，
并补齐效果评估表、赞助权益执行回执、照片索引。不与开盘花束验收混用。
参观邀约、朋友圈九宫格等无单独成片的项目如实标注，不拔高。
"""
from __future__ import annotations

import json
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image as PILImage

TEMPLATE = Path("/home/ubuntu/.cursor/projects/workspace/uploads/____b1b1.xlsx")
OUT = Path("/workspace/deliverables/绿城·潮鸣外滩-营销验收单(含现场照片)-2.xlsx")
OUT_ALIAS = Path("/workspace/deliverables/绿城·潮鸣外滩-营销验收单(含现场照片).xlsx")
PHOTO_DIR = Path("/workspace/deliverables/验收照片")
TMP = Path("/tmp/acceptance-photos")
THUMB = Path("/tmp/acceptance-thumbs-v2")
ZIP_OUT = Path("/workspace/deliverables/下载包/绿城潮鸣外滩-营销验收单(含现场照片)-2.zip")

ALBUM = "https://live.pailixiang.com/album/a12523836160"
VIDEO = "https://pan.baidu.com/s/1S8cNdqnCR_anuVPD6vlhjg"
VIDEO_CODE = "8888"

FONT = "等线"
TITLE_FONT = "黑体"
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")
LEFT_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
GREEN = PatternFill("solid", fgColor="1F6B4A")
GREEN_LIGHT = PatternFill("solid", fgColor="E8F5E9")
GREEN_MID = PatternFill("solid", fgColor="C8E6C9")
AMBER = PatternFill("solid", fgColor="FFF8E1")
WHITE = PatternFill("solid", fgColor="FFFFFF")
HEADER_BAR = PatternFill("solid", fgColor="F5F5F5")
WHITE_FONT = Font(name=TITLE_FONT, size=16, bold=True, color="FFFFFF")
META_JSON = Path("/tmp/acceptance-photos/all_meta.json")


def shoot_time(filename: str) -> str:
    """从拍立享元数据取拍摄时间（941Axxxx.JPG）。"""
    if not getattr(shoot_time, "_map", None):
        mapping = {}
        if META_JSON.exists():
            for item in json.loads(META_JSON.read_text()):
                mapping[item.get("name", "")] = item.get("shoot", "")
        shoot_time._map = mapping
    orig = filename.split("_", 1)[-1] if filename[:1].isdigit() else filename
    if orig.startswith("941A"):
        return shoot_time._map.get(orig, "")
    return shoot_time._map.get(filename, "")

# 首页 9 张：优先「绿城中国 / 潮鸣 / 外滩潮鸣」字样可辨的画面
COVER = [
    ("060_941A7785.JPG", "①议程看板【绿城·潮鸣】晚宴冠名", "议程看板：晚宴场次【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴"),
    ("065_941A7794.JPG", "②1号位展台：绿城礼盒+潮鸣立牌", "1号位展台：绿城礼盒、「潮鸣」立牌、VR 与项目物料"),
    ("001_941A7663.JPG", "③主背景板：晚宴冠名·绿城中国", "主背景板全幅：晚宴冠名战略合作伙伴·绿城中国"),
    ("074_941A7832.JPG", "④双屏：战略合作「绿城·外滩潮鸣」", "主会场双屏底部：战略合作伙伴「绿城·外滩潮鸣」"),
    ("071_941A7821.JPG", "⑤主持：晚宴冠名及战略合作伙伴", "主持环节大屏：「晚宴冠名及战略合作伙伴：绿城中国」"),
    ("055_941A7777.JPG", "⑥主背景板合影核验冠名位", "主背景板合影：晚宴冠名位「绿城中国」可核验"),
    ("388_941A8841.JPG", "⑦晚宴桌卡「潮鳴」", "晚宴桌卡露出「潮鳴」，江景厅圆桌"),
    ("395_941A8859.JPG", "⑧祝酒大屏「绿城·外滩潮鸣」", "晚宴祝酒：弧形大屏露出「绿城·外滩潮鸣」"),
    ("00_banner.jpg", "⑨主视觉KV：战略合作含绿城中国", "主视觉 KV：战略合作伙伴名单含绿城中国"),
]

# 露出页续图（含场地与晚宴氛围；不把无字样的画面写成潮鸣看板）
EXTRA = [
    ("075_941A7835.JPG", "讲台/大屏顶栏合作 logo 带（主办致辞）"),
    ("073_941A7829.JPG", "座席手拎袋物料（嘉宾席）"),
    ("077_941A7839.JPG", "主办致辞画面：大屏顶栏合作 logo"),
    ("056_941A7779.JPG", "主背景板合影（晚宴冠名位特写）"),
    ("p066_941A7798.JPG", "主背景板合影核验晚宴冠名位"),
    ("v04_941A8762.JPG", "晚宴现场举杯"),
    ("v01_941A8878.JPG", "晚宴桌次互动"),
    ("399_941A8881.JPG", "晚宴桌次物料"),
    ("p144_941A8048.JPG", "主会场全景（北外滩·一滴水）"),
    ("p080_941A7864.JPG", "主论坛致辞（讲台合作伙伴 logo 带）"),
    ("394_941A8857.JPG", "晚宴合影（江景厅）"),
    ("068_941A7802.JPG", "场地江景露台（北外滩一滴水，正对陆家嘴）"),
    ("058_941A7782.JPG", "主背景板合影：晚宴冠名战略合作伙伴·绿城中国"),
]

ALIASES = {
    "395_941A8859.JPG": ["v02_941A8859.JPG", "395_941A8859.JPG"],
    "001_941A7663.JPG": ["p000_941A7663.JPG", "001_941A7663.JPG"],
    "065_941A7794.JPG": ["p064_941A7794.JPG", "065_941A7794.JPG"],
}


def resolve_src(name: str) -> Path:
    names = ALIASES.get(name, [name])
    dirs = [PHOTO_DIR, TMP / "greentown", TMP]
    for n in names:
        for d in dirs:
            p = d / n
            if p.exists() and p.stat().st_size > 2000:
                return p
    raise FileNotFoundError(name)


def thumb(src: Path, dest: Path, max_w=520, max_h=390, quality=84) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    im = PILImage.open(src)
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    im.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
    dest = dest.with_suffix(".jpg")
    im.save(dest, "JPEG", quality=quality, optimize=True)
    return dest


def place_image(ws, path: Path, cell: str, width: int, height: int) -> None:
    img = XLImage(str(path))
    img.width = width
    img.height = height
    img.anchor = cell
    ws.add_image(img)


def fill_cell(ws, coord, text, size=10, bold=False, align=WRAP, fill=None, font_name=FONT):
    cell = ws[coord]
    cell.value = text
    cell.font = Font(name=font_name, size=size, bold=bold)
    cell.alignment = align
    cell.border = THIN
    if fill is not None:
        cell.fill = fill


def paint_range(ws, start_row, end_row, start_col, end_col, fill=WHITE):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            if fill is not None:
                cell.fill = fill


def unmerge_if(ws, coord: str) -> None:
    hits = [str(rng) for rng in list(ws.merged_cells.ranges) if coord in rng]
    for rng in hits:
        ws.unmerge_cells(rng)


def setup_print(ws, area: str, landscape: bool = False, fit_height: int = 1) -> None:
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = area
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.2, footer=0.2
    )


def fill_cover(ws) -> None:
    ws.title = "营销验收单"
    fill_cell(ws, "B3", "晚宴冠名战略合作伙伴\n（绿城中国 | 绿城·潮鸣外滩）", 11, True)
    fill_cell(ws, "D3", "人民币 100,000 元整\n（含税 · 四模块整体打包）", 11, True)
    fill_cell(
        ws,
        "B4",
        "主办：上海市杨浦区科技企业联合会\n赞助：上海绿城泓盛建设发展有限公司\n项目品牌：绿城中国 · 绿城·潮鸣外滩",
        10,
    )
    fill_cell(ws, "D4", "2026年5月22日\n13:00 – 20:30\n上海·北外滩·一滴水", 11, True)
    fill_cell(
        ws,
        "B5",
        "【活动】第四届 / 2026人工智能商业化落地与硬核投资破局峰会。\n"
        "【身份】晚宴冠名战略合作伙伴（物料亦作「绿城·外滩潮鸣」「潮鳴」）。\n"
        "【晚宴】【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴。\n"
        "【四模块】①晚宴冠名和专属权益 ¥55,000；②主会场权益 ¥30,000；"
        "③专场项目参观邀约 ¥5,000；④宣发配合 ¥10,000。含税合计 ¥100,000，整体打包验收。\n"
        "【现场已核验】主背景板「晚宴冠名战略合作伙伴·绿城中国」、议程看板晚宴冠名、"
        "1号位展台、主会场双屏「绿城·外滩潮鸣」、晚宴大屏/桌卡「潮鳴」。\n"
        "【证书】事项载明颁发“2026年度智慧人居新质资产领军企业 暨卓越战略合作伙伴”；"
        "本相册未见证书特写，不以照片推定已颁。",
        8,
        align=LEFT_TOP,
    )
    ws.row_dimensions[5].height = 118

    fill_cell(
        ws,
        "B13",
        "【露出】主背景板、议程看板、1号位展台、双屏、晚宴大屏/桌卡已核验，"
        "绿城中国 / 潮鸣 / 外滩潮鸣字样可辨（见本页图①–⑨及「绿城·潮鸣外滩露出」页）。\n"
        "【营销评估】非常好（☐）　良好（☑）　一般（☐）　较差（☐）\n"
        "【打包验收】四模块整体打包；参观邀约无单独成片、朋友圈九宫格/回顾视频未附本表，"
        "见「权益核验清单」黄底项。效果良好，符合约定要求（以已核验露出为准）。",
        9,
        align=LEFT_TOP,
    )
    ws.row_dimensions[13].height = 48
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 28

    fill_cell(
        ws,
        "B16",
        "①本页 9 张均为绿城/潮鸣可辨露出，含图注；"
        "②「权益核验清单」按报价单 4 模块对照照片；"
        "③「绿城·潮鸣外滩露出」页为现场大图；"
        "④「效果评估表」按绿城中国策划活动类效果评估表结构；"
        "⑤「赞助权益执行回执」供绿城复核盖章；"
        "⑥「照片索引」含拍摄时间与模块对照；"
        f"⑦拍立享直播 {ALBUM}（相册约 425 张）；"
        f"⑧全程录像 {VIDEO} 提取码 {VIDEO_CODE}；"
        "⑨费用清单（经办人签字）。本表不与 2026年3月开盘花束验收混用。",
        8,
        align=LEFT,
    )
    ws.row_dimensions[16].height = 84
    fill_cell(ws, "C17", "胡继刚\n13262607888\n（主办执行）\n签字：____________", 10, align=LEFT)
    fill_cell(
        ws,
        "C18",
        "绿城对接：笪浩 18678408669\n营销负责人签字：____________\n日期：____________",
        10,
        align=LEFT,
    )

    unmerge_if(ws, "B6")
    paint_range(ws, 6, 12, 2, 4, WHITE)
    # 3×3：6/8/10 放图，7/9/11 图注，12 总注
    ws.merge_cells("B12:D12")
    fill_cell(
        ws,
        "B12",
        "图①–⑨对应「权益核验清单」；更多露出见「绿城·潮鸣外滩露出」页。江景露台等无品牌字样画面不作为潮鸣看板证据。",
        8,
        align=LEFT,
    )

    photo_rows = (6, 8, 10)
    cap_rows = (7, 9, 11)
    cols = ("B", "C", "D")
    for i, (name, short, _full) in enumerate(COVER):
        r, c = divmod(i, 3)
        fill_cell(ws, f"{cols[c]}{cap_rows[r]}", short, 7, align=LEFT)
        t = thumb(resolve_src(name), THUMB / f"cover_{name}", max_w=420, max_h=300)
        place_image(ws, t, f"{cols[c]}{photo_rows[r]}", 168, 108)

    for r, h in [(6, 86), (7, 32), (8, 86), (9, 32), (10, 86), (11, 32), (12, 22)]:
        ws.row_dimensions[r].height = h

    ws["A20"] = "照片直播（拍立享）"
    ws["A20"].font = Font(name=FONT, size=9, bold=True)
    ws["B20"] = ALBUM
    ws["B20"].hyperlink = ALBUM
    ws["B20"].font = Font(name=FONT, size=9, color="0563C1", underline="single")
    ws.merge_cells("B20:D20")
    ws["A21"] = "全程录像（百度网盘）"
    ws["A21"].font = Font(name=FONT, size=9, bold=True)
    ws["B21"] = f"{VIDEO}  提取码：{VIDEO_CODE}"
    ws["B21"].hyperlink = VIDEO
    ws["B21"].font = Font(name=FONT, size=9, color="0563C1", underline="single")
    ws.merge_cells("B21:D21")

    setup_print(ws, "A2:D18", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_properties.tabColor = "1F6B4A"


def add_rights_sheet(wb) -> None:
    ws = wb.create_sheet("权益核验清单", 1)
    widths = [6, 18, 36, 12, 42, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    fill_cell(ws, "A1", "绿城·潮鸣外滩 · 晚宴冠名四模块权益核验清单", 16, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 6, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    fill_cell(
        ws,
        "A2",
        "活动：2026-05-22 上海·北外滩·一滴水　身份：晚宴冠名战略合作伙伴　"
        "金额：含税 ¥100,000（报价单 4 模块整体打包，不拆子项验收）　"
        "依据：绿城中国-潮鸣外滩-10万元晚宴冠名服务报价单",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 6, GREEN_MID)
    ws.row_dimensions[2].height = 32

    headers = ["序号", "服务模块", "约定内容", "金额（元）", "现场照片核验", "结论"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[3].height = 22

    rows = [
        (
            "1",
            "晚宴冠名和专属权益",
            "晚宴主题冠名、专场PPT/视频、主持人口播、席卡等物料植入",
            "55,000",
            "图①议程看板晚宴冠名【绿城·潮鸣】；图③⑤⑥主背景板「晚宴冠名战略合作伙伴·绿城中国」；"
            "图⑦桌卡「潮鳴」；图⑧祝酒大屏「绿城·外滩潮鸣」。专场PPT宣讲无单独成片，以录像为准。",
            "通过（物料/冠名）",
        ),
        (
            "2",
            "主会场权益",
            "1号位展台、手拎袋项目物料、视频轮播/奖项颁发画面",
            "30,000",
            "图②1号位展台（绿城礼盒+潮鸣立牌+VR）；图④双屏战略合作伙伴「绿城·外滩潮鸣」；"
            "图⑨主视觉KV含绿城中国；露出页手拎袋/讲台顶栏。奖项颁发证书未见特写。",
            "通过（展台/双屏）",
        ),
        (
            "3",
            "专场项目参观邀约",
            "论坛/颁奖结束后主持人口播 + 动线引导，邀约前往案场",
            "5,000",
            "图⑤证明主会场主持人口播位已具备。相册抽样未见「参观邀约/案场接驳」专项成片，"
            "不以本表照片推定口播已执行，请以全程录像及现场执行回执为准。",
            "待录像核验",
        ),
        (
            "4",
            "宣发配合",
            "现场摄影、朋友圈九宫格、回顾视频项目鸣谢、执行回执",
            "10,000",
            "现场摄影：本表露出照片 + 拍立享约 425 张 + 网盘录像。本工作簿「赞助权益执行回执」即执行报告。"
            "朋友圈九宫格、回顾视频片头鸣谢未附，不以照片推定已发。",
            "摄影通过；其余待补",
        ),
        (
            "—",
            "合计（含税）",
            "4 模块整体打包，一次性验收",
            "100,000",
            "已核验项以绿城/潮鸣字样可辨露出为准；黄底项需录像或另行回执，不在本表拔高为已完成。",
            "部分待补",
        ),
        (
            "—",
            "证书",
            "2026年度智慧人居新质资产领军企业 暨卓越战略合作伙伴",
            "—",
            "事项描述已载明颁发。本相册未见证书/颁奖特写，不作为已交付证据。",
            "待补照片",
        ),
    ]
    fills = {
        "通过（物料/冠名）": GREEN_LIGHT,
        "通过（展台/双屏）": GREEN_LIGHT,
        "待录像核验": AMBER,
        "摄影通过；其余待补": AMBER,
        "部分待补": AMBER,
        "待补照片": AMBER,
    }
    for i, vals in enumerate(rows, 4):
        result = vals[-1]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(name=FONT, size=9, bold=(col in (2, 4, 6)))
            cell.alignment = LEFT if col in (3, 5) else WRAP
            cell.border = THIN
            cell.fill = fills.get(result, WHITE) if col == 6 else WHITE
        ws.row_dimensions[i].height = 72

    note = 10
    ws.merge_cells(f"A{note}:F{note}")
    fill_cell(
        ws,
        f"A{note}",
        "验收口径：报价单约定「不逐项贴照片、四模块整体打包」。本清单把照片映射到模块，便于绿城复核，"
        "不等于把未成片的子项改成已交付。营销评估预勾「良好」，请绿城营销负责人复核签字。"
        "本清单仅对应 2026-05-22 峰会，不与 3 月开盘花束验收混用。",
        9,
        False,
        LEFT,
        GREEN_LIGHT,
    )
    paint_range(ws, note, note, 1, 6, GREEN_LIGHT)
    ws.row_dimensions[note].height = 48

    setup_print(ws, f"A1:F{note}", landscape=True, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FFF8E1"


def add_photo_sheet(wb) -> None:
    ws = wb.create_sheet("绿城·潮鸣外滩露出")
    ws.merge_cells("A1:D1")
    fill_cell(
        ws,
        "A1",
        "绿城中国｜绿城·潮鸣外滩（外滩潮鸣 / 潮鳴）现场露出验收照片",
        14,
        True,
        WRAP,
        GREEN,
        TITLE_FONT,
    )
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        f"活动时间：2026-05-22　地点：上海·北外滩·一滴水　"
        f"本页 {len(COVER) + len(EXTRA)} 张（首页 9 张 + 续图）　"
        f"完整相册约 425 张：{ALBUM}　录像：{VIDEO}（提取码 {VIDEO_CODE}）",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws["A2"].hyperlink = ALBUM
    ws.row_dimensions[2].height = 36
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 42

    all_items = [(n, full) for n, _s, full in COVER] + EXTRA
    row = 4
    col_pair = [("A", "B"), ("C", "D")]
    i = 0
    while i < len(all_items):
        ws.row_dimensions[row].height = 168
        ws.row_dimensions[row + 1].height = 36
        for j, (img_col, cap_col) in enumerate(col_pair):
            if i + j >= len(all_items):
                break
            name, cap = all_items[i + j]
            t = thumb(resolve_src(name), THUMB / f"sheet_{name}", max_w=640, max_h=480)
            place_image(ws, t, f"{img_col}{row}", 290, 205)
            cell = ws[f"{img_col}{row + 1}"]
            ws.merge_cells(f"{img_col}{row + 1}:{cap_col}{row + 1}")
            cell.value = f"{i + j + 1}. {cap}"
            cell.font = Font(name=FONT, size=9)
            cell.alignment = LEFT
            cell.border = THIN
        i += 2
        row += 2

    note_row = row + 1
    ws.merge_cells(f"A{note_row}:D{note_row}")
    fill_cell(
        ws,
        f"A{note_row}",
        "说明：优先收录绿城中国、绿城·潮鸣外滩、绿城·外滩潮鸣、潮鳴可辨露出。"
        "江景露台等无品牌字样画面仅证明场地，不作为潮鸣看板证据。"
        "完整 425 张以拍立享为准；MP4 以网盘「5.22」为准。不与开盘花束 4 张混用。",
        9,
        False,
        LEFT,
        WHITE,
    )
    paint_range(ws, note_row, note_row, 1, 4, WHITE)
    ws.row_dimensions[note_row].height = 42
    setup_print(ws, f"A1:D{note_row}", landscape=True, fit_height=0)
    ws.print_title_rows = "1:2"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F6B4A"


def add_eval_sheet(wb) -> None:
    """绿城中国策划活动类效果评估表（峰会晚宴冠名）。"""
    ws = wb.create_sheet("效果评估表", 2)
    for col, width in enumerate([16, 24, 24, 36], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.merge_cells("A1:D1")
    fill_cell(ws, "A1", "绿城中国策划活动类效果评估表", 16, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        "（冠名、活动赞助、活动执行等）　项目：绿城·潮鸣外滩　活动：2026-05-22 晚宴冠名",
        10,
        False,
        WRAP,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws.row_dimensions[2].height = 22

    fill_cell(ws, "A3", "活动主题", 11, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B3:D3")
    fill_cell(ws, "B3", "【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴", 13, True)
    paint_range(ws, 3, 3, 1, 4, WHITE)
    ws["A3"].fill = GREEN_LIGHT
    ws.row_dimensions[3].height = 28

    fill_cell(ws, "A4", "活动举办时间", 11, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B4:D4")
    fill_cell(ws, "B4", "2026年5月22日 13:00–20:30　　上海·北外滩·一滴水", 12, True)
    paint_range(ws, 4, 4, 1, 4, WHITE)
    ws["A4"].fill = GREEN_LIGHT
    ws.row_dimensions[4].height = 28

    ws.merge_cells("A5:A12")
    fill_cell(ws, "A5", "活动\n现场氛围", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 5, 12, 1, 4, WHITE)
    ws["A5"].fill = GREEN_LIGHT
    eval_photos = [COVER[0], COVER[1], COVER[2], COVER[7]]  # 看板、展台、主背景、祝酒大屏
    caps = [eval_photos[0][1], eval_photos[1][1], eval_photos[2][1], eval_photos[3][1]]
    ws.merge_cells("B5:C5")
    fill_cell(ws, "B5", caps[0], 8, align=LEFT)
    fill_cell(ws, "D5", caps[1], 8, align=LEFT)
    ws.merge_cells("B6:C8")
    ws.merge_cells("D6:D8")
    ws.merge_cells("B9:C9")
    fill_cell(ws, "B9", caps[2], 8, align=LEFT)
    fill_cell(ws, "D9", caps[3], 8, align=LEFT)
    ws.merge_cells("B10:C12")
    ws.merge_cells("D10:D12")
    for r, h in [(5, 28), (6, 78), (7, 78), (8, 78), (9, 28), (10, 78), (11, 78), (12, 78)]:
        ws.row_dimensions[r].height = h
    anchors = [("B6", 340, 280), ("D6", 250, 280), ("B10", 340, 280), ("D10", 250, 280)]
    for (name, _s, _f), (cell, w, h) in zip(eval_photos, anchors):
        t = thumb(resolve_src(name), THUMB / f"eval_{name}", max_w=800, max_h=600)
        place_image(ws, t, cell, w, h)

    ws.merge_cells("A13:A16")
    fill_cell(ws, "A13", "策划负责人填写", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 13, 16, 1, 4, WHITE)
    ws["A13"].fill = GREEN_LIGHT
    ws.merge_cells("B13:D13")
    fill_cell(ws, "B13", "总体活动评价：", 11, True, LEFT, GREEN_MID)
    ws.merge_cells("B14:D15")
    fill_cell(
        ws,
        "B14",
        "本次活动为「2026人工智能商业化落地与硬核投资破局峰会」晚宴冠名。"
        "绿城中国 | 绿城·潮鸣外滩作为晚宴冠名战略合作伙伴，主背景板、议程看板、"
        "1号位展台、主会场双屏及晚宴大屏/桌卡均完成品牌露出。"
        "按照约定完成整体策划与执行，现场氛围良好，活动取得圆满成功。",
        11,
        False,
        LEFT_TOP,
        WHITE,
    )
    ws.merge_cells("B16:D16")
    fill_cell(
        ws,
        "B16",
        "策划负责人签名：胡继刚 ____________________　　日期：2026-05-22"
        "　　（请补签）",
        10,
        False,
        LEFT,
        HEADER_BAR,
    )
    ws.row_dimensions[13].height = 24
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 36
    ws.row_dimensions[16].height = 32

    ws.merge_cells("A17:A21")
    fill_cell(ws, "A17", "营销负责人填写", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 17, 21, 1, 4, WHITE)
    ws["A17"].fill = GREEN_LIGHT
    ws.merge_cells("B17:D17")
    fill_cell(ws, "B17", "对活动【总体效果】评估（文字说明）：", 11, True, LEFT, GREEN_MID)
    ws.merge_cells("B18:D18")
    fill_cell(
        ws,
        "B18",
        "绿城中国 / 潮鸣 / 外滩潮鸣字样可辨露出已核验，效果良好，符合约定要求。"
        "参观邀约口播、朋友圈九宫格、回顾视频、证书特写见执行回执待补项，不在此拔高。",
        11,
        False,
        LEFT,
        WHITE,
    )
    ws.merge_cells("B19:D19")
    fill_cell(
        ws,
        "B19",
        "非常好（☐）　　良好（☑）　　一般（☐）　　较差（☐）",
        13,
        True,
        WRAP,
        AMBER,
    )
    ws.merge_cells("B20:D20")
    fill_cell(
        ws,
        "B20",
        "勾选说明：按《基础营销验收单》模板默认口径「效果良好，符合要求」预勾「良好」，"
        "不拔高为「非常好」。请绿城营销负责人复核。",
        9,
        False,
        LEFT,
        AMBER,
    )
    ws.merge_cells("B21:D21")
    fill_cell(
        ws,
        "B21",
        "营销负责人签名：____________________　　日期：____________________"
        "　　对接：笪浩 18678408669（评估表请补签）",
        10,
        False,
        LEFT,
        HEADER_BAR,
    )
    ws.row_dimensions[17].height = 24
    ws.row_dimensions[18].height = 42
    ws.row_dimensions[19].height = 28
    ws.row_dimensions[20].height = 42
    ws.row_dimensions[21].height = 32

    ws.merge_cells("A22:D22")
    fill_cell(
        ws,
        "A22",
        "详见附件：本工作簿「营销验收单」「权益核验清单」「赞助权益执行回执」及拍立享相册、网盘录像。"
        "合作金额含税 ¥100,000，四模块整体打包。本表仅对应 5 月 22 日峰会。",
        9,
        False,
        LEFT,
        WHITE,
    )
    paint_range(ws, 22, 22, 1, 4, WHITE)
    ws.row_dimensions[22].height = 36
    setup_print(ws, "A1:D22", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C8E6C9"


def add_receipt_sheet(wb) -> None:
    """主办向绿城出具的《赞助权益执行回执》，供签字盖章。"""
    ws = wb.create_sheet("赞助权益执行回执", 3)
    for col, w in enumerate([22, 28, 22, 28], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("A1:D1")
    fill_cell(ws, "A1", "赞助权益执行回执", 18, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        "依据《绿城中国-潮鸣外滩-10万元晚宴冠名服务报价单》：大会结束后 7 个自然日内，"
        "由主办向绿城一次性出具本回执，作为四模块整体验收依据。",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws.row_dimensions[2].height = 32

    pairs = [
        ("出具方（主办）", "上海市杨浦区科技企业联合会", "接收方（赞助）", "上海绿城泓盛建设发展有限公司"),
        ("对接人", "胡继刚　13262607888", "对接人", "笪浩　18678408669"),
        ("活动名称", "2026人工智能商业化落地与硬核投资破局峰会", "活动时间/地点", "2026-05-22　上海·北外滩·一滴水"),
        ("合作身份", "晚宴冠名战略合作伙伴", "合作金额", "人民币 100,000 元整（含税）"),
        ("晚宴场次", "【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴", "项目品牌", "绿城中国 · 绿城·潮鸣外滩"),
    ]
    r = 3
    for a, b, c, d in pairs:
        fill_cell(ws, f"A{r}", a, 10, True, WRAP, GREEN_LIGHT)
        fill_cell(ws, f"B{r}", b, 10, False, LEFT, WHITE)
        fill_cell(ws, f"C{r}", c, 10, True, WRAP, GREEN_LIGHT)
        fill_cell(ws, f"D{r}", d, 10, False, LEFT, WHITE)
        ws.row_dimensions[r].height = 28
        r += 1

    ws.merge_cells("A8:D8")
    fill_cell(ws, "A8", "四模块执行确认（整体打包，不拆子项计价）", 11, True, LEFT, GREEN)
    paint_range(ws, 8, 8, 1, 4, GREEN)
    ws["A8"].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws["A8"].fill = GREEN
    ws.row_dimensions[8].height = 24

    headers = ["模块", "金额（元）", "执行结论", "代表性证据"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(9, i, h)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[9].height = 22

    mods = [
        ("1. 晚宴冠名和专属权益", "55,000", "☑ 已执行（冠名/物料）", "议程看板【绿城·潮鸣】晚宴；主背景板「晚宴冠名战略合作伙伴·绿城中国」；晚宴桌卡「潮鳴」；祝酒大屏「绿城·外滩潮鸣」"),
        ("2. 主会场权益", "30,000", "☑ 已执行（展台/双屏）", "1号位展台绿城礼盒+潮鸣立牌+VR；主会场双屏战略合作伙伴「绿城·外滩潮鸣」；主视觉KV含绿城中国"),
        ("3. 专场项目参观邀约", "5,000", "☐ 待录像核验", "主持人口播位已具备。相册未见参观/接驳专项成片，请以全程录像确认是否口播执行。"),
        ("4. 宣发配合", "10,000", "☑ 摄影已交　☐ 其余待补", "现场摄影+本回执（图片/合影/执行报告）。朋友圈九宫格、回顾视频片头鸣谢未附本回执。"),
        ("合计（含税）", "100,000", "部分待补，主体已交", "四模块整体打包验收。已核验项以绿城/潮鸣字样可辨露出为准。"),
    ]
    fills = {
        "☑ 已执行（冠名/物料）": GREEN_LIGHT,
        "☑ 已执行（展台/双屏）": GREEN_LIGHT,
        "☐ 待录像核验": AMBER,
        "☑ 摄影已交　☐ 其余待补": AMBER,
        "部分待补，主体已交": AMBER,
    }
    for i, row in enumerate(mods, 10):
        for col, val in enumerate(row, 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(name=FONT, size=9, bold=(col in (1, 2, 3)))
            cell.alignment = LEFT if col == 4 else WRAP
            cell.border = THIN
            cell.fill = fills.get(row[2], WHITE) if col == 3 else WHITE
        ws.row_dimensions[i].height = 48

    ws.merge_cells("A15:D15")
    fill_cell(
        ws,
        "A15",
        "证据目录：①本工作簿营销验收单（首页 9 张图注）②权益核验清单 ③效果评估表 ④绿城·潮鸣外滩露出大图 "
        f"⑤拍立享 {ALBUM}（约 425 张）⑥网盘录像 {VIDEO} 提取码 {VIDEO_CODE}。"
        "证书特写未见，不以本回执推定证书已颁。",
        9,
        False,
        LEFT,
        WHITE,
    )
    paint_range(ws, 15, 15, 1, 4, WHITE)
    ws.row_dimensions[15].height = 52

    ws.merge_cells("A16:D16")
    fill_cell(
        ws,
        "A16",
        "总体结论：晚宴冠名与主会场露出已按约定落地，效果良好，符合要求（预勾「良好」）。"
        "参观邀约、朋友圈九宫格、回顾视频、证书特写为待补项，不影响四模块整体打包之主体验收，"
        "由绿城营销负责人复核后签字盖章。本回执不与 2026 年 3 月开盘花束验收混用。",
        10,
        False,
        LEFT_TOP,
        GREEN_LIGHT,
    )
    paint_range(ws, 16, 16, 1, 4, GREEN_LIGHT)
    ws.row_dimensions[16].height = 56

    fill_cell(ws, "A17", "主办方（出具）", 10, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B17:B18")
    fill_cell(
        ws,
        "B17",
        "单位：上海市杨浦区科技企业联合会\n授权代表：胡继刚\n签字：____________\n日期：2026-05-22\n（公章）",
        10,
        False,
        LEFT_TOP,
        WHITE,
    )
    fill_cell(ws, "C17", "赞助方（接收）", 10, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("D17:D18")
    fill_cell(
        ws,
        "D17",
        "单位：上海绿城泓盛建设发展有限公司\n授权代表：____________________\n对接：笪浩 18678408669\n签字：____________　日期：____________\n（公章）",
        10,
        False,
        LEFT_TOP,
        WHITE,
    )
    paint_range(ws, 17, 18, 1, 4, WHITE)
    ws["A17"].fill = GREEN_LIGHT
    ws["C17"].fill = GREEN_LIGHT
    ws.merge_cells("A17:A18")
    ws.merge_cells("C17:C18")
    ws.row_dimensions[17].height = 36
    ws.row_dimensions[18].height = 72

    setup_print(ws, "A1:D18", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F6B4A"


def add_index_sheet(wb) -> None:
    ws = wb.create_sheet("照片索引")
    widths = [8, 22, 22, 42, 18, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells("A1:F1")
    fill_cell(ws, "A1", "绿城·潮鸣外滩露出 · 照片索引（含拍摄时间）", 16, True, WRAP, GREEN, TITLE_FONT)
    paint_range(ws, 1, 1, 1, 6, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:F2")
    fill_cell(
        ws,
        "A2",
        "拍摄时间取自拍立享相册元数据。首页 9 张为绿城/潮鸣字样可辨主证据；续图含场地与晚宴氛围。"
        "不收录他方展位（如泰隆银行）或无品牌字样的论坛特写。",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 6, GREEN_MID)
    ws.row_dimensions[2].height = 28
    headers = ["图号", "文件名", "拍摄时间", "露出点", "对应模块", "首页"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN

    cover_mod = [
        "晚宴冠名",
        "主会场",
        "晚宴冠名",
        "主会场",
        "晚宴冠名",
        "晚宴冠名",
        "晚宴冠名",
        "晚宴冠名",
        "主会场/宣发",
    ]
    extra_mod = {
        "075_941A7835.JPG": "主会场",
        "073_941A7829.JPG": "主会场",
        "077_941A7839.JPG": "主会场",
        "056_941A7779.JPG": "晚宴冠名",
        "p066_941A7798.JPG": "晚宴冠名",
        "v04_941A8762.JPG": "晚宴冠名",
        "v01_941A8878.JPG": "晚宴冠名",
        "399_941A8881.JPG": "晚宴冠名",
        "p144_941A8048.JPG": "主会场",
        "p080_941A7864.JPG": "主会场",
        "394_941A8857.JPG": "晚宴冠名",
        "068_941A7802.JPG": "场地（无字样）",
        "058_941A7782.JPG": "晚宴冠名",
    }
    rows = []
    for i, (name, _s, full) in enumerate(COVER, 1):
        rows.append((str(i), name, shoot_time(name), full, cover_mod[i - 1], "是"))
    for j, (name, cap) in enumerate(EXTRA, len(COVER) + 1):
        rows.append((str(j), name, shoot_time(name), cap, extra_mod.get(name, "—"), "否"))
    for i, vals in enumerate(rows, 4):
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = LEFT if col in (2, 4) else WRAP
            cell.border = THIN
            cell.fill = GREEN_LIGHT if vals[5] == "是" else WHITE
        ws.row_dimensions[i].height = 22
    last = 3 + len(rows)
    setup_print(ws, f"A1:F{last}", landscape=True, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C8E6C9"
    ws.auto_filter.ref = f"A3:F{last}"


COMBINED_ZIP = Path("/workspace/deliverables/下载包/绿城潮鸣外滩-营销验收完整包.zip")
BOUQUET_XLSX = Path("/workspace/deliverables/绿城·潮鸣外滩-开盘花束营销验收单.xlsx")


def make_zip() -> None:
    ZIP_OUT.parent.mkdir(parents=True, exist_ok=True)
    used = [n for n, *_ in COVER] + [n for n, _ in EXTRA]
    with zipfile.ZipFile(ZIP_OUT, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(OUT, OUT.name)
        for name in used:
            src = resolve_src(name)
            zf.write(src, f"验收照片-露出精选/{name}")
    with zipfile.ZipFile(COMBINED_ZIP, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(OUT, f"峰会晚宴冠名/{OUT.name}")
        if BOUQUET_XLSX.exists():
            zf.write(BOUQUET_XLSX, f"开盘花束/{BOUQUET_XLSX.name}")
        for name in used:
            src = resolve_src(name)
            zf.write(src, f"峰会晚宴冠名/验收照片-露出精选/{name}")


def main() -> None:
    if not TEMPLATE.exists():
        raise SystemExit(f"缺少模板 {TEMPLATE}")
    PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    THUMB.mkdir(parents=True, exist_ok=True)

    used = [n for n, *_ in COVER] + [n for n, _ in EXTRA]
    for name in used:
        src = resolve_src(name)
        dest = PHOTO_DIR / name
        if not dest.exists():
            dest.write_bytes(src.read_bytes())

    wb = load_workbook(str(TEMPLATE))
    fill_cover(wb.active)
    add_rights_sheet(wb)
    add_eval_sheet(wb)
    add_receipt_sheet(wb)
    add_photo_sheet(wb)
    add_index_sheet(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT))
    wb.save(str(OUT_ALIAS))
    make_zip()
    print(f"saved {OUT} size={OUT.stat().st_size}")
    print(f"alias {OUT_ALIAS} size={OUT_ALIAS.stat().st_size}")
    print(f"zip {ZIP_OUT} size={ZIP_OUT.stat().st_size}")
    print(f"combined {COMBINED_ZIP} size={COMBINED_ZIP.stat().st_size}")


if __name__ == "__main__":
    main()
