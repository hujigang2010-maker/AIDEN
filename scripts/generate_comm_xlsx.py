#!/usr/bin/env python3
"""生成肇事方沟通流程 Excel。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from comm_plan import (
    BANS,
    COMM_GOAL,
    CT_PLAIN,
    FLOW,
    LAWYER_NOW,
    MED_FOR_TALK,
    MED_SUMMARY,
    OPENING_SCRIPT,
    RIDER_REPLIES,
    SCHEMES,
    SHOW_LIST,
)
from content import CT_REPORTS, TALKING_POINTS
from generate_xlsx import (
    GREEN,
    LIGHT,
    NAVY,
    RED,
    THIN,
    WHITE,
    body_font,
    fill,
    freeze,
    style_header,
    style_rows,
    widths,
    write_title,
)

MUST_GET = [
    "微信群：骑手 + 站点 + 保险 + 家属",
    "保险对接人、保单/工号、是否须认定书才能垫付",
    "人民医院（抚顺路院区）认不认；齐鲁已发生费用认不认",
    "已发生费用垫付路径和发票清单",
]


def _page(ws, header_text: str):
    freeze(ws)
    ws.oddHeader.left.text = header_text
    ws.oddFooter.right.text = "第 &P 页 / 共 &N 页 · 内部使用"


def _row_h(ws, row: int, height: int = 48):
    ws.row_dimensions[row].height = height


def build_comm_workbook(output_path: Path) -> None:
    wb = Workbook()

    # —— 怎么用 ——
    ws0 = wb.active
    ws0.title = "怎么用"
    write_title(ws0, "肇事方沟通流程表 · 内部使用 · 2026-08-17", 2)
    ws0.cell(2, 1, "项")
    ws0.cell(2, 2, "内容")
    style_header(ws0, 2, 2)
    cover = [
        ("这一场目标", " ".join(COMM_GOAL)),
        ("现在请不请律师", LAWYER_NOW["answer"] + " " + LAWYER_NOW["why_yes_internal"]),
        ("方案甲（推荐）", f"{SCHEMES[0]['use']}。{SCHEMES[0]['steps']} 律师：{SCHEMES[0]['lawyer']}"),
        ("方案乙", f"{SCHEMES[1]['use']}。{SCHEMES[1]['steps']} 律师：{SCHEMES[1]['lawyer']}"),
        ("方案丙", f"{SCHEMES[2]['use']}。{SCHEMES[2]['steps']} 律师：{SCHEMES[2]['lawyer']}"),
        ("开场稿", OPENING_SCRIPT),
        ("影像结论", " ".join(MED_SUMMARY)),
        ("Word / PDF", "同目录《肇事方沟通方案与体检分析》：开场稿、接话、12 步说明、律师触发条件。"),
    ]
    for i, (k, v) in enumerate(cover, 3):
        ws0.cell(i, 1, k)
        ws0.cell(i, 2, v)
        _row_h(ws0, i, 78 if i <= 8 else 56)
    style_rows(ws0, 3, 10, 2)
    for r in range(3, 11):
        ws0.cell(r, 1).font = body_font(bold=True, color=NAVY)
        ws0.cell(r, 1).fill = fill("F8F1E9")
    widths(ws0, [22, 100])
    _page(ws0, "青岛红枫路事故 · 沟通流程")

    # —— 12 步 ——
    ws = wb.create_sheet("12步沟通顺序", 1)
    write_title(ws, "接下来沟通顺序 · 12 步可勾选 · 没拿到「过关」不要假装过关", 9)
    headers = ["步骤", "何时", "谁出面", "找谁", "做什么", "拿到什么才算过", "过不了就", "律师", "状态"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 9)
    dv = DataValidation(
        type="list",
        formula1='"未开始,进行中,已完成,卡住转方案乙,卡住转方案丙,作废"',
        allow_blank=True,
    )
    dv.error = "请选下拉状态"
    dv.errorTitle = "状态"
    ws.add_data_validation(dv)
    dv.add("I3:I20")
    for i, s in enumerate(FLOW, 3):
        vals = [s["n"], s["when"], s["who"], s["to"], s["do"], s["get"], s["fail"], s["lawyer"], "未开始"]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
        _row_h(ws, i, 72)
    style_rows(ws, 3, 14, 9)
    for r in range(3, 15):
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.cell(r, 9).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths(ws, [8, 16, 16, 22, 42, 28, 28, 28, 16])
    _page(ws, "12 步沟通顺序 · 可勾选")

    # —— 对方接话 ——
    ws2 = wb.create_sheet("对方一句话怎么接", 2)
    write_title(ws2, "骑手常见接话 · 照着回 · 不要自己加戏", 2)
    ws2.cell(2, 1, "对方可能说")
    ws2.cell(2, 2, "你怎么接")
    style_header(ws2, 2, 2)
    for i, (a, b) in enumerate(RIDER_REPLIES, 3):
        ws2.cell(i, 1, a)
        ws2.cell(i, 2, b)
        _row_h(ws2, i, 52)
    style_rows(ws2, 3, 2 + len(RIDER_REPLIES), 2)
    widths(ws2, [36, 90])
    _page(ws2, "对方一句话怎么接")

    # —— 可以说 / 不可以说 ——
    ws3 = wb.create_sheet("可以说与不可以说", 3)
    write_title(ws3, "开场必须说清 · 绝对不能说 · 这一场要拿到", 3)
    for i, h in enumerate(["类型", "原句", "备注"], 1):
        ws3.cell(2, i, h)
    style_header(ws3, 2, 3)
    rows = []
    for x in MUST_GET:
        rows.append(("这一场要拿到", x, "没拿到就不要结束电话"))
    for x in TALKING_POINTS["to_rider"]:
        rows.append(("必须说的态度", x, "对骑手"))
    for x in MED_FOR_TALK:
        rows.append(("伤情口径", x, "只说报告上有的"))
    for x in BANS:
        rows.append(("绝对不说", x, "说了会把谈判或处罚风险推高"))
    for x in TALKING_POINTS["do_not_say"]:
        rows.append(("绝对不说", x, "备忘录禁用清单"))
    for i, (typ, sent, note) in enumerate(rows, 3):
        ws3.cell(i, 1, typ)
        ws3.cell(i, 2, sent)
        ws3.cell(i, 3, note)
        _row_h(ws3, i, 36)
    style_rows(ws3, 3, 2 + len(rows), 3)
    color_map = {
        "这一场要拿到": (NAVY, "D6E3F0"),
        "必须说的态度": (GREEN, "C8E6C9"),
        "伤情口径": (NAVY, LIGHT),
        "绝对不说": (RED, "FFCDD2"),
    }
    for i, (typ, _, _) in enumerate(rows, 3):
        color, bg = color_map[typ]
        ws3.cell(i, 1).font = body_font(bold=True, color=color)
        ws3.cell(i, 1).fill = fill(bg)
    widths(ws3, [18, 78, 28])
    _page(ws3, "可以说与不可以说")

    # —— 律师决策 ——
    ws4 = wb.create_sheet("律师决策", 4)
    write_title(ws4, "现在不请诉讼律师当面谈 · 触发条件出现后再发函/起诉", 3)
    for i, h in enumerate(["阶段 / 触发", "做不做", "说明 / 文件"], 1):
        ws4.cell(2, i, h)
    style_header(ws4, 2, 3)
    ws4.cell(3, 1, "现在（认定书前、保险未进场）")
    ws4.cell(3, 2, "对内岳父把关；不请诉讼律师当面谈；不先发律师函")
    ws4.cell(3, 3, "；".join(LAWYER_NOW["why_not"]))
    _row_h(ws4, 3, 90)
    actions = [
        ("书面催告保险进场，限 3–5 个工作日回复", "律师函 + 送达记录"),
        ("函告拒赔点，要求指定医院并垫付已发生费用", "律师函 + 病历/发票清单"),
        ("函告保留追偿，固定微信/电话记录", "律师函 + 聊天截图"),
        ("书面拒绝总包，声明后续评残和误工另计", "律师见证回复"),
        ("顾问审查理赔清单，尚未必须诉讼代理", "理赔材料清单"),
        ("委托司法鉴定（治疗终结后），用途写交通事故", "鉴定委托书"),
        ("起诉骑手、视情况追加平台；执行盯平台险", "起诉状 + 鉴定意见"),
    ]
    for i, (trig, (act, docn)) in enumerate(zip(LAWYER_NOW["triggers"], actions), 4):
        ws4.cell(i, 1, trig)
        ws4.cell(i, 2, act)
        ws4.cell(i, 3, docn)
        _row_h(ws4, i, 40)
    last = 3 + 1 + len(LAWYER_NOW["triggers"])
    style_rows(ws4, 3, last, 3)
    ws4.cell(3, 1).font = body_font(bold=True, color=RED)
    ws4.cell(3, 1).fill = fill("F8F1E9")
    widths(ws4, [48, 48, 36])
    _page(ws4, "律师决策 · 触发再请")

    # —— 三套方案 ——
    ws5 = wb.create_sheet("三套方案", 5)
    write_title(ws5, "方案甲推荐现在走 · 乙是医院保险卡住 · 丙是失联拒赔后的法律路径", 6)
    for i, h in enumerate(["方案", "何时用", "怎么走", "律师", "风险", "不做"], 1):
        ws5.cell(2, i, h)
    style_header(ws5, 2, 6)
    for i, sch in enumerate(SCHEMES, 3):
        ws5.cell(i, 1, sch["name"])
        ws5.cell(i, 2, sch["use"])
        ws5.cell(i, 3, sch["steps"])
        ws5.cell(i, 4, sch["lawyer"])
        ws5.cell(i, 5, sch["risk"])
        ws5.cell(i, 6, sch["dont"])
        _row_h(ws5, i, 72)
    style_rows(ws5, 3, 5, 6)
    ws5.cell(3, 1).fill = fill("C8E6C9")
    ws5.cell(4, 1).fill = fill("FFF3CD")
    ws5.cell(5, 1).fill = fill("FFCDD2")
    widths(ws5, [28, 28, 42, 28, 32, 28])
    _page(ws5, "三套完整方案")

    # —— 出示材料 ——
    ws6 = wb.create_sheet("出示材料清单", 6)
    write_title(ws6, "可以给对方看的 · 先不要给的已写在最后一条", 3)
    for i, h in enumerate(["序号", "材料", "勾选"], 1):
        ws6.cell(2, i, h)
    style_header(ws6, 2, 3)
    for i, x in enumerate(SHOW_LIST, 3):
        ws6.cell(i, 1, i - 2)
        ws6.cell(i, 2, x)
        ws6.cell(i, 3, "☐")
        _row_h(ws6, i, 32)
    last = 2 + len(SHOW_LIST)
    style_rows(ws6, 3, last, 3)
    ws6.cell(last, 1).fill = fill("FFCDD2")
    ws6.cell(last, 2).fill = fill("FFCDD2")
    ws6.cell(last, 3).fill = fill("FFCDD2")
    widths(ws6, [10, 90, 10])
    _page(ws6, "出示材料清单")

    # —— 影像对照 ——
    ws7 = wb.create_sheet("影像三份对照", 7)
    write_title(ws7, "这不是体检 · 是齐鲁三份已审核 CT · 不能当残级", 6)
    for i, h in enumerate(["检查", "白话结论", "对骑手怎么说", "索赔", "检查号", "报告诊断"], 1):
        ws7.cell(2, i, h)
    style_header(ws7, 2, 6)
    for i, (plain, report) in enumerate(zip(CT_PLAIN, CT_REPORTS), 3):
        ws7.cell(i, 1, plain["title"])
        ws7.cell(i, 2, plain["plain"])
        ws7.cell(i, 3, plain["to_rider"])
        ws7.cell(i, 4, plain["claim"])
        ws7.cell(i, 5, report["no"])
        ws7.cell(i, 6, "；".join(report["diagnosis"]))
        _row_h(ws7, i, 72)
    style_rows(ws7, 3, 5, 6)
    widths(ws7, [28, 36, 36, 16, 16, 36])
    _page(ws7, "影像三份对照")

    # —— 今日记录 ——
    ws8 = wb.create_sheet("今日通话记录", 8)
    write_title(ws8, "每打一通记一行 · 不要只靠脑子", 7)
    for i, h in enumerate(["日期", "对象", "方式", "对方说了什么", "我方承诺", "待办", "截图/录音编号"], 1):
        ws8.cell(2, i, h)
    style_header(ws8, 2, 7)
    for i in range(3, 13):
        for c in range(1, 8):
            cell = ws8.cell(i, c, "")
            cell.font = body_font()
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = THIN
            if i % 2 == 0:
                cell.fill = fill(LIGHT)
        _row_h(ws8, i, 28)
    widths(ws8, [14, 18, 12, 36, 28, 28, 20])
    _page(ws8, "今日通话记录")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print("Wrote", output_path)


if __name__ == "__main__":
    out = Path(__file__).resolve().parent.parent / "deliverables" / "青岛红枫路交通事故_肇事方沟通流程表_20260817.xlsx"
    build_comm_workbook(out)
