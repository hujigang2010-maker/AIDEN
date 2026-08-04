#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 WAIC 2026 相关多份名录、日程、联系人、活动与文档资料整合为一份总表。

输出：output/WAIC2026_全量资源整合总表.xlsx
"""

from __future__ import annotations

import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader

UPLOADS = Path("/home/ubuntu/.cursor/projects/workspace/uploads")
OUT_DIR = Path("/workspace/output")
OUT_PATH = OUT_DIR / "WAIC2026_全量资源整合总表.xlsx"

# 去重后选用的主数据源（按内容哈希已确认重复）
SRC = {
    "品牌联系人总库": UPLOADS / "WAIC2026____________7039.xlsx",
    "参展商汇总": UPLOADS / "WAIC2026_________8923.xlsx",
    "论坛日程": UPLOADS / "WAIC_2026___________________fd26.xlsx"
    if (UPLOADS / "WAIC_2026___________________fd26.xlsx").exists()
    else UPLOADS / "WAIC_2026___________________1__1__b247.xlsx",
    "分级联系": UPLOADS / "WAIC2026_________935c.xlsx",
    "活动转化": UPLOADS / "WAIC2026__________0017.xlsx",
    "演讲视频": UPLOADS / "WAIC2026_______by_0718_18_____e405.xlsx",
    "Ai联盟名录": UPLOADS / "Ai__-______-____-_56ba.xlsx",
    "WAIC联系人精选": UPLOADS / "WAIC______________456a.xlsx",
    "具身智能观察": UPLOADS / "6-WAIC____150_______________a9af.docx",
    "参会攻略": UPLOADS / "2026_WAIC___________AI______86ff.pdf",
    "演讲稿": UPLOADS / "5-WAIC_2026_______9443.pdf",
    "完整议程": UPLOADS / "2-WAIC_2026______8a48.docx",
    "展位图": UPLOADS / "3-WAIC2026______________69be.docx",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="666666")
SECTION_FONT = Font(name="微软雅黑", bold=True, size=12, color="1F4E79")
BODY_FONT = Font(name="微软雅黑", size=10)
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")


def s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def clean_text(text: str) -> str:
    return text.encode("utf-8", errors="replace").decode("utf-8")


def style_header(ws: Worksheet, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autosize(ws: Worksheet, max_width: int = 42, min_width: int = 8) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 200)):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value)
            # 中文按约 1.7 宽度估算
            w = min(max_width, max(min_width, int(len(val) * 1.2) + 2))
            widths[cell.column] = max(widths.get(cell.column, min_width), w)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def write_table(
    ws: Worksheet,
    headers: list[str],
    rows: list[list],
    start_row: int = 1,
    freeze: str | None = None,
) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(headers))
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN
            if (r_idx - start_row) % 2 == 0:
                cell.fill = ALT_FILL
    end = start_row + len(rows)
    if freeze:
        ws.freeze_panes = freeze
    elif start_row == 1:
        ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{max(end, start_row)}"
    ws.row_dimensions[start_row].height = 22
    return end


def read_sheet_table(
    path: Path,
    sheet_name: str,
    header_keywords: tuple[str, ...] | None = None,
    header_row_hint: int | None = None,
) -> tuple[list[str], list[list]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    header_idx = None
    if header_row_hint is not None:
        header_idx = header_row_hint
    elif header_keywords:
        for i, row in enumerate(all_rows):
            cells = [s(c) for c in row]
            if any(k in cells for k in header_keywords):
                # 选择包含最多关键词的行
                score = sum(1 for k in header_keywords if k in cells)
                if score >= max(1, len(header_keywords) // 2):
                    header_idx = i
                    break
    if header_idx is None:
        # 回退：找第一个非空行
        for i, row in enumerate(all_rows):
            if any(s(c) for c in row):
                header_idx = i
                break
    if header_idx is None:
        return [], []

    raw_headers = all_rows[header_idx]
    # 截断尾部空列
    last = 0
    for i, h in enumerate(raw_headers):
        if s(h):
            last = i
    headers = [s(h) or f"列{i+1}" for i, h in enumerate(raw_headers[: last + 1])]

    data: list[list] = []
    for row in all_rows[header_idx + 1 :]:
        vals = [(c if c is not None else "") for c in row[: len(headers)]]
        if not any(s(v) for v in vals):
            continue
        # 补齐长度
        while len(vals) < len(headers):
            vals.append("")
        data.append(vals)
    return headers, data


def add_title_block(ws: Worksheet, title: str, subtitle: str = "") -> int:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        row = 3
    return row + 1


def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("00_使用说明", 0)
    ws["A1"] = "WAIC 2026 全量资源整合总表"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    lines = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("整合目标", "把分散的参展商、联系人、论坛、演讲、活动转化等资料收敛到一份可筛选、可调用的工作簿。"),
        ("去重原则", "同内容文件只保留一份；参展商以「全部汇总」963家为准；品牌/联系人优先采用已去重的资源总库。"),
        ("推荐用法", "先看「01_数据总览」与「02_资料索引」→ 找公司用「03/05」→ 找人用「06/07」→ 找论坛用「09」→ 找视频用「11」。"),
        ("筛选提示", "每个数据表已开启筛选与冻结首行；可用行业、展馆、优先级、赛道等列快速过滤。"),
        ("文档说明", "PDF/DOCX（攻略、演讲稿、议程图、展位图）无法完整表格化，已提取要点进对应 Sheet，原件见资料索引。"),
    ]
    r = 3
    for k, v in lines:
        ws.cell(row=r, column=1, value=k).font = Font(name="微软雅黑", bold=True, size=10)
        ws.cell(row=r, column=1).fill = NOTE_FILL
        ws.cell(row=r, column=2, value=v).font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="工作表导航").font = SECTION_FONT
    r += 1
    nav = [
        ("00_使用说明", "本说明与使用建议"),
        ("01_数据总览", "关键指标一览，快速判断库存规模"),
        ("02_资料索引", "全部源文件清单、去重状态与用途"),
        ("03_WAIC参展商", "WAIC 2026 参展商全量（963家）"),
        ("04_行业统计", "参展商行业分布"),
        ("05_品牌主库", "跨展会品牌主数据（含评分与优先级）"),
        ("06_联系人库", "去重后的联系人/邮箱资源池"),
        ("07_分级联系执行", "A/B级可触达线索执行台账"),
        ("08_触达话术", "按优先级的邮件触达框架"),
        ("09_论坛日程", "175场论坛完整日程"),
        ("10_赛道分类统计", "论坛赛道分布与分析结论"),
        ("11_演讲视频", "官方演讲视频与链接整理"),
        ("12_活动项目池", "会后转化活动项目设计"),
        ("13_品牌转化台账", "品牌邀约/到场/商机记录"),
        ("14_活动设计模板", "活动策划画布"),
        ("15_活动复盘记录", "活动复盘空白台账"),
        ("16_评分分级规则", "品牌匹配评分规则说明"),
        ("17_具身智能场景观察", "H3具身智能企业场景观察结构化摘要"),
        ("18_参会攻略要点", "线下参会攻略结构化要点"),
        ("19_演讲稿目录", "全量演讲稿 PDF 目录索引"),
        ("20_跨展会名录统计", "Ai联盟等多展会名录来源条数"),
    ]
    write_table(ws, ["工作表", "内容说明"], [[a, b] for a, b in nav], start_row=r)
    autosize(ws, max_width=55)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55


def build_source_index(wb: Workbook) -> None:
    ws = wb.create_sheet("02_资料索引")
    start = add_title_block(
        ws,
        "源文件索引与去重说明",
        "共27份上传文件；内容完全相同的副本已合并，仅保留一份进入总表。",
    )
    rows = [
        ["Ai联盟-中国会展名录-无电话版-.xlsx / 副本 / (2)", "跨展会联系人池", "已并入品牌主库+联系人库", "保留", "3份内容相同，只取1份"],
        ["WAIC世界人工智能世博会无电话版.xlsx / 副本", "WAIC精选联系人", "已并入品牌主库+联系人库", "保留", "2份内容相同"],
        ["WAIC2026_参展商信息扫描*.xlsx / 1- / (16) / 副本 等", "WAIC参展商明细", "03_WAIC参展商", "保留", "多份同为963家，数据一致"],
        ["2026WAIC-参展商及行业全量信息汇总.xlsx", "参展商主数据同源", "03_WAIC参展商", "保留", "与扫描表公司清单一致"],
        ["WAIC2026_品牌与联系人资源总库.xlsx", "品牌+联系人主库", "05/06及相关", "主数据源", "已完成跨表去重的最完整CRM"],
        ["WAIC2026_分级联系执行表.xlsx", "A/B线索执行", "07/08", "主数据源", "首批500条可触达线索"],
        ["WAIC2026_活动转化与复盘表.xlsx", "活动转化漏斗", "12~15", "主数据源", "项目池+转化台账"],
        ["WAIC 2026世界人工智能大会日程表（分类导航版）*.xlsx", "论坛日程175场", "09/10", "主数据源", "与WAIC 2027.xlsx内容相同"],
        ["WAIC 2026世界人工智能大会论坛日程统计｜分类导航工具表.xlsx", "日程统计工具", "09/10", "同日程", "与分类导航版同内容"],
        ["WAIC 2027.xlsx", "论坛日程", "09/10", "同日程", "文件名虽为2027，内容为WAIC2026论坛"],
        ["WAIC2026演讲视频整理 by 0718 18点*.xlsx / 4-", "演讲视频链接", "11_演讲视频", "主数据源", "3份内容相同"],
        ["2026 WAIC 线下参会攻略｜脉脉 AI 创作者版.pdf", "参会攻略", "18_参会攻略要点", "要点抽取", "全文见原PDF"],
        ["5-WAIC 2026 全量演讲稿.pdf", "演讲稿全文", "19_演讲稿目录", "目录抽取", "230页全文过大，仅保留目录"],
        ["2-WAIC 2026 完整议程.docx", "议程长图", "02_资料索引备注", "图像文档", "以图片为主，无法表格化"],
        ["3-WAIC2026世界人工智能大会高清展位图.docx", "展位平面图", "02_资料索引备注", "图像文档", "13张图片，建议原件查阅"],
        ["6-WAIC观察：超150家具身智能企业，在卷什么场景.docx", "具身智能观察", "17_具身智能场景观察", "结构化摘要", "已抽取企业与场景"],
    ]
    write_table(
        ws,
        ["源文件（含副本）", "原始用途", "落入本总表位置", "处理状态", "备注"],
        rows,
        start_row=start,
        freeze=f"A{start+1}",
    )
    autosize(ws, max_width=48)


def build_overview(wb: Workbook, stats: dict) -> None:
    ws = wb.create_sheet("01_数据总览", 1)
    start = add_title_block(
        ws,
        "数据总览",
        "以下指标来自去重后的主数据，便于快速判断可调用资源规模。",
    )
    cards = [
        ["WAIC参展商", stats["exhibitors"], "家"],
        ["品牌主库", stats["brands"], "家"],
        ["联系人", stats["contacts"], "人"],
        ["有邮箱联系人", stats["emails"], "人"],
        ["分级线索(A/B)", stats["leads"], "条"],
        ["论坛日程", stats["forums"], "场"],
        ["演讲视频条目", stats["videos"], "条"],
        ["活动项目", stats["acts"], "个"],
        ["转化台账品牌", stats["conversions"], "家"],
        ["源文件（去重前）", 27, "份"],
        ["内容唯一文件", stats["unique_files"], "份"],
    ]
    write_table(ws, ["指标", "数量", "单位"], cards, start_row=start, freeze=f"A{start+1}")

    r = start + len(cards) + 3
    ws.cell(row=r, column=1, value="调用路径建议").font = SECTION_FONT
    tips = [
        ["找某家WAIC展商展位/行业", "03_WAIC参展商 → 筛公司名称/展馆/行业"],
        ["找可发邮件的联系人", "06_联系人库 → 筛邮箱有效=是；或 07_分级联系执行"],
        ["按赛道约论坛", "09_论坛日程 → 筛赛道分类/日期/展馆"],
        ["回看官方演讲", "11_演讲视频 → 按论坛/演讲人筛选"],
        ["推进会后合作", "12_活动项目池 + 13_品牌转化台账"],
        ["看具身智能卷什么场景", "17_具身智能场景观察"],
    ]
    write_table(ws, ["需求", "怎么用"], tips, start_row=r + 1)
    autosize(ws, max_width=60)


def copy_exhibitor(wb: Workbook) -> int:
    headers, data = read_sheet_table(
        SRC["参展商汇总"],
        "全部汇总",
        header_keywords=("序号", "公司名称", "展馆"),
    )
    ws = wb.create_sheet("03_WAIC参展商")
    start = add_title_block(ws, "WAIC 2026 参展商全量汇总", f"共 {len(data)} 家｜来源：参展商信息扫描/全量汇总（去重后）")
    write_table(ws, headers, data, start_row=start, freeze=f"A{start+1}")
    autosize(ws, max_width=36)
    return len(data)


def copy_industry(wb: Workbook) -> None:
    path = SRC["参展商汇总"]
    wb0 = load_workbook(path, read_only=True, data_only=True)
    ws0 = wb0["行业统计"]
    rows = [list(r) for r in ws0.iter_rows(values_only=True)]
    wb0.close()

    # 提取一级大类表
    big = []
    sub = []
    mode = None
    for row in rows:
        cells = [s(c) for c in row]
        joined = "".join(cells)
        if "大类" in cells and "展商数量" in cells:
            mode = "big"
            continue
        if ("细分" in joined or "领域" in cells) and "展商数量" in cells:
            mode = "sub"
            continue
        if mode == "big" and cells[0] and cells[0] not in ("大类", "一级大类", ""):
            if cells[1].replace(".", "", 1).isdigit() or cells[1].endswith("%") or cells[1].isdigit():
                big.append([cells[0], cells[1], cells[2] if len(cells) > 2 else ""])
            elif any(cells[1:]):
                big.append([cells[0], cells[1], cells[2] if len(cells) > 2 else ""])
        if mode == "sub" and cells[0] and "细分" not in cells[0]:
            sub.append(cells[:4])

    ws = wb.create_sheet("04_行业统计")
    start = add_title_block(ws, "参展商行业分布统计", "来自参展商汇总「行业统计」页")
    ws.cell(row=start, column=1, value="一级大类").font = SECTION_FONT
    write_table(ws, ["大类", "展商数量", "占比"], big, start_row=start + 1)
    r = start + 1 + len(big) + 3
    ws.cell(row=r, column=1, value="细分领域").font = SECTION_FONT
    # 若未能解析细分，直接整页粘贴精简版
    if not sub:
        # fallback: dump non-empty rows
        cleaned = []
        for row in rows:
            vals = [s(c) for c in row[:4]]
            if any(vals):
                cleaned.append(vals)
        write_table(ws, ["列1", "列2", "列3", "列4"], cleaned, start_row=r + 1)
    else:
        # normalize header length
        maxc = max(len(x) for x in sub) if sub else 4
        headers = ["细分领域", "展商数量", "占比", "备注"][:maxc]
        write_table(ws, headers, [x[:maxc] for x in sub], start_row=r + 1)
    autosize(ws)


def copy_brand_and_contacts(wb: Workbook) -> tuple[int, int, int]:
    b_headers, b_data = read_sheet_table(
        SRC["品牌联系人总库"],
        "品牌主库",
        header_keywords=("品牌ID", "公司名称", "WAIC参展"),
    )
    c_headers, c_data = read_sheet_table(
        SRC["品牌联系人总库"],
        "联系人库",
        header_keywords=("联系人ID", "品牌ID", "公司名称"),
    )

    ws = wb.create_sheet("05_品牌主库")
    start = add_title_block(
        ws,
        "品牌主库",
        f"共 {len(b_data)} 家｜已融合 WAIC参展商 + Ai联盟多展会名录 + WAIC精选联系人",
    )
    write_table(ws, b_headers, b_data, start_row=start, freeze=f"A{start+1}")
    autosize(ws, max_width=28)

    ws2 = wb.create_sheet("06_联系人库")
    start2 = add_title_block(
        ws2,
        "联系人库",
        f"共 {len(c_data)} 人｜按邮箱或公司+姓名+职位去重，保留来源追溯",
    )
    write_table(ws2, c_headers, c_data, start_row=start2, freeze=f"A{start2+1}")
    autosize(ws2, max_width=28)

    # 有邮箱人数
    email_col = None
    for i, h in enumerate(c_headers):
        if h in ("邮箱", "首选邮箱"):
            email_col = i
            break
    emails = 0
    if email_col is not None:
        for row in c_data:
            ev = s(row[email_col])
            if ev and "@" in ev and ev not in ("/", "-", "无"):
                emails += 1
    return len(b_data), len(c_data), emails


def copy_leads_and_scripts(wb: Workbook) -> int:
    headers, data = read_sheet_table(
        SRC["分级联系"],
        "分级联系清单",
        header_keywords=("线索ID", "优先级", "公司名称"),
    )
    ws = wb.create_sheet("07_分级联系执行")
    start = add_title_block(ws, "分级联系执行清单", f"共 {len(data)} 条｜建议优先推进 A/B 级且有邮箱线索")
    write_table(ws, headers, data, start_row=start, freeze=f"A{start+1}")
    autosize(ws, max_width=28)

    h2, d2 = read_sheet_table(
        SRC["分级联系"],
        "触达话术框架",
        header_keywords=("优先级", "目标", "邮件标题框架"),
    )
    ws2 = wb.create_sheet("08_触达话术")
    start2 = add_title_block(ws2, "触达话术框架", "每封至少替换品牌事实、匹配理由和一个明确的15分钟沟通请求")
    write_table(ws2, h2, d2, start_row=start2, freeze=f"A{start2+1}")
    autosize(ws2, max_width=40)
    return len(data)


def copy_forums(wb: Workbook) -> int:
    headers, data = read_sheet_table(
        SRC["论坛日程"],
        "完整175",
        header_keywords=("序号", "赛道分类", "论坛名称"),
    )
    # 只保留前7列有效字段
    keep = min(7, len(headers))
    headers = headers[:keep]
    data = [r[:keep] for r in data]

    ws = wb.create_sheet("09_论坛日程")
    start = add_title_block(ws, "WAIC 2026 论坛完整日程（175场）", "可按赛道/日期/展馆筛选；场馆分表已合并到本表")
    write_table(ws, headers, data, start_row=start, freeze=f"A{start+1}")
    autosize(ws, max_width=40)

    # 赛道统计
    h2, d2 = read_sheet_table(
        SRC["论坛日程"],
        "赛道分类",
        header_keywords=("排名", "赛道分类", "论坛数量"),
    )
    h3, d3 = read_sheet_table(
        SRC["论坛日程"],
        "核心结论",
        header_keywords=None,
        header_row_hint=0,
    )
    ws2 = wb.create_sheet("10_赛道分类统计")
    start2 = add_title_block(ws2, "论坛赛道统计与核心结论", "")
    ws2.cell(row=start2, column=1, value="赛道分布").font = SECTION_FONT
    end = write_table(ws2, h2, d2, start_row=start2 + 1)
    r = end + 3
    ws2.cell(row=r, column=1, value="核心结论原文摘录").font = SECTION_FONT
    # 核心结论页结构不规整，整理为 条目/内容
    conclusions = []
    for row in d3:
        vals = [s(c) for c in row if s(c)]
        if not vals:
            continue
        if len(vals) == 1:
            conclusions.append([vals[0], ""])
        else:
            conclusions.append([vals[0], " | ".join(vals[1:3])])
    write_table(ws2, ["条目", "内容"], conclusions, start_row=r + 1)
    autosize(ws2, max_width=45)
    return len(data)


def copy_videos(wb: Workbook) -> int:
    headers, data = read_sheet_table(
        SRC["演讲视频"],
        "WAIC2026演讲视频整理",
        header_keywords=("类别(论坛)", "日期", "演讲人"),
    )
    ws = wb.create_sheet("11_演讲视频")
    start = add_title_block(
        ws,
        "WAIC 2026 演讲视频整理",
        "截止 2026-07-18 18:00｜仅收录官方及承办方原汁原味视频/全文",
    )
    write_table(ws, headers, data, start_row=start, freeze=f"A{start+1}")
    # 链接列高亮
    link_col = None
    for i, h in enumerate(headers, 1):
        if "链接" in h:
            link_col = i
            break
    if link_col:
        for r in range(start + 1, start + 1 + len(data)):
            cell = ws.cell(row=r, column=link_col)
            if s(cell.value).startswith("http"):
                cell.hyperlink = s(cell.value)
                cell.font = LINK_FONT
    autosize(ws, max_width=40)

    # 官方视频源说明
    h2, d2 = read_sheet_table(
        SRC["演讲视频"],
        "官方视频源说明",
        header_keywords=("项目", "说明"),
    )
    r = start + len(data) + 4
    ws.cell(row=r, column=1, value="官方视频源说明").font = SECTION_FONT
    write_table(ws, h2, d2, start_row=r + 1)
    return len(data)


def copy_activities(wb: Workbook) -> tuple[int, int]:
    mapping = [
        ("活动项目池", "12_活动项目池", "活动项目池"),
        ("品牌转化台账", "13_品牌转化台账", "品牌转化台账"),
        ("活动设计模板", "14_活动设计模板", "活动设计模板"),
        ("活动复盘记录", "15_活动复盘记录", "活动复盘记录"),
    ]
    counts = {}
    for src_sheet, out_name, title in mapping:
        headers, data = read_sheet_table(
            SRC["活动转化"],
            src_sheet,
            header_keywords=None,
        )
        # 对设计模板，header可能是「模块」
        if src_sheet == "活动设计模板":
            headers, data = read_sheet_table(
                SRC["活动转化"],
                src_sheet,
                header_keywords=("模块", "必须回答的问题"),
            )
        elif src_sheet != "活动复盘记录":
            # 找含ID的表头
            keywords = {
                "活动项目池": ("项目ID", "活动名称"),
                "品牌转化台账": ("转化ID", "品牌ID"),
                "活动复盘记录": ("活动项目ID", "活动日期"),
            }.get(src_sheet, None)
            if keywords:
                headers, data = read_sheet_table(
                    SRC["活动转化"], src_sheet, header_keywords=keywords
                )
        else:
            headers, data = read_sheet_table(
                SRC["活动转化"],
                src_sheet,
                header_keywords=("活动项目ID", "活动日期"),
            )

        ws = wb.create_sheet(out_name)
        start = add_title_block(ws, title, f"共 {len(data)} 行")
        if headers:
            write_table(ws, headers, data, start_row=start, freeze=f"A{start+1}")
        autosize(ws, max_width=28)
        counts[out_name] = len(data)
    return counts.get("12_活动项目池", 0), counts.get("13_品牌转化台账", 0)


def copy_score_rules(wb: Workbook) -> None:
    headers, data = read_sheet_table(
        SRC["品牌联系人总库"],
        "评分与分级规则",
        header_keywords=("维度", "权重"),
    )
    ws = wb.create_sheet("16_评分分级规则")
    start = add_title_block(ws, "评分与分级规则", "用于品牌排队，可按个人资源与活动目标调整黄色输入列")
    write_table(ws, headers, data, start_row=start, freeze=f"A{start+1}")
    autosize(ws, max_width=36)


def build_embodied_sheet(wb: Workbook) -> None:
    doc = Document(SRC["具身智能观察"])
    paras = [
        p.text.strip()
        for p in doc.paragraphs
        if p.text.strip() and p.text.strip() not in ("image.png",)
    ]
    # 去重标题
    if len(paras) >= 2 and paras[0] == paras[1]:
        paras = paras[1:]

    # 手工结构化场景-企业（基于文章内容）
    observations = [
        ["零售服务", "银河通用", "Galbot G1 便利店值守；烤面包、倒饮料早餐制作全流程", "商业零售"],
        ["零售服务", "梅卡曼德", "人形轮式机器人货架取货互动，扫码领赠品", "商业零售"],
        ["零售服务", "穹彻智能", "药房取药递药；洗衣洗烘方案（即将进酒店）", "药房/酒店"],
        ["零售服务", "蚂蚁灵波", "与穹彻智能联合药房场景", "药房"],
        ["零售服务", "擎朗智能", "具身社区：咖啡亭、甜品屋、零售店、酒店洗衣房，零遥操", "社区商业"],
        ["零售服务", "京东×星尘机器人", "智能零售工作站首秀", "零售"],
        ["工业物流", "夸父（相关展项）", "工业场景作业计时展示", "工业"],
        ["工业物流", "银河通用", "Galbot S1 搬运30kg、打螺丝等工业作业", "工业"],
        ["工业物流", "极佳视界", "Maker H01 上下料、搬运、关节组装", "工业装配"],
        ["工业物流", "星动纪元", "物流全栈方案，人形机器人分拣快递", "物流"],
        ["工业物流", "优艾智合", "五台隙锋人形+移动搬运机器人协作线边仓拣配", "工业产线"],
        ["工业物流", "极智嘉", "轮式人形 Gino 1 料箱抓取搬运与仓储拣选", "仓储"],
        ["工业物流", "浙江人形机器人创新中心", "NAVIAI-WA2 轮臂式机器人料箱搬运", "仓储物流"],
        ["零部件", "月泉仿生等", "灵巧手展示；行业讨论性能-成本-可靠性三角，抓夹仍占出货主流", "核心零部件"],
    ]

    sections = []
    current = ""
    for p in paras:
        if re.fullmatch(r"0\d", p):
            continue
        if p in (
            "零售服务，机器人“扎堆儿”",
            "最卷场景：工业、物流",
            "从本体自嗨到零部件上桌",
        ) or (len(p) < 30 and ("场景" in p or "零部件" in p or "零售" in p)):
            current = p
            sections.append([current, ""])
        elif current and len(p) > 40:
            if sections and sections[-1][1] == "":
                sections[-1][1] = p[:300]
            else:
                sections.append([current, p[:300]])

    ws = wb.create_sheet("17_具身智能场景观察")
    start = add_title_block(
        ws,
        "具身智能场景观察（结构化摘要）",
        "来源：WAIC观察：超150家具身智能企业，在卷什么场景｜H3展厅",
    )
    ws.cell(row=start, column=1, value="企业×场景速查").font = SECTION_FONT
    end = write_table(
        ws,
        ["场景大类", "企业/品牌", "展示内容", "落地方向"],
        observations,
        start_row=start + 1,
    )
    r = end + 3
    ws.cell(row=r, column=1, value="文章章节要点").font = SECTION_FONT
    write_table(ws, ["章节", "要点摘要"], sections[:20], start_row=r + 1)
    autosize(ws, max_width=48)


def build_guide_sheet(wb: Workbook) -> None:
    reader = PdfReader(str(SRC["参会攻略"]))
    pages = []
    for page in reader.pages:
        pages.append(clean_text(page.extract_text() or ""))

    full = "\n".join(pages)
    tips = [
        ["时间少怎么逛", "优先：世博 H3 机器人 → H1 大模型/Agent → 西岸 AI 终端 → 张江算力/芯片"],
        ["拍摄建议", "优先拍动态演示、展位号与公司名、能说明产品用途的展板关键词"],
        ["世博展览馆入口", "观众：博成路850北门、周家渡路E1/E2、世博馆路W1；论坛：博成路850北门、周家渡E2"],
        ["世博地铁", "13号线世博大道站1号口；8号线中华艺术宫站3号口；7号线耀华路站1号口"],
        ["世博中心", "浦东新区世博大道1500号2号门；地铁8/13号线同上"],
        ["张江科学会堂", "浦东新区海科路1393号S3门；地铁13号线学林路站4号口"],
        ["西岸国际会展中心", "徐汇区龙耀路7号北门；地铁11号线龙耀路站1号口"],
        ["半天路线", "世博 H3 机器人 → H1 大模型/Agent｜适合现场视频与落地观察"],
        ["体验感路线", "徐汇西岸 AI 终端/内容工具 → 世博 H3 机器人"],
        ["行业深度路线", "世博 H2 算力基建 → 张江芯片/智算 → H4 创投"],
        ["世博展览馆定位", "主展区：大模型、Agent、机器人、算力、创投，最值得投入时间"],
    ]

    # 尝试从正文抽「重点公司」类短句
    company_hits = []
    for m in re.finditer(r"([^\n。；]{2,20}(?:科技|智能|机器人|芯片|大模型)[^\n。；]{0,20})", full):
        t = m.group(1).strip()
        if 4 <= len(t) <= 30 and t not in company_hits:
            company_hits.append(t)
        if len(company_hits) >= 30:
            break

    ws = wb.create_sheet("18_参会攻略要点")
    start = add_title_block(
        ws,
        "线下参会攻略要点（结构化）",
        "来源：2026 WAIC 线下参会攻略｜脉脉 AI 创作者版.pdf｜完整内容请回看原PDF",
    )
    end = write_table(ws, ["主题", "要点"], tips, start_row=start, freeze=f"A{start+1}")
    if company_hits:
        r = end + 3
        ws.cell(row=r, column=1, value="正文提及的公司/主题词（自动抽取，供检索）").font = SECTION_FONT
        write_table(
            ws,
            ["序号", "提及内容"],
            [[i + 1, x] for i, x in enumerate(company_hits)],
            start_row=r + 1,
        )
    autosize(ws, max_width=55)


def build_speech_toc(wb: Workbook) -> None:
    reader = PdfReader(str(SRC["演讲稿"]))
    # 目录通常在前几页（含跨页续行）
    toc_text = []
    for i in range(min(8, len(reader.pages))):
        toc_text.append(clean_text(reader.pages[i].extract_text() or ""))
    text = "\n".join(toc_text)

    # 先按行拼接被 PDF 拆开的标题，再匹配点线+页码
    raw_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    # 去掉罗马页码/目录标题噪声
    raw_lines = [
        ln
        for ln in raw_lines
        if ln not in ("目 录", "目录", "I", "II", "III", "IV", "V")
        and not re.fullmatch(r"[IVX]+", ln)
    ]

    merged: list[str] = []
    buf = ""
    for ln in raw_lines:
        # 新条目起点
        is_new = bool(
            re.match(r"^[一二三四五六七八九十百]+、", ln)
            or re.match(r"^（[一二三四五六七八九十]+）", ln)
            or re.match(r"^\([一二三四五六七八九十]+\)", ln)
        )
        if is_new and buf:
            merged.append(buf)
            buf = ln
        else:
            buf = (buf + ln) if buf else ln
        # 若本段已带点线页码，收束
        if re.search(r"\.{3,}\s*\d+\s*$", buf) or re.search(r"…+\s*\d+\s*$", buf):
            merged.append(buf)
            buf = ""
    if buf:
        merged.append(buf)

    entries = []
    for item in merged:
        item = re.sub(r"\s+", "", item)  # 目录里空格噪声较多
        m = re.search(r"^(.*?)[\.·•…]{2,}(\d+)$", item)
        if m:
            title = m.group(1).strip()
            page = m.group(2)
            if len(title) >= 4:
                entries.append([title, page])
            continue
        # 无点线时：末尾数字当页码
        m2 = re.search(r"^(.*?)(\d+)$", item)
        if m2 and len(m2.group(1)) >= 6:
            entries.append([m2.group(1), m2.group(2)])

    # 去重保序
    seen = set()
    uniq = []
    for a, b in entries:
        if a in seen:
            continue
        seen.add(a)
        uniq.append([a, b])

    ws = wb.create_sheet("19_演讲稿目录")
    start = add_title_block(
        ws,
        "全量演讲稿目录索引",
        f"来源：5-WAIC 2026 全量演讲稿.pdf（共 {len(reader.pages)} 页）｜此处仅目录，便于定位后再打开PDF",
    )
    if not uniq:
        uniq = [["未能自动解析目录，请直接打开原PDF", str(len(reader.pages))]]
    write_table(ws, ["演讲/章节", "页码"], uniq, start_row=start, freeze=f"A{start+1}")
    autosize(ws, max_width=70)


def build_expo_source_stats(wb: Workbook) -> None:
    headers, data = read_sheet_table(
        SRC["Ai联盟名录"],
        "📊来源统计",
        header_keywords=("来源", "条数"),
    )
    # 也汇总各 sheet 行数
    wb0 = load_workbook(SRC["Ai联盟名录"], read_only=True, data_only=True)
    sheet_stats = []
    for name in wb0.sheetnames:
        if name.startswith("📊"):
            continue
        ws0 = wb0[name]
        n = sum(1 for i, row in enumerate(ws0.iter_rows(values_only=True)) if i >= 1 and any(s(c) for c in row))
        sheet_stats.append([name, n])
    wb0.close()

    ws = wb.create_sheet("20_跨展会名录统计")
    start = add_title_block(
        ws,
        "跨展会名录来源统计",
        "来源：Ai联盟-中国会展名录｜明细联系人已去重并入「05品牌主库/06联系人库」",
    )
    ws.cell(row=start, column=1, value="按来源标签统计").font = SECTION_FONT
    end = write_table(ws, headers or ["来源", "条数"], data, start_row=start + 1)
    r = end + 3
    ws.cell(row=r, column=1, value="按工作表统计").font = SECTION_FONT
    write_table(ws, ["工作表", "数据行数"], sheet_stats, start_row=r + 1)
    autosize(ws)


def validate_sources() -> None:
    missing = [k for k, p in SRC.items() if not p.exists()]
    if missing:
        # 论坛日程备选
        if "论坛日程" in missing:
            for cand in UPLOADS.glob("WAIC_2026*.xlsx"):
                SRC["论坛日程"] = cand
                missing.remove("论坛日程")
                break
            for cand in UPLOADS.glob("WAIC_2027*.xlsx"):
                SRC["论坛日程"] = cand
                if "论坛日程" in missing:
                    missing.remove("论坛日程")
                break
    missing = [k for k, p in SRC.items() if not p.exists()]
    if missing:
        raise FileNotFoundError(f"缺少源文件: {missing}")


def main() -> None:
    validate_sources()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # 删除默认 sheet，稍后按顺序创建
    default = wb.active
    wb.remove(default)

    build_readme(wb)

    # 先构建数据表，再回填总览
    n_ex = copy_exhibitor(wb)
    copy_industry(wb)
    n_brands, n_contacts, n_emails = copy_brand_and_contacts(wb)
    n_leads = copy_leads_and_scripts(wb)
    n_forums = copy_forums(wb)
    n_videos = copy_videos(wb)
    n_acts, n_conv = copy_activities(wb)
    copy_score_rules(wb)
    build_embodied_sheet(wb)
    build_guide_sheet(wb)
    build_speech_toc(wb)
    build_expo_source_stats(wb)
    build_source_index(wb)

    # unique file count by md5 of uploads
    import hashlib

    hashes = set()
    for f in UPLOADS.iterdir():
        hashes.add(hashlib.md5(f.read_bytes()).hexdigest())

    stats = {
        "exhibitors": n_ex,
        "brands": n_brands,
        "contacts": n_contacts,
        "emails": n_emails,
        "leads": n_leads,
        "forums": n_forums,
        "videos": n_videos,
        "acts": n_acts,
        "conversions": n_conv,
        "unique_files": len(hashes),
    }
    build_overview(wb, stats)

    # 把资料索引移到靠前：openpyxl 已按创建顺序；调整顺序
    desired = [
        "00_使用说明",
        "01_数据总览",
        "02_资料索引",
        "03_WAIC参展商",
        "04_行业统计",
        "05_品牌主库",
        "06_联系人库",
        "07_分级联系执行",
        "08_触达话术",
        "09_论坛日程",
        "10_赛道分类统计",
        "11_演讲视频",
        "12_活动项目池",
        "13_品牌转化台账",
        "14_活动设计模板",
        "15_活动复盘记录",
        "16_评分分级规则",
        "17_具身智能场景观察",
        "18_参会攻略要点",
        "19_演讲稿目录",
        "20_跨展会名录统计",
    ]
    for i, name in enumerate(desired):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(OUT_PATH)
    print(f"已生成: {OUT_PATH}")
    print("统计:", stats)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
