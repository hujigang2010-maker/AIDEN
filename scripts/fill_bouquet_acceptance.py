"""填写《绿城·潮鸣外滩-开盘花束营销验收单》。

依据《绿城中国策划活动类效果评估表-开盘花束》原文与 4 张现场照片，
完善营销验收单首页布局，并增加「效果评估表」「交付清单」页。
评估表未列金额，本脚本不编造单价或总价。
"""
from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from PIL import Image as PILImage

SRC = Path("/home/ubuntu/.cursor/projects/workspace/uploads/____feef.xlsx")
if not SRC.exists():
    SRC = Path("/home/ubuntu/.cursor/projects/workspace/uploads/____b1b1.xlsx")

DOCX_SRC = Path(
    "/home/ubuntu/.cursor/projects/workspace/uploads/"
    "______________________-_____3__9bd2.docx"
)
IMG_TMP = Path("/tmp/flower-eval")
OUT = Path("/workspace/deliverables/绿城·潮鸣外滩-开盘花束营销验收单.xlsx")
PHOTO_DIR = Path("/workspace/deliverables/开盘花束验收照片")
DOCX_OUT = Path(
    "/workspace/deliverables/附件-绿城中国策划活动类效果评估表-开盘花束.docx"
)
ZIP_OUT = Path("/workspace/deliverables/下载包/绿城潮鸣外滩-开盘花束营销验收单.zip")
THUMB = Path("/tmp/flower-thumbs")

FONT = "等线"
TITLE_FONT = "黑体"
THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")
LEFT_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")

GREEN = PatternFill("solid", fgColor="1F6B4A")
GREEN_LIGHT = PatternFill("solid", fgColor="E8F5E9")
GREEN_MID = PatternFill("solid", fgColor="C8E6C9")
AMBER = PatternFill("solid", fgColor="FFF8E1")
WHITE = PatternFill("solid", fgColor="FFFFFF")
HEADER_BAR = PatternFill("solid", fgColor="F5F5F5")
WHITE_FONT = Font(name=TITLE_FONT, size=16, bold=True, color="FFFFFF")

PHOTOS = [
    ("image1.jpeg", "案场大厅陈列：绿城中国标识 +「潮鸣」背景，花束装箱陈列"),
    ("image2.jpeg", "开盘现场氛围：礼仪接待、花束阵列与拍摄设备"),
    ("image3.jpeg", "花束送达：厢式货车卸货（纸箱内红粉花束）"),
    ("image4.jpeg", "花束清点转运：纸箱「鲜切花 轻拿轻放」，送达车辆 沪A·ZR373"),
]


def ensure_photos() -> None:
    """优先用已入库照片，其次 /tmp，再次从评估表 Word 解压。"""
    PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    missing = [name for name, _ in PHOTOS if not (PHOTO_DIR / name).exists()]
    if not missing:
        return
    for name, _ in PHOTOS:
        dest = PHOTO_DIR / name
        if dest.exists():
            continue
        tmp = IMG_TMP / name
        if tmp.exists():
            dest.write_bytes(tmp.read_bytes())
            continue
        if DOCX_SRC.exists():
            with zipfile.ZipFile(DOCX_SRC) as zf:
                dest.write_bytes(zf.read(f"word/media/{name}"))
        else:
            raise SystemExit(f"缺少现场照片 {name}")


def thumb(src: Path, dest: Path, max_w=720, max_h=540, quality=86) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    im = PILImage.open(src)
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    im.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
    dest = dest.with_suffix(".jpg")
    im.save(dest, "JPEG", quality=quality, optimize=True)
    return dest


def place_image(ws, path: Path, cell: str, width: int, height: int) -> None:
    img = XLImage(str(path))
    img.width = width
    img.height = height
    img.anchor = cell
    ws.add_image(img)


def fill_cell(ws, coord, text, size=10, bold=False, align=WRAP, fill=None, font_name=FONT):
    cell = ws[coord]
    cell.value = text
    cell.font = Font(name=font_name, size=size, bold=bold)
    cell.alignment = align
    cell.border = THIN
    if fill is not None:
        cell.fill = fill


def paint_range(ws, start_row, end_row, start_col, end_col, fill=WHITE):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            if fill is not None:
                cell.fill = fill


def unmerge_if(ws, coord: str) -> None:
    hits = [str(rng) for rng in list(ws.merged_cells.ranges) if coord in rng]
    for rng in hits:
        ws.unmerge_cells(rng)


def setup_print(ws, area: str, landscape: bool = False, fit_height: int = 1) -> None:
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    if ws.sheet_properties.pageSetUpPr is None:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = area
    ws.page_setup.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.2, footer=0.2
    )


def fill_cover(ws) -> None:
    """完善首页：事项/金额/公司/时间/描述/2×2 照片+图注/评估/签字/打印。"""
    ws.title = "营销验收单"

    # 略调 B/C/D 列宽，使照片区左右更接近 2×2（仍保留四列表头）
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 38

    fill_cell(
        ws,
        "B3",
        "绿城潮鳴开盘花束定制\n（潮鸣外滩开盘 / 潮鳴东方示范区开放）",
        11,
        True,
    )
    fill_cell(
        ws,
        "D3",
        "45 束鲜花\n金额：按双方约定\n（评估表未列明单价/总价）",
        10,
        True,
        fill=AMBER,
    )
    fill_cell(
        ws,
        "B4",
        "项目方：上海绿城泓盛建设发展有限公司\n品牌：绿城中国 · 绿城·潮鸣外滩\n地点：潮鳴东方项目示范区",
        10,
    )
    fill_cell(ws, "D4", "2026年3月\n示范区开放：3月26日", 11, True)
    fill_cell(
        ws,
        "B5",
        "【活动】绿城潮鳴开盘花束定制。\n"
        "【节点】潮鳴东方项目示范区开放（2026年3月26日），配套潮鸣外滩开盘。\n"
        "【交付】为潮鸣外滩开盘活动提供 45 束鲜花。\n"
        "【执行】厢式货车送达案场（车牌 沪A·ZR373），纸箱标注「鲜切花 轻拿轻放」；"
        "案场大厅在绿城中国标识及「潮鸣」背景前完成花束陈列。\n"
        "【结论】按照约定完成本次活动的整体策划与执行，活动取得圆满成功。"
        "（依据《绿城中国策划活动类效果评估表》原文，详见「效果评估表」页）",
        9,
        align=LEFT_TOP,
    )
    ws.row_dimensions[5].height = 118

    fill_cell(
        ws,
        "B13",
        "【策划评价】按照约定进行本次活动的整体策划与执行，活动取得圆满成功。\n"
        "【现场核验】花束按时送达并完成陈列；绿城中国标识、「潮鸣」背景露出清晰"
        "（见本页 2×2 照片及「开盘花束现场」页）。\n"
        "【营销评估】非常好（☐）　良好（☑）　一般（☐）　较差（☐）\n"
        "说明：评估表原件勾选栏为空白。本表与模板默认口径「效果良好，符合要求」对齐，勾选「良好」。"
        "请营销负责人复核后在下方签字。金额以双方约定/费用清单为准，本表不编造。",
        10,
        align=LEFT_TOP,
    )
    ws.row_dimensions[13].height = 48
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 28

    fill_cell(
        ws,
        "B16",
        "①本表首页嵌入评估表现场照片 4 张（含图注）；"
        "②「效果评估表」页按绿城中国策划活动类效果评估表结构复刻；"
        "③「交付清单」页列明 45 束、3月26日、送达车辆与品牌露出核验；"
        "④「开盘花束现场」页为大图；"
        "⑤附件：绿城中国策划活动类效果评估表-开盘花束（Word 原件）；"
        "⑥费用清单（经办人签字，金额按约定另附）。",
        9,
        align=LEFT,
    )
    ws.row_dimensions[16].height = 72
    fill_cell(
        ws,
        "C17",
        "签字：____________________\n日期：____________________\n（评估表原件签字栏为空，请补签）",
        10,
        align=LEFT,
    )
    fill_cell(
        ws,
        "C18",
        "签字：____________________\n日期：____________________\n营销评估已按「良好」预勾，请复核补签",
        10,
        align=LEFT,
    )

    # 照片区：解开 B6:D12，改成 2×2 + 图注，保留外框
    unmerge_if(ws, "B6")
    for rng in ("B6:C7", "D6:D7", "B8:C8", "B10:C11", "D10:D11", "B12:C12"):
        ws.merge_cells(rng)
    paint_range(ws, 6, 12, 2, 4, WHITE)

    fill_cell(ws, "B8", "① " + PHOTOS[0][1], 8, align=LEFT)
    fill_cell(ws, "D8", "② " + PHOTOS[1][1], 8, align=LEFT)
    fill_cell(ws, "B12", "③ " + PHOTOS[2][1], 8, align=LEFT)
    fill_cell(ws, "D12", "④ " + PHOTOS[3][1], 8, align=LEFT)

    ws.row_dimensions[6].height = 86
    ws.row_dimensions[7].height = 86
    ws.row_dimensions[8].height = 32
    ws.row_dimensions[9].height = 8
    ws.row_dimensions[10].height = 86
    ws.row_dimensions[11].height = 86
    ws.row_dimensions[12].height = 32

    anchors_sizes = [
        ("B6", 300, 210),
        ("D6", 250, 210),
        ("B10", 300, 210),
        ("D10", 250, 210),
    ]
    for (name, _cap), (cell, w, h) in zip(PHOTOS, anchors_sizes):
        t = thumb(PHOTO_DIR / name, THUMB / f"cover_{name}", max_w=640, max_h=480)
        place_image(ws, t, cell, w, h)

    ws["A20"] = "对应评估表"
    ws["A20"].font = Font(name=FONT, size=9, bold=True)
    ws["B20"] = (
        "绿城中国策划活动类效果评估表（冠名、活动赞助、活动执行等）· 开盘花束；"
        "本工作簿另含「效果评估表」「交付清单」「开盘花束现场」。"
    )
    ws["B20"].font = Font(name=FONT, size=9)
    ws["B20"].alignment = LEFT
    ws.merge_cells("B20:D20")
    ws.row_dimensions[20].height = 28

    setup_print(ws, "A2:D18", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = "1F6B4A"


def add_eval_sheet(wb) -> None:
    """按 Word《效果评估表》结构复刻一页，填入原文并嵌入四图。"""
    ws = wb.create_sheet("效果评估表", 1)
    for col, width in enumerate([16, 24, 24, 36], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.merge_cells("A1:D1")
    fill_cell(ws, "A1", "绿城中国策划活动类效果评估表", 16, True, WRAP, GREEN, TITLE_FONT)
    ws["A1"].font = WHITE_FONT
    ws.row_dimensions[1].height = 32
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].fill = GREEN
    ws["A1"].font = WHITE_FONT

    ws.merge_cells("A2:D2")
    fill_cell(
        ws,
        "A2",
        "（冠名、活动赞助、活动执行等）　项目：上海绿城 · 潮鸣外滩　活动：开盘花束定制",
        10,
        False,
        WRAP,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws.row_dimensions[2].height = 22

    fill_cell(ws, "A3", "活动主题", 11, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B3:D3")
    fill_cell(ws, "B3", "绿城潮鳴开盘花束定制", 13, True, WRAP, WHITE)
    paint_range(ws, 3, 3, 1, 4, WHITE)
    ws["A3"].fill = GREEN_LIGHT
    ws.row_dimensions[3].height = 28

    fill_cell(ws, "A4", "活动举办时间", 11, True, WRAP, GREEN_LIGHT)
    ws.merge_cells("B4:D4")
    fill_cell(
        ws,
        "B4",
        "2026年3月　　潮鳴东方项目示范区开放：3月26日",
        12,
        True,
        WRAP,
        WHITE,
    )
    paint_range(ws, 4, 4, 1, 4, WHITE)
    ws["A4"].fill = GREEN_LIGHT
    ws.row_dimensions[4].height = 28

    ws.merge_cells("A5:A12")
    fill_cell(ws, "A5", "活动\n现场氛围", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 5, 12, 1, 4, WHITE)
    ws["A5"].fill = GREEN_LIGHT
    ws.merge_cells("B5:C5")
    fill_cell(ws, "B5", "① " + PHOTOS[0][1], 8, align=LEFT)
    fill_cell(ws, "D5", "② " + PHOTOS[1][1], 8, align=LEFT)
    ws.merge_cells("B6:C8")
    ws.merge_cells("D6:D8")
    ws.merge_cells("B9:C9")
    fill_cell(ws, "B9", "③ " + PHOTOS[2][1], 8, align=LEFT)
    fill_cell(ws, "D9", "④ " + PHOTOS[3][1], 8, align=LEFT)
    ws.merge_cells("B10:C12")
    ws.merge_cells("D10:D12")

    for r, h in [(5, 28), (6, 78), (7, 78), (8, 78), (9, 28), (10, 78), (11, 78), (12, 78)]:
        ws.row_dimensions[r].height = h

    eval_anchors = [
        ("B6", 340, 280),
        ("D6", 250, 280),
        ("B10", 340, 280),
        ("D10", 250, 280),
    ]
    for (name, _cap), (cell, w, h) in zip(PHOTOS, eval_anchors):
        t = thumb(PHOTO_DIR / name, THUMB / f"eval_{name}", max_w=800, max_h=600)
        place_image(ws, t, cell, w, h)

    ws.merge_cells("A13:A16")
    fill_cell(ws, "A13", "策划负责人填写", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 13, 16, 1, 4, WHITE)
    ws["A13"].fill = GREEN_LIGHT
    ws.merge_cells("B13:D13")
    fill_cell(ws, "B13", "总体活动评价：", 11, True, LEFT, GREEN_MID)
    ws.merge_cells("B14:D15")
    fill_cell(
        ws,
        "B14",
        "本次活动为潮鳴东方项目示范区开放（3月26日）。"
        "为潮鸣外滩开盘活动提供了45束鲜花。"
        "按照约定进行本次活动的整体策划与执行，活动取得圆满成功。",
        12,
        False,
        LEFT_TOP,
        WHITE,
    )
    ws.merge_cells("B16:D16")
    fill_cell(
        ws,
        "B16",
        "策划负责人签名：____________________　　日期：____________________"
        "　　（评估表原件此栏为空，请补签）",
        10,
        False,
        LEFT,
        HEADER_BAR,
    )
    ws.row_dimensions[13].height = 24
    ws.row_dimensions[14].height = 36
    ws.row_dimensions[15].height = 36
    ws.row_dimensions[16].height = 32

    ws.merge_cells("A17:A21")
    fill_cell(ws, "A17", "营销负责人填写", 11, True, WRAP, GREEN_LIGHT)
    paint_range(ws, 17, 21, 1, 4, WHITE)
    ws["A17"].fill = GREEN_LIGHT
    ws.merge_cells("B17:D17")
    fill_cell(ws, "B17", "对活动【总体效果】评估（文字说明）：", 11, True, LEFT, GREEN_MID)
    ws.merge_cells("B18:D18")
    fill_cell(
        ws,
        "B18",
        "现场氛围良好，45 束鲜花按时送达并完成案场陈列，绿城中国 /「潮鸣」品牌露出清晰，"
        "与策划评价「圆满成功」一致。效果良好，符合要求。",
        11,
        False,
        LEFT,
        WHITE,
    )
    ws.merge_cells("B19:D19")
    fill_cell(
        ws,
        "B19",
        "非常好（☐）　　良好（☑）　　一般（☐）　　较差（☐）",
        13,
        True,
        WRAP,
        AMBER,
    )
    ws.merge_cells("B20:D20")
    fill_cell(
        ws,
        "B20",
        "勾选说明：评估表 Word 原件四档勾选栏均为空白；"
        "本表按《基础营销验收单》模板默认口径「效果良好，符合要求」勾选「良好」，"
        "不拔高为「非常好」。请营销负责人复核。",
        9,
        False,
        LEFT,
        AMBER,
    )
    ws.merge_cells("B21:D21")
    fill_cell(
        ws,
        "B21",
        "营销负责人签名：____________________　　日期：____________________"
        "　　（评估表原件此栏为空，请补签）",
        10,
        False,
        LEFT,
        HEADER_BAR,
    )
    ws.row_dimensions[17].height = 24
    ws.row_dimensions[18].height = 42
    ws.row_dimensions[19].height = 28
    ws.row_dimensions[20].height = 42
    ws.row_dimensions[21].height = 32

    ws.merge_cells("A22:D22")
    fill_cell(
        ws,
        "A22",
        "详见附件：绿城中国策划活动类效果评估表-开盘花束（Word 原件已随本工作簿打包）；"
        "费用清单由经办人签字后另附。合作金额评估表未列明，以双方约定为准。",
        9,
        False,
        LEFT,
        WHITE,
    )
    paint_range(ws, 22, 22, 1, 4, WHITE)
    ws.row_dimensions[22].height = 36

    setup_print(ws, "A1:D22", landscape=False, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "C8E6C9"


def add_checklist_sheet(wb) -> None:
    ws = wb.create_sheet("交付清单", 2)
    widths = [8, 16, 36, 28, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:E1")
    fill_cell(
        ws,
        "A1",
        "绿城潮鳴开盘花束定制 · 交付核验清单",
        16,
        True,
        WRAP,
        GREEN,
        TITLE_FONT,
    )
    paint_range(ws, 1, 1, 1, 5, GREEN)
    ws["A1"].font = WHITE_FONT
    ws["A1"].fill = GREEN
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    fill_cell(
        ws,
        "A2",
        "项目：绿城中国 · 绿城·潮鸣外滩（潮鳴东方项目示范区）　"
        "时间：2026年3月（示范区开放 3月26日）　交付：45 束鲜花　"
        "依据：绿城中国策划活动类效果评估表-开盘花束",
        9,
        False,
        LEFT,
        GREEN_MID,
    )
    paint_range(ws, 2, 2, 1, 5, GREEN_MID)
    ws.row_dimensions[2].height = 28

    headers = ["序号", "验收项", "约定 / 载明内容", "现场核验", "结论"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = GREEN
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[3].height = 22

    rows = [
        ("1", "活动主题", "绿城潮鳴开盘花束定制", "与评估表原文一致", "通过"),
        ("2", "举办时间", "2026年3月", "示范区开放日 3月26日已载明", "通过"),
        ("3", "项目 / 案场", "潮鸣外滩开盘；潮鳴东方项目示范区", "案场大厅绿城中国 +「潮鸣」背景（图①②）", "通过"),
        ("4", "交付数量", "45 束鲜花", "评估表载明 45 束；现场照片见装箱/卸货陈列", "通过"),
        ("5", "送达方式", "开盘现场送达、卸货、清点", "厢式货车卸货（图③）", "通过"),
        ("6", "包装要求", "鲜切花轻拿轻放", "纸箱印刷「鲜切花 轻拿轻放」（图④）", "通过"),
        ("7", "送达车辆", "现场转运车辆", "车牌 沪A·ZR373（图④）", "通过"),
        ("8", "现场陈列", "案场大厅花束阵列 / 礼仪接待氛围", "图①②可核验陈列与拍摄现场", "通过"),
        ("9", "品牌露出", "绿城中国标识、「潮鸣」背景字", "图①②案场背景清晰可辨", "通过"),
        (
            "10",
            "策划评价",
            "按约定完成整体策划与执行，圆满成功",
            "评估表策划栏原文已转录至「效果评估表」页",
            "通过",
        ),
        (
            "11",
            "营销评估",
            "非常好 / 良好 / 一般 / 较差",
            "原件勾选栏空白；本表按模板「效果良好，符合要求」预勾「良好」",
            "待补签",
        ),
        (
            "12",
            "合作金额",
            "评估表未列明单价或总价",
            "本表合作金额栏注明「按双方约定」，不编造数字；费用清单另附",
            "待清单",
        ),
        (
            "13",
            "负责人签字",
            "策划负责人、营销负责人签名",
            "评估表原件签字栏为空，验收单已留签字行",
            "待补签",
        ),
    ]
    for i, (idx, item, agreed, evidence, result) in enumerate(rows, 4):
        values = [idx, item, agreed, evidence, result]
        fills = {
            "通过": GREEN_LIGHT,
            "待补签": AMBER,
            "待清单": AMBER,
        }
        for col, val in enumerate(values, 1):
            cell = ws.cell(i, col, val)
            cell.font = Font(name=FONT, size=9, bold=(col == 5))
            cell.alignment = LEFT if col in (3, 4) else WRAP
            cell.border = THIN
            cell.fill = fills.get(result, WHITE) if col == 5 else WHITE
        ws.row_dimensions[i].height = 36

    last = 3 + len(rows)
    note_row = last + 1
    ws.merge_cells(f"A{note_row}:E{note_row}")
    fill_cell(
        ws,
        f"A{note_row}",
        "验收结论：开盘花束按评估表约定完成 45 束鲜花的送达、陈列与现场氛围布置，"
        "策划评价为圆满成功；营销评估预勾「良好」，待绿城营销负责人复核签字。"
        "金额、费用清单不以本表推定。本清单不与 2026-05-22 峰会晚宴冠名验收混用。",
        9,
        False,
        LEFT,
        GREEN_LIGHT,
    )
    paint_range(ws, note_row, note_row, 1, 5, GREEN_LIGHT)
    ws.row_dimensions[note_row].height = 48

    setup_print(ws, f"A1:E{note_row}", landscape=True, fit_height=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "FFF8E1"


def add_photo_sheet(wb) -> None:
    ws = wb.create_sheet("开盘花束现场")
    ws["A1"] = "绿城潮鳴开盘花束定制｜活动现场氛围（评估表附图大图）"
    ws["A1"].font = Font(name=TITLE_FONT, size=14, bold=True, color="FFFFFF")
    ws.merge_cells("A1:D1")
    paint_range(ws, 1, 1, 1, 4, GREEN)
    ws["A1"].fill = GREEN
    ws["A1"].alignment = WRAP
    ws.row_dimensions[1].height = 28

    ws["A2"] = (
        "活动主题：绿城潮鳴开盘花束定制　　举办时间：2026年3月（示范区开放 3月26日）　　"
        "交付：为潮鸣外滩开盘活动提供 45 束鲜花。按约定完成整体策划与执行，活动取得圆满成功。"
    )
    ws["A2"].font = Font(name=FONT, size=10)
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:D2")
    paint_range(ws, 2, 2, 1, 4, GREEN_MID)
    ws["A2"].fill = GREEN_MID
    ws.row_dimensions[2].height = 42
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 42

    row = 4
    for i, (name, cap) in enumerate(PHOTOS):
        col = "A" if i % 2 == 0 else "C"
        if i % 2 == 0:
            ws.row_dimensions[row].height = 188
            ws.row_dimensions[row + 1].height = 40
        t = thumb(PHOTO_DIR / name, THUMB / f"page_{name}", max_w=900, max_h=680)
        place_image(ws, t, f"{col}{row}", 310, 230)
        cap_cell = ws[f"{col}{row + 1}"]
        merge_end = "B" if col == "A" else "D"
        ws.merge_cells(f"{col}{row + 1}:{merge_end}{row + 1}")
        cap_cell.value = f"{i + 1}. {cap}"
        cap_cell.font = Font(name=FONT, size=10)
        cap_cell.alignment = LEFT
        cap_cell.border = THIN
        if i % 2 == 1:
            row += 2

    ws.merge_cells("A8:D8")
    fill_cell(
        ws,
        "A8",
        "策划负责人填写｜总体活动评价：本次活动为潮鳴东方项目示范区开放（3月26日）。"
        "为潮鸣外滩开盘活动提供了 45 束鲜花。按照约定进行本次活动的整体策划与执行，活动取得圆满成功。",
        10,
        False,
        LEFT,
        GREEN_LIGHT,
    )
    paint_range(ws, 8, 8, 1, 4, GREEN_LIGHT)
    ws.row_dimensions[8].height = 48
    ws.merge_cells("A9:D9")
    fill_cell(
        ws,
        "A9",
        "营销负责人填写｜总体效果评估：非常好（☐）  良好（☑）  一般（☐）  较差（☐）"
        "　　营销负责人签名：__________　　日期：__________",
        10,
        True,
        LEFT,
        AMBER,
    )
    paint_range(ws, 9, 9, 1, 4, AMBER)
    ws.row_dimensions[9].height = 32

    setup_print(ws, "A1:D9", landscape=True, fit_height=1)
    ws.print_title_rows = "1:2"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1F6B4A"


def copy_attachment() -> None:
    if DOCX_SRC.exists():
        shutil.copy2(DOCX_SRC, DOCX_OUT)


def make_zip() -> None:
    ZIP_OUT.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(ZIP_OUT, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(OUT, OUT.name)
        if DOCX_OUT.exists():
            zf.write(DOCX_OUT, DOCX_OUT.name)
        for p in sorted(PHOTO_DIR.glob("image*.jpeg")):
            zf.write(p, f"开盘花束验收照片/{p.name}")


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"缺少验收单模板 {SRC}")
    ensure_photos()
    THUMB.mkdir(parents=True, exist_ok=True)
    copy_attachment()

    wb = load_workbook(str(SRC))
    ws = wb.active
    fill_cover(ws)
    add_eval_sheet(wb)
    add_checklist_sheet(wb)
    add_photo_sheet(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT))
    make_zip()
    print(f"saved {OUT} size={OUT.stat().st_size}")
    print(f"zip {ZIP_OUT} size={ZIP_OUT.stat().st_size}")
    print(f"photos {PHOTO_DIR} n={len(list(PHOTO_DIR.glob('image*.jpeg')))}")
    if DOCX_OUT.exists():
        print(f"docx {DOCX_OUT} size={DOCX_OUT.stat().st_size}")


if __name__ == "__main__":
    main()
