#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成「复旦链接」功能对照简表 Excel（便于直接下载查看）"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "output" / "复旦链接-功能对照表.xlsx"

PRIMARY = "57068C"
TEAL = "8900E1"
LIGHT = "F0E6F8"
WHITE = "FFFFFF"
DARK = "1A0A2E"
GREY = "5C3D7A"
ALT = "E8D8F5"


def thin():
    s = Side(style="thin", color="D0D7DE")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(ws, row, n, fill=PRIMARY):
    for i in range(1, n + 1):
        c = ws.cell(row, i)
        c.font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[row].height = 28


def write_rows(ws, start, rows):
    for i, row in enumerate(rows):
        for j, v in enumerate(row, 1):
            c = ws.cell(start + i, j, v)
            c.font = Font(name="微软雅黑", size=10, color=DARK)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.fill = PatternFill("solid", fgColor=ALT if i % 2 else WHITE)
            c.border = thin()
        ws.row_dimensions[start + i].height = 36


def widths(ws, ws_list):
    for i, w in enumerate(ws_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, span, text, sub):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.font = Font(name="微软雅黑", size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c2 = ws.cell(2, 1, sub)
    c2.font = Font(name="微软雅黑", size=10, italic=True, color=GREY)
    c2.fill = PatternFill("solid", fgColor=LIGHT)
    c2.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30


def main():
    wb = Workbook()

    # 1 三源对标
    ws = wb.active
    ws.title = "01 三源对标"
    title(ws, 4, "复旦链接 · 三源对标总表", "复旦大学小程序 + 管院校友中心 + 互动吧 → 复旦链接（结构对标，品牌自有）")
    ws.append([])
    headers = ["对标来源", "借什么（主功能）", "复旦链接落地", "说明"]
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 4)
    write_rows(
        ws,
        4,
        [
            ["复旦大学小程序（eHall）", "服务宫格、办事/查询入口、个人中心", "首页八宫格 + 服务 Tab + 我的", "降低找功能成本"],
            ["管院校友中心小程序", "组织 / 活动 / 终身学习 / 职业发展 / 风采 / 支持 / 关于", "协作组织、活动、课程、机会、风采、支持平台", "主栏目骨架"],
            ["互动吧", "轻松办活动、海报、多票种、验票、裂变、现场互动、私域", "活动详情工具条 + 票夹 + 活动能力页", "活动运营全链路"],
        ],
    )
    widths(ws, [28, 42, 36, 18])
    ws.freeze_panes = "A4"

    # 2 Tab与页面
    ws2 = wb.create_sheet("02 Tab与页面")
    title(ws2, 4, "信息架构：Tab 与关键页面", "Tab：首页｜服务｜活动｜我的")
    for i, h in enumerate(["层级", "页面", "职责", "对标来源"], 1):
        ws2.cell(3, i, h)
    style_header(ws2, 3, 4, TEAL)
    write_rows(
        ws2,
        4,
        [
            ["Tab", "首页", "服务宫格 + 资讯 + 近期活动 + 终身学习", "eHall + 校友中心"],
            ["Tab", "服务", "栏目大厅：组织/学习/机会/风采/支持", "校友中心"],
            ["Tab", "活动", "列表筛选 + 互动吧能力提示", "校友活动 + 互动吧"],
            ["Tab", "我的", "资料、组织、票夹、会员、支持", "校友中心我的"],
            ["子页", "终身学习 / 课程详情", "公开课、在线、定制", "校友终身学习"],
            ["子页", "活动详情", "议程、多票种、海报、电子票、现场互动", "互动吧"],
            ["子页", "协作组织", "联络处 / 社群 / 同学会", "校友组织"],
            ["子页", "平台风采", "故事 / 喜讯 / 观点", "校友风采"],
            ["子页", "产业机会 / 领事检索", "AI·具身·展会；姓名检索无电话邮箱", "职业发展延伸"],
            ["子页", "活动能力", "互动吧十大优点说明", "互动吧"],
        ],
    )
    widths(ws2, [10, 28, 40, 20])
    ws2.freeze_panes = "A4"

    # 3 互动吧能力
    ws3 = wb.create_sheet("03 互动吧能力落地")
    title(ws3, 5, "互动吧特色优点 → 复旦链接落地", "对外不挂互动吧名义，品牌用复旦链接")
    for i, h in enumerate(["特色优点", "要点", "落地方式", "优先级", "价值"], 1):
        ws3.cell(3, i, h)
    style_header(ws3, 3, 5)
    write_rows(
        ws3,
        4,
        [
            ["轻松办活动", "模板化发专业活动页", "论坛/沙龙/产业日/出海/联谊模板", "P0", "降运营门槛"],
            ["海报/邀请函", "一键生成传播物料", "活动详情「生成海报」", "P1", "社交传播"],
            ["裂变传播", "分享与召集官", "复制路径 / 召集官码演示", "P1", "获客"],
            ["多票种售票", "免费/学生/标准/VIP", "活动详情选票 + 支付演示", "P0", "收费闭环"],
            ["验票签到", "电子票核销", "我的票夹 + 核销码", "P1", "到场效率"],
            ["渠道统计", "来源可复盘", "运营后台（规划）", "P1", "投放复盘"],
            ["到场提醒", "提高到场率", "订阅消息（规划）", "P1", "降爽约"],
            ["现场互动", "签到墙/抽奖/投票", "活动详情演示入口", "P2", "现场氛围"],
            ["私域沉淀", "获客→社群→复购", "报名引导 + 会员", "P1", "人脉资产"],
            ["交易保障", "退改/订单/发票", "规则公示 + 订单可查", "P1", "报名安心"],
        ],
    )
    widths(ws3, [16, 22, 32, 10, 14])
    ws3.freeze_panes = "A4"

    # 4 收费
    ws4 = wb.create_sheet("04 收费入口")
    title(ws4, 4, "收费入口一览", "演示环境为模拟支付")
    for i, h in enumerate(["收费类型", "典型场景", "价格示意", "触点"], 1):
        ws4.cell(3, i, h)
    style_header(ws4, 3, 4, TEAL)
    write_rows(
        ws4,
        4,
        [
            ["活动报名费", "论坛/产业日/联谊", "免费 / ¥68–¥1,280", "活动详情"],
            ["公开课学费/定金", "研修与实训营", "¥99–¥6,800+", "课程详情"],
            ["平台会员", "优先报名与资料", "¥365 / 年", "我的"],
            ["支持平台", "支持公共活动与展览", "任意金额", "我的"],
            ["增值对接包", "一对一对接席", "按档报价", "活动/机会"],
        ],
    )
    widths(ws4, [16, 24, 22, 14])
    ws4.freeze_panes = "A4"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"已生成: {OUT}")


if __name__ == "__main__":
    main()
