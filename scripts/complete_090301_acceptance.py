#!/usr/bin/env python3
"""完善验收单090301：插入专场参观三图，修正错位图注与页脚重复。"""

from __future__ import annotations

import shutil
import zipfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

ROOT = Path(__file__).resolve().parents[1]
SRC = Path("/home/ubuntu/.cursor/projects/workspace/uploads/___090301_cd86.xlsx")
PHOTO_SRC = [
    Path("/home/ubuntu/.cursor/projects/workspace/assets/f2575e3b-3010-4151-a4f2-78457883ba8d.jpg"),
    Path("/home/ubuntu/.cursor/projects/workspace/assets/7b484ba0-4d1e-4e54-8c7a-650107586371.jpg"),
    Path("/home/ubuntu/.cursor/projects/workspace/assets/b6959dff-d604-4a5b-83b4-84fd56ce2547.jpg"),
]
OUT = ROOT / "deliverables" / "绿城·潮鸣外滩-营销验收单-090301.xlsx"
PHOTO_DIR = ROOT / "deliverables" / "验收照片"
ZIP_OUT = ROOT / "deliverables" / "下载包" / "绿城潮鸣外滩-营销验收单-090301.zip"

ALBUM = "https://live.pailixiang.com/album/a12523836160"
VIDEO = "https://pan.baidu.com/s/1S8cNdqnCR_anuVPD6vlhjg"
VIDEO_CODE = "8888"

FONT = "等线"
CAPTION_FONT = "Microsoft YaHei"

# 新图为 16:9，与原表图 17–19 的显示尺寸一致
PHOTO_W, PHOTO_H = 1920, 1080
PHOTO_ROW_H, CAPTION_ROW_H = 254.25, 66.0

PHOTO_NAMES = [
    "专场参观-图27-案场展厅沙盘.jpg",
    "专场参观-图28-示范区户外讲解.jpg",
    "专场参观-图29-连廊SYMPHONY江景大宅.jpg",
]

NEW_CAPTIONS = [
    "图 27｜专场参观：绿城·潮鸣外滩案场展厅。嘉宾围观上海城市沙盘（墙面「源流百年」「航运始发港」「领馆聚集区」「世博会客厅」）。——兑现 ③专场项目参观邀约。",
    "图 28｜专场参观：SYMPHONY 潮鸣外滩示范区户外讲解。现场指引楼体与景观（墙面可见 SYMPHONY）。——兑现 ③专场项目参观邀约。",
    "图 29｜专场参观：连廊合影。楼体标识「SYMPHONY SHANGHAI / 中央公园江景大宅」，嘉宾佩戴峰会紫色挂绳。——兑现 ③专场项目参观邀约。",
]

# 按嵌入图片实际画面校正图 16–26（原稿图注与照片错位）
CAPTION_FIXES = {
    38: "图 16｜签到接待台：红桌布、峰会紫色挂绳与潮鸣物料，嘉宾核验入场。——佐证活动到场与接待执行。",
    40: "图 17｜专场参观：SYMPHONY 潮鸣外滩示范区户外讲解（楼体可见 SYMPHONY）。——兑现 ③专场项目参观邀约。",
    42: "图 18｜专场参观：示范区户外交流，嘉宾佩戴峰会紫色挂绳，楼体可见 SYMPHONY。——兑现 ③专场项目参观邀约。",
    44: "图 19｜专场参观：案场会所/休息区嘉宾围聚交流（手持咖啡，窗外江景画面）。——兑现 ③专场项目参观邀约。",
    46: "图 20｜嘉宾席「缪川」名牌特写（圆桌席位牌）。——佐证嘉宾规格与定向邀约落地。",
    48: "图 21｜签到台二维码展板，嘉宾扫码完成入场登记。——佐证活动到场规模与数字化会务。",
    50: "图 22｜签到处接待：红桌布、紫色挂绳与二维码台卡。——佐证活动到场规模。",
    52: "图 23｜签到处「外滩潮鸣」易拉宝，嘉宾扫码核验。——兑现 ②主会场/接待露出。",
    54: "图 24｜接待台「绿城中国 GREENTOWN」台签 +「SYMPHONY SHANGHAI 潮鸣」易拉宝。——兑现 ①②品牌露出。",
    56: "图 25｜接待区主视觉墙「SYMPHONY SHANGHAI | 潮鸣外滩」「中央公园江景大宅」，嘉宾佩戴峰会挂绳。——兑现 ②主会场露出。",
    58: "图 26｜签到走廊「外滩新百年 盛世正潮鸣」易拉宝与接待台。——兑现 ②主会场露出。",
}

B5_TEXT = (
    "【活动】第四届 / 2026人工智能商业化落地与硬核投资破局峰会。\n"
    "【身份】晚宴冠名战略合作伙伴（物料亦作「绿城·外滩潮鸣」「潮鳴」「SYMPHONY 潮鸣外滩」）。\n"
    "【晚宴】【绿城·潮鸣】星耀北外滩——AI领袖定制晚宴。\n"
    "【四模块】①晚宴冠名和专属权益 ¥55,000；②主会场权益 ¥30,000；"
    "③专场项目参观邀约 ¥5,000；④宣发配合 ¥10,000。含税合计 ¥100,000，整体打包验收。\n"
    "【现场已核验】主背景板「晚宴冠名战略合作伙伴·绿城中国」、茶歇「绿城中国」桌牌、"
    "晚宴大屏「外滩新百年 世界正潮鸣」、接待台「绿城中国 / SYMPHONY 潮鸣」。\n"
    "【专场参观已核验】图17–19、图27–29：案场展厅沙盘、SYMPHONY 示范区户外讲解、"
    "连廊「SYMPHONY SHANGHAI / 中央公园江景大宅」。\n"
    "【证书】事项载明颁发“2026年度智慧人居新质资产领军企业 暨卓越战略合作伙伴”；"
    "本表已附颁授画面（图 6–7）。本表 29 张；完整相册与录像见附件。"
    "营销预勾「良好 / 效果良好，符合要求」。本表不与 2026年3月开盘花束验收混用。"
)

EVAL_TEXT = (
    "【露出】主背景板、茶歇桌牌、晚宴大屏、颁授证书、接待台「绿城中国 / 潮鸣」均已核验，"
    "字样可辨（见图1–15、图23–26）。\n"
    "【专场参观】已核验。图17–19、图27–29 为 SYMPHONY 潮鸣外滩案场："
    "展厅沙盘（源流百年 / 航运始发港）、示范区户外讲解、连廊「中央公园江景大宅」。\n"
    "【营销评估】非常好（☐）　良好（☑）　一般（☐）　较差（☐）\n"
    "【打包验收】报价单四模块均已兑现。效果良好，符合约定要求。"
    "本表不与 2026年3月开盘花束验收混用。"
)

ATTACH_TEXT = (
    "①本表现场照片 29 张（图1–15 晚宴/主会场/证书；图16、21–26 签到接待与品牌物料；"
    "图17–19、图27–29 专场项目参观；图20 嘉宾席名牌）。"
    f"②拍立享相册 {ALBUM}（约 425 张）；"
    f"③全程录像 {VIDEO} 提取码 {VIDEO_CODE}。"
    "④报价单四模块：晚宴冠名、主会场、专场项目参观邀约、宣发配合。"
    "⑤费用清单（经办人签字）。本表为单页基础营销验收单，不与开盘花束验收混用。"
)


def copy_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def unmerge_from(ws, min_row: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= min_row:
            ws.unmerge_cells(str(rng))


def add_photo(ws, path: Path, row1: int, col0: int = 1, width: int = PHOTO_W, height: int = PHOTO_H) -> None:
    img = XLImage(str(path))
    img.width = width
    img.height = height
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=col0,
            colOff=pixels_to_EMU(6),
            row=row1 - 1,
            rowOff=pixels_to_EMU(4),
        ),
        ext=XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height)),
    )
    ws.add_image(img)


def main() -> None:
    PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    dest_photos: list[Path] = []
    for src, name in zip(PHOTO_SRC, PHOTO_NAMES):
        dest = PHOTO_DIR / name
        shutil.copy2(src, dest)
        dest_photos.append(dest)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, OUT)
    wb = load_workbook(OUT)
    ws = wb.worksheets[0]

    # 校正已有图注（插行前行号不变）
    for row, text in CAPTION_FIXES.items():
        cell = ws.cell(row, 2)
        cell.value = text
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    sig_exec = ws["C60"].value
    sig_mkt = ws["C61"].value
    label_exec = ws["B60"].value
    label_mkt = ws["B61"].value

    # 解除页脚与照片栏合并，避免插行后错位
    unmerge_from(ws, 59)
    if any(str(r) == "A6:A58" for r in ws.merged_cells.ranges):
        ws.unmerge_cells("A6:A58")

    ws.insert_rows(59, 6)

    sample_photo_row, sample_cap_row = 57, 58
    for i in range(3):
        photo_row = 59 + i * 2
        cap_row = photo_row + 1
        ws.row_dimensions[photo_row].height = PHOTO_ROW_H
        ws.row_dimensions[cap_row].height = CAPTION_ROW_H
        for col in range(1, 5):
            copy_style(ws.cell(sample_photo_row, col), ws.cell(photo_row, col))
            copy_style(ws.cell(sample_cap_row, col), ws.cell(cap_row, col))
            ws.cell(photo_row, col).value = None
            ws.cell(cap_row, col).value = None
        ws.merge_cells(start_row=photo_row, start_column=2, end_row=photo_row, end_column=4)
        ws.merge_cells(start_row=cap_row, start_column=2, end_row=cap_row, end_column=4)
        cap = ws.cell(cap_row, 2)
        cap.value = NEW_CAPTIONS[i]
        cap.font = Font(name=CAPTION_FONT, size=9, color="333333")
        cap.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        add_photo(ws, dest_photos[i], photo_row)

    ws.merge_cells("A6:A64")
    a6 = ws.cell(6, 1)
    a6.value = "验收照片"
    a6.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    a6.font = Font(name=FONT, size=12, bold=True)

    # 清空插行后下移的旧页脚（原 59–64 → 65–70），按空白模板结构重建
    for r in range(65, 78):
        for c in range(1, 5):
            cell = ws.cell(r, c)
            cell.value = None

    # 效果评估：A65:A67 + B65:D67（对应空白模板 A13:A15 / B13:D15）
    for r in range(65, 68):
        for c in range(1, 5):
            copy_style(ws.cell(3, 1) if c == 1 else ws.cell(5, 2), ws.cell(r, c))
    ws.merge_cells("A65:A67")
    ws.merge_cells("B65:D67")
    ws.row_dimensions[65].height = 42
    ws.row_dimensions[66].height = 42
    ws.row_dimensions[67].height = 42
    ws.cell(65, 1).value = "效果评估"
    ws.cell(65, 1).font = Font(name=FONT, size=12, bold=True)
    ws.cell(65, 1).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.cell(65, 2).value = EVAL_TEXT
    ws.cell(65, 2).font = Font(name=FONT, size=9)
    ws.cell(65, 2).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # 附件
    for c in range(1, 5):
        copy_style(ws.cell(3, 1) if c == 1 else ws.cell(5, 2), ws.cell(68, c))
    ws.merge_cells("B68:D68")
    ws.row_dimensions[68].height = 72
    ws.cell(68, 1).value = "附件"
    ws.cell(68, 1).font = Font(name=FONT, size=12, bold=True)
    ws.cell(68, 1).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.cell(68, 2).value = ATTACH_TEXT
    ws.cell(68, 2).font = Font(name=FONT, size=8)
    ws.cell(68, 2).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # 验收确认（只保留一处签字，去掉效果评估下的重复签字）
    for r in (69, 70):
        for c in range(1, 5):
            copy_style(ws.cell(4, 1) if c == 1 else ws.cell(4, 2), ws.cell(r, c))
    ws.merge_cells("A69:A70")
    ws.merge_cells("C69:D69")
    ws.merge_cells("C70:D70")
    ws.row_dimensions[69].height = 48
    ws.row_dimensions[70].height = 48
    ws.cell(69, 1).value = "验收确认"
    ws.cell(69, 1).font = Font(name=FONT, size=12, bold=True)
    ws.cell(69, 1).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.cell(69, 2).value = label_exec or "执行人/策划负责人"
    ws.cell(69, 2).font = Font(name=FONT, size=11)
    ws.cell(69, 2).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.cell(69, 3).value = sig_exec
    ws.cell(69, 3).font = Font(name=FONT, size=10)
    ws.cell(69, 3).alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    ws.cell(70, 2).value = label_mkt or "营销负责人"
    ws.cell(70, 2).font = Font(name=FONT, size=11)
    ws.cell(70, 2).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.cell(70, 3).value = sig_mkt
    ws.cell(70, 3).font = Font(name=FONT, size=10)
    ws.cell(70, 3).alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    ws["B5"] = B5_TEXT
    ws["B5"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[5].height = 118

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.45, right=0.45, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.print_area = "A1:D70"
    ws.sheet_view.showGridLines = False
    ws.oddFooter.center.text = "绿城·潮鸣外滩 营销验收单（29 张现场照片）  第 &P 页 / 共 &N 页"

    wb.save(OUT)

    ZIP_OUT.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(ZIP_OUT, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(OUT, OUT.name)
        for p in dest_photos:
            zf.write(p, f"验收照片/{p.name}")

    print(f"saved {OUT}")
    print(f"zip {ZIP_OUT} ({ZIP_OUT.stat().st_size} bytes)")
    print(f"images {len(ws._images)}")


if __name__ == "__main__":
    main()
