#!/usr/bin/env python3
"""生成伤情、费用、待办与口径工作簿。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from content import (
    ACCIDENT,
    ASR_CORRECTIONS,
    BOSS,
    BOSS_ITEMS,
    BOSS_NOT,
    GAPS,
    CALLS,
    CLOUD_FILM,
    CT_REPORTS,
    DISABILITY,
    HOSPITAL_TRACK,
    ACCIDENT_FOUR,
    TRANSFER_THREE,
    NOW_EIGHT,
    INJURY_NOT_THIS_ACCIDENT,
    INJURY_THIS_ACCIDENT,
    LEGAL_REMINDER,
    QINGXIAN,
    STAGE_FORBID,
    STAGE_PRINCIPLE,
    STAGE_STEPS,
    TALKING_POINTS,
)

NAVY = "0B2F5B"
GOLD = "C4A35A"
LIGHT = "F3F6FA"
RED = "A63D2F"
GREEN = "2F6B4F"
WHITE = "FFFFFF"
THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def header_font():
    return Font(name="微软雅黑", size=11, bold=True, color=WHITE)


def title_font():
    return Font(name="微软雅黑", size=14, bold=True, color=NAVY)


def body_font(bold=False, color="1F2A37"):
    return Font(name="微软雅黑", size=10, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def wrap():
    return Alignment(wrap_text=True, vertical="center")


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = header_font()
        cell.fill = fill(NAVY)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def style_rows(ws, start, end, cols):
    for r in range(start, end + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.font = body_font()
            cell.alignment = wrap()
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = fill(LIGHT)


def widths(ws, values):
    for i, w in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze(ws):
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "1:2"


def write_title(ws, text, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(1, 1, text)
    cell.font = title_font()
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22


def build_workbook(output_path: Path) -> None:
    wb = Workbook()

    # 0 办理顺序
    ws = wb.active
    ws.title = "办理顺序"
    write_title(ws, "现阶段不谈总赔偿数额。先固定监控和用工证据，取得书面认定，确认保险与医院资格，核清费用；鉴定后再算。", 7)
    headers = ["序号", "缓急", "时限", "事项", "谁来办", "怎么办理", "完成标志"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 7)
    urgency_color = {
        "贯穿禁止": RED,
        "紧急": RED,
        "急": "B57A2A",
        "中": NAVY,
        "后": GREEN,
    }
    for i, s in enumerate(STAGE_STEPS, 3):
        ws.cell(i, 1, s["id"])
        ws.cell(i, 2, s["urgency"])
        ws.cell(i, 3, s["when"])
        ws.cell(i, 4, s["name"])
        ws.cell(i, 5, s["who"])
        ws.cell(i, 6, s["how"])
        ws.cell(i, 7, s["done"])
        ws.row_dimensions[i].height = 72
    last_step = 2 + len(STAGE_STEPS)
    style_rows(ws, 3, last_step, 7)
    for i, s in enumerate(STAGE_STEPS, 3):
        ws.cell(i, 2).font = body_font(True, urgency_color.get(s["urgency"], NAVY))
    r = last_step + 2
    ws.cell(r, 1, "贯穿禁止")
    ws.cell(r, 1).font = body_font(True, RED)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    for item in STAGE_FORBID:
        r += 1
        ws.cell(r, 1, item)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 28
        ws.cell(r, 1).font = body_font(True, RED)
    r += 2
    ws.cell(r, 1, STAGE_PRINCIPLE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 48
    ws.cell(r, 1).font = body_font(True, RED)
    widths(ws, [10, 12, 22, 22, 28, 55, 36])
    freeze(ws)

    # 0b 材料核心
    ws = wb.create_sheet("材料核心", 1)
    write_title(ws, "17 份转写去重：转院只三点，事故四条，立即八事。费用按 4 万元逐笔核。", 4)
    headers = ["块", "序号", "事项", "怎么做 / 口径"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 4)
    r = 3
    for t in TRANSFER_THREE:
        ws.cell(r, 1, "转院三点")
        ws.cell(r, 2, t["n"])
        ws.cell(r, 3, t["what"])
        ws.cell(r, 4, t["how"] + " 完成：" + t["done"])
        ws.row_dimensions[r].height = 56
        r += 1
    for a in ACCIDENT_FOUR:
        ws.cell(r, 1, "事故四条")
        ws.cell(r, 2, a["title"])
        ws.cell(r, 3, a["text"])
        ws.cell(r, 4, a.get("note") or "")
        ws.row_dimensions[r].height = 56
        r += 1
    for x in NOW_EIGHT:
        ws.cell(r, 1, "立即八事")
        ws.cell(r, 2, x["n"])
        ws.cell(r, 3, x["what"])
        ws.cell(r, 4, x["who"] + "。" + x["why"])
        ws.row_dimensions[r].height = 56
        r += 1
    style_rows(ws, 3, r - 1, 4)
    widths(ws, [12, 16, 48, 55])
    freeze(ws)

    # 1 伤情鉴定
    ws = wb.create_sheet("伤情鉴定")
    write_title(ws, "齐鲁医院云影像 · 谁受伤、伤了什么（2026-08-17 核对）", 6)
    headers = ["日期", "检查号", "项目 / 科室", "影像所见（报告原文）", "影像诊断（报告原文）", "本次事故？"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 6)
    for i, r in enumerate(CT_REPORTS, 3):
        ws.cell(i, 1, r["date"])
        ws.cell(i, 2, r["no"])
        ws.cell(i, 3, f"{r['item']}\n{r['place']} · {r['status']}")
        ws.cell(i, 4, r["findings"])
        ws.cell(i, 5, "\n".join(r["diagnosis"]))
        ws.cell(i, 6, "是，计入本次" if r["accident_related"] else "否，退变/陈旧，不计入评残")
        if not r["accident_related"]:
            ws.cell(i, 6).font = body_font(True, GREEN)
        else:
            ws.cell(i, 6).font = body_font(True, RED)
    style_rows(ws, 3, 5, 6)
    ws.cell(7, 1, f"患者：{CLOUD_FILM['patient_display']} / {CLOUD_FILM['sex']} / {CLOUD_FILM['age']}岁　　医院：{CLOUD_FILM['hospital']}")
    ws.merge_cells("A7:F7")
    ws.cell(7, 1).font = body_font(True)
    ws.cell(8, 1, CLOUD_FILM["disclaimer"] + " 这是影像诊断，不是伤残鉴定。")
    ws.merge_cells("A8:F8")
    widths(ws, [14, 16, 28, 42, 36, 22])
    ws.row_dimensions[3].height = 72
    ws.row_dimensions[4].height = 56
    ws.row_dimensions[5].height = 48
    freeze(ws)

    # 2 伤残情况
    ws = wb.create_sheet("伤残情况")
    write_title(ws, "伤残：现在不能定级。只准备材料。", 3)
    headers = ["问题", "结论", "家属立刻怎么做"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 3)
    rows = [
        ["谁可能评残", DISABILITY["who"], "只为伤者本人建病历档案，不要给别人做伤残材料"],
        ["现在有没有残级", DISABILITY["now"], "对外禁止说「已经几级」"],
        ["什么时候评", DISABILITY["when"], "先治病，3 个月窗口前不要催鉴定"],
        ["用哪套标准", DISABILITY["standard"], "交通事故用人体损伤致残程度分级，不要套工伤"],
        ["评哪些伤", DISABILITY["what_counts"], "踝关节功能+行走；右足骨刺删除"],
        ["现在能否报价", DISABILITY["cannot_grade"], "面谈不谈残级、不签总包死价"],
    ]
    for i, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
    style_rows(ws, 3, 8, 3)
    ws.cell(10, 1, "评残前材料清单")
    ws.merge_cells("A10:C10")
    ws.cell(10, 1).font = body_font(True, NAVY)
    for i, item in enumerate(DISABILITY["prep"], 11):
        ws.cell(i, 1, i - 10)
        ws.cell(i, 2, item)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    style_rows(ws, 11, 15, 3)
    widths(ws, [18, 70, 36])
    for r in range(3, 9):
        ws.row_dimensions[r].height = 48
    freeze(ws)

    # 3 外伤 vs 退变
    ws = wb.create_sheet("外伤与退变对照")
    write_title(ws, "通话口述 vs 医院 CT：必须改口的清单", 3)
    headers = ["来源", "内容", "处理"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 3)
    r = 3
    ws.cell(r, 1, "CT：本次外伤")
    ws.cell(r, 2, "\n".join(f"• {x}" for x in INJURY_THIS_ACCIDENT))
    ws.cell(r, 3, "对外、对保险、对鉴定都用这套")
    ws.row_dimensions[r].height = 110
    r = 4
    ws.cell(r, 1, "CT：非本次外伤")
    ws.cell(r, 2, "\n".join(f"• {x}" for x in INJURY_NOT_THIS_ACCIDENT))
    ws.cell(r, 3, "不要写成跟骨粉碎骨折")
    ws.row_dimensions[r].height = 70
    r = 5
    for i, (wrong, right) in enumerate(ASR_CORRECTIONS, 5):
        ws.cell(i, 1, "通话须改口")
        ws.cell(i, 2, wrong)
        ws.cell(i, 3, right)
        ws.row_dimensions[i].height = 40
    style_rows(ws, 3, 4 + len(ASR_CORRECTIONS), 3)
    widths(ws, [18, 55, 55])
    freeze(ws)

    # 4 转院报销
    ws = wb.create_sheet("转院报销")
    write_title(ws, "转院录音大量重复。核心只三点：原医院结算、新医院入院、承保公司书面确认二甲。", 3)
    headers = ["序号", "事项", "怎么做 / 完成标志"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 3)
    rows = [
        [t["n"], t["what"], t["how"] + " 完成：" + t["done"]]
        for t in TRANSFER_THREE
    ]
    rows += [
        ["补", "结算铁律", HOSPITAL_TRACK["settle_rule"]],
        ["补", "找哪家保险", HOSPITAL_TRACK["no_motor_insurer"]],
        ["补", "家属口径", HOSPITAL_TRACK["family_rule"]],
        ["补", "不是康复期", HOSPITAL_TRACK["not_rehab"]],
        ["禁", "不要另找三甲", HOSPITAL_TRACK["other_sanjia"]],
    ]
    for i, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
        ws.row_dimensions[i].height = 52
    last = 2 + len(rows)
    style_rows(ws, 3, last, 3)
    widths(ws, [8, 28, 70])
    freeze(ws)

    # 5 费用台账
    ws = wb.create_sheet("费用台账")
    write_title(ws, "用发票分清：用工方垫付现金 3 万，费用口径确认 4 万。对外只报对上票的数。", 8)
    headers = ["日期", "项目", "金额（元）", "已付/未付", "付款人", "有无发票", "票据号", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 8)
    presets = [
        ["2026-08-14", "急诊右小腿CT（10008056847）", "", "已付待核", "家属", "待贴", "", "齐鲁医院"],
        ["2026-08-14", "急诊右足CT（10008056848）", "", "已付待核", "家属", "待贴", "", "齐鲁医院"],
        ["2026-08-15", "住院右踝CT术后复查（10008059257）", "", "已付待核", "家属", "待贴", "", "手足与显微重建外科"],
        ["2026-08-15", "右踝骨折内外固定手术", "", "已付待核", "家属", "待贴", "", "以手术记录为准"],
        ["", "住院押金/预缴", "", "已付待核", "家属", "待贴", "", "费用口径确认 4 万，须对账单"],
        ["", "药品（院内）", "", "已付待核", "家属", "待贴", "", "医保范围内用药，保险可赔、比例待定"],
        ["", "护工费", "", "已付待核", "家属", "待贴", "", "24小时一对一，合同+发票；按责比例报"],
        ["", "护理耗材（垫、便盆等）", "", "已付待核", "家属", "待贴", "", ""],
        ["2026-08-17", "转院救护车", 160, "已付待核", "家属", "待贴", "", "预估150，实收160，必须发票"],
        ["", "用工方垫付现金", 30000, "已付待写收据", "乔刘记商贸/刘孝春", "收据", "", "写明垫付不是借款、不是了结、不冲抵欠薪"],
        ["", "未结工钱", 5000, "未付", "刘孝春", "欠条", "", "另写一张，不乘 30%，不和 3 万垫付混"],
        ["", "责任方已付（美团侧）", 0, "待拉群", "骑手/平台险", "无", "", "保单和限额待骑手拉群"],
    ]
    for i, row in enumerate(presets, 3):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
        ws.cell(i, 3).number_format = "#,##0.00"
    style_rows(ws, 3, 14, 8)
    ws.cell(16, 1, "合计（仅填写了数字的行）")
    ws.cell(16, 3, "=SUM(C3:C14)")
    ws.cell(16, 3).number_format = "#,##0.00"
    ws.cell(16, 1).font = body_font(True)
    ws.cell(16, 3).font = body_font(True, RED)
    dv = DataValidation(type="list", formula1='"已付,未付,已付待核,未发生或待核"', allow_blank=True)
    dv.add("D3:D30")
    ws.add_data_validation(dv)
    dv2 = DataValidation(type="list", formula1='"有,无,待贴,待核"', allow_blank=True)
    dv2.add("F3:F30")
    ws.add_data_validation(dv2)
    widths(ws, [14, 34, 14, 16, 14, 12, 14, 36])
    freeze(ws)
    ws.auto_filter.ref = "A2:H30"

    # 5 待办
    ws = wb.create_sheet("72小时待办")
    write_title(ws, "立即八事。转院录音只看三点。费用按 4 万元逐笔核。不谈总赔偿数额。", 5)
    headers = ["序号", "谁", "做什么", "为什么", "状态"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 5)
    for i, a in enumerate(NOW_EIGHT, 3):
        ws.cell(i, 1, a["n"])
        ws.cell(i, 2, a["who"])
        ws.cell(i, 3, a["what"])
        ws.cell(i, 4, a["why"])
        ws.cell(i, 5, "未完成")
        ws.row_dimensions[i].height = 52
    style_rows(ws, 3, 10, 5)
    dv = DataValidation(type="list", formula1='"未完成,进行中,已完成,律师跟进"', allow_blank=False)
    dv.add("E3:E20")
    ws.add_data_validation(dv)
    widths(ws, [8, 22, 58, 40, 14])
    freeze(ws)

    # 6 口径
    ws = wb.create_sheet("沟通口径卡")
    write_title(ws, "面谈只说这些。全责、残级、跟骨粉碎骨折，全部停用。", 2)
    headers = ["场合", "原话/要点"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 2)
    r = 3
    mapping = [
        ("对交警", TALKING_POINTS["to_police"]),
        ("对骑手/平台", TALKING_POINTS["to_rider"]),
        ("对医院", TALKING_POINTS["to_hospital"]),
        ("对保险公司", TALKING_POINTS["to_insurer"]),
        ("绝对不要说", TALKING_POINTS["do_not_say"]),
    ]
    for title, items in mapping:
        for item in items:
            ws.cell(r, 1, title)
            ws.cell(r, 2, item)
            if title == "绝对不要说":
                ws.cell(r, 2).font = body_font(True, RED)
            r += 1
    style_rows(ws, 3, r - 1, 2)
    widths(ws, [18, 90])
    freeze(ws)

    # 7 责任路径
    ws = wb.create_sheet("责任与程序")
    write_title(ws, "认定书尚未出具。简易或一般由律师看监控后定。", 3)
    headers = ["事项", "交警口径", "家属对策"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 3)
    rows = [
        ["地点", ACCIDENT["place"], "对外用抚顺路和哈尔滨路路口，以认定书为准"],
        ["过程", ACCIDENT["process"], "以监控为准，不补充猜测"],
        ["三轮过错", "；".join(ACCIDENT["tri_faults"]), "承认监控里看得到的问题，不争口头全责"],
        ["骑手过错", "；".join(ACCIDENT["rider_faults"]), "要求主责，保底同等；待认定书"],
        ["责任范围", ACCIDENT["police_range"], "认定书未出；停止使用「对方全责」"],
        ["简易程序", "不审资质，快出认定书", "可以谈，是否走由律师定"],
        ["一般程序", "无青岛市号牌仍有风险；驾驶人有证，无证风险下降", "律师看完监控再选，家属不拍板"],
        ["驾驶资格", "家属确认具备驾驶资格证书", "内部留证，不要主动交给一般程序扩查"],
        ["交警赔钱", "只出认定书，不参与金额", "金额找保险和法院，别逼交警"],
        ["刘孝春身份", BOSS["who"], "乔刘记商贸；雇佣无合同；工伤排除"],
        ["不替美团补", BOSS["not_meituan"], "骑手责任范围内的缺口仍找平台险和骑手"],
        ["可主张己方份", BOSS["yes_own_side"], "无牌+灯坏交人使用；3万垫付先写收据"],
    ]
    for i, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)
        ws.row_dimensions[i].height = 48
    style_rows(ws, 3, 14, 3)
    widths(ws, [16, 55, 40])
    freeze(ws)

    # 7b 刘孝春分担
    ws = wb.create_sheet("刘孝春分担清单")
    write_title(ws, "30%是己方缺口不是刘的固定比例。未结工钱100%由刘付，其余进总损失后再切。", 5)
    headers = ["项目", "现在能否主张", "计算口径", "刘孝春怎么担", "不要搞错"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 5)
    for i, x in enumerate(BOSS_ITEMS, 3):
        ws.cell(i, 1, x["name"])
        ws.cell(i, 2, x["now"])
        ws.cell(i, 3, x["how"])
        ws.cell(i, 4, x["liu"])
        ws.cell(i, 5, x["no"])
        ws.row_dimensions[i].height = 56
    last_item = 2 + len(BOSS_ITEMS)
    style_rows(ws, 3, last_item, 5)
    r = last_item + 2
    ws.cell(r, 1, "不应承担")
    ws.cell(r, 1).font = body_font(True, RED)
    r += 1
    for item in BOSS_NOT:
        ws.cell(r, 1, item)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 28
        r += 1
    ws.cell(r + 1, 1, BOSS["formula"])
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    ws.row_dimensions[r + 1].height = 48
    widths(ws, [22, 22, 42, 42, 36])
    freeze(ws)

    # 7b2 法律关系提醒
    ws = wb.create_sheet("法律关系提醒")
    write_title(ws, "不必等法院先确认雇佣关系。垫付和欠薪必须写清性质。", 2)
    headers = ["序号", "提醒"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 2)
    for i, item in enumerate(LEGAL_REMINDER, 3):
        ws.cell(i, 1, i - 2)
        ws.cell(i, 2, item)
        ws.row_dimensions[i].height = 36
    style_rows(ws, 3, 2 + len(LEGAL_REMINDER), 2)
    widths(ws, [8, 100])
    freeze(ws)

    # 7c 待补充信息
    ws = wb.create_sheet("待补充信息")
    write_title(ws, "8/19 补充后仍缺的事实。没有就不要编进 3D 或对外口径。3D 视频留作待用。", 5)
    headers = ["类别", "已经掌握", "还缺什么", "找谁要", "缓急"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 5)
    for i, g in enumerate(GAPS, 3):
        ws.cell(i, 1, g["cat"])
        ws.cell(i, 2, g["have"])
        ws.cell(i, 3, g["need"])
        ws.cell(i, 4, g["who"])
        ws.cell(i, 5, g["u"])
        ws.row_dimensions[i].height = 48
    style_rows(ws, 3, 2 + len(GAPS), 5)
    widths(ws, [16, 42, 48, 28, 14])
    freeze(ws)

    # 8 录音索引
    ws = wb.create_sheet("录音来源")
    write_title(ws, f"得到大脑转写 {len(CALLS)} 条索引。与 CT 冲突的以 CT 为准。录音≠认定书。3D 视频留作待用。", 4)
    headers = ["日期", "标题", "时长", "本表如何使用"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 4)
    for i, c in enumerate(CALLS, 3):
        ws.cell(i, 1, c["date"])
        ws.cell(i, 2, c["title"])
        ws.cell(i, 3, c["mins"])
        ws.cell(i, 4, c["use"])
    style_rows(ws, 3, 2 + len(CALLS), 4)
    widths(ws, [14, 48, 10, 55])
    freeze(ws)

    # 9 轻弦附录
    ws = wb.create_sheet("附录轻弦洽谈")
    write_title(ws, QINGXIAN["title"], 2)
    headers = ["序号", "要点"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    style_header(ws, 2, 2)
    for i, p in enumerate(QINGXIAN["points"], 3):
        ws.cell(i, 1, i - 2)
        ws.cell(i, 2, p)
    style_rows(ws, 3, 6, 2)
    widths(ws, [8, 90])
    freeze(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    build_workbook(root / "deliverables" / "青岛抚顺路和哈尔滨路路口交通事故_伤情伤残与行动表_20260817.xlsx")
