"""绿城华东区 · 2026下半年 AI 商业化落地合作方案 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path("/workspace/deliverables/绿城华东区-2026下半年AI商业化落地合作方案.xlsx")

GREEN = "006B3F"
GOLD = "C8A25B"
DARK = "1F1F1F"
GREY = "666666"
LIGHT = "EEF4F0"
HEADER_FILL = PatternFill("solid", fgColor=GREEN)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
WHITE_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color=GREEN)
HEAD_FONT = Font(name="微软雅黑", size=11, bold=True, color=DARK)
BODY_FONT = Font(name="微软雅黑", size=10, color=DARK)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True, color=DARK)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN


def style_body(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(r, c)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if r % 2 == 0:
                cell.fill = LIGHT_FILL


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_overview(wb):
    ws = wb.create_sheet("01-方案总览", 0)
    ws["A1"] = "绿城华东区 × 人工智能商业化落地｜2026下半年片区战略合作方案"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    overview = [
        ["项目", "内容"],
        ["方案定位", "服务华东区总经理的片区经营工具：客户链接 + 行政链接 + 交易链接"],
        ["地理焦点", "黄浦、虹口、青浦；主场枢纽为北外滩一滴水"],
        ["活动形态", "办行（闭门）、办大众（公开峰会）、大厂/TMT参访、项目联动日"],
        ["主推赞助档", "铂金·片区首席合作伙伴 ¥300,000（可升级钻石 ¥500,000）"],
        ["对区域总价值", "形成三区共识叙事，沉淀圈层客群与政商互动接口，服务去化与产业落位"],
        ["交付物", "本Excel（档位/权益/日历/联动/ROI）+ 配套PPT洽谈稿"],
        ["协同方", "上海市杨浦区科技企业联合会 / 见微知海新质商业生态"],
        ["对接人", "胡继刚  13262607888"],
        ["用途", "明日华东区总经理一对一沟通，内部洽谈稿，金额与场次可谈"],
    ]
    for r, row in enumerate(overview, 3):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_header(ws, 3, 2)
    style_body(ws, 4, 12, 2)
    ws["A4"].font = BOLD_FONT
    set_widths(ws, [18, 90])
    for r in range(4, 13):
        ws.row_dimensions[r].height = 32


def sheet_tiers(wb):
    ws = wb.create_sheet("02-赞助档位与金额", 1)
    ws["A1"] = "赞助档位建议（金额含税口径可按合同再议；下表为洽谈建议价）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    headers = ["档位", "建议金额(元)", "覆盖场次概览", "品牌高度", "片区覆盖", "适合谁拍板", "备注"]
    rows = [
        ["钻石·华东区年度战略合作伙伴", 500000,
         "旗舰1–2场+沙龙2–3场+参访2场+三区项目日3场", "最高：年度战略伙伴",
         "黄浦+虹口+青浦全覆盖", "华东区总经理", "主品牌叙事最完整，建议年末复盘升级续约"],
        ["铂金·片区首席合作伙伴（主推）", 300000,
         "旗舰1场+沙龙2场+参访1场+项目日2场", "高：片区首席",
         "主场虹口+另1–2区轮动", "华东区总经理", "性价比最高，明日优先锁定"],
        ["黄金·一滴水系列冠名", 150000,
         "一滴水旗舰峰会/晚宴系列", "中高：系列冠名",
         "虹口北外滩为主", "区域办/品牌", "适合先做高光再追加"],
        ["银牌·项目专场赞助", 80000,
         "单片区项目联动日1场+配套传播", "中：项目专场",
         "单区", "项目总+区域办", "可多项目分别认购"],
        ["单场闭门沙龙/晚宴", 50000,
         "单场冠名（人数精、转化强）", "灵活",
         "按场次", "灵活", "试水档，可累计升级"],
        ["单场闭门沙龙/晚宴（高配）", 100000,
         "单场高规格晚宴冠名+参观", "中高",
         "按场次", "灵活", "对标上半年潮鸣外滩模式"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, len(headers))
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if c == 2:
                cell.number_format = '¥#,##0'
    style_body(ws, 4, 9, len(headers))
    # highlight recommended
    for c in range(1, 8):
        ws.cell(5, c).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(5, c).font = BOLD_FONT

    ws["A11"] = "付款建议"
    ws["A11"].font = HEAD_FONT
    ws["A12"] = "签约后10个工作日内支付50%启动；首场旗舰活动前7日支付剩余50%。发票：会议服务费/会务咨询费/赞助费（以合同为准）。"
    ws["A12"].font = BODY_FONT
    ws["A12"].alignment = WRAP
    ws.merge_cells("A12:G12")
    ws.row_dimensions[12].height = 40

    ws["A14"] = "预算分摊示意（铂金30万）"
    ws["A14"].font = HEAD_FONT
    share_headers = ["分摊口径", "金额(元)", "说明"]
    share_rows = [
        ["华东区品牌统筹", 120000, "母品牌与片区共识、行政接口"],
        ["虹口主场项目", 100000, "一滴水主场与夜访承接"],
        ["黄浦/青浦轮动项目", 80000, "两场项目联动日分摊"],
    ]
    for c, h in enumerate(share_headers, 1):
        ws.cell(15, c, h)
    style_header(ws, 15, 3)
    for r, row in enumerate(share_rows, 16):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if c == 2:
                cell.number_format = '¥#,##0'
    style_body(ws, 16, 18, 3)
    set_widths(ws, [32, 14, 42, 22, 22, 18, 36])
    for r in range(4, 10):
        ws.row_dimensions[r].height = 48


def sheet_rights(wb):
    ws = wb.create_sheet("03-权益矩阵", 2)
    ws["A1"] = "权益矩阵（●标配 ○可选/弱化 —不含）详细勾选可用于合同附表"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["权益模块", "权益细项", "钻石50万", "铂金30万", "黄金15万", "银牌8万"]
    rows = [
        ["品牌称号", "绿城华东区战略/首席合作伙伴称号", "● 年度战略", "● 片区首席", "○ 系列合作", "—"],
        ["主场冠名", "一滴水旗舰峰会/晚宴冠名权", "主冠名", "联合冠名", "主冠名", "—"],
        ["露出", "背景板/签到墙/手册/官网合作伙伴位", "LOGO置顶", "LOGO第二位", "LOGO主视觉", "单场LOGO"],
        ["演讲席位", "区域总或指派代表致辞/对话", "● 1–2场", "● 1场", "○", "按场次"],
        ["嘉宾权益", "VIP席位/晚宴主桌", "主桌+VIP通道", "主桌", "VIP席", "嘉宾席"],
        ["展陈", "现场展位/品牌陈列", "优先选位×2", "标准×1–2", "标准×1", "按场"],
        ["参访", "杨浦/虹口大厂或TMT参访随行", "8–12人×2次", "6–8人×1次", "4人×1次", "—"],
        ["片区联动", "黄浦/虹口/青浦项目日露出与接待协同", "三区全覆盖", "2区", "虹口", "1区"],
        ["传播", "会前会中会后图文/短视频传播包", "全年包", "季度包", "单系列", "单场"],
        ["媒体", "合作通稿与重点媒体邀约协助", "●", "●", "○", "—"],
        ["证书", "战略合作/智慧人居相关荣誉证书", "●", "●", "○", "—"],
        ["复盘", "对区域总的线索与行政互动复盘会", "双月", "季度×2", "会后1次", "—"],
        ["线索交接", "企业/个人意向线索书面交接（合规脱敏）", "●", "●", "○", "按场"],
        ["定制", "定制闭门小局（≤20人）", "1场", "○ 可加购", "—", "—"],
        ["升级", "执行期内补差升级更高档位", "—", "可升级钻石", "可升级铂金", "可累计"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, 6)
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_body(ws, 4, 18, 6)
    set_widths(ws, [14, 42, 16, 16, 14, 14])
    for r in range(4, 19):
        ws.row_dimensions[r].height = 28

    ws["A20"] = "单场沙龙/晚宴（5–10万）核心权益：单场冠名、背景板LOGO、致辞或祝酒、VIP席、会后项目参观组织协助、单场传播包。"
    ws["A20"].font = BODY_FONT
    ws["A20"].alignment = WRAP
    ws.merge_cells("A20:F20")
    ws.row_dimensions[20].height = 36


def sheet_calendar(wb):
    ws = wb.create_sheet("04-H2活动日历", 3)
    ws["A1"] = "2026下半年关键节点（建议版，可改期）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    headers = ["序号", "建议月份", "活动名称", "类型", "地点", "联动片区/项目",
               "目标客群", "建议冠名归属", "成功标准"]
    rows = [
        [1, "7月下旬–8月", "一滴水·AI商业化闭门沙龙（暖场）", "办行", "北外滩一滴水",
         "虹口", "AI应用企业主/投资人", "铂金/钻石", "到场30–40人，线索≥15"],
        [2, "9月", "一滴水·人工智能商业化落地公开日暨晚宴", "办大众+晚宴", "北外滩一滴水",
         "虹口主场+黄浦露出", "大厂/TMT/基金/媒体", "钻石主冠/铂金联合", "品牌高光，媒体传播，参观转化"],
        [3, "9–10月", "杨浦·虹口大厂/TMT参访半日", "参访", "杨浦/虹口产业带",
         "三区线索回流", "绿城客户代表+企业高管", "铂金/钻石", "参访成行，对接纪要落地"],
        [4, "10月", "黄浦·城市更新×AI企业主沙龙", "办行+项目日", "黄浦（会所/一滴水）",
         "黄浦项目", "CBD企业主/总部决策人", "铂金轮动/银牌", "到访项目，意向跟进"],
        [5, "11月", "青浦·新城产业导入与人才安居日", "项目联动日", "青浦项目",
         "青浦项目", "产业园区企业/家庭客", "铂金轮动/银牌", "企业包场或家庭日转化"],
        [6, "11月", "第二场大厂/硬科技企业交流", "参访/座谈", "杨浦或虹口",
         "按线索匹配", "硬科技/专精特新", "钻石", "形成1–2个合作意向"],
        [7, "12月", "华东区三区联动私董宴（收官）", "办行", "一滴水或核心会所",
         "区域总主持", "全年高价值客户复盘", "钻石/铂金", "年度资产沉淀与续约意向"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, len(headers))
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_body(ws, 4, 10, len(headers))
    set_widths(ws, [6, 14, 36, 14, 16, 18, 22, 16, 28])
    for r in range(4, 11):
        ws.row_dimensions[r].height = 40


def sheet_district(wb):
    ws = wb.create_sheet("05-片区项目联动", 4)
    ws["A1"] = "片区项目联动矩阵——服务区域总共识，不替代项目总销售"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    headers = ["片区", "角色", "主打叙事", "建议联动动作", "区域总动作", "项目总动作", "频次建议"]
    rows = [
        ["虹口/北外滩", "主场枢纽", "北外滩高端人居×圈层地标",
         "一滴水峰会晚宴后夜访/次日参观", "致辞定调、合影、闭门交流",
         "动线接待、产品讲解、意向跟进", "主场2–3场"],
        ["黄浦", "城市更新/企业客群", "中心城区企业置业与总部客群",
         "企业主沙龙+项目开放日", "邀请行政/协会接口（如有）",
         "CBD客群包场与转化", "1–2场"],
        ["青浦", "新城产业导入", "新城置业、人才安居、企业落位",
         "产业日/家庭日+座谈", "片区叙事背书",
         "企业与家庭客接待转化", "1–2场"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, 7)
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_body(ws, 4, 6, 7)
    set_widths(ws, [14, 14, 28, 30, 26, 26, 12])
    for r in range(4, 7):
        ws.row_dimensions[r].height = 55

    ws["A8"] = "物料露出规则（强化片区共识）"
    ws["A8"].font = HEAD_FONT
    ws["A9"] = "所有主视觉统一出现「绿城华东区」；当期主场项目作为副标题露出。禁止只打单项目品牌而弱化区域。季度复盘材料由区域办归档。"
    ws["A9"].font = BODY_FONT
    ws["A9"].alignment = WRAP
    ws.merge_cells("A9:G9")
    ws.row_dimensions[9].height = 40


def sheet_guests(wb):
    ws = wb.create_sheet("06-嘉宾与资源池", 5)
    ws["A1"] = "嘉宾与资源池方向（按议题动态匹配，不做必到硬承诺）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    headers = ["类别", "示例方向", "邀约场景", "对绿城价值", "备注"]
    rows = [
        ["人工智能大厂", "云计算、大模型应用、智能硬件等业务线负责人", "旗舰峰会/圆桌/参访", "品牌势能+企业客群", "按当季热点匹配"],
        ["TMT与产业互联网", "垂直赛道CEO/高管", "闭门沙龙/对接会", "潜在买家与合作方", "杨浦虹口密度高"],
        ["创投机构", "早期/成长期基金GP、LP代表", "晚宴/私董会", "高净值与企业资源", "适合办行"],
        ["专精特新/硬科技", "杨浦区科技企业联合会会员企业", "沙龙路演/参访", "真实产业内容", "联合会资源可协同"],
        ["行政与平台", "区级协会、园区、商会接口", "公开日/座谈会", "行政链接", "按官方流程协调"],
        ["专业媒体", "财经/地产/科技媒体", "公开日传播", "声量放大", "通稿口径由双方确认"],
        ["绿城侧", "华东区总、区域办、项目总、重点客户", "全场次", "转化闭环", "绿城内部统筹名单"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, 5)
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_body(ws, 4, 10, 5)
    set_widths(ws, [18, 40, 22, 22, 22])
    for r in range(4, 11):
        ws.row_dimensions[r].height = 36


def sheet_roi(wb):
    ws = wb.create_sheet("07-ROI与价值测算", 6)
    ws["A1"] = "ROI与价值测算（管理口径，非财务审计口径）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["价值维度", "铂金30万目标", "钻石50万目标", "衡量方式", "责任方", "说明"]
    rows = [
        ["高质量活动场次", "5–6场", "7–8场", "排期完成率", "双方项目组", "含办行/大众/参访/项目日"],
        ["圈层有效触达", "≥150人", "≥200人", "签到与会后回访", "主办方执行", "去重后有效人数"],
        ["企业/个人线索交接", "≥40条", "≥60条", "脱敏线索表", "主办方→绿城", "合规，不含违规个信"],
        ["项目到访组织", "≥2场联动到访", "≥3场", "项目接待记录", "绿城项目", "转化归绿城销售"],
        ["行政/平台互动", "≥2次有效对接", "≥3–4次", "会议纪要", "双方", "视议题与档期"],
        ["传播素材", "季度传播包", "全年传播包", "物料与报道清单", "主办方", "口径双方确认"],
        ["区域复盘会", "2次", "3次（双月）", "纪要归档", "区域办主持", "服务区域总决策"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, 6)
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_body(ws, 4, 10, 6)
    set_widths(ws, [18, 16, 16, 18, 14, 28])
    for r in range(4, 11):
        ws.row_dimensions[r].height = 30

    ws["A12"] = "单客获取成本粗算（示意）"
    ws["A12"].font = HEAD_FONT
    ws["A13"] = "若铂金30万带来40条有效线索，单线索成本约7,500元；其中进入项目到访的按20%计约8组，则单组到访获客成本约37,500元——对高端改善/企业客群通常可接受。最终以绿城成交口径评估。"
    ws["A13"].font = BODY_FONT
    ws["A13"].alignment = WRAP
    ws.merge_cells("A13:F13")
    ws.row_dimensions[13].height = 48


def sheet_checklist(wb):
    ws = wb.create_sheet("08-推进Checklist", 7)
    ws["A1"] = "明日会后推进清单"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    headers = ["阶段", "动作", "负责人建议", "状态"]
    rows = [
        ["会中", "确认主推档位：铂金30万 / 钻石50万 / 其他", "华东区总经理", "待确认"],
        ["会中", "确认首场月份与主场（建议一滴水）", "区域办", "待确认"],
        ["会中", "确认黄浦/青浦轮动顺序", "区域办+项目总", "待确认"],
        ["会中", "确认区域总亲自站台场次数量", "华东区总经理", "待确认"],
        ["会后3日", "输出正式报价单与排期确认版", "胡继刚侧", "待启动"],
        ["会后5日", "指定绿城对接人（区域办1人+项目接口）", "绿城", "待启动"],
        ["会后7日", "签约意向/合同文本交换", "双方法务/商务", "待启动"],
        ["签约后", "启动首场嘉宾与场地筹备", "主办执行组", "待启动"],
        ["每场后", "7日内提交执行简报与线索交接", "主办执行组", "机制"],
        ["季度", "区域复盘会（线索/行政/下季排期）", "区域总或委托", "机制"],
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, 4)
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_body(ws, 4, 13, 4)
    set_widths(ws, [12, 48, 22, 12])
    for r in range(4, 14):
        ws.row_dimensions[r].height = 26

    ws["A15"] = "三个当场可问的问题（备忘）"
    ws["A15"].font = HEAD_FONT
    ws["A16"] = "1）下半年主推30万还是50万？是否需要拆到项目分摊？\n2）首场定在一滴水的哪个月？黄浦/青浦谁先轮动？\n3）区域总希望亲自站台几场？需要对接哪些行政条线？"
    ws["A16"].font = BODY_FONT
    ws["A16"].alignment = WRAP
    ws.merge_cells("A16:D16")
    ws.row_dimensions[16].height = 70


def main():
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)
    sheet_overview(wb)
    sheet_tiers(wb)
    sheet_rights(wb)
    sheet_calendar(wb)
    sheet_district(wb)
    sheet_guests(wb)
    sheet_roi(wb)
    sheet_checklist(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT))
    print(f"saved: {OUT}")


if __name__ == "__main__":
    main()
