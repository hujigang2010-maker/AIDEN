#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
上海学区板块差异化选房地图
基于「避开新上海人集中新区、选择人口空心化老板块」逻辑，
对全市主要板块做学区竞争与策略契合度标记，并输出 Excel + HTML。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output"
OUT_XLSX = OUT_DIR / "上海学区板块差异化选房地图.xlsx"
OUT_HTML = OUT_DIR / "上海学区板块差异化选房地图.html"

# 标记说明（与纪要策略一一对应）
MARK_META = {
    "★优选锁定": {
        "color": "#0F766E",
        "bg": "#CCFBF1",
        "desc": "人口空心化 / 老上海人聚集，升学竞争相对可控，纪要明确推荐",
    },
    "☆次优关注": {
        "color": "#1D4ED8",
        "bg": "#DBEAFE",
        "desc": "部分符合空心化或老城区特征，可作小学阶段备选或后续置换观察",
    },
    "○可观察": {
        "color": "#A16207",
        "bg": "#FEF9C3",
        "desc": "资源尚可，但人口导入或竞争压力不低，需个案核验对口与年限",
    },
    "△条件不符": {
        "color": "#C2410C",
        "bg": "#FFEDD5",
        "desc": "学区质量尚可，但户籍年限/总价/一步到位要求与当前时间线不匹配",
    },
    "✕策略排除": {
        "color": "#B91C1C",
        "bg": "#FEE2E2",
        "desc": "新上海人导入集中、适龄人口堆积，升学卷度高，纪要建议去掉",
    },
}

# 区级总览（竞争与人口画像）
DISTRICT_ROWS = [
    # 区, 入学对口类型, 人口画像, 区内竞争, 本区保护, 中考相对压力, 策略结论, 标记
    ("黄浦", "学籍对口", "老城区，常住人口相对少，老龄化明显", "低", "强", "相对友好（录取率较高）", "核心优选区：人口与学位匹配更接近「空心化」", "★优选锁定"),
    ("杨浦", "学籍对口", "老工业区转型，滨江有更新，鞍山等老小区存量大", "中低", "强", "中等", "滨江及老板块可挖「学校在、人口出」机会", "★优选锁定"),
    ("虹口", "学籍对口", "老城区为主，北外滩更新中", "中低", "较强", "中等", "部分老板块可观察，整体次于黄浦", "☆次优关注"),
    ("静安", "户籍/学籍混合特征强", "核心区+大宁等导入并存", "中", "强", "中等偏卷", "优质但总价高，性价比与空心化特征一般", "○可观察"),
    ("长宁", "学籍+派位", "古北等外籍/高收入导入，北新泾有老小区", "中", "较强", "中等", "公办均衡，但非纪要主推空心化路径", "○可观察"),
    ("徐汇", "学籍+派位", "衡复/滨江优质，华泾等有导入", "中高", "中等", "顶尖资源强但竞争也强", "复旦附属九年一贯等质量好，但年限与总价不匹配明年入学", "△条件不符"),
    ("普陀", "学籍对口", "真如/桃浦等有人口导入", "中", "中等", "中等", "部分老板块可观察，整体非首选", "○可观察"),
    ("浦东", "户籍对口", "人口体量全市最大，新上海人导入极多", "高（多数板块）", "较强", "分化极大", "仅保留陆家嘴潍坊等少数老板块；其余多数排除", "○可观察"),
    ("闵行", "户籍对口", "新上海人高度集中，适龄儿童多", "极高", "弱（民办开放）", "高", "纪要：去掉大部分；九年一贯年限也不达标", "✕策略排除"),
    ("宝山", "户籍对口", "人口多，教育资源相对滞后", "高", "中等", "高", "新人口集中、竞争大，纪要明确避开", "✕策略排除"),
    ("嘉定", "户籍对口", "新城导入强，户籍净流入高", "高", "尚可", "高", "纪要：去掉嘉定；年限预警趋严", "✕策略排除"),
    ("松江", "户籍对口", "九亭等早期导入极强", "极高（九亭尤甚）", "弱", "高", "纪要：去掉松江；九亭属高竞争早期段位", "✕策略排除"),
    ("青浦", "户籍对口", "赵巷/徐泾等高端导入", "高", "中等", "中高", "人口导入型，非空心化路径", "✕策略排除"),
    ("奉贤", "户籍对口", "南桥等有导入，整体密度低于近郊热门", "中", "中等", "中等", "通勤与资源一般，非本次学区主线", "○可观察"),
    ("金山", "户籍对口", "人口相对少", "较低", "中等", "较低", "竞争小但优质公办与通勤弱，非本次目标", "○可观察"),
    ("崇明", "户籍对口", "人口流出为主", "低", "中等", "低", "空心化极端，但升学资源与通勤不适合主城区就业家庭", "○可观察"),
]

# 板块明细
# 字段: 区, 板块, 代表小学/学区, 代表初中路径, 人口结构, 竞争压力(1-5),
#       空心化契合(1-5), 总价带(万), 年限/政策提示, 策略说明, 标记, 是否入选结果
PLATE_ROWS = [
    # —— 黄浦（优选）——
    ("黄浦", "老西门/蓬莱", "蓬莱路第二小学等", "学籍对口公办初中", "老上海人为主，老龄化", 2, 5, "250-450", "核验当年对口与人户一致", "老城区典型空心化特征，公办均衡，竞争相对可控", "★优选锁定", "是"),
    ("黄浦", "打浦桥", "卢湾二中心等", "学籍对口+区内初中", "老小区+局部更新", 2, 5, "280-500", "核验入户截止与人户一致", "人口少于学位紧张度低，符合「学校在、人口出」", "★优选锁定", "是"),
    ("黄浦", "五里桥", "区属公办小学群", "学籍对口", "老公房为主", 2, 5, "220-400", "核验对口表", "总价更友好的黄浦入口，适合先解决小学", "★优选锁定", "是"),
    ("黄浦", "豫园/小东门", "实验小学等周边", "学籍对口", "老城区居住人口下降", 2, 4, "250-480", "核验对口", "老板块，竞争低于近郊导入区", "★优选锁定", "是"),
    ("黄浦", "人民广场/南京东路", "区属优质公办", "学籍对口", "商业占比高，带娃家庭密度低", 2, 4, "300-600", "总价偏高需精选小户型", "居住属性弱、适龄密度低，策略契合", "☆次优关注", "是"),
    # —— 杨浦 ——
    ("杨浦", "杨浦滨江", "定海/滨江规划校群", "九年一贯及滨江新校规划", "更新中，原住民流出+局部导入", 2, 4, "300-700", "关注新校招生政策落地进度", "纪要点名符合空心化特征；需盯学校兑现", "★优选锁定", "是"),
    ("杨浦", "鞍山/四平", "鞍山等一梯队公办小学", "杨浦优质公办初中链", "老公房存量极大，年轻带娃家庭少", 2, 5, "180-350", "核验对口与年限", "「学校强、房子旧、人口出」典型，低总价挂户逻辑", "★优选锁定", "是"),
    ("杨浦", "黄兴公园/延吉", "区属公办小学", "学籍对口", "老小区为主", 3, 4, "200-380", "核验对口", "次于鞍山但同属老杨浦逻辑", "☆次优关注", "是"),
    ("杨浦", "五角场", "大学城周边公办/民办", "公民办并存", "高校+年轻租住导入", 3, 3, "280-550", "民办与公办路径分开评估", "导入人口多于纯老破小板块，降一档", "○可观察", "否"),
    ("杨浦", "中原", "区属公办", "学籍对口", "成熟居住区，适龄中等", 3, 3, "220-400", "核验对口", "中性板块，非空心化极值", "○可观察", "否"),
    # —— 浦东（仅潍坊等少数入选）——
    ("浦东", "潍坊新村（陆家嘴南）", "明珠小学 / 浦师附小", "东昌南校 / 洋泾菊园潍坊校区等", "老上海人为主，租客少本地生娃", 2, 5, "200-300", "关注五年一户与入户预警；优先世纪大道南侧", "纪要明确推荐；崂山一村至十村可筛；核心入选结果", "★优选锁定", "是"),
    ("浦东", "塘桥", "塘桥周边公办", "洋泾/建平等路径分化", "老小区+局部改善", 3, 3, "250-450", "核验对口分组", "毗邻潍坊，可作观察，竞争高于潍坊老破小", "☆次优关注", "是"),
    ("浦东", "洋泾", "洋泾系小学", "洋泾中学等", "成熟居住，适龄不少", 3, 3, "280-500", "核验年限", "资源好但非空心化主线", "○可观察", "否"),
    ("浦东", "陆家嘴金融城核心", "少量对口小学", "优质初中周边", "商务为主，住宅少", 2, 3, "400-900+", "总价高", "居住供给少，挂户性价比不如潍坊南侧", "○可观察", "否"),
    ("浦东", "花木/世纪公园", "建平系等", "建平等", "改善型+适龄家庭集中", 4, 2, "400-800", "年限普遍不低", "人口导入与改善家庭集中，卷度高", "✕策略排除", "否"),
    ("浦东", "张江", "张江周边公办/实验", "张江/中芯等路径", "高新人才新上海人高度集中", 5, 1, "350-700", "热门户口年限紧", "典型新人口堆叠，纪要：浦东大部分去掉", "✕策略排除", "否"),
    ("浦东", "金桥", "金桥公办群", "区内初中", "产业导入家庭多", 4, 2, "280-500", "核验对口", "导入型，非空心化", "✕策略排除", "否"),
    ("浦东", "三林/御桥", "御桥小学等", "区内初中", "新盘+年轻家庭", 5, 1, "250-450", "部分学校年限不降反升", "适龄堆积，竞争加剧", "✕策略排除", "否"),
    ("浦东", "周浦/康桥", "周浦公办", "区内", "近郊导入", 4, 1, "180-350", "核验", "人口导入型外围", "✕策略排除", "否"),
    ("浦东", "川沙/唐镇", "川沙公办", "区内", "新城导入", 4, 1, "200-400", "核验", "纪要排除浦东大部分", "✕策略排除", "否"),
    ("浦东", "临港", "临港新建校", "区内", "政策导入青年", 4, 1, "150-350", "新校兑现不确定", "远郊导入，非本次主线", "✕策略排除", "否"),
    ("浦东", "北蔡", "北蔡公办", "区内", "成熟+局部导入", 4, 2, "250-450", "核验", "竞争不低，非优选", "✕策略排除", "否"),
    # —— 徐汇 ——
    ("徐汇", "衡复/徐家汇", "区属一梯队小学", "华育/世外/南模等路径", "优质改善家庭集中", 4, 2, "500-1200+", "总价高、竞争强", "资源顶尖但与「降竞争」目标相反", "○可观察", "否"),
    ("徐汇", "徐汇滨江", "滨江配套校", "徐汇优质初中圈", "高端改善", 3, 2, "600-1500+", "房价少数仍坚挺", "投资属性强，非低竞争挂户路径", "○可观察", "否"),
    ("徐汇", "复旦附属九年一贯（徐汇相关）", "复旦附属九年一贯", "小初一贯", "学区溢价明显", 3, 2, "高", "户籍房产一致且需提前约3年落户", "质量不错，但明年入学年限不达标 → 不推荐", "△条件不符", "否"),
    ("徐汇", "华泾/漕河泾", "区属公办", "徐汇路径", "产业导入+居住", 4, 2, "300-600", "核验", "导入压力高于老黄浦", "○可观察", "否"),
    # —— 闵行（大部分排除）——
    ("闵行", "莘庄", "莘庄镇小等", "上宝等民办强区路径", "新上海人高度集中", 5, 1, "250-500", "区内区外双卷", "纪要：去掉闵行大部分", "✕策略排除", "否"),
    ("闵行", "七宝", "七宝实验等", "文来/上闵外等", "改善+鸡娃家庭集中", 5, 1, "300-600", "竞争极高", "高竞争代表板块", "✕策略排除", "否"),
    ("闵行", "古美/梅陇", "区属公办", "闵行路径", "成熟导入居住区", 5, 1, "280-500", "核验", "适龄密度高", "✕策略排除", "否"),
    ("闵行", "颛桥/浦江", "区属公办", "闵行路径", "近郊导入", 4, 1, "200-400", "核验", "导入型", "✕策略排除", "否"),
    ("闵行", "复旦附属九年一贯（闵行）", "复旦附属九年一贯", "小初一贯", "学区溢价", 3, 2, "高", "户籍房产一致+提前约3年", "纪要点名：明年上小学不符合条件", "△条件不符", "否"),
    ("闵行", "华漕", "华漕公办", "闵行路径", "虹桥枢纽导入", 4, 1, "220-450", "核验", "非空心化", "✕策略排除", "否"),
    # —— 松江 ——
    ("松江", "九亭", "九亭小学等", "松江初中/民办", "早期新上海人堆叠样板", 5, 1, "200-400", "卷度极高", "纪要点名：高竞争早期段位，不推荐", "✕策略排除", "否"),
    ("松江", "泗泾", "泗泾公办", "松江路径", "导入居住", 5, 1, "180-350", "核验", "去掉松江", "✕策略排除", "否"),
    ("松江", "松江新城", "实验小学等", "松江路径", "新城导入", 4, 1, "180-400", "核验", "去掉松江", "✕策略排除", "否"),
    ("松江", "新桥", "新桥公办", "松江路径", "导入", 4, 1, "150-300", "核验", "去掉松江", "✕策略排除", "否"),
    # —— 嘉定 ——
    ("嘉定", "嘉定新城", "普通小学等热门校", "嘉定华二等", "户籍净流入高", 5, 1, "200-450", "入户年限预警趋严", "纪要：去掉嘉定", "✕策略排除", "否"),
    ("嘉定", "南翔", "南翔公办", "嘉定路径", "轨交导入", 4, 1, "220-450", "核验", "去掉嘉定", "✕策略排除", "否"),
    ("嘉定", "江桥", "江桥公办", "嘉定路径", "近郊导入", 4, 1, "180-350", "核验", "去掉嘉定", "✕策略排除", "否"),
    ("嘉定", "安亭/马陆", "安亭公办", "嘉定路径", "产业导入", 4, 1, "150-320", "核验", "去掉嘉定", "✕策略排除", "否"),
    # —— 宝山 ——
    ("宝山", "顾村", "顾村公办", "宝山路径", "大型社区导入", 5, 1, "180-350", "核验", "纪要避开新人口集中区", "✕策略排除", "否"),
    ("宝山", "共康/高境", "区属公办", "宝山路径", "近市区导入", 4, 2, "220-400", "核验", "人口压力大、资源相对紧", "✕策略排除", "否"),
    ("宝山", "淞宝/吴淞", "区属公办", "宝山路径", "老工业居住+更新", 3, 3, "180-350", "核验", "略好于顾村，但仍非空心化优选", "○可观察", "否"),
    ("宝山", "罗店/大场", "区属公办", "宝山路径", "外围导入", 4, 1, "150-300", "核验", "排除", "✕策略排除", "否"),
    # —— 其他区代表性板块 ——
    ("虹口", "北外滩", "虹口公办小学", "学籍对口", "更新中，原住民结构变化", 3, 3, "300-700", "核验对口", "可作黄浦外溢观察", "☆次优关注", "是"),
    ("虹口", "四川北路/临平路", "区属公办", "学籍对口", "老城区", 2, 4, "220-400", "核验", "老虹口空心化特征，次优", "☆次优关注", "是"),
    ("静安", "南京西路/静安寺", "一梯队公办", "静安优质路径", "核心区高总价", 3, 2, "500-1200+", "总价门槛高", "资源好但非低竞争低总价路径", "○可观察", "否"),
    ("静安", "大宁", "大宁公办", "静安路径", "改善导入", 4, 2, "350-650", "核验", "导入特征强于空心化", "○可观察", "否"),
    ("长宁", "中山公园", "长宁公办", "学籍+派位", "成熟居住", 3, 3, "350-650", "核验派位规则", "均衡但非纪要主推", "○可观察", "否"),
    ("长宁", "古北", "国际社区周边公办", "长宁路径", "外籍/高收入", 3, 2, "400-900", "居住属性强", "非国内升学空心化主线", "○可观察", "否"),
    ("长宁", "北新泾", "区属公办", "长宁路径", "老公房较多", 3, 3, "220-400", "核验", "可个案看，非首选", "○可观察", "否"),
    ("普陀", "真如", "真如公办", "学籍对口", "副中心导入", 4, 2, "250-450", "核验", "导入压力", "○可观察", "否"),
    ("普陀", "长风/武宁", "区属公办", "普陀路径", "成熟+局部改善", 3, 3, "280-500", "核验", "中性", "○可观察", "否"),
    ("普陀", "桃浦", "桃浦公办", "普陀路径", "更新导入", 4, 2, "200-380", "核验", "导入型", "○可观察", "否"),
    ("青浦", "徐泾", "徐泾公办", "青浦路径", "虹桥商务导入", 4, 1, "250-500", "核验", "高端导入，排除主线", "✕策略排除", "否"),
    ("青浦", "赵巷", "赵巷公办", "青浦路径", "改善别墅盘导入", 4, 1, "300-800+", "核验", "非空心化挂户逻辑", "✕策略排除", "否"),
    ("青浦", "青浦新城", "区属公办", "青浦路径", "新城导入", 4, 1, "150-350", "核验", "排除主线", "✕策略排除", "否"),
    ("奉贤", "南桥", "南桥公办", "奉贤路径", "新城中心", 3, 2, "120-280", "通勤远", "竞争一般但偏离主城区就业与优质初中圈", "○可观察", "否"),
    ("金山", "石化/山阳", "金山公办", "金山路径", "人口相对少", 2, 3, "80-200", "通勤远", "空心但资源与通勤不匹配目标", "○可观察", "否"),
    ("崇明", "城桥", "崇明公办", "崇明路径", "人口流出", 1, 4, "60-180", "通勤极远", "极端空心化，不适合多数家庭", "○可观察", "否"),
]

LOGIC_POINTS = [
    ("核心目标", "购房目的是获得优质上学资格，而非房产增值投资。"),
    ("升学难点", "上海高考本科率约88%（对比山东约33.6%），核心难点在初中升高中；优先布局小学+初中。"),
    ("板块原则", "优先人口空心化区域（居住人口 < 学校配套容量）；避开新上海人集中、孩子扎堆出生的新区。"),
    ("排除范围", "去掉嘉定、松江、闵行（大部分）、浦东（大部分）；九亭属高竞争早期段位。"),
    ("保留范围", "黄浦、杨浦滨江（及鞍山等老板块）、浦东陆家嘴潍坊新村（明珠小学，世纪大道南侧，崂山一村–十村，约200–300万）。"),
    ("学段策略", "不必一步到位买「双优」学区；可先解决小学，小学结束后再置换初中学区。"),
    ("年限红线", "徐汇/闵行复旦附属九年一贯需户籍房产一致且提前约3年；明年入学不符，不推荐。"),
    ("房价周期", "长周期多数板块仍偏弱；上学需求与投资时机可能冲突，需家庭自行取舍。"),
]


def score_fit(competition: int, hollow: int) -> int:
    """策略契合分：空心化高、竞争低 → 分高。满分 100。"""
    return max(0, min(100, hollow * 12 + (6 - competition) * 10 + 8))


def thin_border() -> Border:
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="0F766E")


def mark_fill(mark: str) -> PatternFill:
    hex_bg = MARK_META[mark]["bg"].lstrip("#")
    return PatternFill("solid", fgColor=hex_bg)


def build_workbook() -> Workbook:
    wb = Workbook()

    # —— Sheet1: 入选结果 ——
    ws = wb.active
    ws.title = "入选结果"
    ws["A1"] = "上海学区板块差异化选房 · 入选结果（基于纪要策略）"
    ws["A1"].font = Font(bold=True, size=16, color="0F766E")
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "筛选逻辑：人口空心化 / 老上海人聚集 / 低升学竞争 / 总价相对友好 / 匹配明年入学时间线。"
        "「是」= 本轮建议纳入看房清单；具体对口以当年教育局地段表为准。"
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 40

    headers = [
        "标记",
        "区",
        "板块",
        "代表小学/学区",
        "代表初中路径",
        "参考总价带(万)",
        "策略契合分",
        "入选理由（摘要）",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    selected = [r for r in PLATE_ROWS if r[-1] == "是"]
    selected.sort(key=lambda r: (-score_fit(r[5], r[6]), r[0], r[1]))
    for i, row in enumerate(selected, 4):
        district, plate, primary, middle, _pop, comp, hollow, price, _policy, reason, mark, _sel = row
        vals = [mark, district, plate, primary, middle, price, score_fit(comp, hollow), reason]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(i, col, v)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                cell.fill = mark_fill(mark)
                cell.font = Font(bold=True, color=MARK_META[mark]["color"].lstrip("#"))

    widths = [12, 8, 22, 28, 28, 14, 12, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{3 + len(selected)}"

    # —— Sheet2: 全市板块总表 ——
    ws2 = wb.create_sheet("全市板块标记总表")
    ws2["A1"] = "全市主要板块学区策略标记总表（覆盖核心+近郊代表性板块，非官方168板块全量名录）"
    ws2["A1"].font = Font(bold=True, size=14, color="0F766E")
    ws2.merge_cells("A1:M1")
    h2 = [
        "标记",
        "是否入选",
        "区",
        "板块",
        "代表小学/学区",
        "代表初中路径",
        "人口结构画像",
        "竞争压力(1低-5高)",
        "空心化契合(1低-5高)",
        "策略契合分",
        "参考总价带(万)",
        "年限/政策提示",
        "策略说明",
    ]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(2, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(PLATE_ROWS, 3):
        district, plate, primary, middle, pop, comp, hollow, price, policy, reason, mark, sel = row
        vals = [
            mark,
            sel,
            district,
            plate,
            primary,
            middle,
            pop,
            comp,
            hollow,
            score_fit(comp, hollow),
            price,
            policy,
            reason,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(i, col, v)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                cell.fill = mark_fill(mark)
                cell.font = Font(bold=True, color=MARK_META[mark]["color"].lstrip("#"))
            if col == 2 and v == "是":
                cell.fill = PatternFill("solid", fgColor="BBF7D0")
                cell.font = Font(bold=True, color="166534")

    for i, w in enumerate([12, 10, 8, 26, 26, 26, 28, 12, 12, 12, 14, 28, 42], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:M{2 + len(PLATE_ROWS)}"
    ws2.row_dimensions[2].height = 32

    # —— Sheet3: 区级总览 ——
    ws3 = wb.create_sheet("十六区策略总览")
    ws3["A1"] = "上海十六区 · 升学竞争与人口画像（服务板块筛选）"
    ws3["A1"].font = Font(bold=True, size=14, color="0F766E")
    ws3.merge_cells("A1:H1")
    h3 = ["标记", "区", "入学对口类型", "人口画像", "区内竞争", "本区保护", "中考相对压力", "策略结论"]
    for col, h in enumerate(h3, 1):
        cell = ws3.cell(2, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(DISTRICT_ROWS, 3):
        mark = row[7]
        vals = [mark, row[0], row[1], row[2], row[3], row[4], row[5], row[6]]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(i, col, v)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                cell.fill = mark_fill(mark)
                cell.font = Font(bold=True, color=MARK_META[mark]["color"].lstrip("#"))
    for i, w in enumerate([12, 8, 18, 40, 12, 12, 22, 42], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A3"
    ws3.row_dimensions[2].height = 28

    # —— Sheet4: 标记图例与纪要逻辑 ——
    ws4 = wb.create_sheet("图例与纪要逻辑")
    ws4["A1"] = "标记图例"
    ws4["A1"].font = Font(bold=True, size=14, color="0F766E")
    ws4["A3"] = "标记"
    ws4["B3"] = "含义"
    ws4["A3"].font = Font(bold=True, color="FFFFFF")
    ws4["B3"].font = Font(bold=True, color="FFFFFF")
    ws4["A3"].fill = header_fill()
    ws4["B3"].fill = header_fill()
    for i, (mark, meta) in enumerate(MARK_META.items(), 4):
        ws4.cell(i, 1, mark).fill = mark_fill(mark)
        ws4.cell(i, 1).font = Font(bold=True, color=meta["color"].lstrip("#"))
        ws4.cell(i, 2, meta["desc"])
        ws4.cell(i, 1).border = thin_border()
        ws4.cell(i, 2).border = thin_border()

    ws4["A10"] = "纪要核心逻辑（决策框架）"
    ws4["A10"].font = Font(bold=True, size=14, color="0F766E")
    ws4["A11"] = "维度"
    ws4["B11"] = "内容"
    ws4["A11"].font = Font(bold=True, color="FFFFFF")
    ws4["B11"].font = Font(bold=True, color="FFFFFF")
    ws4["A11"].fill = header_fill()
    ws4["B11"].fill = header_fill()
    for i, (k, v) in enumerate(LOGIC_POINTS, 12):
        ws4.cell(i, 1, k).border = thin_border()
        ws4.cell(i, 2, v).border = thin_border()
        ws4.cell(i, 2).alignment = Alignment(wrap_text=True)

    ws4["A21"] = "使用说明"
    ws4["A21"].font = Font(bold=True, size=14, color="0F766E")
    notes = [
        "1. 本表是决策辅助框架，不是官方对口承诺；买房前必须核对目标学校最新招生简章与对口地段表。",
        "2. 「策略契合分」= 空心化契合×12 + (6−竞争压力)×10 + 8，用于同标记内排序，非房价预测。",
        "3. 入选清单建议看房顺序：潍坊新村（世纪大道南）→ 杨浦鞍山/滨江 → 黄浦五里桥/打浦桥/老西门 → 虹口老板块观察。",
        "4. 小学与初中可分两步走；高中统考，不必提前用学区房锁定。",
        "5. 若家庭更重视资产保值，需单独评估徐汇滨江等少数抗跌板块，那是另一套决策，与本升学策略可冲突。",
    ]
    for i, t in enumerate(notes, 22):
        ws4.cell(i, 1, t)
        ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ws4.cell(i, 1).alignment = Alignment(wrap_text=True)
        ws4.row_dimensions[i].height = 28

    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 88

    # —— Sheet5: 统计 ——
    ws5 = wb.create_sheet("标记统计")
    ws5["A1"] = "标记分布统计"
    ws5["A1"].font = Font(bold=True, size=14, color="0F766E")
    ws5["A3"] = "标记"
    ws5["B3"] = "板块数量"
    ws5["C3"] = "其中入选"
    for col in range(1, 4):
        ws5.cell(3, col).font = Font(bold=True, color="FFFFFF")
        ws5.cell(3, col).fill = header_fill()
    from collections import Counter

    c_all = Counter(r[10] for r in PLATE_ROWS)
    c_sel = Counter(r[10] for r in PLATE_ROWS if r[11] == "是")
    for i, mark in enumerate(MARK_META.keys(), 4):
        ws5.cell(i, 1, mark).fill = mark_fill(mark)
        ws5.cell(i, 2, c_all.get(mark, 0))
        ws5.cell(i, 3, c_sel.get(mark, 0))
    ws5["A10"] = f"板块样本合计：{len(PLATE_ROWS)}"
    ws5["A11"] = f"本轮入选合计：{sum(1 for r in PLATE_ROWS if r[11] == '是')}"
    ws5.column_dimensions["A"].width = 14
    ws5.column_dimensions["B"].width = 12
    ws5.column_dimensions["C"].width = 12

    return wb


def build_html() -> str:
    from collections import Counter
    import json

    selected = [r for r in PLATE_ROWS if r[-1] == "是"]
    selected.sort(key=lambda r: (-score_fit(r[5], r[6]), r[0], r[1]))

    def plate_obj(r):
        d, plate, primary, middle, pop, comp, hollow, price, policy, reason, mark, sel = r
        return {
            "mark": mark,
            "selected": sel,
            "district": d,
            "plate": plate,
            "primary": primary,
            "middle": middle,
            "pop": pop,
            "competition": comp,
            "hollow": hollow,
            "score": score_fit(comp, hollow),
            "price": price,
            "policy": policy,
            "reason": reason,
        }

    data = [plate_obj(r) for r in PLATE_ROWS]
    selected_data = [plate_obj(r) for r in selected]
    c_all = Counter(r[10] for r in PLATE_ROWS)

    mark_cards = "".join(
        f'<div class="mark-card" style="--c:{m["color"]};--bg:{m["bg"]}"><b>{k}</b><span>{m["desc"]}</span></div>'
        for k, m in MARK_META.items()
    )
    logic_html = "".join(f"<li><strong>{k}</strong> — {v}</li>" for k, v in LOGIC_POINTS)
    district_rows = "".join(
        f"<tr><td><span class='tag' data-m='{r[7]}'>{r[7]}</span></td>"
        f"<td>{r[0]}</td><td>{r[1]}</td><td>{r[2]}</td><td>{r[3]}</td><td>{r[4]}</td><td>{r[5]}</td><td>{r[6]}</td></tr>"
        for r in DISTRICT_ROWS
    )
    stats = "".join(
        f'<div class="stat"><div class="n">{c_all.get(k, 0)}</div><div class="l">{k}</div></div>'
        for k in MARK_META
    )

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>上海学区板块差异化选房地图</title>
<style>
  :root {{
    --ink: #0f172a;
    --muted: #64748b;
    --line: #e2e8f0;
    --bg: #f8fafc;
    --card: #ffffff;
    --teal: #0f766e;
    --teal-soft: #ccfbf1;
    --serif: "Source Han Serif SC", "Noto Serif SC", "Songti SC", "SimSun", serif;
    --sans: "PingFang SC", "Noto Sans SC", "Microsoft YaHei", sans-serif;
  }}
  * {{ box-sizing: border-box; }}
  body {{
    margin: 0;
    font-family: var(--sans);
    color: var(--ink);
    background:
      radial-gradient(1200px 600px at 10% -10%, #ccfbf1 0%, transparent 55%),
      radial-gradient(900px 500px at 100% 0%, #e0f2fe 0%, transparent 50%),
      linear-gradient(180deg, #f0fdfa 0%, var(--bg) 40%, #f8fafc 100%);
    min-height: 100vh;
  }}
  .wrap {{ max-width: 1180px; margin: 0 auto; padding: 32px 20px 64px; }}
  header.hero {{
    padding: 28px 0 8px;
    animation: rise .7s ease both;
  }}
  .eyebrow {{
    color: var(--teal);
    font-weight: 600;
    letter-spacing: .08em;
    font-size: 13px;
    text-transform: uppercase;
  }}
  h1 {{
    font-family: var(--serif);
    font-weight: 700;
    font-size: clamp(28px, 4vw, 42px);
    line-height: 1.25;
    margin: 8px 0 12px;
  }}
  .lead {{
    max-width: 720px;
    color: var(--muted);
    font-size: 15px;
    line-height: 1.7;
    margin: 0 0 20px;
  }}
  .chips {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 28px; }}
  .chip {{
    background: var(--card);
    border: 1px solid var(--line);
    padding: 6px 12px;
    border-radius: 999px;
    font-size: 13px;
    color: var(--muted);
  }}
  .chip strong {{ color: var(--teal); }}
  .grid-stats {{
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 10px;
    margin: 18px 0 28px;
  }}
  @media (max-width: 800px) {{
    .grid-stats {{ grid-template-columns: repeat(2, 1fr); }}
  }}
  .stat {{
    background: var(--card);
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 14px 12px;
    text-align: center;
    animation: rise .6s ease both;
  }}
  .stat .n {{ font-size: 28px; font-weight: 700; color: var(--teal); font-family: var(--serif); }}
  .stat .l {{ font-size: 12px; color: var(--muted); margin-top: 4px; }}
  section {{
    background: var(--card);
    border: 1px solid var(--line);
    border-radius: 18px;
    padding: 22px;
    margin-bottom: 18px;
    box-shadow: 0 1px 0 rgba(15, 23, 42, .03);
    animation: rise .75s ease both;
  }}
  section h2 {{
    font-family: var(--serif);
    font-size: 22px;
    margin: 0 0 6px;
  }}
  section .sub {{ color: var(--muted); font-size: 13px; margin-bottom: 16px; }}
  .mark-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 10px;
  }}
  .mark-card {{
    background: var(--bg);
    border-left: 4px solid var(--c);
    border-radius: 10px;
    padding: 12px;
    display: flex;
    flex-direction: column;
    gap: 6px;
    background: var(--bg);
  }}
  .mark-card b {{ color: var(--c); }}
  .mark-card span {{ font-size: 12px; color: var(--muted); line-height: 1.5; }}
  ul.logic {{ margin: 0; padding-left: 18px; }}
  ul.logic li {{ margin: 8px 0; line-height: 1.6; color: #334155; font-size: 14px; }}
  .toolbar {{
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    align-items: center;
    margin-bottom: 14px;
  }}
  input, select {{
    border: 1px solid var(--line);
    border-radius: 10px;
    padding: 9px 12px;
    font: inherit;
    background: #fff;
  }}
  input {{ min-width: 200px; flex: 1; }}
  .btn {{
    border: none;
    background: var(--teal);
    color: #fff;
    border-radius: 10px;
    padding: 9px 14px;
    cursor: pointer;
    font: inherit;
  }}
  .btn.secondary {{ background: #fff; color: var(--teal); border: 1px solid var(--teal); }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }}
  th, td {{
    border-bottom: 1px solid var(--line);
    padding: 10px 8px;
    text-align: left;
    vertical-align: top;
  }}
  th {{
    position: sticky;
    top: 0;
    background: #f0fdfa;
    color: var(--teal);
    font-weight: 600;
    z-index: 1;
  }}
  .table-scroll {{ overflow: auto; max-height: 560px; border: 1px solid var(--line); border-radius: 12px; }}
  .tag {{
    display: inline-block;
    padding: 2px 8px;
    border-radius: 6px;
    font-weight: 700;
    font-size: 12px;
    white-space: nowrap;
  }}
  .tag[data-m="★优选锁定"] {{ background: #ccfbf1; color: #0f766e; }}
  .tag[data-m="☆次优关注"] {{ background: #dbeafe; color: #1d4ed8; }}
  .tag[data-m="○可观察"] {{ background: #fef9c3; color: #a16207; }}
  .tag[data-m="△条件不符"] {{ background: #ffedd5; color: #c2410c; }}
  .tag[data-m="✕策略排除"] {{ background: #fee2e2; color: #b91c1c; }}
  .score {{
    font-weight: 700;
    font-variant-numeric: tabular-nums;
    color: var(--teal);
  }}
  .pick {{
    display: grid;
    gap: 12px;
  }}
  .pick-card {{
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 14px 16px;
    background: linear-gradient(135deg, #ffffff, #f0fdfa);
    display: grid;
    grid-template-columns: 88px 1fr auto;
    gap: 12px;
    align-items: start;
    transition: transform .2s ease, box-shadow .2s ease;
  }}
  .pick-card:hover {{ transform: translateY(-2px); box-shadow: 0 10px 24px rgba(15,118,110,.08); }}
  .rank {{
    font-family: var(--serif);
    font-size: 34px;
    color: var(--teal);
    line-height: 1;
  }}
  .pick-card h3 {{ margin: 0 0 4px; font-size: 16px; }}
  .pick-card p {{ margin: 0; color: var(--muted); font-size: 13px; line-height: 1.55; }}
  .meta {{ font-size: 12px; color: #475569; margin-top: 6px; }}
  footer {{
    text-align: center;
    color: var(--muted);
    font-size: 12px;
    margin-top: 24px;
    line-height: 1.6;
  }}
  @keyframes rise {{
    from {{ opacity: 0; transform: translateY(10px); }}
    to {{ opacity: 1; transform: none; }}
  }}
</style>
</head>
<body>
<div class="wrap">
  <header class="hero">
    <div class="eyebrow">SHANGHAI SCHOOL-DISTRICT MAP</div>
    <h1>上海学区板块差异化选房地图</h1>
    <p class="lead">
      按纪要思路：跳出「人口导入=资产增值」惯性，优先找<strong>人口空心化、老上海人聚集、学位相对宽松</strong>的老板块；
      避开嘉定/松江/闵行大部分/浦东大部分等高竞争导入区。下表给出标记与本轮入选结果。
    </p>
    <div class="chips">
      <span class="chip">样本板块 <strong>{len(PLATE_ROWS)}</strong></span>
      <span class="chip">本轮入选 <strong>{len(selected)}</strong></span>
      <span class="chip">核心推荐 <strong>潍坊 · 黄浦老城 · 杨浦滨江/鞍山</strong></span>
    </div>
    <div class="grid-stats">{stats}</div>
  </header>

  <section>
    <h2>本轮入选看房清单</h2>
    <p class="sub">按策略契合分排序。建议顺序：潍坊新村 → 杨浦鞍山/滨江 → 黄浦老板块 → 虹口观察盘。</p>
    <div class="pick" id="picks"></div>
  </section>

  <section>
    <h2>标记图例</h2>
    <p class="sub">五档标记直接对应纪要中的保留 / 排除 / 年限不符判断。</p>
    <div class="mark-grid">{mark_cards}</div>
  </section>

  <section>
    <h2>纪要决策框架</h2>
    <ul class="logic">{logic_html}</ul>
  </section>

  <section>
    <h2>全市板块筛选台</h2>
    <p class="sub">可按标记、区、关键词过滤；「是否入选」表示本轮是否纳入看房清单。</p>
    <div class="toolbar">
      <input id="q" placeholder="搜索板块 / 学校 / 区…" />
      <select id="markFilter">
        <option value="">全部标记</option>
        {''.join(f'<option value="{k}">{k}</option>' for k in MARK_META)}
      </select>
      <select id="selFilter">
        <option value="">全部</option>
        <option value="是">仅入选</option>
        <option value="否">未入选</option>
      </select>
      <button class="btn" id="resetBtn" type="button">重置</button>
    </div>
    <div class="table-scroll">
      <table>
        <thead>
          <tr>
            <th>标记</th><th>入选</th><th>区</th><th>板块</th><th>小学/学区</th>
            <th>初中路径</th><th>竞争</th><th>空心化</th><th>契合分</th><th>总价带</th><th>说明</th>
          </tr>
        </thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </section>

  <section>
    <h2>十六区策略总览</h2>
    <p class="sub">先选区、再选板块。黄浦 / 杨浦优先；闵行 / 宝山 / 嘉定 / 松江 / 青浦主线排除。</p>
    <div class="table-scroll" style="max-height:420px">
      <table>
        <thead>
          <tr>
            <th>标记</th><th>区</th><th>对口类型</th><th>人口画像</th>
            <th>区内竞争</th><th>本区保护</th><th>中考压力</th><th>策略结论</th>
          </tr>
        </thead>
        <tbody>{district_rows}</tbody>
      </table>
    </div>
  </section>

  <footer>
    本页为决策辅助材料，对口学校、入户年限、五年一户等以各区教育局当年文件为准。<br/>
    同步交付：<code>output/上海学区板块差异化选房地图.xlsx</code>
  </footer>
</div>
<script>
const DATA = {json.dumps(data, ensure_ascii=False)};
const SELECTED = {json.dumps(selected_data, ensure_ascii=False)};

function renderPicks() {{
  const root = document.getElementById('picks');
  root.innerHTML = SELECTED.map((r, i) => `
    <article class="pick-card">
      <div class="rank">${{String(i+1).padStart(2,'0')}}</div>
      <div>
        <h3>${{r.district}} · ${{r.plate}}</h3>
        <p>${{r.reason}}</p>
        <div class="meta">小学：${{r.primary}} ｜ 初中：${{r.middle}} ｜ 总价带：${{r.price}}万 ｜ ${{r.policy}}</div>
      </div>
      <div>
        <span class="tag" data-m="${{r.mark}}">${{r.mark}}</span>
        <div class="score" style="margin-top:8px;text-align:right">${{r.score}}</div>
      </div>
    </article>
  `).join('');
}}

function renderTable() {{
  const q = document.getElementById('q').value.trim().toLowerCase();
  const mark = document.getElementById('markFilter').value;
  const sel = document.getElementById('selFilter').value;
  const rows = DATA.filter(r => {{
    if (mark && r.mark !== mark) return false;
    if (sel && r.selected !== sel) return false;
    if (!q) return true;
    const blob = [r.district, r.plate, r.primary, r.middle, r.reason, r.pop].join(' ').toLowerCase();
    return blob.includes(q);
  }});
  document.getElementById('tbody').innerHTML = rows.map(r => `
    <tr>
      <td><span class="tag" data-m="${{r.mark}}">${{r.mark}}</span></td>
      <td>${{r.selected}}</td>
      <td>${{r.district}}</td>
      <td>${{r.plate}}</td>
      <td>${{r.primary}}</td>
      <td>${{r.middle}}</td>
      <td>${{r.competition}}</td>
      <td>${{r.hollow}}</td>
      <td class="score">${{r.score}}</td>
      <td>${{r.price}}</td>
      <td>${{r.reason}}</td>
    </tr>
  `).join('');
}}

document.getElementById('q').addEventListener('input', renderTable);
document.getElementById('markFilter').addEventListener('change', renderTable);
document.getElementById('selFilter').addEventListener('change', renderTable);
document.getElementById('resetBtn').addEventListener('click', () => {{
  document.getElementById('q').value = '';
  document.getElementById('markFilter').value = '';
  document.getElementById('selFilter').value = '';
  renderTable();
}});

renderPicks();
renderTable();
</script>
</body>
</html>
"""


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUT_XLSX)
    OUT_HTML.write_text(build_html(), encoding="utf-8")
    selected_n = sum(1 for r in PLATE_ROWS if r[-1] == "是")
    print(f"已生成: {OUT_XLSX}")
    print(f"已生成: {OUT_HTML}")
    print(f"板块样本: {len(PLATE_ROWS)} | 本轮入选: {selected_n}")


if __name__ == "__main__":
    main()
