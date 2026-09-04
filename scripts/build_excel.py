"""Generate the Greentown China 100,000 RMB sponsorship Excel workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FONT_NAME = "微软雅黑"

GREEN = "FF006B3F"
DARK_GREEN = "FF004A2C"
GOLD = "FFC8A25B"
LIGHT_GREEN = "FFD8E8DF"
LIGHT_GREY = "FFF2F2F2"
WHITE = "FFFFFFFF"
DARK = "FF1F1F1F"
RED = "FFC0392B"

thin = Side(border_style="thin", color="FFBFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, fill=GREEN, color=WHITE, bold=True, size=11):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=FONT_NAME, color=color, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_body(cell, fill=None, color=DARK, bold=False, size=10, align_h="left"):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=FONT_NAME, color=color, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)
    cell.border = BORDER


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title(ws, title, subtitle, n_cols):
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = PatternFill("solid", fgColor=DARK_GREEN)
    c.font = Font(name=FONT_NAME, color=WHITE, bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.fill = PatternFill("solid", fgColor=GOLD)
    c2.font = Font(name=FONT_NAME, color=WHITE, bold=True, size=11)
    c2.alignment = Alignment(horizontal="center", vertical="center")


def write_table(ws, start_row, headers, rows, fills_by_col=None,
                status_col=None, status_map=None):
    for c_idx, h in enumerate(headers, 1):
        style_header(ws.cell(row=start_row, column=c_idx, value=h))
    ws.row_dimensions[start_row].height = 30

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            fill = LIGHT_GREY if r_idx % 2 == 0 else WHITE
            align = "center" if (status_col and c_idx == status_col) else "left"
            color = DARK
            bold = False
            if status_col and c_idx == status_col and status_map:
                s_fill, s_color = status_map.get(val, (fill, DARK))
                style_body(cell, fill=s_fill, color=s_color, align_h="center", bold=True)
            else:
                style_body(cell, fill=fill, color=color, align_h=align)
        ws.row_dimensions[start_row + r_idx].height = 32


def sheet_overview(wb):
    ws = wb.create_sheet("1.方案总览")
    set_col_widths(ws, [4, 22, 60, 18, 14])
    add_title(
        ws,
        "绿城中国 | 绿城·潮鸣外滩 · 晚宴冠名战略合作伙伴",
        "10 万元赞助专项设计方案 · 总览（2026 AI 商业化峰会）",
        n_cols=5,
    )
    rows = [
        ["1", "赞助方", "绿城中国控股有限公司", "—", "—"],
        ["2", "项目品牌", "绿城·潮鸣外滩", "—", "—"],
        ["3", "合作身份", "晚宴冠名战略合作伙伴 Dinner Title Strategic Partner", "—", "—"],
        ["4", "赞助金额", "人民币 100,000 元（壹拾万元整）", "—", "—"],
        ["5", "合作峰会", "重构与突围 · 2026 AI 商业化落地与硬核投资破局峰会", "—", "—"],
        ["6", "主办方", "北京大学经济学院上海校友会 · 复旦大学住房政策研究中心", "—", "—"],
        ["7", "大会时间", "2026 年 5 月 22 日", "—", "—"],
        ["8", "大会地点", "上海·北外滩·一滴水", "—", "—"],
        ["9", "晚宴规模", "500+ 高净值嘉宾（25% 上市公司高管 / 20% 基金合伙人 / 25% AI 创始团队 / 20% 双校核心校友 / 10% 政府媒体）", "—", "—"],
        ["10", "晚宴时间", "18:10–20:30（含项目专场宣讲 18:15–18:30）", "—", "—"],
        ["11", "项目到访", "尊界 8–10 辆 + 考斯特补位；建议采用「方案 A · 20:30 会后专场」", "—", "—"],
        ["12", "付款方式", "合同签订后 5 个工作日内一次性付款", "—", "—"],
        ["13", "发票内容", "会议服务费 / 赞助费（增值税普票或专票）", "—", "—"],
        ["14", "执行回执", "大会结束后 7 个自然日内主办出具《赞助权益执行回执》", "—", "—"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "维度", "内容", "状态", "负责人"],
        rows=rows,
        status_col=4,
        status_map={
            "—": (LIGHT_GREY, DARK),
        },
    )


def sheet_rights(wb):
    ws = wb.create_sheet("2.权益清单")
    set_col_widths(ws, [4, 16, 14, 38, 30, 14, 14, 12])
    add_title(
        ws,
        "10 万元晚宴冠名 · 权益清单 Rights List",
        "对应钻石 5 万基础叠加「晚宴冠名 + 项目专场 + 接驳到访」",
        n_cols=8,
    )
    rows = [
        ["1", "A 晚宴冠名", "冠名权", "「晚宴冠名战略合作伙伴 绿城中国 | 绿城·潮鸣外滩」全程统称权", "晚宴 KV / 桌卡 / 菜单 / 席卡 logo 植入", "主办+活动公司", "Aiden", "待执行"],
            ["2", "A 晚宴冠名", "口播", "主持人开场 + 川总宣讲引荐 + 散场再次鸣谢 共 3 段", "口播稿（每段 30s）", "主办+绿城", "wangsheng", "待执行"],
            ["3", "B 主会场植入", "主背景板 logo", "钻石级 logo 位 + 「晚宴冠名」副标题位", "主背景板设计稿", "主办", "wangsheng", "待执行"],
        ["4", "B 主会场植入", "议程手册广告", "整版 + 项目折页夹页", "议程手册印刷文件", "印刷单位", "印刷负责人", "待执行"],
        ["5", "B 主会场植入", "宣传片轮播", "嘉宾入场及中场休息时段大屏 16:9 轮播", "30–60s 项目宣传片 mp4", "主办", "wangsheng", "待执行"],
        ["6", "B 主会场植入", "手拎袋折页", "500 份手拎袋植入项目折页 + 285/310 户型图", "三折页 + 户型图（印刷）", "印刷单位+活动公司", "印刷负责人", "待执行"],
        ["7", "B 主会场植入", "白皮书署名", "扉页联合署名「晚宴冠名战略合作伙伴」", "白皮书设计稿", "主办", "wangsheng", "待执行"],
        ["8", "C 现场展位", "电梯口展位", "品牌桌台 + 易拉宝×2 + 户型 KT 板 + 销售 2 人", "易拉宝/折页/KT 板", "绿城+活动公司", "销售负责人", "待执行"],
        ["9", "C 现场展位", "展场内展位", "洽谈圆桌 + 易拉宝×2 + 户型 KT 板 + 销售 2 人", "易拉宝/折页/KT 板", "绿城+活动公司", "销售负责人", "待执行"],
        ["10", "C 现场展位", "展位包装", "桌台轻包装 / 背景挂幔（时间允许升级 KT 板背景墙）", "展位包装方案", "活动公司", "活动公司对接人", "待执行"],
        ["11", "D 项目专场", "川总宣讲 15min", "晚宴开场前 15 分钟项目专场 PPT 演讲", "1080P 16:9 PPT + Q&A 提纲", "绿城", "川总团队", "待执行"],
        ["12", "D 项目专场", "音响 & 话筒", "话筒×2（手持/领夹）+ 切换器调试", "舞台设备", "主办+活动公司", "wangsheng", "待执行"],
        ["13", "E 接驳到访", "尊界车队", "华为尊界 8–10 辆，会后专场接驳意向嘉宾", "车队 + 引导员", "绿城", "项目接待负责人", "待执行"],
        ["14", "E 接驳到访", "考斯特补位", "意向客户超出尊界容量时补位", "考斯特", "绿城", "项目接待负责人", "待执行"],
        ["15", "E 接驳到访", "案场接待", "项目售楼处沙盘 + 样板间 + 灯光秀", "案场接待方案", "绿城", "项目接待负责人", "待执行"],
        ["16", "F 论坛后口播", "主持人口播", "圆桌结束 / 颁奖结束后口播 + 引导动线", "口播稿 + 动线图", "主办", "wangsheng", "待执行"],
        ["17", "G 媒体宣发", "媒体通稿", "≥ 5 家头部财经/地产媒体，标题级或副标题级", "通稿正文 + 媒体名单", "主办媒体组", "组委会媒体组", "待执行"],
        ["18", "G 媒体宣发", "朋友圈九宫格", "9 张视觉，至少 3 张含项目 logo/KV", "九宫格设计稿", "主办媒体组", "组委会媒体组", "待执行"],
        ["19", "G 媒体宣发", "回顾视频", "片头 3s 鸣谢 + 片尾 logo 墙 + 川总宣讲画面", "回顾视频", "主办", "组委会媒体组", "待执行"],
        ["20", "H 长效圈层", "校友产业联盟入册", "双校长三角校友产业联盟战略合作伙伴永久入册 + 牌匾交接", "联盟通讯录 + 牌匾", "主办", "组委会", "待执行"],
        ["21", "H 长效圈层", "执行回执", "大会结束后 7 日内出具《赞助权益执行回执》", "现场图片 + 媒体链接 + 嘉宾合影", "主办", "组委会", "待执行"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "模块", "权益项", "权益描述", "对应物料 / 产出", "责任方", "对接人", "状态"],
        rows=rows,
        status_col=8,
        status_map={
            "待执行": ("FFFFE9B6", "FF8A6D00"),
            "进行中": ("FFCDE5FF", "FF0A3E78"),
            "已完成": ("FFD8E8DF", DARK_GREEN),
            "—": (LIGHT_GREY, DARK),
        },
    )


def sheet_materials(wb):
    ws = wb.create_sheet("3.物料尺寸表")
    set_col_widths(ws, [4, 22, 14, 24, 18, 14, 22, 12])
    add_title(
        ws,
        "印刷 & 视频 & PPT 物料尺寸标准",
        "晚宴像素比待场地方最终确认；主会场已确认 16:9",
        n_cols=8,
    )
    rows = [
        ["1", "主会场宣传片（轮播）", "16:9", "1920×1080（4K 备份 3840×2160）", "≤ 60s", "25/30 fps", "H.264 .mp4，≤ 200MB", "已确认"],
        ["2", "晚宴 LED 宣传片", "16:9（条屏另出）", "1920×1080 或 6144×1080", "30–60s 循环", "25/30 fps", "请场地方提供像素比", "待确认"],
        ["3", "川总专场宣讲 PPT", "16:9", "1920×1080（40×22.5cm）", "≤ 25 张", "—", "封面/封底/沙盘 3D 预留", "待制作"],
        ["4", "晚宴 KV 主视觉", "依场地像素比", "先 1920×1080 主稿，再延展", "—", "—", "RGB；提供 AI 源 + PNG", "待确认"],
        ["5", "项目 logo（矢量）", "矢量", "AI / EPS / SVG（无白底）", "—", "—", "另出 PNG 透明底 4 套尺寸", "待提供"],
        ["6", "项目品牌简介", "—", "≤ 200 字", "—", "—", "txt / docx", "待提供"],
        ["7", "户型折页（285/310）", "印刷", "297×210mm 三折页（出血 3mm）", "—", "300 dpi", "CMYK；活动公司排版", "待制作"],
        ["8", "项目易拉宝", "印刷", "800×2000mm（出血 5mm）", "—", "150 dpi", "CMYK；2 处展位共 4 个", "待制作"],
        ["9", "席卡", "印刷", "90×55mm 双面", "—", "300 dpi", "CMYK；主办统一加 logo + slogan", "待制作"],
        ["10", "桌卡（桌号牌）", "印刷", "A5 双面折立 148×210mm；或亚克力 200×150mm", "—", "300 dpi", "CMYK；活动公司印刷", "待制作"],
        ["11", "菜单", "印刷", "对开 A4 210×285mm，或单页 285×210mm", "—", "300 dpi", "CMYK；活动公司印刷", "待制作"],
        ["12", "桌花卡 / 餐巾条（可选）", "印刷", "150×30mm 条形", "—", "300 dpi", "烫金字条", "可选"],
        ["13", "议程手册广告", "印刷", "整版 A4 210×285mm（出血 3mm）", "—", "300 dpi", "CMYK；夹页可加户型折页", "待制作"],
        ["14", "白皮书署名", "印刷", "扉页联合署名（位置以白皮书设计为准）", "—", "300 dpi", "CMYK；主办设计", "待制作"],
        ["15", "朋友圈九宫格", "传播", "1080×1080 PNG × 9", "—", "—", "媒体组发布；绿城同步素材", "待制作"],
        ["16", "回顾视频", "传播", "1920×1080 16:9 mp4", "≤ 5min", "25/30 fps", "片头 3s 鸣谢 + 片尾 logo 墙", "待制作"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "物料 / 素材", "比例", "推荐尺寸 / 分辨率", "时长", "帧率 / dpi", "备注", "状态"],
        rows=rows,
        status_col=8,
        status_map={
            "已确认": ("FFD8E8DF", DARK_GREEN),
            "待确认": ("FFFFE9B6", "FF8A6D00"),
            "待制作": ("FFFFE9B6", "FF8A6D00"),
            "待提供": ("FFFFE9B6", "FF8A6D00"),
            "可选": ("FFE6E6E6", DARK),
        },
    )


def sheet_timeline(wb):
    ws = wb.create_sheet("4.执行时间表")
    set_col_widths(ws, [4, 12, 22, 38, 36, 14, 12])
    add_title(
        ws,
        "执行时间表 · Timeline（设计制作仅 1 天，关键节点前置）",
        "5/18 合同确认 → 5/19 物料交付 → 5/20 设计确认 → 5/21 印刷 → 5/22 大会执行 → 5/29 宣发回执",
        n_cols=7,
    )
    rows = [
        ["1", "5/18", "确认合作框架及金额", "绿城：盖章合同回传", "主办：出具盖章合同 + 银行账户", "Aiden / 组委会", "待执行"],
        ["2", "5/19", "物料源文件交付（绿城 → 主办）", "logo（AI/EPS/PNG）· 项目宣传片 · 川总 PPT · 户型折页源文件 · 品牌简介 ≤ 200 字", "主办：晚宴 KV / 桌卡 / 菜单 / 席卡设计稿初稿", "Aiden / wangsheng", "待执行"],
        ["3", "5/19", "晚宴场地像素比确认", "—", "场地方提供晚宴 LED / 投影 像素比", "wangsheng", "待执行"],
        ["4", "5/19", "项目参观方案确认", "绿城确认到访方案 A/B/C", "主办协调议程时间", "Aiden / 组委会", "待执行"],
        ["5", "5/20", "设计稿确认", "确认一轮，提出修改意见", "主办：完成印刷物料终稿", "Aiden / 印刷负责人", "待执行"],
        ["6", "5/21", "印刷下单 & 物流", "—", "活动公司：桌卡、菜单、易拉宝、户型折页全部下印", "印刷负责人", "待执行"],
        ["7", "5/21", "尊界车队 & 案场接待预约", "绿城确认 8–10 辆尊界 + 考斯特补位 + 案场接待人员", "主办：现场引导员排期", "项目接待负责人", "待执行"],
        ["8", "5/22 上午", "现场布展", "销售 4 人到场对接展位包装；陈列易拉宝、KT 板、纸质物料", "活动公司：展位、KV、LED 调试到位", "销售负责人 / 活动公司", "待执行"],
        ["9", "5/22 13:30–17:55", "大会执行", "项目宣传片轮播；展位 1V1 沟通；手拎袋折页发放", "主办：议程执行；主持口播", "wangsheng", "待执行"],
        ["10", "5/22 17:55–18:10", "议程衔接 + 主持口播", "意向嘉宾名单对接", "主持人：感谢晚宴冠名战略合作伙伴 + 引导动线", "wangsheng", "待执行"],
        ["11", "5/22 18:10–20:30", "晚宴执行", "川总专场宣讲（18:15–18:30）；桌卡/菜单/KV 全场植入；宣传片循环", "主办：主持人口播 3 段；联合祝酒辞", "wangsheng / 川总", "待执行"],
        ["12", "5/22 20:30+", "项目参观（方案 A）", "尊界 8–10 辆 + 考斯特补位接送意向嘉宾至案场夜场参观（灯光秀+样板间）", "主办：现场引导员配合", "项目接待负责人", "待执行"],
        ["13", "5/23–5/29", "宣发 & 回执", "绿城自媒体矩阵同步传播", "主办：媒体通稿、九宫格、回顾视频；7 日内出执行回执", "组委会媒体组", "待执行"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "日期", "节点", "绿城方动作", "主办 / 活动公司动作", "对接人", "状态"],
        rows=rows,
        status_col=7,
        status_map={
            "待执行": ("FFFFE9B6", "FF8A6D00"),
            "进行中": ("FFCDE5FF", "FF0A3E78"),
            "已完成": ("FFD8E8DF", DARK_GREEN),
        },
    )


def sheet_dinner_rundown(wb):
    ws = wb.create_sheet("5.晚宴流程脚本")
    set_col_widths(ws, [4, 14, 26, 50, 22, 22])
    add_title(
        ws,
        "晚宴现场流程脚本 · Dinner Run-of-Show",
        "18:10–20:30（含川总专场宣讲 15min + 项目宣传片轮播）",
        n_cols=6,
    )
    rows = [
        ["1", "18:00–18:10", "嘉宾入场", "晚宴 KV 点亮；席卡引位；项目宣传片轮播（低音乐）", "LED 16:9 / BGM", "活动公司"],
        ["2", "18:10–18:15", "主持人开场 + 冠名鸣谢", "口播「晚宴冠名战略合作伙伴 绿城中国 | 绿城·潮鸣外滩」", "话筒×2", "主办主持人"],
        ["3", "18:15–18:30", "川总·绿城·潮鸣外滩专场宣讲", "项目 PPT 15min + Q&A 5min；强调 285/310 户型", "PPT 1080P / 话筒 / 切换器", "川总 / 活动公司"],
        ["4", "18:30–18:35", "联合祝酒辞", "主办 + 绿城联合致酒辞", "话筒×2", "主办 / 绿城"],
        ["5", "18:35–19:30", "正餐第一轮", "项目宣传片循环；菜单/桌卡 logo 植入", "宣传片循环 / BGM", "活动公司"],
        ["6", "19:30–20:00", "圈层社交 + 校友联盟介绍", "联盟介绍片；绿城战略合作伙伴永久入册牌匾交接", "LED / 摄影", "主办 / 绿城"],
        ["7", "20:00–20:25", "正餐第二轮 / 自由交流", "销售场内 1V1 沟通；预约项目到访", "BGM", "绿城销售"],
        ["8", "20:25–20:30", "结束鸣谢 & 引导到访", "再次口播鸣谢；引导意向嘉宾至尊界车队（方案 A）", "话筒×1 / 现场引导员", "主办主持人 / 绿城"],
        ["9", "20:30+", "项目专场参观（方案 A）", "尊界接驳至案场夜场参观（灯光秀 + 样板间 + 沙盘讲解）", "尊界 8–10 辆 + 考斯特", "绿城项目接待"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "时间", "环节", "内容 / 物料", "音视频需求", "执行方"],
        rows=rows,
    )


def sheet_contacts(wb):
    ws = wb.create_sheet("6.对接人")
    set_col_widths(ws, [4, 20, 22, 18, 18, 36])
    add_title(
        ws,
        "对接人矩阵 · Contact Matrix",
        "请双方在 5/19 前补充电话/邮箱/微信",
        n_cols=6,
    )
    rows = [
        ["1", "绿城方·总对接", "整体方案 / 付款 / 合同", "Aiden", "—", "合作框架内植入物料整体把控"],
        ["2", "绿城方·项目宣讲", "川总专场 PPT / 品牌口径", "川总团队", "—", "PPT 终稿确认 + 现场宣讲"],
        ["3", "绿城方·现场销售", "电梯口 & 展场内 2 处展位", "销售负责人", "—", "销售 4 人 + 物料统一"],
        ["4", "绿城方·接待 & 接驳", "尊界车队 / 案场接待", "项目接待负责人", "—", "8–10 辆尊界 + 考斯特补位 + 案场接待"],
        ["5", "主办·组委会招商", "权益落地 / 合同执行", "组委会招商组", "13262607888", "微信 13262607888"],
        ["6", "主办·现场执行", "场地 / KV / 印刷 / 流程", "wangsheng", "—", "晚宴 KV 像素比 & 印刷物料植入对接"],
        ["7", "主办·媒体组", "通稿 / 九宫格 / 回顾视频", "组委会媒体组", "—", "口径与绿城品牌部对齐"],
        ["8", "活动公司·印刷", "桌卡 / 菜单 / 手册物料 logo 植入", "印刷负责人", "—", "5/21 前完成印刷"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "角色", "对接事项", "对接人", "电话 / 微信", "备注"],
        rows=rows,
    )


def sheet_budget(wb):
    ws = wb.create_sheet("7.投入产出")
    set_col_widths(ws, [4, 24, 14, 14, 14, 30])
    add_title(
        ws,
        "投入产出测算 · ROI",
        "赞助费 100,000 元；覆盖 500+ 高净值嘉宾",
        n_cols=6,
    )
    rows = [
        ["1", "赞助费（现金）", "项", 1, 100000, "支付至主办指定账户；增值税普票/专票"],
        ["2", "项目印刷物料（折页/易拉宝/KT 板/手拎袋折页）", "套", 1, 0, "已含在赞助费内；由印刷单位统一制作"],
        ["3", "晚宴 KV / 桌卡 / 菜单 / 席卡 logo 植入", "套", 1, 0, "活动公司印刷植入；主办协调"],
        ["4", "项目宣传片（绿城自备）", "条", 1, 0, "绿城提供 mp4 1920×1080"],
        ["5", "川总专场宣讲 PPT（绿城自备）", "份", 1, 0, "绿城提供 16:9 1920×1080"],
        ["6", "尊界 8–10 辆 + 考斯特补位", "次", 1, 0, "绿城自备车队（不计入赞助费）"],
        ["7", "案场接待（夜场参观）", "次", 1, 0, "绿城案场资源（不计入赞助费）"],
        ["8", "现场销售 4 人", "人", 4, 0, "绿城派驻（不计入赞助费）"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "项目", "单位", "数量", "金额（元）", "备注"],
        rows=rows,
    )
    total_row = 4 + len(rows) + 1
    ws.cell(row=total_row, column=1, value="")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    c = ws.cell(row=total_row, column=1, value="合计（赞助费）")
    style_body(c, fill=GREEN, color=WHITE, bold=True, align_h="center", size=11)
    c2 = ws.cell(row=total_row, column=5, value=100000)
    style_body(c2, fill=GREEN, color=WHITE, bold=True, align_h="center", size=11)
    c2.number_format = '"¥"#,##0'
    c3 = ws.cell(row=total_row, column=6, value="人民币壹拾万元整")
    style_body(c3, fill=GREEN, color=WHITE, bold=True, align_h="center", size=11)

    # Output side
    out_row = total_row + 3
    ws.cell(row=out_row, column=1, value="")
    ws.merge_cells(start_row=out_row, start_column=1, end_row=out_row, end_column=6)
    h = ws.cell(row=out_row, column=1, value="预期产出 Expected Output")
    style_body(h, fill=DARK_GREEN, color=WHITE, bold=True, align_h="center", size=12)

    out_rows = [
        ["1", "现场直接触达高净值嘉宾", "人", 500, "—", "500+ 嘉宾全场曝光"],
        ["2", "意向客户登记", "组", 80, "—", "保守口径，电梯口 + 展场内 + 晚宴"],
        ["3", "项目到访意向客户（方案 A 夜场）", "组", 30, "—", "区间 30–60 组，视意向名单确认"],
        ["4", "媒体通稿覆盖", "篇", 5, "—", "主流财经 / 地产媒体"],
        ["5", "媒体二次曝光（社交+门户）", "次", 500000, "—", "保守口径 50w+ 次曝光"],
        ["6", "朋友圈九宫格触达", "人次", 100000, "—", "双校友 + 嘉宾社交圈"],
        ["7", "长效圈层入册", "联盟", 1, "—", "双校长三角校友产业联盟战略合作伙伴永久入册"],
    ]
    write_table(
        ws, start_row=out_row + 1,
        headers=["序号", "产出指标", "单位", "数量", "金额", "备注"],
        rows=out_rows,
    )


def sheet_risks(wb):
    ws = wb.create_sheet("8.风险与待确认")
    set_col_widths(ws, [4, 26, 42, 32, 14, 14])
    add_title(
        ws,
        "风险点 & 待确认事项 · Risks & Pending Items",
        "请双方在 5/19 前逐项闭环",
        n_cols=6,
    )
    rows = [
        ["1", "晚宴 KV 像素比未定", "请场地方提供晚宴 LED / 投影准确像素比（首选 1920×1080；条屏 6144×1080）", "主办 wangsheng 协调场地方反馈", "高", "待确认"],
        ["2", "项目参观时间冲突", "现议程下颁奖 17:55 → 晚宴 18:10，无空余衔接时间，往返+参观 ≥ 60min", "建议采用方案 A · 20:30 会后专场参观", "高", "待确认"],
        ["3", "设计制作时间仅 1 天", "桌卡 / 菜单 / 易拉宝 / 折页须 5/21 前下印", "5/20 完成设计稿确认；5/21 印刷下单", "高", "待执行"],
        ["4", "印刷物料 logo 植入范围", "除晚宴桌卡/菜单外，议程手册、白皮书、手拎袋折页也需植入项目 logo", "Aiden 与活动公司/印刷单位再次确认范围", "中", "待确认"],
        ["5", "现场展位包装升级", "设计制作时间仅 1 天，展位包装可能偏弱", "保底易拉宝 + 项目纸质物料；时间允许升级 KT 板背景墙", "中", "待执行"],
        ["6", "现场展位数量", "总展位多少家？是否影响电梯口和展场内位置分配", "主办 wangsheng 反馈现场排位数", "中", "待确认"],
        ["7", "尊界车队 + 考斯特数量", "意向客户名单未定，影响车队规模", "提前 24h 引导反馈人数，调整考斯特补位数量", "中", "待执行"],
        ["8", "媒体口径", "通稿 / 九宫格文案需与绿城品牌部对齐", "媒体组与绿城品牌部对齐 1 轮", "中", "待执行"],
        ["9", "音响 & 话筒", "川总宣讲需 1080P 16:9 PPT + 话筒×2 + 切换器", "wangsheng 与活动公司确认舞台设备", "中", "待确认"],
        ["10", "晚宴入场券", "10 万级别预计 8 张（销售 4 + 川总 + Aiden + 备 2）", "Aiden 与主办最终核定", "低", "待确认"],
    ]
    write_table(
        ws, start_row=4,
        headers=["序号", "事项", "描述", "解决方案", "优先级", "状态"],
        rows=rows,
        status_col=6,
        status_map={
            "待确认": ("FFFFE9B6", "FF8A6D00"),
            "待执行": ("FFFFE9B6", "FF8A6D00"),
            "已完成": ("FFD8E8DF", DARK_GREEN),
            "已确认": ("FFD8E8DF", DARK_GREEN),
        },
    )


def build():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_overview(wb)
    sheet_rights(wb)
    sheet_materials(wb)
    sheet_timeline(wb)
    sheet_dinner_rundown(wb)
    sheet_contacts(wb)
    sheet_budget(wb)
    sheet_risks(wb)

    out = "/workspace/deliverables/绿城中国-潮鸣外滩-10万元晚宴冠名赞助专项设计方案.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    build()
