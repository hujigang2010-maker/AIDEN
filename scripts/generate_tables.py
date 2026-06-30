# -*- coding: utf-8 -*-
"""生成『群邦·领事会客厅』收费与测算表格（Excel，多工作表）。

用法：
    python3 scripts/generate_tables.py [输出路径.xlsx]
默认输出 output/群邦-领事会客厅-收费与测算表格.xlsx
"""
import sys
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import content as C

NAVY = "FF0B2A4A"
GOLD = "FFC9A24B"
LIGHT = "FFF1F4F8"
WHITE = "FFFFFFFF"

THIN = Side(style="thin", color="FFBFC8D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEAD_FILL = PatternFill("solid", fgColor=NAVY)
TOTAL_FILL = PatternFill("solid", fgColor=GOLD)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)

HEAD_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color=NAVY)
CELL_FONT = Font(name="Microsoft YaHei", size=10.5, color="FF1E2933")
NOTE_FONT = Font(name="Microsoft YaHei", size=10, italic=True, color="FF5B6B7B")
BOLD_CELL = Font(name="Microsoft YaHei", size=10.5, bold=True, color="FF1E2933")

WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_sheet(ws, tdata):
    headers = tdata["headers"]
    rows = tdata["rows"]
    ncol = len(headers)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tc = ws.cell(row=1, column=1, value=tdata["title"])
    tc.font = TITLE_FONT
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # 表头
    hrow = 3
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=hrow, column=ci, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[hrow].height = 26

    # 数据
    for ri, row in enumerate(rows):
        excel_r = hrow + 1 + ri
        is_total = (str(row[0]) == "合计")
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=excel_r, column=ci, value=val)
            c.font = BOLD_CELL if (is_total or ci == 1) else CELL_FONT
            c.alignment = LEFT if ci == 1 else WRAP
            c.border = BORDER
            if is_total:
                c.fill = TOTAL_FILL
            elif ri % 2 == 1:
                c.fill = ALT_FILL
        ws.row_dimensions[excel_r].height = 30

    # 说明
    note = tdata.get("note")
    if note:
        nrow = hrow + 1 + len(rows) + 1
        ws.merge_cells(start_row=nrow, start_column=1, end_row=nrow, end_column=ncol)
        nc = ws.cell(row=nrow, column=1, value="说明：" + note)
        nc.font = NOTE_FONT
        nc.alignment = LEFT
        ws.row_dimensions[nrow].height = 34

    # 列宽
    for ci in range(1, ncol + 1):
        col = get_column_letter(ci)
        maxlen = len(str(headers[ci - 1]))
        for row in rows:
            maxlen = max(maxlen, len(str(row[ci - 1])))
        # 中文按 ~2 宽估算
        ws.column_dimensions[col].width = min(46, max(12, maxlen * 1.9 + 2))

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def write_overview(ws):
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1,
                value=f"{C.PROJECT_NAME} · 收费与测算表（{C.VERSION}）")
    t.font = Font(name="Microsoft YaHei", size=16, bold=True, color=NAVY)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    st = ws.cell(row=2, column=1, value=C.PROJECT_SUBTITLE)
    st.font = Font(name="Microsoft YaHei", size=11, color="FF5B6B7B")

    rows = [
        ("工作表", "内容"),
        ("表一 活动定价", "领事专题活动（小/中/大型）产品与定价"),
        ("表二 国家会客厅", "国家会客厅冠名与会籍定价"),
        ("表三 出海与企业服务", "出海深度游与企业出海对接服务定价"),
        ("表四 会员体系", "会员与会籍分层及权益"),
        ("表五 三方分润", "三方分工与分润机制"),
        ("表六 路线图", "12个月落地路线图"),
        ("表七 收入测算", "第一年收入测算（示意）"),
    ]
    start = 4
    for ri, (a, b) in enumerate(rows):
        r = start + ri
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        if ri == 0:
            ca.fill = HEAD_FILL; ca.font = HEAD_FONT; ca.alignment = WRAP
            cb.fill = HEAD_FILL; cb.font = HEAD_FONT; cb.alignment = WRAP
        else:
            ca.font = BOLD_CELL; ca.alignment = LEFT
            cb.font = CELL_FONT; cb.alignment = LEFT
            fill = ALT_FILL if ri % 2 else PatternFill("solid", fgColor=WHITE)
            ca.fill = fill; cb.fill = fill
        ca.border = BORDER; cb.border = BORDER
        ws.row_dimensions[r].height = 24
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=start + len(rows) + 1, start_column=1,
                   end_row=start + len(rows) + 1, end_column=4)
    nc = ws.cell(row=start + len(rows) + 1, column=1,
                 value="说明：本表所有金额均为人民币，用于商业测算示意，最终以三方框架协议及单项目合同为准。")
    nc.font = NOTE_FONT; nc.alignment = LEFT


def build(path):
    wb = Workbook()
    write_overview(wb.active)
    wb.active.title = "总览"

    sheet_names = [
        "一·活动定价", "二·国家会客厅", "三·出海与企业服务",
        "四·会员体系", "五·三方分润", "六·路线图", "七·收入测算",
    ]
    for name, tdata in zip(sheet_names, C.ALL_TABLES):
        ws = wb.create_sheet(title=name)
        write_sheet(ws, tdata)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"已生成 Excel：{path}（{len(wb.sheetnames)} 个工作表）")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else \
        "output/群邦-领事会客厅-收费与测算表格.xlsx"
    build(out)
