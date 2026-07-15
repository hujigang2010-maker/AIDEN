# -*- coding: utf-8 -*-
"""广州黄埔区九佛TOD「全球自贸365街区」项目——统一(内容级合并)表格 Excel。

面向对方提供的表格内容:内容层面合并去重、表述详细,统一到唯一项目名称。
工作表:
  1. 项目概要      项目名称/提供方/概念/产业方向/服务范围/成果交付/报价合计
  2. 提资清单      合并去重后的详细清单(类别/资料名称/详细说明/优先级)
  3. 服务报价      四阶段报价 + 合计
  4. 服务模块      八大模块详细表述
  5. 附件一 出口TOP20
  6. 附件二 消费类出口TOP20
  7. 附件三 楼层功能布局
商务大气风格:深藏青+金色主题、横幅标题、分区表头、斑马纹、冻结窗格。
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

import content_unified as C
import content_cgc as CGC

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "deliverables")
FONT = "微软雅黑"
NAVY = "0F2A4A"
BLUE = "1F4E79"
GOLD = "C79A3B"
LIGHT = "EEF2F8"
LIGHTGOLD = "F6EEDA"
WHITE = "FFFFFF"
DARK = "222B35"

med = Side(style="thin", color="AAB4C0")
BORDER = Border(left=med, right=med, top=med, bottom=med)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def setup(ws, tab_color=GOLD):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.sheet_properties.tabColor = tab_color


def widths(ws, ws_widths):
    for j, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def banner(ws, ncols, title, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, size=15, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 34
    r = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(row=2, column=1, value=subtitle)
        c.font = Font(name=FONT, size=10, italic=True, color="5A5A5A")
        c.fill = PatternFill("solid", fgColor=LIGHTGOLD)
        c.alignment = CENTER
        ws.row_dimensions[2].height = 20
        r = 3
    return r


def header(ws, row, headers, fill=GOLD):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=FONT, size=10.5, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def drow(ws, row, values, center_cols=(), h=22, zebra=True):
    fill = PatternFill("solid", fgColor=LIGHT) if (zebra and row % 2 == 0) else None
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(name=FONT, size=10, color=DARK)
        c.alignment = CENTER if j in center_cols else LEFT
        c.border = BORDER
        if fill:
            c.fill = fill
    ws.row_dimensions[row].height = h


def total_row(ws, row, values, center_cols=()):
    for j, v in enumerate(values, 1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(name=FONT, size=10.5, bold=True, color=NAVY)
        c.fill = PatternFill("solid", fgColor=LIGHTGOLD)
        c.alignment = CENTER if j in center_cols else LEFT
        c.border = BORDER
    ws.row_dimensions[row].height = 24


def kv_row(ws, row, label, value, ncols=2):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name=FONT, size=10.5, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = CENTER
    c.border = BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    v = ws.cell(row=row, column=2, value=value)
    v.font = Font(name=FONT, size=10, color=DARK)
    v.alignment = LEFT
    for col in range(2, ncols + 1):
        ws.cell(row=row, column=col).border = BORDER


def note(ws, row, ncols, text, h=34):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=9.5, italic=True, color="666666")
    c.alignment = LEFT
    ws.row_dimensions[row].height = h


# ------------------------------------------------------------ 1 项目概要
def sheet_overview(wb):
    ws = wb.active
    ws.title = "项目概要"
    setup(ws)
    NC = 6
    widths(ws, [16, 20, 20, 20, 20, 20])
    r = banner(ws, NC, C.PROJECT_NAME + " · 项目概要",
               "提供方:" + C.PROVIDER_LINE + "　|　日期:" + C.DOC_DATE)
    rows = [
        ("项目名称", C.PROJECT_NAME),
        ("提供方", C.PROVIDER_LINE),
        ("提交对象", C.CLIENT),
        ("核心概念", "\n".join("• " + p for p in C.CONCEPT_POINTS)),
        ("产业方向", C.INDUSTRY_INTRO + "\n" +
         "\n".join(f"• {n}:{d}" for n, d in C.INDUSTRY_GROUPS)),
        ("下阶段重点", C.NEXT_STAGE_TEXT),
        ("服务范围", "、".join(n for n, _ in C.SERVICE_MODULES)),
        ("成果交付", "\n".join("• " + d for d in C.DELIVERABLES)),
        ("服务报价", f"四阶段合计 {C.QUOTATION_TOTAL} 万元(含税参考价,详见“服务报价”工作表)"),
    ]
    for label, value in rows:
        kv_row(ws, r, label, value, ncols=NC)
        nlines = str(value).count("\n") + 1 + sum(len(s) // 60 for s in str(value).split("\n"))
        ws.row_dimensions[r].height = max(24, min(nlines * 16 + 6, 170))
        r += 1


# ------------------------------------------------------------ 2 提资清单
def sheet_info(wb):
    ws = wb.create_sheet("提资清单")
    setup(ws)
    widths(ws, [7, 16, 30, 52, 9])
    r = banner(ws, 5, "提资清单(请委托方提供的资料)",
               "内容已合并去重并细化说明,可直接对外提供")
    header(ws, r, ["序号", "类别", "资料名称", "说明", "优先级"])
    r += 1
    start = r
    for i, (cat, name, desc, prio) in enumerate(C.INFO_REQUEST_ITEMS, 1):
        drow(ws, r, [i, cat, name, desc, prio], center_cols=(1, 5), h=34)
        r += 1
    note(ws, r, 5, "说明：" + "；".join(C.INFO_REQUEST_NOTES), h=48)
    ws.freeze_panes = "A" + str(start)


# ------------------------------------------------------------ 3 服务报价
def sheet_quote(wb):
    ws = wb.create_sheet("服务报价")
    setup(ws)
    widths(ws, [24, 50, 34, 12])
    r = banner(ws, 4, "策划服务报价", "含税参考价(人民币),最终以双方商务洽谈及合同为准")
    header(ws, r, ["工作阶段", "主要工作内容", "主要成果", "报价(万元)"])
    r += 1
    for (stage, work, output, fee) in C.QUOTATION_ITEMS:
        drow(ws, r, [stage, work, output, fee], center_cols=(4,), h=44)
        r += 1
    total_row(ws, r, ["合计", "", "", C.QUOTATION_TOTAL], center_cols=(4,))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2
    for i, nt in enumerate(C.QUOTATION_NOTES):
        note(ws, r, 4, f"{i+1}. {nt}", h=22)
        r += 1


# ------------------------------------------------------------ 4 服务模块
def sheet_modules(wb):
    ws = wb.create_sheet("服务模块")
    setup(ws)
    widths(ws, [6, 24, 78])
    r = banner(ws, 3, "策划服务八大模块(详细说明)")
    header(ws, r, ["序号", "服务模块", "工作内容与详细说明"])
    r += 1
    start = r
    for i, (name, desc) in enumerate(C.SERVICE_MODULES, 1):
        drow(ws, r, [i, name, desc], center_cols=(1,), h=50)
        r += 1
    ws.freeze_panes = "A" + str(start)


# ------------------------------------------------------------ 5-7 市场数据附件
def sheet_export(wb):
    ws = wb.create_sheet("附件一_出口TOP20")
    setup(ws, tab_color=BLUE)
    widths(ws, [6, 22, 22, 16, 46])
    r = banner(ws, 5, "附件一 · 2025年广州出口TOP20品类")
    header(ws, r, ["排名", "品类", "代表品牌", "广州口岸交易额", "核心说明"])
    r += 1
    start = r
    for row in CGC.EXPORT_TOP20:
        drow(ws, r, list(row), center_cols=(1, 4), h=22)
        r += 1
    note(ws, r, 5, "说明：" + CGC.EXPORT_TOP20_NOTE, h=40)
    ws.freeze_panes = "A" + str(start)


def sheet_consumer(wb):
    ws = wb.create_sheet("附件二_消费类TOP20")
    setup(ws, tab_color=BLUE)
    widths(ws, [6, 20, 12, 10, 26, 20])
    r = banner(ws, 6, "附件二 · 2025年前10月广州消费类出口20强（单位：亿元）")
    header(ws, r, ["排名", "品类", "出口额", "同比增速", "核心出口品牌", "核心市场"])
    r += 1
    start = r
    for row in CGC.CONSUMER_TOP20:
        drow(ws, r, list(row), center_cols=(1, 3, 4), h=22)
        r += 1
    note(ws, r, 6, "说明：" + CGC.CONSUMER_TOP20_NOTE, h=40)
    ws.freeze_panes = "A" + str(start)


def sheet_floor(wb):
    ws = wb.create_sheet("附件三_楼层功能布局")
    setup(ws, tab_color=BLUE)
    widths(ws, [16, 26, 42, 40])
    r = banner(ws, 4, "附件三 · 黄埔九佛TOD「全球自贸365街区」楼层功能布局")
    header(ws, r, ["区位楼层", "功能定位", "核心业态 / 服务", "数据支撑与参考案例"])
    r += 1
    for row in CGC.FLOOR_LAYOUT:
        drow(ws, r, list(row), center_cols=(1,), h=42)
        r += 1


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_info(wb)
    sheet_quote(wb)
    sheet_modules(wb)
    sheet_export(wb)
    sheet_consumer(wb)
    sheet_floor(wb)
    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, "广州黄埔九佛TOD全球自贸365街区_提资清单报价与市场数据.xlsx")
    wb.save(out)
    print("saved:", out)


if __name__ == "__main__":
    main()
