# -*- coding: utf-8 -*-
"""生成《下半年单场活动立项计划》配套明细表（Excel）。
含总览 + 每月排期 + 报价体系 + 招商标的 + 10 张单场 OA 立项表。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
import plan_data as D

NAVY = "2E1F47"
BLUE = "5B3E8E"
LTBLUE = "EAE3F5"
GOLD = "B0841A"
GREY = "F6F3FB"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT = "Microsoft YaHei"


def style_header(cell):
    cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_cell(cell, wrap=True, center=False, bold=False, fill=None, color="000000"):
    cell.font = Font(name=FONT, size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def title_row(ws, text, ncols, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, bold=True, color=WHITE, size=14)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    start = 2
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(row=2, column=1, value=sub)
        c.font = Font(name=FONT, italic=True, color="595959", size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18
        start = 3
    return start


def kv(ws, r, k, v, fill_k=LTBLUE, fill_v=WHITE, h=28):
    style_cell(ws.cell(row=r, column=1, value=k), bold=True, center=True, fill=fill_k)
    style_cell(ws.cell(row=r, column=2, value=v), fill=fill_v)
    ws.row_dimensions[r].height = h
    return r + 1


wb = openpyxl.Workbook()

# ================================================================== 1 总览
ws = wb.active
ws.title = "1-10场总览"
headers = ["立项编号", "月份", "拟定时间", "活动主题", "产业板块", "形式/规模", "场地建议", "档位", "报价(万元)"]
widths = [14, 8, 16, 32, 18, 14, 24, 6, 11]
hr = title_row(ws, "东方枢纽 × 复旦大学 × 上海市科技企业联合会  |  下半年 10 场单场立项总览",
               len(headers), "8–12 月每月 2 场；每场独立立项、独立报价、单独走 OA；报价单位：万元")
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for j, h in enumerate(headers, 1):
    style_header(ws.cell(row=hr, column=j, value=h))
r = hr + 1
for a in D.ACTIVITIES:
    tier_letter = a["tier"].split(" ")[0]
    vals = [a["oa_code"], a["month"], a["date"], a["title"], a["sector"],
            a["scale"], a["venue"], tier_letter, a["price"]]
    for j, v in enumerate(vals, 1):
        style_cell(ws.cell(row=r, column=j, value=v), center=j in (1, 2, 8, 9),
                   fill=GREY if r % 2 else WHITE, bold=(j == 9), color=GOLD if j == 9 else "000000")
    ws.row_dimensions[r].height = 36
    r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
style_cell(ws.cell(row=r, column=1, value="10 场合计（单场立项累加）"), center=True, bold=True, fill=LTBLUE)
style_cell(ws.cell(row=r, column=9, value=D.TOTAL_PRICE), center=True, bold=True, fill=LTBLUE, color=GOLD)
ws.freeze_panes = ws.cell(row=hr + 1, column=1)

# ================================================================== 2 月度排期
ws = wb.create_sheet("2-月度排期")
hr = title_row(ws, "月度排期 · 8–12 月每月 2 场", 4, "7 月不排场，预留 OA 与邀约准备；最早一场 8 月上旬")
for i, w in enumerate([10, 8, 40, 36], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for j, h in enumerate(["月份", "场次", "本月主题", "说明"], 1):
    style_header(ws.cell(row=hr, column=j, value=h))
r = hr + 1
for month, cnt, note in D.MONTH_PLAN:
    titles = "；".join(f"{a['no']}.{a['title']}" for a in D.ACTIVITIES if a["month"] == month)
    style_cell(ws.cell(row=r, column=1, value=month), center=True, bold=True, fill=LTBLUE)
    style_cell(ws.cell(row=r, column=2, value=cnt), center=True, fill=GREY if r % 2 else WHITE)
    style_cell(ws.cell(row=r, column=3, value=titles), fill=GREY if r % 2 else WHITE)
    style_cell(ws.cell(row=r, column=4, value=note), fill=GREY if r % 2 else WHITE)
    ws.row_dimensions[r].height = 40
    r += 1

# ================================================================== 3 报价体系
ws = wb.create_sheet("3-报价体系")
hr = title_row(ws, "报价体系 · 两档模型与构成明细（单位：万元）", 3,
               "单场独立报价；上不封顶原则下的合理基准价，以邀约名单/规模确认后核定")
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 24
headers = ["费用构成项"] + list(D.TIERS.keys())
for j, h in enumerate(headers, 1):
    style_header(ws.cell(row=hr, column=j, value=h))
r = hr + 1
for idx, item in enumerate(D.COST_ITEMS):
    style_cell(ws.cell(row=r, column=1, value=item), fill=LTBLUE, bold=True)
    for j, k in enumerate(D.TIERS.keys(), 2):
        style_cell(ws.cell(row=r, column=j, value=D.TIERS[k][idx]), center=True,
                   fill=GREY if r % 2 else WHITE)
    r += 1
style_cell(ws.cell(row=r, column=1, value="单场合计"), bold=True, fill=GOLD, color=WHITE)
for j, k in enumerate(D.TIERS.keys(), 2):
    style_cell(ws.cell(row=r, column=j, value=D.TIER_TOTAL[k]), center=True, bold=True,
               fill=GOLD, color=WHITE)
r += 2
ws.cell(row=r, column=1, value="全年场次与预算汇总").font = Font(name=FONT, bold=True, size=12, color=NAVY)
r += 1
for j, h in enumerate(["档位", "场次", "单价(万元)", "小计(万元)"], 1):
    style_header(ws.cell(row=r, column=j, value=h))
r += 1
tier_counts = {}
for a in D.ACTIVITIES:
    tier_counts[a["tier"]] = tier_counts.get(a["tier"], 0) + 1
for k in D.TIERS.keys():
    cnt = tier_counts.get(k, 0)
    style_cell(ws.cell(row=r, column=1, value=k), fill=GREY)
    style_cell(ws.cell(row=r, column=2, value=cnt), center=True, fill=GREY)
    style_cell(ws.cell(row=r, column=3, value=D.TIER_TOTAL[k]), center=True, fill=GREY)
    style_cell(ws.cell(row=r, column=4, value=round(cnt * D.TIER_TOTAL[k], 1)), center=True, bold=True, fill=GREY)
    r += 1
style_cell(ws.cell(row=r, column=1, value="合计"), bold=True, fill=NAVY, color=WHITE)
style_cell(ws.cell(row=r, column=2, value=D.TOTAL_COUNT), center=True, bold=True, fill=NAVY, color=WHITE)
style_cell(ws.cell(row=r, column=3, value="—"), center=True, fill=NAVY, color=WHITE)
style_cell(ws.cell(row=r, column=4, value=D.TOTAL_PRICE), center=True, bold=True, fill=NAVY, color=WHITE)

# ================================================================== 4 招商标的与资源
ws = wb.create_sheet("4-招商标的与资源")
hr = title_row(ws, "招商标的 · 合作资源 · 执行机制", 2)
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 72
ws.cell(row=hr, column=1, value="〇、招商标的：" + D.PROJECT["name"]).font = Font(name=FONT, bold=True, size=12, color=NAVY)
r = hr + 1
r = kv(ws, r, "项目定位", D.PROJECT["position"], h=32)
r = kv(ws, r, "体量规模", D.PROJECT["area"] + "（非“133 万方”）", fill_v=WHITE, h=28)
style_cell(ws.cell(row=r, column=2), color=GOLD, bold=True)
r = kv(ws, r, "四大产品线", "\n".join(f"• {n}：{d}" for n, d in D.PROJECT["product_lines"]), h=90)
r = kv(ws, r, "销售/租赁模式", "\n".join("• " + x for x in D.PROJECT["model"]), h=66)
r = kv(ws, r, "活动↔产品匹配", D.PROJECT["match"], fill_k=GOLD, h=32)
# fix gold text on match key
ws.cell(row=r - 1, column=1).font = Font(name=FONT, size=10, bold=True, color=WHITE)
r += 1
ws.cell(row=r, column=1, value="一、三方合作定位").font = Font(name=FONT, bold=True, size=12, color=NAVY)
r += 1
for name, role, desc in D.PARTIES:
    r = kv(ws, r, f"{name}\n（{role}）", desc, h=48)
r = kv(ws, r, "备选/补充科技组织", "、".join(D.ALT_TECH_ORGS), h=28)
r += 1
ws.cell(row=r, column=1, value="二、政府资源背书矩阵（市级 + 浦东新区）").font = Font(name=FONT, bold=True, size=12, color=NAVY)
r += 1
for group, items in D.GOV_RESOURCES.items():
    style_cell(ws.cell(row=r, column=1, value=group), bold=True, center=True, fill=BLUE, color=WHITE)
    style_cell(ws.cell(row=r, column=2, value="\n".join("• " + x for x in items)), fill=GREY if r % 2 else WHITE)
    ws.row_dimensions[r].height = 58
    r += 1
style_cell(ws.cell(row=r, column=1, value="联动逻辑"), bold=True, center=True, fill=GOLD, color=WHITE)
style_cell(ws.cell(row=r, column=2, value=D.GOV_TAGLINE), fill=WHITE)
r += 2
ws.cell(row=r, column=1, value="三、单场 OA 执行机制").font = Font(name=FONT, bold=True, size=12, color=NAVY)
r += 1
for j, h in enumerate(["环节", "要点"], 1):
    style_header(ws.cell(row=r, column=j, value=h))
r += 1
for name, desc in D.EXECUTION:
    style_cell(ws.cell(row=r, column=1, value=name), bold=True, center=True, fill=LTBLUE)
    style_cell(ws.cell(row=r, column=2, value=desc), fill=GREY if r % 2 else WHITE)
    ws.row_dimensions[r].height = 30
    r += 1

# ================================================================== 5–14 单场立项表
for a in D.ACTIVITIES:
    name = f"立项{a['no']:02d}-{a['month']}"
    ws = wb.create_sheet(name[:31])
    hr = title_row(ws, f"单场活动立项表  |  {a['oa_code']}", 2,
                   "本表可单独提交 OA：含主题、时间、规模、邀约方向、预算明细")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 78
    r = hr
    r = kv(ws, r, "立项编号", a["oa_code"], fill_k=NAVY)
    ws.cell(row=r - 1, column=1).font = Font(name=FONT, size=10, bold=True, color=WHITE)
    r = kv(ws, r, "活动主题", a["title"], h=30)
    r = kv(ws, r, "拟定时间", a["date"])
    r = kv(ws, r, "所属月份", a["month"])
    r = kv(ws, r, "产业板块", a["sector"])
    r = kv(ws, r, "形式 / 规模", a["scale"])
    r = kv(ws, r, "场地建议", a["venue"])
    r = kv(ws, r, "报价档位", a["tier"])
    r = kv(ws, r, "单场报价（万元）", a["price"], fill_k=GOLD, h=30)
    ws.cell(row=r - 1, column=1).font = Font(name=FONT, size=10, bold=True, color=WHITE)
    ws.cell(row=r - 1, column=2).font = Font(name=FONT, size=14, bold=True, color=GOLD)
    r += 1
    ws.cell(row=r, column=1, value="一、内容建议").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    r = kv(ws, r, "活动内容", "\n".join(f"• {x}" for x in a["content"]), h=70)
    r = kv(ws, r, "拟邀嘉宾/资源", "\n".join(f"• {x}" for x in a["guests"]), h=56)
    r = kv(ws, r, "招商衔接价值", "\n".join(f"• {x}" for x in a["invest"]), h=48)
    r = kv(ws, r, "立项说明", a["oa_note"], fill_k=LTBLUE, h=36)
    r += 1
    ws.cell(row=r, column=1, value="二、单场预算明细（万元）").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    style_header(ws.cell(row=r, column=1, value="费用构成项"))
    style_header(ws.cell(row=r, column=2, value="金额（万元）"))
    r += 1
    for item, val in zip(D.COST_ITEMS, a["costs"]):
        style_cell(ws.cell(row=r, column=1, value=item), fill=LTBLUE, bold=True)
        style_cell(ws.cell(row=r, column=2, value=val), center=True, fill=GREY if r % 2 else WHITE)
        r += 1
    style_cell(ws.cell(row=r, column=1, value="单场合计"), bold=True, fill=NAVY, color=WHITE)
    style_cell(ws.cell(row=r, column=2, value=a["price"]), center=True, bold=True, fill=NAVY, color=WHITE)
    r += 2
    ws.cell(row=r, column=1, value="三、OA 提交要件（勾选）").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    for item in ["☐ 活动主题与时间", "☐ 参会规模与场地", "☐ 邀约名单（企业/嘉宾）",
                 "☐ 参会人员背景（决策层）", "☐ 单场预算明细（本表第二节）", "☐ 指定策划供应商签约路径"]:
        style_cell(ws.cell(row=r, column=1, value="要件"), bold=True, center=True, fill=LTBLUE)
        style_cell(ws.cell(row=r, column=2, value=item), fill=WHITE)
        ws.row_dimensions[r].height = 24
        r += 1

# page setup
for ws in wb.worksheets:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

out = "东方枢纽三方合作计划_明细表.xlsx"
wb.save(out)
print(f"Excel 已生成：{out}  (工作表数: {len(wb.sheetnames)})")
print("Sheets:", wb.sheetnames)
