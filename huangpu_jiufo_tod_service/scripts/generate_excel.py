# -*- coding: utf-8 -*-
"""生成 Excel 工作簿:项目概要 / 提资清单 / 报价单 三个工作表。"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import content as C

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "deliverables")
CN_FONT = "微软雅黑"
ACCENT = "1F4E79"
LIGHT = "D9E2F3"

thin = Side(style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup_print(ws):
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def title_row(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=CN_FONT, size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = CENTER
    ws.row_dimensions[row].height = 30


def header_row(ws, headers, row):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=CN_FONT, size=11, bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def data_cell(ws, row, col, value, bold=False, center=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=CN_FONT, size=11, bold=bold)
    c.alignment = CENTER if center else WRAP
    c.border = BORDER
    return c


def set_widths(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def note_block(ws, row, ncols, title, notes):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name=CN_FONT, size=11, bold=True)
    for i, n in enumerate(notes, 1):
        ws.merge_cells(start_row=row + i, start_column=1,
                       end_row=row + i, end_column=ncols)
        c = ws.cell(row=row + i, column=1, value=f"{i}. {n}")
        c.font = Font(name=CN_FONT, size=10.5)
        c.alignment = WRAP
        ws.row_dimensions[row + i].height = 28


def sheet_overview(wb):
    ws = wb.active
    ws.title = "项目概要"
    setup_print(ws)
    set_widths(ws, [18, 90])
    title_row(ws, C.PROJECT_NAME + " 项目概要", 2)

    rows = [("项目名称", C.PROJECT_NAME),
            ("提供方", C.PROVIDER_LINE),
            ("提交对象", C.CLIENT),
            ("日期", C.DOC_DATE),
            ("核心概念", "\n".join(f"• {p}" for p in C.CONCEPT_POINTS)),
            ("产业方向", C.INDUSTRY_INTRO + "\n" + "\n".join(
                f"• {n}:{d}" for n, d in C.INDUSTRY_GROUPS)),
            ("下阶段重点", C.NEXT_STAGE_TEXT),
            ("服务范围", "\n".join(
                f"{i}. {n}" for i, (n, _) in enumerate(C.SERVICE_MODULES, 1))),
            ("成果交付", "\n".join(f"• {d}" for d in C.DELIVERABLES)),
            ("服务报价", f"合计 {C.QUOTATION_TOTAL} 万元(含税参考价,详见“报价单”工作表)")]
    r = 2
    # 列B宽90字符,约容纳44个全角汉字;按换行+折行估算行高
    chars_per_line = 44
    for label, value in rows:
        data_cell(ws, r, 1, label, bold=True, center=True)
        data_cell(ws, r, 2, value)
        nlines = sum(max(1, -(-len(line) // chars_per_line))
                     for line in str(value).split("\n"))
        ws.row_dimensions[r].height = max(22, min(nlines * 18 + 6, 260))
        r += 1


def sheet_info_request(wb):
    ws = wb.create_sheet("提资清单")
    setup_print(ws)
    set_widths(ws, [8, 16, 34, 44, 10])
    title_row(ws, "提资清单(请委托方提供的资料)", 5)
    header_row(ws, ["序号", "类别", "资料名称", "说明", "优先级"], 2)
    r = 3
    for i, (cat, name, desc, prio) in enumerate(C.INFO_REQUEST_ITEMS, 1):
        data_cell(ws, r, 1, i, center=True)
        data_cell(ws, r, 2, cat, center=True)
        data_cell(ws, r, 3, name)
        data_cell(ws, r, 4, desc)
        data_cell(ws, r, 5, prio, center=True)
        ws.row_dimensions[r].height = 30
        r += 1
    note_block(ws, r + 1, 5, "提资说明:", C.INFO_REQUEST_NOTES)


def sheet_quotation(wb):
    ws = wb.create_sheet("报价单")
    setup_print(ws)
    set_widths(ws, [26, 48, 34, 14])
    title_row(ws, "策划服务报价单", 4)
    header_row(ws, ["工作阶段", "主要工作内容", "主要成果", "报价(万元)"], 2)
    r = 3
    for stage, work, output, fee in C.QUOTATION_ITEMS:
        data_cell(ws, r, 1, stage)
        data_cell(ws, r, 2, work)
        data_cell(ws, r, 3, output)
        data_cell(ws, r, 4, fee, center=True)
        ws.row_dimensions[r].height = 40
        r += 1
    data_cell(ws, r, 1, "合计", bold=True, center=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for col in (2, 3):
        ws.cell(row=r, column=col).border = BORDER
    data_cell(ws, r, 4, C.QUOTATION_TOTAL, bold=True, center=True)
    ws.row_dimensions[r].height = 24
    note_block(ws, r + 2, 4, "报价说明:", C.QUOTATION_NOTES)


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_info_request(wb)
    sheet_quotation(wb)
    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, "广州黄埔九佛TOD全球自贸365街区_提资清单与报价.xlsx")
    wb.save(out)
    print("saved:", out)


if __name__ == "__main__":
    main()
