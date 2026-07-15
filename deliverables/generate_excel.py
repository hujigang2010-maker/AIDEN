# -*- coding: utf-8 -*-
"""
生成《夏春长三角合作方案 - 配套测算与清单》Excel
Sheets: 说明 / 语录解析 / 3个月过渡计划 / 活动内容矩阵 / 差异化对照 /
        分工分润 / 财务测算 / 风险与待办 / KPI与转长期安排
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

# ---- 配色 ----
NAVY = "142A4A"
NAVY2 = "1F3A5F"
GOLD = "C9A227"
GOLDL = "EBE2C4"
TEAL = "2E7D83"
LGREY = "EEF1F5"
CARD = "F6F8FB"
WHITE = "FFFFFF"
GREYTXT = "3C4450"

FONT = "Microsoft YaHei"

thin = Side(style="thin", color="D2DAE3")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, headers, fill=NAVY, height=30):
    ws.row_dimensions[row].height = height
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name=FONT, bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all


def write_row(ws, row, values, wrap=True, bold_cols=(), fills=None, size=10.5,
              aligns=None, height=None):
    if height:
        ws.row_dimensions[row].height = height
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(name=FONT, size=size, bold=(c in bold_cols),
                         color=(NAVY if c in bold_cols else GREYTXT))
        al = "left"
        if aligns and c - 1 < len(aligns):
            al = aligns[c - 1]
        cell.alignment = Alignment(horizontal=al, vertical="center", wrap_text=wrap)
        cell.border = border_all
        if fills and c in fills:
            cell.fill = PatternFill("solid", fgColor=fills[c])


def zebra(ws, first_row, last_row, ncols, color=CARD):
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = PatternFill("solid", fgColor=color)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=FONT, bold=True, size=16, color=WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = Font(name=FONT, size=10, color=NAVY2, italic=True)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c2.fill = PatternFill("solid", fgColor=GOLDL)
    ws.row_dimensions[2].height = 22


wb = Workbook()

# =====================================================
# Sheet 0 — 说明
# =====================================================
ws = wb.active
ws.title = "说明"
ws.sheet_view.showGridLines = False
set_widths(ws, [3, 30, 82])
title_block(ws, "夏春 · 长三角战略合作方案 — 配套测算与清单", "复旦大学住房政策研究中心 × 上海市杨浦区科技企业联合会 × 上海市科技企业联合会   |   2026年7月   |  与 PPT 方案配套使用", 3)
info = [
    ("目标", "作为“夏春财经智识”在长三角区域相关服务的独家 / 优先总代理。"),
    ("合作阶段", "长期深化合作 + 先行约 3 个月的适度过渡期（双方适应与适配）；过渡顺畅、无重大问题即自动转为长期。"),
    ("提报机构", "复旦大学住房政策研究中心；上海市杨浦区科技企业联合会（为主）；上海市科技企业联合会。三家联同开展服务。"),
    ("差异化", "只做长三角区域增量，区隔于夏春现有新 / 老团队，避免存量与客户冲突。"),
    ("内容内核", "以夏春公开访谈 / 公众号 · 视频号观点与双方前期沟通为依据（见“语录解析”）。"),
    ("", ""),
    ("工作表导航", ""),
    ("① 语录解析", "夏春核心语录 → 主题 → 解读 → 长三角合作启示 → 落地场景 → 优先级。"),
    ("② 3个月过渡计划", "过渡期逐月 / 逐周任务、责任方、交付物与里程碑（含状态下拉）。"),
    ("③ 活动内容矩阵", "六大落地场景的客群、频次、规模、参考客单价与毛利。"),
    ("④ 差异化对照", "老团队 / 新团队 / 我方（总代理）多维对照。"),
    ("⑤ 分工分润", "各环节夏春侧与我方职责、收入来源与建议分润比例。"),
    ("⑥ 财务测算", "过渡期三场活动的收入 / 成本 / 毛利 / 分润模型（公式联动，可改参数）。"),
    ("⑦ 风险与待办", "双方前期沟通中的待解决问题、应对措施、责任方与优先级。"),
    ("⑧ KPI与转长期安排", "过渡期与长期的关键参考指标，及过渡顺畅即自动转长期的安排。"),
    ("", ""),
    ("免责声明", "本方案所涉夏春观点摘自公开内容，仅供合作沟通参考，不构成任何投资建议。"),
]
r = 4
for k, v in info:
    ws.cell(row=r, column=2, value=k).font = Font(name=FONT, bold=True, size=11, color=NAVY)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="center", horizontal="left")
    ws.cell(row=r, column=3, value=v).font = Font(name=FONT, size=10.5, color=GREYTXT)
    ws.cell(row=r, column=3).alignment = Alignment(vertical="center", wrap_text=True)
    if k == "工作表导航":
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=GOLD)
        ws.cell(row=r, column=2).font = Font(name=FONT, bold=True, size=11, color=WHITE)
    ws.row_dimensions[r].height = 26
    r += 1

# =====================================================
# Sheet 1 — 语录解析（核心）
# =====================================================
ws = wb.create_sheet("语录解析")
ws.sheet_view.showGridLines = False
set_widths(ws, [5, 40, 15, 16, 34, 34, 8])
title_block(ws, "夏春先生核心观点（语录）解析", "摘自公开访谈 / 公众号 · 视频号 + 双方前期沟通，提炼为长三角合作切入点", 7)
hr = 3
style_header(ws, hr, ["序号", "夏春语录 / 观点原文", "来源", "核心主题", "观点解读", "长三角合作启示 / 落地场景", "优先级"])
quotes = [
    ("2025 年……我个人更倾向于悲观派……需要更加警惕股债市场可能出现的危险。", "访谈·宏观", "宏观判断 / 风控优先",
     "稳健、风控优先，契合长三角地产与高净值客群的避险焦虑。", "“不确定市场下的资产防御”主题闭门沙龙，作为区域首发爆款。", "高"),
    ("每年年底做一张全年资产表现表：前一年涨幅大的减配、跌幅大的增配，做再平衡。", "访谈·资产配置", "标志性再平衡方法论",
     "可复制、可 IP 化的年度框架，适合做长期年度旗舰内容。", "落地“长三角年度资产配置盘点”旗舰活动 + 会员专栏，年度 IP 化。", "高"),
    ("香港对虚拟资产的接受度……速度甚至比美国更快；现货 ETF 允许实物兑换，全球首创。", "访谈·香港/RWA", "香港虚拟资产 / RWA",
     "直接呼应双方共同关注的香港游学 RWA、中东资金窗口期。", "我方作长三角组团总入口，承接“半天讲座 + 半天参访”香港游学。", "高"),
    ("可将加密资产视为美股七巨头之外的‘第八个巨头’，与美股相关性最高约 0.5。", "访谈·另类配置", "行为金融 / 另类资产",
     "独到洞见，精准命中科创企业家与家办人群。", "面向杨浦科创企业家、家族办公室的另类配置主题专场。", "中"),
    ("我们更注重多元化分散、风险控制，寻找相对估值偏低但属同一投资逻辑的资产。", "访谈·方法论", "分散 / 估值方法论",
     "可迁移至不动产再配置，与复旦住房政策研究中心互补。", "联合复旦推出“住房 / 不动产 + 大类资产”跨界研讨与课题。", "中"),
    ("2025 年黄金价格可能突破 3000 美元……受益于央行与个人购金。", "访谈·黄金", "黄金 / 大宗",
     "热门且大众化议题，利于面向大众高净值获客与私域裂变。", "黄金与避险资产公开课 / 直播连麦，作私域拉新与转化钩子。", "中"),
    ("（前期沟通）香港游学需夏春先生全程参与；方案由夏春先生主导决策。", "前期沟通", "IP 稀缺 / 减负",
     "夏春个人时间是核心稀缺资源，须为其减负。", "我方承接全部执行、排期与履约，录播 / 图文复用降低时间投入。", "高"),
    ("（前期沟通）依托复旦 / 科技企业联合会背书突破企业制度限制，落地头部科技企业参访。", "前期沟通", "背书 / 科技参访",
     "我方多方背书正是突破点，与双方前期沟通重点高度契合。", "以复旦 + 市 / 区科技企业联合会背书，落地 AI / 科技前沿头部企业参访。", "高"),
]
r = hr + 1
for i, q in enumerate(quotes, 1):
    pr_fill = {7: (GOLD if q[5] == "高" else (TEAL if q[5] == "中" else LGREY))}
    write_row(ws, r, [i, q[0], q[1], q[2], q[3], q[4], q[5]],
              bold_cols=(1, 4), fills=pr_fill,
              aligns=["center", "left", "center", "left", "left", "left", "center"], height=64)
    if q[5] in ("高", "中"):
        ws.cell(row=r, column=7).font = Font(name=FONT, bold=True, color=WHITE, size=10.5)
    r += 1
zebra(ws, hr + 1, r - 1, 7)
ws.freeze_panes = "A4"

# =====================================================
# Sheet 2 — 3个月过渡计划
# =====================================================
ws = wb.create_sheet("3个月过渡计划")
ws.sheet_view.showGridLines = False
set_widths(ws, [10, 10, 42, 16, 30, 22, 10])
title_block(ws, "3 个月适度过渡期行动计划（双方适应与适配）", "逐月 / 逐周任务、责任方、交付物与里程碑；状态列可下拉更新", 7)
hr = 3
style_header(ws, hr, ["阶段", "周次", "关键任务", "责任方", "交付物", "里程碑", "状态"])
plan = [
    ("第1月\n启动共识", "W1", "对齐目标与边界，确认过渡期独家授权范围", "双方", "合作意向确认", "启动", "待开始"),
    ("第1月\n启动共识", "W2", "起草并会签合作备忘录（MOU）与区域独家授权草案", "我方主起草", "MOU + 授权草案", "★ 签署 MOU", "待开始"),
    ("第1月\n启动共识", "W3", "客户 / 会员画像盘点，明确客户归属与收益分配规则", "双方", "客户画像 & 归属规则", "", "待开始"),
    ("第1月\n启动共识", "W4", "首场活动选题立项（科技参访 / 资产沙龙）与预算", "我方", "首场策划案 & 预算", "首场立项", "待开始"),
    ("第2月\n首战落地", "W5", "招募 / 报名与场地、嘉宾（夏春关键出席）排期", "我方执行", "报名数据 & 排期表", "", "待开始"),
    ("第2月\n首战落地", "W6", "首场活动落地执行（科技参访或资产配置沙龙）", "双方", "活动交付 & 满意度", "★ 首场落地", "待开始"),
    ("第2月\n首战落地", "W7", "香港游学产品设计与预售启动", "我方", "游学方案 & 预售", "游学预售", "待开始"),
    ("第2月\n首战落地", "W8", "搭建私域与 CRM、会员分层与内容本地化专栏", "我方", "CRM & 私域 & 专栏", "", "待开始"),
    ("第3月\n复盘转长期", "W9", "第 2 场活动落地（另类配置 / 家办专场）", "双方", "活动交付", "第2场", "待开始"),
    ("第3月\n复盘转长期", "W10", "第 3 场活动落地并沉淀活动 SOP", "双方", "活动 SOP", "★ 第3场", "待开始"),
    ("第3月\n复盘转长期", "W11", "运作复盘、双方磨合与客户满意度分析", "我方", "复盘报告", "", "待开始"),
    ("第3月\n复盘转长期", "W12", "过渡顺畅、无重大问题，自动转长期并会签区域独家总代理协议", "双方", "长期协议", "★ 转长期", "待开始"),
]
r = hr + 1
phase_colors = {"第1月\n启动共识": TEAL, "第2月\n首战落地": GOLD, "第3月\n复盘转长期": NAVY2}
for p in plan:
    write_row(ws, r, list(p), bold_cols=(1,), aligns=["center", "center", "left", "center", "left", "center", "center"], height=34)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=phase_colors[p[0]])
    ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, color=WHITE, size=9.5)
    if p[5].startswith("★"):
        ws.cell(row=r, column=6).font = Font(name=FONT, bold=True, color=GOLD, size=10.5)
    r += 1
# 状态下拉
dv = DataValidation(type="list", formula1='"待开始,进行中,已完成,风险"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"G{hr+1}:G{r-1}")
ws.freeze_panes = "A4"

# =====================================================
# Sheet 3 — 活动内容矩阵
# =====================================================
ws = wb.create_sheet("活动内容矩阵")
ws.sheet_view.showGridLines = False
set_widths(ws, [5, 20, 30, 22, 12, 12, 14, 12, 22])
title_block(ws, "长三角落地场景与内容矩阵", "结合夏春观点与双方前期沟通；客单价 / 毛利为区间预估，需按实际招商调整", 9)
hr = 3
style_header(ws, hr, ["序号", "活动类型", "主题（结合夏春观点）", "目标客群", "频次", "单场人数", "参考客单价", "预估毛利率", "依托资源"])
matrix = [
    ("科技企业参访", "AI / 科技前沿 · 七巨头视角", "科创企业家 / 投资人", "月 1–2 场", "20–40", "0.5–3 万/人", "60%–90%", "市 / 区科技企业联合会 + 复旦背书"),
    ("资产配置闭门沙龙", "年度资产表 · 再平衡 · 美债美股", "高净值 / 私行客户", "月 1–2 场", "20–50", "0.2–1 万/人", "70%–90%", "夏春 IP + 复旦"),
    ("香港 / 出海游学", "RWA · 虚拟资产 · 中东资金窗口", "高净值 / 企业主", "季度 1–2 团", "15–30", "3–8 万/人", "30%–50%", "夏春全程 + 香港资源"),
    ("家办 / 另类配置专场", "第八个巨头 · 黄金 · 另类资产", "家办 / 超高净值", "季度 1 场", "10–20", "1–5 万/人", "60%–85%", "夏春 IP"),
    ("内容本地化 & 私域", "长三角专栏 · 会员分层运营", "泛高净值 / 粉丝", "持续运营", "—", "会员费 / 订阅", "80%+", "我方私域 + CRM"),
    ("企业内训 / 政企研讨", "宏观形势 · 住房与大类资产", "企业 / 政府 / 高校", "按需", "定制", "5–30 万/场", "50%–80%", "复旦住房政策中心"),
]
r = hr + 1
for i, m in enumerate(matrix, 1):
    write_row(ws, r, [i] + list(m), bold_cols=(2,),
              aligns=["center", "left", "left", "left", "center", "center", "center", "center", "left"], height=40)
    r += 1
zebra(ws, hr + 1, r - 1, 9)
ws.freeze_panes = "A4"

# =====================================================
# Sheet 4 — 差异化对照
# =====================================================
ws = wb.create_sheet("差异化对照")
ws.sheet_view.showGridLines = False
set_widths(ws, [16, 30, 30, 40])
title_block(ws, "差异化定位对照：区别于夏春新 / 老团队", "只做长三角区域增量，边界清晰、互补不重叠（明确主导权 / 客户归属）", 4)
hr = 3
style_header(ws, hr, ["对比维度", "夏春 · 老团队", "夏春 · 新团队", "我方（长三角总代理）"])
diff = [
    ("核心侧重", "内容生产、全国泛在影响", "金融产品与新业务孵化", "长三角区域落地与机构化承接"),
    ("目标客群", "全国粉丝 / 线上受众", "金融端存量客户", "长三角高净值 + 科创企业 + 高校 / 政府"),
    ("主要场景", "视频号 / 公众号内容", "金融配置与产品", "区域活动 / 参访 / 游学组团 / 内训"),
    ("背书资源", "夏春个人 IP", "资本 / 金融资源", "复旦 + 市 / 区科技企业联合会 背书 + 本地网络"),
    ("客户归属", "既有存量粉丝", "既有金融客户", "只做长三角新增增量，签不重叠条款"),
    ("与夏春关系", "内容协同", "产品协同", "区域独家总代理，统一出口与结算"),
    ("对夏春价值", "扩大影响力", "金融变现", "区域变现 + 减负 + 机构化背书"),
]
r = hr + 1
for d in diff:
    write_row(ws, r, list(d), bold_cols=(1,),
              aligns=["center", "left", "left", "left"], height=32,
              fills={4: GOLDL})
    ws.cell(row=r, column=4).font = Font(name=FONT, size=10.5, bold=True, color=NAVY)
    r += 1
ws.freeze_panes = "A4"

# =====================================================
# Sheet 5 — 分工分润
# =====================================================
ws = wb.create_sheet("分工分润")
ws.sheet_view.showGridLines = False
set_widths(ws, [22, 34, 34, 22, 14])
title_block(ws, "分工与分润机制（建议）", "分润比例为建议区间，最终以合作备忘录约定为准", 5)
hr = 3
style_header(ws, hr, ["合作环节", "夏春侧职责", "我方（总代理）职责", "收入来源", "建议分润\n（夏春:我方）"])
split = [
    ("区域线下活动", "核心观点 / 讲座、品牌授权", "获客、场地、执行、履约", "票务 / 赞助", "30% : 70%"),
    ("科技企业参访", "科技视角分享（可录播复用）", "企业对接、组织、接待", "票务 / 企业赞助", "25% : 75%"),
    ("香港 / 出海游学", "全程参与、讲座、关键资源", "组团、行程、履约、结算", "游学费 / 佣金", "40% : 60%"),
    ("内容本地化 & 私域", "内容授权 / 素材", "本地化、私域、会员运营", "会员费 / 订阅", "35% : 65%"),
    ("企业内训 / 政企", "宏观内容 / 出席", "复旦背书、对接、交付", "内训费", "35% : 65%"),
    ("品牌授权（长期）", "IP 与品牌授权", "区域独家运营", "年度授权费 / 分成", "另议（保底 + 分成）"),
]
r = hr + 1
for sdat in split:
    write_row(ws, r, list(sdat), bold_cols=(1,),
              aligns=["center", "left", "left", "center", "center"], height=38)
    ws.cell(row=r, column=5).font = Font(name=FONT, bold=True, color=NAVY, size=10.5)
    r += 1
zebra(ws, hr + 1, r - 1, 5)
note = ws.cell(row=r + 1, column=1,
               value="说明：成本相对固定，毛利率主要取决于客单价与销量；分润在扣除直接活动成本后的毛利基础上分配，按月对账、账目透明。")
ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
note.font = Font(name=FONT, size=9.5, italic=True, color=NAVY2)
note.alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[r + 1].height = 30
ws.freeze_panes = "A4"

# =====================================================
# Sheet 6 — 财务测算（公式联动）
# =====================================================
ws = wb.create_sheet("财务测算")
ws.sheet_view.showGridLines = False
set_widths(ws, [26, 16, 16, 16, 16])
title_block(ws, "过渡期财务测算模型（示例，参数可调）", "蓝底为可修改输入项；其余为公式自动计算。数字为示例假设，非承诺", 5)
hr = 3
style_header(ws, hr, ["项目 \\ 场次", "首场·科技参访", "第2场·资产沙龙", "第3场·家办专场", "合计 / 备注"])

# 输入行：人数、客单价、固定成本、单人变动成本、夏春分润比例
INPUT_FILL = PatternFill("solid", fgColor="DCE6F5")
rows_def = [
    ("报名人数（人）", 30, 40, 15, None, "输入"),
    ("客单价（元/人）", 8000, 5000, 30000, None, "输入"),
    ("固定成本（元/场）", 30000, 20000, 25000, None, "场地/嘉宾/物料"),
    ("单人变动成本（元/人）", 800, 500, 3000, None, "餐饮/资料/接待"),
    ("夏春侧分润比例", 0.30, 0.30, 0.40, None, "对毛利分润"),
]
r = hr + 1
row_index = {}
for name, a, b, c, _, memo in rows_def:
    ws.cell(row=r, column=1, value=name).font = Font(name=FONT, bold=True, size=10.5, color=NAVY)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
    ws.cell(row=r, column=1).border = border_all
    for ci, val in zip((2, 3, 4), (a, b, c)):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.fill = INPUT_FILL
        cell.border = border_all
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name=FONT, size=10.5, color="14315E", bold=True)
        if "比例" in name:
            cell.number_format = "0%"
        elif "人数" not in name:
            cell.number_format = "#,##0"
    memo_cell = ws.cell(row=r, column=5, value=memo)
    memo_cell.font = Font(name=FONT, size=9.5, italic=True, color=NAVY2)
    memo_cell.alignment = Alignment(vertical="center")
    memo_cell.border = border_all
    row_index[name] = r
    ws.row_dimensions[r].height = 24
    r += 1

rp = row_index["报名人数（人）"]
rprice = row_index["客单价（元/人）"]
rfix = row_index["固定成本（元/场）"]
rvar = row_index["单人变动成本（元/人）"]
rshare = row_index["夏春侧分润比例"]

COLS = ("B", "C", "D")


def calc_row(label, per_col_formula, fmt="#,##0", total=True, fill=None, memo=""):
    """per_col_formula: function(col_letter)->formula string (without '=')."""
    global r
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = Font(name=FONT, bold=True, size=10.5, color=(WHITE if fill else NAVY))
    lc.alignment = Alignment(vertical="center")
    lc.border = border_all
    if fill:
        lc.fill = PatternFill("solid", fgColor=fill)
    for ci, col in zip((2, 3, 4), COLS):
        cell = ws.cell(row=r, column=ci, value="=" + per_col_formula(col))
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
        cell.font = Font(name=FONT, size=10.5, color=(WHITE if fill else GREYTXT))
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
    tcell = ws.cell(row=r, column=5)
    tcell.border = border_all
    tcell.alignment = Alignment(horizontal=("center" if total else "left"), vertical="center")
    if total:
        tcell.value = "=SUM(B%d:D%d)" % (r, r)
        tcell.number_format = fmt
        tcell.font = Font(name=FONT, size=10.5, bold=True, color=(WHITE if fill else NAVY))
    else:
        tcell.value = memo
        tcell.font = Font(name=FONT, size=9.5, italic=True, color=(WHITE if fill else NAVY2))
    if fill:
        tcell.fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[r].height = 24
    r += 1
    return r - 1


rev_row = calc_row("营业收入（元）", lambda c: f"{c}{rp}*{c}{rprice}")
cost_row = calc_row("总成本（元）", lambda c: f"{c}{rfix}+{c}{rvar}*{c}{rp}")
gp_row = calc_row("毛利（元）", lambda c: f"{c}{rev_row}-{c}{cost_row}", fill=NAVY2)
calc_row("毛利率", lambda c: f"IF({c}{rev_row}=0,0,{c}{gp_row}/{c}{rev_row})",
         fmt="0.0%", total=False, memo="毛利/收入")
xc_row = calc_row("夏春侧分润（元）", lambda c: f"{c}{gp_row}*{c}{rshare}")
calc_row("我方毛利留存（元）", lambda c: f"{c}{gp_row}-{c}{xc_row}", fill=GOLD)

# 说明脚注
note = ws.cell(row=r + 1, column=1,
               value="说明：以上为示例参数下的测算，蓝色单元格可自行修改（人数 / 客单价 / 成本 / 分润比例），"
                     "其余数值与合计将自动重算。数字仅供沟通参考，不构成收益承诺。")
ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=5)
note.font = Font(name=FONT, size=9.5, italic=True, color=NAVY2)
note.alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "B4"

# =====================================================
# Sheet 7 — 风险与待办
# =====================================================
ws = wb.create_sheet("风险与待办")
ws.sheet_view.showGridLines = False
set_widths(ws, [5, 22, 14, 20, 44, 16, 10, 10])
title_block(ws, "关键问题、风险与待办清单", "多为双方前期沟通中的待解决问题 / 待办事项，过渡期内逐项落定并写入 MOU", 8)
hr = 3
style_header(ws, hr, ["序号", "事项", "类别", "来源", "应对措施", "责任方", "优先级", "状态"])
risks = [
    ("主导权与客户归属界定", "商务", "前期沟通", "明确长三角新增客户归属与交叉销售权益划分，写入 MOU。", "双方", "高", "待开始"),
    ("夏春档期与精力稀缺", "资源", "前期沟通", "我方承接执行与批量排期；录播 / 图文复用，关键场次才需本人。", "我方", "高", "待开始"),
    ("与新 / 老团队边界重叠", "商务", "前期沟通", "只做区域增量，签互不重叠 / 不竞争条款，定期同步避免撞单。", "双方", "高", "待开始"),
    ("异地执行精力有限", "运营", "前期沟通", "由我方长三角本地团队承接落地，夏春侧远程支持。", "我方", "中", "待开始"),
    ("合规与免责", "合规", "夏春惯例", "金融观点仅供参考、不构成投资建议；内容与产品销售分离。", "双方", "高", "待开始"),
    ("参访 / 分配 / 落地案例参考", "交付", "前期沟通", "参考既有成熟活动案例，我方据此做长三角本地化适配与收益分配设计。", "我方", "中", "待开始"),
    ("香港游学完整方案", "交付", "前期沟通", "我方主导长三角组团方案，夏春确认讲座与档期。", "我方", "高", "待开始"),
    ("CRM / 内容（出书）方案", "交付", "前期沟通", "我方搭建区域 CRM 与私域，内容出版按长期规划推进。", "我方", "中", "待开始"),
    ("形成合作备忘录（MOU）", "商务", "前期沟通", "过渡期第 1 月内会签 MOU 与区域独家授权草案。", "双方", "高", "待开始"),
    ("建立微信群同步进度", "协作", "前期沟通", "组建联合工作群，按周同步进度、加微对接细节。", "双方", "中", "待开始"),
]
r = hr + 1
for i, rk in enumerate(risks, 1):
    pr = rk[5]
    prfill = GOLD if pr == "高" else (TEAL if pr == "中" else LGREY)
    write_row(ws, r, [i, rk[0], rk[1], rk[2], rk[3], rk[4], rk[5], rk[6]],
              bold_cols=(2,), fills={7: prfill},
              aligns=["center", "left", "center", "center", "left", "center", "center", "center"], height=36)
    ws.cell(row=r, column=7).font = Font(name=FONT, bold=True, color=WHITE, size=10.5)
    r += 1
zebra(ws, hr + 1, r - 1, 8)
dv2 = DataValidation(type="list", formula1='"待开始,进行中,已完成,已关闭"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add(f"H{hr+1}:H{r-1}")
ws.freeze_panes = "A4"

# =====================================================
# Sheet 8 — KPI 与转长期安排
# =====================================================
ws = wb.create_sheet("KPI与转长期")
ws.sheet_view.showGridLines = False
set_widths(ws, [5, 26, 30, 30, 18])
title_block(ws, "KPI 与转长期安排", "过渡顺畅即自动转为长期区域独家总代理；下列为过渡期与长期的参考指标（非硬性考核门槛），签约前双方确认", 5)
hr = 3
style_header(ws, hr, ["序号", "指标", "过渡期参考值（约3个月）", "长期目标（年度）", "衡量方式"])
kpis = [
    ("落地活动场次", "≥ 3 场", "≥ 24 场 / 年", "活动交付记录"),
    ("累计营收", "≥ 60 万元", "≥ 800 万元 / 年", "财务对账"),
    ("综合毛利率", "≥ 60%", "≥ 65%", "财务测算"),
    ("客户满意度（NPS/评分）", "≥ 4.5 / 5", "≥ 4.6 / 5", "活动问卷"),
    ("新增私域会员", "≥ 1000 人", "≥ 8000 人 / 年", "CRM 统计"),
    ("香港 / 出海游学", "预售 ≥ 1 团", "≥ 4 团 / 年", "报名与成团"),
    ("夏春时间投入产出比", "关键场次出席，其余复用", "IP 化年度框架", "档期记录"),
]
r = hr + 1
for i, k in enumerate(kpis, 1):
    write_row(ws, r, [i] + list(k), bold_cols=(2,),
              aligns=["center", "left", "center", "center", "left"], height=30,
              fills={3: GOLDL})
    ws.cell(row=r, column=3).font = Font(name=FONT, size=10.5, bold=True, color=NAVY)
    r += 1
gate = ws.cell(row=r + 1, column=1,
               value="转长期安排：三个月适度过渡期为双方适应与适配阶段；若运作顺畅、无重大问题，即自动转为长期区域独家总代理并会签正式协议。"
                     "上列指标仅作过渡期健康度参考，非硬性考核门槛。")
ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=5)
gate.font = Font(name=FONT, size=10.5, bold=True, color=WHITE)
gate.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
for rr in (r + 1, r + 2):
    for cc in range(1, 6):
        ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=NAVY)
ws.freeze_panes = "A4"

# 打印版式：横向 + 适应一页宽，避免跨页
for sh in wb.worksheets:
    sh.page_setup.orientation = "landscape"
    sh.page_setup.fitToWidth = 1
    sh.page_setup.fitToHeight = 0
    sh.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sh.page_margins.left = sh.page_margins.right = 0.3
    sh.page_margins.top = sh.page_margins.bottom = 0.4

out = "/workspace/deliverables/夏春长三角合作方案_配套测算与清单.xlsx"
wb.save(out)
print("saved:", out, "sheets:", wb.sheetnames)
